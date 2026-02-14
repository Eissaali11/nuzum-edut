from flask import Blueprint, render_template, request, jsonify, make_response, send_file, redirect, url_for
from flask_login import login_required
from sqlalchemy import func, or_
from datetime import datetime, date, timedelta
from io import BytesIO
from utils.pdf import create_pdf, arabic_text, create_data_table, get_styles
from app import db
from models import Department, Employee, Attendance, Salary, Document, SystemAudit, Vehicle, Fee, VehicleChecklist, VehicleDamageMarker, VehicleChecklistImage, employee_departments
from utils.date_converter import parse_date, format_date_hijri, format_date_gregorian, get_month_name_ar
from utils.excel import generate_employee_excel, generate_salary_excel
from utils.vehicles_export import export_vehicle_pdf, export_vehicle_excel
from utils.pdf_generator import generate_salary_report_pdf
from utils.vehicle_checklist_pdf import create_vehicle_checklist_pdf
# إضافة الاستيرادات المفقودة
import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import pandas as pd
from utils.salary_report_pdf import generate_salary_report_pdf


reports_bp = Blueprint('reports', __name__, url_prefix='/reports')

# طرق تصدير التقارير للنسخة المحمولة
@reports_bp.route('/export/vehicles/<export_type>')
@login_required
def export_vehicles_report(export_type):
    """
    تصدير تقرير المركبات من النسخة المحمولة
    :param export_type: نوع التصدير ('pdf' أو 'excel')
    """
    # نقل معلمات البحث من الطلب الحالي
    vehicle_type = request.args.get('vehicle_type', '')
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    
    if export_type == 'pdf':
        # إعادة توجيه إلى طريقة تصدير PDF الحالية
        return redirect(url_for('reports.vehicles_pdf', vehicle_type=vehicle_type, status=status, search=search))
    elif export_type == 'excel':
        # إعادة توجيه إلى طريقة تصدير Excel الحالية
        return redirect(url_for('reports.vehicles_excel', vehicle_type=vehicle_type, status=status, search=search))
    else:
        # إعادة توجيه إلى صفحة تقارير المركبات المحمولة
        return redirect(url_for('mobile.report_vehicles'))


@reports_bp.route('/vehicle_checklist/<int:checklist_id>/pdf')
@login_required
def vehicle_checklist_pdf(checklist_id):
    """
    تصدير تقرير فحص المركبة إلى PDF مع عرض علامات التلف
    :param checklist_id: معرف سجل الفحص
    """
    try:
        # الحصول على بيانات الفحص
        checklist = VehicleChecklist.query.get_or_404(checklist_id)
        
        # الحصول على بيانات المركبة وفحص قيود العمليات
        vehicle = Vehicle.query.get_or_404(checklist.vehicle_id)
        
        # فحص حالة السيارة - عرض تحذير للسيارات خارج الخدمة
        from utils.vehicle_route_helpers import check_vehicle_operation_restrictions
        restrictions = check_vehicle_operation_restrictions(vehicle)
        if restrictions['blocked']:
            # إضافة تحذير في بداية التقرير ولكن السماح بعرض التشك لست التاريخي
            print(f"تحذير: {restrictions['message']}")
        
        # جمع بيانات عناصر الفحص مرتبة حسب الفئة
        checklist_items = {}
        for item in checklist.checklist_items:
            if item.category not in checklist_items:
                checklist_items[item.category] = []
            
            checklist_items[item.category].append(item)
        
        # الحصول على علامات التلف المرتبطة بهذا الفحص
        damage_markers = VehicleDamageMarker.query.filter_by(checklist_id=checklist_id).all()
        
        # الحصول على صور الفحص المرفقة
        checklist_images = VehicleChecklistImage.query.filter_by(checklist_id=checklist_id).all()
        
        # إنشاء ملف PDF
        pdf_buffer = create_vehicle_checklist_pdf(
            checklist=checklist,
            vehicle=vehicle,
            checklist_items=checklist_items,
            damage_markers=damage_markers,
            checklist_images=checklist_images
        )
        
        # إنشاء استجابة تحميل للملف
        response = make_response(pdf_buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=vehicle_checklist_{checklist_id}.pdf'
        
        return response
        
    except Exception as e:
        # تسجيل الخطأ وإعادة توجيه المستخدم
        print(f"خطأ في إنشاء ملف PDF لفحص المركبة: {str(e)}")
        return redirect(url_for('mobile.vehicle_checklist_details', checklist_id=checklist_id))

@reports_bp.route('/export/fees/<export_type>')
@login_required
def export_fees_report(export_type):
    """
    تصدير تقرير الرسوم من النسخة المحمولة
    :param export_type: نوع التصدير ('pdf' أو 'excel')
    """
    # نقل معلمات البحث من الطلب الحالي
    fee_type = request.args.get('fee_type', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status = request.args.get('status', '')
    
    if export_type == 'pdf':
        # إعادة توجيه إلى طريقة تصدير PDF الحالية
        return redirect(url_for('reports.fees_pdf', fee_type=fee_type, date_from=date_from, date_to=date_to, status=status))
    elif export_type == 'excel':
        # إعادة توجيه إلى طريقة تصدير Excel الحالية
        return redirect(url_for('reports.fees_excel', fee_type=fee_type, date_from=date_from, date_to=date_to, status=status))
    else:
        # إعادة توجيه إلى صفحة تقارير الرسوم المحمولة
        return redirect(url_for('mobile.report_fees'))

@reports_bp.route('/export/employees/<export_type>')
@login_required
def export_employees_report(export_type):
    """
    تصدير تقرير الموظفين من النسخة المحمولة
    :param export_type: نوع التصدير ('pdf' أو 'excel')
    """
    # نقل معلمات البحث من الطلب الحالي
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    
    if export_type == 'pdf':
        # إعادة توجيه إلى طريقة تصدير PDF الحالية
        return redirect(url_for('reports.employees_pdf', department_id=department_id, status=status, search=search))
    elif export_type == 'excel':
        # إعادة توجيه إلى طريقة تصدير Excel الحالية
        return redirect(url_for('reports.employees_excel', department_id=department_id, status=status, search=search))
    else:
        # إعادة توجيه إلى صفحة تقارير الموظفين المحمولة
        return redirect(url_for('mobile.report_employees'))

# دوال متعلقة بتصدير السيارات
@reports_bp.route('/vehicles/pdf')
@login_required
def vehicles_pdf():
    """تصدير تقرير المركبات إلى PDF"""
    # الحصول على معلمات الفلتر
    vehicle_type = request.args.get('vehicle_type', '')
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    
    # إنشاء استعلام المركبات
    query = Vehicle.query
    
    # تطبيق الفلترة على المركبات
    if vehicle_type:
        query = query.filter_by(make=vehicle_type)  # نستخدم make بدلاً من vehicle_type
    
    if status:
        query = query.filter_by(status=status)
    
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Vehicle.plate_number.like(search_term),
                Vehicle.make.like(search_term),
                Vehicle.model.like(search_term),
                Vehicle.color.like(search_term)
            )
        )
    
    # الحصول على المركبات المرتبة حسب الترتيب
    vehicles = query.order_by(Vehicle.plate_number).all()
    
    # تحضير مخرجات التقرير
    buffer = BytesIO()
    
    # إنشاء تقرير PDF للمركبات
    data = []
    for vehicle in vehicles:
        data.append({
            'plate_number': vehicle.plate_number,
            'make': vehicle.make,
            'model': vehicle.model,
            'color': vehicle.color,
            'year': vehicle.year,
            'status': vehicle.status
        })
    
    # استدعاء دالة إنشاء PDF للمركبات
    report_title = "تقرير المركبات"
    export_vehicle_pdf(buffer, data, report_title)
    
    # إرسال الملف كمرفق للتنزيل
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"vehicles_report_{datetime.now().strftime('%Y%m%d')}.pdf",
        mimetype='application/pdf'
    )

@reports_bp.route('/vehicles/excel')
@login_required
def vehicles_excel():
    """تصدير تقرير المركبات إلى Excel"""
    from utils.excel import generate_vehicles_excel
    
    # الحصول على معلمات الفلتر
    vehicle_type = request.args.get('vehicle_type', '')
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    
    # إنشاء استعلام المركبات
    query = Vehicle.query
    
    # تطبيق الفلترة على المركبات
    if vehicle_type:
        query = query.filter_by(make=vehicle_type)  # نستخدم make بدلاً من vehicle_type
    
    if status:
        query = query.filter_by(status=status)
    
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Vehicle.plate_number.like(search_term),
                Vehicle.make.like(search_term),
                Vehicle.model.like(search_term),
                Vehicle.color.like(search_term)
            )
        )
    
    # الحصول على المركبات المرتبة حسب الترتيب
    vehicles = query.order_by(Vehicle.plate_number).all()
    
    # تحضير مخرجات التقرير
    output = BytesIO()
    
    # استدعاء دالة إنشاء Excel الاحترافية للمركبات
    generate_vehicles_excel(vehicles, output)
    
    # إرسال الملف كمرفق للتنزيل
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"vehicles_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports_bp.route('/fees/pdf')
@login_required
def fees_pdf():
    """تصدير تقرير الرسوم إلى PDF"""
    # الحصول على معلمات الفلتر
    fee_type = request.args.get('fee_type', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status = request.args.get('status', '')  # paid/unpaid
    
    # تحويل التواريخ إلى كائنات datetime إذا تم تحديدها
    if date_from:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    if date_to:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # إنشاء استعلام الرسوم
    query = Fee.query
    
    # تطبيق الفلترة على الرسوم
    if fee_type:
        query = query.filter_by(fee_type=fee_type)
    
    if date_from:
        query = query.filter(Fee.due_date >= date_from)
    
    if date_to:
        query = query.filter(Fee.due_date <= date_to)
    
    if status:
        is_paid_bool = (status.lower() == 'paid')
        query = query.filter(Fee.is_paid == is_paid_bool)
    
    # الحصول على قائمة الرسوم المرتبة حسب تاريخ الاستحقاق
    fees = query.order_by(Fee.due_date).all()
    
    # تحضير مخرجات التقرير
    buffer = BytesIO()
    
    # إنشاء مستند PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # تحضير النمط
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Arabic', fontName='Amiri', fontSize=12, alignment=1)) # للتوسيط
    styles.add(ParagraphStyle(name='ArabicTitle', fontName='Amiri-Bold', fontSize=16, alignment=1))
    styles.add(ParagraphStyle(name='ArabicSubTitle', fontName='Amiri-Bold', fontSize=14, alignment=1))
    
    # تسجيل الخطوط
    pdfmetrics.registerFont(TTFont('Amiri', 'static/fonts/Amiri-Regular.ttf'))
    pdfmetrics.registerFont(TTFont('Amiri-Bold', 'static/fonts/Amiri-Bold.ttf'))
    
    # إعداد المحتوى
    content = []
    
    # إضافة عنوان التقرير
    title = Paragraph(arabic_text("تقرير الرسوم"), styles['ArabicTitle'])
    content.append(title)
    content.append(Spacer(1, 20))
    
    # فلاتر البحث
    filter_text = []
    if fee_type:
        filter_text.append(f"نوع الرسوم: {fee_type}")
    if date_from:
        filter_text.append(f"من تاريخ: {date_from.strftime('%Y-%m-%d')}")
    if date_to:
        filter_text.append(f"إلى تاريخ: {date_to.strftime('%Y-%m-%d')}")
    if status:
        status_text = "مدفوعة" if status.lower() == 'paid' else "غير مدفوعة"
        filter_text.append(f"الحالة: {status_text}")
    
    if filter_text:
        filters = Paragraph(arabic_text(" - ".join(filter_text)), styles['Arabic'])
        content.append(filters)
        content.append(Spacer(1, 10))
    
    # جدول الرسوم
    if fees:
        # إنشاء بيانات الجدول
        data = [
            [
                arabic_text("نوع الرسوم"),
                arabic_text("الوصف"),
                arabic_text("المبلغ (ر.س)"),
                arabic_text("تاريخ الاستحقاق"),
                arabic_text("حالة الدفع"),
                arabic_text("المستلم")
            ]
        ]
        
        # إضافة بيانات الرسوم
        for fee in fees:
            is_paid_text = "مدفوعة" if fee.is_paid else "غير مدفوعة"
            data.append([
                arabic_text(fee.fee_type),
                arabic_text(fee.description or ""),
                arabic_text(f"{fee.amount:.2f}"),
                arabic_text(fee.due_date.strftime("%Y-%m-%d") if fee.due_date else ""),
                arabic_text(is_paid_text),
                arabic_text(fee.recipient or "")
            ])
        
        # إنشاء الجدول وتنسيقه
        table = Table(data, colWidths=[doc.width/6] * 6)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Amiri'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        content.append(table)
        content.append(Spacer(1, 20))
        
        # إحصائيات الرسوم
        total_fees = sum(fee.amount for fee in fees if fee.amount)
        total_paid = sum(fee.amount for fee in fees if fee.amount and fee.is_paid)
        total_unpaid = sum(fee.amount for fee in fees if fee.amount and not fee.is_paid)
        
        stats = [
            Paragraph(arabic_text(f"إجمالي الرسوم: {total_fees:.2f} ر.س"), styles['Arabic']),
            Paragraph(arabic_text(f"إجمالي المدفوع: {total_paid:.2f} ر.س"), styles['Arabic']),
            Paragraph(arabic_text(f"إجمالي غير المدفوع: {total_unpaid:.2f} ر.س"), styles['Arabic'])
        ]
        
        for stat in stats:
            content.append(stat)
            content.append(Spacer(1, 5))
    else:
        content.append(Paragraph(arabic_text("لا توجد رسوم مطابقة للمعايير المحددة"), styles['Arabic']))
    
    # التذييل
    content.append(Spacer(1, 20))
    footer_text = Paragraph(
        arabic_text(f"تم إنشاء هذا التقرير بواسطة نُظم - نظام إدارة متكامل في {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
        styles['Arabic']
    )
    content.append(footer_text)
    
    # بناء الوثيقة
    doc.build(content)
    buffer.seek(0)
    
    # إرسال الملف كمرفق للتنزيل
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"fees_report_{datetime.now().strftime('%Y%m%d')}.pdf",
        mimetype='application/pdf'
    )

@reports_bp.route('/fees/excel')
@login_required
def fees_excel():
    """تصدير تقرير الرسوم إلى Excel"""
    # الحصول على معلمات الفلتر
    fee_type = request.args.get('fee_type', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status = request.args.get('status', '')  # paid/unpaid
    
    # تحويل التواريخ إلى كائنات datetime إذا تم تحديدها
    if date_from:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    if date_to:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # إنشاء استعلام الرسوم
    query = Fee.query
    
    # تطبيق الفلترة على الرسوم
    if fee_type:
        query = query.filter_by(fee_type=fee_type)
    
    if date_from:
        query = query.filter(Fee.due_date >= date_from)
    
    if date_to:
        query = query.filter(Fee.due_date <= date_to)
    
    if status:
        is_paid_bool = (status.lower() == 'paid')
        query = query.filter(Fee.is_paid == is_paid_bool)
    
    # الحصول على قائمة الرسوم المرتبة حسب تاريخ الاستحقاق
    fees = query.order_by(Fee.due_date).all()
    
    # تحضير بيانات Excel
    data = []
    for fee in fees:
        is_paid_text = "مدفوعة" if fee.is_paid else "غير مدفوعة"
        data.append({
            'نوع الرسوم': fee.fee_type,
            'الوصف': fee.description or "",
            'المبلغ (ر.س)': fee.amount,
            'تاريخ الاستحقاق': fee.due_date.strftime("%Y-%m-%d") if fee.due_date else "",
            'حالة الدفع': is_paid_text,
            'المستلم': fee.recipient or ""
        })
    
    # إنشاء DataFrame
    df = pd.DataFrame(data)
    
    # تحضير مخرجات التقرير
    output = BytesIO()
    
    # إنشاء ملف Excel
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # كتابة DataFrame إلى ورقة العمل
        df.to_excel(writer, sheet_name='تقرير الرسوم', index=False)
        
        # الحصول على ورقة العمل وكائن المصنف
        workbook = writer.book
        worksheet = writer.sheets['تقرير الرسوم']
        
        # إضافة تنسيقات
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'center',
            'align': 'center',
            'bg_color': '#D7E4BC',
            'border': 1
        })
        
        # تطبيق التنسيق على العناوين
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
        
        # إضافة تنسيق للخلايا
        cell_format = workbook.add_format({
            'align': 'right',
            'valign': 'vcenter',
            'border': 1
        })
        
        # تطبيق التنسيق على الخلايا
        for row in range(1, len(df) + 1):
            for col in range(len(df.columns)):
                worksheet.write(row, col, df.iloc[row-1, col], cell_format)
        
        # تنسيق خاص لعمود المبلغ
        money_format = workbook.add_format({
            'align': 'right',
            'valign': 'vcenter',
            'border': 1,
            'num_format': '#,##0.00 ر.س'
        })
        
        # تطبيق تنسيق المبالغ
        amount_col = df.columns.get_loc('المبلغ (ر.س)')
        for row in range(1, len(df) + 1):
            worksheet.write(row, amount_col, df.iloc[row-1, amount_col], money_format)
        
        # إضافة ورقة للإحصائيات
        stats_df = pd.DataFrame({
            'البيان': ['إجمالي الرسوم', 'إجمالي المدفوع', 'إجمالي غير المدفوع'],
            'القيمة': [
                sum(fee.amount for fee in fees if fee.amount),
                sum(fee.amount for fee in fees if fee.amount and fee.is_paid),
                sum(fee.amount for fee in fees if fee.amount and not fee.is_paid)
            ]
        })
        
        stats_df.to_excel(writer, sheet_name='إحصائيات', index=False)
        stats_worksheet = writer.sheets['إحصائيات']
        
        # تنسيق ورقة الإحصائيات
        for col_num, value in enumerate(stats_df.columns.values):
            stats_worksheet.write(0, col_num, value, header_format)
        
        for row in range(1, len(stats_df) + 1):
            stats_worksheet.write(row, 0, stats_df.iloc[row-1, 0], cell_format)
            stats_worksheet.write(row, 1, stats_df.iloc[row-1, 1], money_format)
        
        # ضبط عرض الأعمدة
        for i, col in enumerate(df.columns):
            column_len = max(df[col].astype(str).map(len).max(), len(col) + 2)
            worksheet.set_column(i, i, column_len)
        
        stats_worksheet.set_column(0, 0, 20)
        stats_worksheet.set_column(1, 1, 15)
    
    # إرسال الملف كمرفق للتنزيل
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"fees_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports_bp.route('/')
def index():
    """الصفحة الرئيسية للتقارير"""
    try:
        departments = Department.query.all()
        return render_template('reports/index.html', departments=departments)
    except Exception as e:
        print(f"Error in reports index: {e}")
        # في حالة وجود خطأ، نرجع صفحة بسيطة
        return render_template('reports/index.html', departments=[])

@reports_bp.route('/employees')
def employees_report():
    """تقرير الموظفين حسب القسم"""
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    query = Employee.query
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter_by(department_id=department_id)
    if status:
        query = query.filter_by(status=status)
    
    employees = query.all()
    departments = Department.query.all()
    
    return render_template('reports/employees.html', 
                          employees=employees, 
                          departments=departments,
                          department_id=department_id,
                          status=status)

@reports_bp.route('/employees/pdf')
def employees_pdf():
    """تصدير تقرير الموظفين إلى PDF"""
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    query = Employee.query
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter_by(department_id=department_id)
        department = Department.query.get(department_id)
        department_name = department.name if department else ""
    else:
        department_name = "جميع الأقسام"
    
    if status:
        query = query.filter_by(status=status)
        if status == 'active':
            status_name = "نشط"
        elif status == 'inactive':
            status_name = "غير نشط"
        elif status == 'on_leave':
            status_name = "في إجازة"
        else:
            status_name = ""
    else:
        status_name = "جميع الحالات"
    
    employees = query.all()
    
    # استخدام المكتبة الموحدة لإنشاء PDF
    from utils.pdf import arabic_text, create_pdf, create_data_table, get_styles
    from reportlab.lib.units import cm
    from reportlab.platypus import Spacer, Paragraph
    
    # تجهيز العناصر
    elements = []
    
    # إضافة العنوان
    styles = get_styles()
    title = f"تقرير الموظفين - {department_name} - {status_name}"
    elements.append(Paragraph(arabic_text(title), styles['title']))
    elements.append(Spacer(1, 20))
    
    # إضافة تاريخ التقرير
    date_text = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d')}"
    elements.append(Paragraph(arabic_text(date_text), styles['arabic']))
    elements.append(Spacer(1, 20))
    
    # إعداد جدول البيانات
    headers = ["الاسم", "الرقم الوظيفي", "الرقم الوطني", "الهاتف", "المسمى الوظيفي", "القسم", "الحالة"]
    data = []
    
    # إضافة بيانات الموظفين
    for emp in employees:
        department_name = emp.department.name if emp.department else "---"
        
        # ترجمة حالة الموظف
        status_map = {
            'active': 'نشط',
            'inactive': 'غير نشط',
            'on_leave': 'في إجازة'
        }
        status_text = status_map.get(emp.status, emp.status)
        
        row = [
            arabic_text(emp.name),
            emp.employee_id,
            emp.national_id,
            emp.mobile,
            arabic_text(emp.job_title),
            arabic_text(department_name),
            arabic_text(status_text)
        ]
        data.append(row)
    
    # إنشاء الجدول
    if data:
        col_widths = [3*cm, 2*cm, 2*cm, 2*cm, 3*cm, 3*cm, 2*cm]
        table = create_data_table(headers, data, col_widths)
        elements.append(table)
    else:
        no_data_text = "لا توجد بيانات متاحة"
        elements.append(Paragraph(arabic_text(no_data_text), styles['arabic']))
    
    # إنشاء ملف PDF
    buffer = create_pdf(elements, landscape_mode=True)
    
    # إنشاء استجابة تحميل
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="employees_report.pdf",
        mimetype='application/pdf'
    )

@reports_bp.route('/employees/excel')
def employees_excel():
    """تصدير تقرير الموظفين إلى Excel"""
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    query = Employee.query
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter_by(department_id=department_id)
    if status:
        query = query.filter_by(status=status)
    
    employees = query.all()
    
    # توليد ملف Excel
    output = generate_employee_excel(employees)
    
    # إنشاء استجابة تحميل
    return send_file(
        output,
        as_attachment=True,
        download_name=f'employees_report_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports_bp.route('/attendance')
def attendance_report():
    """تقرير الحضور والغياب"""
    # الحصول على معلمات الفلتر
    from_date_str = request.args.get('from_date', (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
    to_date_str = request.args.get('to_date', datetime.now().strftime('%Y-%m-%d'))
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    # معالجة التواريخ
    try:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)
    except ValueError:
        from_date = datetime.now() - timedelta(days=7)
        to_date = datetime.now()
    
    # استعلام الحضور
    query = db.session.query(
            Attendance, Employee
        ).join(
            Employee, Attendance.employee_id == Employee.id
        ).filter(
            Attendance.date.between(from_date, to_date)
        )
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
    if status:
        query = query.filter(Attendance.status == status)
    
    # الحصول على النتائج النهائية
    results = query.order_by(Attendance.date.desc()).all()
    
    # الحصول على قائمة الأقسام لعناصر الفلتر
    departments = Department.query.all()
    
    return render_template('reports/attendance.html',
                        results=results,
                        departments=departments,
                        from_date=from_date,
                        to_date=to_date,
                        department_id=department_id,
                        status=status,
                        format_date_gregorian=format_date_gregorian,
                        format_date_hijri=format_date_hijri)

@reports_bp.route('/attendance/pdf')
def attendance_pdf():
    """تصدير تقرير الحضور إلى PDF"""
    # الحصول على معلمات الفلتر
    from_date_str = request.args.get('from_date', (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
    to_date_str = request.args.get('to_date', datetime.now().strftime('%Y-%m-%d'))
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    # معالجة التواريخ
    try:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)
    except ValueError:
        from_date = datetime.now() - timedelta(days=7)
        to_date = datetime.now()
    
    # استعلام الحضور
    query = db.session.query(
            Attendance, Employee
        ).join(
            Employee, Attendance.employee_id == Employee.id
        ).filter(
            Attendance.date.between(from_date, to_date)
        )
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
        department = Department.query.get(department_id)
        department_name = department.name if department else ""
    else:
        department_name = "جميع الأقسام"
    
    if status:
        query = query.filter(Attendance.status == status)
        if status == 'present':
            status_name = "حاضر"
        elif status == 'absent':
            status_name = "غائب"
        elif status == 'leave':
            status_name = "إجازة"
        elif status == 'sick':
            status_name = "مرضي"
        else:
            status_name = ""
    else:
        status_name = "جميع الحالات"
    
    # الحصول على النتائج النهائية
    results = query.order_by(Attendance.date.desc()).all()
    
    # إنشاء ملف PDF
    buffer = BytesIO()
    
    # تسجيل الخط العربي
    try:
        # محاولة تسجيل الخط العربي إذا لم يكن مسجلاً مسبقًا
        pdfmetrics.registerFont(TTFont('Arabic', 'static/fonts/Arial.ttf'))
    except:
        # إذا كان هناك خطأ، نستخدم الخط الافتراضي
        pass
    
    # تعيين أبعاد الصفحة واتجاهها
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # إعداد الأنماط
    styles = getSampleStyleSheet()
    # إنشاء نمط للنص العربي
    arabic_style = ParagraphStyle(
        name='Arabic',
        parent=styles['Normal'],
        fontName='Arabic',
        fontSize=12,
        alignment=1, # وسط
        textColor=colors.black
    )
    
    # إنشاء نمط للعناوين
    title_style = ParagraphStyle(
        name='Title',
        parent=styles['Title'],
        fontName='Arabic',
        fontSize=16,
        alignment=1, # وسط
        textColor=colors.black
    )
    
    # إعداد المحتوى
    elements = []
    
    # إضافة العنوان
    title = f"تقرير الحضور والغياب - {department_name} - {status_name}"
    # تهيئة النص العربي للعرض في PDF
    title = get_display(arabic_reshaper.reshape(title))
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 10))
    
    # إضافة نطاق التاريخ
    date_range = f"الفترة من: {format_date_gregorian(from_date)} إلى: {format_date_gregorian(to_date)}"
    date_range = get_display(arabic_reshaper.reshape(date_range))
    elements.append(Paragraph(date_range, arabic_style))
    elements.append(Spacer(1, 20))
    
    # إعداد جدول البيانات
    headers = ["التاريخ", "الاسم", "الرقم الوظيفي", "وقت الحضور", "وقت الانصراف", "الحالة", "القسم"]
    data = []
    
    # إضافة الرؤوس
    headers_display = [get_display(arabic_reshaper.reshape(h)) for h in headers]
    data.append(headers_display)
    
    # إضافة بيانات الحضور
    for attendance, employee in results:
        department_name = employee.department.name if employee.department else "---"
        
        # ترجمة حالة الحضور
        status_map = {
            'present': 'حاضر',
            'absent': 'غائب',
            'leave': 'إجازة',
            'sick': 'مرضي'
        }
        status_text = status_map.get(attendance.status, attendance.status)
        
        row = [
            format_date_gregorian(attendance.date),
            get_display(arabic_reshaper.reshape(employee.name)),
            employee.employee_id,
            str(attendance.check_in) if attendance.check_in else "---",
            str(attendance.check_out) if attendance.check_out else "---",
            get_display(arabic_reshaper.reshape(status_text)),
            get_display(arabic_reshaper.reshape(department_name))
        ]
        data.append(row)
    
    # إنشاء الجدول
    if data:
        # حساب العرض المناسب للجدول بناءً على حجم الصفحة
        table_width = landscape(A4)[0] - 4*cm  # العرض الإجمالي ناقص الهوامش
        col_widths = [table_width/len(headers)] * len(headers)  # توزيع متساوي
        table = Table(data, colWidths=col_widths)
        
        # إعداد أنماط الجدول
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),  # لون خلفية العناوين
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # لون نص العناوين
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # محاذاة النص
            ('FONTNAME', (0, 0), (-1, 0), 'Arabic'),  # خط العناوين
            ('FONTSIZE', (0, 0), (-1, 0), 12),  # حجم خط العناوين
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # تباعد أسفل العناوين
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # لون خلفية البيانات
            ('FONTNAME', (0, 1), (-1, -1), 'Arabic'),  # خط البيانات
            ('FONTSIZE', (0, 1), (-1, -1), 10),  # حجم خط البيانات
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # حدود الجدول
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # محاذاة النص عموديا
        ])
        
        # تطبيق التناوب في ألوان الصفوف لتحسين القراءة
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
        
        table.setStyle(table_style)
        elements.append(table)
    else:
        no_data_text = get_display(arabic_reshaper.reshape("لا توجد بيانات متاحة"))
        elements.append(Paragraph(no_data_text, arabic_style))
    
    # إضافة معلومات التقرير في أسفل الصفحة
    elements.append(Spacer(1, 20))
    footer_text = f"تاريخ إنشاء التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    footer_text = get_display(arabic_reshaper.reshape(footer_text))
    elements.append(Paragraph(footer_text, arabic_style))
    
    # بناء المستند
    doc.build(elements)
    
    # إعادة المؤشر إلى بداية البايت والإرجاع كملف
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="attendance_report.pdf",
        mimetype='application/pdf'
    )

@reports_bp.route('/attendance/excel')
def attendance_excel():
    """تصدير تقرير الحضور إلى Excel بتصميم احترافي ومنسق مع داش بورد تحليلي"""
    # الحصول على معلمات الفلتر
    from_date_str = request.args.get('from_date', (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
    to_date_str = request.args.get('to_date', datetime.now().strftime('%Y-%m-%d'))
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    # معالجة التواريخ
    try:
        from_date = parse_date(from_date_str)
        to_date = parse_date(to_date_str)
    except ValueError:
        from_date = datetime.now() - timedelta(days=7)
        to_date = datetime.now()
    
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    
    # إنشاء ملف Excel جديد
    output = io.BytesIO()
    wb = Workbook()
    
    # حذف الصفحة الافتراضية
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # ===== 1. صفحة Dashboard الرئيسية =====
    ws_dashboard = wb.create_sheet("📊 لوحة المعلومات", 0)
    
    # جلب البيانات لجميع الأقسام
    departments = Department.query.all() if not department_id else [Department.query.get(department_id)]
    
    # جمع إحصائيات عامة
    total_employees = 0
    total_present = 0
    total_absent = 0
    total_leave = 0
    total_sick = 0
    total_records = 0
    
    department_stats = []
    
    for dept in departments:
        if not dept:
            continue
        
        # جلب موظفي القسم عبر العلاقة الصحيحة (استبعاد المنتهية خدمتهم فقط)
        employees = [emp for emp in dept.employees if emp.status not in ['terminated', 'inactive']]
        dept_employee_count = len(employees)
        
        if dept_employee_count == 0:
            continue
        
        employee_ids = [emp.id for emp in employees]
        
        # جلب سجلات الحضور
        attendance_records = Attendance.query.filter(
            Attendance.employee_id.in_(employee_ids),
            Attendance.date >= from_date,
            Attendance.date <= to_date
        ).all()
        
        # حساب الإحصائيات
        present_count = sum(1 for r in attendance_records if r.status == 'present')
        absent_count = sum(1 for r in attendance_records if r.status == 'absent')
        leave_count = sum(1 for r in attendance_records if r.status == 'leave')
        sick_count = sum(1 for r in attendance_records if r.status == 'sick')
        dept_total = len(attendance_records)
        
        attendance_rate = (present_count / dept_total * 100) if dept_total > 0 else 0
        
        # جمع أسماء الغائبين والإجازات
        absentees = []
        on_leave = []
        sick_list = []
        
        for record in attendance_records:
            employee = next((e for e in employees if e.id == record.employee_id), None)
            if not employee:
                continue
            
            emp_data = {
                'name': employee.name,
                'employee_id': employee.employee_id,
                'date': record.date,
                'notes': record.notes
            }
            
            if record.status == 'absent':
                absentees.append(emp_data)
            elif record.status == 'leave':
                on_leave.append(emp_data)
            elif record.status == 'sick':
                sick_list.append(emp_data)
        
        department_stats.append({
            'name': dept.name,
            'employees': dept_employee_count,
            'present': present_count,
            'absent': absent_count,
            'leave': leave_count,
            'sick': sick_count,
            'total': dept_total,
            'rate': round(attendance_rate, 1),
            'absentees': absentees,
            'on_leave': on_leave,
            'sick_list': sick_list
        })
        
        total_employees += dept_employee_count
        total_present += present_count
        total_absent += absent_count
        total_leave += leave_count
        total_sick += sick_count
        total_records += dept_total
    
    # تنسيقات عامة
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ===== بناء Dashboard الرئيسي =====
    # العنوان
    ws_dashboard.merge_cells('A1:M3')
    title_cell = ws_dashboard['A1']
    title_cell.value = f"📊 تقرير الحضور التحليلي\n{from_date.strftime('%Y/%m/%d')} - {to_date.strftime('%Y/%m/%d')}"
    title_cell.font = Font(size=24, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_dashboard.row_dimensions[1].height = 30
    ws_dashboard.row_dimensions[2].height = 30
    
    # KPIs Row
    kpi_row = 5
    ws_dashboard.row_dimensions[kpi_row].height = 35
    ws_dashboard.row_dimensions[kpi_row + 1].height = 30
    
    # KPI 1: إجمالي الموظفين
    ws_dashboard.merge_cells(f'A{kpi_row}:C{kpi_row+1}')
    kpi1 = ws_dashboard[f'A{kpi_row}']
    kpi1.value = f"👥 إجمالي الموظفين\n{total_employees}"
    kpi1.font = Font(size=16, bold=True, color="FFFFFF")
    kpi1.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    kpi1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    kpi1.border = thick_border
    
    # KPI 2: الحضور
    ws_dashboard.merge_cells(f'D{kpi_row}:F{kpi_row+1}')
    kpi2 = ws_dashboard[f'D{kpi_row}']
    kpi2.value = f"✅ الحضور\n{total_present}"
    kpi2.font = Font(size=16, bold=True, color="FFFFFF")
    kpi2.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    kpi2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    kpi2.border = thick_border
    
    # KPI 3: الغياب
    ws_dashboard.merge_cells(f'G{kpi_row}:I{kpi_row+1}')
    kpi3 = ws_dashboard[f'G{kpi_row}']
    kpi3.value = f"❌ الغياب\n{total_absent}"
    kpi3.font = Font(size=16, bold=True, color="FFFFFF")
    kpi3.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    kpi3.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    kpi3.border = thick_border
    
    # KPI 4: الإجازات
    ws_dashboard.merge_cells(f'J{kpi_row}:L{kpi_row+1}')
    kpi4 = ws_dashboard[f'J{kpi_row}']
    kpi4.value = f"🏖️ الإجازات\n{total_leave}"
    kpi4.font = Font(size=16, bold=True, color="FFFFFF")
    kpi4.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    kpi4.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    kpi4.border = thick_border
    
    # جدول الأقسام
    table_start_row = kpi_row + 3
    ws_dashboard.merge_cells(f'A{table_start_row}:H{table_start_row}')
    ws_dashboard[f'A{table_start_row}'].value = "📋 تفاصيل الأقسام"
    ws_dashboard[f'A{table_start_row}'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_dashboard[f'A{table_start_row}'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws_dashboard[f'A{table_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dashboard.row_dimensions[table_start_row].height = 25
    
    # رؤوس الأعمدة
    headers_row = table_start_row + 1
    headers = ['القسم', 'عدد الموظفين', 'حاضر', 'غائب', 'إجازة', 'مرضي', 'إجمالي السجلات', 'معدل الحضور %']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_dashboard.cell(row=headers_row, column=col_idx)
        cell.value = header
        cell.font = Font(size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws_dashboard.row_dimensions[headers_row].height = 30
    
    # بيانات الأقسام
    data_start_row = headers_row + 1
    for row_idx, dept in enumerate(department_stats, data_start_row):
        values = [dept['name'], dept['employees'], dept['present'], dept['absent'], 
                 dept['leave'], dept['sick'], dept['total'], dept['rate']]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws_dashboard.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    
    # ضبط عرض الأعمدة
    column_widths = [20, 15, 12, 12, 12, 12, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws_dashboard.column_dimensions[get_column_letter(i)].width = width
    
    # ===== صفحة قائمة الغياب =====
    if total_absent > 0:
        ws_absence = wb.create_sheet("📋 قائمة الغياب")
        ws_absence.merge_cells('A1:F3')
        ws_absence['A1'].value = f"📋 قائمة الغياب التفصيلية\n{from_date.strftime('%Y/%m/%d')} - {to_date.strftime('%Y/%m/%d')}"
        ws_absence['A1'].font = Font(size=20, bold=True, color="FFFFFF")
        ws_absence['A1'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        ws_absence['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        current_row = 5
        for dept in department_stats:
            if dept['absentees']:
                ws_absence.merge_cells(f'A{current_row}:F{current_row}')
                ws_absence[f'A{current_row}'].value = f"🏢 {dept['name']} - عدد الغائبين: {len(dept['absentees'])}"
                ws_absence[f'A{current_row}'].font = Font(size=14, bold=True, color="FFFFFF")
                ws_absence[f'A{current_row}'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                ws_absence[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
                
                # رؤوس الأعمدة
                headers = ['#', 'اسم الموظف', 'رقم الموظف', 'التاريخ', 'ملاحظات', 'الحالة']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_absence.cell(row=current_row, column=col_idx)
                    cell.value = header
                    cell.font = Font(size=11, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                current_row += 1
                
                # البيانات
                for idx, emp in enumerate(dept['absentees'], 1):
                    row_data = [idx, emp['name'], emp.get('employee_id', '-'), 
                               emp['date'].strftime('%Y-%m-%d'), emp.get('notes', '-'), 'غائب']
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_absence.cell(row=current_row, column=col_idx)
                        cell.value = value
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                        if current_row % 2 == 0:
                            cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    current_row += 1
                current_row += 1
        
        ws_absence.column_dimensions['A'].width = 8
        ws_absence.column_dimensions['B'].width = 30
        ws_absence.column_dimensions['C'].width = 15
        ws_absence.column_dimensions['D'].width = 15
        ws_absence.column_dimensions['E'].width = 35
        ws_absence.column_dimensions['F'].width = 12
    
    # ===== صفحة قائمة الإجازات =====
    if total_leave > 0:
        ws_leave = wb.create_sheet("🏖️ قائمة الإجازات")
        ws_leave.merge_cells('A1:F3')
        ws_leave['A1'].value = f"🏖️ قائمة الإجازات التفصيلية\n{from_date.strftime('%Y/%m/%d')} - {to_date.strftime('%Y/%m/%d')}"
        ws_leave['A1'].font = Font(size=20, bold=True, color="FFFFFF")
        ws_leave['A1'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        ws_leave['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        current_row = 5
        for dept in department_stats:
            if dept['on_leave']:
                ws_leave.merge_cells(f'A{current_row}:F{current_row}')
                ws_leave[f'A{current_row}'].value = f"🏢 {dept['name']} - عدد الإجازات: {len(dept['on_leave'])}"
                ws_leave[f'A{current_row}'].font = Font(size=14, bold=True, color="FFFFFF")
                ws_leave[f'A{current_row}'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                ws_leave[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
                
                headers = ['#', 'اسم الموظف', 'رقم الموظف', 'التاريخ', 'ملاحظات', 'الحالة']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_leave.cell(row=current_row, column=col_idx)
                    cell.value = header
                    cell.font = Font(size=11, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                current_row += 1
                
                for idx, emp in enumerate(dept['on_leave'], 1):
                    row_data = [idx, emp['name'], emp.get('employee_id', '-'), 
                               emp['date'].strftime('%Y-%m-%d'), emp.get('notes', '-'), 'إجازة']
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_leave.cell(row=current_row, column=col_idx)
                        cell.value = value
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                        if current_row % 2 == 0:
                            cell.fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
                    current_row += 1
                current_row += 1
        
        ws_leave.column_dimensions['A'].width = 8
        ws_leave.column_dimensions['B'].width = 30
        ws_leave.column_dimensions['C'].width = 15
        ws_leave.column_dimensions['D'].width = 15
        ws_leave.column_dimensions['E'].width = 35
        ws_leave.column_dimensions['F'].width = 12
    
    # توليد قائمة بكل الأيام في النطاق الزمني
    date_list = []
    current_date = from_date
    while current_date <= to_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)
    
    # ===== صفحة حضور تفصيلية لكل قسم =====
    for dept_data in department_stats:
        # جلب القسم من قاعدة البيانات
        dept = Department.query.filter_by(name=dept_data['name']).first()
        if not dept:
            continue
        
        # جلب جميع موظفي القسم عبر العلاقة الصحيحة (استبعاد المنتهية خدمتهم فقط)
        employees = [emp for emp in dept.employees if emp.status not in ['terminated', 'inactive']]
        
        if not employees:
            continue
        
        ws_dept = wb.create_sheet(f"🏢 {dept.name[:25]}")
        
        # العنوان
        total_cols = 9 + len(date_list)
        ws_dept.merge_cells(f'A1:{get_column_letter(total_cols)}3')
        ws_dept['A1'].value = f"🏢 تقرير حضور قسم {dept.name}\n{from_date.strftime('%Y/%m/%d')} - {to_date.strftime('%Y/%m/%d')}"
        ws_dept['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws_dept['A1'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        ws_dept['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws_dept.row_dimensions[1].height = 25
        ws_dept.row_dimensions[2].height = 25
        
        # KPIs للقسم
        kpi_row = 5
        ws_dept.row_dimensions[kpi_row].height = 25
        ws_dept.row_dimensions[kpi_row + 1].height = 30
        
        kpis_data = [
            (f'A{kpi_row}:B{kpi_row+1}', 'عدد الموظفين', len(employees), "5B9BD5"),
            (f'C{kpi_row}:D{kpi_row+1}', 'حاضر', dept_data['present'], "70AD47"),
            (f'E{kpi_row}:F{kpi_row+1}', 'غائب', dept_data['absent'], "E74C3C"),
            (f'G{kpi_row}:H{kpi_row+1}', 'إجازة', dept_data['leave'], "F39C12"),
            (f'I{kpi_row}:J{kpi_row+1}', 'مرضي', dept_data['sick'], "3498DB"),
        ]
        
        for cell_range, label, value, color in kpis_data:
            ws_dept.merge_cells(cell_range)
            cell = ws_dept[cell_range.split(':')[0]]
            cell.value = f"{label}\n{value}"
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thick_border
        
        # جدول الحضور
        table_start_row = kpi_row + 3
        
        # رؤوس الأعمدة الثابتة
        headers = ['اسم الموظف', 'رقم الموظف', 'الهوية الوطنية', 'الجوال', 'المسمى الوظيفي', 
                   'الموقع', 'المشروع', 'الإجمالي', 'الحضور']
        
        # إضافة أيام الشهر كرؤوس
        for date in date_list:
            headers.append(f"{date.day}\n{date.strftime('%a')}")
        
        # كتابة رؤوس الأعمدة
        for col_idx, header in enumerate(headers, 1):
            cell = ws_dept.cell(row=table_start_row, column=col_idx)
            cell.value = header
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws_dept.row_dimensions[table_start_row].height = 30
        
        # بيانات الموظفين
        data_start_row = table_start_row + 1
        for emp_idx, employee in enumerate(employees, data_start_row):
            # معلومات الموظف
            ws_dept.cell(row=emp_idx, column=1).value = employee.name
            ws_dept.cell(row=emp_idx, column=2).value = employee.employee_id or '-'
            ws_dept.cell(row=emp_idx, column=3).value = employee.national_id or '-'
            ws_dept.cell(row=emp_idx, column=4).value = employee.mobile or '-'
            ws_dept.cell(row=emp_idx, column=5).value = employee.job_title or '-'
            ws_dept.cell(row=emp_idx, column=6).value = employee.location or '-'
            ws_dept.cell(row=emp_idx, column=7).value = employee.project or '-'
            
            # جلب سجلات الحضور لهذا الموظف
            attendance_records = Attendance.query.filter(
                Attendance.employee_id == employee.id,
                Attendance.date >= from_date,
                Attendance.date <= to_date
            ).all()
            
            # إنشاء dictionary لربط التاريخ بالحالة
            attendance_dict = {record.date: record.status for record in attendance_records}
            
            # حساب إجمالي الحضور
            total_present = sum(1 for status in attendance_dict.values() if status == 'present')
            total_days = len(attendance_dict)
            
            ws_dept.cell(row=emp_idx, column=8).value = total_days
            ws_dept.cell(row=emp_idx, column=9).value = total_present
            
            # ملء أعمدة الأيام
            for day_idx, date in enumerate(date_list, 10):
                cell = ws_dept.cell(row=emp_idx, column=day_idx)
                
                if date in attendance_dict:
                    status = attendance_dict[date]
                    status_map = {
                        'present': ('P', "D5F4E6"),
                        'absent': ('A', "FADBD8"),
                        'leave': ('L', "FCF3CF"),
                        'sick': ('S', "D6EAF8")
                    }
                    symbol, bg_color = status_map.get(status, ('', "FFFFFF"))
                    cell.value = symbol
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                else:
                    cell.value = ""
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                cell.font = Font(size=11, bold=True)
            
            # تطبيق التنسيق على خلايا معلومات الموظف
            for col in range(1, 10):
                cell = ws_dept.cell(row=emp_idx, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # تلوين الصفوف بالتناوب
                if emp_idx % 2 == 0:
                    if col <= 9:
                        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # ضبط عرض الأعمدة
        ws_dept.column_dimensions['A'].width = 25
        ws_dept.column_dimensions['B'].width = 12
        ws_dept.column_dimensions['C'].width = 15
        ws_dept.column_dimensions['D'].width = 12
        ws_dept.column_dimensions['E'].width = 18
        ws_dept.column_dimensions['F'].width = 12
        ws_dept.column_dimensions['G'].width = 12
        ws_dept.column_dimensions['H'].width = 10
        ws_dept.column_dimensions['I'].width = 10
        
        # ضبط عرض أعمدة الأيام
        for col_idx in range(10, 10 + len(date_list)):
            ws_dept.column_dimensions[get_column_letter(col_idx)].width = 5
    
    # حفظ الملف
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'📊_تقرير_الحضور_التحليلي_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports_bp.route('/salaries')
def salaries_report():
    """تقرير الرواتب"""
    # الحصول على معلمات الفلتر
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    month = int(request.args.get('month', current_month))
    year = int(request.args.get('year', current_year))
    department_id = request.args.get('department_id', '')
    
    # استعلام الرواتب
    salaries_query = Salary.query.filter_by(
        month=month,
        year=year
    )
    
    # الحصول على قائمة الموظفين الذين لديهم رواتب (لعرضها في التقرير)
    employee_ids_with_salary = [s.employee_id for s in salaries_query.all()]
    
    # استعلام لجميع الموظفين (لإظهار الموظفين الذين ليس لديهم رواتب مسجلة)
    employees_query = Employee.query.filter(Employee.status == 'active')
    
    if department_id:
        employees_query = employees_query.filter_by(department_id=department_id)
    
    employees = employees_query.all()
    
    # الحصول على الرواتب النهائية مع التفاصيل
    salaries = []
    for employee in employees:
        salary = Salary.query.filter_by(
            employee_id=employee.id,
            month=month,
            year=year
        ).first()
        
        if salary:
            salaries.append({
                'id': salary.id,
                'employee': employee,
                'basic_salary': salary.basic_salary,
                'allowances': salary.allowances,
                'deductions': salary.deductions,
                'bonus': salary.bonus,
                'net_salary': salary.net_salary,
                'has_salary': True
            })
        else:
            salaries.append({
                'id': None,
                'employee': employee,
                'basic_salary': 0,
                'allowances': 0,
                'deductions': 0,
                'bonus': 0,
                'net_salary': 0,
                'has_salary': False
            })
    
    # حساب الإجماليات
    totals = {
        'basic': sum(s['basic_salary'] for s in salaries if s['has_salary']),
        'allowances': sum(s['allowances'] for s in salaries if s['has_salary']),
        'deductions': sum(s['deductions'] for s in salaries if s['has_salary']),
        'bonus': sum(s['bonus'] for s in salaries if s['has_salary']),
        'net': sum(s['net_salary'] for s in salaries if s['has_salary'])
    }
    
    # الحصول على قائمة الأقسام لعناصر الفلتر
    departments = Department.query.all()
    
    return render_template('reports/salaries.html',
                        salaries=salaries,
                        departments=departments,
                        month=month,
                        year=year,
                        department_id=department_id,
                        totals=totals,
                        month_name=get_month_name_ar(month))

@reports_bp.route('/salaries/pdf')
def salaries_pdf():
    """تصدير تقرير الرواتب إلى PDF"""
    # استخدام توليد التقارير عبر وحدة utils.pdf_generator
    from utils.pdf_generator import generate_salary_report_pdf
    from reportlab.lib.units import cm
    
    # الحصول على معلمات الفلتر
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    month = int(request.args.get('month', current_month))
    year = int(request.args.get('year', current_year))
    department_id = request.args.get('department_id', '')
    
    # استعلام الموظفين النشطين
    employees_query = Employee.query.filter(Employee.status == 'active')
    
    # تطبيق فلتر القسم إذا كان محددًا
    if department_id:
        employees_query = employees_query.filter_by(department_id=department_id)
        department = Department.query.get(department_id)
        department_name = department.name if department else ""
    else:
        department_name = "جميع الأقسام"
    
    employees = employees_query.all()
    
    # الحصول على الرواتب النهائية مع التفاصيل
    salaries = []
    for employee in employees:
        salary = Salary.query.filter_by(
            employee_id=employee.id,
            month=month,
            year=year
        ).first()
        
        if salary:
            salaries.append({
                'id': salary.id,
                'employee': employee,
                'basic_salary': salary.basic_salary,
                'allowances': salary.allowances,
                'deductions': salary.deductions,
                'bonus': salary.bonus,
                'net_salary': salary.net_salary,
                'has_salary': True
            })
    
    # حساب الإجماليات
    totals = {
        'basic': sum(s['basic_salary'] for s in salaries if s['has_salary']),
        'allowances': sum(s['allowances'] for s in salaries if s['has_salary']),
        'deductions': sum(s['deductions'] for s in salaries if s['has_salary']),
        'bonus': sum(s['bonus'] for s in salaries if s['has_salary']),
        'net': sum(s['net_salary'] for s in salaries if s['has_salary'])
    }
    
    # توليد PDF باستخدام وحدة PDF المخصصة
    pdf_file = generate_salary_report_pdf(salaries, month, year, department_name, totals)
    
    # إرسال الملف للمستخدم
    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=f"salaries_report_{month}_{year}.pdf",
        mimetype='application/pdf'
    )

@reports_bp.route('/salaries/pdf')
def salaries_report_pdf():
    """
    إنشاء تقرير PDF شامل للرواتب بناءً على الفلاتر.
    الفلاتر الممكنة: year, month, department_id
    """
    try:
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        department_id = request.args.get('department_id')

        query = Salary.query.join(Employee, Salary.employee_id == Employee.id)\
                            .outerjoin(Department, Employee.department_id == Department.id)

        report_params = {}
        department_name = None

        if year:
            query = query.filter(Salary.year == year)
            report_params["year"] = year
        if month:
            query = query.filter(Salary.month == month)
            report_params["month"] = month
        
        if department_id:
            if department_id.isdigit():
                dept_id_int = int(department_id)
                query = query.filter(Employee.department_id == dept_id_int)
                department = Department.query.get(dept_id_int)
                if department:
                    department_name = department.name
                    report_params["department_name"] = department_name
            elif department_id == "all": #  إذا كان المستخدم يريد جميع الأقسام
                pass # لا تقم بتطبيق فلتر القسم

        salaries = query.order_by(Department.name, Employee.name, Salary.year, Salary.month).all()

        if not salaries:
            flash('لا توجد بيانات رواتب تطابق معايير البحث لإنشاء التقرير.', 'warning')
            # يمكنك توجيهه إلى صفحة التقارير الرئيسية أو صفحة أخرى مناسبة
            return redirect(request.referrer or url_for('dashboard.index'))

        pdf_buffer = generate_salary_report_pdf(salaries, report_params)
        
        filename_parts = ["salary_report"]
        if department_name:
            filename_parts.append(department_name.replace(" ", "_"))
        if month:
            filename_parts.append(str(month))
        if year:
            filename_parts.append(str(year))
        filename_parts.append(datetime.now().strftime("%Y%m%d"))
        
        filename = "_".join(filename_parts) + ".pdf"

        return send_file(pdf_buffer, as_attachment=True, download_name=filename, mimetype='application/pdf') # تأكد من استخدام pdf_buffer هنا

    except Exception as e:
        current_app.logger.error(f"خطأ في إنشاء تقرير PDF للرواتب: {e}", exc_info=True)
        flash(f'حدث خطأ أثناء إنشاء تقرير PDF: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('dashboard.index'))

# يمكنك إضافة مسارات أخرى للتقارير هنا (مثل تقارير Excel أو HTML)



@reports_bp.route('/salaries/excel')
def salaries_excel():
    """تصدير تقرير الرواتب إلى Excel"""
    from openpyxl.utils import get_column_letter
    # الحصول على معلمات الفلتر
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    month = int(request.args.get('month', current_month))
    year = int(request.args.get('year', current_year))
    department_id = request.args.get('department_id', '')
    
    # استعلام الموظفين النشطين
    employees_query = Employee.query.filter(Employee.status == 'active')
    
    # تطبيق فلتر القسم إذا كان محددًا
    if department_id:
        employees_query = employees_query.filter_by(department_id=department_id)
    
    employees = employees_query.all()
    
    # الحصول على الرواتب النهائية مع التفاصيل
    salaries = []
    for employee in employees:
        salary = Salary.query.filter_by(
            employee_id=employee.id,
            month=month,
            year=year
        ).first()
        
        if salary:
            salaries.append(salary)
    
    # إنشاء ملف Excel
    import pandas as pd
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # إنشاء ملف Excel جديد
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "تقرير الرواتب"
    
    # تنسيق اسم المشروع
    sheet.merge_cells('A1:G1')
    company_cell = sheet['A1']
    company_cell.value = "نُظم - نظام إدارة متكامل"
    company_cell.font = Font(size=18, bold=True, name='Tajawal')
    company_cell.alignment = Alignment(horizontal='center')
    
    # تنسيق العنوان
    sheet.merge_cells('A2:G2')
    title_cell = sheet['A2']
    title_cell.value = f"تقرير الرواتب لشهر {get_month_name_ar(month)} {year}"
    title_cell.font = Font(size=16, bold=True, name='Tajawal')
    title_cell.alignment = Alignment(horizontal='center')
    
    # إضافة عناوين الأعمدة
    headers = ["اسم الموظف", "الرقم الوظيفي", "القسم", "الراتب الأساسي", "البدلات", "الاستقطاعات", "المكافآت", "صافي الراتب"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, name='Tajawal')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # تنسيق الحدود
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # إضافة البيانات
    for idx, salary in enumerate(salaries, start=3):
        employee = salary.employee
        sheet.cell(row=idx, column=1).value = employee.name
        sheet.cell(row=idx, column=2).value = employee.employee_id
        
        # القسم
        department_name = employee.department.name if employee.department else "---"
        sheet.cell(row=idx, column=3).value = department_name
        
        # تفاصيل الراتب
        sheet.cell(row=idx, column=4).value = salary.basic_salary
        sheet.cell(row=idx, column=5).value = salary.allowances
        sheet.cell(row=idx, column=6).value = salary.deductions
        sheet.cell(row=idx, column=7).value = salary.bonus
        sheet.cell(row=idx, column=8).value = salary.net_salary
        
        # تطبيق التنسيق على كل خلية
        for col in range(1, 9):
            cell = sheet.cell(row=idx, column=col)
            cell.alignment = Alignment(horizontal='center')
            
            # تنسيق الحدود
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
    
    # إضافة صف الإجماليات
    total_row = len(salaries) + 4
    
    sheet.cell(row=total_row, column=1).value = "الإجمالي"
    sheet.cell(row=total_row, column=1).font = Font(bold=True, name='Tajawal')
    sheet.merge_cells(f'A{total_row}:C{total_row}')
    
    # حساب الإجماليات
    basic_total = sum(s.basic_salary for s in salaries)
    allowances_total = sum(s.allowances for s in salaries)
    deductions_total = sum(s.deductions for s in salaries)
    bonus_total = sum(s.bonus for s in salaries)
    net_total = sum(s.net_salary for s in salaries)
    
    sheet.cell(row=total_row, column=4).value = basic_total
    sheet.cell(row=total_row, column=5).value = allowances_total
    sheet.cell(row=total_row, column=6).value = deductions_total
    sheet.cell(row=total_row, column=7).value = bonus_total
    sheet.cell(row=total_row, column=8).value = net_total
    
    # تنسيق صف الإجماليات
    for col in range(1, 9):
        cell = sheet.cell(row=total_row, column=col)
        cell.font = Font(bold=True, name='Tajawal')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        
        # تنسيق الحدود
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    # ضبط عرض الأعمدة بطريقة آمنة
    # تعيين عرض افتراضي لجميع الأعمدة
    try:
        for i in range(1, sheet.max_column + 1):
            column = get_column_letter(i)
            sheet.column_dimensions[column].width = 15
    except Exception as e:
        print(f"خطأ أثناء ضبط عرض الأعمدة: {str(e)}")
    
    # ضبط اتجاه الورقة للعربية (من اليمين لليسار)
    sheet.sheet_view.rightToLeft = True
    
    # حفظ الملف
    workbook.save(output)
    output.seek(0)
    
    # إنشاء استجابة تحميل
    return send_file(
        output,
        as_attachment=True,
        download_name=f'salaries_report_{month}_{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports_bp.route('/documents')
def documents_report():
    """تقرير الوثائق"""
    # الحصول على معلمات الفلتر
    department_id = request.args.get('department_id', '')
    document_type = request.args.get('document_type', '')
    expiring_only = request.args.get('expiring_only', '') == 'true'
    expiry_days = int(request.args.get('expiry_days', 30))
    
    # تحديد تاريخ الانتهاء للمقارنة في حالة "قريبة من الانتهاء"
    cutoff_date = datetime.now().date() + timedelta(days=expiry_days)
    
    # استعلام الوثائق
    query = db.session.query(
            Document, Employee
        ).join(
            Employee, Document.employee_id == Employee.id
        )
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
    if document_type:
        query = query.filter(Document.document_type == document_type)
    if expiring_only:
        query = query.filter(Document.expiry_date <= cutoff_date)
    
    # الحصول على النتائج النهائية
    results = query.order_by(Document.expiry_date).all()
    
    # الحصول على قائمة الأقسام وأنواع الوثائق لعناصر الفلتر
    departments = Department.query.all()
    document_types = [
        {'id': 'national_id', 'name': 'الهوية الوطنية'},
        {'id': 'passport', 'name': 'جواز السفر'},
        {'id': 'driver_license', 'name': 'رخصة القيادة'},
        {'id': 'annual_leave', 'name': 'الإجازة السنوية'},
        {'id': 'health_certificate', 'name': 'الشهادة الصحية'}
    ]
    
    return render_template('reports/documents.html',
                        results=results,
                        departments=departments,
                        document_types=document_types,
                        department_id=department_id,
                        document_type=document_type,
                        expiring_only=expiring_only,
                        expiry_days=expiry_days,
                        format_date_gregorian=format_date_gregorian,
                        format_date_hijri=format_date_hijri)

@reports_bp.route('/documents/pdf')
def documents_pdf():
    """تصدير تقرير الوثائق إلى PDF"""
    # الحصول على معلمات الفلتر
    department_id = request.args.get('department_id', '')
    document_type = request.args.get('document_type', '')
    expiring_only = request.args.get('expiring_only', '') == 'true'
    expiry_days = int(request.args.get('expiry_days', 30))
    
    # تحديد تاريخ الانتهاء للمقارنة في حالة "قريبة من الانتهاء"
    cutoff_date = datetime.now().date() + timedelta(days=expiry_days)
    
    # استعلام الوثائق
    query = db.session.query(
            Document, Employee
        ).join(
            Employee, Document.employee_id == Employee.id
        )
    
    # تطبيق الفلاتر والحصول على أسماء الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
        department = Department.query.get(department_id)
        department_name = department.name if department else ""
    else:
        department_name = "جميع الأقسام"
    
    if document_type:
        query = query.filter(Document.document_type == document_type)
        document_types_map = {
            'national_id': 'الهوية الوطنية',
            'passport': 'جواز السفر',
            'driver_license': 'رخصة القيادة',
            'annual_leave': 'الإجازة السنوية',
            'health_certificate': 'الشهادة الصحية'
        }
        document_type_name = document_types_map.get(document_type, "")
    else:
        document_type_name = "جميع أنواع الوثائق"
    
    if expiring_only:
        query = query.filter(Document.expiry_date <= cutoff_date)
        expiry_status = f"الوثائق التي ستنتهي خلال {expiry_days} يوم"
    else:
        expiry_status = "جميع الوثائق"
    
    # الحصول على النتائج النهائية
    results = query.order_by(Document.expiry_date).all()
    
    # استخدام المكتبة الموحدة لإنشاء PDF
    from utils.pdf import arabic_text, create_pdf, create_data_table, get_styles
    from reportlab.lib.units import cm
    from reportlab.platypus import Spacer, Paragraph
    from reportlab.lib import colors
    
    # إعداد المحتوى باستخدام وحدة PDF الجديدة
    elements = []
    
    # الحصول على أنماط النصوص المحسنة
    styles = get_styles()
    
    # إضافة العنوان
    title = f"تقرير الوثائق - {document_type_name} - {department_name} - {expiry_status}"
    # تهيئة النص العربي للعرض في PDF باستخدام الدالة المحسنة
    elements.append(Paragraph(arabic_text(title), styles['title']))
    elements.append(Spacer(1, 20))
    
    # إعداد جدول البيانات
    headers = ["الموظف", "الرقم الوظيفي", "القسم", "نوع الوثيقة", "رقم الوثيقة", "تاريخ الإصدار", "تاريخ الانتهاء", "الحالة"]
    data = []
    
    # إضافة الرؤوس (مع تطبيق ترميز النص العربي باستخدام الدالة المحسنة)
    headers_display = [arabic_text(h) for h in headers]
    data.append(headers_display)
    
    # ترجمة أنواع الوثائق
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'driver_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'health_certificate': 'الشهادة الصحية'
    }
    
    # إضافة بيانات الوثائق
    today = datetime.now().date()
    for document, employee in results:
        department_name = employee.department.name if employee.department else "---"
        document_type_arabic = document_types_map.get(document.document_type, document.document_type)
        
        # تحديد حالة الوثيقة (سارية، قاربت الانتهاء، منتهية)
        days_to_expiry = (document.expiry_date - today).days
        if days_to_expiry <= 0:
            status = "منتهية"
            status_color = colors.red
        elif days_to_expiry <= expiry_days:
            status = f"تنتهي خلال {days_to_expiry} يوم"
            status_color = colors.orange
        else:
            status = "سارية"
            status_color = colors.green
        
        row = [
            arabic_text(employee.name),
            employee.employee_id,
            arabic_text(department_name),
            arabic_text(document_type_arabic),
            document.document_number,
            format_date_gregorian(document.issue_date),
            format_date_gregorian(document.expiry_date),
            arabic_text(status)
        ]
        data.append(row)
    
    # إنشاء الجدول باستخدام دالة جدول البيانات المحسنة
    if len(data) > 1:  # لدينا بيانات بخلاف الرؤوس
        # إنشاء جدول البيانات باستخدام الدالة المحسنة
        elements.append(create_data_table(headers, data[1:]))
    else:
        elements.append(Paragraph(arabic_text("لا توجد بيانات وثائق متاحة"), styles['normal']))
    
    # إضافة معلومات التقرير في أسفل الصفحة
    elements.append(Spacer(1, 20))
    footer_text = f"تاريخ إنشاء التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    elements.append(Paragraph(arabic_text(footer_text), styles['normal']))
    
    # إنشاء ملف PDF واستخدام دالة الإنشاء المحسنة
    buffer = create_pdf(elements, landscape_mode=True)
    
    # إنشاء استجابة تحميل
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="documents_report.pdf",
        mimetype='application/pdf'
    )

@reports_bp.route('/documents/excel')
def documents_excel():
    """تصدير تقرير الوثائق إلى Excel"""
    # الحصول على معلمات الفلتر
    department_id = request.args.get('department_id', '')
    document_type = request.args.get('document_type', '')
    expiring_only = request.args.get('expiring_only', '') == 'true'
    expiry_days = int(request.args.get('expiry_days', 30))
    
    # تحديد تاريخ الانتهاء للمقارنة في حالة "قريبة من الانتهاء"
    cutoff_date = datetime.now().date() + timedelta(days=expiry_days)
    
    # استعلام الوثائق
    query = db.session.query(
            Document, Employee
        ).join(
            Employee, Document.employee_id == Employee.id
        )
    
    # تطبيق الفلاتر
    if department_id:
        query = query.filter(Employee.department_id == department_id)
    if document_type:
        query = query.filter(Document.document_type == document_type)
    if expiring_only:
        query = query.filter(Document.expiry_date <= cutoff_date)
    
    # الحصول على النتائج النهائية
    results = query.order_by(Document.expiry_date).all()
    
    # إنشاء كائن Pandas DataFrame
    import pandas as pd
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # إنشاء ملف Excel جديد
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "تقرير الوثائق"
    
    # تحديد عنوان التقرير
    title = "تقرير الوثائق"
    if expiring_only:
        title += f" - الوثائق التي ستنتهي خلال {expiry_days} يوم"
    
    # تنسيق اسم المشروع
    sheet.merge_cells('A1:H1')
    company_cell = sheet['A1']
    company_cell.value = "نُظم - نظام إدارة متكامل"
    company_cell.font = Font(size=18, bold=True, name='Tajawal')
    company_cell.alignment = Alignment(horizontal='center')
    
    # تنسيق العنوان
    sheet.merge_cells('A2:H2')
    title_cell = sheet['A2']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, name='Tajawal')
    title_cell.alignment = Alignment(horizontal='center')
    
    # إضافة عناوين الأعمدة
    headers = ["اسم الموظف", "الرقم الوظيفي", "القسم", "نوع الوثيقة", "رقم الوثيقة", "تاريخ الإصدار", "تاريخ الانتهاء", "الحالة", "ملاحظات"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, name='Tajawal')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # تنسيق الحدود
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # ترجمة أنواع الوثائق
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'driver_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'health_certificate': 'الشهادة الصحية'
    }
    
    # إضافة البيانات
    for idx, (document, employee) in enumerate(results, start=3):
        sheet.cell(row=idx, column=1).value = employee.name
        sheet.cell(row=idx, column=2).value = employee.employee_id
        
        # القسم
        department_name = employee.department.name if employee.department else "---"
        sheet.cell(row=idx, column=3).value = department_name
        
        # تفاصيل الوثيقة
        sheet.cell(row=idx, column=4).value = document_types_map.get(document.document_type, document.document_type)
        sheet.cell(row=idx, column=5).value = document.document_number
        sheet.cell(row=idx, column=6).value = format_date_gregorian(document.issue_date)
        sheet.cell(row=idx, column=7).value = format_date_gregorian(document.expiry_date)
        
        # حالة الوثيقة (سارية، منتهية، قريبة من الانتهاء)
        today = datetime.now().date()
        if document.expiry_date < today:
            status = "منتهية"
            status_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # أحمر فاتح
        elif document.expiry_date <= today + timedelta(days=30):
            status = "تنتهي قريباً"
            status_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # أصفر فاتح
        else:
            status = "سارية"
            status_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # أخضر فاتح
        
        status_cell = sheet.cell(row=idx, column=8)
        status_cell.value = status
        status_cell.fill = status_fill
        
        # ملاحظات
        sheet.cell(row=idx, column=9).value = document.notes if document.notes else ""
        
        # تطبيق التنسيق على كل خلية
        for col in range(1, 10):
            cell = sheet.cell(row=idx, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name='Tajawal')
            
            # تنسيق الحدود
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
    
    # ضبط عرض الأعمدة باستخدام الدالة المساعدة
    from utils.excel_utils import adjust_column_width
    adjust_column_width(sheet)
    
    # ضبط اتجاه الورقة للعربية (من اليمين لليسار)
    sheet.sheet_view.rightToLeft = True
    
    # حفظ الملف
    workbook.save(output)
    output.seek(0)
    
    # إنشاء استجابة تحميل
    return send_file(
        output,
        as_attachment=True,
        download_name=f'documents_report_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )