from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file
from flask_login import login_required, current_user
from models import Geofence, GeofenceEvent, GeofenceSession, GeofenceAttendance, Employee, Department, Attendance, EmployeeLocation, employee_departments
from core.extensions import db
from datetime import datetime, timedelta
from utils.geofence_session_manager import SessionManager
from sqlalchemy import func, desc
import re
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

geofences_bp = Blueprint('geofences', __name__, url_prefix='/employees/geofences')


@geofences_bp.route('/')
@login_required
def index():
    """صفحة إدارة الدوائر الجغرافية"""
    geofences = Geofence.query.filter_by(is_active=True).all()
    departments = Department.query.all()
    
    geofences_data = []
    for geofence in geofences:
        employees_inside = geofence.get_department_employees_inside()
        
        # تحويل الموظفين إلى قواميس قابلة للتحويل إلى JSON
        employees_list = []
        for emp_data in employees_inside:
            emp = emp_data.get('employee') if isinstance(emp_data, dict) else emp_data
            employees_list.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id,
                'profile_image': emp.profile_image if hasattr(emp, 'profile_image') else None
            })
        
        geofences_data.append({
            'geofence': {
                'id': geofence.id,
                'name': geofence.name,
                'type': geofence.type,
                'center_latitude': geofence.center_latitude,
                'center_longitude': geofence.center_longitude,
                'radius_meters': geofence.radius_meters,
                'color': geofence.color,
                'description': geofence.description,
                'department_id': geofence.department_id,
                'department': {
                    'id': geofence.department.id,
                    'name': geofence.department.name
                }
            },
            'employees_count': len(employees_inside),
            'employees_inside': employees_list
        })
    
    return render_template(
        'geofences/index.html',
        geofences_data=geofences_data,
        departments=departments
    )


@geofences_bp.route('/create', methods=['POST'])
@login_required
def create():
    """إنشاء دائرة جديدة"""
    try:
        data = request.get_json()
        
        geofence = Geofence(
            name=data['name'],
            type=data.get('type', 'project'),
            description=data.get('description'),
            center_latitude=data['latitude'],
            center_longitude=data['longitude'],
            radius_meters=data['radius'],
            color=data.get('color', '#667eea'),
            department_id=data['department_id'],
            notify_on_entry=data.get('notify_on_entry', False),
            notify_on_exit=data.get('notify_on_exit', False),
            created_by=current_user.id
        )
        
        db.session.add(geofence)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'geofence_id': geofence.id,
            'message': 'تم إنشاء الدائرة بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في إنشاء الدائرة: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>')
@login_required
def view(geofence_id):
    """عرض تفاصيل دائرة معينة"""
    from collections import defaultdict
    import json
    
    geofence = Geofence.query.get_or_404(geofence_id)
    
    employees_inside = geofence.get_department_employees_inside()
    all_employees = geofence.get_all_employees_inside()
    
    recent_events = GeofenceEvent.query.filter_by(
        geofence_id=geofence_id
    ).order_by(GeofenceEvent.recorded_at.desc()).limit(50).all()
    
    # جلب جميع موظفي القسم للإضافة
    department_employees = Employee.query.join(employee_departments).filter(
        employee_departments.c.department_id == geofence.department_id
    ).all()
    
    # الموظفون المتاحون للربط (غير مرتبطين بالدائرة حالياً)
    assigned_employee_ids = [emp.id for emp in geofence.assigned_employees]
    available_employees = [emp for emp in department_employees if emp.id not in assigned_employee_ids]
    
    # دالة لحساب السرعة وتحديد نمط النقل
    def get_transportation_mode(employee_id):
        """حساب السرعة وتحديد ما إذا كان الموظف يمشي أو يقود"""
        locations = EmployeeLocation.query.filter_by(
            employee_id=employee_id
        ).order_by(EmployeeLocation.recorded_at.desc()).limit(2).all()
        
        if len(locations) < 2:
            return 'unknown'
        
        loc1 = locations[0]
        loc2 = locations[1]
        
        time_diff = (loc1.recorded_at - loc2.recorded_at).total_seconds() / 3600
        if time_diff == 0:
            return 'unknown'
        
        # حساب المسافة بين النقطتين (Haversine formula)
        from math import radians, cos, sin, asin, sqrt
        lon1, lat1, lon2, lat2 = float(loc1.longitude), float(loc1.latitude), float(loc2.longitude), float(loc2.latitude)
        dlon = lon2 - lon1
        dlat = lat2 - lat1
        a = sin(radians(dlat)/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(radians(dlon)/2)**2
        c = 2 * asin(sqrt(a))
        km = 6371 * c
        
        speed = km / time_diff
        return 'driving' if speed > 5 else 'walking'
    
    # جلب الجلسات النشطة (الموظفون داخل الدائرة الآن)
    active_sessions = SessionManager.get_active_sessions(geofence_id=geofence_id)
    
    # حساب حالة الحضور لكل جلسة نشطة وتحويلها إلى قاموس
    active_sessions_data = []
    employees_with_sessions_today = set()
    for session in active_sessions:
        session.attendance_status = geofence.get_attendance_status(session)
        employees_with_sessions_today.add(session.employee_id)
        
        # الحصول على آخر موقع للموظف
        latest_location = EmployeeLocation.query.filter_by(
            employee_id=session.employee_id
        ).order_by(EmployeeLocation.recorded_at.desc()).first()
        
        transportation_mode = get_transportation_mode(session.employee_id)
        
        active_sessions_data.append({
            'id': session.id,
            'employee_id': session.employee_id,
            'employee': {
                'id': session.employee.id,
                'name': session.employee.name,
                'employee_id': session.employee.employee_id
            },
            'entry_time': session.entry_time.isoformat() if session.entry_time else None,
            'exit_time': session.exit_time.isoformat() if session.exit_time else None,
            'duration_minutes': session.duration_minutes,
            'attendance_status': session.attendance_status,
            'transportation_mode': transportation_mode,
            'latitude': float(latest_location.latitude) if latest_location else None,
            'longitude': float(latest_location.longitude) if latest_location else None
        })
    
    # حساب الموظفين الغائبين (المرتبطين بالدائرة لكن لم يدخلوا اليوم)
    absent_employees = [emp for emp in geofence.assigned_employees 
                       if emp.id not in employees_with_sessions_today]
    
    # جلب جميع الجلسات مع تفاصيل الموظفين
    all_sessions = GeofenceSession.query.filter_by(
        geofence_id=geofence_id
    ).options(
        db.joinedload(GeofenceSession.employee)
    ).order_by(GeofenceSession.entry_time.desc()).limit(100).all()
    
    # إضافة أوقات محولة إلى توقيت السعودية (+3 ساعات) لكل جلسة
    for session in all_sessions:
        if session.entry_time:
            session.entry_time_sa = session.entry_time + timedelta(hours=3)
        if session.exit_time:
            session.exit_time_sa = session.exit_time + timedelta(hours=3)
    
    # حساب إحصائيات الموظفين
    employee_stats = {}
    for emp in geofence.assigned_employees:
        total_time = SessionManager.get_employee_total_time(emp.id, geofence_id)
        visit_count = SessionManager.get_employee_visit_count(emp.id, geofence_id)
        is_inside = any(s.employee_id == emp.id for s in active_sessions)
        
        employee_stats[emp.id] = {
            'employee': emp,
            'total_time_minutes': total_time,
            'total_time_hours': round(total_time / 60, 1) if total_time else 0,
            'visit_count': visit_count,
            'avg_duration': round(total_time / visit_count, 1) if visit_count > 0 else 0,
            'is_inside': is_inside
        }
    
    # إحصائيات الحضور
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=7)
    month_start = today_start - timedelta(days=30)
    
    total_assigned = len(geofence.assigned_employees)
    assigned_ids = {emp.id for emp in geofence.assigned_employees}
    present_count = len([s for s in active_sessions if s.employee_id in assigned_ids])
    absent_count = total_assigned - present_count
    attendance_rate = (present_count / total_assigned * 100) if total_assigned > 0 else 0
    
    # إحصائيات متقدمة: تأخر، في الوقت، إلخ
    late_count = 0
    on_time_count = 0
    for session in active_sessions:
        status = geofence.get_attendance_status(session)
        if isinstance(status, str) and status.startswith('late'):
            late_count += 1
        elif status == 'on_time':
            on_time_count += 1
    
    # إحصائيات آخر 24 ساعة (كل ساعة)
    hourly_data = []
    for hour in range(24):
        hour_start = today_start + timedelta(hours=hour)
        hour_end = hour_start + timedelta(hours=1)
        
        events_count = GeofenceEvent.query.filter(
            GeofenceEvent.geofence_id == geofence_id,
            GeofenceEvent.recorded_at >= hour_start,
            GeofenceEvent.recorded_at < hour_end
        ).count()
        
        hourly_data.append({
            'hour': hour,
            'count': events_count
        })
    
    # إحصائيات أسبوعية
    weekly_data = []
    days_ar = ['الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت']
    for day in range(7):
        day_start = week_start + timedelta(days=day)
        day_end = day_start + timedelta(days=1)
        
        events_count = GeofenceEvent.query.filter(
            GeofenceEvent.geofence_id == geofence_id,
            GeofenceEvent.recorded_at >= day_start,
            GeofenceEvent.recorded_at < day_end
        ).count()
        
        weekly_data.append({
            'day': days_ar[day_start.weekday()],
            'count': events_count
        })
    
    # تقرير الانتظامية الشهرية
    monthly_report = {}
    for emp in geofence.assigned_employees:
        emp_sessions = GeofenceSession.query.filter(
            GeofenceSession.geofence_id == geofence_id,
            GeofenceSession.employee_id == emp.id,
            GeofenceSession.entry_time >= month_start
        ).all()
        monthly_report[emp.id] = len(emp_sessions)
    
    # أكثر موظف حضوراً
    top_attendees = sorted(monthly_report.items(), key=lambda x: x[1], reverse=True)[:5]
    top_attendees_data = []
    for emp_id, count in top_attendees:
        emp = next((e for e in geofence.assigned_employees if e.id == emp_id), None)
        if emp:
            top_attendees_data.append({'employee': emp, 'attendance_count': count})
    
    # موظفون متأخرون متكررون
    late_employees = {}
    for emp in geofence.assigned_employees:
        late_sessions = GeofenceSession.query.filter(
            GeofenceSession.geofence_id == geofence_id,
            GeofenceSession.employee_id == emp.id,
            GeofenceSession.entry_time >= month_start
        ).all()
        late_count_emp = 0
        for session in late_sessions:
            status = geofence.get_attendance_status(session)
            if isinstance(status, str) and status.startswith('late'):
                late_count_emp += 1
        if late_count_emp > 0:
            late_employees[emp.id] = late_count_emp
    
    frequent_late = sorted(late_employees.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # جلب سجلات الحضور الصباحي والمسائي
    today_date = datetime.utcnow().date()
    geofence_attendance_records = GeofenceAttendance.query.filter(
        GeofenceAttendance.geofence_id == geofence_id,
        GeofenceAttendance.attendance_date == today_date
    ).all()
    
    # تسجيل الحضور الصباحي والمسائي تلقائياً للموظفين الموجودين
    for session in active_sessions:
        emp_id = session.employee_id
        entry_time_sa = session.entry_time + timedelta(hours=3) if session.entry_time else None
        
        # التحقق إذا كان هناك سجل حضور لهذا الموظف اليوم
        attendance_record = GeofenceAttendance.query.filter(
            GeofenceAttendance.geofence_id == geofence_id,
            GeofenceAttendance.employee_id == emp_id,
            GeofenceAttendance.attendance_date == today_date
        ).first()
        
        if not attendance_record:
            attendance_record = GeofenceAttendance(
                geofence_id=geofence_id,
                employee_id=emp_id,
                attendance_date=today_date
            )
            db.session.add(attendance_record)
        
        # تسجيل الحضور الصباحي أو المسائي حسب الساعة
        if entry_time_sa:
            hour = entry_time_sa.hour
            # الصباح: من 5 صباحاً إلى 1 ظهراً (تم توسيع النطاق ليشمل 12-13)
            if 5 <= hour < 13:  
                if not attendance_record.morning_entry:
                    attendance_record.morning_entry = session.entry_time
                    attendance_record.morning_entry_sa = entry_time_sa
            # المساء: من 1 ظهراً إلى 11 مساءً
            elif 13 <= hour < 23:  
                if not attendance_record.evening_entry:
                    attendance_record.evening_entry = session.entry_time
                    attendance_record.evening_entry_sa = entry_time_sa
    
    db.session.commit()
    
    # إعادة جلب السجلات بعد التحديث وتحويلها إلى قواميس
    geofence_attendance_records = GeofenceAttendance.query.filter(
        GeofenceAttendance.geofence_id == geofence_id,
        GeofenceAttendance.attendance_date == today_date
    ).all()
    
    # تحويل السجلات إلى قواامس لتكون قابلة للتحويل إلى JSON
    attendance_records_json = []
    for record in geofence_attendance_records:
        attendance_records_json.append({
            'id': record.id,
            'employee': {
                'id': record.employee.id,
                'name': record.employee.name,
                'employee_id': record.employee.employee_id
            },
            'morning_entry_sa': record.morning_entry_sa.isoformat() if record.morning_entry_sa else None,
            'evening_entry_sa': record.evening_entry_sa.isoformat() if record.evening_entry_sa else None
        })
    
    stats = {
        'total_assigned': total_assigned,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'on_time_count': on_time_count,
        'attendance_rate': round(attendance_rate, 1),
        'hourly_data': json.dumps(hourly_data),
        'weekly_data': json.dumps(weekly_data)
    }
    
    return render_template(
        'geofences/view.html',
        geofence=geofence,
        employees_inside=employees_inside,
        all_employees=all_employees,
        recent_events=recent_events,
        assigned_employees=geofence.assigned_employees,
        available_employees=available_employees,
        absent_employees=absent_employees,
        stats=stats,
        employee_stats=employee_stats,
        all_sessions=all_sessions,
        active_sessions=active_sessions,
        active_sessions_data=active_sessions_data,
        top_attendees=top_attendees_data,
        frequent_late=frequent_late,
        geofence_attendance_records=geofence_attendance_records,
        attendance_records_json=attendance_records_json,
        geofence_settings={
            'attendance_start_time': geofence.attendance_start_time,
            'attendance_required_minutes': geofence.attendance_required_minutes,
            'early_arrival_minutes': 15,
            'late_arrival_minutes': 0
        }
    )


@geofences_bp.route('/<int:geofence_id>/bulk-check-in', methods=['POST'])
@login_required
def bulk_check_in(geofence_id):
    """تسجيل حضور جماعي فقط لموظفي القسم المرتبط بالدائرة والموجودين داخل دوائرهم المخصصة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        employees_inside = geofence.get_department_employees_inside()
        
        if not employees_inside:
            return jsonify({
                'success': False,
                'message': f'لا يوجد موظفين من قسم "{geofence.department.name}" داخل الدائرة حالياً'
            })
        
        checked_in = []
        already_checked = []
        errors = []
        not_assigned = []
        
        for emp_data in employees_inside:
            employee = emp_data['employee']
            location = emp_data['location']
            
            # التحقق من أن الموظف مرتبط بهذه الدائرة
            if employee not in geofence.assigned_employees:
                not_assigned.append(employee.name)
                continue
            
            today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
            existing_attendance = Attendance.query.filter(
                Attendance.employee_id == employee.id,
                Attendance.check_in_time >= today_start
            ).first()
            
            if existing_attendance:
                already_checked.append(employee.name)
                continue
            
            try:
                attendance = Attendance(
                    employee_id=employee.id,
                    check_in_time=datetime.utcnow(),
                    status='present',
                    notes=f'تسجيل جماعي من دائرة: {geofence.name} (قسم: {geofence.department.name})'
                )
                db.session.add(attendance)
                db.session.flush()
                
                event = GeofenceEvent(
                    geofence_id=geofence.id,
                    employee_id=employee.id,
                    event_type='bulk_check_in',
                    location_latitude=location.latitude,
                    location_longitude=location.longitude,
                    distance_from_center=int(emp_data['distance']),
                    source='bulk',
                    attendance_id=attendance.id,
                    notes=f'تسجيل جماعي بواسطة {current_user.username} - قسم: {geofence.department.name}'
                )
                db.session.add(event)
                
                checked_in.append(employee.name)
                
            except Exception as e:
                errors.append(f'{employee.name}: {str(e)}')
        
        db.session.commit()
        
        message_parts = [f'تم تسجيل حضور {len(checked_in)} موظف']
        if not_assigned:
            message_parts.append(f'({len(not_assigned)} موظف غير مرتبط بهذه الدائرة)')
        if already_checked:
            message_parts.append(f'({len(already_checked)} تم تسجيلهم مسبقاً)')
        
        return jsonify({
            'success': True,
            'department_name': geofence.department.name,
            'checked_in_count': len(checked_in),
            'already_checked_count': len(already_checked),
            'not_assigned_count': len(not_assigned),
            'error_count': len(errors),
            'checked_in': checked_in,
            'already_checked': already_checked,
            'not_assigned': not_assigned,
            'errors': errors,
            'message': ' '.join(message_parts)
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في تسجيل الحضور: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/employees')
@login_required
def get_employees(geofence_id):
    """جلب الموظفين داخل الدائرة (API)"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        employees_inside = geofence.get_department_employees_inside()
        all_employees = geofence.get_all_employees_inside()
        
        return jsonify({
            'success': True,
            'department_employees': [{
                'id': emp['employee'].id,
                'name': emp['employee'].name,
                'employee_id': emp['employee'].employee_id,
                'distance': round(emp['distance'], 2),
                'latitude': float(emp['location'].latitude),
                'longitude': float(emp['location'].longitude),
                'profile_image': emp['employee'].profile_image
            } for emp in employees_inside],
            'other_employees': [{
                'id': emp['employee'].id,
                'name': emp['employee'].name,
                'employee_id': emp['employee'].employee_id,
                'distance': round(emp['distance'], 2),
                'latitude': float(emp['location'].latitude),
                'longitude': float(emp['location'].longitude),
                'is_eligible': emp['is_eligible'],
                'profile_image': emp['employee'].profile_image
            } for emp in all_employees if not emp['is_eligible']]
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/events')
@login_required
def get_events(geofence_id):
    """جلب أحداث دائرة معينة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        events = GeofenceEvent.query.filter_by(
            geofence_id=geofence_id
        ).order_by(GeofenceEvent.recorded_at.desc()).limit(100).all()
        
        return jsonify({
            'success': True,
            'events': [{
                'id': event.id,
                'employee_name': event.employee.name,
                'event_type': event.event_type,
                'event_time': event.recorded_at.isoformat(),
                'distance': event.distance_from_center,
                'notes': event.notes
            } for event in events]
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/api/list')
@login_required
def api_list():
    """API: جلب قائمة جميع الدوائر النشطة"""
    try:
        geofences = Geofence.query.filter_by(is_active=True).all()
        
        return jsonify({
            'success': True,
            'geofences': [{
                'id': g.id,
                'name': g.name,
                'type': g.type,
                'center_lat': float(g.center_latitude),
                'center_lng': float(g.center_longitude),
                'radius': g.radius_meters,
                'color': g.color,
                'department_id': g.department_id,
                'department_name': g.department.name
            } for g in geofences]
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/update-settings', methods=['POST'])
@login_required
def update_settings(geofence_id):
    """تحديث إعدادات الحضور للدائرة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        data = request.get_json()
        
        if 'attendance_start_time' in data:
            geofence.attendance_start_time = data['attendance_start_time']
        if 'attendance_required_minutes' in data:
            geofence.attendance_required_minutes = data['attendance_required_minutes']
        
        geofence.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم تحديث إعدادات الحضور بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/auto-record-attendance', methods=['POST'])
@login_required
def auto_record_attendance(geofence_id):
    """تسجيل حضور تلقائي للموظفين الذين استوفوا الشروط"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        # جلب جميع الجلسات النشطة
        active_sessions = SessionManager.get_active_sessions(geofence_id=geofence_id)
        
        recorded_count = 0
        already_recorded = 0
        insufficient_time = 0
        
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        
        for session in active_sessions:
            # التحقق من أن الموظف مرتبط بهذه الدائرة
            if session.employee not in geofence.assigned_employees:
                continue
            
            # التحقق من أن المدة كافية
            if not session.duration_minutes or session.duration_minutes < geofence.attendance_required_minutes:
                insufficient_time += 1
                continue
            
            # التحقق من عدم وجود حضور مسجل مسبقاً اليوم
            existing_attendance = Attendance.query.filter(
                Attendance.employee_id == session.employee_id,
                Attendance.check_in_time >= today_start
            ).first()
            
            if existing_attendance:
                already_recorded += 1
                continue
            
            # حساب حالة الحضور
            status = geofence.get_attendance_status(session)
            
            # تسجيل الحضور
            attendance = Attendance(
                employee_id=session.employee_id,
                check_in_time=session.entry_time,
                status='present',
                notes=f'تسجيل تلقائي من دائرة: {geofence.name} - الحالة: {status}'
            )
            db.session.add(attendance)
            db.session.flush()
            
            # تسجيل حدث
            event = GeofenceEvent(
                geofence_id=geofence.id,
                employee_id=session.employee_id,
                event_type='auto_attendance',
                location_latitude=geofence.center_latitude,
                location_longitude=geofence.center_longitude,
                source='auto',
                attendance_id=attendance.id,
                notes=f'تسجيل حضور تلقائي - المدة: {session.duration_minutes} دقيقة - الحالة: {status}'
            )
            db.session.add(event)
            
            recorded_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'recorded_count': recorded_count,
            'already_recorded': already_recorded,
            'insufficient_time': insufficient_time,
            'message': f'تم تسجيل حضور {recorded_count} موظف'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/export-entry-data', methods=['GET'])
@login_required
def export_entry_data(geofence_id):
    """تصدير بيانات دخول الدائرة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "دخول الدائرة"
        
        # الرأس
        headers = ['الموظف', 'رقم الموظف', 'وقت الدخول', 'وقت الخروج', 'المدة (دقيقة)', 'القسم', 'الدائرة']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        
        # البيانات
        sessions = GeofenceSession.query.filter_by(geofence_id=geofence_id).order_by(GeofenceSession.entry_time.desc()).limit(500).all()
        
        for session in sessions:
            ws.append([
                session.employee.name,
                session.employee.employee_id,
                session.entry_time.strftime('%Y-%m-%d %H:%M:%S') if session.entry_time else '-',
                session.exit_time.strftime('%Y-%m-%d %H:%M:%S') if session.exit_time else 'في الداخل',
                session.duration_minutes or '-',
                geofence.department.name if geofence.department else '-',
                geofence.name
            ])
        
        for col in range(1, 8):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"دخول_{geofence.name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'خطأ في التصدير: {str(e)}', 'danger')
        return redirect(url_for('geofences.view', geofence_id=geofence_id))


@geofences_bp.route('/<int:geofence_id>/export-daily-attendance', methods=['GET'])
@login_required
def export_daily_attendance(geofence_id):
    """تصدير حضور اليوم"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today_start + timedelta(days=1)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "حضور اليوم"
        
        headers = ['الموظف', 'رقم الموظف', 'وقت الدخول', 'الحالة', 'المدة (دقيقة)', 'الملاحظات']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        
        # جلب الحضور لليوم
        attendances = Attendance.query.filter(
            Attendance.check_in_time >= today_start,
            Attendance.check_in_time < today_end
        ).join(Employee).all()
        
        # جلب الجلسات لليوم
        sessions = GeofenceSession.query.filter(
            GeofenceSession.geofence_id == geofence_id,
            GeofenceSession.entry_time >= today_start,
            GeofenceSession.entry_time < today_end
        ).all()
        
        for session in sessions:
            status = geofence.get_attendance_status(session)
            status_ar = 'في الوقت' if status == 'on_time' else ('متأخر' if status.startswith('late_') else 'موجود')
            
            ws.append([
                session.employee.name,
                session.employee.employee_id,
                session.entry_time.strftime('%H:%M:%S') if session.entry_time else '-',
                status_ar,
                session.duration_minutes or '-',
                geofence.name
            ])
        
        for col in range(1, 7):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"حضور_اليوم_{geofence.name}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'خطأ في التصدير: {str(e)}', 'danger')
        return redirect(url_for('geofences.view', geofence_id=geofence_id))


@geofences_bp.route('/<int:geofence_id>/export-entry-timeline', methods=['GET'])
@login_required
def export_entry_timeline(geofence_id):
    """تصدير الموظفين الذين دخلوا اليوم بحسب الوقت"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today_start + timedelta(days=1)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "الدخول بحسب الوقت"
        
        headers = ['الترتيب', 'الموظف', 'رقم الموظف', 'وقت الدخول', 'القسم', 'الحالة']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        
        # جلب الدخول بترتيب زمني
        sessions = GeofenceSession.query.filter(
            GeofenceSession.geofence_id == geofence_id,
            GeofenceSession.entry_time >= today_start,
            GeofenceSession.entry_time < today_end
        ).order_by(GeofenceSession.entry_time.asc()).all()
        
        for idx, session in enumerate(sessions, 1):
            status = geofence.get_attendance_status(session)
            status_ar = 'في الوقت' if status == 'on_time' else ('متأخر' if status.startswith('late_') else 'موجود')
            
            ws.append([
                idx,
                session.employee.name,
                session.employee.employee_id,
                session.entry_time.strftime('%H:%M:%S') if session.entry_time else '-',
                geofence.department.name if geofence.department else '-',
                status_ar
            ])
        
        for col in range(1, 7):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"دخول_بالوقت_{geofence.name}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'خطأ في التصدير: {str(e)}', 'danger')
        return redirect(url_for('geofences.view', geofence_id=geofence_id))


@geofences_bp.route('/<int:geofence_id>/update', methods=['PUT'])
@login_required
def update(geofence_id):
    """تحديث دائرة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        data = request.get_json()
        
        if 'name' in data:
            geofence.name = data['name']
        if 'description' in data:
            geofence.description = data['description']
        if 'radius' in data:
            geofence.radius_meters = data['radius']
        if 'color' in data:
            geofence.color = data['color']
        if 'notify_on_entry' in data:
            geofence.notify_on_entry = data['notify_on_entry']
        if 'notify_on_exit' in data:
            geofence.notify_on_exit = data['notify_on_exit']
        
        geofence.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم تحديث الدائرة بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في التحديث: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/delete', methods=['POST', 'DELETE'])
@login_required
def delete(geofence_id):
    """حذف دائرة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        db.session.delete(geofence)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم حذف الدائرة بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في الحذف: {str(e)}'
        }), 400


@geofences_bp.route('/extract-google-maps-coords', methods=['POST'])
@login_required
def extract_google_maps_coords():
    """استخراج الإحداثيات من روابط Google Maps (بما في ذلك الروابط المختصرة)"""
    try:
        data = request.get_json()
        url = data.get('url', '').strip()
        
        if not url:
            return jsonify({
                'success': False,
                'message': 'الرجاء إدخال رابط Google Maps'
            }), 400
        
        # إذا كان الرابط مختصر (goo.gl أو maps.app.goo.gl)، نحتاج لفتحه للحصول على الرابط الكامل
        if 'goo.gl' in url or 'maps.app.goo.gl' in url:
            try:
                # إرسال طلب للحصول على الرابط الكامل بعد إعادة التوجيه
                response = requests.get(url, allow_redirects=True, timeout=10)
                url = response.url  # الرابط الكامل بعد إعادة التوجيه
            except Exception as e:
                return jsonify({
                    'success': False,
                    'message': f'فشل فتح الرابط المختصر: {str(e)}'
                }), 400
        
        # استخراج الإحداثيات من الرابط
        coords = None
        
        # نمط 1: !3d/!4d (الأكثر دقة - موقع العلامة)
        match = re.search(r'!3d(-?\d+\.?\d*)!4d(-?\d+\.?\d*)', url)
        if match:
            coords = {
                'lat': float(match.group(1)),
                'lng': float(match.group(2))
            }
        
        # نمط 2: @ (مركز الخريطة)
        if not coords:
            match = re.search(r'@(-?\d+\.\d+),(-?\d+\.\d+)', url)
            if match:
                coords = {
                    'lat': float(match.group(1)),
                    'lng': float(match.group(2))
                }
        
        # نمط 3: ?q= أو ll= (روابط قصيرة)
        if not coords:
            match = re.search(r'[?&](q|ll)=(-?\d+\.?\d*),(-?\d+\.?\d*)', url)
            if match:
                coords = {
                    'lat': float(match.group(2)),
                    'lng': float(match.group(3))
                }
        
        if coords:
            return jsonify({
                'success': True,
                'coords': coords,
                'full_url': url
            })
        else:
            return jsonify({
                'success': False,
                'message': 'لم يتم العثور على إحداثيات في الرابط'
            }), 400
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/assign-employees', methods=['POST'])
@login_required
def assign_employees(geofence_id):
    """ربط موظفين بدائرة جغرافية محددة"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        data = request.get_json()
        employee_ids = data.get('employee_ids', [])
        
        if not employee_ids:
            return jsonify({
                'success': False,
                'message': 'الرجاء اختيار موظف واحد على الأقل'
            }), 400
        
        # إضافة الموظفين إلى الدائرة
        for employee_id in employee_ids:
            employee = Employee.query.get(employee_id)
            if employee and employee not in geofence.assigned_employees:
                geofence.assigned_employees.append(employee)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم ربط {len(employee_ids)} موظف بالدائرة بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في ربط الموظفين: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/unassign-employee/<int:employee_id>', methods=['POST'])
@login_required
def unassign_employee(geofence_id, employee_id):
    """إلغاء ربط موظف من دائرة جغرافية"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        employee = Employee.query.get_or_404(employee_id)
        
        if employee in geofence.assigned_employees:
            geofence.assigned_employees.remove(employee)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'تم إلغاء ربط الموظف من الدائرة'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'الموظف غير مرتبط بهذه الدائرة'
            }), 400
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/update', methods=['POST'])
@login_required
def update_geofence(geofence_id):
    """تحديث بيانات الدائرة الجغرافية"""
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        data = request.get_json()
        
        name = data.get('name', '').strip()
        radius_meters = data.get('radius_meters')
        color = data.get('color', '').strip()
        
        if not name:
            return jsonify({
                'success': False,
                'message': 'الرجاء إدخال اسم الدائرة'
            }), 400
        
        if not radius_meters or radius_meters < 10:
            return jsonify({
                'success': False,
                'message': 'نصف القطر يجب أن يكون 10 متر على الأقل'
            }), 400
        
        if not color or not color.startswith('#'):
            return jsonify({
                'success': False,
                'message': 'الرجاء اختيار لون صحيح'
            }), 400
        
        geofence.name = name
        geofence.radius_meters = radius_meters
        geofence.color = color
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم تحديث الدائرة بنجاح'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في التحديث: {str(e)}'
        }), 400


@geofences_bp.route('/<int:geofence_id>/export-events')
@login_required
def export_events(geofence_id):
    """تصدير بيانات الوصول والمغادرة لموظفي الدائرة"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "سجل الحضور والمغادرة"
        
        ws.right_to_left = True
        
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['الحالة', 'نوع الحدث', 'وقت الحضور', 'وقت الخروج', 'رقم الموظف', 'اسم الموظف', 'القسم', 'الدائرة']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        employees_inside = geofence.get_department_employees_inside()
        inside_employee_ids = {emp['employee'].id for emp in employees_inside}
        
        all_assigned_employees = geofence.assigned_employees
        
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        
        row_num = 2
        for employee in all_assigned_employees:
            is_inside = employee.id in inside_employee_ids
            
            entry_event = GeofenceEvent.query.filter_by(
                geofence_id=geofence_id,
                employee_id=employee.id,
                event_type='entry'
            ).filter(
                GeofenceEvent.recorded_at >= today_start
            ).order_by(GeofenceEvent.recorded_at.asc()).first()
            
            bulk_entry_event = GeofenceEvent.query.filter_by(
                geofence_id=geofence_id,
                employee_id=employee.id,
                event_type='bulk_check_in'
            ).filter(
                GeofenceEvent.recorded_at >= today_start
            ).order_by(GeofenceEvent.recorded_at.asc()).first()
            
            exit_event = GeofenceEvent.query.filter_by(
                geofence_id=geofence_id,
                employee_id=employee.id,
                event_type='exit'
            ).filter(
                GeofenceEvent.recorded_at >= today_start
            ).order_by(GeofenceEvent.recorded_at.desc()).first()
            
            first_entry = entry_event if entry_event else bulk_entry_event
            if entry_event and bulk_entry_event:
                first_entry = entry_event if entry_event.recorded_at < bulk_entry_event.recorded_at else bulk_entry_event
            
            entry_time = first_entry.recorded_at.strftime('%H:%M:%S') if first_entry else '-'
            exit_time = exit_event.recorded_at.strftime('%H:%M:%S') if exit_event else '-'
            
            if is_inside:
                status = 'موجود داخل الدائرة'
                event_type = 'حضور'
            else:
                if first_entry:
                    status = 'خارج الدائرة'
                    event_type = 'غادر'
                else:
                    status = 'خارج الحضور'
                    event_type = 'غائب'
            
            ws.cell(row=row_num, column=1, value=status)
            ws.cell(row=row_num, column=2, value=event_type)
            ws.cell(row=row_num, column=3, value=entry_time)
            ws.cell(row=row_num, column=4, value=exit_time)
            ws.cell(row=row_num, column=5, value=employee.employee_id)
            ws.cell(row=row_num, column=6, value=employee.name)
            ws.cell(row=row_num, column=7, value=geofence.department.name if geofence.department else '-')
            ws.cell(row=row_num, column=8, value=geofence.name)
            
            for col in range(1, 9):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if col == 1:
                    if is_inside:
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                        cell.font = Font(color="065F46", bold=True)
                    elif status == 'خارج الحضور':
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = Font(color="991B1B", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                        cell.font = Font(color="92400E", bold=True)
            
            row_num += 1
        
        for col in range(1, 9):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"تقرير_الحضور_{geofence.name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        flash(f'خطأ في تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('geofences.view', geofence_id=geofence_id))


@geofences_bp.route('/<int:geofence_id>/export-attendance', methods=['GET'])
@login_required
def export_attendance(geofence_id):
    """تصدير سجلات الحضور إلى Excel بنفس تنسيق export_events"""
    geofence = Geofence.query.get_or_404(geofence_id)
    
    export_date_str = request.args.get('date')
    if export_date_str:
        export_date = datetime.strptime(export_date_str, '%Y-%m-%d').date()
    else:
        sa_now = datetime.utcnow() + timedelta(hours=3)
        export_date = sa_now.date()
    
    all_employees = geofence.assigned_employees
    
    attendance_records = {}
    records = GeofenceAttendance.query.filter(
        GeofenceAttendance.geofence_id == geofence_id,
        GeofenceAttendance.attendance_date == export_date
    ).all()
    for rec in records:
        attendance_records[rec.employee_id] = rec
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'تقرير الحضور'
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['الحالة', 'نوع الحدث', 'وقت الخروج', 'وقت الدخول', 'رقم الموظف', 'اسم الموظف', 'القسم', 'الدائرة']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row_num = 2
    for idx, employee in enumerate(all_employees, 1):
        rec = attendance_records.get(employee.id)
        
        if rec:
            morning_time = rec.morning_entry_sa.strftime('%H:%M:%S') if rec.morning_entry_sa else '-'
            evening_time = rec.evening_entry_sa.strftime('%H:%M:%S') if rec.evening_entry_sa else '-'
            
            if rec.morning_entry_sa and rec.evening_entry_sa:
                status = 'موظف بدوام كامل'
                event_type = 'حاضر'
            elif rec.morning_entry_sa:
                status = 'موظف بدوام صباحي'
                event_type = 'حاضر'
            elif rec.evening_entry_sa:
                status = 'موظف بدوام مسائي'
                event_type = 'حاضر'
            else:
                status = 'موظف غائب'
                event_type = 'غائب'
        else:
            morning_time = '-'
            evening_time = '-'
            status = 'موظف غائب'
            event_type = 'غائب'
        
        ws.cell(row=row_num, column=1, value=status)
        ws.cell(row=row_num, column=2, value=event_type)
        ws.cell(row=row_num, column=3, value=evening_time)
        ws.cell(row=row_num, column=4, value=morning_time)
        ws.cell(row=row_num, column=5, value=employee.employee_id)
        ws.cell(row=row_num, column=6, value=employee.name)
        ws.cell(row=row_num, column=7, value=geofence.department.name if geofence.department else '-')
        ws.cell(row=row_num, column=8, value=geofence.name)
        
        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if col == 1:
                if event_type == 'حاضر':
                    cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                    cell.font = Font(color='065F46', bold=True)
                else:
                    cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                    cell.font = Font(color='991B1B', bold=True)
        
        row_num += 1
    
    for col in range(1, 9):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"تقرير_الحضور_{geofence.name}_{export_date}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@geofences_bp.route('/<int:geofence_id>/export-sessions', methods=['GET'])
@login_required
def export_sessions(geofence_id):
    """تصدير جلسات الموظفين (وقت الوصول والمغادرة) إلى Excel احترافي"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    try:
        geofence = Geofence.query.get_or_404(geofence_id)
        
        all_sessions = GeofenceSession.query.filter_by(
            geofence_id=geofence_id
        ).options(
            db.joinedload(GeofenceSession.employee)
        ).order_by(GeofenceSession.entry_time.desc()).limit(500).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "جلسات الموظفين"
        ws.sheet_view.rightToLeft = True
        
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        title_font = Font(name='Arial', size=16, bold=True, color="4F46E5")
        border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        active_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        active_font = Font(name='Arial', color="065F46", bold=True)
        inactive_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        inactive_font = Font(name='Arial', color="6B7280", bold=True)
        
        ws.merge_cells('A1:H1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"📊 جلسات الموظفين - {geofence.name}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35
        
        ws.merge_cells('A2:H2')
        subtitle_cell = ws.cell(row=2, column=1)
        sa_now = datetime.utcnow() + timedelta(hours=3)
        subtitle_cell.value = f"تاريخ التصدير: {sa_now.strftime('%Y-%m-%d %H:%M')} | إجمالي الجلسات: {len(all_sessions)}"
        subtitle_cell.font = Font(name='Arial', size=10, color="6B7280")
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 25
        
        headers = ['#', 'رقم الموظف', 'اسم الموظف', 'تاريخ الجلسة', 'وقت الدخول', 'وقت الخروج', 'المدة', 'الحالة']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[4].height = 30
        
        row_num = 5
        for idx, session in enumerate(all_sessions, 1):
            entry_time_sa = session.entry_time + timedelta(hours=3) if session.entry_time else None
            exit_time_sa = session.exit_time + timedelta(hours=3) if session.exit_time else None
            
            session_date = entry_time_sa.strftime('%Y-%m-%d') if entry_time_sa else '-'
            entry_time_str = entry_time_sa.strftime('%I:%M %p') if entry_time_sa else '-'
            exit_time_str = exit_time_sa.strftime('%I:%M %p') if exit_time_sa else 'لا يزال بالداخل'
            
            if session.duration_minutes:
                if session.duration_minutes < 60:
                    duration_str = f"{session.duration_minutes} دقيقة"
                else:
                    hours = session.duration_minutes // 60
                    mins = session.duration_minutes % 60
                    if mins > 0:
                        duration_str = f"{hours} ساعة و {mins} دقيقة"
                    else:
                        duration_str = f"{hours} ساعة"
            else:
                duration_str = 'جاري الحساب...'
            
            is_active = session.exit_time is None
            status_text = '🟢 داخل الدائرة' if is_active else '⚪ غادر'
            
            row_data = [
                idx,
                session.employee.employee_id if session.employee else '-',
                session.employee.name if session.employee else '-',
                session_date,
                entry_time_str,
                exit_time_str,
                duration_str,
                status_text
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Arial', size=11)
                
                if col_num == 8:
                    if is_active:
                        cell.fill = active_fill
                        cell.font = active_font
                    else:
                        cell.fill = inactive_fill
                        cell.font = inactive_font
                
                if row_num % 2 == 0 and col_num != 8:
                    cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            
            ws.row_dimensions[row_num].height = 28
            row_num += 1
        
        column_widths = [8, 15, 25, 15, 15, 18, 20, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        stats_row = row_num + 2
        ws.merge_cells(f'A{stats_row}:H{stats_row}')
        stats_cell = ws.cell(row=stats_row, column=1)
        
        active_count = sum(1 for s in all_sessions if s.exit_time is None)
        total_duration = sum(s.duration_minutes or 0 for s in all_sessions)
        avg_duration = total_duration // len(all_sessions) if all_sessions else 0
        
        stats_cell.value = f"📈 إحصائيات: داخل الدائرة حالياً: {active_count} | إجمالي المدة: {total_duration // 60} ساعة | متوسط المدة: {avg_duration} دقيقة"
        stats_cell.font = Font(name='Arial', size=11, bold=True, color="4F46E5")
        stats_cell.alignment = Alignment(horizontal='center', vertical='center')
        stats_cell.fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
        ws.row_dimensions[stats_row].height = 30
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"جلسات_الموظفين_{geofence.name}_{sa_now.strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        flash(f'خطأ في تصدير الجلسات: {str(e)}', 'danger')
        return redirect(url_for('geofences.view', geofence_id=geofence_id))
