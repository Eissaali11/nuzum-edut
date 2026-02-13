"""
API Endpoints الخارجية - بدون مصادقة
تستخدم للتطبيقات الخارجية مثل تطبيق الأندرويد لتتبع المواقع
محسّنة للأداء مع Rate Limiting و Caching
"""
from flask import Blueprint, request, jsonify
from datetime import datetime, timedelta
from models import (
    Employee, EmployeeLocation, Geofence, GeofenceEvent, GeofenceSession, employee_departments, 
    VehicleHandover, db, Attendance, Salary, EmployeeRequest, EmployeeLiability,
    Document, MobileDevice, SimCard, Department, Vehicle
)
from sqlalchemy import func, and_, or_, extract
from sqlalchemy.orm import joinedload
from datetime import date
import os
import logging
from utils.geofence_session_manager import SessionManager
from time import time

# إنشاء Blueprint
api_external_bp = Blueprint('api_external', __name__, url_prefix='/api/external')

# مفتاح API الثابت (محفوظ في متغير بيئة)
LOCATION_API_KEY = os.environ.get('LOCATION_API_KEY', 'test_location_key_2025')

# إعداد السجلات
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================
# Rate Limiting و Caching
# ============================================
# تخزين آخر موقع وآخر طلب لكل موظف
last_employee_location = {}  # {employee_id: {'lat': x, 'lng': y, 'time': timestamp}}
last_saved_location = {}  # {employee_id: timestamp} - آخر وقت تم فيه حفظ موقع حقيقي
rate_limit_tracker = {}  # {employee_id: [timestamps_of_requests]}

RATE_LIMIT_REQUESTS_PER_SECOND = 5
RATE_LIMIT_WINDOW_SECONDS = 1
MIN_DISTANCE_METERS = 100  # لا تسجل الموقع إذا لم يتغير أكثر من 100 متر
MIN_TIME_BETWEEN_SAVES = 300  # 5 دقائق - الحد الأدنى بين حفظ المواقع المتتالية


# ============================================
# دوال Rate Limiting
# ============================================
def check_rate_limit(employee_id):
    """التحقق من Rate Limit للموظف"""
    current_time = time()
    
    if employee_id not in rate_limit_tracker:
        rate_limit_tracker[employee_id] = []
    
    # حذف الطلبات القديمة خارج النافذة الزمنية
    rate_limit_tracker[employee_id] = [
        t for t in rate_limit_tracker[employee_id] 
        if current_time - t < RATE_LIMIT_WINDOW_SECONDS
    ]
    
    # التحقق من عدد الطلبات
    if len(rate_limit_tracker[employee_id]) >= RATE_LIMIT_REQUESTS_PER_SECOND:
        return False, "تم تجاوز حد الطلبات المسموح به"
    
    # إضافة الطلب الحالي
    rate_limit_tracker[employee_id].append(current_time)
    return True, None


def calculate_distance(lat1, lon1, lat2, lon2):
    """حساب المسافة بين نقطتين بالكيلومتر"""
    from math import radians, sin, cos, sqrt, atan2
    
    R = 6371  # نصف قطر الأرض بالكيلومتر
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    distance = R * c
    
    return distance * 1000  # تحويل لمتر


def is_location_changed(employee_id, latitude, longitude):
    """التحقق مما إذا تغير الموقع بشكل كافي"""
    if employee_id not in last_employee_location:
        return True
    
    last_loc = last_employee_location[employee_id]
    distance = calculate_distance(
        last_loc['lat'], last_loc['lng'],
        latitude, longitude
    )
    
    return distance >= MIN_DISTANCE_METERS


def check_time_since_last_save(employee_id):
    """التحقق من الوقت المنقضي منذ آخر حفظ موقع (كل 5 دقائق)"""
    current_time = time()
    
    if employee_id not in last_saved_location:
        return True  # أول طلب - اقبله
    
    time_elapsed = current_time - last_saved_location[employee_id]
    return time_elapsed >= MIN_TIME_BETWEEN_SAVES


def update_last_saved_time(employee_id):
    """تحديث آخر وقت تم فيه حفظ موقع"""
    last_saved_location[employee_id] = time()


def update_location_cache(employee_id, latitude, longitude):
    """تحديث الموقع المخزن مؤقتاً"""
    last_employee_location[employee_id] = {
        'lat': latitude,
        'lng': longitude,
        'time': time()
    }


def process_geofence_events(employee, latitude, longitude):
    """
    معالجة أحداث الدوائر الجغرافية عند استلام موقع جديد
    يكتشف تلقائياً دخول/خروج الموظف من جميع الدوائر (بغض النظر عن القسم)
    """
    try:
        # جلب جميع الدوائر النشطة (بدون تصفية حسب القسم)
        # هذا يضمن تسجيل جميع الدخولات والخروجات للأمان والمراقبة
        active_geofences = Geofence.query.filter(
            Geofence.is_active == True
        ).all()
        
        for geofence in active_geofences:
            # حساب المسافة من مركز الدائرة
            distance = geofence.calculate_distance(latitude, longitude)
            is_inside = distance <= geofence.radius_meters
            
            # جلب آخر حدث للموظف في هذه الدائرة
            last_event = GeofenceEvent.query.filter_by(
                geofence_id=geofence.id,
                employee_id=employee.id
            ).order_by(GeofenceEvent.recorded_at.desc()).first()
            
            # تحديد نوع الحدث
            event_type = None
            
            if is_inside:
                # داخل الدائرة
                if not last_event or last_event.event_type == 'exit':
                    # دخول جديد
                    event_type = 'enter'
                    logger.info(f"🟢 دخول: {employee.name} دخل دائرة {geofence.name}")
            else:
                # خارج الدائرة
                if last_event and last_event.event_type == 'enter':
                    # خروج جديد
                    event_type = 'exit'
                    logger.info(f"🔴 خروج: {employee.name} خرج من دائرة {geofence.name}")
            
            # تسجيل الحدث
            if event_type:
                event = GeofenceEvent(
                    geofence_id=geofence.id,
                    employee_id=employee.id,
                    event_type=event_type,
                    location_latitude=latitude,
                    location_longitude=longitude,
                    distance_from_center=int(distance),
                    source='auto',
                    notes=f'كشف تلقائي من نظام تتبع المواقع'
                )
                db.session.add(event)
                db.session.flush()  # للحصول على event.id
                
                # إنشاء/تحديث جلسة باستخدام SessionManager
                try:
                    if event_type == 'enter':
                        SessionManager.process_enter_event(employee.id, geofence.id, event)
                    elif event_type == 'exit':
                        SessionManager.process_exit_event(employee.id, geofence.id, event)
                except Exception as e:
                    logger.error(f"خطأ في معالجة جلسة الموظف: {str(e)}")
                
                # إرسال إشعار (اختياري) - يمكن تفعيله لاحقاً
                if (event_type == 'enter' and geofence.notify_on_entry) or \
                   (event_type == 'exit' and geofence.notify_on_exit):
                    # TODO: إضافة إشعارات (SendGrid أو Twilio)
                    logger.info(f"📧 يجب إرسال إشعار لـ {event_type} في {geofence.name}")
        
        db.session.commit()
        
    except Exception as e:
        logger.error(f"خطأ في معالجة أحداث الدوائر الجغرافية: {str(e)}")
        db.session.rollback()


@api_external_bp.route('/employee-location', methods=['POST'])
def receive_employee_location():
    """
    استقبال موقع الموظف من تطبيق الأندرويد (محسّنة للأداء)
    مع Rate Limiting و Caching تلقائي
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'لا توجد بيانات'}), 400
        
        # التحقق من مفتاح API
        if data.get('api_key') != LOCATION_API_KEY:
            return jsonify({'success': False, 'error': 'مفتاح API غير صحيح'}), 401
        
        job_number = data.get('job_number')
        if not job_number:
            return jsonify({'success': False, 'error': 'الرقم الوظيفي مطلوب'}), 400
        
        # البحث عن الموظف
        employee = Employee.query.filter_by(employee_id=job_number).first()
        if not employee:
            return jsonify({'success': False, 'error': 'موظف غير موجود'}), 404
        
        # التحقق من Rate Limit
        allowed, error_msg = check_rate_limit(employee.id)
        if not allowed:
            return jsonify({'success': False, 'error': error_msg}), 429
        
        # التحقق من صحة الإحداثيات
        try:
            lat = float(data.get('latitude'))
            lng = float(data.get('longitude'))
            
            if not (-90 <= lat <= 90) or not (-180 <= lng <= 180):
                return jsonify({'success': False, 'error': 'إحداثيات غير صحيحة'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'إحداثيات غير صحيحة'}), 400
        
        # 🔥 تحسين الأداء: تخطي الموقع إذا لم يتغير بشكل كافي
        if not is_location_changed(employee.id, lat, lng):
            # تحديث الـ cache لكن بدون حفظ في قاعدة البيانات
            update_location_cache(employee.id, lat, lng)
            logger.info(f"📍 CACHED (no distance): {employee.name} ({job_number})")
            return jsonify({
                'success': True,
                'message': 'الموقع لم يتغير (cached)',
                'cached': True
            }), 200
        
        # ⏱️ التحقق من الفاصل الزمني (5 دقائق بين كل حفظ)
        if not check_time_since_last_save(employee.id):
            time_elapsed = time() - last_saved_location.get(employee.id, 0)
            minutes_remaining = (MIN_TIME_BETWEEN_SAVES - time_elapsed) / 60
            # تحديث الـ cache فقط
            update_location_cache(employee.id, lat, lng)
            logger.info(f"⏳ Throttled: {employee.name} ({job_number}) - انتظر {minutes_remaining:.1f} دقيقة")
            return jsonify({
                'success': True,
                'message': f'يجب الانتظار {minutes_remaining:.1f} دقيقة قبل التسجيل التالي',
                'wait_minutes': minutes_remaining,
                'throttled': True
            }), 200
        
        # تحديث الموقع المخزن مؤقتاً
        update_location_cache(employee.id, lat, lng)
        update_last_saved_time(employee.id)
        logger.info(f"✅ SAVED (5-min interval): {employee.name} ({job_number}) - lat: {lat:.4f}, lng: {lng:.4f}")
        
        # تحليل وقت التسجيل
        recorded_at = datetime.utcnow()
        if data.get('recorded_at'):
            try:
                recorded_at = datetime.fromisoformat(data['recorded_at'].replace('Z', '+00:00'))
            except:
                pass
        
        # إنشاء وحفظ سجل الموقع
        location = EmployeeLocation(
            employee_id=employee.id,
            latitude=lat,
            longitude=lng,
            accuracy_m=float(data.get('accuracy')) if data.get('accuracy') else None,
            source='android_app',
            recorded_at=recorded_at,
            received_at=datetime.utcnow(),
            notes=data.get('notes', '')
        )
        
        db.session.add(location)
        db.session.flush()
        
        # ✅ معالجة الدوائر الجغرافية - تسجيل الدخول والخروج
        try:
            process_geofence_events(employee, lat, lng)
        except Exception as e:
            logger.warning(f"تحذير في معالجة الدوائر الجغرافية: {str(e)}")
        
        db.session.commit()
        
        logger.info(f"✅ موقع: {employee.name} ({job_number})")
        
        return jsonify({
            'success': True,
            'message': 'تم حفظ الموقع',
            'data': {
                'employee_name': employee.name,
                'location_id': location.id
            }
        }), 200
        
    except Exception as e:
        logger.error(f"خطأ: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'error': 'خطأ في الخادم'}), 500


@api_external_bp.route('/test', methods=['GET'])
def test_api():
    """نقطة اختبار بسيطة للتأكد من عمل API"""
    return jsonify({
        'success': True,
        'message': 'External API is working!',
        'endpoints': {
            'employee_location': '/api/external/employee-location [POST]',
            'employee_complete_profile': '/api/external/employee-complete-profile [POST]'
        }
    }), 200


def parse_date_filters(data):
    """تحليل فلاتر التاريخ من الطلب"""
    from datetime import datetime
    
    month = data.get('month')  # YYYY-MM format
    start_date = data.get('start_date')  # YYYY-MM-DD
    end_date = data.get('end_date')  # YYYY-MM-DD
    
    # إذا تم إرسال month، استخدمه وتجاهل start/end
    if month:
        try:
            year, month_num = map(int, month.split('-'))
            # أول يوم في الشهر
            start = datetime(year, month_num, 1).date()
            # آخر يوم في الشهر
            import calendar
            last_day = calendar.monthrange(year, month_num)[1]
            end = datetime(year, month_num, last_day).date()
            return start, end
        except (ValueError, AttributeError):
            raise ValueError("تنسيق month غير صحيح. يجب أن يكون YYYY-MM")
    
    # إذا تم إرسال start_date أو end_date
    if start_date or end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
            end = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
            return start, end
        except ValueError:
            raise ValueError("تنسيق التاريخ غير صحيح. يجب أن يكون YYYY-MM-DD")
    
    # افتراضياً: آخر 30 يوم للحضور، آخر 12 شهر للرواتب
    return None, None


def get_employee_data(employee, request_origin=None):
    """جلب معلومات الموظف الكاملة"""
    # جلب القسم الأول
    department = employee.departments[0] if employee.departments else None
    
    # بناء روابط الصور الكاملة
    def build_image_url(image_path):
        if not image_path:
            return None
        if image_path.startswith('http'):
            return image_path
        # استخدام request_origin إذا توفر، وإلا استخدام رابط افتراضي
        if request_origin:
            return f"{request_origin}/static/uploads/{image_path}"
        return f"/static/uploads/{image_path}"
    
    return {
        'job_number': employee.employee_id,
        'name': employee.name,
        'name_en': None,  # غير متوفر في النموذج
        'national_id': employee.national_id,
        'birth_date': employee.birth_date.strftime('%Y-%m-%d') if employee.birth_date else None,
        'hire_date': employee.join_date.strftime('%Y-%m-%d') if employee.join_date else None,
        'nationality': employee.nationality,
        'residence_expiry_date': None,  # يمكن إضافته لاحقاً من Documents
        'sponsor_name': employee.current_sponsor_name,
        'absher_phone': employee.mobilePersonal,
        'department': department.name if department else None,
        'department_en': None,
        'section': None,  # غير متوفر
        'section_en': None,
        'position': employee.job_title,
        'position_en': None,
        'phone': employee.mobile,
        'email': employee.email,
        'address': employee.residence_details,
        'is_driver': employee.employee_type == 'driver',
        'photos': {
            'personal': build_image_url(employee.profile_image),
            'id': build_image_url(employee.national_id_image),
            'license': build_image_url(employee.license_image) if employee.employee_type == 'driver' else None
        }
    }


def get_vehicle_assignments(employee_id):
    """جلب السيارة الحالية والسيارات السابقة للموظف"""
    from models import Vehicle
    
    # جلب جميع عمليات التسليم والاستلام للموظف
    handovers = VehicleHandover.query.filter_by(
        employee_id=employee_id
    ).order_by(
        VehicleHandover.handover_date.desc(),
        VehicleHandover.handover_time.desc()
    ).all()
    
    current_car = None
    previous_cars = []
    processed_vehicles = set()
    
    # بناء map للتسليمات والاستلامات لكل سيارة
    vehicle_operations = {}
    for h in handovers:
        if h.vehicle_id not in vehicle_operations:
            vehicle_operations[h.vehicle_id] = []
        vehicle_operations[h.vehicle_id].append(h)
    
    # معالجة كل سيارة
    for vehicle_id, ops in vehicle_operations.items():
        # ترتيب العمليات حسب التاريخ (الأحدث أولاً)
        ops.sort(key=lambda x: (x.handover_date, x.handover_time or datetime.min.time()), reverse=True)
        
        latest_op = ops[0]
        vehicle = Vehicle.query.get(vehicle_id)
        
        if not vehicle:
            continue
        
        vehicle_data = {
            'car_id': str(vehicle.id),
            'plate_number': vehicle.plate_number,
            'plate_number_en': None,
            'model': f"{vehicle.make} {vehicle.model}",
            'model_en': None,
            'color': vehicle.color,
            'color_en': None,
            'status': vehicle.status,
            'assigned_date': latest_op.handover_date.isoformat() if latest_op.handover_date else None,
            'photo': None,  # يمكن إضافته لاحقاً
            'notes': vehicle.notes
        }
        
        # السيارة الحالية: آخر عملية هي تسليم ولم يتم استلامها بعد
        if latest_op.handover_type == 'delivery' and vehicle_id not in processed_vehicles:
            current_car = vehicle_data.copy()
            current_car.pop('unassigned_date', None)  # السيارة الحالية ليس لها unassigned_date
            processed_vehicles.add(vehicle_id)
        else:
            # السيارات السابقة
            if vehicle_id not in processed_vehicles:
                # البحث عن آخر استلام
                last_receipt = next((op for op in ops if op.handover_type == 'receipt'), None)
                vehicle_data['unassigned_date'] = last_receipt.handover_date.isoformat() if last_receipt and last_receipt.handover_date else None
                previous_cars.append(vehicle_data)
                processed_vehicles.add(vehicle_id)
    
    return current_car, previous_cars


def get_attendance_records(employee_id, start_date, end_date):
    """جلب سجلات الحضور للموظف مع فلترة التواريخ"""
    from models import Attendance as AttendanceModel
    from datetime import datetime, timedelta
    
    query = AttendanceModel.query.filter_by(employee_id=employee_id)
    
    # تطبيق الفلترة
    if start_date:
        query = query.filter(AttendanceModel.date >= start_date)
    if end_date:
        query = query.filter(AttendanceModel.date <= end_date)
    
    # إذا لم يتم تحديد فلترة، جلب آخر 30 يوم
    if not start_date and not end_date:
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=30)
        query = query.filter(AttendanceModel.date >= start_date, AttendanceModel.date <= end_date)
    
    records = query.order_by(AttendanceModel.date.desc()).all()
    
    attendance_list = []
    for att in records:
        # حساب الساعات
        hours_worked = 0.0
        if att.check_in and att.check_out:
            check_in_dt = datetime.combine(att.date, att.check_in)
            check_out_dt = datetime.combine(att.date, att.check_out)
            hours_worked = (check_out_dt - check_in_dt).total_seconds() / 3600
        
        attendance_list.append({
            'date': att.date.strftime('%Y-%m-%d'),
            'check_in': att.check_in.strftime('%H:%M') if att.check_in else None,
            'check_out': att.check_out.strftime('%H:%M') if att.check_out else None,
            'status': att.status,
            'hours_worked': round(hours_worked, 2),
            'late_minutes': 0,  # يمكن حسابه لاحقاً
            'early_leave_minutes': 0,  # يمكن حسابه لاحقاً
            'notes': att.notes
        })
    
    return attendance_list


def get_salary_records(employee_id, start_date, end_date):
    """جلب سجلات الرواتب للموظف مع فلترة التواريخ"""
    from models import Salary as SalaryModel
    from datetime import datetime, date as date_cls
    from dateutil.relativedelta import relativedelta
    
    query = SalaryModel.query.filter_by(employee_id=employee_id)
    
    # تطبيق الفلترة حسب الشهر والسنة
    if start_date:
        query = query.filter(
            db.or_(
                SalaryModel.year > start_date.year,
                db.and_(
                    SalaryModel.year == start_date.year,
                    SalaryModel.month >= start_date.month
                )
            )
        )
    if end_date:
        query = query.filter(
            db.or_(
                SalaryModel.year < end_date.year,
                db.and_(
                    SalaryModel.year == end_date.year,
                    SalaryModel.month <= end_date.month
                )
            )
        )
    
    # إذا لم يتم تحديد فلترة، جلب آخر 12 شهر
    if not start_date and not end_date:
        end_date = datetime.now().date()
        start_date = end_date - relativedelta(months=12)
        query = query.filter(
            db.or_(
                SalaryModel.year > start_date.year,
                db.and_(
                    SalaryModel.year == start_date.year,
                    SalaryModel.month >= start_date.month
                )
            )
        )
    
    records = query.order_by(SalaryModel.year.desc(), SalaryModel.month.desc()).all()
    
    salary_list = []
    for sal in records:
        salary_list.append({
            'salary_id': f"SAL-{sal.year}-{sal.month:02d}",
            'month': f"{sal.year}-{sal.month:02d}",
            'amount': float(sal.net_salary),
            'currency': 'SAR',
            'paid_date': sal.created_at.isoformat() if sal.is_paid and sal.created_at else None,
            'status': 'paid' if sal.is_paid else 'pending',
            'details': {
                'base_salary': float(sal.basic_salary),
                'allowances': float(sal.allowances),
                'deductions': float(sal.deductions),
                'bonuses': float(sal.bonus),
                'overtime': float(sal.overtime_hours * (sal.basic_salary / 30 / 8)) if sal.overtime_hours else 0.0,  # تقدير تقريبي
                'tax': 0.0  # لا توجد ضرائب في السعودية
            },
            'notes': sal.notes
        })
    
    return salary_list


def get_operations_records(employee_id):
    """جلب سجلات العمليات (التسليم/الاستلام) للموظف"""
    from models import Vehicle
    
    # جلب جميع عمليات التسليم والاستلام
    handovers = VehicleHandover.query.filter_by(
        employee_id=employee_id
    ).order_by(VehicleHandover.handover_date.desc()).all()
    
    operations = []
    for h in handovers:
        vehicle = Vehicle.query.get(h.vehicle_id)
        
        operations.append({
            'operation_id': f"OP-{h.id}",
            'type': 'delivery' if h.handover_type == 'delivery' else 'pickup',
            'date': f"{h.handover_date.isoformat()}T{h.handover_time.isoformat() if h.handover_time else '00:00:00'}",
            'car_id': str(h.vehicle_id),
            'car_plate_number': vehicle.plate_number if vehicle else None,
            'client_name': h.supervisor_name or h.person_name,
            'client_phone': h.supervisor_phone_number,
            'address': h.city or h.project_name or '',
            'status': 'completed',  # جميع العمليات المسجلة تعتبر مكتملة
            'notes': h.notes
        })
    
    return operations


def calculate_statistics(attendance, salaries, current_car, previous_cars, operations):
    """حساب الإحصائيات الشاملة"""
    # إحصائيات الحضور
    total_days = len(attendance)
    present_days = len([a for a in attendance if a['status'] in ['present', 'late', 'early_leave']])
    absent_days = len([a for a in attendance if a['status'] == 'absent'])
    late_days = len([a for a in attendance if a['status'] == 'late'])
    early_leave_days = len([a for a in attendance if a['status'] == 'early_leave'])
    total_hours = sum([a['hours_worked'] for a in attendance])
    attendance_rate = round((present_days / total_days * 100) if total_days > 0 else 0.0, 2)
    
    # إحصائيات الرواتب
    total_salaries = len(salaries)
    total_amount = sum([s['amount'] for s in salaries])
    average_amount = round(total_amount / total_salaries if total_salaries > 0 else 0.0, 2)
    last_salary = salaries[0]['amount'] if salaries else 0.0
    last_paid_date = salaries[0]['paid_date'] if salaries and salaries[0]['paid_date'] else None
    
    # إحصائيات السيارات
    all_cars = previous_cars + ([current_car] if current_car else [])
    total_cars = len(all_cars)
    active_cars = len([c for c in all_cars if c['status'] == 'available'])
    maintenance_cars = len([c for c in all_cars if c['status'] == 'in_workshop'])
    retired_cars = len([c for c in all_cars if c['status'] == 'out_of_service'])
    
    # إحصائيات العمليات
    total_operations = len(operations)
    delivery_count = len([o for o in operations if o['type'] == 'delivery'])
    pickup_count = len([o for o in operations if o['type'] == 'pickup'])
    completed_count = len([o for o in operations if o['status'] == 'completed'])
    
    return {
        'attendance': {
            'total_days': total_days,
            'present_days': present_days,
            'absent_days': absent_days,
            'late_days': late_days,
            'early_leave_days': early_leave_days,
            'total_hours': round(total_hours, 2),
            'attendance_rate': attendance_rate
        },
        'salaries': {
            'total_salaries': total_salaries,
            'total_amount': round(total_amount, 2),
            'average_amount': average_amount,
            'last_salary': last_salary,
            'last_paid_date': last_paid_date
        },
        'cars': {
            'current_car': current_car is not None,
            'total_cars': total_cars,
            'active_cars': active_cars,
            'maintenance_cars': maintenance_cars,
            'retired_cars': retired_cars
        },
        'operations': {
            'total_operations': total_operations,
            'delivery_count': delivery_count,
            'pickup_count': pickup_count,
            'completed_count': completed_count
        }
    }


@api_external_bp.route('/employee-complete-profile', methods=['POST'])
def get_employee_complete_profile():
    """
    جلب الملف الشامل للموظف
    يتضمن جميع المعلومات: الموظف، السيارات، الحضور، الرواتب، العمليات، الإحصائيات
    """
    try:
        # الحصول على البيانات
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'طلب فارغ',
                'error': 'No data provided'
            }), 400
        
        # التحقق من مفتاح API
        api_key = data.get('api_key')
        if not api_key or api_key != LOCATION_API_KEY:
            logger.warning(f"محاولة وصول بمفتاح خاطئ إلى employee-complete-profile من {request.remote_addr}")
            return jsonify({
                'success': False,
                'message': 'غير مصرح. يرجى التحقق من المفتاح',
                'error': 'Invalid API key'
            }), 401
        
        # التحقق من job_number
        job_number = data.get('job_number')
        if not job_number:
            return jsonify({
                'success': False,
                'message': 'طلب غير صحيح',
                'error': 'Missing required field: job_number'
            }), 400
        
        # البحث عن الموظف
        employee = Employee.query.filter_by(employee_id=job_number).first()
        
        if not employee:
            logger.warning(f"موظف غير موجود: {job_number}")
            return jsonify({
                'success': False,
                'message': 'الموظف غير موجود',
                'error': 'Employee not found'
            }), 404
        
        # تحليل فلاتر التواريخ
        try:
            start_date, end_date = parse_date_filters(data)
        except ValueError as e:
            return jsonify({
                'success': False,
                'message': 'طلب غير صحيح',
                'error': str(e)
            }), 400
        
        # جلب معلومات الموظف
        request_origin = request.host_url.rstrip('/')
        employee_data = get_employee_data(employee, request_origin)
        
        # جلب السيارات
        current_car, previous_cars = get_vehicle_assignments(employee.id)
        
        # جلب الحضور
        attendance = get_attendance_records(employee.id, start_date, end_date)
        
        # جلب الرواتب
        salaries = get_salary_records(employee.id, start_date, end_date)
        
        # جلب العمليات
        operations = get_operations_records(employee.id)
        
        # حساب الإحصائيات
        statistics = calculate_statistics(attendance, salaries, current_car, previous_cars, operations)
        
        # بناء الاستجابة
        response_data = {
            'employee': employee_data,
            'current_car': current_car,
            'previous_cars': previous_cars,
            'attendance': attendance,
            'salaries': salaries,
            'operations': operations,
            'statistics': statistics
        }
        
        logger.info(f"✅ تم جلب الملف الشامل للموظف {employee.name} ({job_number})")
        
        return jsonify({
            'success': True,
            'message': 'تم جلب البيانات بنجاح',
            'data': response_data
        }), 200
        
    except Exception as e:
        logger.error(f"خطأ في جلب الملف الشامل للموظف: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': 'خطأ في السيرفر',
            'error': 'Internal server error'
        }), 500


@api_external_bp.route('/verify-employee/<employee_id>/<national_id>', methods=['GET'])
def verify_employee(employee_id, national_id):
    """
    التحقق من وجود الموظف باستخدام رقم الموظف الوظيفي ورقم الهوية
    
    نقطة نهاية بسيطة للطرف الثالث للتحقق من هوية الموظف
    لا تحتاج إلى مصادقة
    
    استخدام:
    GET /api/external/verify-employee/EMP001/1234567890
    
    استجابة:
    {
        "exists": true
    }
    أو
    {
        "exists": false
    }
    """
    try:
        # البحث عن الموظف باستخدام الرقم الوظيفي ورقم الهوية
        employee = Employee.query.filter_by(
            employee_id=employee_id,
            national_id=national_id
        ).first()
        
        if employee:
            logger.info(f"✅ تحقق ناجح: الموظف {employee.name} ({employee_id}) موجود")
            return jsonify({'exists': True}), 200
        else:
            logger.info(f"❌ تحقق فاشل: لا يوجد موظف بالرقم الوظيفي {employee_id} ورقم الهوية {national_id}")
            return jsonify({'exists': False}), 200
            
    except Exception as e:
        logger.error(f"خطأ في التحقق من الموظف: {str(e)}")
        return jsonify({
            'exists': False,
            'error': 'حدث خطأ في الخادم'
        }), 500


# ═══════════════════════════════════════════════════════════════════════
# دوال مساعدة لنقطة النهاية الشاملة
# ═══════════════════════════════════════════════════════════════════════

def safe_get(obj, attr, default=None):
    """جلب آمن للبيانات مع معالجة القيم الفارغة"""
    try:
        value = getattr(obj, attr, default)
        return value if value is not None else default
    except:
        return default


def build_full_url(path):
    """بناء رابط كامل للملفات"""
    if not path:
        return None
    if path.startswith('http'):
        return path
    base_url = request.host_url.rstrip('/')
    return f"{base_url}/{path.lstrip('/')}"


def format_date(date_obj):
    """تنسيق التاريخ بشكل آمن"""
    if not date_obj:
        return None
    try:
        return date_obj.strftime('%Y-%m-%d') if hasattr(date_obj, 'strftime') else str(date_obj)
    except:
        return None


def format_datetime(datetime_obj):
    """تنسيق التاريخ والوقت بشكل آمن"""
    if not datetime_obj:
        return None
    try:
        return datetime_obj.strftime('%Y-%m-%d %H:%M:%S') if hasattr(datetime_obj, 'strftime') else str(datetime_obj)
    except:
        return None


@api_external_bp.route('/employee-profile/<employee_id>', methods=['GET'])
def get_complete_employee_profile(employee_id):
    """
    جلب الملف الشخصي الكامل للموظف
    
    نقطة نهاية شاملة تعيد جميع بيانات الموظف في JSON واحد:
    - البيانات الأساسية
    - الحضور والغياب
    - السيارات (الحالية + التاريخ)
    - الرواتب
    - الطلبات
    - الذمم المالية
    - الوثائق
    - الأجهزة والعهدة
    - الإحصائيات الشاملة
    
    استخدام:
    GET /api/external/employee-profile/5216?api_key=YOUR_API_KEY
    
    يحتاج إلى API key للوصول
    """
    try:
        # التحقق من مفتاح API
        api_key = request.args.get('api_key')
        if not api_key or api_key != LOCATION_API_KEY:
            logger.warning(f"❌ محاولة وصول غير مصرح بها إلى employee-profile من {request.remote_addr}")
            return jsonify({
                'success': False,
                'message': 'غير مصرح. يرجى التحقق من المفتاح',
                'error': 'Invalid or missing API key'
            }), 401
        
        # البحث عن الموظف باستخدام الرقم الوظيفي
        employee = Employee.query.filter_by(employee_id=employee_id).first()
        
        if not employee:
            logger.warning(f"❌ موظف غير موجود: {employee_id}")
            return jsonify({
                'success': False,
                'message': 'الموظف غير موجود',
                'error': 'Employee not found'
            }), 404
        
        logger.info(f"🔍 جلب الملف الشامل للموظف: {employee.name} ({employee_id})")
        
        # ═══════════════════════════════════════════════════════════════
        # القسم 1: البيانات الأساسية للموظف
        # ═══════════════════════════════════════════════════════════════
        
        # جلب القسم
        department = None
        if employee.departments:
            dept = employee.departments[0]
            department = safe_get(dept, 'name')
        
        employee_data = {
            'job_number': safe_get(employee, 'employee_id'),
            'name': safe_get(employee, 'name', 'غير محدد'),
            'national_id': safe_get(employee, 'national_id'),
            'birth_date': format_date(safe_get(employee, 'birth_date')),
            'nationality': safe_get(employee, 'nationality'),
            'contract_type': safe_get(employee, 'contract_type'),
            'employee_type': safe_get(employee, 'employee_type', 'regular'),
            'status': safe_get(employee, 'status', 'active'),
            'join_date': format_date(safe_get(employee, 'join_date')),
            'contract_status': safe_get(employee, 'contract_status'),
            'license_status': safe_get(employee, 'license_status'),
            
            'contact': {
                'mobile': safe_get(employee, 'mobile'),
                'mobile_personal': safe_get(employee, 'mobilePersonal'),
                'email': safe_get(employee, 'email')
            },
            
            'work': {
                'department': department,
                'job_title': safe_get(employee, 'job_title'),
                'location': safe_get(employee, 'location'),
                'project': safe_get(employee, 'project')
            },
            
            'salary_info': {
                'basic_salary': safe_get(employee, 'basic_salary', 0.0),
                'daily_wage': safe_get(employee, 'daily_wage', 0.0),
                'attendance_bonus': safe_get(employee, 'attendance_bonus', 0.0)
            },
            
            'images': {
                'profile': build_full_url(safe_get(employee, 'profile_image')),
                'national_id': build_full_url(safe_get(employee, 'national_id_image')),
                'license': build_full_url(safe_get(employee, 'license_image'))
            },
            
            'sponsor': {
                'current_sponsor_name': safe_get(employee, 'current_sponsor_name'),
                'sponsor_id': safe_get(employee, 'sponsor_id'),
                'sponsor_mobile': safe_get(employee, 'sponsor_mobile')
            },
            
            'custody': {
                'has_mobile_custody': safe_get(employee, 'has_mobile_custody', False),
                'mobile_type': safe_get(employee, 'mobile_type'),
                'mobile_imei': safe_get(employee, 'mobile_imei')
            }
        }
        
        # ═══════════════════════════════════════════════════════════════
        # القسم 2: الحضور والغياب
        # ═══════════════════════════════════════════════════════════════
        
        attendance_data = {
            'summary': {
                'total_days': 0,
                'present_days': 0,
                'absent_days': 0,
                'late_days': 0,
                'attendance_rate': 0.0,
                'total_hours_worked': 0.0,
                'average_hours_per_day': 0.0
            },
            'recent_records': [],
            'last_30_days': {
                'present': 0,
                'absent': 0,
                'late': 0
            }
        }
        
        try:
            # جلب سجلات الحضور لآخر 60 يوم
            sixty_days_ago = date.today() - timedelta(days=60)
            thirty_days_ago = date.today() - timedelta(days=30)
            
            attendance_records = Attendance.query.filter(
                Attendance.employee_id == employee.id,
                Attendance.date >= sixty_days_ago
            ).order_by(Attendance.date.desc()).all()
            
            if attendance_records:
                total_days = len(attendance_records)
                present_days = sum(1 for a in attendance_records if safe_get(a, 'status') == 'present')
                absent_days = sum(1 for a in attendance_records if safe_get(a, 'status') == 'absent')
                late_days = sum(1 for a in attendance_records if safe_get(a, 'is_late', False))
                
                total_hours = sum(safe_get(a, 'hours_worked', 0.0) for a in attendance_records)
                avg_hours = total_hours / total_days if total_days > 0 else 0.0
                attendance_rate = (present_days / total_days * 100) if total_days > 0 else 0.0
                
                attendance_data['summary'] = {
                    'total_days': total_days,
                    'present_days': present_days,
                    'absent_days': absent_days,
                    'late_days': late_days,
                    'attendance_rate': round(attendance_rate, 2),
                    'total_hours_worked': round(total_hours, 1),
                    'average_hours_per_day': round(avg_hours, 1)
                }
                
                # آخر 5 سجلات
                attendance_data['recent_records'] = [
                    {
                        'date': format_date(safe_get(a, 'date')),
                        'check_in': a.check_in.strftime('%H:%M') if a.check_in else None,
                        'check_out': a.check_out.strftime('%H:%M') if a.check_out else None,
                        'status': safe_get(a, 'status'),
                        'hours_worked': safe_get(a, 'hours_worked', 0.0),
                        'overtime_hours': safe_get(a, 'overtime_hours', 0.0),
                        'is_late': safe_get(a, 'is_late', False),
                        'notes': safe_get(a, 'notes')
                    }
                    for a in attendance_records[:5]
                ]
                
                # إحصائيات آخر 30 يوم
                last_30_records = [a for a in attendance_records if a.date and a.date >= thirty_days_ago]
                attendance_data['last_30_days'] = {
                    'present': sum(1 for a in last_30_records if safe_get(a, 'status') == 'present'),
                    'absent': sum(1 for a in last_30_records if safe_get(a, 'status') == 'absent'),
                    'late': sum(1 for a in last_30_records if safe_get(a, 'is_late', False))
                }
        except Exception as e:
            logger.warning(f"خطأ في جلب بيانات الحضور: {str(e)}")
        
        # ═══════════════════════════════════════════════════════════════
        # القسم 3: السيارات
        # ═══════════════════════════════════════════════════════════════
        
        vehicles_data = {
            'current': None,
            'history': [],
            'total_vehicles_used': 0
        }
        
        try:
            # جلب جميع السيارات مع التحميل المسبق (لتجنب N+1)
            all_handovers = VehicleHandover.query.options(
                joinedload(VehicleHandover.vehicle)
            ).filter(
                VehicleHandover.employee_id == employee.id
            ).order_by(VehicleHandover.created_at.desc()).all()
            
            vehicles_data['total_vehicles_used'] = len(all_handovers)
            
            # جلب السيارة الحالية (آخر تسليم بدون استلام)
            current_handover = None
            for handover in all_handovers:
                if handover.handover_type == 'delivery' and handover.vehicle_id:
                    current_handover = handover
                    break
            
            if current_handover and current_handover.vehicle:
                vehicle = current_handover.vehicle
                vehicles_data['current'] = {
                    'id': safe_get(vehicle, 'id'),
                    'plate_number': safe_get(vehicle, 'plate_number'),
                    'make': safe_get(vehicle, 'make'),
                    'model': safe_get(vehicle, 'model'),
                    'year': safe_get(vehicle, 'year'),
                    'color': safe_get(vehicle, 'color'),
                    'status': safe_get(vehicle, 'status'),
                    'assigned_date': format_date(safe_get(current_handover, 'created_at')),
                    'handover_type': safe_get(current_handover, 'handover_type')
                }
            
            # بناء تاريخ السيارات (آخر 5، مستثنياً السيارة الحالية)
            history_count = 0
            for handover in all_handovers:
                if history_count >= 5:
                    break
                if handover.vehicle and handover != current_handover:
                    vehicles_data['history'].append({
                        'plate_number': safe_get(handover.vehicle, 'plate_number'),
                        'make': safe_get(handover.vehicle, 'make'),
                        'model': safe_get(handover.vehicle, 'model'),
                        'assigned_date': format_date(safe_get(handover, 'created_at')),
                        'handover_type': safe_get(handover, 'handover_type')
                    })
                    history_count += 1
        except Exception as e:
            logger.warning(f"خطأ في جلب بيانات السيارات: {str(e)}")
        
        # ═══════════════════════════════════════════════════════════════
        # القسم 4: الرواتب
        # ═══════════════════════════════════════════════════════════════
        
        salaries_data = {
            'summary': {
                'total_months': 0,
                'total_paid': 0.0,
                'average_monthly': 0.0,
                'last_salary': None
            },
            'recent_records': [],
            'yearly_total': 0.0
        }
        
        try:
            # جلب الرواتب لآخر 12 شهر
            current_year = date.today().year
            
            salary_records = Salary.query.filter(
                Salary.employee_id == employee.id
            ).order_by(Salary.year.desc(), Salary.month.desc()).limit(12).all()
            
            if salary_records:
                total_paid = sum(safe_get(s, 'net_salary', 0.0) for s in salary_records)
                total_months = len(salary_records)
                avg_monthly = total_paid / total_months if total_months > 0 else 0.0
                
                # آخر راتب
                last_salary_record = salary_records[0]
                
                salaries_data['summary'] = {
                    'total_months': total_months,
                    'total_paid': round(total_paid, 2),
                    'average_monthly': round(avg_monthly, 2),
                    'last_salary': {
                        'month': f"{safe_get(last_salary_record, 'year')}-{safe_get(last_salary_record, 'month'):02d}",
                        'amount': safe_get(last_salary_record, 'net_salary', 0.0),
                        'paid_date': format_date(safe_get(last_salary_record, 'payment_date'))
                    } if last_salary_record else None
                }
                
                # آخر 6 سجلات
                salaries_data['recent_records'] = [
                    {
                        'month': safe_get(s, 'month'),
                        'year': safe_get(s, 'year'),
                        'basic_salary': safe_get(s, 'basic_salary', 0.0),
                        'allowances': safe_get(s, 'allowances', 0.0),
                        'deductions': safe_get(s, 'deductions', 0.0),
                        'overtime_pay': safe_get(s, 'overtime_pay', 0.0),
                        'bonus': safe_get(s, 'bonus', 0.0),
                        'net_salary': safe_get(s, 'net_salary', 0.0),
                        'is_paid': safe_get(s, 'is_paid', False),
                        'payment_date': format_date(safe_get(s, 'payment_date')),
                        'notes': safe_get(s, 'notes')
                    }
                    for s in salary_records[:6]
                ]
                
                # مجموع السنة الحالية
                yearly_records = [s for s in salary_records if safe_get(s, 'year') == current_year]
                salaries_data['yearly_total'] = round(sum(safe_get(s, 'net_salary', 0.0) for s in yearly_records), 2)
        except Exception as e:
            logger.warning(f"خطأ في جلب بيانات الرواتب: {str(e)}")
        
        # ═══════════════════════════════════════════════════════════════
        # القسم 5: الطلبات
        # ═══════════════════════════════════════════════════════════════
        
        requests_data = {
            'summary': {
                'total': 0,
                'pending': 0,
                'approved': 0,
                'rejected': 0
            },
            'by_type': {},
            'recent': []
        }
        
        try:
            employee_requests = EmployeeRequest.query.filter(
                EmployeeRequest.employee_id == employee.id
            ).order_by(EmployeeRequest.created_at.desc()).all()
            
            if employee_requests:
                requests_data['summary']['total'] = len(employee_requests)
                requests_data['summary']['pending'] = sum(1 for r in employee_requests if safe_get(r, 'status') == 'pending')
                requests_data['summary']['approved'] = sum(1 for r in employee_requests if safe_get(r, 'status') == 'approved')
                requests_data['summary']['rejected'] = sum(1 for r in employee_requests if safe_get(r, 'status') == 'rejected')
                
                # تصنيف حسب النوع
                for req in employee_requests:
                    req_type = str(safe_get(req, 'request_type', 'unknown')).lower()
                    requests_data['by_type'][req_type] = requests_data['by_type'].get(req_type, 0) + 1
                
                # آخر 10 طلبات
                requests_data['recent'] = [
                    {
                        'id': safe_get(r, 'id'),
                        'type': str(safe_get(r, 'request_type')),
                        'status': str(safe_get(r, 'status')),
                        'amount': safe_get(r, 'amount', 0.0),
                        'created_at': format_datetime(safe_get(r, 'created_at')),
                        'notes': safe_get(r, 'notes')
                    }
                    for r in employee_requests[:10]
                ]
        except Exception as e:
            logger.warning(f"خطأ في جلب بيانات الطلبات: {str(e)}")
        
        # ═══════════════════════════════════════════════════════════════
        # القسم 6: الذمم المالية
        # ═══════════════════════════════════════════════════════════════
        
        liabilities_data = {
            'summary': {
                'total_active': 0,
                'total_amount': 0.0,
                'total_paid': 0.0,
                'remaining': 0.0
            },
            'records': []
        }
        
        try:
            liabilities = EmployeeLiability.query.filter(
                EmployeeLiability.employee_id == employee.id
            ).order_by(EmployeeLiability.created_at.desc()).all()
            
            if liabilities:
                active_liabilities = [l for l in liabilities if str(safe_get(l, 'status')).lower() == 'active']
                liabilities_data['summary']['total_active'] = len(active_liabilities)
                liabilities_data['summary']['total_amount'] = sum(safe_get(l, 'amount', 0.0) for l in active_liabilities)
                liabilities_data['summary']['total_paid'] = sum(safe_get(l, 'paid_amount', 0.0) for l in active_liabilities)
                liabilities_data['summary']['remaining'] = liabilities_data['summary']['total_amount'] - liabilities_data['summary']['total_paid']
                
                # جميع السجلات
                liabilities_data['records'] = [
                    {
                        'id': safe_get(l, 'id'),
                        'type': str(safe_get(l, 'liability_type')),
                        'amount': safe_get(l, 'amount', 0.0),
                        'paid_amount': safe_get(l, 'paid_amount', 0.0),
                        'remaining': safe_get(l, 'amount', 0.0) - safe_get(l, 'paid_amount', 0.0),
                        'status': str(safe_get(l, 'status')),
                        'description': safe_get(l, 'description'),
                        'due_date': format_date(safe_get(l, 'due_date')),
                        'created_at': format_date(safe_get(l, 'created_at'))
                    }
                    for l in liabilities[:10]
                ]
        except Exception as e:
            logger.warning(f"خطأ في جلب بيانات الذمم: {str(e)}")
        
        # ═══════════════════════════════════════════════════════════════
        # القسم 7: الوثائق
        # ═══════════════════════════════════════════════════════════════
        
        documents_data = {
            'total': 0,
            'valid': 0,
            'expiring_soon': 0,
            'expired': 0,
            'records': []
        }
        
        try:
            documents = Document.query.filter(
                Document.employee_id == employee.id
            ).order_by(Document.created_at.desc()).all()
            
            if documents:
                today = date.today()
                thirty_days_later = today + timedelta(days=30)
                
                documents_data['total'] = len(documents)
                
                for doc in documents:
                    expiry_date = safe_get(doc, 'expiry_date')
                    if expiry_date:
                        if expiry_date < today:
                            documents_data['expired'] += 1
                        elif expiry_date <= thirty_days_later:
                            documents_data['expiring_soon'] += 1
                        else:
                            documents_data['valid'] += 1
                    else:
                        documents_data['valid'] += 1
                
                # جميع الوثائق
                documents_data['records'] = [
                    {
                        'id': safe_get(doc, 'id'),
                        'type': safe_get(doc, 'document_type'),
                        'number': safe_get(doc, 'document_number'),
                        'issue_date': format_date(safe_get(doc, 'issue_date')),
                        'expiry_date': format_date(safe_get(doc, 'expiry_date')),
                        'status': 'expired' if (doc.expiry_date and doc.expiry_date < today) 
                                  else 'expiring_soon' if (doc.expiry_date and doc.expiry_date <= thirty_days_later)
                                  else 'valid',
                        'days_remaining': (doc.expiry_date - today).days if doc.expiry_date else None,
                        'file_url': build_full_url(safe_get(doc, 'file_path'))
                    }
                    for doc in documents[:10]
                ]
        except Exception as e:
            logger.warning(f"خطأ في جلب بيانات الوثائق: {str(e)}")
        
        # ═══════════════════════════════════════════════════════════════
        # القسم 8: الأجهزة والعهدة
        # ═══════════════════════════════════════════════════════════════
        
        devices_data = {
            'mobile_devices': [],
            'sim_cards': []
        }
        
        try:
            mobile_devices = MobileDevice.query.filter(
                MobileDevice.employee_id == employee.id
            ).all()
            
            devices_data['mobile_devices'] = [
                {
                    'brand': safe_get(device, 'brand'),
                    'model': safe_get(device, 'model'),
                    'imei': safe_get(device, 'imei'),
                    'phone_number': safe_get(device, 'phone_number'),
                    'status': safe_get(device, 'status'),
                    'assigned_date': format_date(safe_get(device, 'created_at'))
                }
                for device in mobile_devices
            ]
            
            sim_cards = SimCard.query.filter(
                SimCard.employee_id == employee.id
            ).all()
            
            devices_data['sim_cards'] = [
                {
                    'phone_number': safe_get(sim, 'phone_number'),
                    'carrier': safe_get(sim, 'carrier'),
                    'status': safe_get(sim, 'status'),
                    'data_plan': safe_get(sim, 'data_plan')
                }
                for sim in sim_cards
            ]
        except Exception as e:
            logger.warning(f"خطأ في جلب بيانات الأجهزة: {str(e)}")
        
        # ═══════════════════════════════════════════════════════════════
        # القسم 9: الإحصائيات الشاملة
        # ═══════════════════════════════════════════════════════════════
        
        statistics_data = {
            'performance': {
                'attendance_rate': attendance_data['summary']['attendance_rate'],
                'punctuality_rate': 100.0 - (attendance_data['summary']['late_days'] / attendance_data['summary']['total_days'] * 100) if attendance_data['summary']['total_days'] > 0 else 100.0,
                'average_working_hours': attendance_data['summary']['average_hours_per_day']
            },
            'financial': {
                'total_earnings_ytd': salaries_data['yearly_total'],
                'average_monthly_salary': salaries_data['summary']['average_monthly'],
                'outstanding_liabilities': liabilities_data['summary']['remaining']
            },
            'activity': {
                'days_employed': (date.today() - employee.join_date).days if employee.join_date else 0,
                'total_vehicles_assigned': vehicles_data['total_vehicles_used'],
                'total_requests_submitted': requests_data['summary']['total'],
                'last_activity_date': format_date(date.today())
            }
        }
        
        # ═══════════════════════════════════════════════════════════════
        # بناء الاستجابة النهائية
        # ═══════════════════════════════════════════════════════════════
        
        response_data = {
            'employee': employee_data,
            'attendance': attendance_data,
            'vehicles': vehicles_data,
            'salaries': salaries_data,
            'requests': requests_data,
            'liabilities': liabilities_data,
            'documents': documents_data,
            'devices': devices_data,
            'statistics': statistics_data
        }
        
        logger.info(f"✅ تم جلب الملف الشامل للموظف {employee.name} ({employee_id}) بنجاح")
        
        return jsonify({
            'success': True,
            'message': 'تم جلب بيانات الموظف بنجاح',
            'data': response_data
        }), 200
        
    except Exception as e:
        logger.error(f"❌ خطأ في جلب الملف الشامل للموظف: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': 'حدث خطأ في الخادم',
            'error': str(e)
        }), 500


@api_external_bp.route('/employees/export-excel', methods=['GET'])
def export_all_employees_to_excel():
    """
    تصدير بيانات جميع الموظفين إلى ملف Excel شامل
    نقطة نهاية عامة بدون مصادقة
    
    المعاملات الاختيارية:
    - department_id: تصفية حسب القسم
    - status: تصفية حسب الحالة (active, inactive, on_leave)
    
    مثال:
    GET /api/external/employees/export-excel
    GET /api/external/employees/export-excel?department_id=5
    GET /api/external/employees/export-excel?status=active
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file
        
        logger.info("📊 بدء تصدير بيانات الموظفين إلى Excel")
        
        # بناء استعلام الموظفين
        query = Employee.query
        
        # تطبيق الفلاتر
        department_id = request.args.get('department_id', type=int)
        status_filter = request.args.get('status')
        
        if department_id:
            query = query.join(employee_departments).filter(
                employee_departments.c.department_id == department_id
            )
            logger.info(f"🔍 تطبيق فلتر القسم: {department_id}")
        
        if status_filter:
            query = query.filter(Employee.status == status_filter)
            logger.info(f"🔍 تطبيق فلتر الحالة: {status_filter}")
        
        # جلب الموظفين مع العلاقات
        employees = query.options(
            joinedload(Employee.departments)
        ).order_by(Employee.id).all()
        
        logger.info(f"📋 تم جلب {len(employees)} موظف")
        
        if not employees:
            return jsonify({
                'success': False,
                'message': 'لا يوجد موظفين للتصدير'
            }), 404
        
        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "بيانات الموظفين"
        
        # تعريف الأنماط
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cell_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # تعريف الأعمدة (الحقول الأساسية فقط من جدول Employee)
        columns = [
            ("ID", 8),
            ("الرقم الوظيفي", 15),
            ("الاسم الكامل", 25),
            ("الرقم الوطني", 15),
            ("الجنسية", 15),
            ("تاريخ الميلاد", 12),
            ("العمر", 8),
            ("رقم الجوال الرسمي", 15),
            ("رقم الجوال الشخصي", 15),
            ("البريد الإلكتروني", 25),
            ("الأقسام", 30),
            ("الحالة", 12),
            ("المسمى الوظيفي", 20),
            ("نوع العقد", 12),
            ("نوع الموظف", 12),
            ("تاريخ التعيين", 12),
            ("الموقع", 15),
            ("المشروع", 20),
            ("الراتب الأساسي", 12),
            ("حافز الدوام", 12),
            ("الأجر اليومي", 12),
            ("إجمالي الراتب", 12),
            ("حالة العقد", 15),
            ("حالة الرخصة", 15),
            ("حالة الكفالة", 15),
            ("اسم الكفيل الحالي", 25),
            ("رقم الإيبان", 25),
            ("عنوان السكن", 35),
            ("رابط موقع السكن", 40),
            ("مقاس البنطلون", 12),
            ("مقاس التيشرت", 12),
            ("عهدة جوال", 10),
            ("نوع الجوال", 20),
            ("رقم IMEI", 20),
            ("تاريخ الإنشاء", 15),
            ("آخر تحديث", 15)
        ]
        
        # كتابة الرؤوس
        for col_num, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
            ws.column_dimensions[get_column_letter(col_num)].width = col_width
        
        # دالة مساعدة آمنة للتعامل مع القيم الفارغة
        def safe_value(value, default=""):
            """إرجاع قيمة آمنة أو قيمة افتراضية"""
            if value is None:
                return default
            if isinstance(value, (int, float)):
                return value
            return str(value).strip() if str(value).strip() else default
        
        def safe_date(date_obj, format="%Y-%m-%d"):
            """تنسيق التاريخ بشكل آمن"""
            try:
                if date_obj:
                    return date_obj.strftime(format)
                return ""
            except:
                return ""
        
        def safe_number(value, default=0):
            """إرجاع رقم آمن"""
            try:
                return float(value) if value is not None else default
            except:
                return default
        
        def calculate_age(birth_date):
            """حساب العمر"""
            try:
                if birth_date:
                    today = date.today()
                    return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
                return ""
            except:
                return ""
        
        def get_document_status(expiry_date):
            """تحديد حالة الوثيقة"""
            try:
                if not expiry_date:
                    return "غير محدد"
                days_remaining = (expiry_date - date.today()).days
                if days_remaining < 0:
                    return "منتهي"
                elif days_remaining <= 30:
                    return "قرب الانتهاء"
                else:
                    return "ساري"
            except:
                return "غير محدد"
        
        # كتابة بيانات الموظفين
        row_num = 2
        successful_count = 0
        for emp in employees:
            try:
                # حساب البيانات المركبة بشكل آمن
                
                # الأقسام
                try:
                    departments_names = ", ".join([d.name for d in emp.departments]) if emp.departments else ""
                except:
                    departments_names = ""
                
                # إجمالي الراتب (فقط الراتب الأساسي + حافز الدوام)
                try:
                    total_salary = (
                        safe_number(emp.basic_salary) +
                        safe_number(emp.attendance_bonus)
                    )
                except:
                    total_salary = safe_number(emp.basic_salary)
                
                # تحديد الجنسية بشكل آمن
                try:
                    nationality_name = emp.nationality_obj.name_ar if emp.nationality_obj else safe_value(emp.nationality)
                except:
                    nationality_name = safe_value(emp.nationality)
                
                row_data = [
                    emp.id,
                    safe_value(emp.employee_id),
                    safe_value(emp.name),
                    safe_value(emp.national_id),
                    nationality_name,
                    safe_date(emp.birth_date),
                    calculate_age(emp.birth_date),
                    safe_value(emp.mobile),
                    safe_value(emp.mobilePersonal),
                    safe_value(emp.email),
                    departments_names,
                    safe_value(emp.status),
                    safe_value(emp.job_title),
                    safe_value(emp.contract_type),
                    safe_value(emp.employee_type),
                    safe_date(emp.join_date),
                    safe_value(emp.location),
                    safe_value(emp.project),
                    safe_number(emp.basic_salary),
                    safe_number(emp.attendance_bonus),
                    safe_number(emp.daily_wage),
                    total_salary,
                    safe_value(emp.contract_status),
                    safe_value(emp.license_status),
                    safe_value(emp.sponsorship_status),
                    safe_value(emp.current_sponsor_name),
                    safe_value(emp.bank_iban),
                    safe_value(emp.residence_details),
                    safe_value(emp.residence_location_url),
                    safe_value(emp.pants_size),
                    safe_value(emp.shirt_size),
                    "نعم" if emp.has_mobile_custody else "لا",
                    safe_value(emp.mobile_type),
                    safe_value(emp.mobile_imei),
                    safe_date(emp.created_at, "%Y-%m-%d %H:%M"),
                    safe_date(emp.updated_at, "%Y-%m-%d %H:%M")
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = cell_alignment
                    cell.border = cell_border
                
                row_num += 1
                successful_count += 1
                
            except Exception as e:
                logger.error(f"❌ خطأ في معالجة الموظف {emp.id}: {str(e)}")
                # المتابعة مع الموظف التالي بدون توقف
                continue
        
        # إضافة صف الإجمالي
        summary_row = row_num + 1
        ws.cell(row=summary_row, column=1).value = "الإجمالي"
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=2).value = f"{successful_count} موظف"
        ws.cell(row=summary_row, column=2).font = Font(bold=True, size=12)
        
        # تجميد الصف الأول
        ws.freeze_panes = "A2"
        
        # حفظ الملف في الذاكرة
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # تحديد اسم الملف
        filename = f"employees_full_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"✅ تم إنشاء ملف Excel بنجاح: {filename} ({successful_count} موظف من أصل {len(employees)})")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"❌ خطأ في تصدير الموظفين إلى Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': 'حدث خطأ في إنشاء ملف Excel',
            'error': str(e)
        }), 500
