from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, make_response
from models import (Vehicle, VehicleHandover, VehicleWorkshop, VehicleExternalSafetyCheck, 
                    VehicleMaintenance, Employee, Department, User, UserRole)
from app import db
from flask_login import current_user, login_required
from sqlalchemy import or_, and_, func
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl.utils
import io

vehicle_operations_bp = Blueprint('vehicle_operations', __name__)

@vehicle_operations_bp.route('/')
@login_required
def vehicle_operations_list():
    """صفحة عرض جميع عمليات السيارات مع فلاتر شاملة"""

    # جلب الفلاتر من الطلب
    vehicle_filter = request.args.get('vehicle_filter', '').strip()
    operation_type = request.args.get('operation_type', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    employee_filter = request.args.get('employee_filter', '').strip()
    department_filter = request.args.get('department_filter', '').strip()

    # إعداد قوائم العمليات
    operations = []
    
    # فلترة العمليات حسب القسم المحدد للمستخدم الحالي
    from models import employee_departments
    department_filter_ids = None
    if current_user.assigned_department_id:
        # الحصول على معرفات الموظفين في القسم المحدد
        dept_employee_ids = db.session.query(Employee.id).join(
            employee_departments
        ).join(Department).filter(
            Department.id == current_user.assigned_department_id
        ).all()
        department_filter_ids = [emp.id for emp in dept_employee_ids]
    
    # إضافة سجلات للتشخيص
    print(f"فلاتر البحث: vehicle_filter={vehicle_filter}, operation_type={operation_type}")
    if department_filter_ids:
        print(f"فلترة القسم: {len(department_filter_ids)} موظف في القسم المحدد")

    try:
        # جلب عمليات التسليم والاستلام
        if not operation_type or operation_type == 'handover':
            handover_query = VehicleHandover.query.join(Vehicle, VehicleHandover.vehicle_id == Vehicle.id)
            
            # فلترة حسب القسم المحدد للمستخدم
            if department_filter_ids:
                handover_query = handover_query.filter(VehicleHandover.employee_id.in_(department_filter_ids))
            
            if vehicle_filter:
                handover_query = handover_query.filter(Vehicle.plate_number.ilike(f'%{vehicle_filter}%'))
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                    handover_query = handover_query.filter(VehicleHandover.handover_date >= date_from_obj)
                except ValueError:
                    pass
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                    handover_query = handover_query.filter(VehicleHandover.handover_date <= date_to_obj)
                except ValueError:
                    pass

            if employee_filter:
                handover_query = handover_query.filter(VehicleHandover.person_name.ilike(f'%{employee_filter}%'))

            handovers = handover_query.all()
            print(f"تم العثور على {len(handovers)} سجل تسليم/استلام")
            
            for handover in handovers:
                operations.append({
                    'id': handover.id,
                    'type': 'handover',
                    'type_ar': 'تسليم/استلام',
                    'vehicle_plate': handover.vehicle.plate_number if handover.vehicle else 'غير محدد',
                    'operation_date': handover.handover_date,
                    'person_name': handover.person_name,
                    'details': f"{handover.handover_type or 'تسليم/استلام'} - الوقود: {handover.fuel_level or 'غير محدد'}",
                    'status': 'مكتمل',
                    'department': handover.driver_employee.departments[0].name if handover.driver_employee and handover.driver_employee.departments else 'غير محدد'
                })

        # جلب عمليات الورشة
        if not operation_type or operation_type == 'workshop':
            workshop_query = VehicleWorkshop.query.join(Vehicle, VehicleWorkshop.vehicle_id == Vehicle.id)
            
            if vehicle_filter:
                workshop_query = workshop_query.filter(Vehicle.plate_number.ilike(f'%{vehicle_filter}%'))
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                    workshop_query = workshop_query.filter(VehicleWorkshop.entry_date >= date_from_obj)
                except ValueError:
                    pass
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                    workshop_query = workshop_query.filter(VehicleWorkshop.entry_date <= date_to_obj)
                except ValueError:
                    pass

            # فلترة حسب القسم المحدد للمستخدم (المركبات المسلمة لموظفي القسم)
            if department_filter_ids:
                # فلترة المركبات التي لها تسليم لموظف في القسم المحدد
                dept_vehicle_ids = db.session.query(VehicleHandover.vehicle_id).filter(
                    VehicleHandover.handover_type == 'delivery',
                    VehicleHandover.employee_id.in_(department_filter_ids)
                ).distinct().all()
                dept_vehicle_ids = [v.vehicle_id for v in dept_vehicle_ids]
                if dept_vehicle_ids:
                    workshop_query = workshop_query.filter(Vehicle.id.in_(dept_vehicle_ids))
                else:
                    workshop_query = workshop_query.filter(Vehicle.id == -1)  # قائمة فارغة

            workshops = workshop_query.all()
            print(f"تم العثور على {len(workshops)} سجل ورشة")
            
            for workshop in workshops:
                operations.append({
                    'id': workshop.id,
                    'type': 'workshop',
                    'type_ar': 'ورشة',
                    'vehicle_plate': workshop.vehicle.plate_number if workshop.vehicle else 'غير محدد',
                    'operation_date': workshop.entry_date,
                    'person_name': workshop.technician_name or 'غير محدد',
                    'details': f"الورشة: {workshop.workshop_name or 'غير محدد'} - الحالة: {workshop.repair_status or 'غير محدد'}",
                    'status': workshop.repair_status or 'غير محدد',
                    'department': 'الصيانة'
                })

        # جلب فحوصات السلامة الخارجية
        if not operation_type or operation_type == 'safety_check':
            safety_query = VehicleExternalSafetyCheck.query
            
            if vehicle_filter:
                safety_query = safety_query.filter(VehicleExternalSafetyCheck.vehicle_plate_number.ilike(f'%{vehicle_filter}%'))
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                    safety_query = safety_query.filter(func.date(VehicleExternalSafetyCheck.created_at) >= date_from_obj)
                except ValueError:
                    pass
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                    safety_query = safety_query.filter(func.date(VehicleExternalSafetyCheck.created_at) <= date_to_obj)
                except ValueError:
                    pass

            if employee_filter:
                safety_query = safety_query.filter(VehicleExternalSafetyCheck.driver_name.ilike(f'%{employee_filter}%'))

            # فلترة حسب القسم المحدد للمستخدم
            if department_filter_ids:
                # فلترة فحوصات السلامة للمركبات المسلمة لموظفي القسم المحدد
                dept_vehicle_plates = db.session.query(Vehicle.plate_number).join(
                    VehicleHandover, Vehicle.id == VehicleHandover.vehicle_id
                ).filter(
                    VehicleHandover.handover_type == 'delivery',
                    VehicleHandover.employee_id.in_(department_filter_ids)
                ).distinct().all()
                dept_vehicle_plates = [v.plate_number for v in dept_vehicle_plates]
                if dept_vehicle_plates:
                    safety_query = safety_query.filter(VehicleExternalSafetyCheck.vehicle_plate_number.in_(dept_vehicle_plates))
                else:
                    safety_query = safety_query.filter(VehicleExternalSafetyCheck.id == -1)  # قائمة فارغة

            safety_checks = safety_query.all()
            print(f"تم العثور على {len(safety_checks)} فحص سلامة")
            
            for safety in safety_checks:
                status_ar = {
                    'pending': 'قيد المراجعة',
                    'approved': 'معتمد',
                    'rejected': 'مرفوض'
                }.get(safety.approval_status, 'غير محدد')
                
                operations.append({
                    'id': safety.id,
                    'type': 'safety_check',
                    'type_ar': 'فحص سلامة',
                    'vehicle_plate': safety.vehicle_plate_number,
                    'operation_date': safety.created_at.date() if safety.created_at else None,
                    'person_name': safety.driver_name or 'غير محدد',
                    'details': f"فحص سلامة خارجي - السائق: {safety.driver_name or 'غير محدد'}",
                    'status': status_ar,
                    'department': safety.driver_department or 'غير محدد'
                })

        # جلب عمليات الصيانة (إذا كانت موجودة)
        if not operation_type or operation_type == 'maintenance':
            try:
                maintenance_query = VehicleMaintenance.query.join(Vehicle, VehicleMaintenance.vehicle_id == Vehicle.id)
                
                if vehicle_filter:
                    maintenance_query = maintenance_query.filter(Vehicle.plate_number.ilike(f'%{vehicle_filter}%'))
                
                if date_from:
                    try:
                        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                        maintenance_query = maintenance_query.filter(VehicleMaintenance.date >= date_from_obj)
                    except ValueError:
                        pass
                
                if date_to:
                    try:
                        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                        maintenance_query = maintenance_query.filter(VehicleMaintenance.date <= date_to_obj)
                    except ValueError:
                        pass

                maintenances = maintenance_query.all()
                print(f"تم العثور على {len(maintenances)} سجل صيانة")
                
                for maintenance in maintenances:
                    operations.append({
                        'id': maintenance.id,
                        'type': 'maintenance',
                        'type_ar': 'صيانة',
                        'vehicle_plate': maintenance.vehicle.plate_number if maintenance.vehicle else 'غير محدد',
                        'operation_date': maintenance.date,
                        'person_name': maintenance.technician or 'غير محدد',
                        'details': f"صيانة: {maintenance.type or 'غير محدد'} - التكلفة: {maintenance.cost or 0} ريال",
                        'status': maintenance.status or 'مكتمل',
                        'department': 'الصيانة'
                    })
            except Exception as e:
                # في حالة عدم وجود جدول الصيانة
                pass

        # فلترة حسب القسم
        if department_filter:
            operations = [op for op in operations if department_filter.lower() in op['department'].lower()]

        # ترتيب العمليات حسب التاريخ (الأحدث أولاً)
        def get_sort_date(operation):
            date_val = operation['operation_date']
            if date_val is None:
                return datetime.min.date()
            if isinstance(date_val, datetime):
                return date_val.date()
            return date_val
        
        operations.sort(key=get_sort_date, reverse=True)
        print(f"إجمالي العمليات الموجودة: {len(operations)}")

        # جلب قوائم البيانات للفلاتر
        vehicles = Vehicle.query.all()
        departments = Department.query.all()
        
        # إحصائيات
        total_operations = len(operations)
        handover_count = len([op for op in operations if op['type'] == 'handover'])
        workshop_count = len([op for op in operations if op['type'] == 'workshop'])
        safety_count = len([op for op in operations if op['type'] == 'safety_check'])
        maintenance_count = len([op for op in operations if op['type'] == 'maintenance'])

        return render_template('vehicle_operations_list.html',
                             operations=operations,
                             vehicles=vehicles,
                             departments=departments,
                             vehicle_filter=vehicle_filter,
                             operation_type=operation_type,
                             date_from=date_from,
                             date_to=date_to,
                             employee_filter=employee_filter,
                             department_filter=department_filter,
                             total_operations=total_operations,
                             handover_count=handover_count,
                             workshop_count=workshop_count,
                             safety_count=safety_count,
                             maintenance_count=maintenance_count)

    except Exception as e:
        flash(f'حدث خطأ في جلب البيانات: {str(e)}', 'error')
        return render_template('vehicle_operations_list.html',
                             operations=[],
                             vehicles=[],
                             departments=[],
                             vehicle_filter='',
                             operation_type='',
                             date_from='',
                             date_to='',
                             employee_filter='',
                             department_filter='',
                             total_operations=0,
                             handover_count=0,
                             workshop_count=0,
                             safety_count=0,
                             maintenance_count=0)

@vehicle_operations_bp.route('/test')
def test_page():
    """صفحة اختبار للتحقق من الوصول"""
    from flask import jsonify
    
    # اختبار جلب البيانات من قاعدة البيانات
    vehicle_count = Vehicle.query.count()
    handover_count = VehicleHandover.query.count()
    workshop_count = VehicleWorkshop.query.count()
    safety_count = VehicleExternalSafetyCheck.query.count()
    
    return jsonify({
        'message': 'صفحة العمليات تعمل!',
        'user_authenticated': current_user.is_authenticated if current_user else False,
        'database_stats': {
            'vehicles': vehicle_count,
            'handovers': handover_count,
            'workshops': workshop_count,
            'safety_checks': safety_count
        }
    })

@vehicle_operations_bp.route('/api/vehicle-operations/export')
@login_required
def export_vehicle_operations():
    """تصدير عمليات السيارة إلى Excel"""
    try:
        # جلب الفلاتر من الطلب
        vehicle_filter = request.args.get('vehicle_filter', '').strip()
        operation_type = request.args.get('operation_type', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        employee_filter = request.args.get('employee_filter', '').strip()
        department_filter = request.args.get('department_filter', '').strip()

        # إعداد قوائم العمليات (نفس المنطق من دالة العرض)
        operations = []
        
        # جلب عمليات التسليم والاستلام
        if not operation_type or operation_type == 'handover':
            handover_query = VehicleHandover.query.join(Vehicle, VehicleHandover.vehicle_id == Vehicle.id)
            
            if vehicle_filter:
                handover_query = handover_query.filter(Vehicle.plate_number.ilike(f'%{vehicle_filter}%'))
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                    handover_query = handover_query.filter(func.date(VehicleHandover.handover_date) >= date_from_obj)
                except ValueError:
                    pass
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                    handover_query = handover_query.filter(func.date(VehicleHandover.handover_date) <= date_to_obj)
                except ValueError:
                    pass

            handovers = handover_query.all()
            
            for handover in handovers:
                operations.append({
                    'vehicle_plate': handover.vehicle.plate_number if handover.vehicle else 'غير محدد',
                    'type_ar': 'تسليم/استلام',
                    'operation_date': handover.handover_date,
                    'person_name': handover.person_name,
                    'details': f"{handover.handover_type or 'تسليم/استلام'} - الوقود: {handover.fuel_level or 'غير محدد'}",
                    'status': 'مكتمل',
                    'department': handover.driver_employee.departments[0].name if handover.driver_employee and handover.driver_employee.departments else 'غير محدد'
                })

        # جلب عمليات الورشة
        if not operation_type or operation_type == 'workshop':
            workshop_query = VehicleWorkshop.query.join(Vehicle, VehicleWorkshop.vehicle_id == Vehicle.id)
            
            if vehicle_filter:
                workshop_query = workshop_query.filter(Vehicle.plate_number.ilike(f'%{vehicle_filter}%'))
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                    workshop_query = workshop_query.filter(VehicleWorkshop.entry_date >= date_from_obj)
                except ValueError:
                    pass
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                    workshop_query = workshop_query.filter(VehicleWorkshop.entry_date <= date_to_obj)
                except ValueError:
                    pass

            workshops = workshop_query.all()
            
            for workshop in workshops:
                operations.append({
                    'vehicle_plate': workshop.vehicle.plate_number if workshop.vehicle else 'غير محدد',
                    'type_ar': 'ورشة',
                    'operation_date': workshop.entry_date,
                    'person_name': workshop.technician_name or 'غير محدد',
                    'details': f"الورشة: {workshop.workshop_name or 'غير محدد'} - الحالة: {workshop.repair_status or 'غير محدد'}",
                    'status': workshop.repair_status or 'غير محدد',
                    'department': 'الصيانة'
                })

        # جلب فحوصات السلامة الخارجية
        if not operation_type or operation_type == 'safety_check':
            safety_query = VehicleExternalSafetyCheck.query
            
            if vehicle_filter:
                safety_query = safety_query.filter(VehicleExternalSafetyCheck.vehicle_plate_number.ilike(f'%{vehicle_filter}%'))
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                    safety_query = safety_query.filter(func.date(VehicleExternalSafetyCheck.created_at) >= date_from_obj)
                except ValueError:
                    pass
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                    safety_query = safety_query.filter(func.date(VehicleExternalSafetyCheck.created_at) <= date_to_obj)
                except ValueError:
                    pass

            safety_checks = safety_query.all()
            
            for safety in safety_checks:
                status_ar = {
                    'pending': 'قيد المراجعة',
                    'approved': 'معتمد',
                    'rejected': 'مرفوض'
                }.get(safety.approval_status, 'غير محدد')
                
                operations.append({
                    'vehicle_plate': safety.vehicle_plate_number,
                    'type_ar': 'فحص سلامة',
                    'operation_date': safety.created_at.date() if safety.created_at else None,
                    'person_name': safety.driver_name or 'غير محدد',
                    'details': f"فحص سلامة خارجي - السائق: {safety.driver_name or 'غير محدد'}",
                    'status': status_ar,
                    'department': safety.driver_department or 'غير محدد'
                })

        # فلترة حسب القسم
        if department_filter:
            operations = [op for op in operations if department_filter.lower() in op['department'].lower()]

        # ترتيب العمليات حسب التاريخ
        def get_sort_date(operation):
            date_val = operation['operation_date']
            if date_val is None:
                return datetime.min.date()
            if isinstance(date_val, datetime):
                return date_val.date()
            return date_val
        
        operations.sort(key=get_sort_date, reverse=True)

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "عمليات السيارات"

        # تعيين الاتجاه من اليمين إلى اليسار
        ws.sheet_view.rightToLeft = True

        # إعداد الألوان والخطوط
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A5C', end_color='1E3A5C', fill_type='solid')
        data_font = Font(name='Arial', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # إضافة العناوين
        headers = [
            'رقم السيارة',
            'نوع العملية', 
            'تاريخ العملية',
            'اسم الشخص/الفني',
            'التفاصيل',
            'الحالة',
            'القسم'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # إضافة البيانات
        for row_idx, operation in enumerate(operations, 2):
            data = [
                operation['vehicle_plate'],
                operation['type_ar'],
                operation['operation_date'].strftime('%Y-%m-%d') if operation['operation_date'] else 'غير محدد',
                operation['person_name'],
                operation['details'],
                operation['status'],
                operation['department']
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        # تعديل عرض الأعمدة
        column_widths = [15, 15, 15, 20, 40, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # حفظ الملف في الذاكرة
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # إنشاء الاستجابة
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=vehicle_operations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response

    except Exception as e:
        print(f"خطأ في تصدير عمليات السيارة: {e}")
        return jsonify({'message': f'حدث خطأ أثناء التصدير: {str(e)}'}), 500

@vehicle_operations_bp.route('/export')
def export_simple():
    """تصدير مبسط بدون مصادقة للاختبار"""
    try:
        # جلب الفلاتر من الطلب
        vehicle_filter = request.args.get('vehicle_filter', '').strip()
        operation_type = request.args.get('operation_type', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()

        # إعداد قوائم العمليات
        operations = []
        
        # جلب عمليات التسليم والاستلام
        if not operation_type or operation_type == 'handover':
            handover_query = VehicleHandover.query.join(Vehicle, VehicleHandover.vehicle_id == Vehicle.id)
            
            if vehicle_filter:
                handover_query = handover_query.filter(Vehicle.plate_number.ilike(f'%{vehicle_filter}%'))
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                    handover_query = handover_query.filter(func.date(VehicleHandover.handover_date) >= date_from_obj)
                except ValueError:
                    pass
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                    handover_query = handover_query.filter(func.date(VehicleHandover.handover_date) <= date_to_obj)
                except ValueError:
                    pass

            handovers = handover_query.all()
            
            for handover in handovers:
                operations.append({
                    'vehicle_plate': handover.vehicle.plate_number if handover.vehicle else 'غير محدد',
                    'type_ar': 'تسليم/استلام',
                    'operation_date': handover.handover_date,
                    'person_name': handover.person_name,
                    'details': f"{handover.handover_type or 'تسليم/استلام'} - الوقود: {handover.fuel_level or 'غير محدد'}",
                    'status': 'مكتمل',
                    'department': handover.driver_employee.departments[0].name if handover.driver_employee and handover.driver_employee.departments else 'غير محدد'
                })

        # جلب عمليات الورشة
        if not operation_type or operation_type == 'workshop':
            workshop_query = VehicleWorkshop.query.join(Vehicle, VehicleWorkshop.vehicle_id == Vehicle.id)
            
            if vehicle_filter:
                workshop_query = workshop_query.filter(Vehicle.plate_number.ilike(f'%{vehicle_filter}%'))
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                    workshop_query = workshop_query.filter(VehicleWorkshop.entry_date >= date_from_obj)
                except ValueError:
                    pass
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                    workshop_query = workshop_query.filter(VehicleWorkshop.entry_date <= date_to_obj)
                except ValueError:
                    pass

            workshops = workshop_query.all()
            
            for workshop in workshops:
                operations.append({
                    'vehicle_plate': workshop.vehicle.plate_number if workshop.vehicle else 'غير محدد',
                    'type_ar': 'ورشة',
                    'operation_date': workshop.entry_date,
                    'person_name': workshop.technician_name or 'غير محدد',
                    'details': f"الورشة: {workshop.workshop_name or 'غير محدد'} - الحالة: {workshop.repair_status or 'غير محدد'}",
                    'status': workshop.repair_status or 'غير محدد',
                    'department': 'الصيانة'
                })

        # جلب فحوصات السلامة الخارجية
        if not operation_type or operation_type == 'safety_check':
            safety_query = VehicleExternalSafetyCheck.query
            
            if vehicle_filter:
                safety_query = safety_query.filter(VehicleExternalSafetyCheck.vehicle_plate_number.ilike(f'%{vehicle_filter}%'))
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                    safety_query = safety_query.filter(func.date(VehicleExternalSafetyCheck.created_at) >= date_from_obj)
                except ValueError:
                    pass
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                    safety_query = safety_query.filter(func.date(VehicleExternalSafetyCheck.created_at) <= date_to_obj)
                except ValueError:
                    pass

            safety_checks = safety_query.all()
            
            for safety in safety_checks:
                status_ar = {
                    'pending': 'قيد المراجعة',
                    'approved': 'معتمد',
                    'rejected': 'مرفوض'
                }.get(safety.approval_status, 'غير محدد')
                
                operations.append({
                    'vehicle_plate': safety.vehicle_plate_number,
                    'type_ar': 'فحص سلامة',
                    'operation_date': safety.created_at.date() if safety.created_at else None,
                    'person_name': safety.driver_name or 'غير محدد',
                    'details': f"فحص سلامة خارجي - السائق: {safety.driver_name or 'غير محدد'}",
                    'status': status_ar,
                    'department': safety.driver_department or 'غير محدد'
                })

        # ترتيب العمليات حسب التاريخ
        def get_sort_date(operation):
            date_val = operation['operation_date']
            if date_val is None:
                return datetime.min.date()
            if isinstance(date_val, datetime):
                return date_val.date()
            return date_val
        
        operations.sort(key=get_sort_date, reverse=True)

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "عمليات السيارات"

        # تعيين الاتجاه من اليمين إلى اليسار
        ws.sheet_view.rightToLeft = True

        # إعداد الألوان والخطوط
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A5C', end_color='1E3A5C', fill_type='solid')
        data_font = Font(name='Arial', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # إضافة العناوين
        headers = [
            'رقم السيارة',
            'نوع العملية', 
            'تاريخ العملية',
            'اسم الشخص/الفني',
            'التفاصيل',
            'الحالة',
            'القسم'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # إضافة البيانات
        for row_idx, operation in enumerate(operations, 2):
            data = [
                operation['vehicle_plate'],
                operation['type_ar'],
                operation['operation_date'].strftime('%Y-%m-%d') if operation['operation_date'] else 'غير محدد',
                operation['person_name'],
                operation['details'],
                operation['status'],
                operation['department']
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        # تعديل عرض الأعمدة
        column_widths = [15, 15, 15, 20, 40, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # حفظ الملف في الذاكرة
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # إنشاء الاستجابة
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=vehicle_operations_filtered_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response

    except Exception as e:
        print(f"خطأ في تصدير عمليات السيارة المبسط: {e}")
        return f"خطأ في التصدير: {str(e)}", 500

@vehicle_operations_bp.route('/test-operations')
def test_operations():
    """اختبار عرض العمليات بدون مصادقة"""
    vehicle_filter = request.args.get('vehicle_filter', '').strip()
    
    # جلب عمليات التسليم والاستلام للاختبار
    handover_query = VehicleHandover.query.join(Vehicle, VehicleHandover.vehicle_id == Vehicle.id)
    
    if vehicle_filter:
        handover_query = handover_query.filter(Vehicle.plate_number.ilike(f'%{vehicle_filter}%'))
    
    handovers = handover_query.all()
    
    operations = []
    for handover in handovers:
        operations.append({
            'id': handover.id,
            'type': 'handover',
            'vehicle_plate': handover.vehicle.plate_number if handover.vehicle else 'غير محدد',
            'operation_date': str(handover.handover_date),
            'person_name': handover.person_name,
            'details': f"{handover.handover_type or 'تسليم/استلام'} - الوقود: {handover.fuel_level or 'غير محدد'}",
        })
    
    return jsonify({
        'message': f'تم العثور على {len(operations)} عملية',
        'filter': vehicle_filter,
        'operations': operations
    })

