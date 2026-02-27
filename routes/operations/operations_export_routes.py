"""
Operations export routes:
- Export operation details to Excel
- Export and share operation data
"""

from flask import Blueprint, request, send_file, flash, redirect, url_for, current_app, jsonify
from flask_login import login_required, current_user
from core.extensions import db
from models import OperationRequest, VehicleHandover, VehicleWorkshop, Vehicle, UserRole, Employee, MobileDevice
from datetime import datetime
from utils.audit_logger import log_audit
import io
import os

operations_export_bp = Blueprint('operations_export', __name__, url_prefix='/operations')


def get_status_name(status):
    """تحويل حالة العملية إلى النص العربي"""
    status_names = {
        'pending': 'معلقة',
        'under_review': 'تحت المراجعة',
        'approved': 'موافق عليها',
        'rejected': 'مرفوضة'
    }
    return status_names.get(status, status)


def get_priority_name(priority):
    """تحويل أولوية العملية إلى النص العربي"""
    priority_names = {
        'urgent': 'عاجل',
        'high': 'عالي',
        'normal': 'عادي',
        'low': 'منخفض'
    }
    return priority_names.get(priority, priority)


def get_operation_type_name(operation_type):
    """تحويل نوع العملية إلى النص العربي"""
    type_names = {
        'handover': 'تسليم/استلام',
        'workshop': 'ورشة صيانة',
        'external_authorization': 'تفويض خارجي',
        'safety_inspection': 'فحص سلامة'
    }
    return type_names.get(operation_type, operation_type)


@operations_export_bp.route('/<int:operation_id>/export-excel')
@login_required
def export_operation_excel(operation_id):
    """تصدير جميع بيانات العملية إلى Excel"""
    
    if not current_user._is_admin_role():
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'danger')
        return redirect(url_for('dashboard.index'))
    
    try:
        operation = OperationRequest.query.get_or_404(operation_id)
        related_record = operation.get_related_record()
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        
        # شيت 1: معلومات العملية الأساسية
        ws1 = wb.active
        ws1.title = 'معلومات العملية'
        
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2D5AA0', end_color='2D5AA0', fill_type='solid')
        data_font = Font(name='Arial', size=11)
        alignment = Alignment(horizontal='right', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        operation_headers = ['رقم العملية', 'نوع العملية', 'العنوان', 'الوصف', 'الحالة', 
                           'الأولوية', 'تاريخ الطلب', 'تاريخ المراجعة', 'طالب العملية', 
                           'مراجع العملية', 'ملاحظات المراجعة']
        
        for col_num, header in enumerate(operation_headers, 1):
            cell = ws1.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border
            ws1.column_dimensions[cell.column_letter].width = 15
        
        operation_values = [
            operation.id,
            get_operation_type_name(operation.operation_type),
            operation.title or '',
            operation.description or '',
            get_status_name(operation.status),
            get_priority_name(operation.priority),
            operation.requested_at.strftime('%Y-%m-%d %H:%M:%S') if operation.requested_at else '',
            operation.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if operation.reviewed_at else '',
            operation.requester.name if operation.requester else '',
            operation.reviewer.name if operation.reviewer else '',
            operation.review_notes or ''
        ]
        
        for col_num, value in enumerate(operation_values, 1):
            cell = ws1.cell(row=2, column=col_num, value=value)
            cell.font = data_font
            cell.alignment = alignment
            cell.border = border
            
        # شيت 2: بيانات المركبة
        if operation.vehicle:
            vehicle = operation.vehicle
            ws2 = wb.create_sheet('بيانات المركبة')
            
            vehicle_headers = ['رقم اللوحة', 'نوع المركبة', 'الماركة', 'الموديل', 'السنة', 
                              'اللون', 'الحالة', 'ملاحظات']
            
            for col_num, header in enumerate(vehicle_headers, 1):
                cell = ws2.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws2.column_dimensions[cell.column_letter].width = 18
            
            vehicle_values = [
                vehicle.plate_number or '',
                getattr(vehicle, 'type_of_car', '') or '',
                getattr(vehicle, 'make', '') or '',
                vehicle.model or '',
                str(vehicle.year) if vehicle.year else '',
                vehicle.color or '',
                vehicle.status or '',
                vehicle.notes or ''
            ]
            
            for col_num, value in enumerate(vehicle_values, 1):
                cell = ws2.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
        
        # شيت 3: بيانات السائق/الموظف
        employee = None
        current_driver_info = None
        
        if operation.operation_type == 'handover' and related_record:
            current_driver_info = {
                'name': getattr(related_record, 'person_name', ''),
                'phone': getattr(related_record, 'driver_phone_number', ''),
                'residency_number': getattr(related_record, 'driver_residency_number', '')
            }
            
            if related_record.driver_residency_number:
                employee = Employee.query.filter_by(
                    national_id=related_record.driver_residency_number
                ).first()
            
            if not employee and related_record.person_name:
                employee = Employee.query.filter_by(
                    name=related_record.person_name
                ).first()
        
        if not employee and not current_driver_info and operation.vehicle_id:
            last_handover = VehicleHandover.query.filter_by(
                vehicle_id=operation.vehicle_id,
                handover_type='delivery'
            ).order_by(VehicleHandover.handover_date.desc()).first()
            
            if last_handover:
                current_driver_info = {
                    'name': getattr(last_handover, 'person_name', ''),
                    'phone': getattr(last_handover, 'driver_phone_number', ''),
                    'residency_number': getattr(last_handover, 'driver_residency_number', '')
                }
                
                if last_handover.driver_residency_number:
                    employee = Employee.query.filter_by(
                        national_id=last_handover.driver_residency_number
                    ).first()
                
                if not employee and last_handover.person_name:
                    employee = Employee.query.filter_by(
                        name=last_handover.person_name
                    ).first()
        
        if employee or current_driver_info:
            ws3 = wb.create_sheet('بيانات السائق')
            
            driver_headers = ['اسم', 'رقم الهوية', 'رقم الجوال', 'البريد الإلكتروني', 'الحالة']
            
            for col_num, header in enumerate(driver_headers, 1):
                cell = ws3.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws3.column_dimensions[cell.column_letter].width = 16
            
            if employee:
                driver_values = [
                    employee.name or '',
                    employee.national_id or '',
                    employee.mobile or '',
                    employee.email or '',
                    employee.status or ''
                ]
            elif current_driver_info:
                driver_values = [
                    current_driver_info.get('name', '') or '',
                    current_driver_info.get('residency_number', '') or '',
                    current_driver_info.get('phone', '') or '',
                    operation.requester.email or '',
                    'نشط'
                ]
            else:
                driver_values = [
                    operation.requester.username or '',
                    '',
                    operation.requester.phone or '',
                    operation.requester.email or '',
                    'نشط'
                ]
            
            for col_num, value in enumerate(driver_values, 1):
                cell = ws3.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
        
        # شيت 4: تفاصيل الخلوة (للتسليم)
        if operation.operation_type == 'handover' and related_record and hasattr(related_record, 'handover_type'):
            ws4 = wb.create_sheet('نموذج التسليم-الاستلام')
            
            handover_headers = ['نوع العملية', 'تاريخ', 'المسافة', 'الوقود', 'الملاحظات']
            
            for col_num, header in enumerate(handover_headers, 1):
                cell = ws4.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws4.column_dimensions[cell.column_letter].width = 15
            
            handover_values = [
                'تسليم' if related_record.handover_type == 'delivery' else 'استلام',
                related_record.handover_date.strftime('%Y-%m-%d') if related_record.handover_date else '',
                str(getattr(related_record, 'mileage', '')) if getattr(related_record, 'mileage', None) else '',
                getattr(related_record, 'fuel_level', '') or '',
                getattr(related_record, 'notes', '') or ''
            ]
            
            for col_num, value in enumerate(handover_values, 1):
                cell = ws4.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
        
        # حفظ الملف
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"operation_details_{operation.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        log_audit(
            user_id=current_user.id,
            action='export',
            entity_type='operation_request',
            details=f'تصدير تفاصيل العملية {operation_id} إلى Excel'
        )
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'خطأ في تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('operations_core.view_operation', operation_id=operation_id))
