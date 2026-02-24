"""
Operations sharing routes:
- Share via email
- Share with Outlook (.eml)
- Create shared packages (ZIP)
"""

from flask import Blueprint, request, send_file, flash, redirect, url_for, current_app, jsonify
from flask_login import login_required, current_user
from core.extensions import db
from models import OperationRequest, VehicleHandover, VehicleHandoverImage, VehicleWorkshopImage, Vehicle, UserRole, Employee
from datetime import datetime
from utils.audit_logger import log_audit
import io
import os
import zipfile
import shutil

operations_sharing_bp = Blueprint('operations_sharing', __name__, url_prefix='/operations')


@operations_sharing_bp.route('/<int:operation_id>/send-email', methods=['POST'])
@login_required
def send_operation_email(operation_id):
    """إرسال ملفات العملية عبر الإيميل"""
    
    operation = OperationRequest.query.get_or_404(operation_id)
    
    try:
        data = request.get_json()
        to_email = data.get('email')
        to_name = data.get('name', '')
        include_excel = data.get('include_excel', True)
        include_pdf = data.get('include_pdf', True)
        
        if not to_email:
            return jsonify({'success': False, 'message': 'عنوان الإيميل مطلوب'})
        
        # التحقق من صحة الإيميل
        import re
        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_regex, to_email):
            return jsonify({'success': False, 'message': 'عنوان الإيميل غير صحيح'})
        
        vehicle = operation.vehicle
        if not vehicle:
            return jsonify({'success': False, 'message': 'لا توجد مركبة مرتبطة بهذه العملية'})
        
        vehicle_plate = vehicle.plate_number
        driver_name = getattr(vehicle, 'driver_name', None) or 'غير محدد'
        
        excel_file_path = None
        pdf_file_path = None
        
        try:
            # إنشاء ملف Excel مؤقت
            if include_excel:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                excel_filename = f"operation_{operation_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                excel_file_path = os.path.join('/tmp', excel_filename)
                
                wb = Workbook()
                ws = wb.active
                ws.title = 'تفاصيل العملية'
                
                header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                data_font = Font(name='Arial', size=11)
                alignment = Alignment(horizontal='center', vertical='center')
                border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                
                headers = ['البيان', 'القيمة']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    cell.border = border
                    ws.column_dimensions[cell.column_letter].width = 25
                
                data_rows = [
                    ('رقم العملية', f"#{operation.id}"),
                    ('عنوان العملية', operation.title),
                    ('رقم لوحة المركبة', vehicle_plate),
                    ('السائق الحالي', driver_name),
                    ('تاريخ الطلب', operation.requested_at.strftime('%Y/%m/%d %H:%M') if operation.requested_at else operation.created_at.strftime('%Y/%m/%d %H:%M')),
                    ('طالب العملية', operation.requester.username if operation.requester else 'غير محدد'),
                ]
                
                for row_num, (label, value) in enumerate(data_rows, 2):
                    ws.cell(row=row_num, column=1, value=label).font = Font(name='Arial', size=11, bold=True)
                    ws.cell(row=row_num, column=1).alignment = alignment
                    ws.cell(row=row_num, column=1).border = border
                    
                    ws.cell(row=row_num, column=2, value=value).font = data_font
                    ws.cell(row=row_num, column=2).alignment = alignment
                    ws.cell(row=row_num, column=2).border = border
                
                wb.save(excel_file_path)
            
            # إنشاء ملف PDF مؤقت
            if include_pdf and operation.operation_type == 'handover' and operation.related_record_id:
                try:
                    from utils.simple_pdf_generator import create_vehicle_handover_pdf
                    handover_record = VehicleHandover.query.get(operation.related_record_id)
                    if handover_record:
                        pdf_content = create_vehicle_handover_pdf(handover_record)
                        pdf_filename = f"operation_{operation_id}_report.pdf"
                        pdf_file_path = os.path.join('/tmp', pdf_filename)
                        with open(pdf_file_path, 'wb') as f:
                            f.write(pdf_content.read())
                except Exception as pdf_error:
                    current_app.logger.warning(f"فشل في إنشاء PDF: {str(pdf_error)}")
                    pdf_file_path = None
            
            # إرسال الإيميل
            try:
                from services.email_service import EmailService
                email_service = EmailService()
                
                if operation.operation_type == 'handover' and operation.related_record_id:
                    handover_record = VehicleHandover.query.get(operation.related_record_id)
                    if handover_record:
                        result = email_service.send_handover_operation_email(
                            to_email=to_email,
                            to_name=to_name or 'العميل',
                            handover_record=handover_record,
                            vehicle_plate=vehicle_plate,
                            driver_name=driver_name,
                            excel_file_path=excel_file_path if include_excel else None,
                            pdf_file_path=pdf_file_path if include_pdf else None
                        )
                    else:
                        result = email_service.send_vehicle_operation_files(
                            to_email=to_email,
                            to_name=to_name or 'العميل',
                            operation=operation,
                            vehicle_plate=vehicle_plate,
                            driver_name=driver_name,
                            excel_file_path=excel_file_path if include_excel else None,
                            pdf_file_path=pdf_file_path if include_pdf else None
                        )
                else:
                    result = email_service.send_vehicle_operation_files(
                        to_email=to_email,
                        to_name=to_name or 'العميل',
                        operation=operation,
                        vehicle_plate=vehicle_plate,
                        driver_name=driver_name,
                        excel_file_path=excel_file_path if include_excel else None,
                        pdf_file_path=pdf_file_path if include_pdf else None
                    )
                
                if result.get('success'):
                    current_app.logger.info(f'تم إرسال الإيميل بنجاح إلى {to_email}')
                else:
                    raise Exception(f'فشل إرسال الإيميل: {result.get("message")}')
                    
            except Exception as email_error:
                current_app.logger.warning(f'فشل إرسال الإيميل: {email_error}')
                
                try:
                    from services.fallback_email_service import FallbackEmailService
                    fallback_service = FallbackEmailService()
                    result = fallback_service.send_email(
                        to_email=to_email,
                        subject=f'تفاصيل العملية: {operation.title}',
                        html_content=f'<p>{operation.title}</p>'
                    )
                except Exception as fallback_error:
                    current_app.logger.error(f'فشل النظام الاحتياطي: {fallback_error}')
                    result = {
                        'success': False,
                        'message': f'فشل في إرسال الإيميل: {str(fallback_error)}'
                    }
            
            log_audit(
                user_id=current_user.id,
                action='send_email',
                entity_type='operation_request',
                entity_id=operation.id,
                details=f'إرسال ملفات العملية {operation_id} إلى {to_email}'
            )
            
            return jsonify(result)
            
        finally:
            # حذف الملفات المؤقتة
            if excel_file_path and os.path.exists(excel_file_path):
                try:
                    os.remove(excel_file_path)
                except:
                    pass
            
            if pdf_file_path and os.path.exists(pdf_file_path):
                try:
                    os.remove(pdf_file_path)
                except:
                    pass
                    
    except Exception as e:
        current_app.logger.error(f"خطأ في إرسال الإيميل للعملية {operation_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})


@operations_sharing_bp.route('/<int:operation_id>/share-outlook', methods=['GET'])
@login_required
def share_with_outlook(operation_id):
    """إنشاء ملف .eml لمشاركته مع Outlook"""
    
    operation = OperationRequest.query.get_or_404(operation_id)
    
    if operation.operation_type != 'handover' or not operation.related_record_id:
        flash('هذه الميزة متاحة فقط لعمليات التسليم/الاستلام', 'warning')
        return redirect(url_for('operations_core.view_operation', operation_id=operation_id))
    
    handover_record = VehicleHandover.query.get(operation.related_record_id)
    if not handover_record:
        flash('لم يتم العثور على سجل التسليم/الاستلام', 'danger')
        return redirect(url_for('operations_core.view_operation', operation_id=operation_id))
    
    vehicle = operation.vehicle
    if not vehicle:
        flash('لا توجد مركبة مرتبطة بهذه العملية', 'danger')
        return redirect(url_for('operations_core.view_operation', operation_id=operation_id))
    
    vehicle_plate = vehicle.plate_number
    driver_name = getattr(vehicle, 'driver_name', None) or 'غير محدد'
    
    excel_file_path = None
    pdf_file_path = None
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        excel_filename = f"handover_{operation_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_file_path = os.path.join('/tmp', excel_filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'تفاصيل العملية'
        
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        data_font = Font(name='Arial', size=11)
        alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        
        headers = ['البيان', 'القيمة']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 25
        
        operation_type_text = "تسليم" if handover_record.handover_type == 'delivery' else "استلام"
        
        data_rows = [
            ('رقم العملية', f"#{operation.id}"),
            ('نوع العملية', operation_type_text),
            ('رقم لوحة المركبة', vehicle_plate),
            ('السائق', driver_name),
            ('اسم المستلم', handover_record.person_name),
            ('تاريخ العملية', handover_record.handover_date.strftime('%Y/%m/%d') if handover_record.handover_date else ''),
        ]
        
        if handover_record.notes:
            data_rows.append(('ملاحظات', handover_record.notes))
        
        for row_num, (label, value) in enumerate(data_rows, 2):
            ws.cell(row=row_num, column=1, value=label).font = Font(name='Arial', size=11, bold=True)
            ws.cell(row=row_num, column=1).alignment = alignment
            ws.cell(row=row_num, column=1).border = border
            
            ws.cell(row=row_num, column=2, value=value).font = data_font
            ws.cell(row=row_num, column=2).alignment = alignment
            ws.cell(row=row_num, column=2).border = border
        
        wb.save(excel_file_path)
        
        # إنشاء ملف PDF
        try:
            from utils.simple_pdf_generator import create_vehicle_handover_pdf
            pdf_content = create_vehicle_handover_pdf(handover_record)
            pdf_filename = f"handover_{operation_id}_report.pdf"
            pdf_file_path = os.path.join('/tmp', pdf_filename)
            with open(pdf_file_path, 'wb') as f:
                f.write(pdf_content.read())
        except Exception as pdf_error:
            current_app.logger.warning(f"فشل في إنشاء PDF: {str(pdf_error)}")
            pdf_file_path = None
        
        # إنشاء ملف .eml
        try:
            from services.email_service import EmailService
            email_service = EmailService()
            
            eml_bytes, eml_filename = email_service.build_handover_eml(
                to_email="recipient@example.com",
                to_name="المستلم",
                handover_record=handover_record,
                vehicle_plate=vehicle_plate,
                driver_name=driver_name,
                excel_file_path=excel_file_path,
                pdf_file_path=pdf_file_path
            )
            
            if not eml_bytes:
                flash('فشل في إنشاء ملف البريد الإلكتروني', 'warning')
                return redirect(url_for('operations_core.view_operation', operation_id=operation_id))
                
        except Exception as eml_error:
            current_app.logger.error(f"خطأ في إنشاء ملف .eml: {str(eml_error)}")
            flash(f'فشل في إنشاء ملف البريد الإلكتروني: {str(eml_error)}', 'danger')
            return redirect(url_for('operations_core.view_operation', operation_id=operation_id))
        
        log_audit(
            user_id=current_user.id,
            action='share_outlook',
            entity_type='operation_request',
            entity_id=operation.id,
            details=f'إنشاء ملف .eml للعملية {operation_id}'
        )
        
        return send_file(
            eml_bytes,
            mimetype='message/rfc822',
            as_attachment=True,
            download_name=eml_filename
        )
    
    except Exception as e:
        current_app.logger.error(f"خطأ في إنشاء ملف .eml للعملية {operation_id}: {str(e)}")
        flash(f'حدث خطأ: {str(e)}', 'danger')
        return redirect(url_for('operations_core.view_operation', operation_id=operation_id))
        
    finally:
        if excel_file_path and os.path.exists(excel_file_path):
            try:
                os.remove(excel_file_path)
            except:
                pass
        
        if pdf_file_path and os.path.exists(pdf_file_path):
            try:
                os.remove(pdf_file_path)
            except:
                pass


@operations_sharing_bp.route('/<int:operation_id>/share-data', methods=['GET'])
@login_required
def share_data(operation_id):
    """إرجاع بيانات للمشاركة عبر Web Share API"""
    
    try:
        operation = OperationRequest.query.get_or_404(operation_id)
        from flask import url_for
        
        message_parts = ["السادة المعنيين، تحية طيبة وبعد،\n\n"]
        message_parts.append("مرفق لكم تفاصيل عملية استلام أو تسليم المركبة:\n\n")
        
        if operation.operation_type == 'handover':
            handover = VehicleHandover.query.get(operation.related_record_id) if operation.related_record_id else None
            if handover:
                operation_title = "🔄 تسليم مركبة" if handover.handover_type == 'delivery' else "🔄 استلام مركبة"
            else:
                operation_title = "🔄 عملية تسليم/استلام"
        else:
            operation_title = "📋 عملية"
        
        message_parts.append(f"{operation_title}\n")
        
        if operation.vehicle:
            vehicle_info = f"{operation.vehicle.plate_number}"
            message_parts.append(f"• رقم السيارة: {vehicle_info}\n")
        
        message_parts.append("\n📎 مرفقات:\n")
        message_parts.append("ملف Excel\n")
        message_parts.append("ملف PDF\n")
        
        excel_url = url_for('operations_export.export_operation_excel', operation_id=operation_id, _external=True)
        pdf_url = None
        if operation.operation_type == 'handover' and operation.related_record_id:
            pdf_url = url_for('vehicles.handover_pdf_public', id=operation.related_record_id, _external=True)
        
        return jsonify({
            'success': True,
            'message': ''.join(message_parts),
            'excelUrl': excel_url,
            'pdfUrl': pdf_url
        })
        
    except Exception as e:
        current_app.logger.error(f"خطأ في share_data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@operations_sharing_bp.route('/<int:operation_id>/share-package', methods=['GET'])
@login_required
def share_package(operation_id):
    """إنشاء حزمة ZIP شاملة للمشاركة"""
    
    operation = OperationRequest.query.get_or_404(operation_id)
    
    temp_dir = os.path.join(current_app.static_folder, '.temp', f'operation_{operation_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
    zip_path = None
    
    try:
        os.makedirs(temp_dir, exist_ok=True)
        
        # 1. إنشاء ملف نصي بالتفاصيل
        details_path = os.path.join(temp_dir, 'تفاصيل_العملية.txt')
        with open(details_path, 'w', encoding='utf-8') as f:
            f.write('═' * 50 + '\n')
            f.write(f'          تفاصيل العملية #{operation.id}\n')
            f.write('═' * 50 + '\n\n')
            
            operation_types = {
                'handover': 'تسليم/استلام مركبة',
                'workshop': 'ورشة صيانة',
                'external_authorization': 'تفويض خارجي',
                'safety_inspection': 'فحص سلامة'
            }
            f.write(f'نوع العملية: {operation_types.get(operation.operation_type, operation.operation_type)}\n')
            f.write(f'الحالة: {operation.status}\n')
            f.write(f'التاريخ: {operation.created_at.strftime("%Y/%m/%d %H:%M")}\n\n')
            
            if operation.vehicle:
                f.write('─' * 50 + '\n')
                f.write('معلومات المركبة:\n')
                f.write('─' * 50 + '\n')
                f.write(f'رقم اللوحة: {operation.vehicle.plate_number}\n')
                f.write('\n')
        
        # 2. إنشاء ملف Excel
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            excel_path = os.path.join(temp_dir, f'بيانات_العملية_{operation_id}.xlsx')
            wb = Workbook()
            ws = wb.active
            ws.title = 'تفاصيل العملية'
            
            ws['A1'] = 'البيان'
            ws['B1'] = 'القيمة'
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            
            wb.save(excel_path)
        except Exception as e:
            current_app.logger.warning(f'فشل في إنشاء Excel: {str(e)}')
        
        # 3. إنشاء حزمة ZIP
        zip_dir = os.path.join(current_app.static_folder, '.temp')
        os.makedirs(zip_dir, exist_ok=True)
        zip_path = os.path.join(zip_dir, f'operation_{operation_id}_package.zip')
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zipf.write(file_path, arcname)
        
        log_audit(
            user_id=current_user.id,
            action='share_package',
            entity_type='operation_request',
            entity_id=operation.id,
            details=f'إنشاء حزمة مشاركة للعملية {operation_id}'
        )
        
        return send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'عملية_{operation_id}_شاملة.zip'
        )
    
    except Exception as e:
        current_app.logger.error(f"خطأ في share_package: {str(e)}")
        flash(f'حدث خطأ: {str(e)}', 'danger')
        return redirect(url_for('operations_core.view_operation', operation_id=operation_id))
