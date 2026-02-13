import os
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from werkzeug.utils import secure_filename
from sqlalchemy.exc import IntegrityError
from flask_login import login_required
from app import db
from models import Employee, Department, SystemAudit, Document, Attendance, Salary, Module, Permission, Vehicle, VehicleHandover,User,Nationality, employee_departments, MobileDevice, DeviceAssignment, EmployeeLocation, Geofence
from sqlalchemy import func, or_
from utils.excel import parse_employee_excel, generate_employee_excel, export_employee_attendance_to_excel
from utils.date_converter import parse_date
from utils.user_helpers import require_module_access
from utils.employee_comprehensive_report_updated import generate_employee_comprehensive_pdf, generate_employee_comprehensive_excel
from utils.employee_basic_report import generate_employee_basic_pdf
from utils.audit_logger import log_activity

employees_bp = Blueprint('employees', __name__)

# المجلد المخصص لحفظ صور الموظفين
UPLOAD_FOLDER = 'static/uploads/employees'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}

def allowed_file(filename):
    """التحقق من أن الملف من الأنواع المسموحة"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def verify_employee_image(image_path):
    """التحقق من وجود الملف وإرجاع المسار الصحيح أو None"""
    if not image_path:
        return None
    
    # إذا كان المسار يبدأ بـ static/، لا نضيف static/ مرة أخرى
    if image_path.startswith('static/'):
        full_path = image_path
    else:
        full_path = f'static/{image_path}'
    
    # التحقق من وجود الملف
    if os.path.exists(full_path) and os.path.getsize(full_path) > 0:
        return image_path
    else:
        # الملف غير موجود - حاول البحث عن أحدث ملف للموظف والنوع
        return None

def save_employee_image(file, employee_id, image_type):
    """حفظ صورة الموظف وإرجاع المسار - مع تحقق صارم من النجاح"""
    if not file or not file.filename:
        print(f"❌ لا يوجد ملف للحفظ")
        return None
    
    try:
        # التأكد من وجود المجلد بالمسار الكامل
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        
        # إنشاء اسم ملف فريد
        filename = secure_filename(file.filename)
        name, ext = os.path.splitext(filename)
        
        # إذا لم يكن هناك امتداد، نستنتجه من نوع الملف
        if not ext:
            content_type = file.content_type or ''
            if 'pdf' in content_type:
                ext = '.pdf'
            elif 'jpeg' in content_type or 'jpg' in content_type:
                ext = '.jpg'
            elif 'png' in content_type:
                ext = '.png'
            elif 'gif' in content_type:
                ext = '.gif'
            else:
                ext = '.jpg'
        
        # استخدام معرف الموظف الفعلي + التاريخ الحالي
        unique_filename = f"{employee_id}_{image_type}_{datetime.now().strftime('%Y%m%d_%H%M%S%f')}{ext}"
        filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
        
        # قراءة محتوى الملف أولاً قبل الحفظ
        file_content = file.read()
        file.seek(0)  # إعادة المؤشر
        
        # حفظ الملف
        with open(filepath, 'wb') as f:
            f.write(file_content)
        
        # ✅ تحقق صارم من النجاح (تحقق ثلاثي)
        if not os.path.exists(filepath):
            print(f"❌ الملف غير موجود بعد الحفظ: {filepath}")
            return None
        
        file_size = os.path.getsize(filepath)
        if file_size == 0:
            print(f"⚠️ الملف فارغ: {filepath}")
            return None
        
        # التحقق من أن حجم الملف المحفوظ يطابق حجم الملف الأصلي
        if file_size != len(file_content):
            print(f"⚠️ عدم تطابق حجم الملف: {file_size} != {len(file_content)}")
            return None
        
        # إرجاع المسار النسبي (بدون static/)
        relative_path = f"uploads/employees/{unique_filename}"
        print(f"✅ حفظ نجح: {relative_path} ({file_size} bytes)")
        
        return relative_path
        
    except Exception as e:
        print(f"❌ خطأ في حفظ الصورة: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

@employees_bp.route('/')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def index():
    """List all employees with filtering options"""
    # الحصول على معاملات الفلترة من URL
    department_filter = request.args.get('department', '')
    status_filter = request.args.get('status', '')
    multi_department_filter = request.args.get('multi_department', '')
    no_department_filter = request.args.get('no_department', '')
    duplicate_names_filter = request.args.get('duplicate_names', '')
    
    # بناء الاستعلام الأساسي
    query = Employee.query.options(
        db.joinedload(Employee.departments),
        db.joinedload(Employee.nationality_rel)
    )
    
    # فلترة الموظفين حسب القسم المحدد للمستخدم الحالي
    from flask_login import current_user
    if current_user.assigned_department_id:
        # إذا كان المستخدم مرتبط بقسم محدد، عرض موظفي ذلك القسم فقط
        query = query.join(employee_departments).join(Department).filter(Department.id == current_user.assigned_department_id)
    # إذا لم يكن المستخدم مرتبط بقسم، عرض جميع الموظفين (للمديرين العامين)
    
    # تطبيق فلتر القسم (إضافي للفلترة اليدوية)
    elif department_filter:
        query = query.join(employee_departments).join(Department).filter(Department.id == department_filter)
    
    # تطبيق فلتر الحالة
    if status_filter:
        query = query.filter(Employee.status == status_filter)
    
    # تطبيق فلتر الأسماء المكررة
    if duplicate_names_filter == 'yes':
        # البحث عن الأسماء المكررة
        duplicate_names_subquery = db.session.query(Employee.name, func.count(Employee.name).label('name_count'))\
                                           .group_by(Employee.name)\
                                           .having(func.count(Employee.name) > 1)\
                                           .subquery()
        query = query.join(duplicate_names_subquery, Employee.name == duplicate_names_subquery.c.name)
    
    # تطبيق فلتر الموظفين غير المربوطين بأي قسم
    if no_department_filter == 'yes':
        # الموظفين الذين لا يوجد لديهم أي أقسام
        query = query.outerjoin(employee_departments)\
                     .filter(employee_departments.c.employee_id.is_(None))
    elif multi_department_filter == 'yes':
        # الموظفين الذين لديهم أكثر من قسم
        subquery = db.session.query(employee_departments.c.employee_id, 
                                   func.count(employee_departments.c.department_id).label('dept_count'))\
                            .group_by(employee_departments.c.employee_id)\
                            .having(func.count(employee_departments.c.department_id) > 1)\
                            .subquery()
        query = query.join(subquery, Employee.id == subquery.c.employee_id)
    elif multi_department_filter == 'no':
        # الموظفين الذين لديهم قسم واحد فقط أو لا يوجد لديهم أقسام
        subquery = db.session.query(employee_departments.c.employee_id, 
                                   func.count(employee_departments.c.department_id).label('dept_count'))\
                            .group_by(employee_departments.c.employee_id)\
                            .having(func.count(employee_departments.c.department_id) <= 1)\
                            .subquery()
        query = query.outerjoin(subquery, Employee.id == subquery.c.employee_id)\
                     .filter(or_(subquery.c.employee_id.is_(None), 
                               subquery.c.dept_count <= 1))
    
    employees = query.all()
    
    # الحصول على الأقسام للفلتر - مفلترة حسب صلاحيات المستخدم
    if current_user.assigned_department_id:
        # إذا كان المستخدم مرتبط بقسم محدد، عرض ذلك القسم فقط
        departments = Department.query.filter(Department.id == current_user.assigned_department_id).all()
    else:
        # إذا لم يكن المستخدم مرتبط بقسم، عرض جميع الأقسام (للمديرين العامين)
        departments = Department.query.all()
    
    # حساب إحصائيات الموظفين متعددي الأقسام
    multi_dept_count = db.session.query(Employee.id)\
                                .join(employee_departments)\
                                .group_by(Employee.id)\
                                .having(func.count(employee_departments.c.department_id) > 1)\
                                .count()
    
    # حساب الموظفين بدون أقسام
    no_dept_count = db.session.query(Employee.id)\
                             .outerjoin(employee_departments)\
                             .filter(employee_departments.c.employee_id.is_(None))\
                             .count()
    
    # حساب الموظفين بأسماء مكررة - طريقة مبسطة
    duplicate_names_list = db.session.query(Employee.name)\
                                    .group_by(Employee.name)\
                                    .having(func.count(Employee.name) > 1)\
                                    .all()
    
    duplicate_names_count = 0
    duplicate_names_set = set()
    for name_tuple in duplicate_names_list:
        name = name_tuple[0]
        count = db.session.query(Employee).filter(Employee.name == name).count()
        duplicate_names_count += count
        duplicate_names_set.add(name)
    
    single_dept_count = db.session.query(Employee).count() - multi_dept_count - no_dept_count
    
    return render_template('employees/index.html', 
                         employees=employees, 
                         departments=departments,
                         current_department=department_filter,
                         current_status=status_filter,
                         current_multi_department=multi_department_filter,
                         current_no_department=no_department_filter,
                         current_duplicate_names=duplicate_names_filter,
                         multi_dept_count=multi_dept_count,
                         single_dept_count=single_dept_count,
                         no_dept_count=no_dept_count,
                         duplicate_names_count=duplicate_names_count,
                         duplicate_names_set=duplicate_names_set)

@employees_bp.route('/create', methods=['GET', 'POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.CREATE)
def create():
    """Create a new employee"""
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form['name']
            employee_id = request.form['employee_id']
            national_id = request.form['national_id']
            mobile = request.form['mobile']
            status = request.form['status']
            job_title = request.form['job_title']
            location = request.form['location']
            project = request.form['project']
            email = request.form.get('email', '')
            department_id = request.form.get('department_id', None)
            join_date = parse_date(request.form.get('join_date', ''))
            birth_date = parse_date(request.form.get('birth_date', ''))
            mobilePersonal = request.form.get('mobilePersonal')
            nationality_id = request.form.get('nationality_id')
            contract_status = request.form.get('contract_status')
            license_status = request.form.get('license_status')
            
            # الحقول الجديدة لنوع الموظف والعهدة
            employee_type = request.form.get('employee_type', 'regular')
            has_mobile_custody = 'has_mobile_custody' in request.form
            mobile_type = request.form.get('mobile_type', '') if has_mobile_custody else None
            mobile_imei = request.form.get('mobile_imei', '') if has_mobile_custody else None
            
            # حقول الكفالة الجديدة
            sponsorship_status = request.form.get('sponsorship_status', 'inside')
            current_sponsor_name = request.form.get('current_sponsor_name', '')
            
            # معلومات السكن
            residence_details = request.form.get('residence_details', '').strip() or None
            residence_location_url = request.form.get('residence_location_url', '').strip() or None
            
            # معالجة روابط Google Drive
            housing_drive_links = request.form.get('housing_drive_links', '').strip() or None
            
            # مقاسات الزي الموحد
            pants_size = request.form.get('pants_size', '').strip() or None
            shirt_size = request.form.get('shirt_size', '').strip() or None
            
            # الراتب الأساسي
            basic_salary_str = request.form.get('basic_salary', '').strip()
            basic_salary = float(basic_salary_str) if basic_salary_str else 0.0
            
            # حافز الدوام الكامل
            attendance_bonus_str = request.form.get('attendance_bonus', '').strip()
            attendance_bonus = float(attendance_bonus_str) if attendance_bonus_str else 0.0
            
            selected_dept_ids = {int(dept_id) for dept_id in request.form.getlist('department_ids')}
            
            # Convert empty department_id to None
            if department_id == '':
                department_id = None
                
            # Create new employee
            employee = Employee(
                name=name,
                employee_id=employee_id,
                national_id=national_id,
                mobile=mobile,
                status=status,
                job_title=job_title,
                location=location,
                project=project,
                email=email,
                department_id=department_id,
                join_date=join_date,
                birth_date=birth_date,
                mobilePersonal=mobilePersonal,
                nationality_id=int(nationality_id) if nationality_id else None,
                contract_status=contract_status,
                license_status=license_status,
                employee_type=employee_type,
                has_mobile_custody=has_mobile_custody,
                mobile_type=mobile_type,
                mobile_imei=mobile_imei,
                sponsorship_status=sponsorship_status,
                current_sponsor_name=current_sponsor_name,
                residence_details=residence_details,
                residence_location_url=residence_location_url,
                housing_drive_links=housing_drive_links,
                pants_size=pants_size,
                shirt_size=shirt_size,
                basic_salary=basic_salary,
                attendance_bonus=attendance_bonus
            )
            if selected_dept_ids:
                departments_to_assign = Department.query.filter(Department.id.in_(selected_dept_ids)).all()
                employee.departments.extend(departments_to_assign)
            
            db.session.add(employee)
            db.session.commit()
            
            # معالجة رفع صور السكن بعد حفظ الموظف (للحصول على ID)
            housing_images_files = request.files.getlist('housing_images')
            if housing_images_files and any(f.filename for f in housing_images_files):
                saved_images = []
                for img_file in housing_images_files:
                    if img_file and img_file.filename:
                        try:
                            saved_path = save_employee_image(img_file, employee.id, 'housing')
                            if saved_path:
                                saved_images.append(saved_path)
                        except Exception as img_error:
                            print(f"Error saving housing image: {str(img_error)}")
                
                if saved_images:
                    employee.housing_images = ','.join(saved_images)
                    db.session.commit()
            
            # معالجة رفع ملف العرض الوظيفي
            job_offer_file = request.files.get('job_offer_file')
            if job_offer_file and job_offer_file.filename:
                employee.job_offer_file = save_employee_image(job_offer_file, employee.id, 'job_offer')
                db.session.commit()
            
            # معالجة رفع صورة الجواز
            passport_image_file = request.files.get('passport_image_file')
            if passport_image_file and passport_image_file.filename:
                employee.passport_image_file = save_employee_image(passport_image_file, employee.id, 'passport')
                db.session.commit()
            
            # معالجة رفع شهادة العنوان الوطني
            national_address_file = request.files.get('national_address_file')
            if national_address_file and national_address_file.filename:
                employee.national_address_file = save_employee_image(national_address_file, employee.id, 'national_address')
                db.session.commit()
            
            # معالجة الروابط للوثائق
            job_offer_link = request.form.get('job_offer_link', '').strip() or None
            passport_image_link = request.form.get('passport_image_link', '').strip() or None
            national_address_link = request.form.get('national_address_link', '').strip() or None
            
            if job_offer_link or passport_image_link or national_address_link:
                employee.job_offer_link = job_offer_link
                employee.passport_image_link = passport_image_link
                employee.national_address_link = national_address_link
                db.session.commit()
            
            # Log the action
            log_activity('create', 'Employee', employee.id, f'تم إنشاء موظف جديد: {name}')
            
            flash('تم إنشاء الموظف بنجاح', 'success')
            return redirect(url_for('employees.index'))
        
        except IntegrityError as e:
            db.session.rollback()
            error_message = str(e)
            if "employee_id" in error_message.lower():
                flash(f"هذه المعلومات مسجلة مسبقاً: رقم الموظف موجود بالفعل في النظام", "danger")
            elif "national_id" in error_message.lower():
                flash(f"هذه المعلومات مسجلة مسبقاً: رقم الهوية موجود بالفعل في النظام", "danger")
            else:
                flash("هذه المعلومات مسجلة مسبقاً، لا يمكن تكرار بيانات الموظفين", "danger")
            
            # إرجاع المستخدم للنموذج مع البيانات المدخلة
            departments = Department.query.all()
            nationalities = Nationality.query.order_by(Nationality.name_ar).all()
            from models import ImportedPhoneNumber
            available_phone_numbers = ImportedPhoneNumber.query.filter(
                ImportedPhoneNumber.employee_id.is_(None)
            ).order_by(ImportedPhoneNumber.phone_number).all()
            from models import MobileDevice
            available_imei_numbers = MobileDevice.query.filter(
                MobileDevice.status == 'متاح',
                MobileDevice.employee_id.is_(None)
            ).order_by(MobileDevice.imei).all()
            
            return render_template('employees/create.html', 
                                 departments=departments,
                                 nationalities=nationalities,
                                 available_phone_numbers=available_phone_numbers,
                                 available_imei_numbers=available_imei_numbers,
                                 form_data=request.form)
                                 
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
            
            # إرجاع المستخدم للنموذج مع البيانات المدخلة
            departments = Department.query.all()
            nationalities = Nationality.query.order_by(Nationality.name_ar).all()
            from models import ImportedPhoneNumber
            available_phone_numbers = ImportedPhoneNumber.query.filter(
                ImportedPhoneNumber.employee_id.is_(None)
            ).order_by(ImportedPhoneNumber.phone_number).all()
            from models import MobileDevice
            available_imei_numbers = MobileDevice.query.filter(
                MobileDevice.status == 'متاح',
                MobileDevice.employee_id.is_(None)
            ).order_by(MobileDevice.imei).all()
            
            return render_template('employees/create.html', 
                                 departments=departments,
                                 nationalities=nationalities,
                                 available_phone_numbers=available_phone_numbers,
                                 available_imei_numbers=available_imei_numbers,
                                 form_data=request.form)
    
    # Get all departments for the dropdown
    departments = Department.query.all()
    nationalities = Nationality.query.order_by(Nationality.name_ar).all()
    
    # جلب الأرقام المتاحة فقط (غير المربوطة بأي موظف)
    from models import ImportedPhoneNumber
    available_phone_numbers = ImportedPhoneNumber.query.filter(
        ImportedPhoneNumber.employee_id.is_(None)  # الأرقام المتاحة فقط
    ).order_by(ImportedPhoneNumber.phone_number).all()
    
    # جلب أرقام IMEI المتاحة من إدارة الأجهزة
    from models import MobileDevice
    available_imei_numbers = MobileDevice.query.filter(
        MobileDevice.status == 'متاح',  # الأجهزة المتاحة فقط
        MobileDevice.employee_id.is_(None)  # غير مربوطة بموظف
    ).order_by(MobileDevice.imei).all()
    
    return render_template('employees/create.html', 
                         departments=departments,
                         nationalities=nationalities,
                         available_phone_numbers=available_phone_numbers,
                         available_imei_numbers=available_imei_numbers)



@employees_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def edit(id):
    """
    تعديل بيانات موظف موجود وأقسامه، مع التحقق من البيانات الفريدة،
    والتعامل الآمن مع تحديث العلاقات، ومزامنة المستخدم المرتبط.
    """
    employee = Employee.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # 1. استخراج البيانات الجديدة من النموذج
            new_name = request.form.get('name', '').strip()
            new_employee_id = request.form.get('employee_id', '').strip()
            new_national_id = request.form.get('national_id', '').strip()

            # 2. التحقق من صحة البيانات الفريدة قبل أي تعديل
            # التحقق من الرقم الوظيفي
            existing_employee = Employee.query.filter(Employee.employee_id == new_employee_id, Employee.id != id).first()
            if existing_employee:
                flash(f"رقم الموظف '{new_employee_id}' مستخدم بالفعل.", "danger")
                return redirect(url_for('employees.edit', id=id))

            # التحقق من الرقم الوطني
            existing_national = Employee.query.filter(Employee.national_id == new_national_id, Employee.id != id).first()
            if existing_national:
                flash(f"الرقم الوطني '{new_national_id}' مستخدم بالفعل.", "danger")
                return redirect(url_for('employees.edit', id=id))

            # 3. تحديث البيانات الأساسية للموظف
            employee.name = new_name
            employee.employee_id = new_employee_id
            employee.national_id = new_national_id
            # معالجة رقم الجوال مع دعم الإدخال المخصص
            mobile_value = request.form.get('mobile', '')
            print(f"DEBUG: Received mobile value from form: '{mobile_value}'")
            if mobile_value == 'custom':
                mobile_value = request.form.get('mobile_custom', '')
                print(f"DEBUG: Using custom mobile value: '{mobile_value}'")
            employee.mobile = mobile_value
            print(f"DEBUG: Final mobile value set to employee: '{employee.mobile}'")
            
            # تتبع حالة الموظف القديمة قبل التحديث
            old_status = employee.status
            new_status = request.form.get('status', 'active')
            employee.status = new_status
            
            employee.job_title = request.form.get('job_title', '')
            employee.location = request.form.get('location', '')
            employee.project = request.form.get('project', '')
            employee.email = request.form.get('email', '')
            employee.mobilePersonal = request.form.get('mobilePersonal', '')
            employee.contract_status = request.form.get('contract_status', '')
            employee.license_status = request.form.get('license_status', '')
            nationality_id = request.form.get('nationality_id')
            employee.nationality_id = int(nationality_id) if nationality_id else None
            
            # تحديث الحقول الجديدة لنوع الموظف والعهدة
            employee.employee_type = request.form.get('employee_type', 'regular')
            employee.has_mobile_custody = 'has_mobile_custody' in request.form
            employee.mobile_type = request.form.get('mobile_type', '') if employee.has_mobile_custody else None
            employee.mobile_imei = request.form.get('mobile_imei', '') if employee.has_mobile_custody else None
            
            # تحديث حقول الكفالة
            employee.sponsorship_status = request.form.get('sponsorship_status', 'inside')
            employee.current_sponsor_name = request.form.get('current_sponsor_name', '') if employee.sponsorship_status == 'inside' else None
            
            # تحديث حقول المعلومات البنكية
            employee.bank_iban = request.form.get('bank_iban', '').strip() or None
            
            # تحديث معلومات السكن
            employee.residence_details = request.form.get('residence_details', '').strip() or None
            employee.residence_location_url = request.form.get('residence_location_url', '').strip() or None
            
            # معالجة رفع صور السكن (multiple images)
            housing_images_files = request.files.getlist('housing_images')
            if housing_images_files and any(f.filename for f in housing_images_files):
                saved_images = []
                # الاحتفاظ بالصور القديمة إذا كانت موجودة
                if employee.housing_images:
                    saved_images = [img.strip() for img in employee.housing_images.split(',') if img.strip()]
                
                # حفظ الصور الجديدة
                for img_file in housing_images_files:
                    if img_file and img_file.filename:
                        try:
                            saved_path = save_employee_image(img_file, id, 'housing')
                            if saved_path:
                                saved_images.append(saved_path)
                        except Exception as img_error:
                            print(f"Error saving housing image: {str(img_error)}")
                
                # حفظ قائمة الصور كنص مفصول بفواصل
                employee.housing_images = ','.join(saved_images) if saved_images else None
            
            # معالجة روابط Google Drive
            employee.housing_drive_links = request.form.get('housing_drive_links', '').strip() or None
            
            # تحديث مقاسات الزي الموحد
            employee.pants_size = request.form.get('pants_size', '').strip() or None
            employee.shirt_size = request.form.get('shirt_size', '').strip() or None
            
            # تحديث الراتب الأساسي
            basic_salary_str = request.form.get('basic_salary', '').strip()
            employee.basic_salary = float(basic_salary_str) if basic_salary_str else 0.0
            
            # تحديث حافز الدوام الكامل
            attendance_bonus_str = request.form.get('attendance_bonus', '').strip()
            employee.attendance_bonus = float(attendance_bonus_str) if attendance_bonus_str else 0.0
            
            # معالجة رفع صورة شهادة الإيبان
            bank_iban_image_file = request.files.get('bank_iban_image')
            if bank_iban_image_file and bank_iban_image_file.filename:
                # 💾 لا يتم حذف الصورة القديمة - الاحتفاظ بجميع الملفات بشكل دائم
                # حفظ الصورة الجديدة
                employee.bank_iban_image = save_employee_image(bank_iban_image_file, id, 'iban')
            
            # معالجة رفع ملف العرض الوظيفي
            job_offer_file = request.files.get('job_offer_file')
            if job_offer_file and job_offer_file.filename:
                # 💾 لا يتم حذف الملف القديم - الاحتفاظ بجميع الملفات بشكل دائم
                # حفظ الملف الجديد
                employee.job_offer_file = save_employee_image(job_offer_file, id, 'job_offer')
            
            # معالجة رفع صورة الجواز
            passport_image_file = request.files.get('passport_image_file')
            if passport_image_file and passport_image_file.filename:
                # 💾 لا يتم حذف الصورة القديمة - الاحتفاظ بجميع الملفات بشكل دائم
                # حفظ الصورة الجديدة
                employee.passport_image_file = save_employee_image(passport_image_file, id, 'passport')
            
            # معالجة رفع شهادة العنوان الوطني
            national_address_file = request.files.get('national_address_file')
            if national_address_file and national_address_file.filename:
                # 💾 لا يتم حذف الملف القديم - الاحتفاظ بجميع الملفات بشكل دائم
                # حفظ الملف الجديد
                employee.national_address_file = save_employee_image(national_address_file, id, 'national_address')
            
            # معالجة الروابط للوثائق
            employee.job_offer_link = request.form.get('job_offer_link', '').strip() or None
            employee.passport_image_link = request.form.get('passport_image_link', '').strip() or None
            employee.national_address_link = request.form.get('national_address_link', '').strip() or None
            
            join_date_str = request.form.get('join_date')
            employee.join_date = parse_date(join_date_str) if join_date_str else None
            
            # إضافة معالجة تاريخ الميلاد
            birth_date_str = request.form.get('birth_date')
            employee.birth_date = parse_date(birth_date_str) if birth_date_str else None

            selected_dept_ids = {int(dept_id) for dept_id in request.form.getlist('department_ids')}
            current_dept_ids = {dept.id for dept in employee.departments}

            depts_to_add_ids = selected_dept_ids - current_dept_ids

            if depts_to_add_ids:
                    depts_to_add = Department.query.filter(Department.id.in_(depts_to_add_ids)).all()
                    for dept in depts_to_add:
                        employee.departments.append(dept)
                
            depts_to_remove_ids = current_dept_ids - selected_dept_ids


            if depts_to_remove_ids:
                    depts_to_remove = Department.query.filter(Department.id.in_(depts_to_remove_ids)).all()
                    for dept in depts_to_remove:
                        employee.departments.remove(dept)

            user_linked = User.query.filter_by(employee_id=employee.id).first()

            if user_linked:
                    # الطريقة الأسهل هنا هي فقط تعيين القائمة النهائية بعد تعديلها
                    # بما أننا داخل no_autoflush، يمكننا تعيينها مباشرة
                    # سيقوم SQLAlchemy بحساب الفرق بنفسه عند الـ commit
                    final_departments = Department.query.filter(Department.id.in_(selected_dept_ids)).all()
                    user_linked.departments = final_departments
            
            # 6. إذا تم تغيير الحالة إلى غير نشط، فك ربط جميع أرقام SIM والأجهزة
            if new_status == 'inactive' and old_status != 'inactive':
                try:
                    # استيراد النماذج المطلوبة
                    from models import SimCard, DeviceAssignment, MobileDevice
                    from flask import current_app
                    
                    current_app.logger.info(f"Employee {employee.id} ({employee.name}) became inactive - checking for SIM cards and devices")
                    
                    # البحث عن جميع أرقام SIM المرتبطة مباشرة بهذا الموظف
                    sim_cards = SimCard.query.filter_by(employee_id=employee.id).all()
                    
                    # البحث عن جميع تخصيصات الأجهزة النشطة للموظف
                    device_assignments = DeviceAssignment.query.filter_by(
                        employee_id=employee.id, 
                        is_active=True
                    ).all()
                    
                    total_unlinked = 0
                    
                    # فك ربط أرقام SIM المرتبطة مباشرة
                    current_app.logger.info(f"Found {len(sim_cards)} SIM cards directly linked to employee {employee.id}")
                    
                    if sim_cards:
                        for sim_card in sim_cards:
                            current_app.logger.info(f"Unlinking SIM card {sim_card.phone_number} (ID: {sim_card.id}) from employee {employee.id}")
                            
                            # فك الربط
                            sim_card.employee_id = None
                            sim_card.assigned_date = None
                            sim_card.status = 'متاح'
                            total_unlinked += 1
                            
                            # تسجيل عملية فك الربط
                            try:
                                from utils.audit_logger import log_activity
                                log_activity(
                                    action="unassign_auto",
                                    entity_type="SIM",
                                    entity_id=sim_card.id,
                                    details=f"فك ربط رقم SIM {sim_card.phone_number} تلقائياً بسبب تغيير حالة الموظف {employee.name} إلى غير نشط"
                                )
                            except Exception as audit_e:
                                current_app.logger.error(f"Failed to log SIM audit: {str(audit_e)}")
                    
                    # فك ربط تخصيصات الأجهزة النشطة
                    current_app.logger.info(f"Found {len(device_assignments)} active device assignments for employee {employee.id}")
                    
                    if device_assignments:
                        for assignment in device_assignments:
                            current_app.logger.info(f"Deactivating device assignment {assignment.id} for employee {employee.id}")
                            
                            # إلغاء تنشيط التخصيص
                            assignment.is_active = False
                            assignment.end_date = datetime.now()
                            assignment.end_reason = f'فك ربط تلقائي - تغيير حالة الموظف إلى غير نشط'
                            
                            # فك ربط الجهاز إذا كان موجوداً
                            if assignment.device:
                                assignment.device.employee_id = None
                                assignment.device.status = 'متاح'
                            
                            # فك ربط SIM إذا كان موجوداً
                            if assignment.sim_card:
                                assignment.sim_card.employee_id = None
                                assignment.sim_card.assigned_date = None
                                assignment.sim_card.status = 'متاح'
                                total_unlinked += 1
                            
                            # تسجيل عملية فك الربط
                            try:
                                from utils.audit_logger import log_activity
                                device_info = f"جهاز {assignment.device.brand} {assignment.device.model}" if assignment.device else "بدون جهاز"
                                sim_info = f"رقم {assignment.sim_card.phone_number}" if assignment.sim_card else "بدون رقم"
                                
                                log_activity(
                                    action="unassign_auto",
                                    entity_type="DeviceAssignment",
                                    entity_id=assignment.id,
                                    details=f"فك ربط تخصيص الجهاز تلقائياً ({device_info} - {sim_info}) بسبب تغيير حالة الموظف {employee.name} إلى غير نشط"
                                )
                            except Exception as audit_e:
                                current_app.logger.error(f"Failed to log device assignment audit: {str(audit_e)}")
                    
                    # رسالة نجاح شاملة
                    message_parts = []
                    if len(sim_cards) > 0:
                        message_parts.append(f'{len(sim_cards)} رقم SIM مرتبط مباشرة')
                    if len(device_assignments) > 0:
                        message_parts.append(f'{len(device_assignments)} تخصيص جهاز/رقم')
                    
                    if message_parts:
                        flash(f'تم فك ربط {" و ".join(message_parts)} بالموظف تلقائياً', 'info')
                    
                    current_app.logger.info(f"Successfully processed employee {employee.id} deactivation: {len(sim_cards)} SIM cards, {len(device_assignments)} device assignments")
                
                except Exception as e:
                    current_app.logger.error(f"Error unassigning SIM cards for inactive employee: {str(e)}")
                    flash('تحذير: حدث خطأ في فك ربط أرقام SIM. يرجى فحص الأرقام يدوياً', 'warning')
                    # لا نتوقف عن تحديث حالة الموظف حتى لو فشل فك ربط الأرقام

           
            # 7. حفظ كل التغييرات للموظف والمستخدم دفعة واحدة
            db.session.commit()
            
            # تسجيل عملية التحديث
            try:
                from utils.audit_logger import log_activity
                log_activity('update', 'Employee', employee.id, f'تم تحديث بيانات الموظف: {employee.name}')
            except Exception as audit_e:
                print(f"Failed to log employee update audit: {str(audit_e)}")
                
            flash('تم تحديث بيانات الموظف وأقسامه بنجاح.', 'success')
            
            # التحقق من مصدر الطلب للعودة إلى الصفحة المناسبة
            return_url = request.form.get('return_url')
            if not return_url:
                return_url = request.referrer
            
            if return_url and '/departments/' in return_url:
                # استخراج معرف القسم من الرابط المرجعي
                try:
                    department_id = return_url.split('/departments/')[1].split('/')[0]
                    return redirect(url_for('departments.view', id=department_id))
                except:
                    pass
            
            return redirect(url_for('employees.index'))
        
        except Exception as e:
            # تسجيل الخطأ للمطورين
            flash(f'حدث خطأ غير متوقع أثناء عملية التحديث. يرجى المحاولة مرة أخرى. Error updating employee (ID: {id}): {e}', 'danger')


    # في حالة GET request (عند فتح الصفحة لأول مرة)
    all_departments = Department.query.order_by(Department.name).all()
    all_nationalities = Nationality.query.order_by(Nationality.name_ar).all() # جلب كل الجنسيات
    
    # جلب الأرقام المتاحة فقط (غير المربوطة بأي موظف)
    from models import ImportedPhoneNumber
    available_phone_numbers = ImportedPhoneNumber.query.filter(
        ImportedPhoneNumber.employee_id.is_(None)  # الأرقام المتاحة فقط
    ).order_by(ImportedPhoneNumber.phone_number).all()
    
    # جلب أرقام IMEI المتاحة من إدارة الأجهزة
    from models import MobileDevice
    available_imei_numbers = MobileDevice.query.filter(
        MobileDevice.status == 'متاح',  # الأجهزة المتاحة فقط
        MobileDevice.employee_id.is_(None)  # غير مربوطة بموظف
    ).order_by(MobileDevice.imei).all()
    
    # جلب بيانات الجهاز و SIM المربوط بالموظف من DeviceAssignment
    from models import DeviceAssignment, SimCard
    active_assignment = DeviceAssignment.query.filter_by(
        employee_id=employee.id,
        is_active=True
    ).first()
    
    # بيانات الجهاز و SIM المربوط (سيتم عرضها في الصفحة)
    assigned_device = None
    assigned_sim = None
    
    if active_assignment:
        # جلب الجهاز مباشرة باستخدام device_id
        if active_assignment.device_id:
            assigned_device = MobileDevice.query.get(active_assignment.device_id)
        
        # جلب SIM مباشرة باستخدام sim_card_id
        if active_assignment.sim_card_id:
            assigned_sim = SimCard.query.get(active_assignment.sim_card_id)
    
    print(f"Passing {len(all_nationalities)} nationalities to the template.")
    return render_template('employees/edit.html', 
                         employee=employee, 
                         nationalities=all_nationalities, 
                         departments=all_departments,
                         available_phone_numbers=available_phone_numbers,
                         available_imei_numbers=available_imei_numbers,
                         assigned_device=assigned_device,
                         assigned_sim=assigned_sim)





# @employees_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
# @login_required
# @require_module_access(Module.EMPLOYEES, Permission.EDIT)
# def edit(id):
#     """
#     تعديل بيانات موظف موجود وأقسامه المرتبطة بها، مع مزامنة المستخدم المرتبط.
#     """
#     employee = Employee.query.get_or_404(id)
    
#     if request.method == 'POST':
#         try:
#             # 1. تحديث البيانات الأساسية للموظف
#             employee.name = request.form['name']
#             employee.employee_id = request.form['employee_id']
#             employee.national_id = request.form['national_id']
#             employee.mobile = request.form['mobile']
#             employee.status = request.form['status']
#             employee.job_title = request.form['job_title']
#             employee.location = request.form.get('location', '')
#             employee.project = request.form.get('project', '')
#             employee.email = request.form.get('email', '')
            
#             join_date_str = request.form.get('join_date', '')
#             if join_date_str:
#                 employee.join_date = parse_date(join_date_str) # افترض وجود دالة parse_date

#             # 2. *** تحديث الأقسام المرتبطة (منطق متعدد إلى متعدد) ***
#             # استلام قائمة معرفات الأقسام المحددة من مربعات الاختيار
#             selected_dept_ids = [int(dept_id) for dept_id in request.form.getlist('department_ids')]
            
#             # جلب كائنات الأقسام الفعلية من قاعدة البيانات
#             selected_departments = Department.query.filter(Department.id.in_(selected_dept_ids)).all()
            
#             # تعيين القائمة الجديدة للموظف، وSQLAlchemy سيتولى تحديث جدول الربط
#             employee.departments = selected_departments
            
#             # 3. *** المزامنة التلقائية للمستخدم المرتبط (مهم جداً) ***
#             # ابحث عن المستخدم المرتبط بهذا الموظف (إن وجد)
#             user_linked_to_employee = User.query.filter_by(employee_id=employee.id).first()
#             if user_linked_to_employee:
#                 # إذا وجد مستخدم، قم بمزامنة قائمة أقسامه لتكون مطابقة
#                 user_linked_to_employee.departments = selected_departments
#                 print(f"INFO: Synced departments for linked user: {user_linked_to_employee.name}")
            
#             # 4. حفظ كل التغييرات للموظف والمستخدم في قاعدة البيانات
#             db.session.commit()
            
#             # 5. تسجيل الإجراء والعودة
#             log_activity('update', 'Employee', employee.id, f'تم تحديث بيانات الموظف: {employee.name}')
#             flash('تم تحديث بيانات الموظف وأقسامه بنجاح.', 'success')
#             return redirect(url_for('employees.index'))
        
#         except  Exception as e:
#             db.session.rollback()
#             flash(f"خطأ في التكامل: رقم الموظف أو الرقم الوطني قد يكون مستخدماً بالفعل.{str(e)}", "danger")
#         except Exception as e:
#             db.session.rollback()
#             flash(f'حدث خطأ غير متوقع أثناء التحديث: {str(e)}', 'danger')
#             # من الجيد تسجيل الخطأ الكامل في السجلات للمطورين
#             # current_app.logger.error(f"Error editing employee {id}: {e}")
            
#     # في حالة GET request، جهز البيانات للعرض
#     all_departments = Department.query.order_by(Department.name).all()
#     return render_template('employees/edit.html', employee=employee, departments=all_departments)








@employees_bp.route('/<int:id>/view')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def view(id):
    """View detailed employee information"""
    employee = Employee.query.options(
        db.joinedload(Employee.departments),
        db.joinedload(Employee.nationality_rel)
    ).get_or_404(id)
    
    # Get employee documents
    documents = Document.query.filter_by(employee_id=id).all()
    
    # Get document types in Arabic
    document_types_map = {
        'national_id': 'الهوية الوطنية', 
        'passport': 'جواز السفر', 
        'health_certificate': 'الشهادة الصحية', 
        'work_permit': 'تصريح العمل', 
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'other': 'أخرى'
    }
    
    # Get documents by type for easier display
    documents_by_type = {}
    for doc_type in document_types_map.keys():
        documents_by_type[doc_type] = None
    
    today = datetime.now().date()
    
    for doc in documents:
        # Add expiry status
        days_to_expiry = (doc.expiry_date - today).days
        if days_to_expiry < 0:
            doc.status_class = "danger"
            doc.status_text = "منتهية"
        elif days_to_expiry < 30:
            doc.status_class = "warning"
            doc.status_text = f"تنتهي خلال {days_to_expiry} يوم"
        else:
            doc.status_class = "success"
            doc.status_text = "سارية"
        
        # Store document by type
        documents_by_type[doc.document_type] = doc
    
    # Get all attendance records for this employee
    attendances = Attendance.query.filter_by(employee_id=id).order_by(Attendance.date.desc()).all()
    
    # Get salary records
    salaries = Salary.query.filter_by(employee_id=id).order_by(Salary.year.desc(), Salary.month.desc()).all()
    
    # Get vehicle handover records
    vehicle_handovers = VehicleHandover.query.filter_by(employee_id=id).order_by(VehicleHandover.handover_date.desc()).all()
    
    # Get current vehicle assigned to this employee
    # السيارة الحالية المربوطة بالموظف في جدول السيارات
    current_assigned_vehicle = Vehicle.query.filter_by(driver_name=employee.name).first()
    
    # Get mobile devices assigned to this employee
    mobile_devices = MobileDevice.query.filter_by(employee_id=id).order_by(MobileDevice.assigned_date.desc()).all()
    
    # Get device assignments for this employee
    from models import DeviceAssignment
    device_assignments = DeviceAssignment.query.filter_by(
        employee_id=id, 
        is_active=True
    ).options(
        db.joinedload(DeviceAssignment.device),
        db.joinedload(DeviceAssignment.sim_card)
    ).all()
    
    all_departments = Department.query.order_by(Department.name).all()
    
    # جلب معلومات السكن (العقارات التي يقطن فيها الموظف)
    housing_properties = employee.housing_properties
    
    return render_template('employees/view.html', 
                          employee=employee, 
                          documents=documents,
                          documents_by_type=documents_by_type,
                          document_types_map=document_types_map,
                          attendances=attendances,
                          salaries=salaries,
                          vehicle_handovers=vehicle_handovers,
                          current_assigned_vehicle=current_assigned_vehicle,
                          mobile_devices=mobile_devices,
                          device_assignments=device_assignments,
                          departments=all_departments,
                          housing_properties=housing_properties
                          )

@employees_bp.route('/<int:id>/upload_iban', methods=['POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def upload_iban(id):
    """رفع صورة الإيبان البنكي للموظف"""
    employee = Employee.query.get_or_404(id)
    
    try:
        # الحصول على بيانات الإيبان والملف
        bank_iban = request.form.get('bank_iban', '').strip()
        iban_file = request.files.get('iban_image')
        
        # تحديث رقم الإيبان
        if bank_iban:
            employee.bank_iban = bank_iban
        
        # رفع صورة الإيبان إذا تم اختيارها
        if iban_file and iban_file.filename:
            # 💾 لا يتم حذف الصورة القديمة - الاحتفاظ بجميع الملفات بشكل دائم
            # حفظ الصورة الجديدة
            image_path = save_employee_image(iban_file, employee.id, 'iban')
            if image_path:
                employee.bank_iban_image = image_path
        
        db.session.commit()
        
        # تسجيل العملية
        log_activity('update', 'Employee', employee.id, f'تم تحديث بيانات الإيبان البنكي للموظف: {employee.name}')
        
        flash('تم حفظ بيانات الإيبان البنكي بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حفظ بيانات الإيبان: {str(e)}', 'danger')
    
    return redirect(url_for('employees.view', id=id))

@employees_bp.route('/<int:id>/delete_iban_image', methods=['POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def delete_iban_image(id):
    """حذف صورة الإيبان البنكي للموظف"""
    employee = Employee.query.get_or_404(id)
    
    try:
        if employee.bank_iban_image:
            # 💾 الملف يبقى محفوظاً - نحذف فقط المرجع من قاعدة البيانات
            # لكن نحتفظ بالملف الفعلي في النظام للأمان
            employee.bank_iban_image = None
            db.session.commit()
            
            # تسجيل العملية
            log_activity('delete', 'Employee', employee.id, f'تم إزالة مرجع صورة الإيبان البنكي للموظف: {employee.name} (الملف محفوظ)')
            
            flash('تم إزالة صورة الإيبان البنكي (الملف محفوظ بشكل آمن)', 'success')
        else:
            flash('لا توجد صورة إيبان لحذفها', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف صورة الإيبان: {str(e)}', 'danger')
    
    return redirect(url_for('employees.view', id=id))

@employees_bp.route('/<int:id>/delete_housing_image', methods=['POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def delete_housing_image(id):
    """حذف صورة من صور السكن التوضيحية"""
    employee = Employee.query.get_or_404(id)
    image_path = request.form.get('image_path', '').strip()
    
    try:
        if not image_path:
            flash('لم يتم تحديد الصورة المراد حذفها', 'warning')
            return redirect(url_for('employees.view', id=id))
        
        if employee.housing_images:
            # تحويل القائمة إلى list
            image_list = [img.strip() for img in employee.housing_images.split(',')]
            
            # البحث عن الصورة في القائمة
            clean_image_path = image_path.replace('static/', '')
            image_to_remove = None
            
            for img in image_list:
                if img.replace('static/', '') == clean_image_path:
                    image_to_remove = img
                    break
            
            if image_to_remove:
                # حذف الصورة من القائمة
                image_list.remove(image_to_remove)
                
                # 💾 الملف يبقى محفوظاً - نحذف فقط المرجع من قاعدة البيانات
                # لكن نحتفظ بالملف الفعلي في النظام للأمان
                
                # تحديث قاعدة البيانات
                employee.housing_images = ','.join(image_list) if image_list else None
                db.session.commit()
                
                # تسجيل العملية
                log_activity('delete', 'Employee', employee.id, f'تم إزالة صورة من صور السكن للموظف: {employee.name} (الملف محفوظ)')
                
                flash('تم إزالة الصورة (الملف محفوظ بشكل آمن)', 'success')
            else:
                flash('لم يتم العثور على الصورة في القائمة', 'warning')
        else:
            flash('لا توجد صور سكن لحذفها', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الصورة: {str(e)}', 'danger')
    
    return redirect(url_for('employees.view', id=id))

@employees_bp.route('/<int:id>/confirm_delete')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.DELETE)
def confirm_delete(id):
    """صفحة تأكيد حذف الموظف"""
    employee = Employee.query.get_or_404(id)
    
    # تحديد عنوان الصفحة التي تم تحويلنا منها للعودة إليها عند الإلغاء
    return_url = request.referrer
    if not return_url or '/employees/' in return_url:
        return_url = url_for('employees.index')
    
    return render_template('employees/confirm_delete.html', 
                          employee=employee, 
                          return_url=return_url)

@employees_bp.route('/<int:id>/delete', methods=['GET', 'POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.DELETE)
def delete(id):
    """Delete an employee"""
    from models import EmployeeRequest
    employee = Employee.query.get_or_404(id)
    name = employee.name
    
    # إذا كان الطلب GET، نعرض صفحة التأكيد
    if request.method == 'GET':
        return redirect(url_for('employees.confirm_delete', id=id))
    
    # إذا كان الطلب POST، نتحقق من تأكيد الحذف
    confirmed = request.form.get('confirmed', 'no')
    
    if confirmed != 'yes':
        flash('لم يتم تأكيد عملية الحذف', 'warning')
        return redirect(url_for('employees.view', id=id))
    
    try:
        # التحقق من وجود طلبات معلقة للموظف
        pending_requests = EmployeeRequest.query.filter_by(
            employee_id=id,
            status='PENDING'
        ).count()
        
        if pending_requests > 0:
            flash(f'لا يمكن حذف الموظف لديه {pending_requests} طلب(ات) معلقة. يرجى حذف الطلبات أولاً', 'danger')
            return redirect(url_for('employees.view', id=id))
        
        # حذف جميع البيانات المرتبطة بالموظف بشكل آمن وشامل
        from sqlalchemy import text
        
        # قائمة الجداول والأعمدة المرتبطة بـ employee_id
        tables_to_clean = [
            ('geofence_attendance', 'employee_id', 'delete'),
            ('salary', 'employee_id', 'delete'),
            ('geofence_sessions', 'employee_id', 'delete'),
            ('document', 'employee_id', 'delete'),
            ('attendance', 'employee_id', 'delete'),
            ('government_fee', 'employee_id', 'delete'),
            ('fee', 'employee_id', 'delete'),
            ('employee_departments', 'employee_id', 'delete'),
            ('sim_cards', 'employee_id', 'unlink'),
            ('imported_phone_numbers', 'employee_id', 'unlink'),
            ('device_assignments', 'employee_id', 'delete'),
            ('transactions', 'employee_id', 'delete'),
            ('voicehub_calls', 'employee_id', 'delete'),
            ('property_employees', 'employee_id', 'delete'),
            ('geofence_events', 'employee_id', 'delete'),
            ('employee_geofences', 'employee_id', 'delete'),
            ('employee_locations', 'employee_id', 'delete'),
            ('employee_requests', 'employee_id', 'delete'),
            ('employee_liabilities', 'employee_id', 'delete'),
            ('request_notifications', 'employee_id', 'delete'),
            ('external_authorization', 'employee_id', 'delete'),
            ('mobile_devices', 'employee_id', 'unlink'),
            ('safety_inspection', 'employee_id', 'delete'),
        ]
        
        # تنظيف جميع الجداول
        for table, column, action in tables_to_clean:
            try:
                if action == 'delete':
                    db.session.execute(text(f"DELETE FROM {table} WHERE {column} = :id"), {"id": id})
                elif action == 'unlink':
                    db.session.execute(text(f"UPDATE {table} SET {column} = NULL WHERE {column} = :id"), {"id": id})
                db.session.flush()
            except Exception as e:
                print(f"Warning: Could not clean {table}: {str(e)}")
                db.session.rollback()
        
        # فك ربط vehicle_handover (لديه عمودين)
        try:
            db.session.execute(text("UPDATE vehicle_handover SET employee_id = NULL WHERE employee_id = :id"), {"id": id})
            db.session.execute(text("UPDATE vehicle_handover SET supervisor_employee_id = NULL WHERE supervisor_employee_id = :id"), {"id": id})
            db.session.flush()
        except:
            pass
        
        # إزالة من الدوائر الجغرافية المعينة
        try:
            for geofence in Geofence.query.all():
                if employee in geofence.assigned_employees:
                    geofence.assigned_employees.remove(employee)
            db.session.flush()
        except:
            pass
        
        # حذف الموظف أخيراً
        db.session.delete(employee)
        db.session.commit()
        
        # Log the action
        log_activity('delete', 'Employee', id, f'تم حذف الموظف: {name}')
        
        flash('تم حذف الموظف بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الموظف: {str(e)}', 'danger')
    
    # التحقق من مصدر الطلب للعودة إلى الصفحة المناسبة
    referrer = request.form.get('return_url')
    if referrer and '/departments/' in referrer:
        try:
            department_id = referrer.split('/departments/')[1].split('/')[0]
            return redirect(url_for('departments.view', id=department_id))
        except:
            pass
    
    return redirect(url_for('employees.index'))

@employees_bp.route('/import', methods=['GET', 'POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.CREATE)
def import_excel():
    """Import employees from Excel file"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                print(f"Received file: {file.filename}")
                
                # Parse Excel file
                employees_data = parse_employee_excel(file)
                print(f"Parsed {len(employees_data)} employee records from Excel")
                
                success_count = 0
                error_count = 0
                error_details = []
                
                for index, data in enumerate(employees_data):
                    try:
                        print(f"Processing employee {index+1}: {data.get('name', 'Unknown')}")
                        
                        # Check if employee with same employee_id already exists
                        existing = Employee.query.filter_by(employee_id=data['employee_id']).first()
                        if existing:
                            print(f"Employee with ID {data['employee_id']} already exists")
                            error_count += 1
                            error_details.append(f"الموظف برقم {data['employee_id']} موجود مسبقا")
                            continue
                            
                        # Check if employee with same national_id already exists
                        existing = Employee.query.filter_by(national_id=data['national_id']).first()
                        if existing:
                            print(f"Employee with national ID {data['national_id']} already exists")
                            error_count += 1
                            error_details.append(f"الموظف برقم هوية {data['national_id']} موجود مسبقا")
                            continue
                        
                        # Extract department data separately
                        department_name = data.pop('department', None)
                        
                        # Create employee without department field
                        employee = Employee(**data)
                        db.session.add(employee)
                        db.session.flush()  # Get the ID without committing
                        
                        # Handle department assignment if provided
                        if department_name:
                            department = Department.query.filter_by(name=department_name).first()
                            if department:
                                employee.departments.append(department)
                            else:
                                # Create new department if it doesn't exist
                                new_department = Department(name=department_name)
                                db.session.add(new_department)
                                db.session.flush()
                                employee.departments.append(new_department)
                        
                        db.session.commit()
                        success_count += 1
                        print(f"Successfully added employee: {data.get('name')}")
                    except Exception as e:
                        db.session.rollback()
                        error_count += 1
                        print(f"Error adding employee {index+1}: {str(e)}")
                        error_details.append(f"خطأ في السجل {index+1}: {str(e)}")
                
                # Log the import
                error_detail_str = ", ".join(error_details[:5])
                if len(error_details) > 5:
                    error_detail_str += f" وغيرها من الأخطاء..."
                
                details = f'تم استيراد {success_count} موظف بنجاح و {error_count} فشل'
                if error_details:
                    details += f". أخطاء: {error_detail_str}"
                    
                audit = SystemAudit(
                    action='import',
                    entity_type='employee',
                    entity_id=0,
                    details=details
                )
                db.session.add(audit)
                db.session.commit()
                
                if error_count > 0:
                    flash(f'تم استيراد {success_count} موظف بنجاح و {error_count} فشل. {error_detail_str}', 'warning')
                else:
                    flash(f'تم استيراد {success_count} موظف بنجاح', 'success')
                return redirect(url_for('employees.index'))
            except Exception as e:
                flash(f'حدث خطأ أثناء استيراد الملف: {str(e)}', 'danger')
        else:
            flash('الملف يجب أن يكون بصيغة Excel (.xlsx, .xls)', 'danger')
    
    return render_template('employees/import.html')

@employees_bp.route('/import/template')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def import_template():
    """Download Excel template for employee import with all comprehensive fields"""
    try:
        import pandas as pd
        
        # إنشاء قالب Excel مع جميع الحقول المطلوبة والاختيارية
        template_data = {
            'الاسم الكامل': ['محمد أحمد علي', 'فاطمة سالم محمد'],
            'رقم الموظف': ['EMP001', 'EMP002'],
            'رقم الهوية الوطنية': ['1234567890', '0987654321'],
            'رقم الجوال': ['0501234567', '0509876543'],
            'الجوال الشخصي': ['0551234567', ''],
            'المسمى الوظيفي': ['مطور برمجيات', 'محاسبة'],
            'الحالة الوظيفية': ['active', 'active'],
            'الموقع': ['الرياض', 'جدة'],
            'المشروع': ['مشروع الرياض', 'مشروع جدة'],
            'البريد الإلكتروني': ['mohamed@company.com', 'fatima@company.com'],
            'الأقسام': ['تقنية المعلومات', 'المحاسبة'],
            'تاريخ الانضمام': ['2024-01-15', '2024-02-01'],
            'تاريخ انتهاء الإقامة': ['2025-12-31', '2025-11-30'],
            'حالة العقد': ['محدد المدة', 'دائم'],
            'حالة الرخصة': ['سارية', 'سارية'],
            'الجنسية': ['سعودي', 'مصري'],
            'ملاحظات': ['موظف متميز', '']
        }
        
        # إنشاء DataFrame
        df = pd.DataFrame(template_data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # كتابة البيانات النموذجية
            df.to_excel(writer, sheet_name='البيانات النموذجية', index=False)
            
            # إنشاء ورقة فارغة للاستخدام
            empty_df = pd.DataFrame(columns=template_data.keys())
            empty_df.to_excel(writer, sheet_name='استيراد الموظفين', index=False)
            
            # إنشاء ورقة التعليمات
            instructions_data = {
                'العمود': list(template_data.keys()),
                'مطلوب/اختياري': ['مطلوب', 'مطلوب', 'مطلوب', 'مطلوب', 'اختياري', 'مطلوب', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري'],
                'التنسيق المطلوب': [
                    'نص',
                    'نص فريد',
                    'رقم من 10 أرقام',
                    'رقم جوال سعودي',
                    'رقم جوال (اختياري)',
                    'نص',
                    'active/inactive/on_leave',
                    'نص',
                    'نص',
                    'بريد إلكتروني صحيح',
                    'اسم القسم',
                    'YYYY-MM-DD',
                    'YYYY-MM-DD',
                    'نص',
                    'نص',
                    'اسم الجنسية',
                    'نص (اختياري)'
                ]
            }
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='التعليمات', index=False)
        
        output.seek(0)
        
        # تسجيل العملية
        audit = SystemAudit(
            action='download_template',
            entity_type='employee_import',
            entity_id=0,
            details='تم تحميل قالب استيراد الموظفين المحسن'
        )
        db.session.add(audit)
        db.session.commit()
        
        return send_file(
            output,
            download_name='قالب_استيراد_الموظفين_شامل.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ في إنشاء القالب: {str(e)}', 'danger')
        return redirect(url_for('employees.import_excel'))

@employees_bp.route('/import/empty_template')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def empty_import_template():
    """Download empty Excel template for employee import"""
    try:
        import pandas as pd
        
        # إنشاء قالب فارغ مع جميع الحقول المطلوبة
        empty_template_data = {
            'الاسم الكامل': [],
            'رقم الموظف': [],
            'رقم الهوية الوطنية': [],
            'رقم الجوال': [],
            'الجوال الشخصي': [],
            'المسمى الوظيفي': [],
            'الحالة الوظيفية': [],
            'الموقع': [],
            'المشروع': [],
            'البريد الإلكتروني': [],
            'الأقسام': [],
            'تاريخ الانضمام': [],
            'تاريخ انتهاء الإقامة': [],
            'حالة العقد': [],
            'حالة الرخصة': [],
            'الجنسية': [],
            'ملاحظات': []
        }
        
        # إنشاء DataFrame فارغ
        df = pd.DataFrame(empty_template_data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # كتابة القالب الفارغ
            df.to_excel(writer, sheet_name='استيراد الموظفين', index=False)
            
            # إنشاء ورقة التعليمات
            instructions_data = {
                'العمود': [
                    'الاسم الكامل', 'رقم الموظف', 'رقم الهوية الوطنية', 'رقم الجوال', 
                    'الجوال الشخصي', 'المسمى الوظيفي', 'الحالة الوظيفية', 'الموقع', 
                    'المشروع', 'البريد الإلكتروني', 'الأقسام', 'تاريخ الانضمام', 
                    'تاريخ انتهاء الإقامة', 'حالة العقد', 'حالة الرخصة', 'الجنسية', 'ملاحظات'
                ],
                'مطلوب/اختياري': [
                    'مطلوب', 'مطلوب', 'مطلوب', 'مطلوب', 'اختياري', 'مطلوب', 
                    'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري', 
                    'اختياري', 'اختياري', 'اختياري', 'اختياري', 'اختياري'
                ],
                'التنسيق المطلوب': [
                    'نص', 'نص فريد', 'رقم من 10 أرقام', 'رقم جوال سعودي', 
                    'رقم جوال (اختياري)', 'نص', 'active/inactive/on_leave', 'نص', 
                    'نص', 'بريد إلكتروني صحيح', 'اسم القسم', 'YYYY-MM-DD', 
                    'YYYY-MM-DD', 'نص', 'نص', 'اسم الجنسية', 'نص (اختياري)'
                ],
                'مثال': [
                    'محمد أحمد علي', 'EMP001', '1234567890', '0501234567',
                    '0551234567', 'مطور برمجيات', 'active', 'الرياض',
                    'مشروع الرياض', 'mohamed@company.com', 'تقنية المعلومات', '2024-01-15',
                    '2025-12-31', 'محدد المدة', 'سارية', 'سعودي', 'موظف متميز'
                ]
            }
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='التعليمات والأمثلة', index=False)
        
        output.seek(0)
        
        # تسجيل العملية
        audit = SystemAudit(
            action='download_empty_template',
            entity_type='employee_import',
            entity_id=0,
            details='تم تحميل نموذج فارغ لاستيراد الموظفين'
        )
        db.session.add(audit)
        db.session.commit()
        
        return send_file(
            output,
            download_name='نموذج_استيراد_الموظفين_فارغ.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ في إنشاء النموذج الفارغ: {str(e)}', 'danger')
        return redirect(url_for('employees.import_excel'))

@employees_bp.route('/<int:id>/update_status', methods=['POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def update_status(id):
    """تحديث حالة الموظف"""
    employee = Employee.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            new_status = request.form.get('status')
            if new_status not in ['active', 'inactive', 'on_leave']:
                flash('حالة غير صالحة', 'danger')
                return redirect(url_for('employees.view', id=id))
            
            old_status = employee.status
            employee.status = new_status
            
            note = request.form.get('note', '')
            
            # إذا تم تغيير الحالة إلى غير نشط، فك ربط جميع أرقام SIM والأجهزة
            if new_status == 'inactive' and old_status != 'inactive':
                try:
                    # استيراد النماذج المطلوبة
                    from models import SimCard, DeviceAssignment, MobileDevice
                    from flask import current_app
                    
                    current_app.logger.info(f"Checking SIM cards and devices for employee {employee.id} ({employee.name}) who became inactive")
                    
                    # 1. البحث عن جميع أرقام SIM المرتبطة مباشرة بهذا الموظف
                    sim_cards = SimCard.query.filter_by(employee_id=employee.id).all()
                    
                    # 2. البحث عن جميع تخصيصات الأجهزة النشطة للموظف
                    device_assignments = DeviceAssignment.query.filter_by(
                        employee_id=employee.id, 
                        is_active=True
                    ).all()
                    
                    total_unlinked = 0
                    
                    # فك ربط أرقام SIM المرتبطة مباشرة
                    current_app.logger.info(f"Found {len(sim_cards)} SIM cards directly linked to employee {employee.id}")
                    
                    if sim_cards:
                        for sim_card in sim_cards:
                            current_app.logger.info(f"Unlinking SIM card {sim_card.phone_number} (ID: {sim_card.id}) from employee {employee.id}")
                            
                            # فك الربط
                            sim_card.employee_id = None
                            sim_card.assigned_date = None
                            sim_card.status = 'متاح'
                            total_unlinked += 1
                            
                            # تسجيل عملية فك الربط
                            try:
                                from utils.audit_logger import log_activity
                                log_activity(
                                    action="unassign_auto",
                                    entity_type="SIM",
                                    entity_id=sim_card.id,
                                    details=f"فك ربط رقم SIM {sim_card.phone_number} تلقائياً بسبب تغيير حالة الموظف {employee.name} إلى غير نشط"
                                )
                            except Exception as audit_e:
                                current_app.logger.error(f"Failed to log SIM audit: {str(audit_e)}")
                    
                    # فك ربط تخصيصات الأجهزة النشطة
                    current_app.logger.info(f"Found {len(device_assignments)} active device assignments for employee {employee.id}")
                    
                    if device_assignments:
                        for assignment in device_assignments:
                            current_app.logger.info(f"Deactivating device assignment {assignment.id} for employee {employee.id}")
                            
                            # إلغاء تنشيط التخصيص
                            assignment.is_active = False
                            assignment.end_date = datetime.now()
                            assignment.end_reason = f'فك ربط تلقائي - تغيير حالة الموظف إلى غير نشط'
                            
                            # فك ربط الجهاز إذا كان موجوداً
                            if assignment.device:
                                assignment.device.employee_id = None
                                assignment.device.status = 'متاح'
                            
                            # فك ربط SIM إذا كان موجوداً
                            if assignment.sim_card:
                                assignment.sim_card.employee_id = None
                                assignment.sim_card.assigned_date = None
                                assignment.sim_card.status = 'متاح'
                                total_unlinked += 1
                            
                            # تسجيل عملية فك الربط
                            try:
                                from utils.audit_logger import log_activity
                                device_info = f"جهاز {assignment.device.brand} {assignment.device.model}" if assignment.device else "بدون جهاز"
                                sim_info = f"رقم {assignment.sim_card.phone_number}" if assignment.sim_card else "بدون رقم"
                                
                                log_activity(
                                    action="unassign_auto",
                                    entity_type="DeviceAssignment",
                                    entity_id=assignment.id,
                                    details=f"فك ربط تخصيص الجهاز تلقائياً ({device_info} - {sim_info}) بسبب تغيير حالة الموظف {employee.name} إلى غير نشط"
                                )
                            except Exception as audit_e:
                                current_app.logger.error(f"Failed to log device assignment audit: {str(audit_e)}")
                    
                    # حفظ التغييرات في قاعدة البيانات
                    db.session.commit()
                    
                    # رسالة نجاح شاملة
                    message_parts = []
                    if len(sim_cards) > 0:
                        message_parts.append(f'{len(sim_cards)} رقم SIM مرتبط مباشرة')
                    if len(device_assignments) > 0:
                        message_parts.append(f'{len(device_assignments)} تخصيص جهاز/رقم')
                    
                    if message_parts:
                        flash(f'تم فك ربط {" و ".join(message_parts)} بالموظف تلقائياً', 'info')
                    
                    current_app.logger.info(f"Successfully processed employee {employee.id} deactivation: {len(sim_cards)} SIM cards, {len(device_assignments)} device assignments")
                
                except Exception as e:
                    current_app.logger.error(f"Error unassigning SIM cards for inactive employee: {str(e)}")
                    db.session.rollback()
                    flash('تحذير: حدث خطأ في فك ربط أرقام SIM. يرجى فحص الأرقام يدوياً', 'warning')
                    # لا نتوقف عن تحديث حالة الموظف حتى لو فشل فك ربط الأرقام
            
            # توثيق التغيير في السجل
            status_names = {
                'active': 'نشط',
                'inactive': 'غير نشط',
                'on_leave': 'في إجازة'
            }
            
            details = f'تم تغيير حالة الموظف {employee.name} من "{status_names.get(old_status, old_status)}" إلى "{status_names.get(new_status, new_status)}"'
            if note:
                details += f" - ملاحظات: {note}"
                
            # تسجيل العملية
            audit = SystemAudit(
                action='update_status',
                entity_type='employee',
                entity_id=employee.id,
                details=details
            )
            db.session.add(audit)
            db.session.commit()
            
            flash(f'تم تحديث حالة الموظف إلى {status_names.get(new_status, new_status)} بنجاح', 'success')
            
            # العودة إلى الصفحة السابقة
            referrer = request.referrer
            if referrer and '/departments/' in referrer:
                department_id = referrer.split('/departments/')[1].split('/')[0]
                return redirect(url_for('departments.view', id=department_id))
            
            return redirect(url_for('employees.view', id=id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث حالة الموظف: {str(e)}', 'danger')
            return redirect(url_for('employees.view', id=id))

@employees_bp.route('/export')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def export_excel():
    """Export employees to Excel file"""
    try:
        employees = Employee.query.options(
            db.joinedload(Employee.departments),
            db.joinedload(Employee.nationality_rel)
        ).all()
        output = generate_employee_excel(employees)
        
        # Log the export
        audit = SystemAudit(
            action='export',
            entity_type='employee',
            entity_id=0,
            details=f'تم تصدير {len(employees)} موظف إلى ملف Excel'
        )
        db.session.add(audit)
        db.session.commit()
        
        return send_file(
            BytesIO(output.getvalue()),
            download_name='employees.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('employees.index'))

@employees_bp.route('/export_comprehensive')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def export_comprehensive():
    """تصدير شامل لبيانات الموظفين مع جميع التفاصيل والعُهد والمعلومات البنكية"""
    try:
        from utils.basic_comprehensive_export import generate_comprehensive_employee_excel
        
        employees = Employee.query.options(
            db.joinedload(Employee.departments),
            db.joinedload(Employee.nationality_rel),
            db.joinedload(Employee.salaries),
            db.joinedload(Employee.attendances),
            db.joinedload(Employee.documents)
        ).all()
        
        output = generate_comprehensive_employee_excel(employees)
        
        # تسجيل العملية
        audit = SystemAudit(
            action='export_comprehensive',
            entity_type='employee',
            entity_id=0,
            details=f'تم التصدير الشامل لبيانات {len(employees)} موظف مع جميع التفاصيل'
        )
        db.session.add(audit)
        db.session.commit()
        
        # إنشاء اسم الملف مع التاريخ
        current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'تصدير_شامل_الموظفين_{current_date}.xlsx'
        
        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        import traceback
        print(f"Error in comprehensive export: {str(e)}")
        print(traceback.format_exc())
        flash(f'حدث خطأ أثناء التصدير الشامل: {str(e)}', 'danger')
        return redirect(url_for('employees.index'))
        
@employees_bp.route('/<int:id>/export_attendance_excel')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def export_attendance_excel(id):
    """تصدير بيانات الحضور كملف إكسل"""
    try:
        # الحصول على بيانات الموظف
        employee = Employee.query.get_or_404(id)
        
        # الحصول على الشهر والسنة من معاملات الطلب
        month = request.args.get('month')
        year = request.args.get('year')
        
        # تحويل البيانات إلى أرقام صحيحة إذا كانت موجودة
        if month:
            try:
                month = int(month)
            except (ValueError, TypeError):
                flash('قيمة الشهر غير صالحة، تم استخدام الشهر الحالي', 'warning')
                month = None
                
        if year:
            try:
                year = int(year)
            except (ValueError, TypeError):
                flash('قيمة السنة غير صالحة، تم استخدام السنة الحالية', 'warning')
                year = None
        
        # توليد ملف الإكسل
        output = export_employee_attendance_to_excel(employee, month, year)
        
        # تعيين اسم الملف مع التاريخ الحالي
        current_date = datetime.now().strftime('%Y%m%d')
        
        # إضافة الشهر والسنة إلى اسم الملف إذا كانا موجودين
        if month and year:
            filename = f"attendance_{employee.name}_{year}_{month}_{current_date}.xlsx"
        else:
            # استخدام الشهر والسنة الحالية إذا لم يتم توفيرهما
            current_month = datetime.now().month
            current_year = datetime.now().year
            filename = f"attendance_{employee.name}_{current_year}_{current_month}_{current_date}.xlsx"
        
        # تسجيل الإجراء
        audit = SystemAudit(
            action='export',
            entity_type='attendance',
            entity_id=employee.id,
            details=f'تم تصدير سجل الحضور للموظف: {employee.name}'
        )
        db.session.add(audit)
        db.session.commit()
        
        # إرسال ملف الإكسل
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        # طباعة تتبع الخطأ في سجل الخادم للمساعدة في التشخيص
        import traceback
        print(f"Error exporting attendance: {str(e)}")
        print(traceback.format_exc())
        
        flash(f'حدث خطأ أثناء تصدير ملف الحضور: {str(e)}', 'danger')
        return redirect(url_for('employees.view', id=id))

@employees_bp.route('/<int:id>/upload_image', methods=['POST'])
@login_required
@require_module_access(Module.EMPLOYEES, Permission.EDIT)
def upload_image(id):
    """رفع صورة للموظف (الصورة الشخصية، صورة الهوية، أو صورة الرخصة)"""
    employee = Employee.query.get_or_404(id)
    
    image_type = request.form.get('image_type')
    if not image_type or image_type not in ['profile', 'national_id', 'license']:
        flash('نوع الصورة غير صحيح', 'danger')
        return redirect(url_for('employees.view', id=id))
    
    if 'image' not in request.files:
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('employees.view', id=id))
    
    file = request.files['image']
    if file.filename == '':
        flash('لم يتم اختيار ملف', 'danger')
        return redirect(url_for('employees.view', id=id))
    
    # حفظ الصورة - التحقق الصارم (استخدام معرف الموظف الفعلي وليس رقم الموظف)
    image_path = save_employee_image(file, employee.id, image_type)
    
    if not image_path:
        flash('❌ فشل في رفع الصورة. تأكد من أن الملف من النوع المسموح', 'danger')
        return redirect(url_for('employees.view', id=id))
    
    try:
        # 1️⃣ حفظ الصورة الجديدة في قاعدة البيانات أولاً
        old_path = None
        if image_type == 'profile':
            old_path = employee.profile_image
            employee.profile_image = image_path
        elif image_type == 'national_id':
            old_path = employee.national_id_image
            employee.national_id_image = image_path
        elif image_type == 'license':
            old_path = employee.license_image
            employee.license_image = image_path
        
        # 2️⃣ تأكيد التغييرات في قاعدة البيانات
        db.session.commit()
        print(f"✅ DB: تم حفظ {image_path} في قاعدة البيانات")
        
        # 💾 الملف القديم يبقى محفوظاً - لا يتم حذف الملفات الفعلية
        if old_path:
            print(f"💾 الملف القديم محفوظ للأمان: {old_path}")
        
        # رسالة النجاح
        success_messages = {
            'profile': '✅ تم رفع الصورة الشخصية بنجاح',
            'national_id': '✅ تم رفع صورة الهوية بنجاح',
            'license': '✅ تم رفع صورة الرخصة بنجاح'
        }
        flash(success_messages.get(image_type, '✅ تم رفع الصورة بنجاح'), 'success')
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ خطأ: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'❌ خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('employees.view', id=id))


@employees_bp.route('/<int:id>/basic_report')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def basic_report(id):
    """تقرير المعلومات الأساسية للموظف"""
    try:
                # طباعة رسالة تشخيصية
        print("بدء إنشاء التقرير الشامل للموظف")
        
        # التحقق من وجود الموظف
        employee = Employee.query.get_or_404(id)
        print(f"تم العثور على الموظف: {employee.name}")
        
        # إنشاء ملف PDF
        print("استدعاء دالة إنشاء PDF")
        pdf_buffer = generate_employee_basic_pdf(id)
        print("تم استلام ناتج ملف PDF")
        
        if not pdf_buffer:
            flash('لم يتم العثور على بيانات كافية لإنشاء التقرير', 'warning')
            return redirect(url_for('employees.view', id=id))
        
        if pdf_buffer:
            employee = Employee.query.get_or_404(id)
            current_date = datetime.now().strftime('%Y%m%d')
            filename = f'تقرير_أساسي_{employee.name}_{current_date}.pdf'
            
            # تسجيل الإجراء
            audit = SystemAudit(
                action='export',
                entity_type='employee_basic_report',
                entity_id=employee.id,
                details=f'تم تصدير التقرير الأساسي للموظف: {employee.name}'
            )
            db.session.add(audit)
            db.session.commit()
            
            return send_file(
                pdf_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        else:
            flash('خطأ في إنشاء ملف PDF', 'danger')
            return redirect(url_for('employees.view', id=id))
    except Exception as e:
        flash(f'خطأ في تصدير PDF: {str(e)}', 'danger')
        return redirect(url_for('employees.view', id=id))


@employees_bp.route('/<int:id>/comprehensive_report')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def comprehensive_report(id):
    """تقرير شامل عن الموظف بصيغة PDF"""
    try:
        # طباعة رسالة تشخيصية
        print("بدء إنشاء التقرير الشامل للموظف")
        
        # التحقق من وجود الموظف
        employee = Employee.query.get_or_404(id)
        print(f"تم العثور على الموظف: {employee.name}")
        
        # إنشاء ملف PDF
        print("استدعاء دالة إنشاء PDF")
        output = generate_employee_comprehensive_pdf(id)
        print("تم استلام ناتج ملف PDF")
        
        if not output:
            flash('لم يتم العثور على بيانات كافية لإنشاء التقرير', 'warning')
            return redirect(url_for('employees.view', id=id))
        
        # اسم الملف المُصدَّر
        filename = f"تقرير_شامل_{employee.name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        print(f"اسم الملف: {filename}")
        
        # تسجيل عملية التصدير
        audit = SystemAudit(
            action='export',
            entity_type='employee_report',
            entity_id=employee.id,
            details=f'تم إنشاء تقرير شامل للموظف: {employee.name}'
        )
        db.session.add(audit)
        db.session.commit()
        print("تم تسجيل العملية في سجل النظام")
        
        # طباعة نوع ناتج الملف للتشخيص
        print(f"نوع ناتج الملف: {type(output)}")
        print(f"حجم البيانات: {output.getbuffer().nbytes} بايت")
        
        # إرسال ملف PDF
        print("إرسال الملف للمتصفح")
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        # طباعة تتبع الخطأ في سجل الخادم للمساعدة في التشخيص
        import traceback
        print(f"Error generating comprehensive report: {str(e)}")
        print(traceback.format_exc())
        
        flash(f'حدث خطأ أثناء إنشاء التقرير الشامل: {str(e)}', 'danger')
        return redirect(url_for('employees.view', id=id))


@employees_bp.route('/<int:id>/comprehensive_report_excel')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def comprehensive_report_excel(id):
    """تقرير شامل عن الموظف بصيغة Excel"""
    try:
        # التحقق من وجود الموظف
        employee = Employee.query.get_or_404(id)
        
        # إنشاء ملف Excel
        output = generate_employee_comprehensive_excel(id)
        
        if not output:
            flash('لم يتم العثور على بيانات كافية لإنشاء التقرير', 'warning')
            return redirect(url_for('employees.view', id=id))
        
        # اسم الملف المُصدَّر
        filename = f"تقرير_شامل_{employee.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # تسجيل عملية التصدير
        audit = SystemAudit(
            action='export',
            entity_type='employee_report_excel',
            entity_id=employee.id,
            details=f'تم تصدير تقرير شامل (إكسل) للموظف: {employee.name}'
        )
        db.session.add(audit)
        db.session.commit()
        
        # إرسال ملف الإكسل
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        # طباعة تتبع الخطأ في سجل الخادم للمساعدة في التشخيص
        import traceback
        print(f"Error generating comprehensive Excel report: {str(e)}")
        print(traceback.format_exc())
        
        flash(f'حدث خطأ أثناء إنشاء التقرير الشامل (إكسل): {str(e)}', 'danger')
        return redirect(url_for('employees.view', id=id))


@employees_bp.route('/tracking')
@login_required
def tracking():
    """صفحة تتبع مواقع الموظفين عبر GPS"""
    from flask_login import current_user
    
    # TODO: لاحقاً فعّل هذا القيد للمديرين فقط
    # if current_user.role != 'admin':
    #     flash('هذه الصفحة متاحة للمديرين فقط', 'danger')
    #     return redirect(url_for('dashboard.index'))
    
    # الحصول على معاملات الفلترة
    department_filter = request.args.get('department', '')
    search_query = request.args.get('search', '')
    
    # بناء استعلام الموظفين النشطين فقط
    query = Employee.query.filter(Employee.status == 'active').options(
        db.joinedload(Employee.departments)
    )
    
    # تطبيق فلتر القسم
    if department_filter:
        query = query.join(employee_departments).join(Department).filter(Department.id == department_filter)
    
    # تطبيق فلتر البحث (اسم أو رقم وظيفي)
    if search_query:
        query = query.filter(
            or_(
                Employee.name.contains(search_query),
                Employee.employee_id.contains(search_query)
            )
        )
    
    # جلب جميع الموظفين
    all_employees = query.all()
    employee_ids = [emp.id for emp in all_employees]
    
    # جلب آخر موقع لكل موظف باستخدام window function في استعلام واحد
    from sqlalchemy import func as sql_func, and_
    from sqlalchemy.orm import aliased
    
    # استخدام subquery مع ROW_NUMBER للحصول على آخر موقع لكل موظف
    latest_locations_subq = db.session.query(
        EmployeeLocation.employee_id,
        EmployeeLocation.id.label('location_id'),
        sql_func.row_number().over(
            partition_by=EmployeeLocation.employee_id,
            order_by=EmployeeLocation.recorded_at.desc()
        ).label('rn')
    ).filter(
        EmployeeLocation.employee_id.in_(employee_ids)
    ).subquery()
    
    # جلب المواقع الفعلية مع البيانات الكاملة
    latest_locations_query = db.session.query(
        EmployeeLocation
    ).join(
        latest_locations_subq,
        and_(
            EmployeeLocation.id == latest_locations_subq.c.location_id,
            latest_locations_subq.c.rn == 1
        )
    ).all()
    
    # بناء dictionary للمواقع حسب employee_id
    locations_by_employee = {loc.employee_id: loc for loc in latest_locations_query}
    
    # معالجة الموظفين وحساب الحالات
    employee_locations = {}
    employees_with_location = []
    employees_without_location = []
    
    for emp in all_employees:
        latest_location = locations_by_employee.get(emp.id)
        
        if latest_location:
            # حساب عمر الموقع بالدقائق والساعات
            age_seconds = (datetime.utcnow() - latest_location.recorded_at).total_seconds()
            age_minutes = age_seconds / 60
            age_hours = age_seconds / 3600
            
            # تحديد حالة الاتصال والألوان
            if age_minutes < 5:
                color = 'green'
                status_text = 'متصل'
                connection_status = 'connected'
            elif age_minutes < 30:
                color = 'orange'
                status_text = 'نشط مؤخراً'
                connection_status = 'recently_active'
            elif age_hours < 6:
                color = 'red'
                status_text = 'غير متصل'
                connection_status = 'disconnected'
            else:
                color = 'gray'
                status_text = 'غير نشط'
                connection_status = 'inactive'
            
            employee_locations[emp.id] = {
                'latitude': latest_location.latitude,
                'longitude': latest_location.longitude,
                'accuracy': getattr(latest_location, 'accuracy_m', None), 
                'recorded_at': latest_location.recorded_at,
                'age_minutes': age_minutes,
                'age_hours': age_hours,
                'color': color,
                'status_text': status_text,
                'connection_status': connection_status,
                'vehicle_id': latest_location.vehicle_id
            }
            employees_with_location.append(emp)
        else:
            employees_without_location.append(emp)
    
    # ترتيب: الموظفون الذين لديهم موقع أولاً
    employees = employees_with_location + employees_without_location
    
    # جلب كل الـ geofences النشطة مرة واحدة لتجنب N+1
    all_geofences = Geofence.query.filter_by(is_active=True).all()
    
    # جلب كل الـ vehicles مرة واحدة
    vehicle_ids = [loc_data['vehicle_id'] for loc_data in employee_locations.values() if loc_data.get('vehicle_id')]
    vehicles_dict = {}
    if vehicle_ids:
        vehicles = Vehicle.query.filter(Vehicle.id.in_(vehicle_ids)).all()
        vehicles_dict = {v.id: v for v in vehicles}
    
    # حساب المسافات للـ geofences لكل موظف
    from math import radians, sin, cos, sqrt, atan2
    
    for emp_id, location_data in employee_locations.items():
        # استخدام البيانات المحفوظة بدلاً من جلبها مرة أخرى
        latest_location = locations_by_employee.get(emp_id)
        
        if latest_location:
            # Check all geofences to see if employee is inside any of them
            for gf in all_geofences:
                # Calculate distance using Haversine formula
                R = 6371000  # Earth radius in meters
                lat1, lon1 = radians(float(latest_location.latitude)), radians(float(latest_location.longitude))
                lat2, lon2 = radians(float(gf.center_latitude)), radians(float(gf.center_longitude))
                dlat = lat2 - lat1
                dlon = lon2 - lon1
                a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
                c = 2 * atan2(sqrt(a), sqrt(1-a))
                distance = R * c
                
                # If employee is within geofence radius
                if distance <= gf.radius_meters:
                    location_data['geofence_name'] = gf.name
                    break
            
            # Get vehicle name if assigned
            if latest_location.vehicle_id and latest_location.vehicle_id in vehicles_dict:
                vehicle = vehicles_dict[latest_location.vehicle_id]
                location_data['vehicle_name'] = vehicle.plate_number
    
    # تحويل الموظفين إلى قواميس لكي يمكن تحويلها إلى JSON
    employees_data = []
    for emp in employees:
        # Get employee department
        dept_name = emp.departments[0].name if emp.departments else 'غير محدد'
        
        # Get location data if exists
        location_data = employee_locations.get(emp.id)
        
        photo_url = None
        if emp.profile_image:
            # المسارات في قاعدة البيانات الآن بدون static/ في البداية
            # مثال: uploads/employees/1212_profile_20251111_075113.jpg
            photo_url = f"/{emp.profile_image}"
        
        emp_dict = {
            'id': emp.id,
            'name': emp.name,
            'employee_number': emp.employee_id,
            'photo_url': photo_url,
            'department_name': dept_name
        }
        employees_data.append(emp_dict)
    
    # تحويل employee_locations لكي تكون serializable
    employee_locations_json = {}
    for emp_id, loc_data in employee_locations.items():
        employee_locations_json[emp_id] = {
            'latitude': float(loc_data['latitude']),
            'longitude': float(loc_data['longitude']),
            'color': loc_data['color'],
            'status_text': loc_data['status_text'],
            'connection_status': loc_data.get('connection_status', 'disconnected'),
            'age_minutes': loc_data.get('age_minutes', 0),
            'geofence_name': loc_data.get('geofence_name'),
            'vehicle_name': loc_data.get('vehicle_name')
        }
    
    # جلب جميع الدوائر الجغرافية النشطة
    geofences = Geofence.query.filter_by(is_active=True).all()
    geofences_data = []
    for gf in geofences:
        geofences_data.append({
            'id': gf.id,
            'name': gf.name,
            'latitude': float(gf.center_latitude),
            'longitude': float(gf.center_longitude),
            'radius': gf.radius_meters
        })
    
    # جلب جميع الأقسام للفلترة
    departments = Department.query.all()
    
    import json
    
    return render_template(
        'employees/tracking.html',
        employees=employees,
        employee_locations=employee_locations,
        employees_json=json.dumps(employees_data, ensure_ascii=False),
        employee_locations_json=json.dumps(employee_locations_json, ensure_ascii=False),
        geofences_json=json.dumps(geofences_data, ensure_ascii=False),
        departments=departments,
        department_filter=department_filter,
        search_query=search_query
    )




@employees_bp.route('/tracking-dashboard')
@login_required
def tracking_dashboard():
    """لوحة تحكم مختصرة لإحصائيات التتبع المباشر"""
    from math import radians, sin, cos, sqrt, atan2
    from sqlalchemy import func as sql_func, and_
    
    cutoff_time_active = datetime.utcnow() - timedelta(hours=1)
    cutoff_time_location = datetime.utcnow() - timedelta(hours=24)
    
    # جلب جميع الموظفين مع الأقسام
    all_employees = Employee.query.options(db.joinedload(Employee.departments)).all()
    employee_ids = [emp.id for emp in all_employees]
    
    # جلب آخر موقع لكل موظف باستخدام window function
    latest_locations_subq = db.session.query(
        EmployeeLocation.employee_id,
        EmployeeLocation.id.label('location_id'),
        sql_func.row_number().over(
            partition_by=EmployeeLocation.employee_id,
            order_by=EmployeeLocation.recorded_at.desc()
        ).label('rn')
    ).filter(
        EmployeeLocation.employee_id.in_(employee_ids)
    ).subquery()
    
    latest_locations_query = db.session.query(
        EmployeeLocation
    ).join(
        latest_locations_subq,
        and_(
            EmployeeLocation.id == latest_locations_subq.c.location_id,
            latest_locations_subq.c.rn == 1
        )
    ).all()
    
    # بناء dictionary للمواقع
    locations_by_employee = {loc.employee_id: loc for loc in latest_locations_query}
    
    # معالجة الموظفين
    active_employees = []
    inactive_employees = []
    employees_with_vehicles = []
    
    for emp in all_employees:
        latest_location = locations_by_employee.get(emp.id)
        
        if latest_location and latest_location.recorded_at >= cutoff_time_active:
            active_employees.append({
                'employee': emp,
                'location': latest_location,
                'departments': [d.name for d in emp.departments]
            })
            
            if latest_location.vehicle_id:
                employees_with_vehicles.append({
                    'employee': emp,
                    'location': latest_location,
                    'vehicle': latest_location.vehicle
                })
        else:
            inactive_employees.append(emp)
    
    # جلب الـ geofences مرة واحدة
    all_geofences = Geofence.query.filter_by(is_active=True).all()
    
    employees_inside_geofences = []
    employees_outside_geofences = []
    geofence_stats = []
    employees_inside_any_geofence = set()
    
    # حساب المسافات لكل geofence
    for geofence in all_geofences:
        inside_count = 0
        inside_employees = []
        
        for emp in all_employees:
            latest_location = locations_by_employee.get(emp.id)
            
            if latest_location and latest_location.recorded_at >= cutoff_time_location:
                lat1, lon1 = float(latest_location.latitude), float(latest_location.longitude)
                lat2, lon2 = float(geofence.center_latitude), float(geofence.center_longitude)
                
                R = 6371000
                dlat = radians(lat2 - lat1)
                dlon = radians(lon2 - lon1)
                a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
                c = 2 * atan2(sqrt(a), sqrt(1-a))
                distance = R * c
                
                if distance <= geofence.radius_meters:
                    inside_count += 1
                    inside_employees.append({
                        'employee': emp,
                        'location': latest_location,
                        'distance': distance
                    })
                    employees_inside_any_geofence.add(emp.id)
        
        geofence_stats.append({
            'geofence': geofence,
            'inside_count': inside_count,
            'inside_employees': inside_employees
        })
    
    # تحديد الموظفين داخل وخارج الـ geofences
    for emp in all_employees:
        latest_location = locations_by_employee.get(emp.id)
        
        if latest_location and latest_location.recorded_at >= cutoff_time_location:
            if emp.id in employees_inside_any_geofence:
                if not any(e['employee'].id == emp.id for e in employees_inside_geofences):
                    employees_inside_geofences.append({
                        'employee': emp,
                        'location': latest_location
                    })
            else:
                employees_outside_geofences.append({
                    'employee': emp,
                    'location': latest_location
                })
    
    stats = {
        'total_employees': len(all_employees),
        'active_count': len(active_employees),
        'inactive_count': len(inactive_employees),
        'with_vehicles_count': len(employees_with_vehicles),
        'inside_geofences_count': len(employees_inside_geofences),
        'outside_geofences_count': len(employees_outside_geofences),
        'total_geofences': len(all_geofences)
    }
    
    return render_template(
        'employees/tracking_dashboard.html',
        stats=stats,
        active_employees=active_employees,
        inactive_employees=inactive_employees,
        employees_with_vehicles=employees_with_vehicles,
        employees_inside_geofences=employees_inside_geofences,
        employees_outside_geofences=employees_outside_geofences,
        geofence_stats=geofence_stats
    )


def format_time_12hr_arabic(dt):
    """تحويل الوقت إلى نظام 12 ساعة بالعربية"""
    hour = dt.hour
    minute = dt.minute
    second = dt.second
    
    if hour == 0:
        hour_12 = 12
        period = "ليلاً"
    elif hour < 12:
        hour_12 = hour
        period = "صباحاً"
    elif hour == 12:
        hour_12 = 12
        period = "ظهراً"
    else:
        hour_12 = hour - 12
        period = "مساءً"
    
    date_str = dt.strftime('%Y-%m-%d')
    return f"{date_str} {hour_12:02d}:{minute:02d}:{second:02d} {period}"


@employees_bp.route('/<int:id>/track-history')
@login_required
def track_history(id):
    """صفحة تتبع تحركات موظف واحد خلال 24 ساعة"""
    from flask import jsonify
    
    employee = Employee.query.get_or_404(id)
    
    employee_photo_url = None
    if employee.profile_image:
        if employee.profile_image.startswith('http'):
            employee_photo_url = employee.profile_image
        elif employee.profile_image.startswith('static/'):
            employee_photo_url = url_for('static', filename=employee.profile_image.replace('static/', ''), _external=False)
        elif employee.profile_image.startswith('uploads/'):
            employee_photo_url = url_for('static', filename=employee.profile_image, _external=False)
        else:
            employee_photo_url = url_for('static', filename=f'uploads/{employee.profile_image}', _external=False)
    
    cutoff_time = datetime.utcnow() - timedelta(hours=24)
    
    locations = EmployeeLocation.query.filter(
        EmployeeLocation.employee_id == id,
        EmployeeLocation.recorded_at >= cutoff_time
    ).order_by(EmployeeLocation.recorded_at.asc()).all()
    
    locations_data = []
    for loc in locations:
        loc_dict = {
            'latitude': float(loc.latitude),
            'longitude': float(loc.longitude),
            'speed': float(loc.speed_kmh) if loc.speed_kmh else 0,
            'vehicle_id': loc.vehicle_id,
            'recorded_at': format_time_12hr_arabic(loc.recorded_at),
            'accuracy': float(loc.accuracy_m) if loc.accuracy_m else None
        }
        
        if loc.vehicle_id and loc.vehicle:
            loc_dict['vehicle'] = {
                'id': loc.vehicle.id,
                'plate_number': loc.vehicle.plate_number,
                'make': loc.vehicle.make,
                'model': loc.vehicle.model
            }
        
        locations_data.append(loc_dict)
    
    departments = Department.query.all()
    
    return render_template(
        'employees/track_history.html',
        employee=employee,
        employee_photo_url=employee_photo_url,
        locations=locations_data,
        departments=departments
    )


@employees_bp.route('/<int:employee_id>/track-history/export-pdf')
@login_required
def export_track_history_pdf(employee_id):
    """تصدير سجل التحركات إلى PDF مع روابط قابلة للنقر"""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display
    import requests
    
    def prepare_arabic(text):
        """دالة مساعدة لمعالجة النص العربي بشكل صحيح"""
        if not text:
            return ""
        return get_display(reshape(str(text)))
    
    employee = Employee.query.get_or_404(employee_id)
    
    cutoff_time = datetime.utcnow() - timedelta(hours=24)
    locations = EmployeeLocation.query.filter(
        EmployeeLocation.employee_id == employee_id,
        EmployeeLocation.recorded_at >= cutoff_time
    ).order_by(EmployeeLocation.recorded_at.asc()).all()
    
    pdfmetrics.registerFont(TTFont('Amiri', 'static/fonts/Amiri-Regular.ttf'))
    pdfmetrics.registerFont(TTFont('AmiriBold', 'static/fonts/Amiri-Bold.ttf'))
    pdfmetrics.registerFontFamily('Amiri', normal='Amiri', bold='AmiriBold')
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    story = []
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=getSampleStyleSheet()['Heading1'],
        fontName='AmiriBold',
        fontSize=24,
        textColor=colors.HexColor('#1e1b4b'),
        alignment=TA_CENTER,
        spaceAfter=15,
        spaceBefore=10,
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=getSampleStyleSheet()['Heading2'],
        fontName='AmiriBold',
        fontSize=16,
        textColor=colors.HexColor('#4f46e5'),
        alignment=TA_RIGHT,
        spaceAfter=12,
        spaceBefore=8,
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=getSampleStyleSheet()['Normal'],
        fontName='Amiri',
        fontSize=12,
        alignment=TA_RIGHT,
        rightIndent=0,
        leftIndent=0,
        textColor=colors.HexColor('#374151'),
        leading=18,
    )
    
    title_text = prepare_arabic(f"سجل تحركات الموظف - {employee.name}")
    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 0.5*cm))
    
    info_data = [
        [prepare_arabic('رقم الموظف:'), prepare_arabic(str(employee.employee_id))],
        [prepare_arabic('الاسم:'), prepare_arabic(employee.name)],
        [prepare_arabic('عدد النقاط:'), str(len(locations))],
        [prepare_arabic('التاريخ:'), datetime.now().strftime('%Y-%m-%d %H:%M')],
    ]
    
    if employee.departments:
        info_data.insert(2, [prepare_arabic('القسم:'), prepare_arabic(employee.departments[0].name)])
    
    info_table = Table(info_data, colWidths=[4.5*cm, 12*cm])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Amiri', 12),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#4f46e5')),
        ('BACKGROUND', (1, 0), (1, -1), colors.white),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e1b4b')),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1.5, colors.HexColor('#c7d2fe')),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('ROUNDEDCORNERS', [5, 5, 5, 5]),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 1*cm))
    
    if locations and len(locations) > 0:
        max_speed = max([float(loc.speed_kmh) if loc.speed_kmh else 0 for loc in locations])
        total_distance = 0
        vehicle_count = sum([1 for loc in locations if loc.vehicle_id])
        
        for i in range(1, len(locations)):
            prev = locations[i-1]
            curr = locations[i]
            lat1, lon1 = float(prev.latitude), float(prev.longitude)
            lat2, lon2 = float(curr.latitude), float(curr.longitude)
            
            from math import radians, sin, cos, sqrt, atan2
            R = 6371
            dlat = radians(lat2 - lat1)
            dlon = radians(lon2 - lon1)
            a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
            c = 2 * atan2(sqrt(a), sqrt(1-a))
            total_distance += R * c
        
        subtitle = prepare_arabic('إحصائيات التحركات')
        story.append(Paragraph(subtitle, subtitle_style))
        story.append(Spacer(1, 0.3*cm))
        
        stats_data = [
            [prepare_arabic('إجمالي المسافة:'), f"{total_distance:.2f} " + prepare_arabic('كم')],
            [prepare_arabic('أقصى سرعة:'), f"{max_speed:.1f} " + prepare_arabic('كم/س')],
            [prepare_arabic('عدد النقاط على سيارة:'), str(vehicle_count)],
        ]
        
        stats_table = Table(stats_data, colWidths=[4.5*cm, 12*cm])
        stats_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Amiri', 12),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#10b981')),
            ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#d1fae5')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#065f46')),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1.5, colors.HexColor('#6ee7b7')),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 1*cm))
        
        subtitle2 = prepare_arabic('سجل التحركات التفصيلي')
        story.append(Paragraph(subtitle2, subtitle_style))
        story.append(Spacer(1, 0.3*cm))
        
        data = [[
            prepare_arabic('#'),
            prepare_arabic('الوقت'),
            prepare_arabic('الإحداثيات'),
            prepare_arabic('السرعة'),
            prepare_arabic('السيارة'),
        ]]
        
        for idx, loc in enumerate(locations, 1):
            coords = f"{float(loc.latitude):.6f}, {float(loc.longitude):.6f}"
            coords_link = f'<link href="https://www.google.com/maps?q={float(loc.latitude)},{float(loc.longitude)}" color="#2563eb"><u>{coords}</u></link>'
            
            speed_val = f"{float(loc.speed_kmh):.1f} " + prepare_arabic('كم/س') if loc.speed_kmh and float(loc.speed_kmh) > 0 else "-"
            
            vehicle_info = "-"
            if loc.vehicle_id and loc.vehicle:
                vehicle_info = prepare_arabic(f"{loc.vehicle.plate_number} - {loc.vehicle.make}")
            
            time_str = format_time_12hr_arabic(loc.recorded_at)
            
            data.append([
                str(idx),
                time_str,
                Paragraph(coords_link, normal_style),
                speed_val,
                vehicle_info,
            ])
        
        table = Table(data, colWidths=[1.2*cm, 3.5*cm, 5*cm, 3*cm, 4.5*cm])
        table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'AmiriBold', 13),
            ('FONT', (0, 1), (-1, -1), 'Amiri', 11),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e1b4b')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#c7d2fe')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eef2ff')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(table)
    else:
        no_data_text = prepare_arabic('لا توجد بيانات تتبع خلال آخر 24 ساعة')
        story.append(Paragraph(no_data_text, normal_style))
    
    doc.build(story)
    buffer.seek(0)
    
    filename = f"track_history_{employee.employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


@employees_bp.route('/<int:employee_id>/track-history/export-excel')
@login_required
def export_track_history_excel(employee_id):
    """تصدير سجل التحركات إلى Excel بتصميم احترافي مع صورة الخريطة"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill, Color
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.chart import BarChart, Reference, LineChart
    import requests
    from io import BytesIO
    from math import radians, sin, cos, sqrt, atan2
    import os
    
    employee = Employee.query.get_or_404(employee_id)
    
    cutoff_time = datetime.utcnow() - timedelta(hours=24)
    locations = EmployeeLocation.query.filter(
        EmployeeLocation.employee_id == employee_id,
        EmployeeLocation.recorded_at >= cutoff_time
    ).order_by(EmployeeLocation.recorded_at.asc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "سجل التحركات"
    ws.right_to_left = True
    
    ws.merge_cells('A1:J1')
    ws['A1'] = f"📍 سجل تحركات الموظف - {employee.name}"
    ws['A1'].font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
    ws['A1'].fill = GradientFill(stop=("4F46E5", "7C3AED"))
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:J2')
    ws['A2'] = f"التقرير الشامل - {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(name='Arial', size=12, italic=True, color='6366F1')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    current_row = 4
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "📋 معلومات الموظف"
    ws[f'A{current_row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws[f'A{current_row}'].fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 28
    
    current_row += 1
    info_data = [
        ['رقم الموظف:', employee.employee_id, 'الاسم:', employee.name],
        ['القسم:', employee.departments[0].name if employee.departments else '-', 'تاريخ التقرير:', format_time_12hr_arabic(datetime.now())],
    ]
    
    for row_data in info_data:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin', color='C7D2FE'),
                right=Side(style='thin', color='C7D2FE'),
                top=Side(style='thin', color='C7D2FE'),
                bottom=Side(style='thin', color='C7D2FE')
            )
            if col_idx % 2 == 1:
                cell.font = Font(name='Arial', size=11, bold=True, color='312E81')
                cell.fill = PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid')
            else:
                cell.font = Font(name='Arial', size=11, color='1E1B4B')
                cell.fill = PatternFill(start_color='F5F7FF', end_color='F5F7FF', fill_type='solid')
        current_row += 1
    
    current_row += 1
    
    if locations and len(locations) > 0:
        max_speed = max([float(loc.speed_kmh) if loc.speed_kmh else 0 for loc in locations])
        total_distance = 0
        vehicle_count = sum([1 for loc in locations if loc.vehicle_id])
        
        for i in range(1, len(locations)):
            prev = locations[i-1]
            curr = locations[i]
            lat1, lon1 = float(prev.latitude), float(prev.longitude)
            lat2, lon2 = float(curr.latitude), float(curr.longitude)
            
            R = 6371
            dlat = radians(lat2 - lat1)
            dlon = radians(lon2 - lon1)
            a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
            c = 2 * atan2(sqrt(a), sqrt(1-a))
            total_distance += R * c
        
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "📊 إحصائيات التحركات"
        ws[f'A{current_row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws[f'A{current_row}'].fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 28
        
        current_row += 1
        stats_data = [
            ['عدد نقاط التتبع:', len(locations), 'إجمالي المسافة:', f"{total_distance:.2f} كم"],
            ['أقصى سرعة:', f"{max_speed:.1f} كم/س", 'نقاط على سيارة:', vehicle_count],
        ]
        
        for row_data in stats_data:
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='medium', color='10B981'),
                    right=Side(style='medium', color='10B981'),
                    top=Side(style='thin', color='D1FAE5'),
                    bottom=Side(style='thin', color='D1FAE5')
                )
                if col_idx % 2 == 1:
                    cell.font = Font(name='Arial', size=12, bold=True, color='065F46')
                    cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                else:
                    cell.font = Font(name='Arial', size=13, bold=True, color='059669')
                    cell.fill = PatternFill(start_color='A7F3D0', end_color='A7F3D0', fill_type='solid')
            current_row += 1
        
        current_row += 1
        
        map_row = current_row
        ws.merge_cells(f'F{map_row}:J{map_row}')
        ws[f'F{map_row}'] = "🗺️ خريطة المسار"
        ws[f'F{map_row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws[f'F{map_row}'].fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
        ws[f'F{map_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[map_row].height = 28
        
        if len(locations) > 0:
            lats = [float(loc.latitude) for loc in locations]
            lons = [float(loc.longitude) for loc in locations]
            center_lat = sum(lats) / len(lats)
            center_lon = sum(lons) / len(lons)
            
            min_lat, max_lat = min(lats), max(lats)
            min_lon, max_lon = min(lons), max(lons)
            
            try:
                zoom_level = 12
                lat_diff = max_lat - min_lat
                lon_diff = max_lon - min_lon
                
                if lat_diff > 0.5 or lon_diff > 0.5:
                    zoom_level = 10
                elif lat_diff < 0.05 and lon_diff < 0.05:
                    zoom_level = 14
                
                markers = ""
                for i, (lat, lon) in enumerate(zip(lats, lons)):
                    if i == 0:
                        markers += f"{center_lat},{center_lon},lightblue1"
                    elif i == len(lats) - 1:
                        markers += f"|{lat},{lon},lightblue2"
                    elif i % 5 == 0:
                        markers += f"|{lat},{lon},lightblue3"
                
                map_url = f"https://staticmap.openstreetmap.de/staticmap.php?center={center_lat},{center_lon}&zoom={zoom_level}&size=800x500&maptype=mapnik&markers={markers}"
                
                response = requests.get(map_url, timeout=15, headers={'User-Agent': 'Mozilla/5.0'})
                if response.status_code == 200 and len(response.content) > 1000:
                    # حفظ الصورة بشكل دائم بدلاً من المؤقت
                    map_upload_folder = os.path.join(UPLOAD_FOLDER, 'maps')
                    os.makedirs(map_upload_folder, exist_ok=True)
                    
                    map_filename = f"track_map_{employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                    map_filepath = os.path.join(map_upload_folder, map_filename)
                    
                    with open(map_filepath, 'wb') as f:
                        f.write(response.content)
                    
                    # التحقق من أن الملف تم حفظه بنجاح
                    if os.path.exists(map_filepath) and os.path.getsize(map_filepath) > 0:
                        img = XLImage(map_filepath)
                        img.width = 450
                        img.height = 300
                        
                        ws.add_image(img, f'F{map_row + 1}')
                        
                        for i in range(map_row + 1, map_row + 16):
                            ws.row_dimensions[i].height = 20
                    else:
                        raise Exception("فشل في حفظ صورة الخريطة")
                else:
                    raise Exception("فشل تحميل الخريطة")
            except Exception as e:
                ws.merge_cells(f'F{map_row + 1}:J{map_row + 5}')
                ws[f'F{map_row + 1}'] = f"🗺️ عرض الخريطة\n\nانقر هنا لفتح الخريطة في Google Maps"
                ws[f'F{map_row + 1}'].font = Font(name='Arial', size=12, bold=True, color='2563EB', underline='single')
                ws[f'F{map_row + 1}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                google_maps_url = f"https://www.google.com/maps/dir/{lats[0]},{lons[0]}/{lats[-1]},{lons[-1]}"
                ws[f'F{map_row + 1}'].hyperlink = google_maps_url
                ws[f'F{map_row + 1}'].fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
                
                for i in range(map_row + 1, map_row + 6):
                    ws.row_dimensions[i].height = 25
        
        table_start_row = max(current_row + 1, map_row + 17)
        
        ws.merge_cells(f'A{table_start_row}:J{table_start_row}')
        ws[f'A{table_start_row}'] = "📝 سجل التحركات التفصيلي"
        ws[f'A{table_start_row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws[f'A{table_start_row}'].fill = PatternFill(start_color='EC4899', end_color='EC4899', fill_type='solid')
        ws[f'A{table_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[table_start_row].height = 28
        
        table_start_row += 1
        
        headers = ['#', 'الوقت', 'خط العرض', 'خط الطول', 'السرعة (كم/س)', 'الحالة', 'السيارة', 'الدقة (م)', 'رابط الموقع', 'ملاحظات']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start_row, column=col_idx)
            cell.value = header
            cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='medium', color='312E81'),
                right=Side(style='medium', color='312E81'),
                top=Side(style='medium', color='312E81'),
                bottom=Side(style='medium', color='312E81')
            )
        ws.row_dimensions[table_start_row].height = 30
        
        for idx, loc in enumerate(locations, 1):
            row = table_start_row + idx
            
            speed_val = float(loc.speed_kmh) if loc.speed_kmh else 0
            
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = format_time_12hr_arabic(loc.recorded_at)
            ws.cell(row=row, column=3).value = float(loc.latitude)
            ws.cell(row=row, column=4).value = float(loc.longitude)
            ws.cell(row=row, column=5).value = f"{speed_val:.1f}" if speed_val > 0 else "-"
            
            if speed_val > 100:
                ws.cell(row=row, column=6).value = "⚠️ سرعة عالية"
                status_color = 'FEE2E2'
            elif speed_val > 60:
                ws.cell(row=row, column=6).value = "⚡ متوسطة"
                status_color = 'FEF3C7'
            elif speed_val > 0:
                ws.cell(row=row, column=6).value = "✅ عادية"
                status_color = 'D1FAE5'
            else:
                ws.cell(row=row, column=6).value = "⏸️ متوقف"
                status_color = 'E0E7FF'
            
            if loc.vehicle_id and loc.vehicle:
                ws.cell(row=row, column=7).value = f"🚗 {loc.vehicle.plate_number} - {loc.vehicle.make}"
            else:
                ws.cell(row=row, column=7).value = "-"
            
            ws.cell(row=row, column=8).value = f"{float(loc.accuracy_m):.1f}" if loc.accuracy_m else "-"
            
            maps_link = f"https://www.google.com/maps?q={float(loc.latitude)},{float(loc.longitude)}"
            ws.cell(row=row, column=9).value = "📍 عرض الموقع"
            ws.cell(row=row, column=9).hyperlink = maps_link
            ws.cell(row=row, column=9).font = Font(name='Arial', size=10, color='2563EB', underline='single', bold=True)
            
            if speed_val > 120:
                ws.cell(row=row, column=10).value = "⚠️ تجاوز السرعة القصوى"
            elif loc.accuracy_m and float(loc.accuracy_m) > 50:
                ws.cell(row=row, column=10).value = "⚠️ دقة منخفضة"
            else:
                ws.cell(row=row, column=10).value = "-"
            
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin', color='C7D2FE'),
                    right=Side(style='thin', color='C7D2FE'),
                    top=Side(style='thin', color='C7D2FE'),
                    bottom=Side(style='thin', color='C7D2FE')
                )
                
                if col == 6:
                    cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
                elif idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F5F7FF', end_color='F5F7FF', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                if col == 5 and speed_val > 100:
                    cell.font = Font(name='Arial', size=11, bold=True, color='DC2626')
            
            ws.row_dimensions[row].height = 22
    else:
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = "⚠️ لا توجد بيانات تتبع خلال آخر 24 ساعة"
        ws[f'A{current_row}'].font = Font(name='Arial', size=14, bold=True, color='DC2626')
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 40
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"track_history_{employee.employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
