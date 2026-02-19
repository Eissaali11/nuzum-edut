"""
لوحة معلومات Power BI احترافية - نُظم
تحليلات متقدمة للحضور والوثائق والسيارات
تصدير Excel بتصميم احترافي
"""
from flask import Blueprint, render_template, request, jsonify, Response, send_file
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from core.extensions import db
from models import Employee, Attendance, Document, Vehicle, Department
from sqlalchemy import func, or_, and_, case
from utils.user_helpers import require_module_access
from models import Module, Permission
import csv
from io import StringIO, BytesIO
import json

powerbi_bp = Blueprint('powerbi', __name__, url_prefix='/powerbi')

@powerbi_bp.route('/')
@login_required
def dashboard():
    """الصفحة الرئيسية للوحة معلومات Power BI الاحترافية - بيانات حقيقية"""
    from datetime import datetime, timedelta
    
    # البيانات الأساسية
    departments = Department.query.all()
    total_vehicles = Vehicle.query.count()
    total_documents = Document.query.count()
    
    # فلاتر التاريخ
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    department_id = request.args.get('department_id')
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except:
            date_from = datetime.now().date() - timedelta(days=30)
    else:
        date_from = datetime.now().date() - timedelta(days=30)
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except:
            date_to = datetime.now().date()
    else:
        date_to = datetime.now().date()
    
    # تصحيح التواريخ إذا كانت معكوسة
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    
    # تحديد الأقسام المستهدفة حسب الفلتر
    if department_id:
        try:
            dept_id = int(department_id)
            target_departments = [d for d in departments if d.id == dept_id]
        except:
            target_departments = departments
    else:
        target_departments = departments
    
    # جلب موظفي الأقسام المستهدفة النشطين
    if department_id:
        try:
            dept_id = int(department_id)
            active_emp_query = db.session.query(Employee.id).join(
                Employee.departments
            ).filter(
                Department.id == dept_id,
                Employee.status == 'active'
            )
            active_emp_ids = [e[0] for e in active_emp_query.all()]
        except:
            active_emp_ids = [e.id for e in Employee.query.filter(Employee.status == 'active').all()]
    else:
        active_emp_ids = [e.id for e in Employee.query.filter(Employee.status == 'active').all()]
    
    # جلب الموظفين النشطين فقط الذين لهم حضور في الفترة المحددة
    active_employee_ids_with_attendance = db.session.query(Attendance.employee_id).filter(
        Attendance.date >= date_from,
        Attendance.date <= date_to,
        Attendance.employee_id.in_(active_emp_ids) if active_emp_ids else Attendance.employee_id.isnot(None)
    ).distinct().all()
    active_employee_ids_with_attendance = [e[0] for e in active_employee_ids_with_attendance]
    
    # عدد الموظفين النشطين الذين لهم حضور
    active_employees_count = Employee.query.filter(
        Employee.status == 'active',
        Employee.id.in_(active_employee_ids_with_attendance)
    ).count()
    
    total_employees = active_employees_count
    
    # سجلات الحضور للموظفين النشطين فقط
    attendance_records = Attendance.query.filter(
        Attendance.date >= date_from,
        Attendance.date <= date_to,
        Attendance.employee_id.in_(active_emp_ids) if active_emp_ids else Attendance.employee_id.isnot(None)
    ).all()
    
    attendance_stats = {
        'present': sum(1 for a in attendance_records if a.status == 'present'),
        'absent': sum(1 for a in attendance_records if a.status in ['absent', 'غائب']),
        'leave': sum(1 for a in attendance_records if a.status == 'leave'),
        'sick': sum(1 for a in attendance_records if a.status == 'sick'),
        'total': len(attendance_records)
    }
    attendance_stats['rate'] = round((attendance_stats['present'] / attendance_stats['total']) * 100, 1) if attendance_stats['total'] > 0 else 0
    
    # إحصائيات السيارات
    vehicles = Vehicle.query.all()
    vehicle_stats = {}
    for v in vehicles:
        status = v.status or 'unknown'
        vehicle_stats[status] = vehicle_stats.get(status, 0) + 1
    
    # الحضور حسب القسم - الموظفين النشطين فقط الذين لهم حضور
    dept_attendance = []
    for dept in target_departments:
        # جلب موظفي القسم النشطين فقط
        emp_ids = db.session.query(Employee.id).join(
            Employee.departments
        ).filter(
            Department.id == dept.id,
            Employee.status == 'active'
        ).all()
        emp_ids = [e[0] for e in emp_ids]
        
        if not emp_ids:
            continue
        
        # الموظفين الذين لهم حضور فعلي في الفترة
        emp_ids_with_attendance = db.session.query(Attendance.employee_id).filter(
            Attendance.date >= date_from,
            Attendance.date <= date_to,
            Attendance.employee_id.in_(emp_ids)
        ).distinct().all()
        emp_ids_with_attendance = [e[0] for e in emp_ids_with_attendance]
        
        if not emp_ids_with_attendance:
            continue
            
        # حساب الحضور
        dept_records = Attendance.query.filter(
            Attendance.date >= date_from,
            Attendance.date <= date_to,
            Attendance.employee_id.in_(emp_ids_with_attendance)
        ).all()
        
        present = sum(1 for a in dept_records if a.status == 'present')
        absent = sum(1 for a in dept_records if a.status in ['absent', 'غائب'])
        total = len(dept_records)
        rate = round((present / total) * 100, 1) if total > 0 else 0
        
        # حساب عدد أيام الحضور الفريدة
        present_days = len(set(a.date for a in dept_records if a.status == 'present'))
        
        dept_attendance.append({
            'name': dept.name,
            'employee_count': len(emp_ids_with_attendance),
            'present': present,
            'absent': absent,
            'present_days': present_days,
            'total': total,
            'rate': rate
        })
    
    # ترتيب حسب نسبة الحضور
    dept_attendance.sort(key=lambda x: x['rate'], reverse=True)
    
    # طباعة للتصحيح
    import logging
    logging.info(f"[POWERBI] dept_attendance count: {len(dept_attendance)}")
    for da in dept_attendance[:3]:
        logging.info(f"[POWERBI] dept: {da['name']}, emp: {da['employee_count']}, rate: {da['rate']}")
    
    # توزيع الموظفين حسب القسم (للرسم البياني) - الموظفين النشطين الذين لهم حضور فقط
    dept_distribution = []
    for dept in departments:
        emp_count = 0
        for da in dept_attendance:
            if da['name'] == dept.name:
                emp_count = da['employee_count']
                break
        if emp_count > 0:
            dept_distribution.append({
                'name': dept.name,
                'count': emp_count
            })
    
    # ترتيب حسب العدد
    dept_distribution.sort(key=lambda x: x['count'], reverse=True)
    
    # إحصائيات الوثائق
    today = datetime.now().date()
    thirty_days = today + timedelta(days=30)
    
    docs = Document.query.all()
    doc_stats = {
        'valid': 0,
        'expiring': 0,
        'expired': 0,
        'total': len(docs)
    }
    
    for doc in docs:
        if hasattr(doc, 'expiry_date') and doc.expiry_date:
            if doc.expiry_date < today:
                doc_stats['expired'] += 1
            elif doc.expiry_date <= thirty_days:
                doc_stats['expiring'] += 1
            else:
                doc_stats['valid'] += 1
        else:
            doc_stats['valid'] += 1
    
    return render_template('powerbi/dashboard_enhanced.html',
        departments=departments,
        total_employees=total_employees,
        total_vehicles=total_vehicles,
        total_documents=total_documents,
        attendance_stats=attendance_stats,
        vehicle_stats=vehicle_stats,
        dept_attendance=dept_attendance,
        dept_distribution=dept_distribution,
        doc_stats=doc_stats,
        date_from=date_from,
        date_to=date_to
    )

@powerbi_bp.route('/api/attendance-summary')
@login_required
@require_module_access(Module.ATTENDANCE, Permission.VIEW)
def attendance_summary():
    """ملخص الحضور مع تحليلات متقدمة"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    department_id = request.args.get('department_id')
    
    try:
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = datetime.now().date() - timedelta(days=30)
        
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = datetime.now().date()
        
        query = Attendance.query.filter(
            Attendance.date >= date_from,
            Attendance.date <= date_to
        )
        
        if department_id:
            employee_ids = [e.id for e in Employee.query.filter_by(department_id=department_id).all()]
            query = query.filter(Attendance.employee_id.in_(employee_ids))
        
        attendance_records = query.all()
        
        present = sum(1 for a in attendance_records if a.status == 'present')
        absent = sum(1 for a in attendance_records if a.status in ['absent', 'غائب'])
        leave = sum(1 for a in attendance_records if a.status == 'leave')
        sick = sum(1 for a in attendance_records if a.status == 'sick')
        total = len(attendance_records)
        
        attendance_rate = round((present / total) * 100, 1) if total > 0 else 0
        absence_rate = round((absent / total) * 100, 1) if total > 0 else 0
        
        prev_date_from = date_from - timedelta(days=30)
        prev_query = Attendance.query.filter(
            Attendance.date >= prev_date_from,
            Attendance.date < date_from
        )
        if department_id:
            prev_query = prev_query.filter(Attendance.employee_id.in_(employee_ids))
        
        prev_records = prev_query.all()
        prev_present = sum(1 for a in prev_records if a.status == 'present')
        prev_total = len(prev_records)
        prev_rate = round((prev_present / prev_total) * 100, 1) if prev_total > 0 else 0
        
        trend = round(attendance_rate - prev_rate, 1)
        trend_direction = 'up' if trend > 0 else 'down' if trend < 0 else 'stable'
        
        return jsonify({
            'success': True,
            'data': {
                'present': present,
                'absent': absent,
                'leave': leave,
                'sick': sick,
                'total': total,
                'attendance_rate': attendance_rate,
                'absence_rate': absence_rate,
                'trend': {
                    'value': abs(trend),
                    'direction': trend_direction,
                    'previous_rate': prev_rate
                },
                'date_range': {
                    'from': date_from.strftime('%Y-%m-%d'),
                    'to': date_to.strftime('%Y-%m-%d')
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/attendance-by-department')
@login_required
@require_module_access(Module.ATTENDANCE, Permission.VIEW)
def attendance_by_department():
    """الحضور حسب القسم مع تحليل مفصل"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    try:
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = datetime.now().date() - timedelta(days=30)
        
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = datetime.now().date()
        
        departments_data = []
        departments = Department.query.all()
        
        for dept in departments:
            employees = Employee.query.filter_by(department_id=dept.id).all()
            employee_ids = [e.id for e in employees]
            
            if not employee_ids:
                continue
            
            attendance_records = Attendance.query.filter(
                Attendance.date >= date_from,
                Attendance.date <= date_to,
                Attendance.employee_id.in_(employee_ids)
            ).all()
            
            present = sum(1 for a in attendance_records if a.status == 'present')
            absent = sum(1 for a in attendance_records if a.status in ['absent', 'غائب'])
            leave = sum(1 for a in attendance_records if a.status == 'leave')
            sick = sum(1 for a in attendance_records if a.status == 'sick')
            total = len(attendance_records)
            
            attendance_rate = round((present / total) * 100, 1) if total > 0 else 0
            
            performance = 'excellent' if attendance_rate >= 90 else 'good' if attendance_rate >= 75 else 'average' if attendance_rate >= 60 else 'poor'
            
            departments_data.append({
                'department': dept.name,
                'department_id': dept.id,
                'employee_count': len(employees),
                'present': present,
                'absent': absent,
                'leave': leave,
                'sick': sick,
                'total': total,
                'attendance_rate': attendance_rate,
                'performance': performance
            })
        
        departments_data.sort(key=lambda x: x['attendance_rate'], reverse=True)
        
        return jsonify({
            'success': True,
            'data': departments_data,
            'summary': {
                'total_departments': len(departments_data),
                'avg_attendance_rate': round(sum(d['attendance_rate'] for d in departments_data) / len(departments_data), 1) if departments_data else 0,
                'best_department': departments_data[0]['department'] if departments_data else None,
                'worst_department': departments_data[-1]['department'] if departments_data else None
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/documents-status')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def documents_status():
    """حالة الوثائق المطلوبة مع تحليل شامل"""
    try:
        required_docs = [
            {'type': 'الهوية الوطنية', 'priority': 'high'},
            {'type': 'جواز السفر', 'priority': 'high'},
            {'type': 'رخصة القيادة', 'priority': 'medium'},
            {'type': 'وثيقة التأمين', 'priority': 'medium'}
        ]
        
        total_employees = Employee.query.count()
        documents_summary = []
        
        today = datetime.now().date()
        thirty_days_later = today + timedelta(days=30)
        
        for doc_info in required_docs:
            doc_type = doc_info['type']
            
            docs = Document.query.filter(
                Document.document_type.ilike(f'%{doc_type}%')
            ).all()
            
            available = len(docs)
            missing = total_employees - available
            
            expiring_soon = 0
            expired = 0
            valid = 0
            
            for doc in docs:
                if hasattr(doc, 'expiry_date') and doc.expiry_date:
                    if doc.expiry_date < today:
                        expired += 1
                    elif doc.expiry_date <= thirty_days_later:
                        expiring_soon += 1
                    else:
                        valid += 1
                else:
                    valid += 1
            
            documents_summary.append({
                'type': doc_type,
                'priority': doc_info['priority'],
                'available': available,
                'missing': missing,
                'valid': valid,
                'expiring_soon': expiring_soon,
                'expired': expired,
                'completion_rate': round((available / total_employees * 100), 1) if total_employees > 0 else 0,
                'health_score': round(((valid) / available * 100), 1) if available > 0 else 0
            })
        
        total_available = sum(d['available'] for d in documents_summary)
        total_required = total_employees * len(required_docs)
        overall_completion = round((total_available / total_required * 100), 1) if total_required > 0 else 0
        
        return jsonify({
            'success': True,
            'data': documents_summary,
            'total_employees': total_employees,
            'overall_completion': overall_completion,
            'total_expiring_soon': sum(d['expiring_soon'] for d in documents_summary),
            'total_expired': sum(d['expired'] for d in documents_summary),
            'total_missing': sum(d['missing'] for d in documents_summary)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/employee-documents')
@login_required
@require_module_access(Module.EMPLOYEES, Permission.VIEW)
def employee_documents():
    """قائمة الموظفين والوثائق الناقصة مع تفاصيل"""
    department_id = request.args.get('department_id')
    limit = request.args.get('limit', 50, type=int)
    
    try:
        query = Employee.query
        
        if department_id:
            query = query.filter_by(department_id=department_id)
        
        employees = query.limit(limit).all()
        
        required_docs = ['الهوية الوطنية', 'جواز السفر', 'رخصة القيادة', 'وثيقة التأمين']
        employees_data = []
        
        for emp in employees:
            docs = Document.query.filter_by(employee_id=emp.id).all()
            doc_types = [d.document_type for d in docs]
            
            missing_docs = []
            for req_doc in required_docs:
                if not any(req_doc in dt for dt in doc_types):
                    missing_docs.append(req_doc)
            
            completion_rate = round(((len(required_docs) - len(missing_docs)) / len(required_docs)) * 100, 0)
            
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id or '-',
                'department': emp.department.name if emp.department else 'بدون قسم',
                'total_docs': len(docs),
                'missing_docs': missing_docs,
                'missing_count': len(missing_docs),
                'documents_complete': len(missing_docs) == 0,
                'completion_rate': completion_rate
            })
        
        employees_data.sort(key=lambda x: x['missing_count'], reverse=True)
        
        complete_count = sum(1 for e in employees_data if e['documents_complete'])
        incomplete_count = len(employees_data) - complete_count
        
        return jsonify({
            'success': True,
            'data': employees_data,
            'summary': {
                'total': len(employees_data),
                'complete': complete_count,
                'incomplete': incomplete_count,
                'completion_rate': round((complete_count / len(employees_data)) * 100, 1) if employees_data else 0
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/vehicles-summary')
@login_required
@require_module_access(Module.VEHICLES, Permission.VIEW)
def vehicles_summary():
    """ملخص حالة السيارات مع تحليل الأسطول"""
    try:
        vehicles = Vehicle.query.all()
        
        statuses = {}
        brands = {}
        years = {}
        
        for vehicle in vehicles:
            status = vehicle.status or 'unknown'
            statuses[status] = statuses.get(status, 0) + 1
            
            if hasattr(vehicle, 'make') and vehicle.make:
                brands[vehicle.make] = brands.get(vehicle.make, 0) + 1
            
            if hasattr(vehicle, 'year') and vehicle.year:
                years[str(vehicle.year)] = years.get(str(vehicle.year), 0) + 1
        
        vehicles_by_status = []
        status_labels = {
            'in_project': 'في المشروع',
            'in_workshop': 'في الورشة',
            'out_of_service': 'خارج الخدمة',
            'accident': 'حادث',
            'unknown': 'غير محدد'
        }
        
        for status, count in statuses.items():
            vehicles_by_status.append({
                'status': status_labels.get(status, status),
                'status_key': status,
                'count': count,
                'percentage': round((count / len(vehicles) * 100), 1) if vehicles else 0
            })
        
        vehicles_by_brand = [{'brand': b, 'count': c} for b, c in sorted(brands.items(), key=lambda x: x[1], reverse=True)]
        
        total = len(vehicles)
        in_project = statuses.get('in_project', 0)
        in_workshop = statuses.get('in_workshop', 0)
        out_of_service = statuses.get('out_of_service', 0)
        accident = statuses.get('accident', 0)
        
        fleet_health = 'excellent' if in_project >= total * 0.8 else 'good' if in_project >= total * 0.6 else 'average' if in_project >= total * 0.4 else 'poor'
        
        return jsonify({
            'success': True,
            'data': {
                'by_status': vehicles_by_status,
                'by_brand': vehicles_by_brand[:5],
                'total_vehicles': total,
                'in_project': in_project,
                'in_workshop': in_workshop,
                'out_of_service': out_of_service,
                'accident': accident,
                'utilization_rate': round((in_project / total) * 100, 1) if total > 0 else 0,
                'fleet_health': fleet_health
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/vehicle-operations-summary')
@login_required
@require_module_access(Module.VEHICLES, Permission.VIEW)
def vehicle_operations_summary():
    """ملخص عمليات السيارات"""
    try:
        vehicles = Vehicle.query.all()
        
        handovers = sum(1 for v in vehicles if v.handover_records)
        in_workshop = sum(1 for v in vehicles if v.status == 'in_workshop')
        in_project = sum(1 for v in vehicles if v.status == 'in_project')
        out_of_service = sum(1 for v in vehicles if v.status == 'out_of_service')
        accident = sum(1 for v in vehicles if v.status == 'accident')
        
        operations_data = [
            {'type': 'في المشروع', 'count': in_project, 'color': '#38ef7d'},
            {'type': 'مستلمة', 'count': handovers, 'color': '#667eea'},
            {'type': 'في الورشة', 'count': in_workshop, 'color': '#fbbf24'},
            {'type': 'خارج الخدمة', 'count': out_of_service, 'color': '#ef4444'},
            {'type': 'حادث', 'count': accident, 'color': '#ff6b6b'}
        ]
        
        return jsonify({
            'success': True,
            'data': {
                'by_type': operations_data,
                'total_vehicles': len(vehicles),
                'summary': {
                    'active_percentage': round((in_project / len(vehicles)) * 100, 1) if vehicles else 0,
                    'handover_percentage': round((handovers / len(vehicles)) * 100, 1) if vehicles else 0
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/export-data')
@login_required
def export_data():
    """تصدير البيانات بصيغة Excel احترافية بتصميم Power BI - نفس تصميم الصفحة"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference
    from openpyxl.chart.label import DataLabelList
    from sqlalchemy import func
    
    data_type = request.args.get('type', 'all')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    department_id = request.args.get('department_id')
    
    try:
        wb = Workbook()
        
        # ألوان التصميم المميز - Professional Dark Theme
        dark_bg = PatternFill(start_color="0A0E17", end_color="0A0E17", fill_type="solid")
        card_fill = PatternFill(start_color="131B2E", end_color="131B2E", fill_type="solid")
        card_fill_alt = PatternFill(start_color="1A2540", end_color="1A2540", fill_type="solid")
        header_fill = PatternFill(start_color="0D1321", end_color="0D1321", fill_type="solid")
        title_bg = PatternFill(start_color="00D4AA", end_color="00D4AA", fill_type="solid")
        
        # ألوان cyan/teal gradient effect
        cyan_fill = PatternFill(start_color="00D4FF", end_color="00D4FF", fill_type="solid")
        teal_fill = PatternFill(start_color="00F5D4", end_color="00F5D4", fill_type="solid")
        cyan_dark = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
        
        # ألوان الحالات - Professional
        green_fill = PatternFill(start_color="00FF88", end_color="00FF88", fill_type="solid")
        green_dark = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        red_fill = PatternFill(start_color="FF4757", end_color="FF4757", fill_type="solid")
        orange_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        blue_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        purple_fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type="solid")
        
        # خطوط احترافية
        title_font = Font(bold=True, color="00D4FF", size=28, name='Arial')
        subtitle_font = Font(color="8892A0", size=11, name='Arial')
        section_font = Font(bold=True, color="00D4FF", size=16, name='Arial')
        cyan_font = Font(bold=True, color="00D4FF", size=13, name='Arial')
        white_font = Font(bold=True, color="FFFFFF", size=12, name='Arial')
        dark_font = Font(bold=True, color="0A0E17", size=12, name='Arial')
        kpi_value_font = Font(bold=True, color="00D4FF", size=32, name='Arial')
        kpi_label_font = Font(bold=True, color="8892A0", size=11, name='Arial')
        data_font = Font(color="E8EAED", size=11, name='Arial')
        
        # حدود احترافية
        cyan_border = Border(
            left=Side(style='medium', color='00D4FF'),
            right=Side(style='medium', color='00D4FF'),
            top=Side(style='medium', color='00D4FF'),
            bottom=Side(style='medium', color='00D4FF')
        )
        cyan_border_thin = Border(
            left=Side(style='thin', color='00D4FF'),
            right=Side(style='thin', color='00D4FF'),
            top=Side(style='thin', color='00D4FF'),
            bottom=Side(style='thin', color='00D4FF')
        )
        thin_border = Border(
            left=Side(style='thin', color='2D3748'),
            right=Side(style='thin', color='2D3748'),
            top=Side(style='thin', color='2D3748'),
            bottom=Side(style='thin', color='2D3748')
        )
        accent_border = Border(
            bottom=Side(style='medium', color='00D4FF')
        )
        
        # معالجة التواريخ
        d_from = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else (datetime.now().date() - timedelta(days=30))
        d_to = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else datetime.now().date()
        
        # جلب الموظفين النشطين الذين لهم حضور في الفترة
        active_employee_ids_with_attendance = db.session.query(Attendance.employee_id).filter(
            Attendance.date >= d_from,
            Attendance.date <= d_to
        ).distinct().all()
        active_employee_ids_with_attendance = [e[0] for e in active_employee_ids_with_attendance]
        
        # عدد الموظفين النشطين الذين لهم حضور
        total_employees = Employee.query.filter(
            Employee.status == 'active',
            Employee.id.in_(active_employee_ids_with_attendance)
        ).count()
        
        # باقي الإحصائيات
        total_vehicles = Vehicle.query.count()
        in_project_vehicles = Vehicle.query.filter_by(status='in_project').count()
        in_workshop_vehicles = Vehicle.query.filter_by(status='in_workshop').count()
        out_of_service_vehicles = Vehicle.query.filter_by(status='out_of_service').count()
        accident_vehicles = Vehicle.query.filter_by(status='accident').count()
        total_documents = Document.query.count()
        total_departments = Department.query.count()
        
        # سجلات الحضور للموظفين النشطين فقط
        active_emp_ids = [e.id for e in Employee.query.filter(Employee.status == 'active').all()]
        
        attendance_records = Attendance.query.filter(
            Attendance.date >= d_from,
            Attendance.date <= d_to,
            Attendance.employee_id.in_(active_emp_ids)
        ).all()
        
        att_data = {
            'present': sum(1 for a in attendance_records if a.status == 'present'),
            'absent': sum(1 for a in attendance_records if a.status in ['absent', 'غائب']),
            'leave': sum(1 for a in attendance_records if a.status == 'leave'),
            'sick': sum(1 for a in attendance_records if a.status == 'sick')
        }
        total_attendance = sum(att_data.values())
        
        ws = wb.active
        ws.title = "Power BI Dashboard"
        ws.sheet_view.rightToLeft = True
        
        # تعبئة الخلفية الداكنة لكل الصفحة
        for row in range(1, 70):
            for col in range(1, 22):
                ws.cell(row=row, column=col).fill = dark_bg
        
        # عرض الأعمدة المتناسق
        column_widths = [18, 16, 14, 14, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # === الترويسة الرئيسية ===
        ws.merge_cells('A1:U1')
        ws['A1'] = "لوحة التحليلات الاحترافية | Power BI Dashboard"
        ws['A1'].font = title_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].border = accent_border
        ws.row_dimensions[1].height = 55
        
        # خط فاصل cyan
        for col in range(1, 22):
            ws.cell(row=2, column=col).fill = cyan_fill
            ws.cell(row=2, column=col).border = Border()
        ws.row_dimensions[2].height = 4
        
        # معلومات التقرير
        ws.merge_cells('A3:U3')
        ws['A3'] = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   الفترة: {d_from} إلى {d_to}   |   نُـظـم - نظام إدارة الموظفين"
        ws['A3'].font = subtitle_font
        ws['A3'].fill = card_fill
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 30
        
        # وصف البيانات
        ws.merge_cells('A4:U4')
        ws['A4'] = "البيانات تعكس الموظفين النشطين فقط الذين لديهم سجلات حضور في الفترة المحددة"
        ws['A4'].font = Font(color="00F5D4", size=10, italic=True, name='Arial')
        ws['A4'].fill = dark_bg
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[4].height = 22
        
        # === بطاقات KPI ===
        kpi_data = [
            ("الموظفين النشطين", total_employees, "👥", "A", cyan_fill),
            ("إجمالي السيارات", total_vehicles, "🚗", "E", teal_fill),
            ("في المشاريع", in_project_vehicles, "✅", "I", green_fill),
            ("الأقسام", total_departments, "🏢", "M", purple_fill),
            ("الوثائق", total_documents, "📄", "Q", orange_fill),
        ]
        
        ws.row_dimensions[5].height = 8
        ws.row_dimensions[6].height = 30
        ws.row_dimensions[7].height = 45
        ws.row_dimensions[8].height = 8
        
        for label, value, icon, start_col, accent_fill in kpi_data:
            col_idx = ord(start_col) - ord('A') + 1
            end_col = get_column_letter(col_idx + 3)
            
            # خط علوي ملون
            for c in range(col_idx, col_idx + 4):
                ws.cell(row=5, column=c).fill = accent_fill
            
            # التسمية
            ws.merge_cells(f'{start_col}6:{end_col}6')
            ws[f'{start_col}6'] = f"{icon}  {label}"
            ws[f'{start_col}6'].font = kpi_label_font
            ws[f'{start_col}6'].fill = card_fill
            ws[f'{start_col}6'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{start_col}6'].border = cyan_border_thin
            
            # القيمة
            ws.merge_cells(f'{start_col}7:{end_col}7')
            ws[f'{start_col}7'] = value
            ws[f'{start_col}7'].font = kpi_value_font
            ws[f'{start_col}7'].fill = card_fill
            ws[f'{start_col}7'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{start_col}7'].border = cyan_border_thin
            
            # خط سفلي
            for c in range(col_idx, col_idx + 4):
                ws.cell(row=8, column=c).fill = card_fill_alt
        
        # === قسم الحضور ===
        ws.merge_cells('A10:I10')
        ws['A10'] = "📊 توزيع حالات الحضور - الموظفين النشطين"
        ws['A10'].font = section_font
        ws['A10'].fill = header_fill
        ws['A10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A10'].border = accent_border
        ws.row_dimensions[10].height = 35
        
        att_headers = ['الحالة', 'العدد', 'النسبة', 'الرسم البياني']
        for i, h in enumerate(att_headers):
            cell = ws.cell(row=11, column=i+1)
            cell.value = h
            cell.font = cyan_font
            cell.fill = card_fill_alt
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cyan_border_thin
        ws.row_dimensions[11].height = 28
        
        att_rows = [
            ('حاضر ✅', att_data['present'], green_fill, "10B981"),
            ('غائب ❌', att_data['absent'], red_fill, "FF4757"),
            ('إجازة 📋', att_data['leave'], blue_fill, "3B82F6"),
            ('مريض 🏥', att_data['sick'], orange_fill, "FFD700")
        ]
        
        for idx, (label, count, fill, bar_color) in enumerate(att_rows, start=12):
            pct = round((count / total_attendance * 100), 1) if total_attendance > 0 else 0
            
            ws.cell(row=idx, column=1).value = label
            ws.cell(row=idx, column=1).font = white_font
            ws.cell(row=idx, column=1).fill = card_fill
            ws.cell(row=idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=idx, column=1).border = thin_border
            
            ws.cell(row=idx, column=2).value = count
            ws.cell(row=idx, column=2).font = Font(bold=True, color="00D4FF", size=14, name='Arial')
            ws.cell(row=idx, column=2).fill = card_fill
            ws.cell(row=idx, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=idx, column=2).border = thin_border
            
            ws.cell(row=idx, column=3).value = f"{pct}%"
            ws.cell(row=idx, column=3).font = Font(bold=True, color="00F5D4", size=12, name='Arial')
            ws.cell(row=idx, column=3).fill = card_fill
            ws.cell(row=idx, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=idx, column=3).border = thin_border
            
            bar_width = int(pct / 4) if pct > 0 else 1
            ws.cell(row=idx, column=4).value = "█" * bar_width
            ws.cell(row=idx, column=4).font = Font(color=bar_color, size=12, name='Arial')
            ws.cell(row=idx, column=4).fill = card_fill
            ws.cell(row=idx, column=4).border = thin_border
            
            ws.row_dimensions[idx].height = 26
        
        pie1 = PieChart()
        pie1.title = "توزيع الحضور"
        labels1 = Reference(ws, min_col=1, min_row=12, max_row=15)
        data1 = Reference(ws, min_col=2, min_row=11, max_row=15)
        pie1.add_data(data1, titles_from_data=True)
        pie1.set_categories(labels1)
        pie1.width = 10
        pie1.height = 7
        pie1.dataLabels = DataLabelList()
        pie1.dataLabels.showPercent = True
        pie1.dataLabels.showCatName = True
        ws.add_chart(pie1, "F11")
        
        # === قسم السيارات === (يبدأ من صف 24 لإعطاء مسافة للرسم البياني)
        veh_start_row = 24
        ws.merge_cells(f'A{veh_start_row}:E{veh_start_row}')
        ws[f'A{veh_start_row}'] = "🚗 حالة أسطول السيارات"
        ws[f'A{veh_start_row}'].font = section_font
        ws[f'A{veh_start_row}'].fill = header_fill
        ws[f'A{veh_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{veh_start_row}'].border = accent_border
        ws.row_dimensions[veh_start_row].height = 35
        
        veh_headers = ['الحالة', 'العدد', 'النسبة', 'الرسم']
        for i, h in enumerate(veh_headers):
            cell = ws.cell(row=veh_start_row+1, column=i+1)
            cell.value = h
            cell.font = cyan_font
            cell.fill = card_fill_alt
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cyan_border_thin
        ws.row_dimensions[veh_start_row+1].height = 28
        
        veh_rows = [
            ('في المشروع 🟢', in_project_vehicles, "10B981"),
            ('في الورشة 🟡', in_workshop_vehicles, "FFD700"),
            ('خارج الخدمة 🔴', out_of_service_vehicles, "FF4757"),
            ('حادث ⚠️', accident_vehicles, "7B68EE")
        ]
        
        for idx, (label, count, bar_color) in enumerate(veh_rows, start=veh_start_row+2):
            pct = round((count / total_vehicles * 100), 1) if total_vehicles > 0 else 0
            
            ws.cell(row=idx, column=1).value = label
            ws.cell(row=idx, column=1).font = white_font
            ws.cell(row=idx, column=1).fill = card_fill
            ws.cell(row=idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=idx, column=1).border = thin_border
            
            ws.cell(row=idx, column=2).value = count
            ws.cell(row=idx, column=2).font = Font(bold=True, color="00D4FF", size=14, name='Arial')
            ws.cell(row=idx, column=2).fill = card_fill
            ws.cell(row=idx, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=idx, column=2).border = thin_border
            
            ws.cell(row=idx, column=3).value = f"{pct}%"
            ws.cell(row=idx, column=3).font = Font(bold=True, color="00F5D4", size=12, name='Arial')
            ws.cell(row=idx, column=3).fill = card_fill
            ws.cell(row=idx, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=idx, column=3).border = thin_border
            
            bar_width = int(pct / 4) if pct > 0 else 1
            ws.cell(row=idx, column=4).value = "█" * bar_width
            ws.cell(row=idx, column=4).font = Font(color=bar_color, size=12, name='Arial')
            ws.cell(row=idx, column=4).fill = card_fill
            ws.cell(row=idx, column=4).border = thin_border
            
            ws.row_dimensions[idx].height = 26
        
        veh_end_row = veh_start_row + 5
        
        doughnut1 = DoughnutChart()
        doughnut1.title = "حالة الأسطول"
        labels2 = Reference(ws, min_col=1, min_row=veh_start_row+2, max_row=veh_end_row)
        data2 = Reference(ws, min_col=2, min_row=veh_start_row+1, max_row=veh_end_row)
        doughnut1.add_data(data2, titles_from_data=True)
        doughnut1.set_categories(labels2)
        doughnut1.width = 10
        doughnut1.height = 7
        doughnut1.dataLabels = DataLabelList()
        doughnut1.dataLabels.showPercent = True
        ws.add_chart(doughnut1, f"F{veh_start_row}")
        
        # === قسم الأقسام === (يبدأ بعد قسم السيارات مع مسافة)
        departments = Department.query.all()
        
        dept_start_row = veh_end_row + 10
        ws.merge_cells(f'A{dept_start_row}:E{dept_start_row}')
        ws[f'A{dept_start_row}'] = "🏢 نسبة الحضور حسب القسم - الموظفين النشطين"
        ws[f'A{dept_start_row}'].font = section_font
        ws[f'A{dept_start_row}'].fill = header_fill
        ws[f'A{dept_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{dept_start_row}'].border = accent_border
        ws.row_dimensions[dept_start_row].height = 35
        
        dept_headers = ['القسم', 'الموظفين', 'الحضور', 'النسبة', 'التقييم']
        for i, h in enumerate(dept_headers):
            cell = ws.cell(row=dept_start_row+1, column=i+1)
            cell.value = h
            cell.font = cyan_font
            cell.fill = card_fill_alt
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cyan_border_thin
        ws.row_dimensions[dept_start_row+1].height = 28
        
        dept_row = dept_start_row + 2
        for dept in departments[:10]:
            # جلب موظفي القسم النشطين فقط
            emp_ids = db.session.query(Employee.id).join(
                Employee.departments
            ).filter(
                Department.id == dept.id,
                Employee.status == 'active'
            ).all()
            emp_ids = [e[0] for e in emp_ids]
            
            if not emp_ids:
                continue
            
            # الموظفين الذين لهم حضور فعلي في الفترة
            emp_ids_with_attendance = db.session.query(Attendance.employee_id).filter(
                Attendance.date >= d_from,
                Attendance.date <= d_to,
                Attendance.employee_id.in_(emp_ids)
            ).distinct().all()
            emp_ids_with_attendance = [e[0] for e in emp_ids_with_attendance]
            
            if not emp_ids_with_attendance:
                continue
            
            dept_attendance = Attendance.query.filter(
                Attendance.date >= d_from,
                Attendance.date <= d_to,
                Attendance.employee_id.in_(emp_ids_with_attendance)
            ).all()
            
            present = sum(1 for a in dept_attendance if a.status == 'present')
            total = len(dept_attendance)
            rate = round((present / total) * 100) if total > 0 else 0
            
            if rate >= 90:
                rating = "ممتاز ⭐"
                rating_fill = green_fill
            elif rate >= 75:
                rating = "جيد 👍"
                rating_fill = teal_fill
            elif rate >= 60:
                rating = "متوسط ⚡"
                rating_fill = orange_fill
            else:
                rating = "يحتاج تحسين ⚠️"
                rating_fill = red_fill
            
            ws.cell(row=dept_row, column=1).value = dept.name
            ws.cell(row=dept_row, column=1).font = data_font
            ws.cell(row=dept_row, column=1).fill = card_fill
            ws.cell(row=dept_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=dept_row, column=1).border = thin_border
            
            ws.cell(row=dept_row, column=2).value = len(emp_ids_with_attendance)
            ws.cell(row=dept_row, column=2).font = Font(bold=True, color="00D4FF", size=12, name='Arial')
            ws.cell(row=dept_row, column=2).fill = card_fill
            ws.cell(row=dept_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=dept_row, column=2).border = thin_border
            
            ws.cell(row=dept_row, column=3).value = present
            ws.cell(row=dept_row, column=3).font = Font(bold=True, color="00F5D4", size=12, name='Arial')
            ws.cell(row=dept_row, column=3).fill = card_fill
            ws.cell(row=dept_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=dept_row, column=3).border = thin_border
            
            ws.cell(row=dept_row, column=4).value = f"{rate}%"
            ws.cell(row=dept_row, column=4).font = Font(bold=True, color="FFFFFF", size=12, name='Arial')
            ws.cell(row=dept_row, column=4).fill = card_fill
            ws.cell(row=dept_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=dept_row, column=4).border = thin_border
            
            ws.cell(row=dept_row, column=5).value = rating
            ws.cell(row=dept_row, column=5).font = dark_font
            ws.cell(row=dept_row, column=5).fill = rating_fill
            ws.cell(row=dept_row, column=5).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=dept_row, column=5).border = thin_border
            
            ws.row_dimensions[dept_row].height = 26
            dept_row += 1
        
        if dept_row > dept_start_row + 2:
            bar1 = BarChart()
            bar1.type = "col"
            bar1.style = 12
            bar1.title = "نسبة الحضور بالأقسام"
            bar1.y_axis.title = "النسبة %"
            
            data_bar = Reference(ws, min_col=4, min_row=dept_start_row+1, max_row=dept_row-1)
            cats_bar = Reference(ws, min_col=1, min_row=dept_start_row+2, max_row=dept_row-1)
            bar1.add_data(data_bar, titles_from_data=True)
            bar1.set_categories(cats_bar)
            bar1.width = 10
            bar1.height = 7
            bar1.dataLabels = DataLabelList()
            bar1.dataLabels.showVal = True
            ws.add_chart(bar1, f"G{dept_start_row}")
        
        # === قسم الوثائق === (يبدأ بعد الأقسام مع مسافة كافية للرسم)
        doc_counts = db.session.query(
            Document.employee_id,
            func.count(Document.id)
        ).group_by(Document.employee_id).all()
        
        complete_docs = sum(1 for _, cnt in doc_counts if cnt >= 4)
        incomplete_docs = max(0, total_employees - complete_docs)
        
        doc_start = dept_row + 10
        ws.merge_cells(f'A{doc_start}:E{doc_start}')
        ws[f'A{doc_start}'] = "📄 حالة اكتمال الوثائق - الموظفين النشطين"
        ws[f'A{doc_start}'].font = section_font
        ws[f'A{doc_start}'].fill = header_fill
        ws[f'A{doc_start}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{doc_start}'].border = accent_border
        ws.row_dimensions[doc_start].height = 35
        
        doc_headers = ['الحالة', 'العدد', 'النسبة', 'الرسم البياني']
        for i, h in enumerate(doc_headers, 1):
            ws.cell(row=doc_start+1, column=i).value = h
            ws.cell(row=doc_start+1, column=i).font = cyan_font
            ws.cell(row=doc_start+1, column=i).fill = card_fill_alt
            ws.cell(row=doc_start+1, column=i).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=doc_start+1, column=i).border = cyan_border_thin
        ws.row_dimensions[doc_start+1].height = 28
        
        complete_pct = round((complete_docs / total_employees * 100), 1) if total_employees > 0 else 0
        incomplete_pct = round((incomplete_docs / total_employees * 100), 1) if total_employees > 0 else 0
        
        ws.cell(row=doc_start+2, column=1).value = "مكتمل ✅"
        ws.cell(row=doc_start+2, column=2).value = complete_docs
        ws.cell(row=doc_start+2, column=3).value = f"{complete_pct}%"
        bar_width = int(complete_pct / 4) if complete_pct > 0 else 1
        ws.cell(row=doc_start+2, column=4).value = "█" * bar_width
        ws.cell(row=doc_start+2, column=1).font = white_font
        ws.cell(row=doc_start+2, column=2).font = Font(bold=True, color="00D4FF", size=14, name='Arial')
        ws.cell(row=doc_start+2, column=3).font = Font(bold=True, color="00F5D4", size=12, name='Arial')
        ws.cell(row=doc_start+2, column=4).font = Font(color="10B981", size=12, name='Arial')
        for i in range(1, 5):
            ws.cell(row=doc_start+2, column=i).fill = card_fill
            ws.cell(row=doc_start+2, column=i).border = thin_border
            ws.cell(row=doc_start+2, column=i).alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[doc_start+2].height = 26
        
        ws.cell(row=doc_start+3, column=1).value = "ناقص ⚠️"
        ws.cell(row=doc_start+3, column=2).value = incomplete_docs
        ws.cell(row=doc_start+3, column=3).value = f"{incomplete_pct}%"
        bar_width = int(incomplete_pct / 4) if incomplete_pct > 0 else 1
        ws.cell(row=doc_start+3, column=4).value = "█" * bar_width
        ws.cell(row=doc_start+3, column=1).font = white_font
        ws.cell(row=doc_start+3, column=2).font = Font(bold=True, color="00D4FF", size=14, name='Arial')
        ws.cell(row=doc_start+3, column=3).font = Font(bold=True, color="FFD700", size=12, name='Arial')
        ws.cell(row=doc_start+3, column=4).font = Font(color="FF4757", size=12, name='Arial')
        for i in range(1, 5):
            ws.cell(row=doc_start+3, column=i).fill = card_fill
            ws.cell(row=doc_start+3, column=i).border = thin_border
            ws.cell(row=doc_start+3, column=i).alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[doc_start+3].height = 26
        
        pie3 = PieChart()
        pie3.title = "اكتمال الوثائق"
        labels3 = Reference(ws, min_col=1, min_row=doc_start+2, max_row=doc_start+3)
        data3 = Reference(ws, min_col=2, min_row=doc_start+1, max_row=doc_start+3)
        pie3.add_data(data3, titles_from_data=True)
        pie3.set_categories(labels3)
        pie3.width = 10
        pie3.height = 7
        pie3.dataLabels = DataLabelList()
        pie3.dataLabels.showPercent = True
        ws.add_chart(pie3, f"G{doc_start}")
        
        # ألوان للتقارير التفصيلية
        detail_header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        detail_header_font = Font(bold=True, color="00D4AA", size=12)
        success_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        warning_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        danger_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        info_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        alt_row_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        if data_type in ['attendance', 'all']:
            ws_att = wb.create_sheet("تقرير الحضور")
            ws_att.sheet_view.rightToLeft = True
            
            ws_att.merge_cells('A1:G1')
            ws_att['A1'] = "📋 تقرير الحضور التفصيلي - الموظفين النشطين"
            ws_att['A1'].font = title_font
            ws_att['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_att.row_dimensions[1].height = 30
            
            ws_att.merge_cells('A2:G2')
            ws_att['A2'] = f"الفترة: {d_from} إلى {d_to}"
            ws_att['A2'].font = subtitle_font
            ws_att['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            headers = ['التاريخ', 'اسم الموظف', 'الرقم الوظيفي', 'القسم', 'الحالة', 'وقت الحضور', 'وقت الانصراف']
            for col, header in enumerate(headers, start=1):
                cell = ws_att.cell(row=4, column=col)
                cell.value = header
                cell.font = detail_header_font
                cell.fill = detail_header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_att.row_dimensions[4].height = 25
            
            # جلب الحضور للموظفين النشطين فقط
            query = Attendance.query.filter(
                Attendance.employee_id.in_(active_emp_ids)
            ).order_by(Attendance.date.desc())
            query = query.filter(Attendance.date >= d_from, Attendance.date <= d_to)
            
            status_translations = {
                'present': 'حاضر',
                'absent': 'غائب',
                'late': 'متأخر',
                'excused': 'معذور',
                'leave': 'إجازة',
                'sick': 'مريض'
            }
            
            att_records = query.limit(500).all()
            for row_idx, record in enumerate(att_records, start=5):
                emp = Employee.query.get(record.employee_id) if record.employee_id else None
                status = record.status or 'unknown'
                
                data_row = [
                    record.date.strftime('%Y-%m-%d') if record.date else '',
                    emp.name if emp else 'غير معروف',
                    emp.employee_id if emp else '',
                    emp.department.name if emp and emp.department else '',
                    status_translations.get(status, status),
                    record.time.strftime('%H:%M') if hasattr(record, 'time') and record.time else '',
                    record.checkout_time.strftime('%H:%M') if hasattr(record, 'checkout_time') and record.checkout_time else ''
                ]
                
                for col, value in enumerate(data_row, start=1):
                    cell = ws_att.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if col == 5:
                        if status == 'present':
                            cell.fill = success_fill
                        elif status == 'absent':
                            cell.fill = danger_fill
                        elif status == 'late':
                            cell.fill = warning_fill
                        elif status == 'excused':
                            cell.fill = info_fill
                    elif row_idx % 2 == 0:
                        cell.fill = alt_row_fill
            
            for col in range(1, 8):
                ws_att.column_dimensions[get_column_letter(col)].width = 18
        
        if data_type in ['employees', 'all']:
            ws_emp = wb.create_sheet("تقرير الموظفين")
            ws_emp.sheet_view.rightToLeft = True
            
            ws_emp.merge_cells('A1:F1')
            ws_emp['A1'] = "👥 تقرير الموظفين والوثائق"
            ws_emp['A1'].font = title_font
            ws_emp['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_emp.row_dimensions[1].height = 30
            
            headers = ['الرقم', 'اسم الموظف', 'الرقم الوظيفي', 'القسم', 'عدد الوثائق', 'حالة الوثائق']
            for col, header in enumerate(headers, start=1):
                cell = ws_emp.cell(row=3, column=col)
                cell.value = header
                cell.font = detail_header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_emp.row_dimensions[3].height = 25
            
            employees = Employee.query.all()
            
            doc_counts = db.session.query(
                Document.employee_id,
                func.count(Document.id).label('count')
            ).group_by(Document.employee_id).all()
            doc_count_map = {e_id: cnt for e_id, cnt in doc_counts}
            
            for row_idx, emp in enumerate(employees, start=4):
                docs_count = doc_count_map.get(emp.id, 0)
                status = 'مكتمل ✅' if docs_count >= 4 else 'ناقص ⚠️'
                
                data_row = [
                    emp.id,
                    emp.name,
                    emp.employee_id or '-',
                    emp.department.name if emp.department else 'بدون قسم',
                    docs_count,
                    status
                ]
                
                for col, value in enumerate(data_row, start=1):
                    cell = ws_emp.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if col == 6:
                        cell.fill = success_fill if docs_count >= 4 else warning_fill
                    elif row_idx % 2 == 0:
                        cell.fill = alt_row_fill
            
            for col in range(1, 7):
                ws_emp.column_dimensions[get_column_letter(col)].width = 18
        
        if data_type in ['vehicles', 'all']:
            ws_veh = wb.create_sheet("تقرير السيارات")
            ws_veh.sheet_view.rightToLeft = True
            
            ws_veh.merge_cells('A1:F1')
            ws_veh['A1'] = "🚗 تقرير أسطول السيارات"
            ws_veh['A1'].font = title_font
            ws_veh['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_veh.row_dimensions[1].height = 30
            
            headers = ['رقم اللوحة', 'الماركة', 'الموديل', 'السنة', 'الحالة', 'حالة التسليم']
            for col, header in enumerate(headers, start=1):
                cell = ws_veh.cell(row=3, column=col)
                cell.value = header
                cell.font = detail_header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_veh.row_dimensions[3].height = 25
            
            status_translations = {
                'working': 'نشط',
                'maintenance': 'صيانة',
                'inactive': 'غير نشط'
            }
            
            vehicles = Vehicle.query.all()
            for row_idx, v in enumerate(vehicles, start=4):
                status = v.status or 'unknown'
                
                data_row = [
                    v.plate_number if hasattr(v, 'plate_number') else '',
                    v.make if hasattr(v, 'make') else '',
                    v.model if hasattr(v, 'model') else '',
                    v.year if hasattr(v, 'year') else '',
                    status_translations.get(status, status),
                    'مستلمة ✅' if v.handover_records else 'غير مستلمة'
                ]
                
                for col, value in enumerate(data_row, start=1):
                    cell = ws_veh.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if col == 5:
                        if status == 'working':
                            cell.fill = success_fill
                        elif status == 'maintenance':
                            cell.fill = warning_fill
                        elif status == 'inactive':
                            cell.fill = danger_fill
                    elif row_idx % 2 == 0:
                        cell.fill = alt_row_fill
            
            for col in range(1, 7):
                ws_veh.column_dimensions[get_column_letter(col)].width = 16
        
        if data_type == 'all':
            ws_dept = wb.create_sheet("تحليل الأقسام")
            ws_dept.sheet_view.rightToLeft = True
            
            ws_dept.merge_cells('A1:G1')
            ws_dept['A1'] = "🏢 تحليل الحضور حسب الأقسام"
            ws_dept['A1'].font = title_font
            ws_dept['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_dept.row_dimensions[1].height = 30
            
            headers = ['القسم', 'عدد الموظفين', 'حاضر', 'غائب', 'متأخر', 'نسبة الحضور', 'التقييم']
            for col, header in enumerate(headers, start=1):
                cell = ws_dept.cell(row=3, column=col)
                cell.value = header
                cell.font = detail_header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_dept.row_dimensions[3].height = 25
            
            if date_from:
                d_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            else:
                d_from = datetime.now().date() - timedelta(days=30)
            
            if date_to:
                d_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            else:
                d_to = datetime.now().date()
            
            departments = Department.query.all()
            
            all_employees = Employee.query.all()
            dept_employees = {}
            for e in all_employees:
                if e.department_id:
                    if e.department_id not in dept_employees:
                        dept_employees[e.department_id] = []
                    dept_employees[e.department_id].append(e.id)
            
            all_attendance = Attendance.query.filter(
                Attendance.date >= d_from,
                Attendance.date <= d_to
            ).all()
            
            dept_attendance = {}
            for a in all_attendance:
                for d_id, emp_ids in dept_employees.items():
                    if a.employee_id in emp_ids:
                        if d_id not in dept_attendance:
                            dept_attendance[d_id] = {'present': 0, 'absent': 0, 'late': 0, 'total': 0}
                        dept_attendance[d_id]['total'] += 1
                        if a.status == 'present':
                            dept_attendance[d_id]['present'] += 1
                        elif a.status == 'absent':
                            dept_attendance[d_id]['absent'] += 1
                        elif a.status == 'late':
                            dept_attendance[d_id]['late'] += 1
                        break
            
            actual_row = 4
            for dept in departments:
                emp_count = len(dept_employees.get(dept.id, []))
                if emp_count == 0:
                    continue
                
                stats = dept_attendance.get(dept.id, {'present': 0, 'absent': 0, 'late': 0, 'total': 0})
                present = stats['present']
                absent = stats['absent']
                late = stats['late']
                total = stats['total']
                rate = round((present / total) * 100, 1) if total > 0 else 0
                
                if rate >= 90:
                    performance = 'ممتاز ⭐'
                    perf_fill = success_fill
                elif rate >= 75:
                    performance = 'جيد 👍'
                    perf_fill = info_fill
                elif rate >= 60:
                    performance = 'متوسط ⚡'
                    perf_fill = warning_fill
                else:
                    performance = 'يحتاج تحسين ⚠️'
                    perf_fill = danger_fill
                
                data_row = [dept.name, len(employees), present, absent, late, f'{rate}%', performance]
                
                for col, value in enumerate(data_row, start=1):
                    cell = ws_dept.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if col == 7:
                        cell.fill = perf_fill
                    elif row_idx % 2 == 0:
                        cell.fill = alt_row_fill
            
            for col in range(1, 8):
                ws_dept.column_dimensions[get_column_letter(col)].width = 16
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"powerbi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@powerbi_bp.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    """إحصائيات شاملة للوحة المعلومات"""
    try:
        total_employees = Employee.query.count()
        total_vehicles = Vehicle.query.count()
        total_documents = Document.query.count()
        total_departments = Department.query.count()
        
        today = datetime.now().date()
        today_attendance = Attendance.query.filter(Attendance.date == today).all()
        today_present = sum(1 for a in today_attendance if a.status == 'present')
        
        working_vehicles = Vehicle.query.filter_by(status='working').count()
        
        return jsonify({
            'success': True,
            'data': {
                'employees': {
                    'total': total_employees,
                    'present_today': today_present
                },
                'vehicles': {
                    'total': total_vehicles,
                    'working': working_vehicles
                },
                'documents': {
                    'total': total_documents
                },
                'departments': {
                    'total': total_departments
                },
                'last_updated': datetime.now().isoformat()
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@powerbi_bp.route('/export-pdf')
@login_required
def export_pdf():
    """تصدير لوحة المعلومات كـ PDF احترافي"""
    try:
        from weasyprint import HTML, CSS
        from flask import url_for, render_template_string
        import tempfile
        
        # جلب البيانات
        departments = Department.query.all()
        total_vehicles = Vehicle.query.count()
        total_documents = Document.query.count()
        
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        department_id = request.args.get('department_id')
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except:
                date_from = datetime.now().date() - timedelta(days=30)
        else:
            date_from = datetime.now().date() - timedelta(days=30)
        
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except:
                date_to = datetime.now().date()
        else:
            date_to = datetime.now().date()
        
        if date_from > date_to:
            date_from, date_to = date_to, date_from
        
        active_employee_ids_with_attendance = db.session.query(Attendance.employee_id).filter(
            Attendance.date >= date_from,
            Attendance.date <= date_to
        ).distinct().all()
        active_employee_ids_with_attendance = [e[0] for e in active_employee_ids_with_attendance]
        
        active_employees_count = Employee.query.filter(
            Employee.status == 'active',
            Employee.id.in_(active_employee_ids_with_attendance)
        ).count()
        
        total_employees = active_employees_count
        
        active_emp_ids = [e.id for e in Employee.query.filter(Employee.status == 'active').all()]
        
        attendance_records = Attendance.query.filter(
            Attendance.date >= date_from,
            Attendance.date <= date_to,
            Attendance.employee_id.in_(active_emp_ids)
        ).all()
        
        attendance_stats = {
            'present': sum(1 for a in attendance_records if a.status == 'present'),
            'absent': sum(1 for a in attendance_records if a.status in ['absent', 'غائب']),
            'leave': sum(1 for a in attendance_records if a.status == 'leave'),
            'sick': sum(1 for a in attendance_records if a.status == 'sick'),
            'total': len(attendance_records)
        }
        attendance_stats['rate'] = round((attendance_stats['present'] / attendance_stats['total']) * 100, 1) if attendance_stats['total'] > 0 else 0
        
        dept_attendance = []
        for dept in departments:
            emp_ids = db.session.query(Employee.id).join(
                Employee.departments
            ).filter(
                Department.id == dept.id,
                Employee.status == 'active'
            ).all()
            emp_ids = [e[0] for e in emp_ids]
            
            if not emp_ids:
                continue
            
            emp_ids_with_attendance = db.session.query(Attendance.employee_id).filter(
                Attendance.date >= date_from,
                Attendance.date <= date_to,
                Attendance.employee_id.in_(emp_ids)
            ).distinct().all()
            emp_ids_with_attendance = [e[0] for e in emp_ids_with_attendance]
            
            if not emp_ids_with_attendance:
                continue
            
            dept_records = Attendance.query.filter(
                Attendance.date >= date_from,
                Attendance.date <= date_to,
                Attendance.employee_id.in_(emp_ids_with_attendance)
            ).all()
            
            present_count = sum(1 for a in dept_records if a.status == 'present')
            total_count = len(dept_records)
            rate = round((present_count / total_count) * 100, 1) if total_count > 0 else 0
            
            dept_attendance.append({
                'name': dept.name,
                'employee_count': len(emp_ids_with_attendance),
                'present': present_count,
                'absent': sum(1 for a in dept_records if a.status in ['absent', 'غائب']),
                'rate': rate
            })
        
        html_content = f'''
        <!DOCTYPE html>
        <html dir="rtl" lang="ar">
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{
                    size: A4 landscape;
                    margin: 1cm;
                }}
                
                * {{
                    box-sizing: border-box;
                    font-family: 'Noto Sans Arabic', 'Tahoma', 'Arial', sans-serif;
                }}
                
                body {{
                    direction: rtl;
                    background: linear-gradient(180deg, #0a0e17 0%, #0d1321 100%);
                    color: #e8eaed;
                    margin: 0;
                    padding: 20px;
                }}
                
                .header {{
                    background: linear-gradient(135deg, #131b2e 0%, rgba(0, 212, 255, 0.05) 100%);
                    border: 2px solid rgba(0, 212, 255, 0.3);
                    border-radius: 15px;
                    padding: 20px;
                    margin-bottom: 20px;
                    text-align: center;
                }}
                
                .header h1 {{
                    color: #00d4ff;
                    margin: 0 0 10px 0;
                    font-size: 28px;
                }}
                
                .header p {{
                    color: #8892a0;
                    margin: 5px 0;
                    font-size: 14px;
                }}
                
                .kpi-grid {{
                    display: grid;
                    grid-template-columns: repeat(5, 1fr);
                    gap: 15px;
                    margin-bottom: 25px;
                }}
                
                .kpi-card {{
                    background: #131b2e;
                    border: 1px solid rgba(0, 212, 255, 0.2);
                    border-radius: 12px;
                    padding: 15px;
                    text-align: center;
                }}
                
                .kpi-value {{
                    font-size: 32px;
                    font-weight: bold;
                    color: #00d4ff;
                    margin-bottom: 5px;
                }}
                
                .kpi-label {{
                    color: #8892a0;
                    font-size: 12px;
                }}
                
                .section {{
                    background: #131b2e;
                    border: 1px solid rgba(0, 212, 255, 0.2);
                    border-radius: 12px;
                    padding: 20px;
                    margin-bottom: 20px;
                }}
                
                .section-title {{
                    color: #00d4ff;
                    font-size: 18px;
                    margin: 0 0 15px 0;
                    border-bottom: 1px solid rgba(0, 212, 255, 0.2);
                    padding-bottom: 10px;
                }}
                
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    font-size: 11px;
                }}
                
                th {{
                    background: rgba(0, 212, 255, 0.1);
                    color: #00d4ff;
                    padding: 10px;
                    text-align: center;
                    border: 1px solid rgba(0, 212, 255, 0.2);
                }}
                
                td {{
                    padding: 8px;
                    text-align: center;
                    border: 1px solid rgba(0, 212, 255, 0.1);
                    color: #e8eaed;
                }}
                
                tr:nth-child(even) {{
                    background: rgba(0, 212, 255, 0.03);
                }}
                
                .rate-good {{ color: #00ff88; }}
                .rate-medium {{ color: #ffd700; }}
                .rate-low {{ color: #ff4757; }}
                
                .footer {{
                    text-align: center;
                    color: #8892a0;
                    font-size: 10px;
                    margin-top: 20px;
                    padding-top: 10px;
                    border-top: 1px solid rgba(0, 212, 255, 0.2);
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>📊 لوحة التحليلات الاحترافية - نُظم</h1>
                <p>تقرير شامل للفترة من {date_from.strftime('%Y-%m-%d')} إلى {date_to.strftime('%Y-%m-%d')}</p>
                <p>تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
            </div>
            
            <div class="kpi-grid">
                <div class="kpi-card">
                    <div class="kpi-value">{total_employees}</div>
                    <div class="kpi-label">إجمالي الموظفين</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-value">{len(departments)}</div>
                    <div class="kpi-label">الأقسام</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-value">{total_vehicles}</div>
                    <div class="kpi-label">السيارات</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-value">{total_documents}</div>
                    <div class="kpi-label">الوثائق</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-value">{attendance_stats['rate']}%</div>
                    <div class="kpi-label">نسبة الحضور</div>
                </div>
            </div>
            
            <div class="section">
                <h2 class="section-title">📈 إحصائيات الحضور</h2>
                <table>
                    <tr>
                        <th>الحضور</th>
                        <th>الغياب</th>
                        <th>الإجازات</th>
                        <th>المرضى</th>
                        <th>الإجمالي</th>
                    </tr>
                    <tr>
                        <td class="rate-good">{attendance_stats['present']}</td>
                        <td class="rate-low">{attendance_stats['absent']}</td>
                        <td>{attendance_stats['leave']}</td>
                        <td>{attendance_stats['sick']}</td>
                        <td>{attendance_stats['total']}</td>
                    </tr>
                </table>
            </div>
            
            <div class="section">
                <h2 class="section-title">🏢 الحضور حسب القسم</h2>
                <table>
                    <tr>
                        <th>القسم</th>
                        <th>عدد الموظفين</th>
                        <th>الحضور</th>
                        <th>الغياب</th>
                        <th>نسبة الحضور</th>
                    </tr>
                    {''.join(f"""
                    <tr>
                        <td><strong>{dept['name']}</strong></td>
                        <td>{dept['employee_count']}</td>
                        <td class="rate-good">{dept['present']}</td>
                        <td class="rate-low">{dept['absent']}</td>
                        <td class="{'rate-good' if dept['rate'] >= 80 else 'rate-medium' if dept['rate'] >= 60 else 'rate-low'}">{dept['rate']}%</td>
                    </tr>
                    """ for dept in dept_attendance)}
                </table>
            </div>
            
            <div class="footer">
                نُظم - نظام إدارة الموظفين المتكامل | تم إنشاء هذا التقرير تلقائياً
            </div>
        </body>
        </html>
        '''
        
        pdf = HTML(string=html_content).write_pdf()
        
        output = BytesIO(pdf)
        output.seek(0)
        
        filename = f"powerbi_dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 400
