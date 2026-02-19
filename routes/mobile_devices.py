from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy import or_, and_, func
import pandas as pd
import io
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import make_response
import os
import uuid

from core.extensions import db
from models import MobileDevice, Employee, Department, AuditLog, employee_departments, ImportedPhoneNumber

# إنشاء Blueprint
mobile_devices_bp = Blueprint('mobile_devices', __name__)

def log_activity(action, entity_type=None, entity_id=None, details=None):
    """دالة تسجيل العمليات في سجل التدقيق"""
    try:
        audit = AuditLog()
        audit.action = action
        audit.entity_type = entity_type
        audit.entity_id = entity_id
        audit.details = details
        audit.user_id = current_user.id if current_user.is_authenticated else None
        audit.timestamp = datetime.utcnow()
        
        db.session.add(audit)
        db.session.commit()
    except Exception as e:
        # لا نريد أن تفشل العملية بسبب فشل التسجيل
        db.session.rollback()
        pass

@mobile_devices_bp.route('/')
@login_required
def index():
    """صفحة إدارة الأجهزة المحمولة الرئيسية"""
    try:
        # جلب معاملات البحث والفلترة
        search = request.args.get('search', '')
        phone_search = request.args.get('phone_search', '')
        status_filter = request.args.get('status', '')
        employee_filter = request.args.get('employee_id', '')
        device_brand = request.args.get('device_brand', '')
        page = request.args.get('page', 1, type=int)
        per_page = 20
        
        # بناء الاستعلام الأساسي
        query = MobileDevice.query
        
        # تطبيق فلاتر البحث
        if search:
            query = query.filter(
                or_(
                    MobileDevice.imei.contains(search),
                    MobileDevice.email.contains(search),
                    MobileDevice.device_model.contains(search)
                )
            )
        
        # البحث المخصص برقم الهاتف
        if phone_search:
            query = query.filter(MobileDevice.phone_number.contains(phone_search))
        
        if status_filter:
            if status_filter == 'متاح':
                # فلتر الأجهزة المتاحة: فقط الأجهزة غير المربوطة أو المربوطة بموظفين غير نشطين
                query = query.filter(
                    or_(
                        # أجهزة غير مربوطة تماماً
                        MobileDevice.employee_id.is_(None),
                        # أجهزة مربوطة بموظفين غير نشطين (أو مرحلين أو منتهين)
                        and_(
                            MobileDevice.employee_id.isnot(None),
                            MobileDevice.employee.has(
                                and_(
                                    Employee.status != 'نشط',
                                    Employee.status != 'active'
                                )
                            )
                        )
                    )
                )
            elif status_filter == 'مرتبط':
                # فلتر الأجهزة المرتبطة فقط مع موظفين نشطين
                query = query.filter(
                    and_(
                        MobileDevice.employee_id.isnot(None),
                        MobileDevice.employee.has(
                            or_(
                                Employee.status == 'نشط',
                                Employee.status == 'active'
                            )
                        )
                    )
                )
            else:
                # باقي الفلاتر العادية (مثل مربوطة بموظفين غير نشطين)
                query = query.filter(MobileDevice.status == status_filter)
            
        if employee_filter and employee_filter.isdigit():
            query = query.filter(MobileDevice.employee_id == int(employee_filter))
            
        if device_brand:
            query = query.filter(MobileDevice.device_brand == device_brand)
        
        # تطبيق الترتيب والترقيم
        devices = query.order_by(MobileDevice.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        # جلب البيانات الإضافية للفلاتر
        employees = Employee.query.filter_by(status='نشط').order_by(Employee.name).all()
        device_brands = db.session.query(MobileDevice.device_brand).filter(
            MobileDevice.device_brand.isnot(None)
        ).distinct().all()
        device_brands = [brand[0] for brand in device_brands if brand[0]]
        
        # حساب الإحصائيات
        total_devices = MobileDevice.query.count()
        
        # الأجهزة المربوطة فعلياً (فقط مع موظفين نشطين)
        assigned_devices = MobileDevice.query.filter(
            and_(
                MobileDevice.employee_id.isnot(None),
                MobileDevice.employee.has(
                    or_(
                        Employee.status == 'نشط',
                        Employee.status == 'active'
                    )
                )
            )
        ).count()
        
        # الأجهزة المتاحة: إما غير مربوطة أو مربوطة بموظف غير نشط
        available_devices = MobileDevice.query.filter(
            or_(
                # أجهزة غير مربوطة تماماً
                MobileDevice.employee_id.is_(None),
                # أجهزة مربوطة بموظفين غير نشطين
                and_(
                    MobileDevice.employee_id.isnot(None),
                    MobileDevice.employee.has(
                        and_(
                            Employee.status != 'نشط',
                            Employee.status != 'active'
                        )
                    )
                )
            )
        ).count()
        
        # الأجهزة المربوطة بموظفين غير نشطين
        inactive_employee_devices = MobileDevice.query.filter(
            and_(
                MobileDevice.employee_id.isnot(None),
                MobileDevice.employee.has(
                    and_(
                        Employee.status != 'نشط',
                        Employee.status != 'active'
                    )
                )
            )
        ).count()
        
        return render_template('mobile_devices/index.html',
                             devices=devices,
                             employees=employees,
                             device_brands=device_brands,
                             search=search,
                             phone_search=phone_search,
                             status_filter=status_filter,
                             employee_filter=employee_filter,
                             device_brand=device_brand,
                             total_devices=total_devices,
                             assigned_devices=assigned_devices,
                             available_devices=available_devices,
                             inactive_employee_devices=inactive_employee_devices)
                             
    except Exception as e:
        flash(f'حدث خطأ أثناء تحميل الأجهزة: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@mobile_devices_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    """إضافة جهاز جديد"""
    if request.method == 'POST':
        try:
            # استخراج البيانات من النموذج
            phone_number = request.form.get('phone_number', '').strip()
            imported_phone_id = request.form.get('imported_phone_id', '')
            imei = request.form.get('imei', '').strip()
            email = request.form.get('email', '').strip()
            device_model = request.form.get('device_model', '').strip()
            device_brand = request.form.get('device_brand', '').strip()
            notes = request.form.get('notes', '').strip()
            
            # إذا تم اختيار رقم من الأرقام المستوردة
            if imported_phone_id and imported_phone_id.isdigit():
                imported_phone = ImportedPhoneNumber.query.get(int(imported_phone_id))
                if imported_phone and not imported_phone.is_used:
                    phone_number = imported_phone.phone_number
            
            # التحقق من البيانات المطلوبة
            if not phone_number or not imei:
                flash('رقم الهاتف ورقم الـ IMEI مطلوبان', 'danger')
                return redirect(url_for('mobile_devices.create'))
            
            # التحقق من عدم تكرار الـ IMEI
            existing_device = MobileDevice.query.filter_by(imei=imei).first()
            if existing_device:
                flash('رقم الـ IMEI موجود مسبقاً في النظام', 'danger')
                return redirect(url_for('mobile_devices.create'))
            
            # إنشاء الجهاز الجديد
            device = MobileDevice()
            device.phone_number = phone_number
            device.imei = imei
            device.email = email if email else None
            device.device_model = device_model if device_model else None
            device.device_brand = device_brand if device_brand else None
            device.notes = notes if notes else None
            device.status = 'متاح'
            
            db.session.add(device)
            db.session.commit()
            
            # تحديد الرقم المستورد كمستخدم إذا تم استخدامه
            if imported_phone_id and imported_phone_id.isdigit():
                imported_phone = ImportedPhoneNumber.query.get(int(imported_phone_id))
                if imported_phone:
                    imported_phone.mark_as_used()
            
            # تسجيل العملية
            log_activity(
                action='إضافة جهاز محمول',
                entity_type='MobileDevice',
                entity_id=device.id,
                details=f'تم إضافة جهاز جديد: {imei} - {phone_number}'
            )
            
            flash('تم إضافة الجهاز بنجاح', 'success')
            return redirect(url_for('mobile_devices.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء إضافة الجهاز: {str(e)}', 'danger')
    
    # جلب الأرقام المستوردة المتاحة
    imported_numbers = ImportedPhoneNumber.get_available_numbers()
    
    return render_template('mobile_devices/create.html', imported_numbers=imported_numbers)

@mobile_devices_bp.route('/<int:device_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(device_id):
    """تعديل بيانات الجهاز"""
    device = MobileDevice.query.get_or_404(device_id)
    
    if request.method == 'POST':
        try:
            # حفظ البيانات القديمة للمقارنة
            old_imei = device.imei
            
            # استخراج البيانات الجديدة
            phone_number = request.form.get('phone_number', '').strip()
            imei = request.form.get('imei', '').strip()
            email = request.form.get('email', '').strip()
            device_model = request.form.get('device_model', '').strip()
            device_brand = request.form.get('device_brand', '').strip()
            status = request.form.get('status', '').strip()
            notes = request.form.get('notes', '').strip()
            
            # التحقق من البيانات المطلوبة
            if not phone_number or not imei:
                flash('رقم الهاتف ورقم الـ IMEI مطلوبان', 'danger')
                return redirect(url_for('mobile_devices.edit', device_id=device_id))
            
            # التحقق من عدم تكرار الـ IMEI (إذا تم تغييره)
            if imei != old_imei:
                existing_device = MobileDevice.query.filter_by(imei=imei).first()
                if existing_device:
                    flash('رقم الـ IMEI موجود مسبقاً في النظام', 'danger')
                    return redirect(url_for('mobile_devices.edit', device_id=device_id))
            
            # تحديث البيانات
            device.phone_number = phone_number
            device.imei = imei
            device.email = email if email else None
            device.device_model = device_model if device_model else None
            device.device_brand = device_brand if device_brand else None
            device.status = status
            device.notes = notes if notes else None
            device.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            # تسجيل العملية
            audit = AuditLog()
            audit.action = 'تحديث جهاز محمول'
            audit.entity_type = 'MobileDevice'
            audit.entity_id = device.id
            audit.details = f'تم تحديث بيانات الجهاز: {imei} - {phone_number}'
            audit.user_id = current_user.id if current_user.is_authenticated else None
            audit.timestamp = datetime.utcnow()
            
            db.session.add(audit)
            db.session.commit()
            
            flash('تم تحديث بيانات الجهاز بنجاح', 'success')
            return redirect(url_for('mobile_devices.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث الجهاز: {str(e)}', 'danger')
    
    # جلب قائمة الموظفين النشطين للربط
    active_employees = Employee.query.filter(Employee.status.in_(['نشط', 'active'])).order_by(Employee.name).all()
    
    return render_template('mobile_devices/edit.html', device=device, active_employees=active_employees)

@mobile_devices_bp.route('/<int:device_id>/delete', methods=['POST'])
@login_required
def delete(device_id):
    """حذف الجهاز"""
    try:
        device = MobileDevice.query.get_or_404(device_id)
        
        # حفظ البيانات للتسجيل
        imei = device.imei
        phone_number = device.phone_number
        
        # حذف الجهاز
        db.session.delete(device)
        db.session.commit()
        
        # تسجيل العملية
        audit = AuditLog()
        audit.action = 'حذف جهاز محمول'
        audit.entity_type = 'MobileDevice'
        audit.entity_id = device_id
        audit.details = f'تم حذف الجهاز: {imei} - {phone_number}'
        audit.user_id = current_user.id if current_user.is_authenticated else None
        audit.timestamp = datetime.utcnow()
        
        db.session.add(audit)
        db.session.commit()
        
        flash('تم حذف الجهاز بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الجهاز: {str(e)}', 'danger')
    
    return redirect(url_for('mobile_devices.index'))

@mobile_devices_bp.route('/import-phone-numbers', methods=['GET', 'POST'])
@login_required
def import_phone_numbers():
    """استيراد أرقام الهواتف من ملف Excel"""
    if request.method == 'POST':
        try:
            file = request.files.get('excel_file')
            if not file or file.filename == '':
                flash('الرجاء اختيار ملف Excel', 'danger')
                return redirect(url_for('mobile_devices.import_phone_numbers'))
            
            # قراءة ملف Excel
            df = pd.read_excel(file)
            
            # التحقق من وجود عمود phone_number أو أرقام الهواتف
            phone_column = None
            description_column = None
            
            # البحث عن العمود المناسب لأرقام الهواتف
            for col in df.columns:
                if 'phone' in col.lower() or 'هاتف' in col or 'رقم' in col:
                    phone_column = col
                    break
            
            # البحث عن عمود الوصف أو الاسم
            for col in df.columns:
                if 'name' in col.lower() or 'description' in col.lower() or 'اسم' in col or 'وصف' in col:
                    description_column = col
                    break
            
            # إذا لم نجد عمود أرقام الهواتف، نستخدم العمود الأول
            if phone_column is None:
                phone_column = df.columns[0]
            
            imported_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # استخراج رقم الهاتف
                    phone_number = str(row[phone_column]).strip()
                    
                    # تنظيف رقم الهاتف من المسافات والرموز غير المرغوبة
                    phone_number = phone_number.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                    
                    if not phone_number or phone_number == 'nan':
                        continue
                    
                    # التحقق من عدم تكرار رقم الهاتف
                    existing_number = ImportedPhoneNumber.query.filter_by(phone_number=phone_number).first()
                    if existing_number:
                        continue  # تجاهل الأرقام المكررة
                    
                    # استخراج الوصف إذا كان موجوداً
                    description = ''
                    if description_column and pd.notna(row.get(description_column)):
                        description = str(row[description_column]).strip()
                    
                    # إنشاء رقم هاتف مستورد جديد
                    imported_phone = ImportedPhoneNumber()
                    imported_phone.phone_number = phone_number
                    imported_phone.description = description if description else None
                    imported_phone.imported_by = current_user.id if current_user.is_authenticated else None
                    
                    db.session.add(imported_phone)
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f'الصف {index + 1}: {str(e)}')
            
            # حفظ البيانات
            if imported_count > 0:
                db.session.commit()
                
                # تسجيل العملية
                log_activity(
                    action='استيراد أرقام هواتف',
                    entity_type='ImportedPhoneNumber',
                    details=f'تم استيراد {imported_count} رقم هاتف من ملف Excel'
                )
            
            # رسائل النتائج
            if imported_count > 0:
                flash(f'تم استيراد {imported_count} رقم هاتف بنجاح', 'success')
            else:
                flash('لم يتم استيراد أي أرقام جديدة', 'warning')
            
            if errors:
                flash(f'حدثت أخطاء في {len(errors)} سجل', 'warning')
            
            return redirect(url_for('mobile_devices.dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء استيراد الملف: {str(e)}', 'danger')
    
    return render_template('mobile_devices/import_phone_numbers.html')

@mobile_devices_bp.route('/download-phone-template')
@login_required
def download_phone_template():
    """تحميل نموذج Excel لاستيراد أرقام الهواتف"""
    try:
        # إنشاء Workbook جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "نموذج أرقام الهواتف"
        
        # تنسيق الرأس
        header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # رؤوس الأعمدة
        headers = ['phone_number', 'description']
        header_names = ['رقم الهاتف', 'الوصف (اختياري)']
        
        # إضافة رؤوس الأعمدة الإنجليزية
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # إضافة رؤوس الأعمدة العربية
        for col, header_name in enumerate(header_names, 1):
            cell = ws.cell(row=2, column=col, value=header_name)
            cell.fill = PatternFill(start_color="20c997", end_color="20c997", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = header_alignment
        
        # إضافة بيانات تجريبية
        sample_data = [
            ['966501234567', 'أحمد محمد علي'],
            ['966522334455', 'فاطمة عبدالله'],
            ['966533445566', 'خالد السالم'],
            ['966544556677', 'نورا أحمد'],
            ['966555667788', 'محمد عبدالله'],
            ['966566778899', 'سارة محمد'],
            ['966577889900', 'عبدالرحمن خالد'],
            ['966588990011', 'مريم يوسف'],
            ['966599001122', 'عمر الحسن'],
            ['966500112233', 'هند عبدالعزيز']
        ]
        
        # إضافة البيانات التجريبية
        for row_idx, data in enumerate(sample_data, 3):
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1:  # عمود رقم الهاتف
                    cell.alignment = Alignment(horizontal="center")
                else:  # عمود الوصف
                    cell.alignment = Alignment(horizontal="right")
        
        # تعديل عرض الأعمدة
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        
        # إضافة ورقة تعليمات
        instructions_sheet = wb.create_sheet("التعليمات")
        instructions = [
            "تعليمات استيراد أرقام الهواتف:",
            "",
            "1. الأعمدة المطلوبة:",
            "   - phone_number: رقم الهاتف (مطلوب)",
            "   - description: الوصف أو اسم صاحب الرقم (اختياري)",
            "",
            "2. تنسيق رقم الهاتف:",
            "   - يجب أن يبدأ برمز الدولة (مثال: 966 للسعودية)",
            "   - مثال صحيح: 966501234567",
            "   - مثال صحيح: 966522334455",
            "",
            "3. ملاحظات مهمة:",
            "   - النظام يتجاهل الأرقام المكررة تلقائياً",
            "   - سيتم تنظيف الأرقام من المسافات والرموز غير المرغوبة",
            "   - يمكن ترك عمود الوصف فارغاً",
            "   - الحد الأقصى: 1000 رقم في الملف الواحد",
            "",
            "4. أسماء الأعمدة المقبولة:",
            "   - للهاتف: phone, phone_number, هاتف, رقم",
            "   - للوصف: name, description, اسم, وصف",
            "",
            "5. خطوات الاستيراد:",
            "   - احذف هذه الصفوف التجريبية",
            "   - أدخل أرقامك الحقيقية",
            "   - احفظ الملف",
            "   - ارفعه في صفحة الاستيراد"
        ]
        
        for row_idx, instruction in enumerate(instructions, 1):
            cell = instructions_sheet.cell(row=row_idx, column=1, value=instruction)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="2c3e50")
            elif instruction.startswith(("1.", "2.", "3.", "4.", "5.")):
                cell.font = Font(bold=True, size=12, color="34495e")
            else:
                cell.font = Font(size=11, color="2c3e50")
        
        instructions_sheet.column_dimensions['A'].width = 60
        
        # حفظ في ذاكرة مؤقتة
        from io import BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # إنشاء استجابة التحميل
        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # استخدام filename* لدعم الأحرف العربية
        response.headers['Content-Disposition'] = "attachment; filename*=UTF-8''phone_numbers_template.xlsx"
        
        # تسجيل العملية
        log_activity(
            action='تحميل نموذج أرقام الهواتف',
            entity_type='Template',
            details='تم تحميل نموذج Excel لاستيراد أرقام الهواتف'
        )
        
        return response
        
    except Exception as e:
        flash(f'حدث خطأ أثناء إنشاء النموذج: {str(e)}', 'danger')
        return redirect(url_for('mobile_devices.import_phone_numbers'))

@mobile_devices_bp.route('/bulk-unlink', methods=['POST'])
@login_required
def bulk_unlink():
    """فك ربط مجموعة من الأجهزة أو الكل"""
    try:
        action = request.form.get('action')
        device_ids = request.form.getlist('device_ids')
        
        if action == 'unlink_all':
            # فك ربط جميع الأجهزة
            devices = MobileDevice.query.filter(MobileDevice.employee_id.isnot(None)).all()
            count = len(devices)
            
            for device in devices:
                device.employee_id = None
                device.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            # تسجيل العملية
            log_activity(
                action='فك ربط جميع الأجهزة',
                entity_type='MobileDevice',
                details=f'تم فك ربط {count} جهاز من الموظفين'
            )
            
            flash(f'تم فك ربط {count} جهاز بنجاح', 'success')
            
        elif action == 'unlink_selected' and device_ids:
            # فك ربط الأجهزة المحددة
            devices = MobileDevice.query.filter(
                MobileDevice.id.in_(device_ids),
                MobileDevice.employee_id.isnot(None)
            ).all()
            
            count = len(devices)
            device_numbers = []
            
            for device in devices:
                device_numbers.append(device.phone_number)
                device.employee_id = None
                device.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            # تسجيل العملية
            log_activity(
                action='فك ربط أجهزة محددة',
                entity_type='MobileDevice',
                details=f'تم فك ربط {count} جهاز: {", ".join(device_numbers[:5])}{"..." if count > 5 else ""}'
            )
            
            flash(f'تم فك ربط {count} جهاز محدد بنجاح', 'success')
        else:
            flash('يرجى تحديد عملية صحيحة أو اختيار أجهزة للفك', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء فك الربط: {str(e)}', 'danger')
    
    return redirect(url_for('mobile_devices.index'))

@mobile_devices_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_excel():
    """استيراد الأجهزة من ملف Excel"""
    if request.method == 'POST':
        try:
            file = request.files.get('excel_file')
            if not file or file.filename == '':
                flash('الرجاء اختيار ملف Excel', 'danger')
                return redirect(url_for('mobile_devices.import_excel'))
            
            # قراءة ملف Excel
            df = pd.read_excel(file)
            
            # التحقق من الأعمدة المطلوبة
            required_columns = ['phone_number', 'imei']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                flash(f'الأعمدة المطلوبة مفقودة: {", ".join(missing_columns)}', 'danger')
                return redirect(url_for('mobile_devices.import_excel'))
            
            imported_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # التحقق من البيانات الأساسية
                    phone_number = str(row['phone_number']).strip()
                    imei = str(row['imei']).strip()
                    
                    if not phone_number or not imei:
                        errors.append(f'الصف {index + 1}: رقم الهاتف أو IMEI فارغ')
                        continue
                    
                    # التحقق من تكرار IMEI
                    if MobileDevice.query.filter_by(imei=imei).first():
                        errors.append(f'الصف {index + 1}: IMEI {imei} موجود مسبقاً')
                        continue
                    
                    # إنشاء الجهاز
                    device = MobileDevice()
                    device.phone_number = phone_number
                    device.imei = imei
                    device.email = str(row.get('email', '')).strip() if pd.notna(row.get('email')) else None
                    device.device_model = str(row.get('device_model', '')).strip() if pd.notna(row.get('device_model')) else None
                    device.device_brand = str(row.get('device_brand', '')).strip() if pd.notna(row.get('device_brand')) else None
                    device.status = 'متاح'
                    
                    db.session.add(device)
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f'الصف {index + 1}: {str(e)}')
            
            # حفظ البيانات
            if imported_count > 0:
                db.session.commit()
                
                # تسجيل العملية
                audit = AuditLog()
                audit.action = 'استيراد أجهزة محمولة'
                audit.entity_type = 'MobileDevice'
                audit.details = f'تم استيراد {imported_count} جهاز من ملف Excel'
                audit.user_id = current_user.id if current_user.is_authenticated else None
                audit.timestamp = datetime.utcnow()
                
                db.session.add(audit)
                db.session.commit()
            
            # رسائل النتائج
            if imported_count > 0:
                flash(f'تم استيراد {imported_count} جهاز بنجاح', 'success')
            
            if errors:
                flash(f'حدثت أخطاء في {len(errors)} سجل', 'warning')
                for error in errors[:10]:  # عرض أول 10 أخطاء
                    flash(error, 'danger')
            
            return redirect(url_for('mobile_devices.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء استيراد الملف: {str(e)}', 'danger')
    
    return render_template('mobile_devices/import.html')

@mobile_devices_bp.route('/export')
@login_required
def export_excel():
    """تصدير الأجهزة إلى ملف Excel"""
    try:
        # جلب جميع الأجهزة
        devices = MobileDevice.query.order_by(MobileDevice.created_at.desc()).all()
        
        # إنشاء Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "الأجهزة المحمولة"
        
        # تنسيق الرأس
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # رؤوس الأعمدة
        headers = [
            'رقم الهاتف', 'رقم IMEI', 'الإيميل', 'نوع الجهاز', 'ماركة الجهاز',
            'الموظف المربوط', 'الحالة', 'تاريخ الربط', 'تاريخ الإضافة', 'ملاحظات'
        ]
        
        # إضافة رؤوس الأعمدة
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # إضافة البيانات
        for row, device in enumerate(devices, 2):
            employee_name = device.employee.name if device.employee else ''
            assigned_date = device.assigned_date.strftime('%Y-%m-%d') if device.assigned_date else ''
            created_date = device.created_at.strftime('%Y-%m-%d') if device.created_at else ''
            
            ws.cell(row=row, column=1, value=device.phone_number)
            ws.cell(row=row, column=2, value=device.imei)
            ws.cell(row=row, column=3, value=device.email or '')
            ws.cell(row=row, column=4, value=device.device_model or '')
            ws.cell(row=row, column=5, value=device.device_brand or '')
            ws.cell(row=row, column=6, value=employee_name)
            ws.cell(row=row, column=7, value=device.status)
            ws.cell(row=row, column=8, value=assigned_date)
            ws.cell(row=row, column=9, value=created_date)
            ws.cell(row=row, column=10, value=device.notes or '')
        
        # تعديل عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # حفظ الملف في الذاكرة
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"mobile_devices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('mobile_devices.index'))

@mobile_devices_bp.route('/assign')
@login_required
def assign():
    """صفحة ربط الأجهزة بالموظفين"""
    try:
        # معاملات البحث والفلترة
        search = request.args.get('search', '').strip()
        department_filter = request.args.get('department', '').strip()
        device_search = request.args.get('device_search', '').strip()
        
        # جلب الأجهزة غير المربوطة أو المربوطة بموظفين غير نشطين
        devices_query = MobileDevice.query.outerjoin(Employee).filter(
            or_(
                MobileDevice.employee_id.is_(None),
                ~Employee.status.in_(['نشط', 'active'])
            )
        ).filter(MobileDevice.status != 'معطل')
        
        # فلترة الأجهزة
        if device_search:
            devices_query = devices_query.filter(
                or_(
                    MobileDevice.phone_number.contains(device_search),
                    MobileDevice.imei.contains(device_search),
                    MobileDevice.device_model.contains(device_search)
                )
            )
        
        available_devices = devices_query.all()
        
        # جلب الموظفين النشطين
        employees_query = Employee.query.filter(Employee.status.in_(['نشط', 'active']))
        
        # تطبيق فلتر القسم أولاً إذا كان محدداً
        if department_filter and department_filter.isdigit():
            employees_query = employees_query.join(employee_departments).filter(
                employee_departments.c.department_id == int(department_filter)
            )
        
        # تطبيق البحث النصي إذا كان محدداً
        if search:
            employees_query = employees_query.filter(
                or_(
                    Employee.name.contains(search),
                    Employee.employee_id.contains(search), 
                    Employee.national_id.contains(search),
                    Employee.mobile.contains(search)
                )
            )
        
        active_employees = employees_query.order_by(Employee.name).all()
        
        # جلب جميع الأقسام للفلترة
        departments = Department.query.order_by(Department.name).all()
        
        return render_template('mobile_devices/assign.html',
                             available_devices=available_devices,
                             active_employees=active_employees,
                             departments=departments,
                             search=search,
                             department_filter=department_filter,
                             device_search=device_search)
                             
    except Exception as e:
        flash(f'حدث خطأ أثناء تحميل صفحة الربط: {str(e)}', 'danger')
        return redirect(url_for('mobile_devices.index'))

@mobile_devices_bp.route('/<int:device_id>/assign/<int:employee_id>', methods=['POST'])
@login_required
def assign_device(device_id, employee_id):
    """ربط الجهاز بالموظف"""
    try:
        device = MobileDevice.query.get_or_404(device_id)
        employee = Employee.query.get_or_404(employee_id)
        
        # التحقق من إمكانية الربط
        if device.status == 'معطل':
            flash('لا يمكن ربط جهاز معطل', 'danger')
            return redirect(url_for('mobile_devices.assign'))
        
        if employee.status not in ['نشط', 'active']:
            flash('لا يمكن ربط الجهاز بموظف غير نشط', 'danger')
            return redirect(url_for('mobile_devices.assign'))
        
        # فك الربط السابق إذا وجد
        old_employee_name = None
        if device.employee_id:
            old_employee = Employee.query.get(device.employee_id)
            old_employee_name = old_employee.name if old_employee else 'غير معروف'
        
        # ربط الجهاز بالموظف الجديد
        device.employee_id = employee_id
        device.status = 'مرتبط'
        device.assigned_date = datetime.utcnow()
        device.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        # تسجيل العملية
        details = f'تم ربط الجهاز {device.imei} بالموظف {employee.name}'
        if old_employee_name:
            details += f' (كان مربوطاً بـ {old_employee_name})'
            
        audit = AuditLog()
        audit.action = 'ربط جهاز محمول'
        audit.entity_type = 'MobileDevice'
        audit.entity_id = device.id
        audit.details = details
        audit.user_id = current_user.id if current_user.is_authenticated else None
        audit.timestamp = datetime.utcnow()
        
        db.session.add(audit)
        db.session.commit()
        
        # التحقق من نوع الطلب
        if request.headers.get('Content-Type') == 'application/json':
            return jsonify({'success': True, 'message': f'تم ربط الجهاز بالموظف {employee.name} بنجاح'})
        else:
            flash(f'تم ربط الجهاز بالموظف {employee.name} بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        if request.headers.get('Content-Type') == 'application/json':
            return jsonify({'success': False, 'message': f'خطأ: {str(e)}'}), 500
        else:
            flash(f'حدث خطأ أثناء ربط الجهاز: {str(e)}', 'danger')
    
    return redirect(url_for('mobile_devices.assign'))

@mobile_devices_bp.route('/<int:device_id>/unassign', methods=['POST'])
@login_required
def unassign_device(device_id):
    """فك ربط الجهاز من الموظف"""
    try:
        device = MobileDevice.query.get_or_404(device_id)
        
        if not device.employee_id:
            return jsonify({'success': False, 'message': 'الجهاز غير مربوط بأي موظف'}), 400
        
        # حفظ اسم الموظف للتسجيل
        employee = Employee.query.get(device.employee_id)
        employee_name = employee.name if employee else 'غير معروف'
        
        # فك الربط
        device.employee_id = None
        device.status = 'متاح'
        device.assigned_date = None
        device.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        # تسجيل العملية
        audit = AuditLog()
        audit.action = 'فك ربط جهاز محمول'
        audit.entity_type = 'MobileDevice'
        audit.entity_id = device.id
        audit.details = f'تم فك ربط الجهاز {device.imei} من الموظف {employee_name}'
        audit.user_id = current_user.id if current_user.is_authenticated else None
        audit.timestamp = datetime.utcnow()
        
        db.session.add(audit)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'تم فك ربط الجهاز من الموظف {employee_name} بنجاح'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'خطأ: {str(e)}'}), 500

@mobile_devices_bp.route('/dashboard')
@login_required
def dashboard():
    """لوحة معلومات الأجهزة المحمولة مع فلاتر وبحث"""
    try:
        # معاملات البحث والفلترة
        search = request.args.get('search', '').strip()
        employee_filter = request.args.get('employee', '')
        action_filter = request.args.get('action', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        phone_filter = request.args.get('phone', '').strip()
        page = request.args.get('page', 1, type=int)
        per_page = 20
        
        # حساب الإحصائيات الأساسية
        total_devices = MobileDevice.query.count()
        assigned_devices = MobileDevice.query.filter(MobileDevice.employee_id.isnot(None)).count()
        available_devices = MobileDevice.query.filter(
            and_(
                MobileDevice.employee_id.is_(None),
                MobileDevice.status == 'متاح'
            )
        ).count()
        
        # الأجهزة المربوطة بموظفين غير نشطين
        inactive_employee_devices = db.session.query(MobileDevice).join(Employee).filter(
            and_(
                MobileDevice.employee_id.isnot(None),
                Employee.status.notin_(['active', 'نشط'])
            )
        ).all()
        
        # الأجهزة غير المربوطة
        unassigned_devices = MobileDevice.query.filter(
            and_(
                MobileDevice.employee_id.is_(None),
                MobileDevice.status != 'معطل'
            )
        ).all()
        
        # إحصائيات حسب الماركة
        brand_stats = db.session.query(
            MobileDevice.device_brand,
            func.count(MobileDevice.id).label('count')
        ).filter(
            MobileDevice.device_brand.isnot(None)
        ).group_by(MobileDevice.device_brand).all()
        
        # بناء استعلام العمليات مع الفلاتر
        activities_query = db.session.query(
            AuditLog.action,
            AuditLog.details,
            AuditLog.timestamp,
            Employee.name.label('employee_name'),
            Employee.id.label('employee_id'),
            MobileDevice.phone_number,
            MobileDevice.imei
        ).outerjoin(
            # ربط بالموظف المسجل للعملية
            Employee, AuditLog.user_id == Employee.id
        ).outerjoin(
            # ربط بالجهاز من خلال entity_id
            MobileDevice, and_(
                AuditLog.entity_type == 'MobileDevice',
                AuditLog.entity_id == MobileDevice.id
            )
        ).filter(
            or_(
                AuditLog.action.contains('جهاز محمول'),
                AuditLog.action.contains('ربط جهاز'),
                AuditLog.action.contains('فك ربط'),
                AuditLog.action.contains('إضافة جهاز'),
                AuditLog.action.contains('تعديل جهاز'),
                AuditLog.action.contains('حذف جهاز')
            )
        )
        
        # تطبيق فلاتر البحث
        if search:
            activities_query = activities_query.filter(
                or_(
                    AuditLog.details.contains(search),
                    Employee.name.contains(search),
                    MobileDevice.phone_number.contains(search),
                    MobileDevice.imei.contains(search)
                )
            )
        
        if employee_filter:
            activities_query = activities_query.filter(Employee.name.contains(employee_filter))
        
        if action_filter:
            activities_query = activities_query.filter(AuditLog.action.contains(action_filter))
        
        if phone_filter:
            activities_query = activities_query.filter(MobileDevice.phone_number.contains(phone_filter))
        
        # فلترة التاريخ
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                activities_query = activities_query.filter(
                    func.date(AuditLog.timestamp) >= date_from_obj.date()
                )
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                activities_query = activities_query.filter(
                    func.date(AuditLog.timestamp) <= date_to_obj.date()
                )
            except ValueError:
                pass
        
        # ترتيب وتطبيق التصفح
        activities_query = activities_query.order_by(AuditLog.timestamp.desc())
        activities = activities_query.paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        # جلب جميع الموظفين للفلترة
        all_employees = Employee.query.order_by(Employee.name).all()
        
        # أنواع الإجراءات المتاحة
        action_types = [
            'ربط جهاز محمول',
            'فك ربط جهاز محمول', 
            'إضافة جهاز محمول',
            'تعديل جهاز محمول',
            'حذف جهاز محمول',
            'استيراد أجهزة محمولة'
        ]
        
        return render_template('mobile_devices/dashboard.html',
                             total_devices=total_devices,
                             assigned_devices=assigned_devices,
                             available_devices=available_devices,
                             inactive_employee_devices=inactive_employee_devices,
                             unassigned_devices=unassigned_devices,
                             brand_stats=brand_stats,
                             activities=activities,
                             all_employees=all_employees,
                             action_types=action_types,
                             search=search,
                             employee_filter=employee_filter,
                             action_filter=action_filter,
                             phone_filter=phone_filter,
                             date_from=date_from,
                             date_to=date_to)
                             
    except Exception as e:
        flash(f'حدث خطأ أثناء تحميل لوحة المعلومات: {str(e)}', 'danger')
        return redirect(url_for('mobile_devices.index'))

@mobile_devices_bp.route('/export-dashboard-excel')
@login_required
def export_dashboard_excel():
    """تصدير بيانات العمليات المفلترة من Dashboard إلى Excel"""
    try:
        # معاملات البحث والفلترة
        search = request.args.get('search', '').strip()
        employee_filter = request.args.get('employee', '')
        action_filter = request.args.get('action', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        phone_filter = request.args.get('phone', '').strip()
        
        # بناء استعلام العمليات مع الفلاتر (نفس المنطق من dashboard)
        activities_query = db.session.query(
            AuditLog.action,
            AuditLog.details,
            AuditLog.timestamp,
            Employee.name.label('employee_name'),
            MobileDevice.phone_number,
            MobileDevice.imei
        ).outerjoin(
            Employee, AuditLog.user_id == Employee.id
        ).outerjoin(
            MobileDevice, and_(
                AuditLog.entity_type == 'MobileDevice',
                AuditLog.entity_id == MobileDevice.id
            )
        ).filter(
            or_(
                AuditLog.action.contains('جهاز محمول'),
                AuditLog.action.contains('ربط جهاز'),
                AuditLog.action.contains('فك ربط'),
                AuditLog.action.contains('إضافة جهاز'),
                AuditLog.action.contains('تعديل جهاز'),
                AuditLog.action.contains('حذف جهاز')
            )
        )
        
        # تطبيق فلاتر البحث
        if search:
            activities_query = activities_query.filter(
                or_(
                    AuditLog.details.contains(search),
                    Employee.name.contains(search),
                    MobileDevice.phone_number.contains(search),
                    MobileDevice.imei.contains(search)
                )
            )
        
        if employee_filter:
            activities_query = activities_query.filter(Employee.name.contains(employee_filter))
        
        if action_filter:
            activities_query = activities_query.filter(AuditLog.action.contains(action_filter))
        
        if phone_filter:
            activities_query = activities_query.filter(MobileDevice.phone_number.contains(phone_filter))
        
        # فلترة التاريخ
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                activities_query = activities_query.filter(
                    func.date(AuditLog.timestamp) >= date_from_obj.date()
                )
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                activities_query = activities_query.filter(
                    func.date(AuditLog.timestamp) <= date_to_obj.date()
                )
            except ValueError:
                pass
        
        # جلب جميع النتائج
        activities = activities_query.order_by(AuditLog.timestamp.desc()).all()
        
        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "عمليات الأجهزة المحمولة"
        
        # إعداد رؤوس الأعمدة
        headers = [
            'نوع العملية',
            'التفاصيل',
            'الموظف المنفذ',
            'رقم الهاتف',
            'IMEI',
            'التاريخ والوقت'
        ]
        
        # كتابة الرؤوس
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # كتابة البيانات
        for row, activity in enumerate(activities, 2):
            ws.cell(row=row, column=1, value=activity.action or 'غير محدد')
            ws.cell(row=row, column=2, value=activity.details or 'غير محدد')
            ws.cell(row=row, column=3, value=activity.employee_name or 'غير محدد')
            ws.cell(row=row, column=4, value=activity.phone_number or '-')
            ws.cell(row=row, column=5, value=activity.imei or '-')
            ws.cell(row=row, column=6, value=activity.timestamp.strftime('%Y-%m-%d %H:%M') if activity.timestamp else 'غير محدد')
        
        # تنسيق العرض
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # إعداد الاستجابة
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'عمليات_الأجهزة_المحمولة_{current_time}.xlsx'
        
        # حفظ الملف في الذاكرة
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # تسجيل العملية
        log_activity(
            action='تصدير عمليات الأجهزة المحمولة',
            entity_type='MobileDevice',
            details=f'تم تصدير {len(activities)} عملية إلى ملف Excel مع الفلاتر المطبقة'
        )
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('mobile_devices.dashboard'))