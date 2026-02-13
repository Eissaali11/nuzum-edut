from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response, send_file
from werkzeug.utils import secure_filename
from sqlalchemy import func, or_
from sqlalchemy.orm import selectinload
from datetime import datetime, timedelta
import io
from io import BytesIO
import csv
import xlsxwriter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask_login import current_user, login_required
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
import arabic_reshaper
from bidi.algorithm import get_display
from app import db
from models import Document, Employee, Department, SystemAudit
from utils.excel import parse_document_excel
from utils.date_converter import parse_date, format_date_hijri, format_date_gregorian
from utils.audit_logger import log_activity
import json

documents_bp = Blueprint('documents', __name__)

@documents_bp.route('/get_sponsorship_employees', methods=['POST'])
def get_sponsorship_employees():
    """Get employees filtered by sponsorship status"""
    try:
        sponsorship_filter = request.form.get('sponsorship_filter')
        
        if not sponsorship_filter:
            return jsonify({'success': False, 'message': 'يرجى تحديد نوع الكفالة'})
        
        # Query employees based on sponsorship status
        if sponsorship_filter == 'on_sponsorship':
            employees = Employee.query.filter(Employee.sponsorship_status == 'على الكفالة').all()
        elif sponsorship_filter == 'off_sponsorship':
            employees = Employee.query.filter(Employee.sponsorship_status == 'خارج الكفالة').all()
        else:
            return jsonify({'success': False, 'message': 'نوع الكفالة غير صحيح'})
        
        # Format employee data
        employees_data = []
        for emp in employees:
            dept_names = ', '.join([dept.name for dept in emp.departments]) if emp.departments else 'غير محدد'
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id,
                'national_id': emp.national_id,
                'sponsorship_status': emp.sponsorship_status,
                'department_names': dept_names
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@documents_bp.route('/department-bulk-create', methods=['GET', 'POST'])
@login_required
def department_bulk_create():
    """صفحة إنشاء وثائق القسم الكامل"""
    try:
        # جلب البيانات الأساسية
        departments = Department.query.all()
        employees = Employee.query.options(selectinload(Employee.departments)).all()
        document_types = [
            'national_id', 'passport', 'health_certificate', 
            'work_permit', 'education_certificate'
        ]
        
        if request.method == 'POST':
            # معالجة طلب الحفظ
            save_type = request.form.get('save_type')
            department_id = request.form.get('department_id')
            document_type = request.form.get('document_type')
            
            if save_type == 'individual':
                # حفظ موظف واحد
                employee_id = request.form.get('employee_id')
                document_number = request.form.get('document_number')
                issue_date = request.form.get('issue_date')
                expiry_date = request.form.get('expiry_date')
                notes = request.form.get('notes', '')
                
                if not all([employee_id, document_type, document_number]):
                    return jsonify({
                        'success': False, 
                        'message': 'يرجى إدخال جميع البيانات المطلوبة'
                    })
                
                # إنشاء الوثيقة
                document = Document(
                    employee_id=employee_id,
                    document_type=document_type,
                    document_number=document_number,
                    issue_date=parse_date(issue_date) if issue_date else None,
                    expiry_date=parse_date(expiry_date) if expiry_date else None,
                    notes=notes
                )
                
                db.session.add(document)
                
                # تسجيل العملية
                employee = Employee.query.get(employee_id)
                audit = SystemAudit(
                    action='create',
                    entity_type='document',
                    entity_id=document.id,
                    details=f'تم إضافة وثيقة {document_type} للموظف {employee.name if employee else "غير محدد"}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'message': 'تم حفظ الوثيقة بنجاح'
                })
            
            elif save_type == 'bulk':
                # حفظ جماعي
                import json
                employees_data = json.loads(request.form.get('employees_data', '[]'))
                
                if not employees_data:
                    return jsonify({
                        'success': False,
                        'message': 'لا توجد بيانات للحفظ'
                    })
                
                saved_count = 0
                for emp_data in employees_data:
                    if emp_data.get('document_number'):
                        document = Document(
                            employee_id=emp_data['employee_id'],
                            document_type=document_type,
                            document_number=emp_data['document_number'],
                            issue_date=parse_date(emp_data['issue_date']) if emp_data.get('issue_date') else None,
                            expiry_date=parse_date(emp_data['expiry_date']) if emp_data.get('expiry_date') else None,
                            notes=emp_data.get('notes', '')
                        )
                        db.session.add(document)
                        saved_count += 1
                
                # تسجيل العملية
                department = Department.query.get(department_id)
                audit = SystemAudit(
                    action='bulk_create',
                    entity_type='document',
                    entity_id=department_id,
                    details=f'تم إضافة {saved_count} وثيقة من نوع {document_type} لقسم {department.name if department else "غير محدد"}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'message': f'تم حفظ {saved_count} وثيقة بنجاح',
                    'redirect': url_for('documents.index')
                })
        
        return render_template('documents/department_bulk_create.html',
                             departments=departments,
                             employees=employees,
                             document_types=document_types)
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@documents_bp.route('/get_employees_by_sponsorship', methods=['POST'])
def get_employees_by_sponsorship():
    """Get employees filtered by sponsorship status"""
    try:
        data = request.get_json()
        sponsorship_type = data.get('sponsorship_type')  # 'internal' or 'external'
        
        # Query employees based on sponsorship status
        if sponsorship_type == 'internal':
            employees = Employee.query.filter(Employee.sponsorship_status == 'على الكفالة').all()
        elif sponsorship_type == 'external':
            employees = Employee.query.filter(Employee.sponsorship_status == 'خارج الكفالة').all()
        else:
            return jsonify({'success': False, 'message': 'نوع الكفالة غير صحيح'})
        
        # Format employee data
        employees_data = []
        for emp in employees:
            dept_names = ', '.join([dept.name for dept in emp.departments]) if emp.departments else 'غير محدد'
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id,
                'national_id': emp.national_id,
                'department_names': dept_names
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@documents_bp.route('/get_employees_by_department_and_sponsorship', methods=['POST'])
def get_employees_by_department_and_sponsorship():
    """Get employees filtered by department and sponsorship status"""
    try:
        data = request.get_json()
        department_id = data.get('department_id')
        sponsorship_type = data.get('sponsorship_type')
        
        # Build base query
        query = Employee.query.options(selectinload(Employee.departments))
        
        # Filter by department
        if department_id:
            query = query.filter(Employee.departments.any(Department.id == department_id))
        
        # Filter by sponsorship status
        if sponsorship_type == 'on_sponsorship':
            query = query.filter(Employee.sponsorship_status == 'على الكفالة')
        elif sponsorship_type == 'off_sponsorship':
            query = query.filter(Employee.sponsorship_status == 'خارج الكفالة')
        
        employees = query.all()
        
        # Format employee data
        employees_data = []
        for emp in employees:
            dept_names = ', '.join([dept.name for dept in emp.departments]) if emp.departments else 'غير محدد'
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id,
                'national_id': emp.national_id,
                'department_names': dept_names,
                'sponsorship_status': sponsorship_type
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@documents_bp.route('/save_individual_document', methods=['POST'])
def save_individual_document():
    """Save individual document for sponsorship-based addition"""
    try:
        data = request.get_json()
        
        # Create new document
        document = Document(
            employee_id=data['employee_id'],
            document_type=data['document_type'],
            document_number=data['document_number'],
            issue_date=datetime.strptime(data['issue_date'], '%Y-%m-%d').date() if data['issue_date'] else None,
            expiry_date=datetime.strptime(data['expiry_date'], '%Y-%m-%d').date() if data['expiry_date'] else None,
            notes=data.get('notes', ''),
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        db.session.add(document)
        db.session.commit()
        
        # Log the activity
        log_activity(
            action='create',
            entity_type='document',
            entity_id=document.id,
            details=f'تم إضافة وثيقة {data["document_type"]} للموظف {data["employee_id"]} فردياً'
        )
        
        return jsonify({
            'success': True,
            'message': 'تم حفظ الوثيقة بنجاح'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'حدث خطأ في حفظ الوثيقة: {str(e)}'
        })

@documents_bp.route('/save_bulk_documents', methods=['POST'])
def save_bulk_documents():
    """Save bulk documents for advanced filtering"""
    try:
        data = request.get_json()
        document_type = data['document_type']
        documents_data = data['documents']
        
        saved_count = 0
        
        for doc_data in documents_data:
            if doc_data.get('document_number') or doc_data.get('issue_date') or doc_data.get('expiry_date'):
                document = Document(
                    employee_id=doc_data['employee_id'],
                    document_type=document_type,
                    document_number=doc_data.get('document_number', ''),
                    issue_date=datetime.strptime(doc_data['issue_date'], '%Y-%m-%d').date() if doc_data.get('issue_date') else None,
                    expiry_date=datetime.strptime(doc_data['expiry_date'], '%Y-%m-%d').date() if doc_data.get('expiry_date') else None,
                    notes=doc_data.get('notes', ''),
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                
                db.session.add(document)
                saved_count += 1
        
        db.session.commit()
        
        # Log the activity
        log_activity(
            action='create',
            entity_type='document',
            entity_id=0,
            details=f'تم إضافة {saved_count} وثيقة من نوع {document_type} بشكل جماعي'
        )
        
        return jsonify({
            'success': True,
            'message': f'تم حفظ {saved_count} وثيقة بنجاح',
            'saved_count': saved_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'حدث خطأ في حفظ الوثائق: {str(e)}'
        })

# Duplicate route code removed - using the one above

@documents_bp.route('/')
def index():
    """List document records with filtering options"""
    # Get filter parameters
    document_type = request.args.get('document_type', '')
    employee_id = request.args.get('employee_id', '')
    department_id = request.args.get('department_id', '')
    sponsorship_status = request.args.get('sponsorship_status', '')
    status_filter = request.args.get('expiring', '')  # Fixed parameter name
    show_all = request.args.get('show_all', 'false')
    
    # Build query
    query = Document.query
    
    # Apply filters
    if document_type:
        query = query.filter(Document.document_type == document_type)
    
    if employee_id and employee_id.isdigit():
        query = query.filter(Document.employee_id == int(employee_id))
    
    # تصفية حسب القسم والكفالة (نحتاج للـ join مع Employee)
    if department_id and department_id.isdigit():
        # فلترة الوثائق للموظفين في قسم محدد
        dept_employees = Employee.query.join(Employee.departments).filter_by(id=int(department_id)).all()
        dept_employee_ids = [emp.id for emp in dept_employees]
        if dept_employee_ids:
            query = query.filter(Document.employee_id.in_(dept_employee_ids))
        else:
            # لا توجد موظفين في هذا القسم
            query = query.filter(False)
    
    if sponsorship_status:
        query = query.join(Employee).filter(Employee.sponsorship_status == sponsorship_status)
    
    # تطبيق فلتر حالة الصلاحية
    today = datetime.now().date()
    
    if status_filter == 'expired':
        # الوثائق المنتهية فقط
        query = query.filter(
            Document.expiry_date.isnot(None),
            Document.expiry_date < today
        )
    elif status_filter == 'expiring_30':
        # الوثائق التي تنتهي خلال 30 يوم
        future_date = today + timedelta(days=30)
        query = query.filter(
            Document.expiry_date.isnot(None),
            Document.expiry_date >= today,
            Document.expiry_date <= future_date
        )
    elif status_filter == 'expiring_60':
        # الوثائق التي تنتهي خلال 60 يوم
        future_date = today + timedelta(days=60)
        query = query.filter(
            Document.expiry_date.isnot(None),
            Document.expiry_date >= today,
            Document.expiry_date <= future_date
        )
    elif status_filter == 'expiring_90':
        # الوثائق التي تنتهي خلال 90 يوم
        future_date = today + timedelta(days=90)
        query = query.filter(
            Document.expiry_date.isnot(None),
            Document.expiry_date >= today,
            Document.expiry_date <= future_date
        )
    elif status_filter == 'valid':
        # الوثائق السارية (أكثر من 30 يوم للانتهاء)
        future_date = today + timedelta(days=30)
        query = query.filter(
            or_(
                Document.expiry_date.is_(None),  # الوثائق بدون تاريخ انتهاء
                Document.expiry_date > future_date  # الوثائق التي تنتهي بعد أكثر من 30 يوم
            )
        )
    elif show_all.lower() != 'true':
        # العرض الافتراضي: الوثائق المنتهية أو القريبة من الانتهاء (خلال 30 يوم)
        future_date_30_days = today + timedelta(days=30)
        query = query.filter(
            Document.expiry_date.isnot(None),
            Document.expiry_date <= future_date_30_days
        )
    
    # Execute query
    documents = query.all()
    
    # احسب عدد الوثائق الكلي والمنتهية والقريبة من الانتهاء
    total_docs = Document.query.count()
    # حساب عدد الوثائق المنتهية (يجب أن يكون لها تاريخ انتهاء حتى تعتبر منتهية)
    expired_docs = Document.query.filter(
        Document.expiry_date.isnot(None),
        Document.expiry_date < today
    ).count()
    # حساب عدد الوثائق التي ستنتهي قريباً
    expiring_soon = Document.query.filter(
        Document.expiry_date.isnot(None),
        Document.expiry_date <= today + timedelta(days=30),
        Document.expiry_date >= today
    ).count()
    # عدد الوثائق الآمنة والتي لها تاريخ انتهاء
    docs_with_expiry = Document.query.filter(Document.expiry_date.isnot(None)).count()
    safe_docs = docs_with_expiry - expired_docs - expiring_soon
    
    # Get all employees for filter dropdown
    employees = Employee.query.all()
    
    # Get all departments for filter dropdown
    departments = Department.query.all()
    
    # Get document types for filter dropdown
    document_types = [
        'national_id', 'passport', 'health_certificate', 
        'work_permit', 'education_certificate', 'driving_license',
        'annual_leave', 'other'
    ]
    
    return render_template('documents/index.html',
                          documents=documents,
                          employees=employees,
                          departments=departments,
                          document_types=document_types,
                          selected_type=document_type,
                          selected_employee=employee_id,
                          selected_department=department_id,
                          selected_sponsorship=sponsorship_status,
                          selected_status_filter=status_filter,
                          show_all=show_all.lower() == 'true',
                          total_docs=total_docs,
                          expired_docs=expired_docs,
                          expiring_soon=expiring_soon,
                          safe_docs=safe_docs,
                          valid_docs=safe_docs,
                          status_filter=status_filter,
                          today=today,
                          now=datetime.now())

@documents_bp.route('/export_excel')
@login_required
def export_excel():
    """تصدير الوثائق إلى ملف Excel حسب الفلاتر المطبقة"""
    try:
        # Get filter parameters (same as index route)
        document_type = request.args.get('document_type', '')
        employee_id = request.args.get('employee_id', '')
        department_id = request.args.get('department_id', '')
        sponsorship_status = request.args.get('sponsorship_status', '')
        status_filter = request.args.get('expiring', '')
        show_all = request.args.get('show_all', 'false')
        
        # Build query (same logic as index route)
        query = Document.query.options(
            selectinload(Document.employee).selectinload(Employee.departments)
        )
        
        # Apply filters
        if document_type:
            query = query.filter(Document.document_type == document_type)
        
        if employee_id and employee_id.isdigit():
            query = query.filter(Document.employee_id == int(employee_id))
        
        if department_id and department_id.isdigit():
            dept_employees = Employee.query.join(Employee.departments).filter_by(id=int(department_id)).all()
            dept_employee_ids = [emp.id for emp in dept_employees]
            if dept_employee_ids:
                query = query.filter(Document.employee_id.in_(dept_employee_ids))
            else:
                query = query.filter(False)
        
        if sponsorship_status:
            query = query.join(Employee).filter(Employee.sponsorship_status == sponsorship_status)
        
        # تطبيق فلتر حالة الصلاحية
        today = datetime.now().date()
        
        if status_filter == 'expired':
            query = query.filter(
                Document.expiry_date.isnot(None),
                Document.expiry_date < today
            )
        elif status_filter == 'expiring_30':
            future_date = today + timedelta(days=30)
            query = query.filter(
                Document.expiry_date.isnot(None),
                Document.expiry_date >= today,
                Document.expiry_date <= future_date
            )
        elif status_filter == 'expiring_60':
            future_date = today + timedelta(days=60)
            query = query.filter(
                Document.expiry_date.isnot(None),
                Document.expiry_date >= today,
                Document.expiry_date <= future_date
            )
        elif status_filter == 'expiring_90':
            future_date = today + timedelta(days=90)
            query = query.filter(
                Document.expiry_date.isnot(None),
                Document.expiry_date >= today,
                Document.expiry_date <= future_date
            )
        elif status_filter == 'valid':
            future_date = today + timedelta(days=30)
            query = query.filter(
                or_(
                    Document.expiry_date.is_(None),
                    Document.expiry_date > future_date
                )
            )
        elif show_all.lower() != 'true':
            future_date_30_days = today + timedelta(days=30)
            query = query.filter(
                Document.expiry_date.isnot(None),
                Document.expiry_date <= future_date_30_days
            )
        
        # Execute query
        documents = query.all()
        
        # تحضير البيانات للتصدير
        data = []
        
        # خريطة أنواع الوثائق
        document_types_map = {
            'national_id': 'الهوية الوطنية',
            'passport': 'جواز السفر',
            'health_certificate': 'الشهادة الصحية',
            'work_permit': 'تصريح العمل',
            'education_certificate': 'الشهادة الدراسية',
            'driving_license': 'رخصة القيادة',
            'annual_leave': 'الإجازة السنوية',
            'other': 'أخرى'
        }
        
        for doc in documents:
            # حساب حالة الوثيقة
            status = 'غير محدد'
            days_remaining = ''
            
            if doc.expiry_date:
                diff = (doc.expiry_date - today).days
                if diff < 0:
                    status = 'منتهية الصلاحية'
                    days_remaining = f'منتهية منذ {abs(diff)} يوم'
                elif diff <= 30:
                    status = 'تنتهي قريباً'
                    days_remaining = f'{diff} يوم متبقي'
                else:
                    status = 'سارية'
                    days_remaining = f'{diff} يوم متبقي'
            else:
                status = 'بدون تاريخ انتهاء'
                days_remaining = 'غير محدد'
            
            # معلومات الأقسام
            departments_list = ', '.join([dept.name for dept in doc.employee.departments]) if doc.employee.departments else 'غير محدد'
            
            row = {
                'نوع الوثيقة': document_types_map.get(doc.document_type, doc.document_type),
                'رقم الوثيقة': doc.document_number or '',
                'اسم الموظف': doc.employee.name if doc.employee else '',
                'رقم الموظف': doc.employee.employee_id if doc.employee else '',
                'رقم الهوية': doc.employee.national_id if doc.employee else '',
                'الأقسام': departments_list,
                'الجوال': doc.employee.mobile if doc.employee else '',
                'المنصب': doc.employee.job_title if doc.employee else '',
                'الحالة الوظيفية': doc.employee.status if doc.employee else '',
                'حالة الكفالة': doc.employee.sponsorship_status if doc.employee else '',
                'تاريخ الإصدار': doc.issue_date.strftime('%Y-%m-%d') if doc.issue_date else '',
                'تاريخ الانتهاء': doc.expiry_date.strftime('%Y-%m-%d') if doc.expiry_date else '',
                'حالة الوثيقة': status,
                'الأيام المتبقية': days_remaining,
                'ملاحظات': doc.notes or '',
                'تاريخ الإنشاء': doc.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(doc, 'created_at') and doc.created_at else '',
                'آخر تحديث': doc.updated_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(doc, 'updated_at') and doc.updated_at else ''
            }
            data.append(row)
        
        # إنشاء DataFrame
        df = pd.DataFrame(data)
        
        if df.empty:
            flash('لا توجد وثائق لتصديرها حسب الفلاتر المحددة', 'warning')
            return redirect(request.referrer or url_for('documents.index'))
        
        # إنشاء ملف Excel
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='الوثائق', index=False, startrow=2)
            
            workbook = writer.book
            worksheet = writer.sheets['الوثائق']
            
            # تنسيق الرأس
            header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # تنسيق العنوان الرئيسي
            title_font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
            title_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
            title_alignment = Alignment(horizontal='center', vertical='center')
            
            # إضافة العنوان الرئيسي
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = f'تقرير الوثائق - تم التصدير في {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = title_alignment
            
            # تطبيق التنسيق على رأس الجدول
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
                # تعديل عرض العمود
                column_width = max(len(str(column)), 15)
                if column_width > 50:
                    column_width = 50
                worksheet.column_dimensions[cell.column_letter].width = column_width
            
            # تطبيق الحدود على الجدول
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # تطبيق الحدود على كامل البيانات
            for row in range(1, len(df) + 4):  # +4 للعنوان والرأس
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).border = thin_border
            
            # تطبيق التنسيق على البيانات
            data_alignment = Alignment(horizontal='center', vertical='center')
            for row in range(4, len(df) + 4):  # البيانات تبدأ من الصف الرابع
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = data_alignment
                    
                    # تلوين الصفوف بالتناوب
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # إعداد اسم الملف
        filter_parts = []
        if document_type:
            filter_parts.append(f'نوع_{document_type}')
        if status_filter:
            filter_parts.append(f'حالة_{status_filter}')
        if department_id:
            dept = Department.query.get(department_id)
            if dept:
                filter_parts.append(f'قسم_{dept.name}')
        
        filename_suffix = '_'.join(filter_parts) if filter_parts else 'جميع_الوثائق'
        filename = f'تقرير_الوثائق_{filename_suffix}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        # تسجيل العملية
        audit = SystemAudit(
            action='export_excel',
            entity_type='document',
            entity_id=0,
            details=f'تم تصدير {len(documents)} وثيقة إلى ملف Excel - الفلاتر: {", ".join(filter_parts) if filter_parts else "بدون فلاتر"}'
        )
        db.session.add(audit)
        db.session.commit()
        
        output.seek(0)
        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('documents.index'))

@documents_bp.route('/create', methods=['GET', 'POST'])
def create():
    """Create a new document record"""
    if request.method == 'POST':
        try:
            # تحقق من وجود CSRF token
            if 'csrf_token' not in request.form:
                flash('خطأ في التحقق من الأمان. يرجى المحاولة مرة أخرى.', 'danger')
                return redirect(url_for('documents.create'))
                
            document_type = request.form['document_type']
            document_number = request.form.get('document_number', '')
            issue_date_str = request.form.get('issue_date', '')
            expiry_date_str = request.form.get('expiry_date', '')
            notes = request.form.get('notes', '')
            add_type = request.form.get('add_type', 'single')
            
            # Parse dates (فقط إذا تم إدخالها)
            issue_date = parse_date(issue_date_str) if issue_date_str else None
            expiry_date = parse_date(expiry_date_str) if expiry_date_str else None
            
            # تحديد ما إذا كان الإضافة لموظف واحد أو لقسم كامل
            if add_type == 'single':
                # إضافة وثيقة لموظف واحد
                employee_id = request.form.get('employee_id')
                if not employee_id:
                    flash('يرجى اختيار الموظف', 'danger')
                    return redirect(url_for('documents.create'))
                
                # Create new document record
                document = Document(
                    employee_id=employee_id,
                    document_type=document_type,
                    document_number=document_number,
                    issue_date=issue_date,
                    expiry_date=expiry_date,
                    notes=notes
                )
                
                db.session.add(document)
                
                # Log the action
                employee = Employee.query.get(employee_id)
                audit = SystemAudit(
                    action='create',
                    entity_type='document',
                    entity_id=employee_id,
                    details=f'تم إضافة وثيقة جديدة من نوع {document_type} للموظف: {employee.name}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                flash('تم إضافة الوثيقة بنجاح', 'success')
            
            elif add_type == 'sponsorship_single':
                # حفظ موظف واحد من قائمة الكفالة
                employee_id = request.form.get('employee_id')
                
                if not employee_id:
                    return jsonify({'success': False, 'message': 'يرجى اختيار الموظف'})
                
                # Create new document record
                document = Document(
                    employee_id=employee_id,
                    document_type=document_type,
                    document_number=document_number,
                    issue_date=issue_date,
                    expiry_date=expiry_date,
                    notes=notes
                )
                
                db.session.add(document)
                
                # Log the action
                employee = Employee.query.get(employee_id)
                audit = SystemAudit(
                    action='create',
                    entity_type='document',
                    entity_id=employee_id,
                    details=f'تم إضافة وثيقة {document_type} للموظف: {employee.name} (حفظ فردي من قائمة الكفالة)',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({'success': True, 'message': 'تم حفظ الوثيقة بنجاح'})
            
            elif add_type == 'department_bulk':
                # حفظ جميع وثائق القسم
                import json
                employees_data = json.loads(request.form.get('employees_data', '[]'))
                department_id = request.form.get('department_id')
                
                if not employees_data:
                    return jsonify({'success': False, 'message': 'لا توجد بيانات للحفظ'})
                
                saved_count = 0
                for emp_data in employees_data:
                    if emp_data.get('document_number'):
                        document = Document(
                            employee_id=emp_data['employee_id'],
                            document_type=document_type,
                            document_number=emp_data['document_number'],
                            issue_date=parse_date(emp_data['issue_date']) if emp_data.get('issue_date') else None,
                            expiry_date=parse_date(emp_data['expiry_date']) if emp_data.get('expiry_date') else None,
                            notes=emp_data.get('notes', '')
                        )
                        
                        db.session.add(document)
                        saved_count += 1
                
                # Log the action
                department = Department.query.get(department_id)
                audit = SystemAudit(
                    action='bulk_create',
                    entity_type='document',
                    entity_id=department_id,
                    details=f'تم إضافة {saved_count} وثيقة من نوع {document_type} لقسم {department.name if department else "غير محدد"}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({'success': True, 'message': f'تم حفظ {saved_count} وثيقة بنجاح'})
            
            elif add_type == 'sponsorship_bulk':
                # حفظ جميع بيانات الكفالة
                import json
                employees_data = json.loads(request.form.get('employees', '[]'))
                
                if not employees_data:
                    return jsonify({'success': False, 'message': 'لا توجد بيانات للحفظ'})
                
                saved_count = 0
                for emp_data in employees_data:
                    # تحقق من وجود بيانات للحفظ
                    if emp_data.get('documentNumber'):
                        document = Document(
                            employee_id=emp_data['id'],
                            document_type=document_type,
                            document_number=emp_data['documentNumber'],
                            issue_date=parse_date(emp_data['issueDate']) if emp_data.get('issueDate') else None,
                            expiry_date=parse_date(emp_data['expiryDate']) if emp_data.get('expiryDate') else None,
                            notes=emp_data.get('notes', '')
                        )
                        
                        db.session.add(document)
                        saved_count += 1
                
                # Log the action
                audit = SystemAudit(
                    action='bulk_create',
                    entity_type='document',
                    entity_id=None,
                    details=f'تم إضافة {saved_count} وثيقة من نوع {document_type} (حفظ جماعي من قائمة الكفالة)',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({'success': True, 'message': f'تم حفظ {saved_count} وثيقة بنجاح'})
            
            elif add_type == 'sponsorship_individual':
                # حفظ موظف واحد من الإضافة الجماعية
                employee_id = request.form.get('employee_id')
                sponsorship_type = request.form.get('sponsorship_type')
                
                if not employee_id:
                    return jsonify({'success': False, 'message': 'يرجى اختيار الموظف'})
                
                # Create new document record
                document = Document(
                    employee_id=employee_id,
                    document_type=document_type,
                    document_number=document_number,
                    issue_date=issue_date,
                    expiry_date=expiry_date,
                    notes=notes
                )
                
                db.session.add(document)
                
                # Log the action
                employee = Employee.query.get(employee_id)
                audit = SystemAudit(
                    action='create',
                    entity_type='document',
                    entity_id=employee_id,
                    details=f'تم إضافة وثيقة {document_type} (كفالة: {sponsorship_type}) للموظف: {employee.name}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({'success': True, 'message': 'تم حفظ الوثيقة بنجاح'})
            
            else:
                # إضافة وثيقة لقسم كامل
                department_id = request.form.get('department_id')
                if not department_id:
                    flash('يرجى اختيار القسم', 'danger')
                    return redirect(url_for('documents.create'))
                
                # الحصول على جميع موظفي القسم
                department = Department.query.get(department_id)
                employees = department.employees
                
                if not employees:
                    flash(f'لا يوجد موظفين في قسم "{department.name}"', 'warning')
                    return redirect(url_for('documents.create'))
                
                # إنشاء وثيقة لكل موظف في القسم
                document_count = 0
                for employee in employees:
                    # التحقق من عدم وجود وثيقة من نفس النوع للموظف (منع التكرار)
                    existing_document = Document.query.filter_by(
                        employee_id=employee.id,
                        document_type=document_type
                    ).first()
                    
                    # تخطي الموظف إذا كان لديه وثيقة من نفس النوع بالفعل
                    if existing_document:
                        continue
                        
                    # استخدام رقم الهوية الوطنية للموظف إذا كان متوفراً
                    # أو استخدام رقم الموظف التسلسلي إذا كان رقم الهوية غير متوفر
                    national_id = None
                    
                    # ابحث عن وثيقة هوية وطنية مسجلة للموظف
                    existing_national_id = Document.query.filter_by(
                        employee_id=employee.id,
                        document_type='national_id'
                    ).first()
                    
                    if existing_national_id:
                        national_id = existing_national_id.document_number
                    
                    # إذا لم نجد رقم هوية، نستخدم الرقم الوظيفي (المسلسل) للموظف
                    document_number_to_use = national_id if national_id else f"ID-{employee.employee_id}"
                    
                    document = Document(
                        employee_id=employee.id,
                        document_type=document_type,
                        document_number=document_number_to_use,
                        issue_date=issue_date,
                        expiry_date=expiry_date,
                        notes=notes
                    )
                    
                    db.session.add(document)
                    document_count += 1
                
                # إنشاء سجل تدقيق للعملية
                audit = SystemAudit(
                    action='create_bulk',
                    entity_type='document',
                    entity_id=department_id,
                    details=f'تم إضافة {document_count} وثيقة من نوع {document_type} لقسم: {department.name}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                flash(f'تم إضافة {document_count} وثائق بنجاح لجميع موظفي قسم "{department.name}"', 'success')
            
            return redirect(url_for('documents.index'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # Get all employees for dropdown with their departments loaded
    employees = Employee.query.options(db.selectinload(Employee.departments)).all()
    
    # Get all departments for dropdown
    departments = Department.query.all()
    
    # Get document types for dropdown
    document_types = [
        'national_id', 'passport', 'health_certificate', 
        'work_permit', 'education_certificate', 'driving_license',
        'annual_leave', 'other'
    ]
    
    # Default dates
    today = datetime.now().date()
    hijri_today = format_date_hijri(today)
    gregorian_today = format_date_gregorian(today)
    
    return render_template('documents/create.html',
                          employees=employees,
                          departments=departments,
                          document_types=document_types,
                          today=today,
                          hijri_today=hijri_today,
                          gregorian_today=gregorian_today)

@documents_bp.route('/<int:id>/confirm-delete')
def confirm_delete(id):
    """صفحة تأكيد حذف وثيقة"""
    document = Document.query.get_or_404(id)
    
    # تحويل أنواع الوثائق إلى عربي للعرض
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'health_certificate': 'الشهادة الصحية',
        'work_permit': 'تصريح العمل',
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'residency_permit': 'تصريح الإقامة',
        'visa': 'تأشيرة',
        'insurance': 'التأمين',
        'contract': 'العقد',
        'certification': 'شهادة مهنية',
        'training_certificate': 'شهادة تدريب',
        'other': 'أخرى'
    }
    
    # الحصول على اسم نوع الوثيقة بالعربية
    document_type_ar = document_types_map.get(document.document_type, document.document_type)
    
    return render_template('documents/confirm_delete.html', 
                          document=document, 
                          document_type_ar=document_type_ar)

@documents_bp.route('/<int:id>/delete', methods=['POST'])
def delete(id):
    """Delete a document record"""
    # تحقق من وجود CSRF token
    if 'csrf_token' not in request.form:
        flash('خطأ في التحقق من الأمان. يرجى المحاولة مرة أخرى.', 'danger')
        return redirect(url_for('documents.index'))
    
    document = Document.query.get_or_404(id)
    employee_name = document.employee.name
    document_type = document.document_type
    
    try:
        db.session.delete(document)
        
        # Log the action
        audit = SystemAudit(
            action='delete',
            entity_type='document',
            entity_id=id,
            details=f'تم حذف وثيقة من نوع {document_type} للموظف: {employee_name}',
            user_id=current_user.id if current_user.is_authenticated else None
        )
        db.session.add(audit)
        db.session.commit()
        
        flash('تم حذف الوثيقة بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الوثيقة: {str(e)}', 'danger')
    
    return redirect(url_for('documents.index'))

@documents_bp.route('/<int:id>/update_expiry', methods=['GET', 'POST'])
def update_expiry(id):
    """تحديث تاريخ انتهاء الوثيقة"""
    document = Document.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # تحقق من وجود CSRF token
            if 'csrf_token' not in request.form:
                flash('خطأ في التحقق من الأمان. يرجى المحاولة مرة أخرى.', 'danger')
                return redirect(url_for('documents.update_expiry', id=id))
            
            expiry_date_str = request.form['expiry_date']
            # تحليل التاريخ
            expiry_date = parse_date(expiry_date_str)
            
            # حفظ التاريخ القديم للسجل
            old_expiry_date = document.expiry_date
            
            # تحديث تاريخ الانتهاء
            document.expiry_date = expiry_date
            
            # إضافة سجل للتدقيق
            audit = SystemAudit(
                action='update',
                entity_type='document',
                entity_id=id,
                details=f'تم تحديث تاريخ انتهاء وثيقة {document.document_type} للموظف: {document.employee.name} من {old_expiry_date} إلى {expiry_date}',
                user_id=current_user.id if current_user.is_authenticated else None
            )
            db.session.add(audit)
            db.session.commit()
            
            flash('تم تحديث تاريخ انتهاء الوثيقة بنجاح', 'success')
            return redirect(url_for('documents.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث تاريخ انتهاء الوثيقة: {str(e)}', 'danger')
            return redirect(url_for('documents.update_expiry', id=id))
    
    # Get document types for dropdown
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'health_certificate': 'الشهادة الصحية',
        'work_permit': 'تصريح العمل',
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'other': 'أخرى'
    }
    
    # احصل على اسم نوع الوثيقة بالعربية
    doc_type_ar = document_types_map.get(document.document_type, document.document_type)
    
    # Default dates
    today = datetime.now().date()
    hijri_today = format_date_hijri(today)
    gregorian_today = format_date_gregorian(today)
    
    return render_template('documents/update_expiry.html',
                          document=document,
                          document_type_ar=doc_type_ar,
                          today=today,
                          hijri_today=hijri_today,
                          gregorian_today=gregorian_today)

@documents_bp.route('/import', methods=['GET', 'POST'])
def import_excel():
    """Import document records from Excel file"""
    if request.method == 'POST':
        # تحقق من وجود CSRF token
        if 'csrf_token' not in request.form:
            flash('خطأ في التحقق من الأمان. يرجى المحاولة مرة أخرى.', 'danger')
            return redirect(request.url)
            
        if 'file' not in request.files:
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                # Parse Excel file
                documents_data = parse_document_excel(file)
                success_count = 0
                error_count = 0
                
                for data in documents_data:
                    try:
                        document = Document(**data)
                        db.session.add(document)
                        db.session.commit()
                        success_count += 1
                    except Exception:
                        db.session.rollback()
                        error_count += 1
                
                # Log the import
                audit = SystemAudit(
                    action='import',
                    entity_type='document',
                    entity_id=0,
                    details=f'تم استيراد {success_count} وثيقة بنجاح و {error_count} فشل',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                if error_count > 0:
                    flash(f'تم استيراد {success_count} وثيقة بنجاح و {error_count} فشل', 'warning')
                else:
                    flash(f'تم استيراد {success_count} وثيقة بنجاح', 'success')
                return redirect(url_for('documents.index'))
            except Exception as e:
                flash(f'حدث خطأ أثناء استيراد الملف: {str(e)}', 'danger')
        else:
            flash('الملف يجب أن يكون بصيغة Excel (.xlsx, .xls)', 'danger')
    
    return render_template('documents/import.html')

@documents_bp.route('/expiring')
def expiring():
    """Show documents that are about to expire or already expired"""
    days = int(request.args.get('days', '30'))
    document_type = request.args.get('document_type', '')
    status = request.args.get('status', 'expiring')  # 'expiring' or 'expired'
    employee_id = request.args.get('employee_id', '')
    department_id = request.args.get('department_id', '')
    sponsorship_status = request.args.get('sponsorship_status', '')
    
    # Calculate expiry date range
    today = datetime.now().date()
    future_date = today + timedelta(days=days)
    
    # Build query for documents based on status
    query = Document.query.filter(Document.expiry_date.isnot(None))  # فقط الوثائق التي لها تاريخ انتهاء
    
    if status == 'expired':
        # Get documents that have already expired
        query = query.filter(Document.expiry_date < today)
    else:
        # Get documents that are about to expire
        query = query.filter(
            Document.expiry_date <= future_date,
            Document.expiry_date >= today
        )
    
    # Apply document type filter if provided
    if document_type:
        query = query.filter(Document.document_type == document_type)
    
    # Apply employee filter if provided
    if employee_id and employee_id.isdigit():
        query = query.filter(Document.employee_id == int(employee_id))
    
    # Apply filters that require Employee join
    needs_employee_join = department_id or sponsorship_status
    
    if needs_employee_join:
        query = query.join(Employee)
        
        if department_id and department_id.isdigit():
            query = query.filter(Employee.department_id == int(department_id))
        
        if sponsorship_status:
            query = query.filter(Employee.sponsorship_status == sponsorship_status)
    
    # Execute query
    documents = query.all()
    
    # Calculate days to expiry for each document
    today = datetime.now().date()
    for doc in documents:
        if doc.expiry_date:
            doc.days_to_expiry = (doc.expiry_date - today).days
        else:
            doc.days_to_expiry = None
    
    # Get document types for filter dropdown
    document_types = [
        'national_id', 'passport', 'health_certificate', 
        'work_permit', 'education_certificate', 'driving_license',
        'annual_leave', 'other'
    ]
    
    # Get all employees and departments for filter dropdowns
    employees = Employee.query.all()
    departments = Department.query.all()
    
    return render_template('documents/expiring.html',
                          documents=documents,
                          days=days,
                          document_types=document_types,
                          employees=employees,
                          departments=departments,
                          selected_type=document_type,
                          selected_employee=employee_id,
                          selected_department=department_id,
                          selected_sponsorship=sponsorship_status,
                          status=status)

@documents_bp.route('/expiry_stats')
def expiry_stats():
    """Get document expiry statistics"""
    # Calculate expiry date ranges
    today = datetime.now().date()
    thirty_days = today + timedelta(days=30)
    sixty_days = today + timedelta(days=60)
    ninety_days = today + timedelta(days=90)
    
    # استبعاد الوثائق التي ليس لها تاريخ انتهاء
    base_query = Document.query.filter(Document.expiry_date.isnot(None))
    
    # Get count of documents expiring in different periods
    expiring_30 = base_query.filter(
        Document.expiry_date <= thirty_days,
        Document.expiry_date >= today
    ).count()
    
    expiring_60 = base_query.filter(
        Document.expiry_date <= sixty_days,
        Document.expiry_date > thirty_days
    ).count()
    
    expiring_90 = base_query.filter(
        Document.expiry_date <= ninety_days,
        Document.expiry_date > sixty_days
    ).count()
    
    expired = base_query.filter(
        Document.expiry_date < today
    ).count()
    
    # Get document counts by type
    type_counts = db.session.query(
        Document.document_type,
        func.count(Document.id).label('count')
    ).group_by(Document.document_type).all()
    
    # Format for response
    type_stats = {}
    for doc_type, count in type_counts:
        type_stats[doc_type] = count
    
    return jsonify({
        'expiring_30': expiring_30,
        'expiring_60': expiring_60,
        'expiring_90': expiring_90,
        'expired': expired,
        'type_stats': type_stats
    })

@documents_bp.route('/employee/<int:employee_id>/export_pdf')
def export_employee_documents_pdf(employee_id):
    """Export employee documents to PDF"""
    employee = Employee.query.get_or_404(employee_id)
    documents = Document.query.filter_by(employee_id=employee_id).all()
    
    # إنشاء ملف PDF
    buffer = BytesIO()
    
    # تسجيل الخط العربي
    try:
        # محاولة تسجيل الخط العربي إذا لم يكن مسجلاً مسبقًا
        pdfmetrics.registerFont(TTFont('Arabic', 'static/fonts/Arial.ttf'))
    except:
        # إذا كان هناك خطأ، نستخدم الخط الافتراضي
        pass
    
    # تعيين أبعاد الصفحة واتجاهها
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # إعداد الأنماط
    styles = getSampleStyleSheet()
    # إنشاء نمط للنص العربي
    arabic_style = ParagraphStyle(
        name='Arabic',
        parent=styles['Normal'],
        fontName='Arabic',
        fontSize=10,
        alignment=2, # يمين (RTL)
        textColor=colors.black
    )
    
    # إنشاء نمط للعناوين
    title_style = ParagraphStyle(
        name='Title',
        parent=styles['Title'],
        fontName='Arabic',
        fontSize=16,
        alignment=1, # وسط
        textColor=colors.black
    )
    
    # إنشاء نمط للعناوين الفرعية
    subtitle_style = ParagraphStyle(
        name='Subtitle',
        parent=styles['Heading2'],
        fontName='Arabic',
        fontSize=14,
        alignment=2, # يمين (RTL)
        textColor=colors.blue
    )
    
    # إعداد المحتوى
    elements = []
    
    # إضافة العنوان
    title = f"وثائق الموظف: {employee.name}"
    # تهيئة النص العربي للعرض في PDF
    title = get_display(arabic_reshaper.reshape(title))
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 20))
    
    # إضافة بيانات الموظف في جدول
    employee_data = [
        [get_display(arabic_reshaper.reshape("بيانات الموظف")), "", get_display(arabic_reshaper.reshape("معلومات العمل")), ""],
        [
            get_display(arabic_reshaper.reshape("الاسم:")), 
            get_display(arabic_reshaper.reshape(employee.name)), 
            get_display(arabic_reshaper.reshape("المسمى الوظيفي:")), 
            get_display(arabic_reshaper.reshape(employee.job_title))
        ],
        [
            get_display(arabic_reshaper.reshape("الرقم الوظيفي:")), 
            employee.employee_id, 
            get_display(arabic_reshaper.reshape("القسم:")), 
            get_display(arabic_reshaper.reshape(', '.join([dept.name for dept in employee.departments]) if employee.departments else '-'))
        ],
        [
            get_display(arabic_reshaper.reshape("رقم الهوية:")), 
            employee.national_id, 
            get_display(arabic_reshaper.reshape("الحالة:")), 
            get_display(arabic_reshaper.reshape(employee.status))
        ],
        [
            get_display(arabic_reshaper.reshape("رقم الجوال:")), 
            employee.mobile, 
            get_display(arabic_reshaper.reshape("الموقع:")), 
            get_display(arabic_reshaper.reshape(employee.location or '-'))
        ]
    ]
    
    # إنشاء جدول بيانات الموظف
    employee_table = Table(employee_data, colWidths=[3*cm, 5*cm, 3*cm, 5*cm])
    employee_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.lightblue),
        ('BACKGROUND', (2, 0), (3, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (3, 0), colors.black),
        ('FONTNAME', (0, 0), (3, 0), 'Arabic'),
        ('FONTSIZE', (0, 0), (3, 0), 12),
        ('SPAN', (0, 0), (1, 0)),
        ('SPAN', (2, 0), (3, 0)),
        ('ALIGN', (0, 0), (3, 0), 'CENTER'),
        ('VALIGN', (0, 0), (3, 4), 'MIDDLE'),
        ('GRID', (0, 0), (3, 4), 0.5, colors.grey),
        ('BOX', (0, 0), (1, 4), 1, colors.black),
        ('BOX', (2, 0), (3, 4), 1, colors.black),
    ]))
    elements.append(employee_table)
    elements.append(Spacer(1, 20))
    
    # إضافة عنوان قائمة الوثائق
    subtitle = get_display(arabic_reshaper.reshape("قائمة الوثائق"))
    elements.append(Paragraph(subtitle, subtitle_style))
    elements.append(Spacer(1, 10))
    
    # إنشاء جدول الوثائق
    headers = [
        get_display(arabic_reshaper.reshape("نوع الوثيقة")),
        get_display(arabic_reshaper.reshape("رقم الوثيقة")),
        get_display(arabic_reshaper.reshape("تاريخ الإصدار")),
        get_display(arabic_reshaper.reshape("تاريخ الانتهاء")),
        get_display(arabic_reshaper.reshape("الحالة")),
        get_display(arabic_reshaper.reshape("ملاحظات"))
    ]
    
    data = [headers]
    
    # إضافة صفوف الوثائق
    today = datetime.now().date()
    
    # ترجمة أنواع الوثائق
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'health_certificate': 'الشهادة الصحية',
        'work_permit': 'تصريح العمل',
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية'
    }
    
    for doc in documents:
        # الحصول على نوع الوثيقة بالعربية
        doc_type_ar = document_types_map.get(doc.document_type, doc.document_type)
        
        # التحقق من حالة انتهاء الصلاحية
        days_to_expiry = (doc.expiry_date - today).days
        if days_to_expiry < 0:
            status_text = "منتهية"
        elif days_to_expiry < 30:
            status_text = f"تنتهي خلال {days_to_expiry} يوم"
        else:
            status_text = "سارية"
        
        # إضافة صف للجدول
        row = [
            get_display(arabic_reshaper.reshape(doc_type_ar)),
            doc.document_number,
            format_date_gregorian(doc.issue_date),
            format_date_gregorian(doc.expiry_date),
            get_display(arabic_reshaper.reshape(status_text)),
            get_display(arabic_reshaper.reshape(doc.notes or '-'))
        ]
        data.append(row)
    
    # إنشاء جدول الوثائق إذا كان هناك وثائق
    if len(data) > 1:
        # حساب عرض الأعمدة بناءً على عرض الصفحة
        table_width = A4[0] - 4*cm  # العرض الإجمالي ناقص الهوامش
        col_widths = [3.5*cm, 3*cm, 2.5*cm, 2.5*cm, 3*cm, 3*cm]
        documents_table = Table(data, colWidths=col_widths)
        
        # تنسيق الجدول
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        # تطبيق التناوب في ألوان الصفوف
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
            
            # إضافة ألوان حالة انتهاء الصلاحية
            days_to_expiry = (documents[i-1].expiry_date - today).days
            if days_to_expiry < 0:
                table_style.add('TEXTCOLOR', (4, i), (4, i), colors.red)
                table_style.add('FONTSIZE', (4, i), (4, i), 10)
            elif days_to_expiry < 30:
                table_style.add('TEXTCOLOR', (4, i), (4, i), colors.orange)
        
        documents_table.setStyle(table_style)
        elements.append(documents_table)
    else:
        # إذا لم تكن هناك وثائق
        no_data_text = get_display(arabic_reshaper.reshape("لا توجد وثائق مسجلة لهذا الموظف"))
        elements.append(Paragraph(no_data_text, arabic_style))
    
    # إضافة معلومات التقرير في أسفل الصفحة
    elements.append(Spacer(1, 30))
    footer_text = f"تم إنشاء هذا التقرير بتاريخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    footer_text = get_display(arabic_reshaper.reshape(footer_text))
    elements.append(Paragraph(footer_text, arabic_style))
    
    # بناء المستند
    doc.build(elements)
    
    # إعادة المؤشر إلى بداية البايت
    buffer.seek(0)
    
    # إنشاء استجابة تحميل
    buffer.seek(0)
    return make_response(send_file(
        buffer,
        as_attachment=True,
        download_name=f'employee_{employee_id}_documents.pdf',
        mimetype='application/pdf'
    ))

@documents_bp.route('/employee/<int:employee_id>/export_excel')
def export_employee_documents_excel(employee_id):
    """Export employee documents to Excel"""
    employee = Employee.query.get_or_404(employee_id)
    documents = Document.query.filter_by(employee_id=employee_id).all()
    
    # Create Excel in memory
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet("الوثائق")
    
    # Add formatting
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#D3E0EA',
        'border': 1,
        'font_size': 13
    })
    
    # RTL format for workbook
    worksheet.right_to_left()
    
    # Add cell formats
    cell_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'font_size': 11
    })
    
    date_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'font_size': 11,
        'num_format': 'dd/mm/yyyy'
    })
    
    # Write headers
    headers = ['نوع الوثيقة', 'رقم الوثيقة', 'تاريخ الإصدار', 'تاريخ الانتهاء', 'ملاحظات']
    for col_num, data in enumerate(headers):
        worksheet.write(0, col_num, data, header_format)
    
    # Adjust column widths
    worksheet.set_column(0, 0, 20)  # نوع الوثيقة
    worksheet.set_column(1, 1, 20)  # رقم الوثيقة
    worksheet.set_column(2, 2, 15)  # تاريخ الإصدار
    worksheet.set_column(3, 3, 15)  # تاريخ الانتهاء
    worksheet.set_column(4, 4, 30)  # ملاحظات
    
    # Map for document types
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'health_certificate': 'الشهادة الصحية',
        'work_permit': 'تصريح العمل',
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية'
    }
    
    # Write data
    for row_num, doc in enumerate(documents, 1):
        # Get document type in Arabic
        doc_type_ar = document_types_map.get(doc.document_type, doc.document_type)
            
        worksheet.write(row_num, 0, doc_type_ar, cell_format)
        worksheet.write(row_num, 1, doc.document_number, cell_format)
        worksheet.write_datetime(row_num, 2, doc.issue_date, date_format)
        worksheet.write_datetime(row_num, 3, doc.expiry_date, date_format)
        worksheet.write(row_num, 4, doc.notes or '', cell_format)
    
    # Add title with employee info
    info_worksheet = workbook.add_worksheet("معلومات الموظف")
    info_worksheet.right_to_left()
    
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#B8D9EB',
        'border': 2
    })
    
    info_worksheet.merge_range('A1:B1', f'بيانات الموظف: {employee.name}', title_format)
    info_worksheet.set_column(0, 0, 20)
    info_worksheet.set_column(1, 1, 30)
    
    field_format = workbook.add_format({
        'bold': True,
        'align': 'right',
        'valign': 'vcenter',
        'bg_color': '#F0F0F0',
        'border': 1
    })
    
    info_fields = [
        ['الاسم', employee.name],
        ['الرقم الوظيفي', employee.employee_id],
        ['رقم الهوية', employee.national_id],
        ['رقم الجوال', employee.mobile],
        ['القسم', ', '.join([dept.name for dept in employee.departments]) if employee.departments else ''],
        ['المسمى الوظيفي', employee.job_title],
        ['الحالة', employee.status],
        ['الموقع', employee.location or '']
    ]
    
    for row_num, (field, value) in enumerate(info_fields):
        info_worksheet.write(row_num + 1, 0, field, field_format)
        info_worksheet.write(row_num + 1, 1, value, cell_format)
    
    # Close workbook
    workbook.close()
    
    # Create response
    output.seek(0)
    return make_response(send_file(
        output,
        as_attachment=True,
        download_name=f'employee_{employee_id}_documents.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ))


    
    # Create Excel in memory
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    
    # Get today's date for remaining days calculation
    today = datetime.now().date()
    
    # Create document status lists
    expired_docs = []
    expiring_soon_docs = []
    valid_docs = []
    
    # Categorize documents
    for doc in documents:
        # تخطي المستندات بدون تاريخ انتهاء
        if not doc.expiry_date:
            # نضع المستندات بدون تاريخ انتهاء في قائمة منفصلة
            valid_docs.append(doc)
            continue
            
        days_remaining = (doc.expiry_date - today).days
        if days_remaining < 0:
            expired_docs.append(doc)
        elif days_remaining < 30:
            expiring_soon_docs.append(doc)
        else:
            valid_docs.append(doc)
    
    # Map for document types
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'health_certificate': 'الشهادة الصحية',
        'work_permit': 'تصريح العمل',
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية'
    }
    
    # Format definitions
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#D3E0EA',
        'border': 1,
        'font_size': 13
    })
    
    cell_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'font_size': 11
    })
    
    date_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'font_size': 11,
        'num_format': 'dd/mm/yyyy'
    })
    
    expired_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'font_size': 11,
        'bg_color': '#FFC7CE',  # Light red
        'font_color': '#9C0006'  # Dark red
    })
    
    warning_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'font_size': 11,
        'bg_color': '#FFEB9C',  # Light yellow
        'font_color': '#9C6500'  # Dark orange
    })
    
    valid_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'font_size': 11,
        'bg_color': '#C6EFCE',  # Light green
        'font_color': '#006100'  # Dark green
    })
    
    # Headers definition
    headers = ['اسم الموظف', 'الرقم الوظيفي', 'القسم', 'نوع الوثيقة', 'رقم الوثيقة', 'تاريخ الإصدار', 'تاريخ الانتهاء', 'الأيام المتبقية', 'حالة الكفالة', 'ملاحظات']
    
    # Add main worksheet with all documents (sorted by expiry date)
    main_worksheet = workbook.add_worksheet("جميع الوثائق")
    main_worksheet.right_to_left()
    
    # Set column widths for main worksheet
    main_worksheet.set_column(0, 0, 25)  # اسم الموظف
    main_worksheet.set_column(1, 1, 15)  # الرقم الوظيفي
    main_worksheet.set_column(2, 2, 20)  # القسم
    main_worksheet.set_column(3, 3, 20)  # نوع الوثيقة
    main_worksheet.set_column(4, 4, 20)  # رقم الوثيقة
    main_worksheet.set_column(5, 5, 15)  # تاريخ الإصدار
    main_worksheet.set_column(6, 6, 15)  # تاريخ الانتهاء
    main_worksheet.set_column(7, 7, 15)  # الأيام المتبقية
    main_worksheet.set_column(8, 8, 15)  # حالة الكفالة
    main_worksheet.set_column(9, 9, 30)  # ملاحظات
    
    # Write headers for main worksheet
    for col_num, data in enumerate(headers):
        main_worksheet.write(0, col_num, data, header_format)
    
    # Write all documents to main worksheet
    row_num = 1
    # نرتب المستندات حسب تاريخ الانتهاء، مع وضع المستندات التي ليس لها تاريخ انتهاء في النهاية
    def sort_key(doc):
        if doc.expiry_date:
            return (0, doc.expiry_date)  # المستندات ذات تاريخ انتهاء أولاً (مفتاح فرز = 0)
        else:
            return (1, None)  # المستندات بدون تاريخ انتهاء في النهاية (مفتاح فرز = 1)
            
    for doc in sorted(documents, key=sort_key):
        # Get employee information
        employee_name = doc.employee.name if doc.employee else "غير متوفر"
        employee_id = doc.employee.employee_id if doc.employee else "غير متوفر"
        department_name = ', '.join([dept.name for dept in doc.employee.departments]) if doc.employee and doc.employee.departments else "غير متوفر"
        
        # Get document type in Arabic
        doc_type_ar = document_types_map.get(doc.document_type, doc.document_type)
        
        # Calculate remaining days
        days_format = cell_format
        if not doc.expiry_date:
            days_remaining = "غير محدد"
            days_format = cell_format  # استخدام التنسيق الافتراضي
        else:
            days_remaining = (doc.expiry_date - today).days
            # Determine format for days remaining
            if days_remaining < 0:
                days_format = expired_format
            elif days_remaining < 30:
                days_format = warning_format
            else:
                days_format = valid_format
        
        # Write data
        main_worksheet.write(row_num, 0, employee_name, cell_format)
        main_worksheet.write(row_num, 1, employee_id, cell_format)
        main_worksheet.write(row_num, 2, department_name, cell_format)
        main_worksheet.write(row_num, 3, doc_type_ar, cell_format)
        main_worksheet.write(row_num, 4, doc.document_number, cell_format)
        
        # كتابة تاريخ الإصدار - قد يكون فارغاً
        if doc.issue_date:
            main_worksheet.write_datetime(row_num, 5, doc.issue_date, date_format)
        else:
            main_worksheet.write(row_num, 5, "غير محدد", cell_format)
            
        # كتابة تاريخ الانتهاء - قد يكون فارغاً
        if doc.expiry_date:
            main_worksheet.write_datetime(row_num, 6, doc.expiry_date, date_format)
        else:
            main_worksheet.write(row_num, 6, "غير محدد", cell_format)
            
        main_worksheet.write(row_num, 7, days_remaining, days_format)
        
        # كتابة حالة الكفالة
        sponsorship_status_ar = "غير محدد"
        if doc.employee and doc.employee.sponsorship_status:
            if doc.employee.sponsorship_status == 'inside':
                sponsorship_status_ar = "على الكفالة"
            elif doc.employee.sponsorship_status == 'outside':
                sponsorship_status_ar = "خارج الكفالة"
        main_worksheet.write(row_num, 8, sponsorship_status_ar, cell_format)
        
        main_worksheet.write(row_num, 9, doc.notes or '', cell_format)
        row_num += 1
        
    # إنشاء تنسيق للمستندات بدون تاريخ انتهاء (نقلناه هنا ليكون متاحاً للدالة الفرعية)
    no_expiry_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'font_size': 11,
        'bg_color': '#D3D3D3',
        'font_color': '#666666'
    })
    
    # Function to create a worksheet for a category of documents
    def create_category_worksheet(docs, name, title_bg_color, sheet_icon=''):
        if not docs:  # Skip if no documents in this category
            return
            
        ws = workbook.add_worksheet(f"{sheet_icon}{name}")
        ws.right_to_left()
        
        # Set column widths
        for i, width in enumerate([25, 15, 20, 20, 20, 15, 15, 15, 15, 30]):
            ws.set_column(i, i, width)
        
        # Custom title format for this category
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': title_bg_color,
            'border': 2
        })
        
        # Add title
        ws.merge_range('A1:J1', f"{name} ({len(docs)})", title_format)
        
        # Write headers
        for col_num, data in enumerate(headers):
            ws.write(1, col_num, data, header_format)
        
        # Write data
        # تعريف دالة الترتيب نفسها
        def sort_key(doc):
            if doc.expiry_date:
                return (0, doc.expiry_date)  # المستندات ذات تاريخ انتهاء أولاً 
            else:
                return (1, None)  # المستندات بدون تاريخ انتهاء في النهاية
                
        for row_num, doc in enumerate(sorted(docs, key=sort_key), 2):
            # Get employee information
            employee_name = doc.employee.name if doc.employee else "غير متوفر"
            employee_id = doc.employee.employee_id if doc.employee else "غير متوفر"
            department_name = ', '.join([dept.name for dept in doc.employee.departments]) if doc.employee and doc.employee.departments else "غير متوفر"
            
            # Get document type in Arabic
            doc_type_ar = document_types_map.get(doc.document_type, doc.document_type)
            
            # Calculate remaining days
            days_format = cell_format
            if not doc.expiry_date:
                days_remaining = "غير محدد"
                days_format = no_expiry_format
            else:
                days_remaining = (doc.expiry_date - today).days
                # Determine format for days remaining
                if days_remaining < 0:
                    days_format = expired_format
                elif days_remaining < 30:
                    days_format = warning_format
                else:
                    days_format = valid_format
            
            # Write data
            ws.write(row_num, 0, employee_name, cell_format)
            ws.write(row_num, 1, employee_id, cell_format)
            ws.write(row_num, 2, department_name, cell_format)
            ws.write(row_num, 3, doc_type_ar, cell_format)
            ws.write(row_num, 4, doc.document_number, cell_format)
            
            # كتابة تاريخ الإصدار - قد يكون فارغاً
            if doc.issue_date:
                ws.write_datetime(row_num, 5, doc.issue_date, date_format)
            else:
                ws.write(row_num, 5, "غير محدد", cell_format)
                
            # كتابة تاريخ الانتهاء - قد يكون فارغاً
            if doc.expiry_date:
                ws.write_datetime(row_num, 6, doc.expiry_date, date_format)
            else:
                ws.write(row_num, 6, "غير محدد", cell_format)
                
            # كتابة الأيام المتبقية
            ws.write(row_num, 7, days_remaining, days_format)
            
            # كتابة حالة الكفالة
            sponsorship_status_ar = "غير محدد"
            if doc.employee and doc.employee.sponsorship_status:
                if doc.employee.sponsorship_status == 'inside':
                    sponsorship_status_ar = "على الكفالة"
                elif doc.employee.sponsorship_status == 'outside':
                    sponsorship_status_ar = "خارج الكفالة"
            ws.write(row_num, 8, sponsorship_status_ar, cell_format)
            
            ws.write(row_num, 9, doc.notes or '', cell_format)
    
    # Create worksheets for each category
    create_category_worksheet(expired_docs, "وثائق منتهية", '#FFD9D9', '🔴 ')
    create_category_worksheet(expiring_soon_docs, "وثائق تنتهي قريباً", '#FFF4D9', '🟠 ')
    create_category_worksheet(valid_docs, "وثائق سارية", '#E8FFE8', '🟢 ')
    
    # Add statistics worksheet
    stats_worksheet = workbook.add_worksheet("إحصائيات")
    stats_worksheet.right_to_left()
    
    # Set up statistics
    expired_count = sum(1 for doc in documents if doc.expiry_date and (doc.expiry_date - today).days < 0)
    expiring_30_count = sum(1 for doc in documents if doc.expiry_date and 0 <= (doc.expiry_date - today).days < 30)
    expiring_60_count = sum(1 for doc in documents if doc.expiry_date and 30 <= (doc.expiry_date - today).days < 60)
    expiring_90_count = sum(1 for doc in documents if doc.expiry_date and 60 <= (doc.expiry_date - today).days < 90)
    valid_count = sum(1 for doc in documents if doc.expiry_date and (doc.expiry_date - today).days >= 90)
    no_expiry_count = sum(1 for doc in documents if not doc.expiry_date)
    
    # Document counts by type
    doc_type_counts = {}
    for doc in documents:
        doc_type = document_types_map.get(doc.document_type, doc.document_type)
        if doc_type in doc_type_counts:
            doc_type_counts[doc_type] += 1
        else:
            doc_type_counts[doc_type] = 1
    
    # Write statistics
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#B8D9EB',
        'border': 2
    })
    
    stats_worksheet.merge_range('A1:B1', 'إحصائيات الوثائق', title_format)
    stats_worksheet.set_column(0, 0, 25)
    stats_worksheet.set_column(1, 1, 15)
    
    stat_header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#E6E6E6',
        'border': 1,
        'font_size': 12
    })
    
    # تنسيق للحقول
    field_format = workbook.add_format({
        'bold': True,
        'align': 'right',
        'valign': 'vcenter',
        'bg_color': '#F0F0F0',
        'border': 1
    })
    
    # Write expiry statistics
    stats_worksheet.write(2, 0, 'حالة صلاحية الوثائق', stat_header_format)
    stats_worksheet.write(2, 1, 'العدد', stat_header_format)
    
    row = 3
    # تنسيق خاص لإجمالي الوثائق
    total_format = workbook.add_format({
        'bold': True, 
        'border': 1, 
        'bg_color': '#D9D9D9'
    })
    
    # إعادة استخدام تنسيق المستندات بدون تاريخ انتهاء المعرف أعلاه
    
    stats_data = [
        ['وثائق منتهية', expired_count, expired_format],
        ['تنتهي خلال 30 يوم', expiring_30_count, warning_format],
        ['تنتهي خلال 60 يوم', expiring_60_count, cell_format],
        ['تنتهي خلال 90 يوم', expiring_90_count, cell_format],
        ['صالحة لأكثر من 90 يوم', valid_count, cell_format],
        ['بدون تاريخ انتهاء', no_expiry_count, no_expiry_format],
        ['المجموع', len(documents), total_format]
    ]
    
    for label, count, fmt in stats_data:
        stats_worksheet.write(row, 0, label, field_format)
        stats_worksheet.write(row, 1, count, fmt)
        row += 1
    
    # Add some space
    row += 2
    
    # Write document type statistics
    stats_worksheet.write(row, 0, 'أنواع الوثائق', stat_header_format)
    stats_worksheet.write(row, 1, 'العدد', stat_header_format)
    row += 1
    
    # Sort document types by count (descending)
    sorted_doc_types = sorted(doc_type_counts.items(), key=lambda x: x[1], reverse=True)
    
    for doc_type, count in sorted_doc_types:
        stats_worksheet.write(row, 0, doc_type, cell_format)
        stats_worksheet.write(row, 1, count, cell_format)
        row += 1
    
    # Close workbook
    workbook.close()
    
    # Create response
    output.seek(0)
    
    # Generate a descriptive filename
    filename_parts = []
    if document_type:
        filename_parts.append(document_types_map.get(document_type, document_type))
    if status_filter:
        status_names = {
            'expired': 'المنتهية',
            'expiring_30': 'تنتهي_خلال_30_يوم',
            'expiring_60': 'تنتهي_خلال_60_يوم', 
            'expiring_90': 'تنتهي_خلال_90_يوم',
            'valid': 'السارية'
        }
        filename_parts.append(status_names.get(status_filter, status_filter))
    if not filename_parts:
        filename_parts.append("جميع_الوثائق")
    
    filename = "_".join(filename_parts) + ".xlsx"
    
    return make_response(send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ))
    
@documents_bp.route('/export_expiring_excel')
def export_expiring_excel():
    """تصدير الوثائق المنتهية/على وشك الانتهاء إلى ملف إكسل"""
    # جلب معايير التصفية من الطلب
    days = int(request.args.get('days', '30'))
    document_type = request.args.get('document_type', '')
    status = request.args.get('status', 'expiring')  # 'expiring' or 'expired'
    employee_id = request.args.get('employee_id', '')
    department_id = request.args.get('department_id', '')
    sponsorship_status = request.args.get('sponsorship_status', '')
    
    # تحديد نطاق التاريخ
    today = datetime.now().date()
    future_date = today + timedelta(days=days)
    
    # بناء الاستعلام بناءً على الحالة
    query = Document.query.filter(Document.expiry_date.isnot(None))
    
    if status == 'expired':
        # الوثائق المنتهية
        query = query.filter(Document.expiry_date < today)
        title = "الوثائق المنتهية"
    else:
        # الوثائق التي على وشك الانتهاء
        query = query.filter(
            Document.expiry_date <= future_date,
            Document.expiry_date >= today
        )
        title = f"الوثائق التي ستنتهي خلال {days} يوم"
    
    # فلتر نوع الوثيقة
    if document_type:
        query = query.filter(Document.document_type == document_type)
    
    # فلتر الموظف
    if employee_id and employee_id.isdigit():
        query = query.filter(Document.employee_id == int(employee_id))
    
    # تطبيق الفلاتر التي تحتاج join مع Employee
    needs_employee_join = department_id or sponsorship_status
    
    if needs_employee_join:
        query = query.join(Employee)
        
        if department_id and department_id.isdigit():
            query = query.filter(Employee.department_id == int(department_id))
        
        if sponsorship_status:
            query = query.filter(Employee.sponsorship_status == sponsorship_status)
    
    # تنفيذ الاستعلام مع تحميل بيانات الموظف والأقسام
    query = query.options(selectinload(Document.employee).selectinload(Employee.departments))
    documents = query.all()
    
    # حساب الأيام المتبقية للانتهاء لكل وثيقة
    for doc in documents:
        if doc.expiry_date:
            doc.days_to_expiry = (doc.expiry_date - today).days
        else:
            doc.days_to_expiry = None
    
    # إنشاء ملف اكسل في الذاكرة
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    
    # إضافة ورقة العمل الرئيسية وتعيين اتجاهها من اليمين إلى اليسار
    worksheet = workbook.add_worksheet(title)
    worksheet.right_to_left()
    
    # إضافة التنسيقات
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#D3E0EA',
        'border': 1,
        'font_size': 13
    })
    
    # تنسيق للخلايا العادية
    cell_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })
    
    # تنسيق للتواريخ
    date_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'num_format': 'yyyy/mm/dd'
    })
    
    # تنسيق خاص للوثائق المنتهية
    expired_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'bg_color': '#FFD9D9'
    })
    
    # تنسيق خاص للوثائق التي ستنتهي قريبا
    warning_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'bg_color': '#FFF4D9'
    })

    # تنسيق خاص للوثائق بدون تاريخ انتهاء
    no_expiry_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'bg_color': '#E6E6E6'
    })
    
    # ترجمة أنواع الوثائق
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'health_certificate': 'الشهادة الصحية',
        'work_permit': 'تصريح العمل',
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'other': 'أخرى'
    }
    
    # كتابة عنوان الملف
    worksheet.merge_range('A1:I1', title, header_format)
    
    # كتابة رؤوس الأعمدة
    headers = [
        'الموظف',
        'القسم',
        'نوع الوثيقة',
        'رقم الوثيقة',
        'تاريخ الإصدار',
        'تاريخ الانتهاء',
        'المدة المتبقية',
        'حالة الكفالة',
        'ملاحظات'
    ]
    
    for col_num, header in enumerate(headers):
        worksheet.write(1, col_num, header, header_format)
    
    # ضبط عرض الأعمدة
    worksheet.set_column(0, 0, 25)  # الموظف
    worksheet.set_column(1, 1, 20)  # القسم
    worksheet.set_column(2, 2, 20)  # نوع الوثيقة
    worksheet.set_column(3, 3, 20)  # رقم الوثيقة
    worksheet.set_column(4, 4, 15)  # تاريخ الإصدار
    worksheet.set_column(5, 5, 15)  # تاريخ الانتهاء
    worksheet.set_column(6, 6, 15)  # المدة المتبقية
    worksheet.set_column(7, 7, 15)  # حالة الكفالة
    worksheet.set_column(8, 8, 30)  # ملاحظات
    
    # كتابة البيانات
    for row_num, doc in enumerate(documents, 2):
        # الحصول على اسم الموظف والأقسام
        employee_name = doc.employee.name if doc.employee else "غير محدد"
        # استخدام علاقة many-to-many للأقسام
        department_name = ', '.join([dept.name for dept in doc.employee.departments]) if doc.employee and doc.employee.departments else "غير محدد"
        
        # الحصول على نوع الوثيقة بالعربية
        doc_type_ar = document_types_map.get(doc.document_type, doc.document_type)
        
        # تحديد تنسيق الخلية بناءً على حالة انتهاء الوثيقة
        days_format = cell_format
        if doc.days_to_expiry is not None:
            if doc.days_to_expiry < 0:
                days_format = expired_format
            elif doc.days_to_expiry < 30:
                days_format = warning_format
        else:
            days_format = no_expiry_format
        
        # كتابة بيانات الوثيقة
        worksheet.write(row_num, 0, employee_name, cell_format)
        worksheet.write(row_num, 1, department_name, cell_format)
        worksheet.write(row_num, 2, doc_type_ar, cell_format)
        worksheet.write(row_num, 3, doc.document_number, cell_format)
        
        # كتابة تاريخ الإصدار - قد يكون فارغاً
        if doc.issue_date:
            worksheet.write_datetime(row_num, 4, doc.issue_date, date_format)
        else:
            worksheet.write(row_num, 4, "غير محدد", cell_format)
            
        # كتابة تاريخ الانتهاء - قد يكون فارغاً
        if doc.expiry_date:
            worksheet.write_datetime(row_num, 5, doc.expiry_date, date_format)
        else:
            worksheet.write(row_num, 5, "غير محدد", cell_format)
            
        # كتابة الأيام المتبقية
        if doc.days_to_expiry is not None:
            days_display = doc.days_to_expiry
            if doc.days_to_expiry < 0:
                days_display = f"منتهية منذ {-doc.days_to_expiry} يوم"
            else:
                days_display = f"{doc.days_to_expiry} يوم"
            worksheet.write(row_num, 6, days_display, days_format)
        else:
            worksheet.write(row_num, 6, "غير محدد", no_expiry_format)
        
        # كتابة حالة الكفالة
        sponsorship_status_ar = "غير محدد"
        if doc.employee and doc.employee.sponsorship_status:
            if doc.employee.sponsorship_status == 'inside':
                sponsorship_status_ar = "على الكفالة"
            elif doc.employee.sponsorship_status == 'outside':
                sponsorship_status_ar = "خارج الكفالة"
        worksheet.write(row_num, 7, sponsorship_status_ar, cell_format)
            
        worksheet.write(row_num, 8, doc.notes or '', cell_format)
    
    # إغلاق المصنف
    workbook.close()
    
    # إنشاء استجابة
    output.seek(0)
    
    # توليد اسم ملف وصفي
    filename_parts = [status]
    if document_type:
        filename_parts.append(document_types_map.get(document_type, document_type))
    if status == 'expiring':
        filename_parts.append(f"خلال_{days}_يوم")
    
    filename = "_".join(filename_parts) + ".xlsx"
    
    return make_response(send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ))



