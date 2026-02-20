from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy import func, extract, or_
from datetime import datetime, time, timedelta, date
from core.extensions import db
from models import Attendance, Employee, Department, SystemAudit, employee_departments
from utils.date_converter import parse_date, format_date_hijri, format_date_gregorian
from utils.excel import export_attendance_by_department
from utils.excel_dashboard import export_attendance_by_department_with_dashboard
from utils.audit_logger import log_activity
from services.attendance_analytics import AttendanceAnalytics
from services.attendance_engine import AttendanceEngine
import calendar
import logging
import time as time_module
from utils.decorators import module_access_required
import pandas as pd
from io import BytesIO

logger = logging.getLogger(__name__)
attendance_bp = Blueprint('attendance', __name__)


def create_absence_notification(user_id, employee_name, absence_date, department_name, employee_id):
    """إشعار غياب موظف"""
    from models import Notification, User
    
    notification = Notification(
        user_id=user_id,
        notification_type='absence',
        title=f'غياب موظف - {employee_name}',
        description=f'تم تسجيل غياب الموظف {employee_name} من قسم {department_name} بتاريخ {absence_date}',
        related_entity_type='attendance',
        related_entity_id=employee_id,
        priority='normal',
        action_url=url_for('attendance.index')
    )
    db.session.add(notification)
    return notification


@attendance_bp.route('/test-notifications', methods=['GET', 'POST'])
def test_absence_notifications():
    """اختبار إنشاء إشعارات الغياب لجميع المستخدمين"""
    try:
        from models import Notification, User
        
        today = datetime.now().date()
        
        # الحصول على موظفين غائبين اليوم
        absent_employees = Employee.query.join(Attendance).filter(
            func.date(Attendance.date) == today,
            Attendance.status == 'absent'
        ).limit(5).all()
        
        if not absent_employees:
            # إذا لم يوجد غياب، نأخذ أي موظفين للاختبار
            absent_employees = Employee.query.limit(3).all()
        
        if not absent_employees:
            return jsonify({'success': False, 'message': 'لا توجد سجلات غياب'}), 404
        
        all_users = User.query.all()
        
        notification_count = 0
        for emp in absent_employees:
            dept_name = emp.departments[0].name if emp.departments else 'غير محدد'
            
            for user in all_users:
                try:
                    create_absence_notification(
                        user_id=user.id,
                        employee_name=emp.name or 'غير محدد',
                        absence_date=today.strftime('%Y-%m-%d'),
                        department_name=dept_name,
                        employee_id=emp.id
                    )
                    notification_count += 1
                except Exception as e:
                    pass
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء {notification_count} إشعار لـ {len(absent_employees)} موظف',
            'employees_count': len(absent_employees),
            'users_count': len(all_users)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


def format_time_12h_ar(dt):
    """تحويل الوقت من 24 ساعة إلى 12 ساعة بصيغة عربية (صباح/مساء)"""
    if not dt:
        return '-'
    hour = dt.hour
    minute = dt.minute
    second = dt.second
    
    # تحديد صباح أو مساء
    period = 'ص' if hour < 12 else 'م'
    
    # تحويل الساعة
    if hour > 12:
        hour = hour - 12
    elif hour == 0:
        hour = 12
    
    return f'{hour}:{minute:02d}:{second:02d} {period}'

def format_time_12h_ar_short(dt):
    """تحويل الوقت من 24 ساعة إلى 12 ساعة بصيغة قصيرة (بدون ثوانٍ)"""
    if not dt:
        return '-'
    hour = dt.hour
    minute = dt.minute
    
    # تحديد صباح أو مساء
    period = 'ص' if hour < 12 else 'م'
    
    # تحويل الساعة
    if hour > 12:
        hour = hour - 12
    elif hour == 0:
        hour = 12
    
    return f'{hour}:{minute:02d} {period}'

@attendance_bp.route('/')
def index():
    """List attendance records with filtering options - shows all employees"""
    # Get filter parameters
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    department_id = request.args.get('department_id', '')
    status = request.args.get('status', '')
    
    # Parse date
    try:
        date = parse_date(date_str)
    except ValueError:
        date = datetime.now().date()
    
    # Get departments for filter dropdown based on user permissions
    from flask_login import current_user
    
    if current_user.is_authenticated:
        departments = current_user.get_accessible_departments()
        
        # Auto-filter to user's assigned department if they have one
        if current_user.assigned_department_id and not department_id:
            department_id = str(current_user.assigned_department_id)
    else:
        departments = Department.query.all()
    
    # Build employee query (active employees only)
    employee_query = Employee.query.filter_by(status='active')
    
    # Apply department filter if specified
    if department_id and department_id != '':
        employee_query = employee_query.join(employee_departments).filter(
            employee_departments.c.department_id == int(department_id)
        )
    
    # Get all active employees
    employees = employee_query.all()
    
    # Get actual attendance records for this date
    attendance_query = Attendance.query.filter(Attendance.date == date)
    if department_id and department_id != '':
        attendance_query = attendance_query.join(Employee).join(employee_departments).filter(
            employee_departments.c.department_id == int(department_id)
        )
    
    # Build a map of employee_id -> attendance record
    attendance_map = {att.employee_id: att for att in attendance_query.all()}
    
    # Build unified attendance list
    unified_attendances = []
    for emp in employees:
        att_record = attendance_map.get(emp.id)
        if att_record:
            # Employee has a record - use it
            record = {
                'id': att_record.id,
                'employee': emp,
                'date': date,
                'status': att_record.status,
                'check_in': att_record.check_in,
                'check_out': att_record.check_out,
                'notes': att_record.notes,
                'sick_leave_file': att_record.sick_leave_file,
                'has_record': True
            }
        else:
            # Employee has no record - mark as absent
            record = {
                'id': None,
                'employee': emp,
                'date': date,
                'status': 'absent',
                'check_in': None,
                'check_out': None,
                'notes': None,
                'sick_leave_file': None,
                'has_record': False
            }
        
        unified_attendances.append(record)
    
    # Apply status filter on unified list
    if status and status != '':
        unified_attendances = [rec for rec in unified_attendances if rec['status'] == status]
    
    # Calculate statistics from unified list
    present_count = sum(1 for rec in unified_attendances if rec['status'] == 'present')
    absent_count = sum(1 for rec in unified_attendances if rec['status'] == 'absent')
    leave_count = sum(1 for rec in unified_attendances if rec['status'] == 'leave')
    sick_count = sum(1 for rec in unified_attendances if rec['status'] == 'sick')
    
    # Format date for display in both calendars
    hijri_date = format_date_hijri(date)
    gregorian_date = format_date_gregorian(date)
    
    return render_template('attendance/index.html', 
                          attendances=unified_attendances,
                          departments=departments,
                          date=date,
                          hijri_date=hijri_date,
                          gregorian_date=gregorian_date,
                          selected_department=department_id,
                          selected_status=status,
                          present_count=present_count,
                          absent_count=absent_count,
                          leave_count=leave_count,
                          sick_count=sick_count)

@attendance_bp.route('/record', methods=['GET', 'POST'])
def record():
    """Record attendance for individual employees"""
    if request.method == 'POST':
        try:
            employee_id = request.form['employee_id']
            date_str = request.form['date']
            status = request.form['status']
            
            # Parse date
            date = parse_date(date_str)
            
            # Check if attendance record already exists
            existing = Attendance.query.filter_by(
                employee_id=employee_id,
                date=date
            ).first()
            
            if existing:
                # Update existing record
                existing.status = status
                existing.notes = request.form.get('notes', '')
                
                # Process check-in and check-out times if present
                if status == 'present':
                    check_in_str = request.form.get('check_in', '')
                    check_out_str = request.form.get('check_out', '')
                    
                    if check_in_str:
                        hours, minutes = map(int, check_in_str.split(':'))
                        existing.check_in = time(hours, minutes)
                    
                    if check_out_str:
                        hours, minutes = map(int, check_out_str.split(':'))
                        existing.check_out = time(hours, minutes)
                else:
                    existing.check_in = None
                    existing.check_out = None
                
                db.session.commit()
                
                # تسجيل العملية في سجل النشاط
                employee = Employee.query.get(employee_id)
                if employee:
                    log_attendance_activity(
                        action='update',
                        attendance_data={
                            'id': existing.id,
                            'employee_id': employee_id,
                            'date': date.isoformat(),
                            'status': status
                        },
                        employee_name=employee.name
                    )
                
                flash('تم تحديث سجل الحضور بنجاح', 'success')
            else:
                # Create new attendance record
                new_attendance = Attendance(
                    employee_id=employee_id,
                    date=date,
                    status=status,
                    notes=request.form.get('notes', '')
                )
                
                # Process check-in and check-out times if present and status is 'present'
                if status == 'present':
                    check_in_str = request.form.get('check_in', '')
                    check_out_str = request.form.get('check_out', '')
                    
                    if check_in_str:
                        hours, minutes = map(int, check_in_str.split(':'))
                        new_attendance.check_in = time(hours, minutes)
                    
                    if check_out_str:
                        hours, minutes = map(int, check_out_str.split(':'))
                        new_attendance.check_out = time(hours, minutes)
                
                db.session.add(new_attendance)
                db.session.commit()
                
                # تسجيل العملية في سجل النشاط
                employee = Employee.query.get(employee_id)
                if employee:
                    log_attendance_activity(
                        action='create',
                        attendance_data={
                            'id': new_attendance.id,
                            'employee_id': employee_id,
                            'date': date.isoformat(),
                            'status': status
                        },
                        employee_name=employee.name
                    )
                
                flash('تم تسجيل الحضور بنجاح', 'success')
            
            return redirect(url_for('attendance.index', date=date_str))
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # الحصول على الموظفين النشطين حسب صلاحيات المستخدم
    from flask_login import current_user
    
    employees = []
    if current_user.is_authenticated:
        try:
            user_role = current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role)
            print(f"المستخدم الحالي: {current_user.name}, الدور: {user_role}, القسم المخصص: {current_user.assigned_department_id}")
            
            if user_role in ['ADMIN', 'MANAGER', 'SUPERVISOR']:
                # المديرون والمشرفون يمكنهم رؤية جميع الموظفين (استبعاد المنتهية خدمتهم فقط)
                employees = Employee.query.filter(~Employee.status.in_(['terminated', 'inactive'])).order_by(Employee.name).all()
                print(f"تم العثور على {len(employees)} موظف للمدير/المشرف")
            elif current_user.assigned_department_id:
                # المستخدمون مع قسم مخصص يرون موظفي قسمهم فقط
                dept = Department.query.get(current_user.assigned_department_id)
                if dept:
                    employees = [emp for emp in dept.employees if emp.status not in ['terminated', 'inactive']]
                    employees.sort(key=lambda x: x.name)
                else:
                    employees = []
                print(f"تم العثور على {len(employees)} موظف للقسم {current_user.assigned_department_id}")
            else:
                # كحل بديل، عرض جميع الموظفين للمستخدمين المسجلين (استبعاد المنتهية خدمتهم فقط)
                employees = Employee.query.filter(~Employee.status.in_(['terminated', 'inactive'])).order_by(Employee.name).all()
                print(f"عرض جميع الموظفين كحل بديل: {len(employees)} موظف")
        except Exception as e:
            print(f"خطأ في تحديد صلاحيات المستخدم: {e}")
            # كحل بديل في حالة الخطأ، عرض جميع الموظفين (استبعاد المنتهية خدمتهم فقط)
            employees = Employee.query.filter(~Employee.status.in_(['terminated', 'inactive'])).order_by(Employee.name).all()
    else:
        print("المستخدم غير مسجل دخول")
    
    # Default to today's date
    today = datetime.now().date()
    hijri_date = format_date_hijri(today)
    gregorian_date = format_date_gregorian(today)
    
    return render_template('attendance/record.html', 
                          employees=employees, 
                          today=today,
                          hijri_date=hijri_date,
                          gregorian_date=gregorian_date)

@attendance_bp.route('/department', methods=['GET', 'POST'])
def department_attendance():
    """Record attendance for an entire department at once"""
    if request.method == 'POST':
        try:
            department_id = request.form['department_id']
            date_str = request.form['date']
            status = request.form['status']
            
            # التحقق من صلاحيات المستخدم للوصول إلى هذا القسم
            from flask_login import current_user
            
            if current_user.is_authenticated and not current_user.can_access_department(int(department_id)):
                flash('ليس لديك صلاحية لتسجيل حضور هذا القسم', 'error')
                return redirect(url_for('attendance.department_attendance'))
            
            # Parse date
            date = parse_date(date_str)
            
            # Get all employees in the department using many-to-many relationship
            department = Department.query.get_or_404(department_id)
            employees = [emp for emp in department.employees if emp.status not in ['terminated', 'inactive']]
            
            count = 0
            for employee in employees:
                # Check if attendance record already exists
                existing = Attendance.query.filter_by(
                    employee_id=employee.id,
                    date=date
                ).first()
                
                if existing:
                    # Update existing record
                    existing.status = status
                    if status != 'present':
                        existing.check_in = None
                        existing.check_out = None
                else:
                    # Create new attendance record
                    new_attendance = Attendance(
                        employee_id=employee.id,
                        date=date,
                        status=status
                    )
                    db.session.add(new_attendance)
                
                count += 1
            
            # تسجيل العملية في سجل النشاط
            department = Department.query.get(department_id)
            if department:
                log_attendance_activity(
                    action='bulk_create',
                    attendance_data={
                        'department_id': department_id,
                        'date': date.isoformat(),
                        'status': status,
                        'count': count
                    },
                    employee_name=f"جميع موظفي قسم {department.name}"
                )
            
            flash(f'تم تسجيل الحضور لـ {count} موظف بنجاح', 'success')
            return redirect(url_for('attendance.index', date=date_str))
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # Get departments based on user permissions
    from flask_login import current_user
    
    if current_user.is_authenticated:
        departments = current_user.get_accessible_departments()
    else:
        departments = Department.query.all()
    
    # Default to today's date
    today = datetime.now().date()
    hijri_date = format_date_hijri(today)
    gregorian_date = format_date_gregorian(today)
    
    return render_template('attendance/department.html', 
                          departments=departments,
                          today=today,
                          hijri_date=hijri_date,
                          gregorian_date=gregorian_date)

@attendance_bp.route('/bulk-record', methods=['GET', 'POST'])
def bulk_record():
    """تسجيل الحضور الجماعي للموظفين بفترات مختلفة"""
    from flask_login import current_user
    
    # التحقق من تسجيل الدخول
    if not current_user.is_authenticated:
        flash('يجب تسجيل الدخول أولاً', 'error')
        return redirect(url_for('auth.login'))
    
    # التحقق من وجود قسم مخصص للمستخدم
    if not current_user.assigned_department_id and current_user.role.value != 'admin':
        flash('يجب تخصيص قسم لك لتتمكن من تسجيل الحضور', 'error')
        return redirect(url_for('attendance.index'))
    
    if request.method == 'POST':
        try:
            period_type = request.form['period_type']
            default_status = request.form['default_status']
            employee_ids = request.form.getlist('employee_ids')
            skip_weekends = 'skip_weekends' in request.form
            overwrite_existing = 'overwrite_existing' in request.form
            
            if not employee_ids:
                flash('يجب اختيار موظف واحد على الأقل', 'error')
                return redirect(url_for('attendance.bulk_record'))
            
            # تحديد التواريخ حسب نوع الفترة
            dates = []
            
            if period_type == 'daily':
                single_date = parse_date(request.form['single_date'])
                dates = [single_date]
                
            elif period_type == 'weekly':
                week_start = parse_date(request.form['week_start'])
                dates = [week_start + timedelta(days=i) for i in range(7)]
                
            elif period_type == 'monthly':
                month_year = request.form['month_year']
                year, month = map(int, month_year.split('-'))
                import calendar
                days_in_month = calendar.monthrange(year, month)[1]
                dates = [date(year, month, day) for day in range(1, days_in_month + 1)]
                
            elif period_type == 'custom':
                start_date = parse_date(request.form['start_date'])
                end_date = parse_date(request.form['end_date'])
                current_date = start_date
                while current_date <= end_date:
                    dates.append(current_date)
                    current_date += timedelta(days=1)
            
            # تصفية عطلة نهاية الأسبوع إذا كان مطلوباً
            if skip_weekends:
                dates = [d for d in dates if d.weekday() not in [4, 5]]  # الجمعة والسبت
            
            # تسجيل الحضور
            count = 0
            for employee_id in employee_ids:
                for date_obj in dates:
                    # التحقق من وجود سجل سابق
                    existing = Attendance.query.filter_by(
                        employee_id=employee_id,
                        date=date_obj
                    ).first()
                    
                    if existing:
                        if overwrite_existing:
                            existing.status = default_status
                            count += 1
                    else:
                        attendance = Attendance(
                            employee_id=employee_id,
                            date=date_obj,
                            status=default_status
                        )
                        db.session.add(attendance)
                        count += 1
            
            db.session.commit()
            flash(f'تم تسجيل {count} سجل حضور بنجاح', 'success')
            return redirect(url_for('attendance.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # الحصول على موظفي القسم المخصص للمستخدم
    try:
        print(f"معلومات المستخدم: ID={current_user.id}, assigned_department_id={getattr(current_user, 'assigned_department_id', None)}")
        
        if hasattr(current_user, 'role') and current_user.role and current_user.role.value == 'admin':
            # المديرون العامون يمكنهم رؤية جميع الموظفين
            employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
            print(f"مدير عام - تم جلب {len(employees)} موظف")
        elif current_user.assigned_department_id:
            # المستخدمون مع قسم مخصص يرون موظفي قسمهم فقط
            employees = Employee.query.filter_by(
                status='active',
                department_id=current_user.assigned_department_id
            ).order_by(Employee.name).all()
            print(f"مستخدم قسم {current_user.assigned_department_id} - تم جلب {len(employees)} موظف")
        else:
            # جرب جلب جميع الموظفين النشطين للاختبار
            employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
            print(f"مستخدم بدون قسم - تم جلب {len(employees)} موظف للاختبار")
    except Exception as e:
        print(f"خطأ في جلب الموظفين: {str(e)}")
        employees = []
    
    today = datetime.now().date()
    
    return render_template('attendance/bulk_record.html', 
                         employees=employees,
                         today=today)

@attendance_bp.route('/all-departments-simple', methods=['GET', 'POST'])
def all_departments_attendance_simple():
    """تسجيل حضور لعدة أقسام لفترة زمنية محددة - نسخة مبسطة"""
    # التاريخ الافتراضي هو اليوم
    today = datetime.now().date()
    
    # CSRF Token - نستخدم المعالج الافتراضي
    from flask_wtf.csrf import generate_csrf
    form = {"csrf_token": generate_csrf()}
    
    if request.method == 'POST':
        try:
            # 1. استلام البيانات من النموذج
            department_ids = request.form.getlist('department_ids')
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            status = request.form.get('status', 'present')  # القيمة الافتراضية هي حاضر
            
            # 2. التحقق من البيانات المدخلة
            if not department_ids:
                flash('الرجاء اختيار قسم واحد على الأقل', 'warning')
                return redirect(url_for('attendance.all_departments_attendance_simple'))
                
            try:
                # التأكد من أن السلاسل ليست None
                if not start_date_str or not end_date_str:
                    flash('تنسيق التاريخ غير صحيح', 'danger')
                    return redirect(url_for('attendance.all_departments_attendance_simple'))
                
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('تنسيق التاريخ غير صحيح', 'danger')
                return redirect(url_for('attendance.all_departments_attendance_simple'))
                
            if end_date < start_date:
                flash('تاريخ النهاية يجب أن يكون بعد تاريخ البداية أو مساوياً له', 'warning')
                return redirect(url_for('attendance.all_departments_attendance_simple'))
                
            # 3. تهيئة المتغيرات للإحصائيات
            total_departments = 0
            total_employees = 0
            total_records = 0
            
            # 4. حساب عدد الأيام
            delta = end_date - start_date
            days_count = delta.days + 1  # لتضمين اليوم الأخير
            
            # 5. معالجة البيانات
            for dept_id in department_ids:
                try:
                    # التحقق من وجود القسم
                    department = Department.query.get(int(dept_id))
                    if not department:
                        continue
                        
                    total_departments += 1
                    
                    # الحصول على الموظفين النشطين في القسم
                    employees = Employee.query.filter_by(
                        department_id=int(dept_id),
                        status='active'
                    ).all()
                    
                    # عدد موظفي القسم
                    dept_employees_count = len(employees)
                    total_employees += dept_employees_count
                    dept_records = 0
                    
                    # معالجة كل يوم في نطاق التاريخ
                    curr_date = start_date
                    while curr_date <= end_date:
                        for employee in employees:
                            # التحقق من وجود سجل حضور سابق
                            existing = Attendance.query.filter_by(
                                employee_id=employee.id,
                                date=curr_date
                            ).first()
                            
                            if existing:
                                # تحديث السجل الموجود
                                existing.status = status
                                if status != 'present':
                                    existing.check_in = None
                                    existing.check_out = None
                            else:
                                # إنشاء سجل جديد
                                new_attendance = Attendance()
                                new_attendance.employee_id = employee.id
                                new_attendance.date = curr_date
                                new_attendance.status = status
                                if status == 'present':
                                    # يمكن إضافة وقت الدخول والخروج الافتراضي إذا كان حاضر
                                    pass
                                db.session.add(new_attendance)
                                
                            dept_records += 1
                            
                        # الانتقال لليوم التالي
                        curr_date += timedelta(days=1)
                    
                    total_records += dept_records
                    
                    try:
                        # تسجيل النشاط للقسم في سجل النظام
                        from flask_login import current_user
                        user_id = current_user.id if hasattr(current_user, 'id') else None
                        
                        SystemAudit.create_audit_record(
                            user_id=user_id,
                            action='mass_attendance',
                            entity_type='department',
                            entity_id=department.id,
                            entity_name=department.name,
                            details=f'تم تسجيل حضور لقسم {department.name} للفترة من {start_date} إلى {end_date} لعدد {dept_employees_count} موظف'
                        )
                    except Exception as audit_error:
                        # تخطي خطأ السجل
                        print(f"خطأ في تسجيل النشاط: {str(audit_error)}")
                        
                except Exception as dept_error:
                    print(f"خطأ في معالجة القسم رقم {dept_id}: {str(dept_error)}")
                    continue
            
            # حفظ التغييرات في قاعدة البيانات
            db.session.commit()
            
            # تسجيل العملية
            departments_names = [Department.query.get(dept_id).name for dept_id in department_ids if Department.query.get(dept_id)]
            log_activity('create', 'BulkAttendance', None, 
                        f"تم تسجيل حضور جماعي لـ {total_departments} قسم ({', '.join(departments_names[:3])}{'...' if len(departments_names) > 3 else ''}) - {total_employees} موظف عن {days_count} يوم")
            
            # عرض رسالة نجاح
            flash(f'تم تسجيل الحضور بنجاح لـ {total_departments} قسم و {total_employees} موظف عن {days_count} يوم (إجمالي {total_records} سجل)', 'success')
            return redirect(url_for('attendance.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء معالجة البيانات: {str(e)}', 'danger')
            print(f"خطأ: {str(e)}")
    
    # الحصول على الأقسام مع عدد الموظفين النشطين لكل قسم
    departments = []
    all_departments = Department.query.all()
    for dept in all_departments:
        active_count = Employee.query.filter_by(department_id=dept.id, status='active').count()
        # أضف جميع الأقسام حتى لو لم يكن لديها موظفين نشطين
        dept.active_employees_count = active_count
        departments.append(dept)
    
    return render_template('attendance/all_departments_simple.html',
                          departments=departments,
                          today=today,
                          form=form)

@attendance_bp.route('/all-departments', methods=['GET', 'POST'])
def all_departments_attendance():
    """تسجيل حضور لعدة أقسام لفترة زمنية محددة"""
    # التاريخ الافتراضي هو اليوم
    today = datetime.now().date()
    hijri_date = format_date_hijri(today)
    gregorian_date = format_date_gregorian(today)
    
    if request.method == 'POST':
        try:
            department_ids = request.form.getlist('department_ids')
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            status = request.form.get('status')
            
            if not department_ids:
                flash('يجب اختيار قسم واحد على الأقل.', 'danger')
                return redirect(url_for('attendance.all_departments_attendance'))
                
            if not start_date_str or not end_date_str:
                flash('يجب تحديد تاريخ البداية والنهاية.', 'danger')
                return redirect(url_for('attendance.all_departments_attendance'))
                
            if not status:
                status = 'present'  # الحالة الافتراضية هي حاضر
                
            # تحليل التواريخ
            try:
                start_date = parse_date(start_date_str)
                end_date = parse_date(end_date_str)
                
                if not start_date or not end_date:
                    flash('تاريخ غير صالح، يرجى التحقق من التنسيق.', 'danger')
                    return redirect(url_for('attendance.all_departments_attendance'))
                
                # التحقق من صحة النطاق
                if end_date < start_date:
                    flash('تاريخ النهاية يجب أن يكون بعد تاريخ البداية أو مساوياً له.', 'danger')
                    return redirect(url_for('attendance.all_departments_attendance'))
            except (ValueError, TypeError) as e:
                flash(f'خطأ في تنسيق التاريخ: {str(e)}', 'danger')
                return redirect(url_for('attendance.all_departments_attendance'))
                
            # تهيئة المتغيرات للإحصائيات
            total_departments = len(department_ids)
            total_employees = 0
            total_records = 0
            
            # حساب عدد الأيام
            delta = end_date - start_date
            days_count = delta.days + 1  # لتضمين اليوم الأخير
            
            try:
                # العمل على كل قسم من الأقسام المحددة
                for department_id in department_ids:
                    try:
                        # التحويل إلى عدد صحيح
                        dept_id = int(department_id)
                        
                        # الحصول على القسم للتأكد من وجوده
                        department = Department.query.get(dept_id)
                        if not department:
                            continue
                            
                        # الحصول على جميع الموظفين في القسم
                        employees = Employee.query.filter_by(
                            department_id=dept_id,
                            status='active'
                        ).all()
                        
                        # عدد موظفي القسم
                        department_employee_count = len(employees)
                        total_employees += department_employee_count
                        
                        # التحضير لكل يوم في النطاق المحدد
                        current_date = start_date
                        department_records = 0
                        
                        while current_date <= end_date:
                            day_count = 0
                            
                            for employee in employees:
                                try:
                                    # التحقق من وجود سجل حضور مسبق
                                    existing = Attendance.query.filter_by(
                                        employee_id=employee.id,
                                        date=current_date
                                    ).first()
                                    
                                    if existing:
                                        # تحديث السجل الموجود
                                        existing.status = status
                                        if status != 'present':
                                            existing.check_in = None
                                            existing.check_out = None
                                    else:
                                        # إنشاء سجل حضور جديد
                                        new_attendance = Attendance(
                                            employee_id=employee.id,
                                            date=current_date,
                                            status=status
                                        )
                                        db.session.add(new_attendance)
                                    
                                    day_count += 1
                                except Exception as emp_error:
                                    # تسجيل الخطأ والاستمرار مع الموظف التالي
                                    print(f"خطأ مع الموظف {employee.id}: {str(emp_error)}")
                                    continue
                            
                            # الانتقال إلى اليوم التالي
                            current_date += timedelta(days=1)
                            department_records += day_count
                        
                        total_records += department_records
                        
                        # تسجيل العملية للقسم
                        log_activity('create', 'DepartmentAttendance', department.id, 
                                   f'تم تسجيل حضور لقسم {department.name} للفترة من {start_date} إلى {end_date} لعدد {department_employee_count} موظف')
                    
                    except Exception as dept_error:
                        # تسجيل الخطأ والاستمرار مع القسم التالي
                        print(f"خطأ مع القسم {department_id}: {str(dept_error)}")
                        continue
                
                # حفظ جميع التغييرات
                db.session.commit()
                
                # رسالة نجاح مفصلة
                flash(f'تم تسجيل الحضور لـ {total_departments} قسم و {total_employees} موظف عن {days_count} يوم بنجاح (إجمالي {total_records} سجل)', 'success')
                return redirect(url_for('attendance.index', date=start_date_str))
            
            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ أثناء معالجة الأقسام: {str(e)}', 'danger')
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ عام: {str(e)}', 'danger')
    
    # الحصول على جميع الأقسام مع عدد الموظفين النشطين لكل قسم
    try:
        departments = []
        all_departments = Department.query.all()
        for dept in all_departments:
            active_count = Employee.query.filter_by(department_id=dept.id, status='active').count()
            # أضف جميع الأقسام حتى لو لم يكن لديها موظفين نشطين
            dept.active_employees_count = active_count
            departments.append(dept)
    except Exception as e:
        departments = []
        flash(f'خطأ في تحميل الأقسام: {str(e)}', 'warning')
    
    return render_template('attendance/all_departments.html', 
                          departments=departments,
                          today=today,
                          hijri_date=hijri_date,
                          gregorian_date=gregorian_date)

@attendance_bp.route('/multi-day-department', methods=['GET', 'POST'])
def multi_day_department_attendance():
    """تسجيل حضور لقسم بالكامل لفترة زمنية محددة"""
    if request.method == 'POST':
        try:
            department_id = request.form['department_id']
            start_date_str = request.form['start_date']
            end_date_str = request.form['end_date']
            status = request.form['status']
            
            # تحليل التواريخ
            try:
                start_date = parse_date(start_date_str)
                end_date = parse_date(end_date_str)
                
                if not start_date or not end_date:
                    raise ValueError("تاريخ غير صالح")
                
                # التحقق من صحة النطاق
                if end_date < start_date:
                    flash('تاريخ النهاية يجب أن يكون بعد تاريخ البداية أو مساوياً له.', 'danger')
                    return redirect(url_for('attendance.multi_day_department_attendance'))
            except (ValueError, TypeError) as e:
                flash(f'خطأ في تنسيق التاريخ: {str(e)}', 'danger')
                return redirect(url_for('attendance.multi_day_department_attendance'))
                
            # الحصول على جميع الموظفين في القسم
            employees = Employee.query.filter_by(
                department_id=department_id,
                status='active'
            ).all()
            
            # حساب عدد الأيام
            delta = end_date - start_date
            days_count = delta.days + 1  # لتضمين اليوم الأخير
            
            # التحضير لكل يوم في النطاق المحدد
            total_count = 0
            current_date = start_date
            
            while current_date <= end_date:
                day_count = 0
                
                for employee in employees:
                    # التحقق من وجود سجل حضور مسبق
                    existing = Attendance.query.filter_by(
                        employee_id=employee.id,
                        date=current_date
                    ).first()
                    
                    if existing:
                        # تحديث السجل الموجود
                        existing.status = status
                        if status != 'present':
                            existing.check_in = None
                            existing.check_out = None
                    else:
                        # إنشاء سجل حضور جديد
                        new_attendance = Attendance(
                            employee_id=employee.id,
                            date=current_date,
                            status=status
                        )
                        db.session.add(new_attendance)
                    
                    day_count += 1
                
                # الانتقال إلى اليوم التالي
                current_date += timedelta(days=1)
                total_count += day_count
            
            # تسجيل العملية
            department = Department.query.get(department_id)
            if department:
                log_activity('create', 'MultiDayDepartmentAttendance', department.id,
                           f'تم تسجيل حضور لقسم {department.name} للفترة من {start_date} إلى {end_date} لعدد {len(employees)} موظف و {days_count} يوم ({total_count} سجل)')
            
            flash(f'تم تسجيل الحضور لـ {len(employees)} موظف عن {days_count} يوم بنجاح (إجمالي {total_count} سجل)', 'success')
            return redirect(url_for('attendance.index', date=start_date_str))
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # الحصول على جميع الأقسام
    departments = Department.query.all()
    
    # التاريخ الافتراضي هو اليوم
    today = datetime.now().date()
    hijri_date = format_date_hijri(today)
    gregorian_date = format_date_gregorian(today)
    
    return render_template('attendance/multi_day_department.html', 
                          departments=departments,
                          today=today,
                          hijri_date=hijri_date,
                          gregorian_date=gregorian_date)

@attendance_bp.route('/delete/<int:id>/confirm', methods=['GET'])
def confirm_delete_attendance(id):
    """عرض صفحة تأكيد حذف سجل الحضور"""
    attendance = Attendance.query.get_or_404(id)
    return render_template('attendance/confirm_delete.html', attendance=attendance)

@attendance_bp.route('/delete/<int:id>', methods=['POST'])
def delete_attendance(id):
    """Delete an attendance record"""
    attendance = Attendance.query.get_or_404(id)
    
    try:
        # Get associated employee
        employee = Employee.query.get(attendance.employee_id)
        
        # Delete attendance record
        db.session.delete(attendance)
        
        # Log the action
        entity_name = employee.name if employee else f'ID: {id}'
        SystemAudit.create_audit_record(
            user_id=None,  # يمكن تعديلها لاستخدام current_user.id
            action='delete',
            entity_type='attendance',
            entity_id=id,
            entity_name=entity_name,
            details=f'تم حذف سجل حضور للموظف: {employee.name if employee else "غير معروف"} بتاريخ {attendance.date}'
        )
        db.session.commit()
        
        flash('تم حذف سجل الحضور بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('attendance.index', date=attendance.date))

@attendance_bp.route('/bulk_delete', methods=['POST'])
def bulk_delete_attendance():
    """حذف سجلات حضور متعددة"""
    from flask import jsonify
    
    try:
        data = request.json
        attendance_ids = data.get('attendance_ids', [])
        
        if not attendance_ids:
            return jsonify({
                'success': False,
                'message': 'لا توجد سجلات محددة للحذف'
            }), 400
        
        deleted_count = 0
        errors = []
        
        for attendance_id in attendance_ids:
            try:
                attendance = Attendance.query.get(attendance_id)
                if attendance:
                    employee = Employee.query.get(attendance.employee_id)
                    entity_name = employee.name if employee else f'ID: {attendance_id}'
                    
                    # حذف السجل
                    db.session.delete(attendance)
                    
                    # تسجيل في Audit
                    SystemAudit.create_audit_record(
                        user_id=None,
                        action='delete',
                        entity_type='attendance',
                        entity_id=attendance_id,
                        entity_name=entity_name,
                        details=f'حذف جماعي - تم حذف سجل حضور للموظف: {employee.name if employee else "غير معروف"} بتاريخ {attendance.date}'
                    )
                    
                    deleted_count += 1
                else:
                    errors.append(f'السجل {attendance_id} غير موجود')
                    
            except Exception as e:
                errors.append(f'خطأ في حذف السجل {attendance_id}: {str(e)}')
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'errors': errors
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        }), 500

@attendance_bp.route('/stats')
def stats():
    """Get attendance statistics for a date range"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    department_id = request.args.get('department_id', '')
    
    try:
        start_date = parse_date(start_date_str) if start_date_str else datetime.now().date().replace(day=1)
        end_date = parse_date(end_date_str) if end_date_str else datetime.now().date()
    except ValueError:
        start_date = datetime.now().date().replace(day=1)
        end_date = datetime.now().date()
    
    query = db.session.query(
        Attendance.status,
        func.count(Attendance.id).label('count')
    ).filter(
        Attendance.date >= start_date,
        Attendance.date <= end_date
    )
    
    if department_id and department_id != '':
        query = query.join(Employee).filter(Employee.department_id == department_id)
    
    stats = query.group_by(Attendance.status).all()
    
    # Convert to a dict for easier consumption by charts
    result = {'present': 0, 'absent': 0, 'leave': 0, 'sick': 0}
    for status, count in stats:
        result[status] = count
    
    return jsonify(result)

@attendance_bp.route('/export/excel', methods=['POST', 'GET'])
def export_excel():
    """تصدير بيانات الحضور إلى ملف Excel"""
    try:
        # الحصول على البيانات من النموذج حسب طريقة الطلب
        if request.method == 'POST':
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            department_id = request.form.get('department_id')
        else:  # GET
            start_date_str = request.args.get('start_date')
            end_date_str = request.args.get('end_date')
            department_id = request.args.get('department_id')
        
        # التحقق من المدخلات
        if not start_date_str:
            flash('تاريخ البداية مطلوب', 'danger')
            return redirect(url_for('attendance.export_page'))
        
        # تحليل التواريخ
        try:
            start_date = parse_date(start_date_str)
            if end_date_str:
                end_date = parse_date(end_date_str)
            else:
                end_date = datetime.now().date()
        except (ValueError, TypeError):
            flash('تاريخ غير صالح', 'danger')
            return redirect(url_for('attendance.export_page'))
        
        # التحقق من اختيار القسم
        if department_id and department_id != '':
            # تصدير قسم واحد فقط
            department = Department.query.get(department_id)
            if not department:
                flash('القسم غير موجود', 'danger')
                return redirect(url_for('attendance.export_page'))
            
            # جلب جميع سجلات الحضور للفترة المحددة أولاً
            attendances = Attendance.query.filter(
                Attendance.date.between(start_date, end_date)
            ).all()
            
            # جلب معرفات الموظفين الذين لديهم حضور
            employee_ids_with_attendance = set([att.employee_id for att in attendances])
            
            # جلب الموظفين في القسم (استبعاد المنتهية خدمتهم فقط) + الموظفين المنتهية خدمتهم الذين لديهم حضور
            department_employee_ids = [emp.id for emp in department.employees]
            employees_to_export = Employee.query.filter(
                Employee.id.in_(department_employee_ids),
                or_(
                    ~Employee.status.in_(['terminated', 'inactive']),
                    Employee.id.in_(employee_ids_with_attendance)
                )
            ).all()
            
            # فلترة سجلات الحضور لموظفي هذا القسم فقط
            attendances = [att for att in attendances if att.employee_id in department_employee_ids]
            
            excel_file = export_attendance_by_department(employees_to_export, attendances, start_date, end_date)
            
            if end_date_str:
                filename = f'سجل الحضور - {department.name} - {start_date_str} إلى {end_date_str}.xlsx'
            else:
                filename = f'سجل الحضور - {department.name} - {start_date_str}.xlsx'
        else:
            # تصدير جميع الأقسام
            departments = Department.query.all()
            
            # جلب جميع سجلات الحضور للفترة المحددة أولاً
            attendances = Attendance.query.filter(
                Attendance.date.between(start_date, end_date)
            ).all()
            
            # جلب معرفات الموظفين الذين لديهم حضور
            employee_ids_with_attendance = set([att.employee_id for att in attendances])
            
            # جلب الموظفين (استبعاد المنتهية خدمتهم فقط) + الموظفين المنتهية خدمتهم الذين لديهم حضور في الفترة
            all_employees = Employee.query.filter(
                or_(
                    ~Employee.status.in_(['terminated', 'inactive']),
                    Employee.id.in_(employee_ids_with_attendance)
                )
            ).all()
            
            excel_file = export_attendance_by_department_with_dashboard(all_employees, attendances, start_date, end_date)
            
            if end_date_str:
                filename = f'سجل الحضور - جميع الأقسام - {start_date_str} إلى {end_date_str}.xlsx'
            else:
                filename = f'سجل الحضور - جميع الأقسام - {start_date_str}.xlsx'
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('attendance.export_page'))

@attendance_bp.route('/export')
def export_page():
    """صفحة تصدير بيانات الحضور إلى ملف Excel"""
    departments = Department.query.all()
    today = datetime.now().date()
    start_of_month = today.replace(day=1)
    
    hijri_today = format_date_hijri(today)
    gregorian_today = format_date_gregorian(today)
    
    hijri_start = format_date_hijri(start_of_month)
    gregorian_start = format_date_gregorian(start_of_month)
    
    return render_template('attendance/export.html',
                          departments=departments,
                          today=today,
                          start_of_month=start_of_month,
                          hijri_today=hijri_today,
                          gregorian_today=gregorian_today,
                          hijri_start=hijri_start,
                          gregorian_start=gregorian_start)

@attendance_bp.route('/api/departments/<int:department_id>/employees')
def get_department_employees(department_id):
    """API endpoint to get all employees in a department"""
    try:
        # الحصول على القسم أولاً
        department = Department.query.get_or_404(department_id)
        
        # جلب جميع الموظفين النشطين في هذا القسم باستخدام العلاقة many-to-many
        employees = [emp for emp in department.employees if emp.status == 'active']
        
        employee_data = []
        for employee in employees:
            employee_data.append({
                'id': employee.id,
                'name': employee.name,
                'employee_id': employee.employee_id,
                'job_title': employee.job_title or 'غير محدد',
                'status': employee.status
            })
        
        logger.info(f"تم جلب {len(employee_data)} موظف نشط من القسم {department_id} ({department.name})")
        return jsonify(employee_data)
    
    except Exception as e:
        logger.error(f"خطأ في جلب موظفي القسم {department_id}: {str(e)}")
        return jsonify({'error': str(e)}), 500
        
@attendance_bp.route('/dashboard')
def dashboard():
    """لوحة معلومات الحضور مع إحصائيات يومية وأسبوعية وشهرية"""
    
    # إضافة آلية إعادة المحاولة للتعامل مع أخطاء الاتصال المؤقتة
    max_retries = 3  # عدد محاولات إعادة الاتصال
    retry_count = 0
    retry_delay = 1  # ثانية واحدة للمحاولة الأولى
    
    while retry_count < max_retries:
        try:
            # إعادة تحميل جميع البيانات من قاعدة البيانات للتأكد من أن البيانات محدثة
            db.session.expire_all()
            
            # 1. الحصول على المشروع المحدد (إذا وجد)
            project_name = request.args.get('project', None)
            
            # 2. الحصول على التاريخ الحالي
            today = datetime.now().date()
            current_month = today.month
            current_year = today.year
            
            # 3. حساب تاريخ بداية ونهاية الأسبوع الحالي بناءً على تاريخ بداية الشهر
            start_of_month = today.replace(day=1)  # أول يوم في الشهر الحالي
            
            # نحسب عدد الأيام منذ بداية الشهر حتى اليوم الحالي
            days_since_month_start = (today - start_of_month).days
            
            # نحسب عدد الأسابيع الكاملة منذ بداية الشهر (كل أسبوع 7 أيام)
            weeks_since_month_start = days_since_month_start // 7
            
            # نحسب بداية الأسبوع الحالي (بناءً على أسابيع من بداية الشهر)
            start_of_week = start_of_month + timedelta(days=weeks_since_month_start * 7)
            
            # نهاية الأسبوع بعد 6 أيام من البداية
            end_of_week = start_of_week + timedelta(days=6)
            
            # إذا كانت نهاية الأسبوع بعد نهاية الشهر، نجعلها آخر يوم في الشهر
            last_day = calendar.monthrange(current_year, current_month)[1]
            end_of_month = today.replace(day=last_day)
            if end_of_week > end_of_month:
                end_of_week = end_of_month
            
            # 4. حساب تاريخ بداية ونهاية الشهر الحالي
            start_of_month = today.replace(day=1)
            last_day = calendar.monthrange(current_year, current_month)[1]
            end_of_month = today.replace(day=last_day)
            
            # 5. إنشاء قاعدة الاستعلام
            query_base = db.session.query(
                Attendance.status,
                func.count(Attendance.id).label('count')
            )
            
            # 6. إحصائيات الحضور حسب المشروع أو عامة
            # تعريف قائمة معرفات الموظفين (سيكون None للكل)
            employee_ids = None
            
            if project_name:
                # استعلام للموظفين في مشروع محدد
                # نحتاج للحصول على قائمة الموظفين المرتبطين بالمشروع
                project_employees = db.session.query(Employee.id).filter(
                    Employee.project == project_name,
                    ~Employee.status.in_(['terminated', 'inactive'])
                ).all()
                
                # تحويل النتائج إلى قائمة بسيطة من المعرفات
                employee_ids = [emp[0] for emp in project_employees]
            
            # تعريف ونهيئة متغيرات إحصائية
            daily_stats = []
            weekly_stats = []
            monthly_stats = []
            
            # إذا كان هناك مشروع محدد ولا يوجد موظفين فيه، نترك الإحصائيات فارغة
            if project_name and not employee_ids:
                # لا يوجد موظفين في هذا المشروع، نترك الإحصائيات فارغة
                pass
            else:
                # بناء استعلامات الإحصائيات إما لجميع الموظفين أو لموظفي مشروع محدد
                if employee_ids:
                    # إحصائيات الموظفين في المشروع المحدد
                    
                    # إحصائيات اليوم
                    daily_stats = query_base.filter(
                        Attendance.date == today,
                        Attendance.employee_id.in_(employee_ids)
                    ).group_by(Attendance.status).all()
                    
                    # إحصائيات الأسبوع
                    weekly_stats = query_base.filter(
                        Attendance.date >= start_of_week,
                        Attendance.date <= end_of_week,
                        Attendance.employee_id.in_(employee_ids)
                    ).group_by(Attendance.status).all()
                    
                    # إحصائيات الشهر
                    monthly_stats = query_base.filter(
                        Attendance.date >= start_of_month,
                        Attendance.date <= end_of_month,
                        Attendance.employee_id.in_(employee_ids)
                    ).group_by(Attendance.status).all()
                else:
                    # إحصائيات عامة لجميع الموظفين
                    
                    # إحصائيات اليوم
                    daily_stats = query_base.filter(
                        Attendance.date == today
                    ).group_by(Attendance.status).all()
                    
                    # إحصائيات الأسبوع
                    weekly_stats = query_base.filter(
                        Attendance.date >= start_of_week,
                        Attendance.date <= end_of_week
                    ).group_by(Attendance.status).all()
                    
                    # إحصائيات الشهر
                    monthly_stats = query_base.filter(
                        Attendance.date >= start_of_month,
                        Attendance.date <= end_of_month
                    ).group_by(Attendance.status).all()
            
            # 7. إحصائيات الحضور اليومي خلال الشهر الحالي لعرضها في المخطط البياني
            daily_attendance_data = []
            
            for day in range(1, last_day + 1):
                current_date = date(current_year, current_month, day)
                
                # تخطي التواريخ المستقبلية
                if current_date > today:
                    break
                    
                # استخدام employee_ids مباشرة بعد التأكد من أنه تم تعريفه في خطوة سابقة
                if employee_ids:
                    present_count = db.session.query(func.count(Attendance.id)).filter(
                        Attendance.date == current_date,
                        Attendance.status == 'present',
                        Attendance.employee_id.in_(employee_ids)
                    ).scalar() or 0
                    
                    absent_count = db.session.query(func.count(Attendance.id)).filter(
                        Attendance.date == current_date,
                        Attendance.status == 'absent',
                        Attendance.employee_id.in_(employee_ids)
                    ).scalar() or 0
                else:
                    present_count = db.session.query(func.count(Attendance.id)).filter(
                        Attendance.date == current_date,
                        Attendance.status == 'present'
                    ).scalar() or 0
                    
                    absent_count = db.session.query(func.count(Attendance.id)).filter(
                        Attendance.date == current_date,
                        Attendance.status == 'absent'
                    ).scalar() or 0
                    
                daily_attendance_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day': str(day),
                    'present': present_count,
                    'absent': absent_count
                })
                
            # 8. الحصول على قائمة المشاريع للفلتر (استبعاد المنتهية خدمتهم فقط)
            active_projects = db.session.query(Employee.project).filter(
                ~Employee.status.in_(['terminated', 'inactive']),
                Employee.project.isnot(None)
            ).distinct().all()
            
            active_projects = [project[0] for project in active_projects if project[0]]
            
            # 9. تحويل البيانات إلى قاموس
            def stats_to_dict(stats_data):
                result = {'present': 0, 'absent': 0, 'leave': 0, 'sick': 0}
                for item in stats_data:
                    result[item[0]] = item[1]
                return result
            
            daily_stats_dict = stats_to_dict(daily_stats)
            weekly_stats_dict = stats_to_dict(weekly_stats)
            monthly_stats_dict = stats_to_dict(monthly_stats)
            
            # 10. إعداد البيانات للمخططات البيانية
            # 10.أ. مخطط توزيع الحضور اليومي
            daily_chart_data = {
                'labels': ['حاضر', 'غائب', 'إجازة', 'مرضي'],
                'datasets': [{
                    'data': [
                        daily_stats_dict['present'],
                        daily_stats_dict['absent'],
                        daily_stats_dict['leave'],
                        daily_stats_dict['sick']
                    ],
                    'backgroundColor': ['#28a745', '#dc3545', '#ffc107', '#17a2b8']
                }]
            }
            
            # 10.ب. مخطط توزيع الحضور الأسبوعي
            weekly_chart_data = {
                'labels': ['حاضر', 'غائب', 'إجازة', 'مرضي'],
                'datasets': [{
                    'data': [
                        weekly_stats_dict['present'],
                        weekly_stats_dict['absent'],
                        weekly_stats_dict['leave'],
                        weekly_stats_dict['sick']
                    ],
                    'backgroundColor': ['#28a745', '#dc3545', '#ffc107', '#17a2b8']
                }]
            }
            
            # 10.ج. مخطط توزيع الحضور الشهري
            monthly_chart_data = {
                'labels': ['حاضر', 'غائب', 'إجازة', 'مرضي'],
                'datasets': [{
                    'data': [
                        monthly_stats_dict['present'],
                        monthly_stats_dict['absent'],
                        monthly_stats_dict['leave'],
                        monthly_stats_dict['sick']
                    ],
                    'backgroundColor': ['#28a745', '#dc3545', '#ffc107', '#17a2b8']
                }]
            }
            
            # 10.د. مخطط الحضور اليومي خلال الشهر
            daily_trend_chart_data = {
                'labels': [item['day'] for item in daily_attendance_data],
                'datasets': [
                    {
                        'label': 'الحضور',
                        'data': [item['present'] for item in daily_attendance_data],
                        'backgroundColor': 'rgba(40, 167, 69, 0.2)',
                        'borderColor': 'rgba(40, 167, 69, 1)',
                        'borderWidth': 1,
                        'tension': 0.4
                    },
                    {
                        'label': 'الغياب',
                        'data': [item['absent'] for item in daily_attendance_data],
                        'backgroundColor': 'rgba(220, 53, 69, 0.2)',
                        'borderColor': 'rgba(220, 53, 69, 1)',
                        'borderWidth': 1,
                        'tension': 0.4
                    }
                ]
            }
            
            # 11. حساب معدل الحضور
            # إجمالي سجلات الحضور اليومية
            total_days = (
                daily_stats_dict['present'] + 
                daily_stats_dict['absent'] + 
                daily_stats_dict['leave'] + 
                daily_stats_dict['sick']
            )
            
            # إجمالي سجلات الحضور المتوقعة لليوم (جميع الموظفين النشطين)
            # حساب إجمالي الموظفين النشطين يتم في سطور لاحقة من الكود
            
            daily_attendance_rate = 0
            if total_days > 0:
                daily_attendance_rate = round((daily_stats_dict['present'] / total_days) * 100)
            
            # حساب إجمالي الموظفين - استخدام علاقة many-to-many الصحيحة (استبعاد المنتهية خدمتهم فقط)
            if employee_ids:
                active_employees_count = len(employee_ids)
            else:
                # عد الموظفين المرتبطين بأقسام عبر علاقة many-to-many
                active_employees_count = db.session.query(func.count(func.distinct(Employee.id))).join(
                    employee_departments, Employee.id == employee_departments.c.employee_id
                ).filter(
                    ~Employee.status.in_(['terminated', 'inactive'])
                ).scalar() or 0
            
            # حساب كامل الأسبوع (7 أيام) × عدد الموظفين النشطين
            # حساب عدد الأيام في الأسبوع (من بداية الأسبوع إلى نهايته)
            days_in_week = (end_of_week - start_of_week).days + 1
            
            # إجمالي سجلات الحضور والغياب في الأسبوع
            total_days_week = (
                weekly_stats_dict['present'] + 
                weekly_stats_dict['absent'] + 
                weekly_stats_dict['leave'] + 
                weekly_stats_dict['sick']
            )
            
            # حساب إجمالي سجلات الحضور المفترضة للأسبوع
            expected_days_week = days_in_week * active_employees_count
            
            weekly_attendance_rate = 0
            if total_days_week > 0:
                weekly_attendance_rate = round((weekly_stats_dict['present'] / total_days_week) * 100)
            
            # حساب معدل الحضور الشهري
            # إجمالي سجلات الحضور والغياب في الشهر
            total_days_month = (
                monthly_stats_dict['present'] + 
                monthly_stats_dict['absent'] + 
                monthly_stats_dict['leave'] + 
                monthly_stats_dict['sick']
            )
            
            # حساب عدد الأيام في الشهر حتى اليوم الحالي
            days_in_month = (today - start_of_month).days + 1
            
            # حساب إجمالي سجلات الحضور المفترضة للشهر
            expected_days_month = days_in_month * active_employees_count
            
            monthly_attendance_rate = 0
            if total_days_month > 0:
                monthly_attendance_rate = round((monthly_stats_dict['present'] / total_days_month) * 100)
            
            # 12. تنسيق التواريخ للعرض
            formatted_today = {
                'gregorian': format_date_gregorian(today),
                'hijri': format_date_hijri(today)
            }
            
            formatted_start_of_week = {
                'gregorian': format_date_gregorian(start_of_week),
                'hijri': format_date_hijri(start_of_week)
            }
            
            formatted_end_of_week = {
                'gregorian': format_date_gregorian(end_of_week),
                'hijri': format_date_hijri(end_of_week)
            }
            
            formatted_start_of_month = {
                'gregorian': format_date_gregorian(start_of_month),
                'hijri': format_date_hijri(start_of_month)
            }
            
            formatted_end_of_month = {
                'gregorian': format_date_gregorian(end_of_month),
                'hijri': format_date_hijri(end_of_month)
            }
            
            # 13. إعداد اسم الشهر الحالي
            month_names = [
                'يناير', 'فبراير', 'مارس', 'إبريل', 'مايو', 'يونيو',
                'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'
            ]
            current_month_name = month_names[current_month - 1]
            
            # 14. جلب بيانات الغياب التفصيلية بأسماء الموظفين لكل قسم
            daily_summary = AttendanceAnalytics.get_department_summary(
                start_date=today,
                end_date=today,
                project_name=project_name
            )
            
            monthly_summary = AttendanceAnalytics.get_department_summary(
                start_date=start_of_month,
                end_date=end_of_month,
                project_name=project_name
            )
            
            # 15. إعداد البيانات للعرض على الصفحة
            return render_template('attendance/dashboard_new.html',
                                today=today,
                                current_month=current_month,
                                current_year=current_year,
                                current_month_name=current_month_name,
                                formatted_today=formatted_today,
                                formatted_start_of_week=formatted_start_of_week,
                                formatted_end_of_week=formatted_end_of_week,
                                formatted_start_of_month=formatted_start_of_month,
                                formatted_end_of_month=formatted_end_of_month,
                                start_of_week=start_of_week,
                                end_of_week=end_of_week,
                                start_of_month=start_of_month,
                                end_of_month=end_of_month,
                                daily_stats=daily_stats_dict,
                                weekly_stats=weekly_stats_dict,
                                monthly_stats=monthly_stats_dict,
                                daily_chart_data=daily_chart_data,
                                weekly_chart_data=weekly_chart_data,
                                monthly_chart_data=monthly_chart_data,
                                daily_trend_chart_data=daily_trend_chart_data,
                                daily_attendance_rate=daily_attendance_rate,
                                weekly_attendance_rate=weekly_attendance_rate,
                                monthly_attendance_rate=monthly_attendance_rate,
                                active_employees_count=active_employees_count,
                                active_projects=active_projects,
                                selected_project=project_name,
                                daily_summary=daily_summary,
                                monthly_summary=monthly_summary)
                                
            # Si todo funciona bien, sal del bucle
            break
            
        except Exception as e:
            # Si hay un error, incrementa el contador y espera
            retry_count += 1
            logger.error(f"Error al cargar el dashboard (intento {retry_count}): {str(e)}")
            
            if retry_count < max_retries:
                # Espera un tiempo exponencial antes de reintentar
                time_module.sleep(retry_delay)
                retry_delay *= 2  # Duplica el tiempo de espera para el próximo intento
            else:
                # Si se han agotado los reintentos, muestra un mensaje de error
                logger.critical(f"Error al cargar el dashboard después de {max_retries} intentos: {str(e)}")
                return render_template('error.html', 
                                      error_title="خطأ في الاتصال",
                                      error_message="حدث خطأ أثناء الاتصال بقاعدة البيانات. الرجاء المحاولة مرة أخرى.",
                                      error_details=str(e))

@attendance_bp.route('/employee/<int:employee_id>')
def employee_attendance(employee_id):
    """عرض سجلات الحضور التفصيلية للموظف مرتبة حسب الشهر والسنة - Dashboard مميز"""
    # الحصول على الموظف
    employee = Employee.query.get_or_404(employee_id)
    
    # الحصول على التاريخ الحالي
    today = datetime.now().date()
    
    # الحصول على السنة والشهر من URL أو استخدام الحالي
    selected_year = request.args.get('year', today.year, type=int)
    selected_month = request.args.get('month', today.month, type=int)
    
    # تحديد فترة الاستعلام (الشهر المختار)
    year = selected_year
    month = selected_month
    start_of_month = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_of_month = date(year, month, last_day)
    
    # الحصول على سجلات الحضور للشهر المختار
    attendances = Attendance.query.filter(
        Attendance.employee_id == employee_id,
        Attendance.date >= start_of_month,
        Attendance.date <= end_of_month
    ).order_by(Attendance.date).all()
    
    # تنظيم السجلات حسب اليوم للتقويم
    attendance_by_day = {}
    for record in attendances:
        attendance_by_day[record.date.day] = record
    
    # حساب الإحصائيات الشاملة
    present_count = sum(1 for a in attendances if a.status == 'present')
    absent_count = sum(1 for a in attendances if a.status == 'absent')
    leave_count = sum(1 for a in attendances if a.status == 'leave')
    sick_count = sum(1 for a in attendances if a.status == 'sick')
    total_records = len(attendances)
    
    # حساب النسب المئوية
    present_percentage = (present_count / total_records * 100) if total_records > 0 else 0
    absent_percentage = (absent_count / total_records * 100) if total_records > 0 else 0
    leave_percentage = (leave_count / total_records * 100) if total_records > 0 else 0
    sick_percentage = (sick_count / total_records * 100) if total_records > 0 else 0
    
    # حساب معدل الحضور
    attendance_rate = round(present_percentage, 1) if total_records > 0 else 0
    
    # جمع كل الفترات المتاحة (السنوات والأشهر التي لديها سجلات)
    all_records = Attendance.query.filter(Attendance.employee_id == employee_id).all()
    attendance_periods = {}
    for record in all_records:
        if record.date.year not in attendance_periods:
            attendance_periods[record.date.year] = set()
        attendance_periods[record.date.year].add(record.date.month)
    
    # معلومات التقويم
    first_day_weekday = calendar.monthrange(year, month)[0]
    days_in_month = last_day
    
    # تنسيق التواريخ للعرض
    hijri_today = format_date_hijri(today)
    gregorian_today = format_date_gregorian(today)
    
    return render_template('attendance/employee_attendance.html',
                          employee=employee,
                          attendances=attendances,
                          attendance_by_day=attendance_by_day,
                          year=year,
                          month=month,
                          selected_year=selected_year,
                          selected_month=selected_month,
                          first_day_weekday=first_day_weekday,
                          days_in_month=days_in_month,
                          attendance_periods=attendance_periods,
                          present_count=present_count,
                          absent_count=absent_count,
                          leave_count=leave_count,
                          sick_count=sick_count,
                          total_records=total_records,
                          present_percentage=present_percentage,
                          absent_percentage=absent_percentage,
                          leave_percentage=leave_percentage,
                          sick_percentage=sick_percentage,
                          attendance_rate=attendance_rate,
                          today=today,
                          hijri_today=hijri_today,
                          gregorian_today=gregorian_today)

@attendance_bp.route('/department-stats')
def department_stats():
    """API لجلب إحصائيات الحضور حسب الأقسام"""
    period = request.args.get('period', 'monthly')  # weekly أو monthly
    project_name = request.args.get('project', None)
    
    today = datetime.now().date()
    
    # تحديد الفترة الزمنية - استخدام البيانات الشهرية الحقيقية
    start_date = today.replace(day=1)  # بداية الشهر الحالي
    end_date = today  # حتى اليوم الحالي
    
    # جلب الأقسام المسموح بالوصول إليها حسب صلاحيات المستخدم
    from flask_login import current_user
    
    if current_user.is_authenticated:
        # إذا كان المستخدم مسجل دخوله، عرض الأقسام المسموحة فقط
        departments = current_user.get_accessible_departments()
    else:
        # إذا لم يكن مسجل دخوله، عرض جميع الأقسام (للعرض العام)
        departments = Department.query.all()
    
    department_stats = []
    
    for dept in departments:
        # جلب الموظفين في القسم - استخدام علاقة many-to-many (استبعاد المنتهية خدمتهم فقط)
        employees = [emp for emp in dept.employees if emp.status not in ['terminated', 'inactive']]
        
        # فلترة حسب المشروع إذا تم تحديده
        if project_name:
            employees = [emp for emp in employees if emp.project == project_name]
        
        total_employees = len(employees)
        
        # عرض جميع الأقسام حتى لو كانت فارغة لضمان الشمولية
        # if total_employees == 0:
        #     continue
        
        # حساب الإحصائيات
        employee_ids = [emp.id for emp in employees]
        
        # جلب سجلات الحضور للفترة المحددة
        attendance_records = []
        if employee_ids:
            attendance_records = Attendance.query.filter(
                Attendance.employee_id.in_(employee_ids),
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).all()
        
        # حساب الإحصائيات
        present_count = sum(1 for record in attendance_records if record.status == 'present')
        absent_count = sum(1 for record in attendance_records if record.status == 'absent')
        leave_count = sum(1 for record in attendance_records if record.status == 'leave')
        sick_count = sum(1 for record in attendance_records if record.status == 'sick')
        total_records = len(attendance_records)
        
        # حساب الأيام والسجلات المتوقعة
        working_days = (end_date - start_date).days + 1
        expected_total_records = total_employees * working_days
        
        # للفترة الشهرية، نحسب أيام العمل الفعلية (عدا الجمع والسبوت)
        if period == 'monthly':
            working_days_actual = 0
            current = start_date
            while current <= end_date:
                # حساب أيام العمل (الأحد-الخميس في النظام السعودي)
                if current.weekday() < 5:  # 0-4 (الاثنين-الجمعة) نحسبها أيام عمل
                    working_days_actual += 1
                current += timedelta(days=1)
            working_days = working_days_actual
        
        # حساب معدل الحضور بناء على السجلات الفعلية الموجودة
        if total_records > 0:
            attendance_rate = (present_count / total_records) * 100
        else:
            attendance_rate = 0
        
        department_stats.append({
            'id': dept.id,
            'name': dept.name,
            'total_employees': total_employees,
            'present': present_count,
            'absent': absent_count,
            'leave': leave_count,
            'sick': sick_count,
            'attendance_rate': round(attendance_rate, 1),
            'total_records': total_records,
            'working_days': working_days,
            'expected_records': expected_total_records
        })
    
    # ترتيب الأقسام حسب معدل الحضور (تنازلي)
    department_stats.sort(key=lambda x: x['attendance_rate'], reverse=True)
    
    return jsonify({
        'departments': department_stats,
        'period': period,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'project': project_name
    })

@attendance_bp.route('/export-excel-dashboard')
def export_excel_dashboard():
    """تصدير لوحة المعلومات إلى Excel مع تصميم داش بورد خيالي ومبهر + تفاصيل الغياب"""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file
        
        # الحصول على المعاملات
        selected_department = request.args.get('department', None)
        selected_project = request.args.get('project', None)
        
        today = datetime.now().date()
        start_date = today.replace(day=1)
        end_date = today
        
        # استخدام خدمة analytics لجلب البيانات التفصيلية
        summary = AttendanceAnalytics.get_department_summary(
            start_date=start_date,
            end_date=end_date,
            project_name=selected_project
        )
        
        # تحويل البيانات للصيغة المطلوبة
        department_data = []
        for dept in summary['departments']:
            if selected_department and dept['name'] != selected_department:
                continue
            
            department_data.append({
                'name': dept['name'],
                'employees': dept['total_employees'],
                'present': dept['present'],
                'absent': dept['absent'],
                'leave': dept['leave'],
                'sick': dept['sick'],
                'total': dept['total_records'],
                'rate': dept['attendance_rate'],
                'absentees': dept['absentees'],  # بيانات الغياب التفصيلية
                'on_leave': dept['on_leave'],
                'sick_employees': dept['sick_employees']
            })
        
        # الإجماليات
        total_employees = summary['total_employees']
        total_present = summary['total_present']
        total_absent = summary['total_absent']
        total_leave = summary['total_leave']
        total_sick = summary['total_sick']
        total_records = summary['total_records']
        
        if not department_data:
            wb = Workbook()
            ws = wb.active
            ws.title = "لا توجد بيانات"
            ws['A1'] = "لا توجد بيانات للعرض"
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name=f'تقرير_الحضور_{today.strftime("%Y%m%d")}.xlsx')
        
        # إنشاء ملف Excel خيالي
        wb = Workbook()
        ws = wb.active
        ws.title = "📊 لوحة المعلومات"
        
        # === العنوان الرئيسي الفاخر ===
        ws.merge_cells('A1:M3')
        title_cell = ws['A1']
        title_cell.value = f"📊 لوحة معلومات الحضور والغياب\n{start_date.strftime('%Y/%m/%d')} - {end_date.strftime('%Y/%m/%d')}"
        title_cell.font = Font(size=24, bold=True, color="FFFFFF", name="Arial")
        title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30
        
        # === بطاقات KPI الملونة ===
        kpi_row = 5
        ws.row_dimensions[kpi_row].height = 35
        ws.row_dimensions[kpi_row + 1].height = 30
        
        # تعريف الحدود
        thick_border = Border(
            left=Side(style='thick', color='FFFFFF'),
            right=Side(style='thick', color='FFFFFF'),
            top=Side(style='thick', color='FFFFFF'),
            bottom=Side(style='thick', color='FFFFFF')
        )
        
        # KPI 1: إجمالي الموظفين
        ws.merge_cells(f'A{kpi_row}:C{kpi_row+1}')
        kpi1 = ws[f'A{kpi_row}']
        kpi1.value = f"👥 إجمالي الموظفين\n{total_employees}"
        kpi1.font = Font(size=16, bold=True, color="FFFFFF")
        kpi1.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        kpi1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        kpi1.border = thick_border
        
        # KPI 2: الحضور
        ws.merge_cells(f'D{kpi_row}:F{kpi_row+1}')
        kpi2 = ws[f'D{kpi_row}']
        kpi2.value = f"✅ الحضور\n{total_present}"
        kpi2.font = Font(size=16, bold=True, color="FFFFFF")
        kpi2.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        kpi2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        kpi2.border = thick_border
        
        # KPI 3: الغياب
        ws.merge_cells(f'G{kpi_row}:I{kpi_row+1}')
        kpi3 = ws[f'G{kpi_row}']
        kpi3.value = f"❌ الغياب\n{total_absent}"
        kpi3.font = Font(size=16, bold=True, color="FFFFFF")
        kpi3.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        kpi3.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        kpi3.border = thick_border
        
        # KPI 4: الإجازات
        ws.merge_cells(f'J{kpi_row}:L{kpi_row+1}')
        kpi4 = ws[f'J{kpi_row}']
        kpi4.value = f"🏖️ الإجازات\n{total_leave}"
        kpi4.font = Font(size=16, bold=True, color="FFFFFF")
        kpi4.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        kpi4.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        kpi4.border = thick_border
        
        # KPI 5: المرضي
        ws.merge_cells(f'M{kpi_row}:O{kpi_row+1}')
        kpi5 = ws[f'M{kpi_row}']
        kpi5.value = f"🏥 المرضي\n{total_sick}"
        kpi5.font = Font(size=16, bold=True, color="FFFFFF")
        kpi5.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        kpi5.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        kpi5.border = thick_border
        
        # === جدول البيانات التفصيلي ===
        table_start_row = kpi_row + 3
        
        # عنوان الجدول
        ws.merge_cells(f'A{table_start_row}:H{table_start_row}')
        table_title = ws[f'A{table_start_row}']
        table_title.value = "📋 تفاصيل الأقسام"
        table_title.font = Font(size=14, bold=True, color="FFFFFF")
        table_title.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        table_title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[table_start_row].height = 25
        
        # رؤوس الأعمدة
        headers_row = table_start_row + 1
        headers = ['القسم', 'عدد الموظفين', 'حاضر', 'غائب', 'إجازة', 'مرضي', 'إجمالي السجلات', 'معدل الحضور %']
        header_colors = ['34495E', '5B9BD5', '70AD47', 'E74C3C', 'F39C12', '3498DB', '95A5A6', '16A085']
        
        for col_idx, (header, color) in enumerate(zip(headers, header_colors), 1):
            cell = ws.cell(row=headers_row, column=col_idx)
            cell.value = header
            cell.font = Font(size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        ws.row_dimensions[headers_row].height = 30
        
        # البيانات
        data_start_row = headers_row + 1
        for row_idx, dept in enumerate(department_data, data_start_row):
            values = [dept['name'], dept['employees'], dept['present'], dept['absent'], 
                     dept['leave'], dept['sick'], dept['total'], dept['rate']]
            
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # تلوين بديل للصفوف
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                
                # تلوين خاص للأعمدة
                if col_idx == 3:  # حاضر
                    cell.fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                    cell.font = Font(bold=True, color="27AE60")
                elif col_idx == 4:  # غائب
                    cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    cell.font = Font(bold=True, color="C0392B")
                elif col_idx == 5:  # إجازة
                    cell.fill = PatternFill(start_color="FEF5E7", end_color="FEF5E7", fill_type="solid")
                    cell.font = Font(bold=True, color="D68910")
                elif col_idx == 6:  # مرضي
                    cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                    cell.font = Font(bold=True, color="2874A6")
                elif col_idx == 8:  # معدل الحضور
                    cell.font = Font(bold=True, size=11)
                    if value >= 90:
                        cell.fill = PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type="solid")
                    elif value >= 75:
                        cell.fill = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
        
        # === الرسوم البيانية ===
        chart_row = data_start_row + len(department_data) + 2
        
        # 1. مخطط عمودي - مقارنة الحضور بين الأقسام
        bar_chart = BarChart()
        bar_chart.title = "مقارنة حالات الحضور بين الأقسام"
        bar_chart.style = 13
        bar_chart.y_axis.title = 'العدد'
        bar_chart.x_axis.title = 'الأقسام'
        bar_chart.height = 12
        bar_chart.width = 20
        
        data_ref = Reference(ws, min_col=3, min_row=headers_row, max_row=data_start_row + len(department_data) - 1, max_col=6)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_start_row + len(department_data) - 1)
        
        bar_chart.add_data(data_ref, titles_from_data=True)
        bar_chart.set_categories(cats_ref)
        ws.add_chart(bar_chart, f"A{chart_row}")
        
        # 2. مخطط دائري - نسب الحضور الإجمالية
        pie_chart = PieChart()
        pie_chart.title = "نسب حالات الحضور الإجمالية"
        pie_chart.height = 12
        pie_chart.width = 15
        
        # إنشاء بيانات الإجمالي
        summary_row = data_start_row + len(department_data) + 15
        ws[f'K{summary_row}'] = 'حاضر'
        ws[f'L{summary_row}'] = total_present
        ws[f'K{summary_row+1}'] = 'غائب'
        ws[f'L{summary_row+1}'] = total_absent
        ws[f'K{summary_row+2}'] = 'إجازة'
        ws[f'L{summary_row+2}'] = total_leave
        ws[f'K{summary_row+3}'] = 'مرضي'
        ws[f'L{summary_row+3}'] = total_sick
        
        pie_data = Reference(ws, min_col=12, min_row=summary_row, max_row=summary_row+3)
        pie_labels = Reference(ws, min_col=11, min_row=summary_row, max_row=summary_row+3)
        pie_chart.add_data(pie_data, titles_from_data=False)
        pie_chart.set_categories(pie_labels)
        ws.add_chart(pie_chart, f"K{chart_row}")
        
        # تعيين عرض الأعمدة
        column_widths = [20, 15, 12, 12, 12, 12, 18, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # ===== Sheet جديد: قائمة الغياب التفصيلية =====
        if total_absent > 0:
            ws_absence = wb.create_sheet(title="📋 قائمة الغياب")
            
            # العنوان الرئيسي
            ws_absence.merge_cells('A1:F3')
            title_cell = ws_absence['A1']
            title_cell.value = f"📋 قائمة الغياب التفصيلية\n{start_date.strftime('%Y/%m/%d')} - {end_date.strftime('%Y/%m/%d')}"
            title_cell.font = Font(size=20, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws_absence.row_dimensions[1].height = 25
            ws_absence.row_dimensions[2].height = 25
            
            current_row = 5
            
            # إضافة بيانات الغياب لكل قسم
            for dept in department_data:
                if dept['absentees']:
                    # عنوان القسم
                    ws_absence.merge_cells(f'A{current_row}:F{current_row}')
                    dept_header = ws_absence[f'A{current_row}']
                    dept_header.value = f"🏢 {dept['name']} - عدد الغائبين: {len(dept['absentees'])}"
                    dept_header.font = Font(size=14, bold=True, color="FFFFFF")
                    dept_header.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                    dept_header.alignment = Alignment(horizontal='center', vertical='center')
                    ws_absence.row_dimensions[current_row].height = 25
                    current_row += 1
                    
                    # رؤوس الأعمدة
                    headers = ['#', 'اسم الموظف', 'رقم الموظف', 'التاريخ', 'ملاحظات', 'الحالة']
                    header_colors = ['95A5A6', '3498DB', '9B59B6', 'E67E22', 'E74C3C', 'E74C3C']
                    for col_idx, (header, color) in enumerate(zip(headers, header_colors), 1):
                        cell = ws_absence.cell(row=current_row, column=col_idx)
                        cell.value = header
                        cell.font = Font(size=11, bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                    ws_absence.row_dimensions[current_row].height = 25
                    current_row += 1
                    
                    # بيانات الموظفين الغائبين
                    for idx, emp in enumerate(dept['absentees'], 1):
                        row_data = [
                            idx,
                            emp['name'],
                            emp.get('employee_id', '-'),
                            emp['date'].strftime('%Y-%m-%d') if emp.get('date') else '-',
                            emp.get('notes', '-'),
                            'غائب'
                        ]
                        
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws_absence.cell(row=current_row, column=col_idx)
                            cell.value = value
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = Border(
                                left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin')
                            )
                            
                            # تلوين بديل للصفوف
                            if current_row % 2 == 0:
                                cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                        
                        current_row += 1
                    
                    # مسافة بين الأقسام
                    current_row += 1
            
            # تعيين عرض الأعمدة
            ws_absence.column_dimensions['A'].width = 8
            ws_absence.column_dimensions['B'].width = 30
            ws_absence.column_dimensions['C'].width = 15
            ws_absence.column_dimensions['D'].width = 15
            ws_absence.column_dimensions['E'].width = 35
            ws_absence.column_dimensions['F'].width = 12
        
        # ===== Sheet جديد: قائمة الإجازات =====
        if total_leave > 0:
            ws_leave = wb.create_sheet(title="🏖️ قائمة الإجازات")
            
            # العنوان الرئيسي
            ws_leave.merge_cells('A1:F3')
            title_cell = ws_leave['A1']
            title_cell.value = f"🏖️ قائمة الإجازات التفصيلية\n{start_date.strftime('%Y/%m/%d')} - {end_date.strftime('%Y/%m/%d')}"
            title_cell.font = Font(size=20, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws_leave.row_dimensions[1].height = 25
            ws_leave.row_dimensions[2].height = 25
            
            current_row = 5
            
            # إضافة بيانات الإجازات لكل قسم
            for dept in department_data:
                if dept.get('on_leave'):
                    # عنوان القسم
                    ws_leave.merge_cells(f'A{current_row}:F{current_row}')
                    dept_header = ws_leave[f'A{current_row}']
                    dept_header.value = f"🏢 {dept['name']} - عدد الإجازات: {len(dept['on_leave'])}"
                    dept_header.font = Font(size=14, bold=True, color="FFFFFF")
                    dept_header.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                    dept_header.alignment = Alignment(horizontal='center', vertical='center')
                    ws_leave.row_dimensions[current_row].height = 25
                    current_row += 1
                    
                    # رؤوس الأعمدة
                    headers = ['#', 'اسم الموظف', 'رقم الموظف', 'التاريخ', 'ملاحظات', 'الحالة']
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws_leave.cell(row=current_row, column=col_idx)
                        cell.value = header
                        cell.font = Font(size=11, bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                    ws_leave.row_dimensions[current_row].height = 25
                    current_row += 1
                    
                    # بيانات الموظفين
                    for idx, emp in enumerate(dept['on_leave'], 1):
                        row_data = [
                            idx,
                            emp['name'],
                            emp.get('employee_id', '-'),
                            emp['date'].strftime('%Y-%m-%d') if emp.get('date') else '-',
                            emp.get('notes', '-'),
                            'إجازة'
                        ]
                        
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws_leave.cell(row=current_row, column=col_idx)
                            cell.value = value
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = Border(
                                left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin')
                            )
                            
                            if current_row % 2 == 0:
                                cell.fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
                        
                        current_row += 1
                    
                    current_row += 1
            
            # تعيين عرض الأعمدة
            ws_leave.column_dimensions['A'].width = 8
            ws_leave.column_dimensions['B'].width = 30
            ws_leave.column_dimensions['C'].width = 15
            ws_leave.column_dimensions['D'].width = 15
            ws_leave.column_dimensions['E'].width = 35
            ws_leave.column_dimensions['F'].width = 12
        
        # حفظ الملف
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'📊_لوحة_معلومات_الحضور_{today.strftime("%Y%m%d")}.xlsx'
        if selected_department:
            filename = f'📊_{selected_department}_{today.strftime("%Y%m%d")}.xlsx'
            
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"خطأ في تصدير Excel: {str(e)}")
        flash('حدث خطأ أثناء تصدير الملف', 'error')
        return redirect(url_for('attendance.dashboard'))

@attendance_bp.route('/department-details')
def department_details():
    """صفحة تفاصيل الحضور لقسم معين"""
    department_name = request.args.get('department')
    period = request.args.get('period', 'weekly')
    project_name = request.args.get('project', None)
    
    if not department_name:
        flash('يجب تحديد القسم', 'error')
        return redirect(url_for('attendance.dashboard'))
    
    # جلب القسم
    department = Department.query.filter_by(name=department_name).first()
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('attendance.dashboard'))
    
    today = datetime.now().date()
    
    # تحديد الفترة الزمنية - دائماً عرض الشهر الكامل للتفاصيل
    start_date = today.replace(day=1)  # بداية الشهر الحالي
    end_date = today  # حتى اليوم الحالي
    
    # إنشاء قائمة بجميع أيام الشهر حتى اليوم
    date_range = []
    current = start_date
    while current <= end_date:
        date_range.append(current)
        current += timedelta(days=1)
    
    # جلب الموظفين النشطين في القسم
    employees_query = Employee.query.filter_by(
        department_id=department.id,
        status='active'
    )
    
    if project_name and project_name != 'None' and project_name.strip():
        employees_query = employees_query.filter_by(project=project_name)
    
    employees = employees_query.all()
    
    # تسجيل عدد الموظفين للتشخيص
    print(f"تفاصيل القسم - عدد الموظفين المجلوبين: {len(employees)} للقسم {department.name}")
    for emp in employees:
        print(f"  - {emp.name} (ID: {emp.id})")
    
    # جلب سجلات الحضور للموظفين في الفترة المحددة
    employee_attendance = {}
    for employee in employees:
        attendance_records = Attendance.query.filter(
            Attendance.employee_id == employee.id,
            Attendance.date >= start_date,
            Attendance.date <= end_date
        ).order_by(Attendance.date).all()
        
        employee_attendance[employee.id] = {
            'employee': employee,
            'records': attendance_records,
            'stats': {
                'present': sum(1 for r in attendance_records if r.status == 'present'),
                'absent': sum(1 for r in attendance_records if r.status == 'absent'),
                'leave': sum(1 for r in attendance_records if r.status == 'leave'),
                'sick': sum(1 for r in attendance_records if r.status == 'sick')
            }
        }
    
    # حساب الإحصائيات الإجمالية للقسم
    total_stats = {
        'total_employees': len(employees),
        'present': 0,
        'absent': 0,
        'leave': 0,
        'sick': 0,
        'total_records': 0,
        'working_days': len(date_range),
        'attendance_rate': 0
    }
    
    for emp_data in employee_attendance.values():
        total_stats['present'] += emp_data['stats']['present']
        total_stats['absent'] += emp_data['stats']['absent']
        total_stats['leave'] += emp_data['stats']['leave']
        total_stats['sick'] += emp_data['stats']['sick']
        total_stats['total_records'] += len(emp_data['records'])
    
    # حساب معدل الحضور
    if total_stats['total_records'] > 0:
        total_stats['attendance_rate'] = round((total_stats['present'] / total_stats['total_records']) * 100, 1)
    
    # حساب الإحصائيات اليومية للقسم
    daily_stats = {}
    for date in date_range:
        daily_count = {
            'present': 0,
            'absent': 0,
            'leave': 0,
            'sick': 0,
            'total': 0
        }
        
        for emp_data in employee_attendance.values():
            for record in emp_data['records']:
                if record.date == date:
                    daily_count[record.status] += 1
                    daily_count['total'] += 1
                    break
        
        daily_stats[date] = daily_count
    
    # إحصائيات أسبوعية
    weekly_stats = []
    week_start = start_date
    while week_start <= end_date:
        week_end = min(week_start + timedelta(days=6), end_date)
        
        week_data = {
            'start_date': week_start,
            'end_date': week_end,
            'present': 0,
            'absent': 0,
            'leave': 0,
            'sick': 0
        }
        
        current = week_start
        while current <= week_end:
            if current in daily_stats:
                week_data['present'] += daily_stats[current]['present']
                week_data['absent'] += daily_stats[current]['absent']
                week_data['leave'] += daily_stats[current]['leave']
                week_data['sick'] += daily_stats[current]['sick']
            current += timedelta(days=1)
        
        weekly_stats.append(week_data)
        week_start += timedelta(days=7)
    
    return render_template('attendance/department_details_enhanced.html',
                          department=department,
                          employee_attendance=employee_attendance,
                          date_range=date_range,
                          daily_stats=daily_stats,
                          weekly_stats=weekly_stats,
                          total_stats=total_stats,
                          period='monthly',  # دائماً عرض شهري
                          start_date=start_date,
                          end_date=end_date,
                          project_name=project_name)

@attendance_bp.route('/export-excel-department')
def export_excel_department():
    """تصدير تفاصيل القسم إلى Excel مع لوحة معلومات تفصيلية مميزة"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
        
        department_name = request.args.get('department')
        selected_project = request.args.get('project', None)
        
        if not department_name:
            flash('يجب تحديد القسم', 'error')
            return redirect(url_for('attendance.dashboard'))
        
        # جلب القسم
        department = Department.query.filter_by(name=department_name).first()
        if not department:
            flash('القسم غير موجود', 'error')
            return redirect(url_for('attendance.dashboard'))
        
        today = datetime.now().date()
        start_date = today.replace(day=1)
        end_date = today
        
        # إنشاء قائمة بجميع أيام الشهر حتى اليوم
        date_range = []
        current = start_date
        while current <= end_date:
            date_range.append(current)
            current += timedelta(days=1)
        
        # جلب الموظفين والبيانات - استخدام علاقة many-to-many
        all_employees = [emp for emp in department.employees if emp.status not in ['terminated', 'inactive']]
        
        if selected_project and selected_project != 'None' and selected_project.strip():
            employees = [emp for emp in all_employees if emp.project == selected_project]
        else:
            employees = all_employees
        
        # إنشاء ملف Excel
        wb = Workbook()
        
        # صفحة لوحة المعلومات الرئيسية
        ws_dashboard = wb.active
        ws_dashboard.title = "لوحة المعلومات"
        
        # العنوان الرئيسي
        title = f"لوحة معلومات قسم {department.name}"
        if selected_project and selected_project != 'None':
            title += f" - مشروع {selected_project}"
        
        ws_dashboard['A1'] = title
        ws_dashboard['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws_dashboard['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard.merge_cells('A1:H3')
        
        # معلومات الفترة
        period_info = f"الفترة: من {start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')}"
        ws_dashboard['A4'] = period_info
        ws_dashboard['A4'].font = Font(size=12, bold=True)
        ws_dashboard['A4'].alignment = Alignment(horizontal='center')
        ws_dashboard.merge_cells('A4:H4')
        
        # جمع بيانات الموظفين والحضور
        employee_data = []
        total_stats = {
            'total_employees': len(employees),
            'present': 0,
            'absent': 0,
            'leave': 0,
            'sick': 0,
            'total_records': 0
        }
        
        for employee in employees:
            attendance_records = Attendance.query.filter(
                Attendance.employee_id == employee.id,
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).all()
            
            stats = {
                'present': sum(1 for r in attendance_records if r.status == 'present'),
                'absent': sum(1 for r in attendance_records if r.status == 'absent'),
                'leave': sum(1 for r in attendance_records if r.status == 'leave'),
                'sick': sum(1 for r in attendance_records if r.status == 'sick')
            }
            
            total_records = sum(stats.values())
            attendance_rate = (stats['present'] / total_records * 100) if total_records > 0 else 0
            
            employee_data.append({
                'الاسم': employee.name,
                'الهوية': employee.employee_id or 'غير محدد',
                'حاضر': stats['present'],
                'غائب': stats['absent'],
                'إجازة': stats['leave'],
                'مرضي': stats['sick'],
                'الإجمالي': total_records,
                'المعدل %': round(attendance_rate, 1)
            })
            
            # إضافة للإحصائيات الإجمالية
            total_stats['present'] += stats['present']
            total_stats['absent'] += stats['absent']
            total_stats['leave'] += stats['leave']
            total_stats['sick'] += stats['sick']
            total_stats['total_records'] += total_records
        
        # حساب معدل الحضور الإجمالي
        overall_rate = (total_stats['present'] / total_stats['total_records'] * 100) if total_stats['total_records'] > 0 else 0
        
        # الإحصائيات الإجمالية في لوحة المعلومات
        stats_row = 6
        
        # عناوين الإحصائيات
        stats_headers = ['إجمالي الموظفين', 'إجمالي الحضور', 'إجمالي الغياب', 'إجمالي الإجازات', 'إجمالي المرضي', 'معدل الحضور %', 'أيام العمل']
        stats_values = [total_stats['total_employees'], total_stats['present'], total_stats['absent'], 
                       total_stats['leave'], total_stats['sick'], round(overall_rate, 1), len(date_range)]
        
        for col, (header, value) in enumerate(zip(stats_headers, stats_values), 1):
            # العنوان
            header_cell = ws_dashboard.cell(row=stats_row, column=col, value=header)
            header_cell.font = Font(bold=True, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # القيمة
            value_cell = ws_dashboard.cell(row=stats_row + 1, column=col, value=value)
            value_cell.font = Font(bold=True, size=14)
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # تلوين حسب النوع
            if 'حضور' in header:
                value_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif 'غياب' in header:
                value_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif 'إجازة' in header:
                value_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif 'مرضي' in header:
                value_cell.fill = PatternFill(start_color="9CC3F7", end_color="9CC3F7", fill_type="solid")
            else:
                value_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # جدول تفاصيل الموظفين
        table_start_row = stats_row + 4
        
        # عنوان الجدول
        ws_dashboard[f'A{table_start_row}'] = "تفاصيل حضور الموظفين"
        ws_dashboard[f'A{table_start_row}'].font = Font(size=14, bold=True)
        ws_dashboard.merge_cells(f'A{table_start_row}:H{table_start_row}')
        
        # رؤوس الجدول
        headers = list(employee_data[0].keys()) if employee_data else []
        header_row = table_start_row + 1
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_dashboard.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # بيانات الموظفين
        for row_num, emp_data in enumerate(employee_data, header_row + 1):
            for col_num, value in enumerate(emp_data.values(), 1):
                cell = ws_dashboard.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # تلوين صفوف متناوبة
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # تعديل عرض الأعمدة
        ws_dashboard.column_dimensions['A'].width = 25
        ws_dashboard.column_dimensions['B'].width = 15
        ws_dashboard.column_dimensions['C'].width = 12
        ws_dashboard.column_dimensions['D'].width = 12
        ws_dashboard.column_dimensions['E'].width = 12
        ws_dashboard.column_dimensions['F'].width = 12
        ws_dashboard.column_dimensions['G'].width = 12
        ws_dashboard.column_dimensions['H'].width = 12
        
        # صفحة التفاصيل اليومية
        ws_daily = wb.create_sheet("التفاصيل اليومية")
        
        # عنوان صفحة التفاصيل اليومية
        ws_daily['A1'] = f"التفاصيل اليومية - قسم {department.name}"
        ws_daily['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws_daily['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws_daily['A1'].alignment = Alignment(horizontal='center')
        ws_daily.merge_cells('A1:AF3')
        
        # رؤوس الأعمدة للتفاصيل اليومية
        ws_daily['A5'] = 'اسم الموظف'
        ws_daily['B5'] = 'رقم الهوية'
        
        # تواريخ الشهر
        for col_num, date in enumerate(date_range, 3):
            cell = ws_daily.cell(row=5, column=col_num, value=date.strftime('%d-%m'))
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', textRotation=90)
        
        # بيانات الحضور اليومي
        for row_num, employee in enumerate(employees, 6):
            ws_daily.cell(row=row_num, column=1, value=employee.name)
            ws_daily.cell(row=row_num, column=2, value=employee.employee_id or 'غير محدد')
            
            # جلب سجلات الحضور للموظف
            attendance_records = Attendance.query.filter(
                Attendance.employee_id == employee.id,
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).all()
            
            attendance_dict = {record.date: record.status for record in attendance_records}
            
            for col_num, date in enumerate(date_range, 3):
                status = attendance_dict.get(date, '-')
                cell = ws_daily.cell(row=row_num, column=col_num, value=status)
                
                # تلوين الخلايا حسب الحالة
                if status == 'present':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.value = "✓"
                elif status == 'absent':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.value = "✗"
                elif status == 'leave':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.value = "إ"
                elif status == 'sick':
                    cell.fill = PatternFill(start_color="9CC3F7", end_color="9CC3F7", fill_type="solid")
                    cell.value = "م"
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # تنسيق أعمدة التفاصيل اليومية
        ws_daily.column_dimensions['A'].width = 30
        ws_daily.column_dimensions['B'].width = 15
        
        # تنسيق أعمدة التواريخ (تحديد عرض بسيط لتجنب مشاكل الأعمدة الكثيرة)
        try:
            for col_num in range(3, min(3 + len(date_range), 26)):  # حد أقصى 26 عمود
                if col_num <= 26:  # A-Z فقط
                    col_letter = chr(64 + col_num)
                    ws_daily.column_dimensions[col_letter].width = 4
        except Exception:
            pass  # تجاهل أخطاء الأعمدة
        
        # حفظ الملف
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'تفاصيل_قسم_{department.name}_{today.strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"خطأ في تصدير تفاصيل القسم: {str(e)}")
        flash('حدث خطأ أثناء تصدير الملف', 'error')
        return redirect(url_for('attendance.dashboard'))

@attendance_bp.route('/department/view', methods=['GET'])
def department_attendance_view():
    """عرض حضور الأقسام خلال فترة زمنية محددة"""
    from flask_login import current_user
    from sqlalchemy import or_
    
    # الحصول على معاملات الفلترة
    department_id = request.args.get('department_id', '')
    search_query = request.args.get('search_query', '').strip()
    status_filter = request.args.get('status_filter', '')
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    working_days_required = request.args.get('working_days_required', '26')
    
    # تحديد التواريخ الافتراضية (آخر 30 يوم)
    if not start_date_str:
        start_date = datetime.now().date() - timedelta(days=30)
    else:
        start_date = parse_date(start_date_str)
    
    if not end_date_str:
        end_date = datetime.now().date()
    else:
        end_date = parse_date(end_date_str)
    
    # الحصول على الأقسام حسب صلاحيات المستخدم
    if current_user.is_authenticated:
        departments = current_user.get_accessible_departments()
    else:
        departments = Department.query.all()
    
    # بناء الاستعلام
    query = Attendance.query.join(Employee).filter(
        Attendance.date >= start_date,
        Attendance.date <= end_date
    )
    
    # تطبيق فلتر القسم
    if department_id:
        query = query.join(employee_departments).filter(
            employee_departments.c.department_id == int(department_id)
        )
    
    # تطبيق البحث عن الموظف (الاسم، رقم الهوية، الرقم الوظيفي)
    if search_query:
        query = query.filter(
            or_(
                Employee.name.ilike(f'%{search_query}%'),
                Employee.employee_id.ilike(f'%{search_query}%'),
                Employee.national_id.ilike(f'%{search_query}%')
            )
        )
    
    # تطبيق فلتر الحالة
    if status_filter:
        query = query.filter(Attendance.status == status_filter)
    
    # الحصول على السجلات مرتبة حسب التاريخ والموظف
    attendances = query.order_by(Attendance.date.desc(), Employee.name).all()
    
    # حساب الإحصائيات
    total_count = len(attendances)
    present_count = sum(1 for a in attendances if a.status == 'present')
    absent_count = sum(1 for a in attendances if a.status == 'absent')
    leave_count = sum(1 for a in attendances if a.status == 'leave')
    sick_count = sum(1 for a in attendances if a.status == 'sick')
    
    return render_template('attendance/department_period.html',
                          departments=departments,
                          department_id=department_id,
                          search_query=search_query,
                          status_filter=status_filter,
                          start_date=start_date,
                          end_date=end_date,
                          attendances=attendances,
                          total_count=total_count,
                          present_count=present_count,
                          absent_count=absent_count,
                          leave_count=leave_count,
                          sick_count=sick_count,
                          working_days_required=working_days_required)

@attendance_bp.route('/department/export-data', methods=['GET'])
def export_department_data():
    """تصدير بيانات الحضور حسب الفلاتر مع تصميم احترافي"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    try:
        department_id = request.args.get('department_id', '')
        search_query = request.args.get('search_query', '').strip()
        status_filter = request.args.get('status_filter', '')
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')
        
        # تحديد التواريخ
        if not start_date_str:
            start_date = datetime.now().date() - timedelta(days=30)
        else:
            start_date = parse_date(start_date_str)
        
        if not end_date_str:
            end_date = datetime.now().date()
        else:
            end_date = parse_date(end_date_str)
        
        # بناء الاستعلام
        query = Attendance.query.join(Employee).filter(
            Attendance.date >= start_date,
            Attendance.date <= end_date
        )
        
        # تطبيق الفلاتر
        if department_id:
            query = query.join(employee_departments).filter(
                employee_departments.c.department_id == int(department_id)
            )
        
        if search_query:
            query = query.filter(
                or_(
                    Employee.name.ilike(f'%{search_query}%'),
                    Employee.employee_id.ilike(f'%{search_query}%'),
                    Employee.national_id.ilike(f'%{search_query}%')
                )
            )
        
        if status_filter:
            query = query.filter(Attendance.status == status_filter)
        
        attendances = query.order_by(Attendance.date.desc(), Employee.name).all()
        
        # حساب الإحصائيات
        total_count = len(attendances)
        present_count = sum(1 for a in attendances if a.status == 'present')
        absent_count = sum(1 for a in attendances if a.status == 'absent')
        leave_count = sum(1 for a in attendances if a.status == 'leave')
        sick_count = sum(1 for a in attendances if a.status == 'sick')
        
        # إنشاء ملف Excel
        wb = Workbook()
        
        # ========== شيت البيانات الرئيسي - نموذج P/A ==========
        ws = wb.active
        ws.title = "بيانات الحضور"
        
        # الحصول على جميع التواريخ والموظفين الفريدين
        all_dates = sorted(set(att.date for att in attendances if att.date))
        employees_dict = {}
        for att in attendances:
            if att.employee.id not in employees_dict:
                employees_dict[att.employee.id] = att.employee
        
        # ترتيب الموظفين حسب الاسم
        sorted_employees = sorted(employees_dict.values(), key=lambda e: e.name)
        
        # بناء قاموس للوصول السريع: (employee_id, date) -> status
        attendance_map = {}
        for att in attendances:
            key = (att.employee.id, att.date)
            attendance_map[key] = att.status
        
        # تحديد عرض الأعمدة
        ws.column_dimensions['A'].width = 20  # الاسم
        ws.column_dimensions['B'].width = 16  # الرقم الوظيفي
        ws.column_dimensions['C'].width = 16  # رقم الهوية
        ws.column_dimensions['D'].width = 22  # القسم
        
        for col_idx in range(len(all_dates)):
            col_letter = get_column_letter(col_idx + 5)
            ws.column_dimensions[col_letter].width = 4
        
        # إنشاء رأس الجدول - مع اسم الشهر (3 حروف) واليوم بالإنجليزي
        header_row = ['الموظف', 'الرقم الوظيفي', 'رقم الهوية', 'القسم'] + [
            d.strftime('%b %d') for d in all_dates
        ]
        ws.append(header_row)
        
        # تنسيق الرأس الاحترافي
        header_fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 28
        
        # إضافة بيانات الموظفين
        row_num = 2
        for emp in sorted_employees:
            department_name = ', '.join([d.name for d in emp.departments]) if emp.departments else '-'
            
            row_data = [
                emp.name,
                emp.employee_id or '-',
                emp.national_id if hasattr(emp, 'national_id') and emp.national_id else '-',
                department_name
            ]
            
            for date in all_dates:
                status = attendance_map.get((emp.id, date), '')
                
                # تحويل الحالة إلى P أو A أو S
                if status == 'present':
                    row_data.append('P')
                elif status == 'absent':
                    row_data.append('A')
                elif status == 'leave' or status == 'sick':
                    row_data.append('S')
                else:
                    row_data.append('')
            
            ws.append(row_data)
            
            # تنسيق الصفوف
            for col_idx in range(1, len(row_data) + 1):
                cell_obj = ws.cell(row=row_num, column=col_idx)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_obj.border = Border(
                    left=Side(style='thin', color='E0E0E0'),
                    right=Side(style='thin', color='E0E0E0'),
                    top=Side(style='thin', color='E0E0E0'),
                    bottom=Side(style='thin', color='E0E0E0')
                )
                
                if col_idx <= 4:
                    # أعمدة بيانات الموظف (الاسم، الرقم الوظيفي، رقم الهوية، القسم)
                    cell_obj.font = Font(bold=True if col_idx == 1 else False)
                    if row_num % 2 == 0:
                        cell_obj.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                else:
                    # أعمدة الحضور (P/A/S)
                    value = cell_obj.value
                    if value == 'P':
                        cell_obj.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        cell_obj.font = Font(bold=True, color="155724", size=11)
                    elif value == 'A':
                        cell_obj.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                        cell_obj.font = Font(bold=True, color="721C24", size=11)
                    elif value == 'S':
                        cell_obj.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                        cell_obj.font = Font(bold=True, color="856404", size=11)
                    else:
                        if row_num % 2 == 0:
                            cell_obj.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            ws.row_dimensions[row_num].height = 20
            row_num += 1
        
        # ========== شيت الإحصائيات ==========
        stats_ws = wb.create_sheet("الإحصائيات")
        
        # عنوان الإحصائيات
        stats_ws.merge_cells('A1:D1')
        title_cell = stats_ws['A1']
        title_cell.value = "ملخص إحصائيات الحضور"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        stats_ws.row_dimensions[1].height = 32
        
        # إعدادات الأعمدة
        stats_ws.column_dimensions['A'].width = 25
        stats_ws.column_dimensions['B'].width = 18
        stats_ws.column_dimensions['C'].width = 18
        stats_ws.column_dimensions['D'].width = 18
        
        # البيانات الإحصائية
        stats_data = [
            ['', 'العدد', 'النسبة %', 'الحالة'],
            ['إجمالي السجلات', total_count, '100%', 'total'],
            ['موظفون حاضرون', present_count, f'{int((present_count/max(total_count,1))*100)}%', 'present'],
            ['موظفون غائبون', absent_count, f'{int((absent_count/max(total_count,1))*100)}%', 'absent'],
            ['إجازات', leave_count, f'{int((leave_count/max(total_count,1))*100)}%', 'leave'],
            ['مرضي', sick_count, f'{int((sick_count/max(total_count,1))*100)}%', 'sick']
        ]
        
        # إضافة البيانات الإحصائية
        for idx, row_data in enumerate(stats_data, 2):
            for col_idx, value in enumerate(row_data[:3], 1):
                cell = stats_ws.cell(row=idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if idx == 2:  # رأس الإحصائيات
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
                elif row_data[3] == 'total':
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                elif row_data[3] == 'present':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    cell.font = Font(bold=True, color="155724")
                elif row_data[3] == 'absent':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    cell.font = Font(bold=True, color="721C24")
                elif row_data[3] == 'leave':
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    cell.font = Font(bold=True, color="856404")
                elif row_data[3] == 'sick':
                    cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
                    cell.font = Font(bold=True, color="0C5460")
                
                stats_ws.row_dimensions[idx].height = 26
        
        # إنشاء رسم بياني دائري للإحصائيات
        if total_count > 0:
            pie = PieChart()
            pie.title = "توزيع حالات الحضور"
            pie.style = 10
            labels = Reference(stats_ws, min_col=1, min_row=3, max_row=6)
            data = Reference(stats_ws, min_col=2, min_row=2, max_row=6)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 12
            pie.width = 16
            stats_ws.add_chart(pie, "A9")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"حضور_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    
    except Exception as e:
        logger.error(f"خطأ في تصدير البيانات: {str(e)}")
        flash(f'خطأ في التصدير: {str(e)}', 'error')
        return redirect(url_for('attendance.department_attendance_view'))


@attendance_bp.route('/department/export-period', methods=['GET'])
def export_department_period():
    """تصدير حضور قسم خلال فترة زمنية إلى Excel مع dashboard احترافي"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    try:
        department_id = request.args.get('department_id')
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        if not all([start_date_str, end_date_str]):
            flash('يجب تحديد الفترة الزمنية', 'error')
            return redirect(url_for('attendance.department_attendance_view'))
        
        # تحليل التواريخ
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
        
        # الحصول على القسم إذا تم تحديده
        department = None
        department_name = "جميع الأقسام"
        if department_id:
            department = Department.query.get_or_404(department_id)
            department_name = department.name
        
        # الحصول على بيانات الحضور
        query = Attendance.query.join(Employee).filter(
            Attendance.date >= start_date,
            Attendance.date <= end_date
        )
        
        if department_id:
            query = query.join(employee_departments).filter(
                employee_departments.c.department_id == int(department_id)
            )
        
        attendances = query.order_by(Employee.name, Attendance.date).all()
        
        if not attendances:
            flash('لا توجد بيانات للتصدير في هذه الفترة', 'warning')
            return redirect(url_for('attendance.department_attendance_view'))
        
        # إنشاء workbook
        wb = Workbook()
        
        # ========== ورقة Dashboard ==========
        ws_dashboard = wb.active
        ws_dashboard.title = "Dashboard"
        
        # تنسيقات احترافية
        title_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=20)
        subtitle_fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
        subtitle_font = Font(bold=True, color="FFFFFF", size=14)
        
        kpi_label_format = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        kpi_success_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        kpi_danger_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        kpi_warning_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
        kpi_info_fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
        
        # عنوان Dashboard
        ws_dashboard.merge_cells('A1:H1')
        ws_dashboard['A1'] = f'تقرير حضور قسم {department_name}'
        ws_dashboard['A1'].fill = title_fill
        ws_dashboard['A1'].font = title_font
        ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard.row_dimensions[1].height = 35
        
        # معلومات الفترة
        ws_dashboard.merge_cells('A2:H2')
        ws_dashboard['A2'] = f'الفترة: من {start_date.strftime("%Y-%m-%d")} إلى {end_date.strftime("%Y-%m-%d")}'
        ws_dashboard['A2'].fill = subtitle_fill
        ws_dashboard['A2'].font = subtitle_font
        ws_dashboard['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard.row_dimensions[2].height = 25
        
        # حساب الإحصائيات
        total_count = len(attendances)
        present_count = sum(1 for a in attendances if a.status == 'present')
        absent_count = sum(1 for a in attendances if a.status == 'absent')
        leave_count = sum(1 for a in attendances if a.status == 'leave')
        sick_count = sum(1 for a in attendances if a.status == 'sick')
        attendance_rate = (present_count / total_count * 100) if total_count > 0 else 0
        
        # عدد الموظفين الفريدين
        unique_employees = set(a.employee_id for a in attendances)
        employee_count = len(unique_employees)
        
        # عدد الأيام
        days_count = (end_date - start_date).days + 1
        
        # KPIs - الصف 4
        kpis = [
            ('A4', 'B4', 'إجمالي السجلات', total_count, kpi_label_format, kpi_info_fill),
            ('C4', 'D4', 'عدد الموظفين', employee_count, kpi_label_format, kpi_info_fill),
            ('E4', 'F4', 'عدد الأيام', days_count, kpi_label_format, kpi_info_fill),
            ('G4', 'H4', 'نسبة الحضور', f'{attendance_rate:.1f}%', kpi_label_format, kpi_success_fill),
        ]
        
        for start_cell, end_cell, label, value, label_fill, value_fill in kpis:
            # Label
            ws_dashboard.merge_cells(f'{start_cell}:{start_cell}')
            ws_dashboard[start_cell] = label
            ws_dashboard[start_cell].fill = label_fill
            ws_dashboard[start_cell].font = Font(bold=True, size=11)
            ws_dashboard[start_cell].alignment = Alignment(horizontal='center', vertical='center')
            ws_dashboard[start_cell].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='thin')
            )
            
            # Value
            ws_dashboard.merge_cells(f'{end_cell}:{end_cell}')
            ws_dashboard[end_cell] = value
            ws_dashboard[end_cell].fill = value_fill
            ws_dashboard[end_cell].font = Font(bold=True, size=16, color='1e3c72')
            ws_dashboard[end_cell].alignment = Alignment(horizontal='center', vertical='center')
            ws_dashboard[end_cell].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='thin'), bottom=Side(style='medium')
            )
        
        ws_dashboard.row_dimensions[4].height = 30
        
        # إحصائيات الحضور - الصف 6
        ws_dashboard.merge_cells('A6:H6')
        ws_dashboard['A6'] = 'توزيع حالات الحضور'
        ws_dashboard['A6'].fill = subtitle_fill
        ws_dashboard['A6'].font = Font(bold=True, color="FFFFFF", size=13)
        ws_dashboard['A6'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard.row_dimensions[6].height = 25
        
        # جدول الإحصائيات للرسم البياني - يبدأ من الصف 8
        stats_data = [
            ('A8', 'B8', 'حاضر', present_count, kpi_success_fill),
            ('C8', 'D8', 'غائب', absent_count, kpi_danger_fill),
            ('E8', 'F8', 'إجازة', leave_count, kpi_warning_fill),
            ('G8', 'H8', 'مرضي', sick_count, kpi_info_fill),
        ]
        
        for start_cell, end_cell, label, value, fill in stats_data:
            # Label
            ws_dashboard.merge_cells(f'{start_cell}:{start_cell}')
            ws_dashboard[start_cell] = label
            ws_dashboard[start_cell].fill = fill
            ws_dashboard[start_cell].font = Font(bold=True, size=12)
            ws_dashboard[start_cell].alignment = Alignment(horizontal='center', vertical='center')
            ws_dashboard[start_cell].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='thin')
            )
            
            # Value
            ws_dashboard.merge_cells(f'{end_cell}:{end_cell}')
            ws_dashboard[end_cell] = value
            ws_dashboard[end_cell].fill = fill
            ws_dashboard[end_cell].font = Font(bold=True, size=18)
            ws_dashboard[end_cell].alignment = Alignment(horizontal='center', vertical='center')
            ws_dashboard[end_cell].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='thin'), bottom=Side(style='medium')
            )
        
        ws_dashboard.row_dimensions[8].height = 30
        
        # بيانات للمخطط الدائري
        ws_dashboard['A10'] = 'الحالة'
        ws_dashboard['B10'] = 'العدد'
        ws_dashboard['A10'].font = Font(bold=True)
        ws_dashboard['B10'].font = Font(bold=True)
        
        chart_data = [
            ('حاضر', present_count),
            ('غائب', absent_count),
            ('إجازة', leave_count),
            ('مرضي', sick_count),
        ]
        
        for idx, (label, value) in enumerate(chart_data, start=11):
            ws_dashboard[f'A{idx}'] = label
            ws_dashboard[f'B{idx}'] = value
        
        # إنشاء مخطط دائري
        pie_chart = PieChart()
        pie_chart.title = "توزيع حالات الحضور"
        pie_chart.style = 10
        pie_chart.height = 10
        pie_chart.width = 15
        
        labels = Reference(ws_dashboard, min_col=1, min_row=11, max_row=14)
        data = Reference(ws_dashboard, min_col=2, min_row=10, max_row=14)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        
        # إضافة تسميات البيانات
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showVal = True
        
        ws_dashboard.add_chart(pie_chart, "J6")
        
        # تعديل عرض الأعمدة
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws_dashboard.column_dimensions[col].width = 15
        
        # ========== ورقة البيانات التفصيلية ==========
        ws_data = wb.create_sheet("البيانات التفصيلية")
        
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        present_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        absent_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        leave_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
        sick_fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
        
        # العنوان
        ws_data.merge_cells('A1:H1')
        ws_data['A1'] = f'سجلات حضور قسم {department_name}'
        ws_data['A1'].fill = title_fill
        ws_data['A1'].font = title_font
        ws_data['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_data.row_dimensions[1].height = 30
        
        # رؤوس الأعمدة
        headers = ['الموظف', 'الرقم الوظيفي', 'التاريخ', 'الحالة', 'وقت الدخول', 'وقت الخروج', 'ساعات العمل', 'ملاحظات']
        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # البيانات
        row_num = 4
        for attendance in attendances:
            ws_data.cell(row=row_num, column=1, value=attendance.employee.name)
            ws_data.cell(row=row_num, column=2, value=attendance.employee.employee_id or '-')
            ws_data.cell(row=row_num, column=3, value=attendance.date.strftime('%Y-%m-%d'))
            
            # حالة الحضور مع التلوين
            status_cell = ws_data.cell(row=row_num, column=4)
            if attendance.status == 'present':
                status_cell.value = 'حاضر'
                status_cell.fill = present_fill
            elif attendance.status == 'absent':
                status_cell.value = 'غائب'
                status_cell.fill = absent_fill
            elif attendance.status == 'leave':
                status_cell.value = 'إجازة'
                status_cell.fill = leave_fill
            elif attendance.status == 'sick':
                status_cell.value = 'مرضي'
                status_cell.fill = sick_fill
            
            status_cell.alignment = Alignment(horizontal='center')
            
            ws_data.cell(row=row_num, column=5, value=attendance.check_in.strftime('%H:%M') if attendance.check_in else '-')
            ws_data.cell(row=row_num, column=6, value=attendance.check_out.strftime('%H:%M') if attendance.check_out else '-')
            
            # حساب ساعات العمل
            if attendance.check_in and attendance.check_out:
                hours = (attendance.check_out.hour - attendance.check_in.hour) + (attendance.check_out.minute - attendance.check_in.minute) / 60.0
                ws_data.cell(row=row_num, column=7, value=f"{hours:.1f}")
            else:
                ws_data.cell(row=row_num, column=7, value='-')
            
            ws_data.cell(row=row_num, column=8, value=attendance.notes or '-')
            
            # إضافة حدود للخلايا
            for col in range(1, 9):
                ws_data.cell(row=row_num, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                ws_data.cell(row=row_num, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            row_num += 1
        
        # تنسيق عرض الأعمدة
        ws_data.column_dimensions['A'].width = 25
        ws_data.column_dimensions['B'].width = 15
        ws_data.column_dimensions['C'].width = 15
        ws_data.column_dimensions['D'].width = 12
        ws_data.column_dimensions['E'].width = 12
        ws_data.column_dimensions['F'].width = 12
        ws_data.column_dimensions['G'].width = 15
        ws_data.column_dimensions['H'].width = 30
        
        # ========== ورقة سجل الحضور (Matrix View) ==========
        ws_matrix = wb.create_sheet("سجل الحضور")
        
        # تجميع البيانات حسب الموظف والتاريخ
        employee_attendance_map = {}
        unique_dates = set()
        
        for att in attendances:
            emp_id = att.employee_id
            if emp_id not in employee_attendance_map:
                employee_attendance_map[emp_id] = {
                    'employee': att.employee,
                    'dates': {}
                }
            employee_attendance_map[emp_id]['dates'][att.date] = att.status
            unique_dates.add(att.date)
        
        # ترتيب التواريخ
        sorted_dates = sorted(list(unique_dates))
        
        # العنوان
        ws_matrix.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(sorted_dates) + 8)
        ws_matrix.cell(row=1, column=1, value=f'سجل حضور قسم {department_name}')
        ws_matrix.cell(row=1, column=1).fill = title_fill
        ws_matrix.cell(row=1, column=1).font = title_font
        ws_matrix.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws_matrix.row_dimensions[1].height = 30
        
        # الفترة
        ws_matrix.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(sorted_dates) + 8)
        ws_matrix.cell(row=2, column=1, value=f'الفترة: من {start_date.strftime("%Y-%m-%d")} إلى {end_date.strftime("%Y-%m-%d")}')
        ws_matrix.cell(row=2, column=1).fill = subtitle_fill
        ws_matrix.cell(row=2, column=1).font = subtitle_font
        ws_matrix.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws_matrix.row_dimensions[2].height = 25
        
        # رؤوس الأعمدة الثابتة
        fixed_headers = ['Name', 'ID Number', 'Emp. No', 'Job Title', 'No. Mobile', 'Location', 'Project', 'Total']
        header_row = 4
        
        for col_num, header in enumerate(fixed_headers, 1):
            cell = ws_matrix.cell(row=header_row, column=col_num, value=header)
            cell.fill = PatternFill(start_color="00B0B0", end_color="00B0B0", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # رؤوس الأعمدة للتواريخ
        for idx, date in enumerate(sorted_dates, start=9):
            # صف اليوم (اسم اليوم بالإنجليزي)
            day_name = date.strftime('%a')
            cell = ws_matrix.cell(row=header_row, column=idx, value=day_name)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            ws_matrix.column_dimensions[get_column_letter(idx)].width = 4
        
        # صف التواريخ
        for idx, date in enumerate(sorted_dates, start=9):
            cell = ws_matrix.cell(row=header_row + 1, column=idx, value=date.strftime('%d/%m/%Y'))
            cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # بيانات الموظفين
        data_row = header_row + 2
        for emp_id, emp_data in sorted(employee_attendance_map.items(), key=lambda x: x[1]['employee'].name):
            employee = emp_data['employee']
            
            # البيانات الثابتة
            ws_matrix.cell(row=data_row, column=1, value=employee.name)
            ws_matrix.cell(row=data_row, column=2, value=employee.national_id or '-')
            ws_matrix.cell(row=data_row, column=3, value=employee.employee_id or '-')
            ws_matrix.cell(row=data_row, column=4, value=employee.job_title or '-')
            ws_matrix.cell(row=data_row, column=5, value=employee.mobile or '-')
            ws_matrix.cell(row=data_row, column=6, value=employee.location or '-')
            ws_matrix.cell(row=data_row, column=7, value=employee.project or '-')
            
            # حساب المجموع
            total_days = len(emp_data['dates'])
            ws_matrix.cell(row=data_row, column=8, value=total_days)
            ws_matrix.cell(row=data_row, column=8).font = Font(bold=True)
            ws_matrix.cell(row=data_row, column=8).alignment = Alignment(horizontal='center', vertical='center')
            
            # حالات الحضور
            for idx, date in enumerate(sorted_dates, start=9):
                status = emp_data['dates'].get(date, '')
                cell = ws_matrix.cell(row=data_row, column=idx)
                
                if status == 'present':
                    cell.value = 'P'
                    cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                    cell.font = Font(bold=True, color="155724")
                elif status == 'absent':
                    cell.value = 'A'
                    cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                    cell.font = Font(bold=True, color="721c24")
                elif status == 'leave':
                    cell.value = 'L'
                    cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
                    cell.font = Font(bold=True, color="856404")
                elif status == 'sick':
                    cell.value = 'S'
                    cell.fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
                    cell.font = Font(bold=True, color="0c5460")
                else:
                    cell.value = ''
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            # تنسيق الصف
            for col in range(1, 9):
                cell = ws_matrix.cell(row=data_row, column=col)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            data_row += 1
        
        # تنسيق عرض الأعمدة الثابتة
        ws_matrix.column_dimensions['A'].width = 25
        ws_matrix.column_dimensions['B'].width = 15
        ws_matrix.column_dimensions['C'].width = 12
        ws_matrix.column_dimensions['D'].width = 20
        ws_matrix.column_dimensions['E'].width = 15
        ws_matrix.column_dimensions['F'].width = 15
        ws_matrix.column_dimensions['G'].width = 15
        ws_matrix.column_dimensions['H'].width = 8
        
        # تجميد الأعمدة الثابتة والصفوف العليا
        ws_matrix.freeze_panes = ws_matrix.cell(row=header_row + 2, column=9)
        
        # حفظ الملف
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'تقرير_حضور_{department_name}_{start_date.strftime("%Y%m%d")}_إلى_{end_date.strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"خطأ في تصدير حضور الفترة: {str(e)}")
        import traceback
        traceback.print_exc()
        flash('حدث خطأ أثناء تصدير الملف', 'error')
        return redirect(url_for('attendance.department_attendance_view'))

@attendance_bp.route('/department/bulk', methods=['GET', 'POST'])
def department_bulk_attendance():
    """تسجيل حضور قسم كامل لفترة زمنية محددة"""
    from flask_login import current_user
    
    if request.method == 'POST':
        try:
            department_id = request.form['department_id']
            start_date_str = request.form['start_date']
            end_date_str = request.form['end_date']
            status = request.form['status']
            skip_weekends = 'skip_weekends' in request.form
            overwrite_existing = 'overwrite_existing' in request.form
            
            # التحقق من صلاحيات المستخدم
            if current_user.is_authenticated and not current_user.can_access_department(int(department_id)):
                flash('ليس لديك صلاحية لتسجيل حضور هذا القسم', 'error')
                return redirect(url_for('attendance.department_bulk_attendance'))
            
            # تحليل التواريخ
            start_date = parse_date(start_date_str)
            end_date = parse_date(end_date_str)
            
            # التحقق من صحة الفترة
            if start_date > end_date:
                flash('تاريخ البداية يجب أن يكون قبل تاريخ النهاية', 'error')
                return redirect(url_for('attendance.department_bulk_attendance'))
            
            # الحد الأقصى للفترة (90 يوم)
            if (end_date - start_date).days > 90:
                flash('الفترة الزمنية لا يمكن أن تتجاوز 90 يوم', 'error')
                return redirect(url_for('attendance.department_bulk_attendance'))
            
            # الحصول على القسم والموظفين
            department = Department.query.get_or_404(department_id)
            employees = [emp for emp in department.employees if emp.status not in ['terminated', 'inactive']]
            
            if not employees:
                flash('لا يوجد موظفين نشطين في هذا القسم', 'warning')
                return redirect(url_for('attendance.department_bulk_attendance'))
            
            # إنشاء قائمة التواريخ
            date_list = []
            current_date = start_date
            while current_date <= end_date:
                # تخطي عطلة نهاية الأسبوع إذا تم تحديد الخيار
                if skip_weekends:
                    # 4 = الجمعة، 5 = السبت في Python (0=الإثنين)
                    if current_date.weekday() in [4, 5]:
                        current_date += timedelta(days=1)
                        continue
                
                date_list.append(current_date)
                current_date += timedelta(days=1)
            
            # حساب الإحصائيات
            created_count = 0
            updated_count = 0
            skipped_count = 0
            
            # تسجيل الحضور لكل موظف في كل يوم
            for employee in employees:
                for attendance_date in date_list:
                    # البحث عن سجل موجود
                    existing = Attendance.query.filter_by(
                        employee_id=employee.id,
                        date=attendance_date
                    ).first()
                    
                    if existing:
                        if overwrite_existing:
                            # تحديث السجل الموجود
                            existing.status = status
                            if status != 'present':
                                existing.check_in = None
                                existing.check_out = None
                            updated_count += 1
                        else:
                            # تخطي السجل الموجود
                            skipped_count += 1
                    else:
                        # إنشاء سجل جديد
                        new_attendance = Attendance(
                            employee_id=employee.id,
                            date=attendance_date,
                            status=status
                        )
                        db.session.add(new_attendance)
                        created_count += 1
            
            # حفظ التغييرات
            db.session.commit()
            
            # تسجيل العملية في سجل النشاط
            log_attendance_activity(
                action='bulk_create_period',
                attendance_data={
                    'department_id': department_id,
                    'start_date': start_date.isoformat(),
                    'end_date': end_date.isoformat(),
                    'status': status,
                    'created': created_count,
                    'updated': updated_count,
                    'skipped': skipped_count
                },
                employee_name=f"جميع موظفي قسم {department.name}"
            )
            
            # رسالة النجاح
            message_parts = []
            if created_count > 0:
                message_parts.append(f'تم إنشاء {created_count} سجل جديد')
            if updated_count > 0:
                message_parts.append(f'تم تحديث {updated_count} سجل')
            if skipped_count > 0:
                message_parts.append(f'تم تخطي {skipped_count} سجل موجود')
            
            flash(' | '.join(message_parts), 'success')
            return redirect(url_for('attendance.index'))
            
        except Exception as e:
            db.session.rollback()
            print(f"خطأ في تسجيل الحضور الجماعي: {str(e)}")
            flash(f'حدث خطأ: {str(e)}', 'danger')
            return redirect(url_for('attendance.department_bulk_attendance'))
    
    # GET request
    if current_user.is_authenticated:
        departments = current_user.get_accessible_departments()
    else:
        departments = Department.query.all()
    
    return render_template('attendance/department_bulk.html', departments=departments)

@attendance_bp.route('/edit/<int:id>', methods=['GET'])
def edit_attendance_page(id):
    """عرض صفحة تعديل سجل الحضور"""
    from flask_login import current_user
    
    # الحصول على سجل الحضور
    attendance = Attendance.query.get_or_404(id)
    
    # التحقق من صلاحيات المستخدم
    if current_user.is_authenticated:
        employee_departments = [dept.id for dept in attendance.employee.departments]
        if employee_departments and not any(current_user.can_access_department(dept_id) for dept_id in employee_departments):
            flash('ليس لديك صلاحية لتعديل هذا السجل', 'error')
            return redirect(url_for('attendance.department_attendance_view'))
    
    return render_template('attendance/edit_attendance.html', attendance=attendance)

@attendance_bp.route('/edit/<int:id>', methods=['POST'])
def update_attendance_page(id):
    """تحديث سجل حضور موجود من صفحة التعديل"""
    from flask_login import current_user
    from datetime import time as dt_time
    
    try:
        # الحصول على سجل الحضور
        attendance = Attendance.query.get_or_404(id)
        
        # التحقق من صلاحيات المستخدم
        if current_user.is_authenticated:
            employee_departments = [dept.id for dept in attendance.employee.departments]
            if employee_departments and not any(current_user.can_access_department(dept_id) for dept_id in employee_departments):
                flash('ليس لديك صلاحية لتعديل هذا السجل', 'error')
                return redirect(url_for('attendance.department_attendance_view'))
        
        # الحصول على البيانات
        status = request.form.get('status')
        check_in_str = request.form.get('check_in', '')
        check_out_str = request.form.get('check_out', '')
        notes = request.form.get('notes', '')
        
        # تحديث الحالة
        old_status = attendance.status
        attendance.status = status
        attendance.notes = notes if notes else None
        
        # معالجة رفع ملف الإجازة المرضية
        if status == 'sick' and 'sick_leave_file' in request.files:
            file = request.files['sick_leave_file']
            if file and file.filename:
                from utils.storage_helper import upload_image
                import os
                from werkzeug.utils import secure_filename
                
                # 💾 لا حذف للملفات القديمة - الاحتفاظ بجميع الملفات للأمان
                # حفظ الملف الجديد
                filename = secure_filename(file.filename)
                file_path = upload_image(file, 'sick_leaves', filename)
                if file_path:
                    # إزالة "static/" من البداية لأن url_for('static') سيضيفها تلقائياً
                    if file_path.startswith('static/'):
                        file_path = file_path[7:]  # إزالة "static/"
                    attendance.sick_leave_file = file_path
        elif status != 'sick':
            # إذا تم تغيير الحالة من مرضي إلى حالة أخرى، نزيل المرجع من قاعدة البيانات
            # 💾 لا حذف للملفات الفعلية - الاحتفاظ بجميع الملفات للأمان
            if attendance.sick_leave_file:
                attendance.sick_leave_file = None
        
        # معالجة أوقات الدخول والخروج
        if status == 'present':
            if check_in_str:
                try:
                    hours, minutes = map(int, check_in_str.split(':'))
                    attendance.check_in = dt_time(hours, minutes)
                except:
                    attendance.check_in = None
            
            if check_out_str:
                try:
                    hours, minutes = map(int, check_out_str.split(':'))
                    attendance.check_out = dt_time(hours, minutes)
                except:
                    attendance.check_out = None
        else:
            # إذا لم تكن الحالة حاضر، نحذف أوقات الدخول والخروج
            attendance.check_in = None
            attendance.check_out = None
        
        # حفظ التغييرات
        db.session.commit()
        
        # تسجيل العملية في سجل النشاط
        log_attendance_activity(
            action='update',
            attendance_data={
                'id': attendance.id,
                'employee_id': attendance.employee_id,
                'date': attendance.date.isoformat(),
                'old_status': old_status,
                'new_status': status
            },
            employee_name=attendance.employee.name
        )
        
        flash('تم تحديث سجل الحضور بنجاح', 'success')
        
        # العودة لصفحة العرض مع المعاملات
        department_id = request.args.get('department_id', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        return redirect(url_for('attendance.department_attendance_view', 
                               department_id=department_id,
                               start_date=start_date,
                               end_date=end_date))
        
    except Exception as e:
        db.session.rollback()
        print(f"خطأ في تحديث سجل الحضور: {str(e)}")
        flash(f'حدث خطأ: {str(e)}', 'danger')
        return redirect(url_for('attendance.edit_attendance_page', id=id))

@attendance_bp.route('/departments-circles-overview')
@login_required
def departments_circles_overview():
    """لوحة تحكم شاملة تعرض الأقسام والدوائر وبيانات الحضور مع فلاتر"""
    date_str = request.args.get('date')
    department_filter = request.args.get('department_id')
    
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # حساب التاريخ: من اليوم الحالي أو التاريخ المختار
    # نجلب بيانات الحضور من آخر 18 ساعة من الصباح (6 صباحاً)
    now = datetime.now()
    today_date = datetime.now().date()
    eighteen_hours_ago_date = (now - timedelta(hours=18)).date()
    
    # نطاق التواريخ المراد البحث فيها
    start_date = min(selected_date, eighteen_hours_ago_date)
    end_date = today_date
    
    all_departments = Department.query.order_by(Department.name).all()
    
    if department_filter:
        try:
            departments = [Department.query.get(int(department_filter))]
            if not departments[0]:
                departments = all_departments
        except (ValueError, TypeError):
            departments = all_departments
    else:
        departments = all_departments
    
    departments_data = []
    
    for dept in departments:
        # جلب جميع الموظفين النشطين في القسم
        active_employees = [emp for emp in dept.employees if emp.status == 'active']
        
        # تجميع الموظفين حسب الدوائر الجغرافية - مع تنظيف المسافات الزائدة
        locations_dict = {}
        employees_without_location = []
        
        for emp in active_employees:
            if emp.location:
                # تنظيف المسافات وتوحيد الأسماء
                location_name = emp.location.strip()
                if location_name not in locations_dict:
                    locations_dict[location_name] = []
                locations_dict[location_name].append(emp)
            else:
                employees_without_location.append(emp)
        
        circles_data = []
        total_dept_present = 0
        total_dept_absent = 0
        total_dept_leave = 0
        total_dept_sick = 0
        
        # إذا لم تكن هناك دوائر جغرافية
        if not locations_dict:
            # إذا تم اختيار قسم معين، نعرضه مع رسالة أنه لا توجد دوائر
            if department_filter:
                departments_data.append({
                    'name': dept.name,
                    'id': dept.id,
                    'total_employees': len(active_employees),
                    'total_present': 0,
                    'total_absent': 0,
                    'total_leave': 0,
                    'total_sick': 0,
                    'circles': [],
                    'no_circles': True
                })
            # وإذا كان عرض جميع الأقسام، نتخطى هذا القسم
            continue
        
        # معالجة كل دائرة جغرافية
        for location in sorted(locations_dict.keys()):
            emp_in_circle = locations_dict[location]
            emp_ids = [e.id for e in emp_in_circle]
            
            if emp_ids:
                # جلب بيانات الحضور لهذه الدائرة من آخر 18 ساعة
                present = db.session.query(func.count(Attendance.id)).filter(
                    Attendance.employee_id.in_(emp_ids),
                    Attendance.date >= start_date,
                    Attendance.date <= end_date,
                    Attendance.status == 'present'
                ).scalar() or 0
                
                absent = db.session.query(func.count(Attendance.id)).filter(
                    Attendance.employee_id.in_(emp_ids),
                    Attendance.date >= start_date,
                    Attendance.date <= end_date,
                    Attendance.status == 'absent'
                ).scalar() or 0
                
                leave = db.session.query(func.count(Attendance.id)).filter(
                    Attendance.employee_id.in_(emp_ids),
                    Attendance.date >= start_date,
                    Attendance.date <= end_date,
                    Attendance.status == 'leave'
                ).scalar() or 0
                
                sick = db.session.query(func.count(Attendance.id)).filter(
                    Attendance.employee_id.in_(emp_ids),
                    Attendance.date >= start_date,
                    Attendance.date <= end_date,
                    Attendance.status == 'sick'
                ).scalar() or 0
                
                not_registered = len(emp_ids) - (present + absent + leave + sick)
            else:
                present = absent = leave = sick = not_registered = 0
            
            total_dept_present += present
            total_dept_absent += absent
            total_dept_leave += leave
            total_dept_sick += sick
            
            # جلب تفاصيل الموظفين في هذه الدائرة (آخر سجل خلال 18 ساعة)
            employees_details = []
            accessed_count = 0
            accessed_employees = []
            
            for emp in emp_in_circle:
                attendance = Attendance.query.filter(
                    Attendance.employee_id == emp.id,
                    Attendance.date >= start_date,
                    Attendance.date <= end_date
                ).order_by(Attendance.date.desc()).first()
                
                # جلب بيانات التتبع الجغرافي (GPS)
                emp_location = db.session.query(EmployeeLocation).filter(
                    EmployeeLocation.employee_id == emp.id,
                    EmployeeLocation.recorded_at >= datetime.combine(start_date, time(0, 0, 0)),
                    EmployeeLocation.recorded_at <= datetime.combine(end_date, time(23, 59, 59))
                ).order_by(EmployeeLocation.recorded_at.desc()).first()
                
                # جلب بيانات الوصول إلى الدوائر (GeofenceSession)
                geo_session = db.session.query(GeofenceSession).filter(
                    GeofenceSession.employee_id == emp.id,
                    GeofenceSession.entry_time >= datetime.combine(start_date, time(0, 0, 0)),
                    GeofenceSession.entry_time <= datetime.combine(end_date, time(23, 59, 59))
                ).order_by(GeofenceSession.entry_time.desc()).first()
                
                # تحديد ما إذا كان الموظف قد دخل الدائرة
                accessed = geo_session is not None
                if accessed:
                    accessed_count += 1
                    accessed_employees.append(emp.name)
                
                duration_minutes = geo_session.duration_minutes if geo_session and geo_session.duration_minutes else 0
                
                emp_data = {
                    'id': emp.id,
                    'name': emp.name,
                    'employee_id': emp.employee_id,
                    'status': attendance.status if attendance else 'لم يتم التسجيل',
                    'check_in': attendance.check_in.strftime('%H:%M') if attendance and attendance.check_in else '-',
                    'check_out': attendance.check_out.strftime('%H:%M') if attendance and attendance.check_out else '-',
                    # بيانات التتبع الجغرافي
                    'gps_latitude': emp_location.latitude if emp_location else None,
                    'gps_longitude': emp_location.longitude if emp_location else None,
                    'gps_recorded_at': emp_location.recorded_at if emp_location else None,
                    # بيانات الوصول إلى الدائرة
                    'accessed_circle': accessed,
                    'circle_enter_time': geo_session.entry_time.strftime('%H:%M:%S') if geo_session and geo_session.entry_time else None,
                    'circle_exit_time': geo_session.exit_time.strftime('%H:%M:%S') if geo_session and geo_session.exit_time else None,
                    'duration_minutes': duration_minutes,
                    'duration_display': f'{duration_minutes // 60}س {duration_minutes % 60}د' if duration_minutes > 0 else '-',
                }
                employees_details.append(emp_data)
            
            # إضافة بيانات الدائرة مع معلومات الوصول
            circles_data.append({
                'name': location,
                'total': len(emp_in_circle),
                'present': present,
                'absent': absent,
                'leave': leave,
                'sick': sick,
                'not_registered': not_registered,
                'accessed_count': accessed_count,
                'accessed_employees': ', '.join(accessed_employees) if accessed_employees else 'لا أحد',
                'employees': employees_details
            })
        
        # معالجة الموظفين بدون دائرة جغرافية محددة
        if employees_without_location:
            emp_ids = [e.id for e in employees_without_location]
            present = db.session.query(func.count(Attendance.id)).filter(
                Attendance.employee_id.in_(emp_ids),
                Attendance.date >= start_date,
                Attendance.date <= end_date,
                Attendance.status == 'present'
            ).scalar() or 0
            
            absent = db.session.query(func.count(Attendance.id)).filter(
                Attendance.employee_id.in_(emp_ids),
                Attendance.date >= start_date,
                Attendance.date <= end_date,
                Attendance.status == 'absent'
            ).scalar() or 0
            
            leave = db.session.query(func.count(Attendance.id)).filter(
                Attendance.employee_id.in_(emp_ids),
                Attendance.date >= start_date,
                Attendance.date <= end_date,
                Attendance.status == 'leave'
            ).scalar() or 0
            
            sick = db.session.query(func.count(Attendance.id)).filter(
                Attendance.employee_id.in_(emp_ids),
                Attendance.date >= start_date,
                Attendance.date <= end_date,
                Attendance.status == 'sick'
            ).scalar() or 0
            
            not_registered = len(emp_ids) - (present + absent + leave + sick)
            
            total_dept_present += present
            total_dept_absent += absent
            total_dept_leave += leave
            total_dept_sick += sick
            
            employees_details = []
            for emp in employees_without_location:
                attendance = Attendance.query.filter(
                    Attendance.employee_id == emp.id,
                    Attendance.date >= start_date,
                    Attendance.date <= end_date
                ).order_by(Attendance.date.desc()).first()
                
                emp_data = {
                    'name': emp.name,
                    'employee_id': emp.employee_id,
                    'status': attendance.status if attendance else 'لم يتم التسجيل',
                    'check_in': attendance.check_in.strftime('%H:%M') if attendance and attendance.check_in else '-',
                    'check_out': attendance.check_out.strftime('%H:%M') if attendance and attendance.check_out else '-',
                }
                employees_details.append(emp_data)
            
            circles_data.append({
                'name': '🔵 بدون دائرة محددة',
                'total': len(employees_without_location),
                'present': present,
                'absent': absent,
                'leave': leave,
                'sick': sick,
                'not_registered': not_registered,
                'employees': employees_details
            })
        
        # إضافة بيانات القسم
        departments_data.append({
            'name': dept.name,
            'id': dept.id,
            'total_employees': len(active_employees),
            'total_present': total_dept_present,
            'total_absent': total_dept_absent,
            'total_leave': total_dept_leave,
            'total_sick': total_dept_sick,
            'circles': circles_data
        })
    
    # جلب بيانات التتبع الجغرافي للموظفين
    from datetime import datetime as dt
    from sqlalchemy import and_
    
    # جلب آخر موقع لكل موظف نشط
    from sqlalchemy.orm import aliased
    from sqlalchemy import func as sql_func
    
    all_active_emp_ids = [e_data['id'] for dept in departments_data for circle in dept['circles'] for e_data in circle['employees']]
    
    # بناء dictionary لبيانات التتبع
    locations_by_employee = {}
    if all_active_emp_ids:
        latest_locations_subq = db.session.query(
            EmployeeLocation.employee_id,
            EmployeeLocation.id.label('location_id'),
            sql_func.row_number().over(
                partition_by=EmployeeLocation.employee_id,
                order_by=EmployeeLocation.recorded_at.desc()
            ).label('rn')
        ).filter(
            EmployeeLocation.employee_id.in_(all_active_emp_ids)
        ).subquery()
        
        latest_locations = db.session.query(EmployeeLocation).join(
            latest_locations_subq,
            and_(
                EmployeeLocation.id == latest_locations_subq.c.location_id,
                latest_locations_subq.c.rn == 1
            )
        ).all()
        
        for loc in latest_locations:
            locations_by_employee[loc.employee_id] = {
                'latitude': loc.latitude,
                'longitude': loc.longitude,
                'recorded_at': loc.recorded_at
            }
    
    # إذا تم اختيار قسم معين، تمرير فقط هذا القسم في all_departments للفلتر
    departments_for_filter = departments if department_filter else all_departments
    
    return render_template(
        'attendance/departments_circles_overview.html',
        departments_data=departments_data,
        all_departments=all_departments,
        selected_date=selected_date,
        selected_date_formatted=format_date_hijri(selected_date),
        selected_department_id=int(department_filter) if department_filter else None,
        locations_by_employee=locations_by_employee
    )

@attendance_bp.route('/circle-accessed-details/<int:department_id>/<circle_name>')
@login_required
def circle_accessed_details(department_id, circle_name):
    """صفحة منفصلة لعرض تفاصيل الموظفين الواصلين للدائرة"""
    date_str = request.args.get('date')
    
    # منطقة زمنية السعودية (UTC+3)
    saudi_tz = timedelta(hours=3)
    
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # حساب نطاق التواريخ - من التاريخ المختار إلى اليوم
    start_date = selected_date
    end_date = datetime.now().date()
    
    dept = Department.query.get(department_id)
    if not dept:
        flash('القسم غير موجود', 'danger')
        return redirect(url_for('attendance.departments_circles_overview'))
    
    # جلب الموظفين النشطين في القسم والدائرة المحددة
    active_employees = [emp for emp in dept.employees if emp.status == 'active' and emp.location and emp.location.strip() == circle_name]
    
    employees_accessed = []
    
    for emp in active_employees:
        # جلب جميع جلسات الدائرة الجغرافية (GeofenceSession)
        geo_sessions = db.session.query(GeofenceSession).filter(
            GeofenceSession.employee_id == emp.id,
            GeofenceSession.entry_time >= datetime.combine(start_date, time(0, 0, 0)),
            GeofenceSession.entry_time <= datetime.combine(end_date, time(23, 59, 59))
        ).order_by(GeofenceSession.entry_time.asc()).all()
        
        if geo_sessions:
            emp_location = db.session.query(EmployeeLocation).filter(
                EmployeeLocation.employee_id == emp.id,
                EmployeeLocation.recorded_at >= datetime.combine(start_date, time(0, 0, 0)),
                EmployeeLocation.recorded_at <= datetime.combine(end_date, time(23, 59, 59))
            ).order_by(EmployeeLocation.recorded_at.desc()).first()
            
            # استخراج دخول وخروج الصباح والمساء من GeofenceSession
            morning_check_in = None
            morning_check_out = None
            evening_check_in = None
            evening_check_out = None
            
            # معالجة جميع جلسات الدائرة
            for geo in geo_sessions:
                # استخراج أول دخول صباحي من الدائرة
                if geo.entry_time and not morning_check_in:
                    entry_sa = geo.entry_time + saudi_tz
                    if entry_sa.hour < 12:
                        morning_check_in = entry_sa
                
                # استخراج أول دخول مسائي من الدائرة
                if geo.entry_time and not evening_check_in:
                    entry_sa = geo.entry_time + saudi_tz
                    if entry_sa.hour >= 12:
                        evening_check_in = entry_sa
                
                # استخراج أخر خروج صباحي من الدائرة
                if geo.exit_time:
                    exit_sa = geo.exit_time + saudi_tz
                    if exit_sa.hour < 12:
                        morning_check_out = exit_sa
                
                # استخراج أخر خروج مسائي من الدائرة
                if geo.exit_time:
                    exit_sa = geo.exit_time + saudi_tz
                    if exit_sa.hour >= 12:
                        evening_check_out = exit_sa
            
            # جلب آخر جلسة للحصول على البيانات الأساسية
            latest_geo = geo_sessions[-1]
            
            # جلب سجلات الحضور (للحالة فقط)
            attendance = Attendance.query.filter(
                Attendance.employee_id == emp.id,
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).order_by(Attendance.date.desc()).first()
            
            duration_minutes = latest_geo.duration_minutes if latest_geo.duration_minutes else 0
            
            # تحويل الأوقات إلى توقيت السعودية
            entry_time_sa = (latest_geo.entry_time + saudi_tz) if latest_geo.entry_time else None
            exit_time_sa = (latest_geo.exit_time + saudi_tz) if latest_geo.exit_time else None
            
            # حساب ساعات العمل من الدخول الصباحي إلى الخروج المسائي
            work_hours = 0
            work_minutes = 0
            if morning_check_in and evening_check_out:
                total_seconds = (evening_check_out - morning_check_in).total_seconds()
                if total_seconds > 0:
                    work_hours = int(total_seconds // 3600)
                    work_minutes = int((total_seconds % 3600) // 60)
            
            employees_accessed.append({
                'name': emp.name,
                'employee_id': emp.employee_id,
                'status': 'حضور' if (attendance and attendance.status) else 'حضور من الدائرة',
                'check_in': '-',
                'check_out': '-',
                'morning_check_in': format_time_12h_ar_short(morning_check_in) if morning_check_in else '-',
                'morning_check_in_hours': morning_check_in.hour if morning_check_in else None,
                'morning_check_in_minutes': morning_check_in.minute if morning_check_in else None,
                'morning_check_out': format_time_12h_ar_short(morning_check_out) if morning_check_out else '-',
                'evening_check_in': format_time_12h_ar_short(evening_check_in) if evening_check_in else '-',
                'evening_check_out': format_time_12h_ar_short(evening_check_out) if evening_check_out else '-',
                'evening_check_out_hours': evening_check_out.hour if evening_check_out else None,
                'evening_check_out_minutes': evening_check_out.minute if evening_check_out else None,
                'work_hours': work_hours,
                'work_minutes': work_minutes,
                'work_hours_display': f'{work_hours} س {work_minutes} د' if (work_hours > 0 or work_minutes > 0) else '-',
                'circle_enter_time': format_time_12h_ar(entry_time_sa) if entry_time_sa else None,
                'circle_exit_time': format_time_12h_ar(exit_time_sa) if exit_time_sa else None,
                'duration_minutes': duration_minutes,
                'duration_display': f'{duration_minutes // 60}س {duration_minutes % 60}د' if duration_minutes > 0 else '-',
                'gps_latitude': emp_location.latitude if emp_location else None,
                'gps_longitude': emp_location.longitude if emp_location else None,
                'profile_image': emp.profile_image,
            })
    
    return render_template(
        'attendance/circle_accessed_details.html',
        circle_name=circle_name,
        department_name=dept.name,
        department_id=department_id,
        employees_accessed=employees_accessed,
        selected_date=selected_date,
        selected_date_formatted=format_date_hijri(selected_date)
    )

@attendance_bp.route('/circle-accessed-details/<int:department_id>/<circle_name>/export-excel')
@login_required
def export_circle_details_excel(department_id, circle_name):
    """تصدير تفاصيل الموظفين الواصلين للدائرة إلى ملف Excel"""
    date_str = request.args.get('date')
    export_type = request.args.get('type', 'all')  # 'all' أو employee_id
    employee_id_filter = request.args.get('employee_id', None)
    
    # منطقة زمنية السعودية (UTC+3)
    saudi_tz = timedelta(hours=3)
    
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # حساب نطاق التواريخ - من التاريخ المختار إلى اليوم
    start_date = selected_date
    end_date = datetime.now().date()
    
    dept = Department.query.get(department_id)
    if not dept:
        flash('القسم غير موجود', 'danger')
        return redirect(url_for('attendance.departments_circles_overview'))
    
    # جلب الموظفين النشطين في القسم والدائرة المحددة
    active_employees = [emp for emp in dept.employees if emp.status == 'active' and emp.location and emp.location.strip() == circle_name]
    
    # إذا كان هناك فلتر موظف معين
    if export_type == 'single' and employee_id_filter:
        try:
            active_employees = [emp for emp in active_employees if emp.employee_id == str(employee_id_filter)]
        except:
            pass
    
    data = []
    
    for emp in active_employees:
        # جلب جميع جلسات الدائرة الجغرافية (GeofenceSession)
        geo_sessions = db.session.query(GeofenceSession).filter(
            GeofenceSession.employee_id == emp.id,
            GeofenceSession.entry_time >= datetime.combine(start_date, time(0, 0, 0)),
            GeofenceSession.entry_time <= datetime.combine(end_date, time(23, 59, 59))
        ).order_by(GeofenceSession.entry_time.asc()).all()
        
        if geo_sessions:
            # استخراج دخول وخروج الصباح والمساء من GeofenceSession
            morning_check_in = None
            morning_check_out = None
            evening_check_in = None
            evening_check_out = None
            
            # معالجة جميع جلسات الدائرة
            for geo in geo_sessions:
                # استخراج أول دخول صباحي من الدائرة
                if geo.entry_time and not morning_check_in:
                    entry_sa = geo.entry_time + saudi_tz
                    if entry_sa.hour < 12:
                        morning_check_in = entry_sa
                
                # استخراج أول دخول مسائي من الدائرة
                if geo.entry_time and not evening_check_in:
                    entry_sa = geo.entry_time + saudi_tz
                    if entry_sa.hour >= 12:
                        evening_check_in = entry_sa
                
                # استخراج أخر خروج صباحي من الدائرة
                if geo.exit_time:
                    exit_sa = geo.exit_time + saudi_tz
                    if exit_sa.hour < 12:
                        morning_check_out = exit_sa
                
                # استخراج أخر خروج مسائي من الدائرة
                if geo.exit_time:
                    exit_sa = geo.exit_time + saudi_tz
                    if exit_sa.hour >= 12:
                        evening_check_out = exit_sa
            
            # جلب آخر جلسة للحصول على البيانات الأساسية
            latest_geo = geo_sessions[-1]
            
            # جلب آخر موقع للموظف
            emp_location = db.session.query(EmployeeLocation).filter(
                EmployeeLocation.employee_id == emp.id,
                EmployeeLocation.recorded_at >= datetime.combine(start_date, time(0, 0, 0)),
                EmployeeLocation.recorded_at <= datetime.combine(end_date, time(23, 59, 59))
            ).order_by(EmployeeLocation.recorded_at.desc()).first()
            
            # جلب آخر سجل حضور (للحالة فقط)
            attendance = Attendance.query.filter(
                Attendance.employee_id == emp.id,
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).order_by(Attendance.date.desc()).first()
            
            duration_minutes = latest_geo.duration_minutes if latest_geo.duration_minutes else 0
            
            # تحويل الأوقات إلى توقيت السعودية
            entry_time_sa = (latest_geo.entry_time + saudi_tz) if latest_geo.entry_time else None
            exit_time_sa = (latest_geo.exit_time + saudi_tz) if latest_geo.exit_time else None
            
            data.append({
                'اسم الموظف': emp.name,
                'رقم الموظف': emp.employee_id,
                'رقم الهوية': emp.national_id,
                'رقم الجوال': emp.mobile,
                'الوظيفة': emp.job_title,
                'حالة الحضور': 'حضور من الدائرة' if geo_sessions else 'لم يتم التسجيل',
                'دخول الدائرة': format_time_12h_ar(entry_time_sa) if entry_time_sa else '-',
                'خروج الدائرة': format_time_12h_ar(exit_time_sa) if exit_time_sa else '-',
                'المدة بالدقائق': duration_minutes,
                'المدة (ساعات ودقائق)': f'{duration_minutes // 60}س {duration_minutes % 60}د' if duration_minutes > 0 else '-',
                'دخول صباحي': format_time_12h_ar_short(morning_check_in) if morning_check_in else '-',
                'خروج صباحي': format_time_12h_ar_short(morning_check_out) if morning_check_out else '-',
                'دخول مسائي': format_time_12h_ar_short(evening_check_in) if evening_check_in else '-',
                'خروج مسائي': format_time_12h_ar_short(evening_check_out) if evening_check_out else '-',
                'GPS - Latitude': emp_location.latitude if emp_location else '-',
                'GPS - Longitude': emp_location.longitude if emp_location else '-',
            })
    
    if not data:
        flash('لا توجد بيانات للتصدير', 'warning')
        return redirect(url_for('attendance.circle_accessed_details', department_id=department_id, circle_name=circle_name, date=date_str))
    
    # إنشاء DataFrame
    df = pd.DataFrame(data)
    
    # إنشاء ملف Excel في الذاكرة
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='البيانات', index=False, startrow=0)
        
        # تنسيق الـ worksheet
        worksheet = writer.sheets['البيانات']
        worksheet.sheet_properties.orientation = 'rtl'
        
        # تعديل عرض الأعمدة
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # اسم الملف
    if export_type == 'single' and employee_id_filter:
        emp = next((e for e in active_employees if e.employee_id == str(employee_id_filter)), None)
        filename = f'تفاصيل_الموظف_{emp.name if emp else "unknown"}_{selected_date.strftime("%Y-%m-%d")}.xlsx'
    else:
        filename = f'تفاصيل_الدائرة_{circle_name}_{selected_date.strftime("%Y-%m-%d")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@attendance_bp.route('/mark-circle-employees-attendance/<int:department_id>/<circle_name>', methods=['POST'])
@login_required
def mark_circle_employees_attendance(department_id, circle_name):
    """تسجيل الموظفين الواصلين للدائرة كحاضرين مع أوقات الدخول والخروج"""
    try:
        date_str = request.args.get('date')
        saudi_tz = timedelta(hours=3)
        
        try:
            if date_str:
                selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                selected_date = datetime.now().date()
        except ValueError:
            selected_date = datetime.now().date()
        
        dept = Department.query.get(department_id)
        if not dept:
            return jsonify({'success': False, 'message': 'القسم غير موجود'}), 404
        
        # جلب الموظفين النشطين في القسم والدائرة المحددة
        active_employees = [emp for emp in dept.employees if emp.status == 'active' and emp.location and emp.location.strip() == circle_name]
        
        # جلب جميع جلسات الدائرة من التاريخ المختار إلى اليوم
        start_date = selected_date
        end_date = datetime.now().date()
        
        marked_count = 0
        
        for emp in active_employees:
            # جلب جميع جلسات الدائرة الجغرافية لهذا الموظف
            geo_sessions = db.session.query(GeofenceSession).filter(
                GeofenceSession.employee_id == emp.id,
                GeofenceSession.entry_time >= datetime.combine(start_date, time(0, 0, 0)),
                GeofenceSession.entry_time <= datetime.combine(end_date, time(23, 59, 59))
            ).order_by(GeofenceSession.entry_time.asc()).all()
            
            if geo_sessions:
                # استخراج أول دخول صباحي وآخر خروج مسائي
                morning_check_in = None
                evening_check_out = None
                
                for geo in geo_sessions:
                    # أول دخول صباحي
                    if geo.entry_time and not morning_check_in:
                        entry_sa = geo.entry_time + saudi_tz
                        if entry_sa.hour < 12:
                            morning_check_in = entry_sa  # احفظ الوقت بعد التحويل للتوقيت السعودي
                    
                    # آخر خروج مسائي
                    if geo.exit_time:
                        exit_sa = geo.exit_time + saudi_tz
                        if exit_sa.hour >= 12:
                            evening_check_out = exit_sa  # احفظ الوقت بعد التحويل للتوقيت السعودي
                
                # إذا لم نجد دخول صباحي، نأخذ أول دخول عام + تحويل للتوقيت السعودي
                if not morning_check_in and geo_sessions:
                    morning_check_in = geo_sessions[0].entry_time + saudi_tz
                
                # إذا لم نجد خروج مسائي، نأخذ آخر خروج عام + تحويل للتوقيت السعودي
                if not evening_check_out and geo_sessions:
                    evening_check_out = geo_sessions[-1].exit_time + saudi_tz
                
                # حفظ الأوقات بعد تحويلها للتوقيت السعودي
                check_in_time = None
                check_out_time = None
                
                if morning_check_in:
                    if isinstance(morning_check_in, datetime):
                        check_in_time = morning_check_in.time()
                    else:
                        check_in_time = morning_check_in
                
                if evening_check_out:
                    if isinstance(evening_check_out, datetime):
                        check_out_time = evening_check_out.time()
                    else:
                        check_out_time = evening_check_out
                
                # تحديث أو إنشاء سجل الحضور
                existing_attendance = Attendance.query.filter(
                    Attendance.employee_id == emp.id,
                    Attendance.date == selected_date
                ).first()
                
                if existing_attendance:
                    # تحديث السجل الموجود
                    if check_in_time:
                        existing_attendance.check_in = check_in_time
                    if check_out_time:
                        existing_attendance.check_out = check_out_time
                    existing_attendance.status = 'present'
                    existing_attendance.updated_at = datetime.utcnow()
                else:
                    # إنشاء سجل حضور جديد
                    attendance = Attendance(
                        employee_id=emp.id,
                        date=selected_date,
                        status='present',
                        check_in=check_in_time,
                        check_out=check_out_time,
                    )
                    db.session.add(attendance)
                
                marked_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم تسجيل {marked_count} موظف كحاضرين مع أوقات الدخول والخروج',
            'count': marked_count
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"خطأ في تسجيل الحضور: {str(e)}")
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'}), 500