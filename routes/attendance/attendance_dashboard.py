"""
وحدة لوحة معلومات الحضور المحسنة
-----------------------------
تعرض إحصائيات الحضور حسب القسم والتاريخ مع إمكانية تصدير البيانات
"""

from datetime import datetime, timedelta
import pandas as pd
from sqlalchemy import func, case, literal, or_
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
import io
import os
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

from models import Department, Employee, Attendance, Module, employee_departments
from core.extensions import db
from utils.decorators import module_access_required
from utils.date_converter import format_date_hijri, format_date_gregorian

# إنشاء blueprint للوحة معلومات الحضور
attendance_dashboard_bp = Blueprint('attendance_dashboard', __name__)

@attendance_dashboard_bp.route('/departments-circles-overview')
@login_required
@module_access_required(Module.ATTENDANCE)
def departments_circles_overview():
    """لوحة تحكم شاملة تعرض الأقسام والدوائر وبيانات الحضور مع فلاتر"""
    date_str = request.args.get('date')
    department_filter = request.args.get('department_id')  # None = جميع الأقسام
    
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # جمع جميع الأقسام للفلتر
    all_departments = Department.query.order_by(Department.name).all()
    
    # تحديد الأقسام المراد عرضها بناءً على الفلتر
    if department_filter:
        try:
            departments = [Department.query.get(int(department_filter))]
        except (ValueError, TypeError):
            departments = all_departments
    else:
        departments = all_departments
    
    departments_data = []
    
    for dept in departments:
        # الموظفين النشطين في القسم
        active_employees = [emp for emp in dept.employees if emp.status == 'active']
        
        # جميع الدوائر المختلفة في هذا القسم
        locations = set()
        for emp in active_employees:
            if emp.location:
                locations.add(emp.location)
        
        circles_data = []
        total_dept_present = 0
        total_dept_absent = 0
        total_dept_leave = 0
        total_dept_sick = 0
        
        for location in sorted(locations):
            # الموظفين في هذه الدائرة من هذا القسم
            emp_in_circle = [e for e in active_employees if e.location == location]
            emp_ids = [e.id for e in emp_in_circle]
            
            if emp_ids:
                # حسابات الحضور
                present = db.session.query(func.count(Attendance.id)).filter(
                    Attendance.employee_id.in_(emp_ids),
                    Attendance.date == selected_date,
                    Attendance.status == 'present'
                ).scalar() or 0
                
                absent = db.session.query(func.count(Attendance.id)).filter(
                    Attendance.employee_id.in_(emp_ids),
                    Attendance.date == selected_date,
                    Attendance.status == 'absent'
                ).scalar() or 0
                
                leave = db.session.query(func.count(Attendance.id)).filter(
                    Attendance.employee_id.in_(emp_ids),
                    Attendance.date == selected_date,
                    Attendance.status == 'leave'
                ).scalar() or 0
                
                sick = db.session.query(func.count(Attendance.id)).filter(
                    Attendance.employee_id.in_(emp_ids),
                    Attendance.date == selected_date,
                    Attendance.status == 'sick'
                ).scalar() or 0
                
                not_registered = len(emp_ids) - (present + absent + leave + sick)
            else:
                present = absent = leave = sick = not_registered = 0
            
            total_dept_present += present
            total_dept_absent += absent
            total_dept_leave += leave
            total_dept_sick += sick
            
            # تفاصيل الموظفين في هذه الدائرة
            employees_details = []
            for emp in emp_in_circle:
                attendance = Attendance.query.filter_by(
                    employee_id=emp.id,
                    date=selected_date
                ).first()
                
                emp_data = {
                    'name': emp.name,
                    'employee_id': emp.employee_id,
                    'status': attendance.status if attendance else 'لم يتم التسجيل',
                    'check_in': attendance.check_in.strftime('%H:%M') if attendance and attendance.check_in else '-',
                    'check_out': attendance.check_out.strftime('%H:%M') if attendance and attendance.check_out else '-',
                }
                employees_details.append(emp_data)
            
            circles_data.append({
                'name': location,
                'total': len(emp_in_circle),
                'present': present,
                'absent': absent,
                'leave': leave,
                'sick': sick,
                'not_registered': not_registered,
                'employees': employees_details
            })
        
        departments_data.append({
            'name': dept.name,
            'id': dept.id,
            'total_employees': len(active_employees),
            'total_present': total_dept_present,
            'total_absent': total_dept_absent,
            'total_leave': total_dept_leave,
            'total_sick': total_dept_sick,
            'circles': circles_data
        })
    
    return render_template(
        'attendance/departments_circles_overview.html',
        departments_data=departments_data,
        all_departments=all_departments,
        selected_date=selected_date,
        selected_date_formatted=format_date_hijri(selected_date),
        selected_department_id=int(department_filter) if department_filter else None
    )

@attendance_dashboard_bp.route('/')
@login_required
@module_access_required(Module.ATTENDANCE)
def index():
    """صفحة لوحة معلومات الحضور الرئيسية"""
    # التحقق من التاريخ المحدد أو استخدام اليوم الحالي
    date_str = request.args.get('date')
    department_id = request.args.get('department_id')
    
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # الحصول على جميع الأقسام
    departments = Department.query.order_by(Department.name).all()
    
    # إحصائيات إجمالية - عدد الموظفين النشطين فقط
    total_employees = Employee.query.filter_by(status='active').count()
    
    # حساب إحصائيات الحضور الإجمالية
    total_present = db.session.query(func.count(Attendance.id)).join(
        Employee, Employee.id == Attendance.employee_id
    ).filter(
        Employee.status == 'active',
        Attendance.date == selected_date,
        Attendance.status == 'present'
    ).scalar() or 0
    
    total_absent = db.session.query(func.count(Attendance.id)).join(
        Employee, Employee.id == Attendance.employee_id
    ).filter(
        Employee.status == 'active',
        Attendance.date == selected_date,
        Attendance.status == 'absent'
    ).scalar() or 0
    
    total_leave = db.session.query(func.count(Attendance.id)).join(
        Employee, Employee.id == Attendance.employee_id
    ).filter(
        Employee.status == 'active',
        Attendance.date == selected_date,
        Attendance.status == 'leave'
    ).scalar() or 0
    
    total_sick = db.session.query(func.count(Attendance.id)).join(
        Employee, Employee.id == Attendance.employee_id
    ).filter(
        Employee.status == 'active',
        Attendance.date == selected_date,
        Attendance.status == 'sick'
    ).scalar() or 0
    
    # حساب عدد الموظفين الذين لم يتم تسجيل حضورهم
    total_registered = total_present + total_absent + total_leave + total_sick
    total_missing = total_employees - total_registered if total_registered <= total_employees else 0
    
    # إحصائيات حسب القسم
    department_stats = []
    
    # جمع إحصائيات كل قسم
    for dept in departments:
        # عدد الموظفين في القسم باستخدام علاقة many-to-many
        active_employees = [emp for emp in dept.employees if emp.status == 'active']
        employees_count = len(active_employees)
        employee_ids = [emp.id for emp in active_employees]
        
        # عدد الحاضرين
        if employee_ids:
            present_count = db.session.query(func.count(Attendance.id)).filter(
                Attendance.employee_id.in_(employee_ids),
                Attendance.date == selected_date,
                Attendance.status == 'present'
            ).scalar() or 0
        else:
            present_count = 0
        
        # عدد الغائبين
        if employee_ids:
            absent_count = db.session.query(func.count(Attendance.id)).filter(
                Attendance.employee_id.in_(employee_ids),
                Attendance.date == selected_date,
                Attendance.status == 'absent'
            ).scalar() or 0
        else:
            absent_count = 0
        
        # عدد الإجازات
        if employee_ids:
            leave_count = db.session.query(func.count(Attendance.id)).filter(
                Attendance.employee_id.in_(employee_ids),
                Attendance.date == selected_date,
                Attendance.status == 'leave'
            ).scalar() or 0
        else:
            leave_count = 0
        
        # عدد المرضي
        if employee_ids:
            sick_count = db.session.query(func.count(Attendance.id)).filter(
                Attendance.employee_id.in_(employee_ids),
                Attendance.date == selected_date,
                Attendance.status == 'sick'
            ).scalar() or 0
        else:
            sick_count = 0
        
        # إجمالي السجلات
        total_records = present_count + absent_count + leave_count + sick_count
        
        # حساب نسب الحضور والغياب
        if employees_count > 0:
            present_percentage = round((present_count / employees_count) * 100, 1)
            absent_percentage = round((absent_count / employees_count) * 100, 1)
            missing_count = employees_count - total_records
            missing_percentage = round((missing_count / employees_count) * 100, 1) if missing_count > 0 else 0
        else:
            present_percentage = 0
            absent_percentage = 0
            missing_count = 0
            missing_percentage = 0
        
        # إضافة إحصائيات القسم للقائمة
        department_stats.append({
            'id': dept.id,
            'name': dept.name,
            'employees_count': employees_count,
            'present_count': present_count,
            'absent_count': absent_count,
            'leave_count': leave_count,
            'sick_count': sick_count,
            'missing_count': missing_count,
            'present_percentage': present_percentage,
            'absent_percentage': absent_percentage,
            'missing_percentage': missing_percentage
        })
    
    # الموظفون الغائبون
    absent_employees = []
    
    # إذا تم تحديد قسم معين، احصل على معرفات الموظفين النشطين في ذلك القسم
    if department_id:
        selected_dept = Department.query.get(int(department_id))
        if selected_dept:
            dept_employee_ids = [emp.id for emp in selected_dept.employees if emp.status == 'active']
            
            # استعلام الموظفين الغائبين في القسم المحدد فقط
            absent_records = db.session.query(Attendance, Employee).join(
                Employee, Employee.id == Attendance.employee_id
            ).filter(
                Attendance.date == selected_date,
                Attendance.status.in_(['absent', 'leave', 'sick']),
                Attendance.employee_id.in_(dept_employee_ids)
            ).all()
            
            # تجهيز بيانات الموظفين الغائبين
            for attendance, employee in absent_records:
                # الحصول على أول قسم للموظف لعرضه
                emp_dept = employee.departments[0] if employee.departments else None
                if emp_dept:
                    absent_employees.append({
                        'id': employee.id,
                        'name': employee.name,
                        'employee_id': employee.employee_id,
                        'department': emp_dept.name,
                        'department_id': emp_dept.id,
                        'status': attendance.status,
                        'notes': attendance.notes
                    })
        else:
            absent_records = []
    else:
        # استعلام جميع الموظفين الغائبين النشطين
        absent_records = db.session.query(Attendance, Employee).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Attendance.date == selected_date,
            Attendance.status.in_(['absent', 'leave', 'sick']),
            Employee.status == 'active'
        ).all()
        
        # تجهيز بيانات الموظفين الغائبين
        for attendance, employee in absent_records:
            # الحصول على أول قسم للموظف لعرضه
            emp_dept = employee.departments[0] if employee.departments else None
            if emp_dept:
                absent_employees.append({
                    'id': employee.id,
                    'name': employee.name,
                    'employee_id': employee.employee_id,
                    'department': emp_dept.name,
                    'department_id': emp_dept.id,
                    'status': attendance.status,
                    'notes': attendance.notes
                })
    
    # تجهيز المعلومات حسب القسم
    absent_by_department = {}
    for emp in absent_employees:
        dept_id = emp['department_id']
        if dept_id not in absent_by_department:
            absent_by_department[dept_id] = {
                'name': emp['department'],
                'employees': []
            }
        absent_by_department[dept_id]['employees'].append(emp)
    
    # تنسيق التواريخ
    hijri_date = format_date_hijri(selected_date)
    gregorian_date = format_date_gregorian(selected_date)
    
    # تقديم الصفحة
    return render_template(
        'attendance/enhanced_dashboard.html',
        selected_date=selected_date,
        hijri_date=hijri_date,
        gregorian_date=gregorian_date,
        departments=departments,
        selected_department_id=int(department_id) if department_id else None,
        total_employees=total_employees,
        total_present=total_present,
        total_absent=total_absent,
        total_leave=total_leave,
        total_sick=total_sick,
        total_missing=total_missing,
        department_stats=department_stats,
        absent_employees=absent_employees,
        absent_by_department=absent_by_department
    )

@attendance_dashboard_bp.route('/data')
@login_required
@module_access_required(Module.ATTENDANCE)
def dashboard_data():
    """API لبيانات لوحة المعلومات"""
    date_str = request.args.get('date')
    
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # الحصول على بيانات الإحصائيات
    departments = Department.query.all()
    
    # بيانات للرسم البياني
    department_names = []
    present_counts = []
    absent_counts = []
    leave_counts = []
    sick_counts = []
    missing_counts = []
    
    for dept in departments:
        # إضافة اسم القسم
        department_names.append(dept.name)
        
        # عدد الموظفين في القسم
        employees_count = Employee.query.filter_by(
            department_id=dept.id, 
            status='active'
        ).count()
        
        # عدد الحاضرين
        present_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'present'
        ).scalar() or 0
        
        # عدد الغائبين
        absent_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'absent'
        ).scalar() or 0
        
        # عدد الإجازات
        leave_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'leave'
        ).scalar() or 0
        
        # عدد المرضي
        sick_count = db.session.query(func.count(Attendance.id)).join(
            Employee, Employee.id == Attendance.employee_id
        ).filter(
            Employee.department_id == dept.id,
            Employee.status == 'active',
            Attendance.date == selected_date,
            Attendance.status == 'sick'
        ).scalar() or 0
        
        # إجمالي السجلات
        total_records = present_count + absent_count + leave_count + sick_count
        
        # حساب عدد الموظفين بدون تسجيل
        missing_count = employees_count - total_records
        if missing_count < 0:
            missing_count = 0
        
        # إضافة الإحصائيات للقوائم
        present_counts.append(present_count)
        absent_counts.append(absent_count)
        leave_counts.append(leave_count)
        sick_counts.append(sick_count)
        missing_counts.append(missing_count)
    
    # تجهيز البيانات للرسم البياني
    chart_data = {
        'labels': department_names,
        'datasets': [
            {
                'label': 'حاضر',
                'data': present_counts,
                'backgroundColor': 'rgba(40, 167, 69, 0.7)',
                'borderColor': 'rgb(40, 167, 69)',
                'borderWidth': 1
            },
            {
                'label': 'غائب',
                'data': absent_counts,
                'backgroundColor': 'rgba(220, 53, 69, 0.7)',
                'borderColor': 'rgb(220, 53, 69)',
                'borderWidth': 1
            },
            {
                'label': 'إجازة',
                'data': leave_counts,
                'backgroundColor': 'rgba(255, 193, 7, 0.7)',
                'borderColor': 'rgb(255, 193, 7)',
                'borderWidth': 1
            },
            {
                'label': 'مرضي',
                'data': sick_counts,
                'backgroundColor': 'rgba(23, 162, 184, 0.7)',
                'borderColor': 'rgb(23, 162, 184)',
                'borderWidth': 1
            },
            {
                'label': 'غير مسجل',
                'data': missing_counts,
                'backgroundColor': 'rgba(108, 117, 125, 0.7)',
                'borderColor': 'rgb(108, 117, 125)',
                'borderWidth': 1
            }
        ]
    }
    
    return jsonify(chart_data)

@attendance_dashboard_bp.route('/export-excel')
@login_required
@module_access_required(Module.ATTENDANCE)
def export_excel():
    """تصدير بيانات الحضور لملف إكسل مع داشبورد ورسم بياني"""
    date_str = request.args.get('date')
    department_id = request.args.get('department_id')
    
    try:
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            selected_date = datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # استخراج بيانات الحضور التفصيلية
    detailed_query = db.session.query(
        Employee.name.label('اسم الموظف'),
        Employee.employee_id.label('الرقم الوظيفي'),
        Department.name.label('القسم'),
        Attendance.status.label('الحالة'),
        Attendance.check_in.label('وقت الحضور'),
        Attendance.check_out.label('وقت الانصراف'),
        Attendance.notes.label('ملاحظات')
    ).join(
        Employee, Employee.id == Attendance.employee_id
    ).join(
        employee_departments, employee_departments.c.employee_id == Employee.id
    ).join(
        Department, Department.id == employee_departments.c.department_id
    ).filter(
        Attendance.date == selected_date,
        Employee.status == 'active'
    )
    
    # إضافة فلتر القسم إذا تم تحديده
    if department_id:
        detailed_query = detailed_query.filter(Department.id == department_id)
    
    # تنفيذ الاستعلام التفصيلي
    detailed_results = detailed_query.all()
    
    # إنشاء DataFrame من النتائج التفصيلية
    detailed_df = pd.DataFrame(detailed_results)


# ==============================
# لوحة المراقبة الحية (Real-Time)
# ==============================

@attendance_dashboard_bp.route('/live-monitoring')
@login_required
@module_access_required(Module.ATTENDANCE)
def live_monitoring():
    """لوحة المراقبة الحية مع البيانات البيومترية وGPS"""
    today = datetime.now().date()
    
    # إحصائيات أساسية
    total_employees = Employee.query.filter_by(status='active').count()
    
    # الحضور اليوم
    checked_in = db.session.query(func.count(Attendance.id)).join(Employee).filter(
        Attendance.date == today,
        Attendance.check_in.isnot(None),
        Employee.status == 'active'
    ).scalar() or 0
    
    # الانصراف اليوم
    checked_out = db.session.query(func.count(Attendance.id)).join(Employee).filter(
        Attendance.date == today,
        Attendance.check_out.isnot(None),
        Employee.status == 'active'
    ).scalar() or 0
    
    # الموظفون الحاليون (حضروا ولم ينصرفوا)
    currently_present = checked_in - checked_out
    
    # إحصائيات بيومترية
    with_biometric = db.session.query(func.count(Attendance.id)).filter(
        Attendance.date == today,
        Attendance.check_in.isnot(None),
        or_(
            Attendance.check_in_confidence.isnot(None),
            Attendance.check_in_liveness_score.isnot(None)
        )
    ).scalar() or 0
    
    # متوسط الثقة
    avg_confidence = db.session.query(func.avg(Attendance.check_in_confidence)).filter(
        Attendance.date == today,
        Attendance.check_in_confidence.isnot(None)
    ).scalar()
    
    # متوسط اختبار الحياة
    avg_liveness = db.session.query(func.avg(Attendance.check_in_liveness_score)).filter(
        Attendance.date == today,
        Attendance.check_in_liveness_score.isnot(None)
    ).scalar()
    
    # إحصائيات GPS
    with_gps = db.session.query(func.count(Attendance.id)).filter(
        Attendance.date == today,
        Attendance.check_in_latitude.isnot(None),
        Attendance.check_in_longitude.isnot(None)
    ).scalar() or 0
    
    # متوسط دقة GPS
    avg_accuracy = db.session.query(func.avg(Attendance.check_in_accuracy)).filter(
        Attendance.date == today,
        Attendance.check_in_accuracy.isnot(None)
    ).scalar()
    
    # آخر عمليات الحضور
    recent_checkins = db.session.query(Attendance, Employee).join(
        Employee, Attendance.employee_id == Employee.id
    ).filter(
        Attendance.date == today,
        Attendance.check_in.isnot(None)
    ).order_by(Attendance.check_in.desc()).limit(15).all()
    
    # تحويل لقائمة
    recent_list = []
    for attendance, employee in recent_checkins:
        dept_name = None
        if employee.departments:
            dept_name = employee.departments[0].name
        
        recent_list.append({
            'id': attendance.id,
            'employee_id': employee.employee_id,
            'employee_name': employee.name,
            'department': dept_name or '---',
            'check_in_time': attendance.check_in.strftime('%H:%M:%S') if attendance.check_in else None,
            'check_out_time': attendance.check_out.strftime('%H:%M:%S') if attendance.check_out else None,
            'confidence': round(float(attendance.check_in_confidence), 2) if attendance.check_in_confidence else None,
            'liveness': round(float(attendance.check_in_liveness_score), 2) if attendance.check_in_liveness_score else None,
            'has_face_image': bool(attendance.check_in_face_image),
            'latitude': float(attendance.check_in_latitude) if attendance.check_in_latitude else None,
            'longitude': float(attendance.check_in_longitude) if attendance.check_in_longitude else None,
            'accuracy': round(float(attendance.check_in_accuracy), 1) if attendance.check_in_accuracy else None,
            'status': 'present' if not attendance.check_out else 'checked_out'
        })
    
    # توزيع الحضور حسب القسم
    departments = Department.query.all()
    dept_breakdown = []
    
    for dept in departments:
        active_emps = [e for e in dept.employees if e.status == 'active']
        total = len(active_emps)
        
        if total == 0:
            continue
        
        emp_ids = [e.id for e in active_emps]
        present = db.session.query(func.count(Attendance.id)).filter(
            Attendance.employee_id.in_(emp_ids),
            Attendance.date == today,
            Attendance.check_in.isnot(None)
        ).scalar() or 0
        
        dept_breakdown.append({
            'name': dept.name,
            'total': total,
            'present': present,
            'absent': total - present,
            'percentage': round((present / total * 100) if total > 0 else 0, 1)
        })
    
    return render_template(
        'attendance/live_monitoring.html',
        today=today,
        gregorian_date=format_date_gregorian(today),
        hijri_date=format_date_hijri(today),
        total_employees=total_employees,
        checked_in=checked_in,
        checked_out=checked_out,
        currently_present=currently_present,
        with_biometric=with_biometric,
        avg_confidence=round(float(avg_confidence), 2) if avg_confidence else 0,
        avg_liveness=round(float(avg_liveness), 2) if avg_liveness else 0,
        with_gps=with_gps,
        avg_accuracy=round(float(avg_accuracy), 1) if avg_accuracy else 0,
        recent_checkins=recent_list,
        dept_breakdown=dept_breakdown
    )


@attendance_dashboard_bp.route('/api/live-stats')
@login_required
@module_access_required(Module.ATTENDANCE)
def api_live_stats():
    """API endpoint للحصول على الإحصائيات الحية"""
    try:
        today = datetime.now().date()
        
        # إحصائيات أساسية
        total_employees = Employee.query.filter_by(status='active').count()
        
        checked_in = db.session.query(func.count(Attendance.id)).join(Employee).filter(
            Attendance.date == today,
            Attendance.check_in.isnot(None),
            Employee.status == 'active'
        ).scalar() or 0
        
        checked_out = db.session.query(func.count(Attendance.id)).join(Employee).filter(
            Attendance.date == today,
            Attendance.check_out.isnot(None),
            Employee.status == 'active'
        ).scalar() or 0
        
        currently_present = checked_in - checked_out
        
        # إحصائيات بيومترية
        with_biometric = db.session.query(func.count(Attendance.id)).filter(
            Attendance.date == today,
            or_(
                Attendance.check_in_confidence.isnot(None),
                Attendance.check_in_liveness_score.isnot(None)
            )
        ).scalar() or 0
        
        avg_confidence = db.session.query(func.avg(Attendance.check_in_confidence)).filter(
            Attendance.date == today,
            Attendance.check_in_confidence.isnot(None)
        ).scalar()
        
        avg_liveness = db.session.query(func.avg(Attendance.check_in_liveness_score)).filter(
            Attendance.date == today,
            Attendance.check_in_liveness_score.isnot(None)
        ).scalar()
        
        # إحصائيات GPS
        with_gps = db.session.query(func.count(Attendance.id)).filter(
            Attendance.date == today,
            Attendance.check_in_latitude.isnot(None)
        ).scalar() or 0
        
        avg_accuracy = db.session.query(func.avg(Attendance.check_in_accuracy)).filter(
            Attendance.date == today,
            Attendance.check_in_accuracy.isnot(None)
        ).scalar()
        
        # آخر الحضور
        recent = db.session.query(Attendance, Employee).join(
            Employee, Attendance.employee_id == Employee.id
        ).filter(
            Attendance.date == today,
            Attendance.check_in.isnot(None)
        ).order_by(Attendance.check_in.desc()).limit(10).all()
        
        recent_list = []
        for attendance, employee in recent:
            dept_name = employee.departments[0].name if employee.departments else '---'
            recent_list.append({
                'employee_id': employee.employee_id,
                'employee_name': employee.name,
                'department': dept_name,
                'check_in_time': attendance.check_in.strftime('%H:%M:%S'),
                'check_out_time': attendance.check_out.strftime('%H:%M:%S') if attendance.check_out else None,
                'confidence': round(float(attendance.check_in_confidence), 2) if attendance.check_in_confidence else None,
                'liveness': round(float(attendance.check_in_liveness_score), 2) if attendance.check_in_liveness_score else None,
                'has_face_image': bool(attendance.check_in_face_image),
                'status': 'present' if not attendance.check_out else 'checked_out'
            })
        
        return jsonify({
            'success': True,
            'timestamp': datetime.now().isoformat(),
            'stats': {
                'total_employees': total_employees,
                'checked_in': checked_in,
                'checked_out': checked_out,
                'currently_present': currently_present,
                'absent': total_employees - checked_in
            },
            'biometric_stats': {
                'with_biometric': with_biometric,
                'avg_confidence': round(float(avg_confidence), 2) if avg_confidence else 0,
                'avg_liveness': round(float(avg_liveness), 2) if avg_liveness else 0
            },
            'location_stats': {
                'with_gps': with_gps,
                'avg_accuracy': round(float(avg_accuracy), 1) if avg_accuracy else 0
            },
            'recent_checkins': recent_list
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    
    # معالجة الحالة
    def format_status(status):
        if status == 'present':
            return 'حاضر'
        elif status == 'absent':
            return 'غائب'
        elif status == 'leave':
            return 'إجازة'
        elif status == 'sick':
            return 'مرضي'
        else:
            return status
    
    if 'الحالة' in detailed_df.columns and not detailed_df.empty:
        detailed_df['الحالة'] = detailed_df['الحالة'].apply(format_status)
    
    # معالجة أوقات الحضور والانصراف
    for col in ['وقت الحضور', 'وقت الانصراف']:
        if col in detailed_df.columns and not detailed_df.empty:
            detailed_df[col] = detailed_df[col].apply(lambda x: x.strftime('%H:%M') if x else '-')
    
    # استخراج إحصائيات الحضور حسب القسم
    departments = Department.query.all()
    
    # إنشاء بيانات الداشبورد
    dashboard_data = []
    
    # حساب الإحصائيات لكل قسم
    for dept in departments:
        # عدد الموظفين في القسم - استخدام علاقة many-to-many
        active_employees = [emp for emp in dept.employees if emp.status == 'active']
        employees_count = len(active_employees)
        employee_ids = [emp.id for emp in active_employees]
        
        # عدد الحاضرين
        present_count = db.session.query(func.count(Attendance.id)).filter(
            Attendance.employee_id.in_(employee_ids),
            Attendance.date == selected_date,
            Attendance.status == 'present'
        ).scalar() or 0 if employee_ids else 0
        
        # عدد الغائبين
        absent_count = db.session.query(func.count(Attendance.id)).filter(
            Attendance.employee_id.in_(employee_ids),
            Attendance.date == selected_date,
            Attendance.status == 'absent'
        ).scalar() or 0 if employee_ids else 0
        
        # عدد الإجازات
        leave_count = db.session.query(func.count(Attendance.id)).filter(
            Attendance.employee_id.in_(employee_ids),
            Attendance.date == selected_date,
            Attendance.status == 'leave'
        ).scalar() or 0 if employee_ids else 0
        
        # عدد المرضي
        sick_count = db.session.query(func.count(Attendance.id)).filter(
            Attendance.employee_id.in_(employee_ids),
            Attendance.date == selected_date,
            Attendance.status == 'sick'
        ).scalar() or 0 if employee_ids else 0
        
        # غير مسجلين
        total_records = present_count + absent_count + leave_count + sick_count
        missing_count = employees_count - total_records
        if missing_count < 0:
            missing_count = 0
        
        # نسبة الحضور
        if employees_count > 0:
            present_percentage = round((present_count / employees_count) * 100, 1)
            absent_percentage = round((absent_count / employees_count) * 100, 1)
        else:
            present_percentage = 0
            absent_percentage = 0
        
        # إضافة بيانات القسم
        dashboard_data.append({
            'القسم': dept.name,
            'عدد الموظفين': employees_count,
            'الحاضرون': present_count,
            'الغائبون': absent_count,
            'إجازة': leave_count,
            'مرضي': sick_count,
            'غير مسجلين': missing_count,
            'نسبة الحضور %': present_percentage,
            'نسبة الغياب %': absent_percentage
        })
    
    # إنشاء DataFrame للداشبورد
    dashboard_df = pd.DataFrame(dashboard_data)
    
    # إنشاء قاموس لعدد الموظفين في كل قسم (من dashboard_data)
    dept_employee_counts = {}
    for row in dashboard_data:
        dept_employee_counts[row['القسم']] = row['عدد الموظفين']
    
    # إنشاء بيانات تقرير الحضور حسب التاريخ والقسم
    overview_query = db.session.query(
        Attendance.date.label('date'),
        Department.name.label('department'),
        func.sum(case((Attendance.status == 'present', 1), else_=0)).label('attend'),
        func.sum(case((Attendance.status == 'leave', 1), else_=0)).label('day_off'),
        func.sum(case((Attendance.status == 'absent', 1), else_=0)).label('absent')
    ).join(
        Employee, Employee.id == Attendance.employee_id
    ).join(
        employee_departments, employee_departments.c.employee_id == Employee.id
    ).join(
        Department, Department.id == employee_departments.c.department_id
    ).filter(
        Attendance.date == selected_date,
        Employee.status == 'active'
    ).group_by(
        Attendance.date,
        Department.name
    ).order_by(
        Attendance.date,
        Department.name
    )
    
    # إضافة فلتر القسم إذا تم تحديده
    if department_id:
        overview_query = overview_query.filter(
            employee_departments.c.department_id == department_id
        )
    
    overview_results = overview_query.all()
    
    # تحويل النتائج إلى قائمة
    overview_data = []
    for row in overview_results:
        dept_name = row.department
        total = dept_employee_counts.get(dept_name, 0)
        attend = row.attend or 0
        percentage = round((attend / total * 100), 0) if total > 0 else 0
        
        overview_data.append({
            'DATE': row.date.strftime('%d-%b-%Y') if row.date else '-',
            'PROJECT': dept_name,
            'LOCATION': 'qassim',
            'TOTAL EMP': total,
            'ATTEND': attend,
            'DAY OFF': row.day_off or 0,
            'ABSENT': row.absent or 0,
            'PERCENTAGE': f"{int(percentage)}%"
        })
    
    # إنشاء DataFrame للتقرير الإجمالي
    overview_df = pd.DataFrame(overview_data)
    
    # تحديد اسم الملف
    date_str = selected_date.strftime('%Y-%m-%d')
    filename = f"تقرير_الحضور_{date_str}.xlsx"
    
    # حفظ في مجلد دائم
    export_folder = os.path.join(UPLOAD_FOLDER, 'exports', 'attendance')
    os.makedirs(export_folder, exist_ok=True)
    temp_file_path = os.path.join(export_folder, f"attendance_{date_str}_{datetime.now().strftime('%H%M%S')}.xlsx")
    
    # حفظ البيانات إلى الملف
    with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
        # إضافة ورقة داشبورد
        dashboard_df.to_excel(writer, sheet_name='لوحة الإحصائيات', index=False)
        
        # إضافة ورقة تقرير الحضور التفصيلي
        if not overview_df.empty:
            overview_df.to_excel(writer, sheet_name='Attendance Overview', index=False)
        
        # الوصول إلى الكائن workbook
        workbook = writer.book
        
        # إضافة ورقة الرسم البياني
        chart_sheet = workbook.create_sheet(title='الرسم البياني')
        
        # إضافة بيانات للرسم البياني
        for r, row in enumerate(dashboard_data, start=1):
            chart_sheet.cell(row=r, column=1, value=row['القسم'])
            chart_sheet.cell(row=r, column=2, value=row['الحاضرون'])
            chart_sheet.cell(row=r, column=3, value=row['الغائبون'])
            chart_sheet.cell(row=r, column=4, value=row['إجازة'])
            chart_sheet.cell(row=r, column=5, value=row['مرضي'])
        
        # إضافة العناوين
        chart_sheet.cell(row=1, column=1, value='القسم')
        chart_sheet.cell(row=1, column=2, value='الحاضرون')
        chart_sheet.cell(row=1, column=3, value='الغائبون')
        chart_sheet.cell(row=1, column=4, value='إجازة')
        chart_sheet.cell(row=1, column=5, value='مرضي')
        
        # تنسيق ورقة الإحصائيات
        dashboard_sheet = writer.sheets['لوحة الإحصائيات']
        for col in dashboard_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            dashboard_sheet.column_dimensions[column].width = adjusted_width
        
        # إضافة إطار وتلوين للعناوين
        for col_num, column_title in enumerate(dashboard_df.columns, 1):
            cell = dashboard_sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            
        # تلوين الخلايا حسب القيمة (للنسب المئوية)
        for row_num, row in enumerate(dashboard_df.iterrows(), 2):
            # الحصول على مؤشر العمود
            present_col = dashboard_df.columns.get_indexer(['نسبة الحضور %'])[0] + 1
            absent_col = dashboard_df.columns.get_indexer(['نسبة الغياب %'])[0] + 1
            
            present_cell = dashboard_sheet.cell(row=row_num, column=present_col)
            absent_cell = dashboard_sheet.cell(row=row_num, column=absent_col)
            
            present_value = row[1]['نسبة الحضور %']
            
            # تلوين نسبة الحضور
            if present_value >= 90:
                present_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif present_value >= 70:
                present_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                present_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # تنسيق ورقة Attendance Overview
        if not overview_df.empty and 'Attendance Overview' in writer.sheets:
            overview_sheet = writer.sheets['Attendance Overview']
            
            # تنسيق العناوين - لون سماوي مثل الصورة
            for col_num in range(1, len(overview_df.columns) + 1):
                cell = overview_sheet.cell(row=1, column=col_num)
                cell.fill = PatternFill(start_color='17C5BC', end_color='17C5BC', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # ضبط عرض الأعمدة
            overview_sheet.column_dimensions['A'].width = 15  # DATE
            overview_sheet.column_dimensions['B'].width = 15  # PROJECT
            overview_sheet.column_dimensions['C'].width = 15  # LOCATION
            overview_sheet.column_dimensions['D'].width = 12  # TOTAL EMP
            overview_sheet.column_dimensions['E'].width = 12  # ATTEND
            overview_sheet.column_dimensions['F'].width = 12  # DAY OFF
            overview_sheet.column_dimensions['G'].width = 12  # ABSENT
            overview_sheet.column_dimensions['H'].width = 15  # PERCENTAGE
            
            # محاذاة البيانات في الوسط
            for row_num in range(2, len(overview_df) + 2):
                for col_num in range(1, len(overview_df.columns) + 1):
                    cell = overview_sheet.cell(row=row_num, column=col_num)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # إنشاء مخطط شريطي (Bar Chart)
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "توزيع الحضور حسب الأقسام"
        chart.y_axis.title = "عدد الموظفين"
        chart.x_axis.title = "القسم"
        
        # تحديد نطاق البيانات
        rows_count = len(dashboard_data) + 1  # +1 للعناوين
        
        # إضافة البيانات للمخطط
        data = Reference(chart_sheet, min_col=2, min_row=1, max_row=rows_count, max_col=5)
        cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=rows_count)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # إضافة المخطط إلى الورقة
        chart_sheet.add_chart(chart, "A10")
    
    # إرسال الملف كملف للتحميل
    return send_file(
        temp_file_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )