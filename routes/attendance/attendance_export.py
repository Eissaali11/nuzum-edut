"""
Attendance Export Routes
========================
Extracted from _attendance_main.py as part of modularization.
Handles all attendance export operations (Excel, PDF).

Routes:
    - GET/POST /export/excel           : Export attendance to Excel
    - GET      /export                  : Export page (form)
    - GET      /export-excel-dashboard : Export dashboard summary
    - GET      /export-excel-department: Export department details  
    - GET      /department/export-data  : Export with filters (P/A format)
    - GET      /department/export-period: Export department over period (professional dashboard)
"""

from flask import request, flash, redirect, url_for, render_template, send_file
from flask_login import current_user
from datetime import datetime, timedelta
from sqlalchemy import or_
import logging

from core.extensions import db
from models import Attendance, Employee, Department, employee_departments
from utils.date_converter import parse_date, format_date_hijri, format_date_gregorian
from utils.excel import export_attendance_by_department
from utils.excel_dashboard import export_attendance_by_department_with_dashboard

logger = logging.getLogger(__name__)


def export_excel():
    """تصدير بيانات الحضور إلى ملف Excel"""
    try:
        if request.method == 'POST':
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            department_id = request.form.get('department_id')
        else:  # GET
            start_date_str = request.args.get('start_date')
            end_date_str = request.args.get('end_date')
            department_id = request.args.get('department_id')
        
        if not start_date_str:
            flash('تاريخ البداية مطلوب', 'danger')
            return redirect(url_for('attendance.export_page'))
        
        try:
            start_date = parse_date(start_date_str)
            if end_date_str:
                end_date = parse_date(end_date_str)
            else:
                end_date = datetime.now().date()
        except (ValueError, TypeError):
            flash('تاريخ غير صالح', 'danger')
            return redirect(url_for('attendance.export_page'))
        
        if department_id and department_id != '':
            department = Department.query.get(department_id)
            if not department:
                flash('القسم غير موجود', 'danger')
                return redirect(url_for('attendance.export_page'))
            
            attendances = Attendance.query.filter(
                Attendance.date.between(start_date, end_date)
            ).all()
            
            employee_ids_with_attendance = set([att.employee_id for att in attendances])
            
            department_employee_ids = [emp.id for emp in department.employees]
            employees_to_export = Employee.query.filter(
                Employee.id.in_(department_employee_ids),
                or_(
                    ~Employee.status.in_(['terminated', 'inactive']),
                    Employee.id.in_(employee_ids_with_attendance)
                )
            ).all()
            
            attendances = [att for att in attendances if att.employee_id in department_employee_ids]
            
            excel_file = export_attendance_by_department(employees_to_export, attendances, start_date, end_date)
            
            if end_date_str:
                filename = f'سجل الحضور - {department.name} - {start_date_str} إلى {end_date_str}.xlsx'
            else:
                filename = f'سجل الحضور - {department.name} - {start_date_str}.xlsx'
        else:
            departments = Department.query.all()
            
            attendances = Attendance.query.filter(
                Attendance.date.between(start_date, end_date)
            ).all()
            
            employee_ids_with_attendance = set([att.employee_id for att in attendances])
            
            all_employees = Employee.query.filter(
                or_(
                    ~Employee.status.in_(['terminated', 'inactive']),
                    Employee.id.in_(employee_ids_with_attendance)
                )
            ).all()
            
            excel_file = export_attendance_by_department_with_dashboard(all_employees, attendances, start_date, end_date)
            
            if end_date_str:
                filename = f'سجل الحضور - جميع الأقسام - {start_date_str} إلى {end_date_str}.xlsx'
            else:
                filename = f'سجل الحضور - جميع الأقسام - {start_date_str}.xlsx'
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('attendance.export_page'))


def export_page():
    """صفحة تصدير بيانات الحضور إلى ملف Excel"""
    departments = Department.query.all()
    today = datetime.now().date()
    start_of_month = today.replace(day=1)
    
    hijri_today = format_date_hijri(today)
    gregorian_today = format_date_gregorian(today)
    
    hijri_start = format_date_hijri(start_of_month)
    gregorian_start = format_date_gregorian(start_of_month)
    
    return render_template('attendance/export.html',
                          departments=departments,
                          today=today,
                          start_of_month=start_of_month,
                          hijri_today=hijri_today,
                          gregorian_today=gregorian_today,
                          hijri_start=hijri_start,
                          gregorian_start=gregorian_start)


def export_excel_dashboard():
    """تصدير لوحة المعلومات إلى Excel"""
    try:
        from services.attendance_reports import AttendanceReportService
        
        selected_department = request.args.get('department', None)
        selected_project = request.args.get('project', None)
        
        result = AttendanceReportService.export_dashboard_summary(selected_department, selected_project)
        
        return send_file(
            result['buffer'],
            mimetype=result['mimetype'],
            as_attachment=True,
            download_name=result['filename']
        )
        
    except Exception as e:
        logger.error(f"Export Excel Dashboard Error: {type(e).__name__}: {str(e)}", exc_info=True)
        flash('حدث خطأ أثناء تصدير الملف', 'error')
        return redirect(url_for('attendance.dashboard'))


def export_excel_department():
    """تصدير تفاصيل القسم إلى Excel"""
    logger.info("[EXPORT] export_excel_department route called")
    try:
        department_name = request.args.get('department')
        logger.info(f"[EXPORT] Department requested: {department_name}")
        selected_project = request.args.get('project', None)
        
        if not department_name:
            logger.warning("[EXPORT] No department provided")
            flash('يجب تحديد القسم', 'error')
            return redirect(url_for('attendance.dashboard'))
        
        logger.info("[EXPORT] Loading AttendanceReportService")
        from services.attendance_reports import AttendanceReportService
        
        logger.info(f"[EXPORT] Calling export_department_details for: {department_name}")
        result = AttendanceReportService.export_department_details(department_name, selected_project)
        logger.info(f"[EXPORT] Service returned result with filename: {result.get('filename')}")
        
        logger.info("[EXPORT] Sending file...")
        return send_file(
            result['buffer'],
            mimetype=result['mimetype'],
            as_attachment=True,
            download_name=result['filename']
        )
        
    except Exception as e:
        logger.error(f"Export Excel Department Error: {type(e).__name__}: {str(e)}", exc_info=True)
        flash(f'خطأ: {str(e)}', 'error')
        return redirect(url_for('attendance.dashboard'))


def export_department_data():
    """تصدير بيانات الحضور حسب الفلاتر مع تصميم احترافي"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    try:
        department_id = request.args.get('department_id', '')
        search_query = request.args.get('search_query', '').strip()
        status_filter = request.args.get('status_filter', '')
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')
        
        if not start_date_str:
            start_date = datetime.now().date() - timedelta(days=30)
        else:
            start_date = parse_date(start_date_str)
        
        if not end_date_str:
            end_date = datetime.now().date()
        else:
            end_date = parse_date(end_date_str)
        
        query = Attendance.query.join(Employee).filter(
            Attendance.date >= start_date,
            Attendance.date <= end_date
        )
        
        if department_id:
            query = query.join(employee_departments).filter(
                employee_departments.c.department_id == int(department_id)
            )
        
        if search_query:
            query = query.filter(
                or_(
                    Employee.name.ilike(f'%{search_query}%'),
                    Employee.employee_id.ilike(f'%{search_query}%'),
                    Employee.national_id.ilike(f'%{search_query}%')
                )
            )
        
        if status_filter:
            query = query.filter(Attendance.status == status_filter)
        
        attendances = query.order_by(Attendance.date.desc(), Employee.name).all()
        
        total_count = len(attendances)
        present_count = sum(1 for a in attendances if a.status == 'present')
        absent_count = sum(1 for a in attendances if a.status == 'absent')
        leave_count = sum(1 for a in attendances if a.status == 'leave')
        sick_count = sum(1 for a in attendances if a.status == 'sick')
        
        wb = Workbook()
        
        ws = wb.active
        ws.title = "بيانات الحضور"
        
        all_dates = sorted(set(att.date for att in attendances if att.date))
        employees_dict = {}
        for att in attendances:
            if att.employee.id not in employees_dict:
                employees_dict[att.employee.id] = att.employee
        
        sorted_employees = sorted(employees_dict.values(), key=lambda e: e.name)
        
        attendance_map = {}
        for att in attendances:
            key = (att.employee.id, att.date)
            attendance_map[key] = att.status
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 22
        
        for col_idx in range(len(all_dates)):
            col_letter = get_column_letter(col_idx + 5)
            ws.column_dimensions[col_letter].width = 4
        
        header_row = ['الموظف', 'الرقم الوظيفي', 'رقم الهوية', 'القسم'] + [
            d.strftime('%b %d') for d in all_dates
        ]
        ws.append(header_row)
        
        header_fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 28
        
        row_num = 2
        for emp in sorted_employees:
            department_name = ', '.join([d.name for d in emp.departments]) if emp.departments else '-'
            
            row_data = [
                emp.name,
                emp.employee_id or '-',
                emp.national_id if hasattr(emp, 'national_id') and emp.national_id else '-',
                department_name
            ]
            
            for date in all_dates:
                status = attendance_map.get((emp.id, date), '')
                
                if status == 'present':
                    row_data.append('P')
                elif status == 'absent':
                    row_data.append('A')
                elif status == 'leave' or status == 'sick':
                    row_data.append('S')
                else:
                    row_data.append('')
            
            ws.append(row_data)
            
            for col_idx in range(1, len(row_data) + 1):
                cell_obj = ws.cell(row=row_num, column=col_idx)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_obj.border = Border(
                    left=Side(style='thin', color='E0E0E0'),
                    right=Side(style='thin', color='E0E0E0'),
                    top=Side(style='thin', color='E0E0E0'),
                    bottom=Side(style='thin', color='E0E0E0')
                )
                
                if col_idx <= 4:
                    cell_obj.font = Font(bold=True if col_idx == 1 else False)
                    if row_num % 2 == 0:
                        cell_obj.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                else:
                    value = cell_obj.value
                    if value == 'P':
                        cell_obj.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        cell_obj.font = Font(bold=True, color="155724", size=11)
                    elif value == 'A':
                        cell_obj.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                        cell_obj.font = Font(bold=True, color="721C24", size=11)
                    elif value == 'S':
                        cell_obj.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                        cell_obj.font = Font(bold=True, color="856404", size=11)
                    else:
                        if row_num % 2 == 0:
                            cell_obj.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            ws.row_dimensions[row_num].height = 20
            row_num += 1
        
        stats_ws = wb.create_sheet("الإحصائيات")
        
        stats_ws.merge_cells('A1:D1')
        title_cell = stats_ws['A1']
        title_cell.value = "ملخص إحصائيات الحضور"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        stats_ws.row_dimensions[1].height = 32
        
        stats_ws.column_dimensions['A'].width = 25
        stats_ws.column_dimensions['B'].width = 18
        stats_ws.column_dimensions['C'].width = 18
        stats_ws.column_dimensions['D'].width = 18
        
        stats_data = [
            ['', 'العدد', 'النسبة %', 'الحالة'],
            ['إجمالي السجلات', total_count, '100%', 'total'],
            ['موظفون حاضرون', present_count, f'{int((present_count/max(total_count,1))*100)}%', 'present'],
            ['موظفون غائبون', absent_count, f'{int((absent_count/max(total_count,1))*100)}%', 'absent'],
            ['إجازات', leave_count, f'{int((leave_count/max(total_count,1))*100)}%', 'leave'],
            ['مرضي', sick_count, f'{int((sick_count/max(total_count,1))*100)}%', 'sick']
        ]
        
        for idx, row_data in enumerate(stats_data, 2):
            for col_idx, value in enumerate(row_data[:3], 1):
                cell = stats_ws.cell(row=idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if idx == 2:
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
                elif row_data[3] == 'total':
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                elif row_data[3] == 'present':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    cell.font = Font(bold=True, color="155724")
                elif row_data[3] == 'absent':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    cell.font = Font(bold=True, color="721C24")
                elif row_data[3] == 'leave':
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    cell.font = Font(bold=True, color="856404")
                elif row_data[3] == 'sick':
                    cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
                    cell.font = Font(bold=True, color="0C5460")
                
                stats_ws.row_dimensions[idx].height = 26
        
        if total_count > 0:
            pie = PieChart()
            pie.title = "توزيع حالات الحضور"
            pie.style = 10
            labels = Reference(stats_ws, min_col=1, min_row=3, max_row=6)
            data = Reference(stats_ws, min_col=2, min_row=2, max_row=6)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 12
            pie.width = 16
            stats_ws.add_chart(pie, "A9")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"حضور_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    
    except Exception as e:
        logger.error(f"خطأ في تصدير البيانات: {str(e)}")
        flash(f'خطأ في التصدير: {str(e)}', 'error')
        return redirect(url_for('attendance.department_attendance_view'))


def export_department_period():
    """تصدير حضور قسم خلال فترة زمنية إلى Excel مع dashboard احترافي"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    try:
        department_id = request.args.get('department_id')
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        if not all([start_date_str, end_date_str]):
            flash('يجب تحديد الفترة الزمنية', 'error')
            return redirect(url_for('attendance.department_attendance_view'))
        
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
        
        department = None
        department_name = "جميع الأقسام"
        if department_id:
            department = Department.query.get_or_404(department_id)
            department_name = department.name
        
        query = Attendance.query.join(Employee).filter(
            Attendance.date >= start_date,
            Attendance.date <= end_date
        )
        
        if department_id:
            query = query.join(employee_departments).filter(
                employee_departments.c.department_id == int(department_id)
            )
        
        attendances = query.order_by(Employee.name, Attendance.date).all()
        
        if not attendances:
            flash('لا توجد بيانات للتصدير في هذه الفترة', 'warning')
            return redirect(url_for('attendance.department_attendance_view'))
        
        wb = Workbook()
        
        ws_dashboard = wb.active
        ws_dashboard.title = "Dashboard"
        
        title_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=20)
        subtitle_fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
        subtitle_font = Font(bold=True, color="FFFFFF", size=14)
        
        kpi_label_format = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        kpi_success_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        kpi_danger_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        kpi_warning_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
        kpi_info_fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
        
        ws_dashboard.merge_cells('A1:H1')
        ws_dashboard['A1'] = f'تقرير حضور قسم {department_name}'
        ws_dashboard['A1'].fill = title_fill
        ws_dashboard['A1'].font = title_font
        ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard.row_dimensions[1].height = 35
        
        ws_dashboard.merge_cells('A2:H2')
        ws_dashboard['A2'] = f'الفترة: من {start_date.strftime("%Y-%m-%d")} إلى {end_date.strftime("%Y-%m-%d")}'
        ws_dashboard['A2'].fill = subtitle_fill
        ws_dashboard['A2'].font = subtitle_font
        ws_dashboard['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard.row_dimensions[2].height = 25
        
        total_count = len(attendances)
        present_count = sum(1 for a in attendances if a.status == 'present')
        absent_count = sum(1 for a in attendances if a.status == 'absent')
        leave_count = sum(1 for a in attendances if a.status == 'leave')
        sick_count = sum(1 for a in attendances if a.status == 'sick')
        attendance_rate = (present_count / total_count * 100) if total_count > 0 else 0
        
        unique_employees = set(a.employee_id for a in attendances)
        employee_count = len(unique_employees)
        
        days_count = (end_date - start_date).days + 1
        
        kpis = [
            ('A4', 'B4', 'إجمالي السجلات', total_count, kpi_label_format, kpi_info_fill),
            ('C4', 'D4', 'عدد الموظفين', employee_count, kpi_label_format, kpi_info_fill),
            ('E4', 'F4', 'عدد الأيام', days_count, kpi_label_format, kpi_info_fill),
            ('G4', 'H4', 'نسبة الحضور', f'{attendance_rate:.1f}%', kpi_label_format, kpi_success_fill),
        ]
        
        for start_cell, end_cell, label, value, label_fill, value_fill in kpis:
            ws_dashboard.merge_cells(f'{start_cell}:{start_cell}')
            ws_dashboard[start_cell] = label
            ws_dashboard[start_cell].fill = label_fill
            ws_dashboard[start_cell].font = Font(bold=True, size=11)
            ws_dashboard[start_cell].alignment = Alignment(horizontal='center', vertical='center')
            ws_dashboard[start_cell].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='thin')
            )
            
            ws_dashboard.merge_cells(f'{end_cell}:{end_cell}')
            ws_dashboard[end_cell] = value
            ws_dashboard[end_cell].fill = value_fill
            ws_dashboard[end_cell].font = Font(bold=True, size=16, color='1e3c72')
            ws_dashboard[end_cell].alignment = Alignment(horizontal='center', vertical='center')
            ws_dashboard[end_cell].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='thin'), bottom=Side(style='medium')
            )
        
        ws_dashboard.row_dimensions[4].height = 30
        
        ws_dashboard.merge_cells('A6:H6')
        ws_dashboard['A6'] = 'توزيع حالات الحضور'
        ws_dashboard['A6'].fill = subtitle_fill
        ws_dashboard['A6'].font = Font(bold=True, color="FFFFFF", size=13)
        ws_dashboard['A6'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard.row_dimensions[6].height = 25
        
        stats_data = [
            ('A8', 'B8', 'حاضر', present_count, kpi_success_fill),
            ('C8', 'D8', 'غائب', absent_count, kpi_danger_fill),
            ('E8', 'F8', 'إجازة', leave_count, kpi_warning_fill),
            ('G8', 'H8', 'مرضي', sick_count, kpi_info_fill),
        ]
        
        for start_cell, end_cell, label, value, fill in stats_data:
            ws_dashboard.merge_cells(f'{start_cell}:{start_cell}')
            ws_dashboard[start_cell] = label
            ws_dashboard[start_cell].fill = fill
            ws_dashboard[start_cell].font = Font(bold=True, size=12)
            ws_dashboard[start_cell].alignment = Alignment(horizontal='center', vertical='center')
            ws_dashboard[start_cell].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='thin')
            )
            
            ws_dashboard.merge_cells(f'{end_cell}:{end_cell}')
            ws_dashboard[end_cell] = value
            ws_dashboard[end_cell].fill = fill
            ws_dashboard[end_cell].font = Font(bold=True, size=18)
            ws_dashboard[end_cell].alignment = Alignment(horizontal='center', vertical='center')
            ws_dashboard[end_cell].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='thin'), bottom=Side(style='medium')
            )
        
        ws_dashboard.row_dimensions[8].height = 30
        
        ws_dashboard['A10'] = 'الحالة'
        ws_dashboard['B10'] = 'العدد'
        ws_dashboard['A10'].font = Font(bold=True)
        ws_dashboard['B10'].font = Font(bold=True)
        
        chart_data = [
            ('حاضر', present_count),
            ('غائب', absent_count),
            ('إجازة', leave_count),
            ('مرضي', sick_count),
        ]
        
        for idx, (label, value) in enumerate(chart_data, start=11):
            ws_dashboard[f'A{idx}'] = label
            ws_dashboard[f'B{idx}'] = value
        
        pie_chart = PieChart()
        pie_chart.title = "توزيع حالات الحضور"
        pie_chart.style = 10
        pie_chart.height = 10
        pie_chart.width = 15
        
        labels = Reference(ws_dashboard, min_col=1, min_row=11, max_row=14)
        data = Reference(ws_dashboard, min_col=2, min_row=10, max_row=14)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showVal = True
        
        ws_dashboard.add_chart(pie_chart, "J6")
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws_dashboard.column_dimensions[col].width = 15
        
        ws_data = wb.create_sheet("البيانات التفصيلية")
        
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        present_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        absent_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        leave_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
        sick_fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
        
        ws_data.merge_cells('A1:H1')
        ws_data['A1'] = f'سجلات حضور قسم {department_name}'
        ws_data['A1'].fill = title_fill
        ws_data['A1'].font = title_font
        ws_data['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_data.row_dimensions[1].height = 30
        
        headers = ['الموظف', 'الرقم الوظيفي', 'التاريخ', 'الحالة', 'وقت الدخول', 'وقت الخروج', 'ساعات العمل', 'ملاحظات']
        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        row_num = 4
        for attendance in attendances:
            ws_data.cell(row=row_num, column=1, value=attendance.employee.name)
            ws_data.cell(row=row_num, column=2, value=attendance.employee.employee_id or '-')
            ws_data.cell(row=row_num, column=3, value=attendance.date.strftime('%Y-%m-%d'))
            
            status_cell = ws_data.cell(row=row_num, column=4)
            if attendance.status == 'present':
                status_cell.value = 'حاضر'
                status_cell.fill = present_fill
            elif attendance.status == 'absent':
                status_cell.value = 'غائب'
                status_cell.fill = absent_fill
            elif attendance.status == 'leave':
                status_cell.value = 'إجازة'
                status_cell.fill = leave_fill
            elif attendance.status == 'sick':
                status_cell.value = 'مرضي'
                status_cell.fill = sick_fill
            
            status_cell.alignment = Alignment(horizontal='center')
            
            ws_data.cell(row=row_num, column=5, value=attendance.check_in.strftime('%H:%M') if attendance.check_in else '-')
            ws_data.cell(row=row_num, column=6, value=attendance.check_out.strftime('%H:%M') if attendance.check_out else '-')
            
            if attendance.check_in and attendance.check_out:
                hours = (attendance.check_out.hour - attendance.check_in.hour) + (attendance.check_out.minute - attendance.check_in.minute) / 60.0
                ws_data.cell(row=row_num, column=7, value=f"{hours:.1f}")
            else:
                ws_data.cell(row=row_num, column=7, value='-')
            
            ws_data.cell(row=row_num, column=8, value=attendance.notes or '-')
            
            for col in range(1, 9):
                ws_data.cell(row=row_num, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                ws_data.cell(row=row_num, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            row_num += 1
        
        ws_data.column_dimensions['A'].width = 25
        ws_data.column_dimensions['B'].width = 15
        ws_data.column_dimensions['C'].width = 15
        ws_data.column_dimensions['D'].width = 12
        ws_data.column_dimensions['E'].width = 12
        ws_data.column_dimensions['F'].width = 12
        ws_data.column_dimensions['G'].width = 15
        ws_data.column_dimensions['H'].width = 30
        
        ws_matrix = wb.create_sheet("سجل الحضور")
        
        employee_attendance_map = {}
        unique_dates = set()
        
        for att in attendances:
            emp_id = att.employee_id
            if emp_id not in employee_attendance_map:
                employee_attendance_map[emp_id] = {
                    'employee': att.employee,
                    'dates': {}
                }
            employee_attendance_map[emp_id]['dates'][att.date] = att.status
            unique_dates.add(att.date)
        
        sorted_dates = sorted(list(unique_dates))
        
        ws_matrix.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(sorted_dates) + 8)
        ws_matrix.cell(row=1, column=1, value=f'سجل حضور قسم {department_name}')
        ws_matrix.cell(row=1, column=1).fill = title_fill
        ws_matrix.cell(row=1, column=1).font = title_font
        ws_matrix.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws_matrix.row_dimensions[1].height = 30
        
        ws_matrix.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(sorted_dates) + 8)
        ws_matrix.cell(row=2, column=1, value=f'الفترة: من {start_date.strftime("%Y-%m-%d")} إلى {end_date.strftime("%Y-%m-%d")}')
        ws_matrix.cell(row=2, column=1).fill = subtitle_fill
        ws_matrix.cell(row=2, column=1).font = subtitle_font
        ws_matrix.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws_matrix.row_dimensions[2].height = 25
        
        fixed_headers = ['Name', 'ID Number', 'Emp. No', 'Job Title', 'No. Mobile', 'Location', 'Project', 'Total']
        header_row = 4
        
        for col_num, header in enumerate(fixed_headers, 1):
            cell = ws_matrix.cell(row=header_row, column=col_num, value=header)
            cell.fill = PatternFill(start_color="00B0B0", end_color="00B0B0", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        for idx, date in enumerate(sorted_dates, start=9):
            day_name = date.strftime('%a')
            cell = ws_matrix.cell(row=header_row, column=idx, value=day_name)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            ws_matrix.column_dimensions[get_column_letter(idx)].width = 4
        
        for idx, date in enumerate(sorted_dates, start=9):
            cell = ws_matrix.cell(row=header_row + 1, column=idx, value=date.strftime('%d/%m/%Y'))
            cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        data_row = header_row + 2
        for emp_id, emp_data in sorted(employee_attendance_map.items(), key=lambda x: x[1]['employee'].name):
            employee = emp_data['employee']
            
            ws_matrix.cell(row=data_row, column=1, value=employee.name)
            ws_matrix.cell(row=data_row, column=2, value=employee.national_id or '-')
            ws_matrix.cell(row=data_row, column=3, value=employee.employee_id or '-')
            ws_matrix.cell(row=data_row, column=4, value=employee.job_title or '-')
            ws_matrix.cell(row=data_row, column=5, value=employee.mobile or '-')
            ws_matrix.cell(row=data_row, column=6, value=employee.location or '-')
            ws_matrix.cell(row=data_row, column=7, value=employee.project or '-')
            
            total_days = len(emp_data['dates'])
            ws_matrix.cell(row=data_row, column=8, value=total_days)
            ws_matrix.cell(row=data_row, column=8).font = Font(bold=True)
            ws_matrix.cell(row=data_row, column=8).alignment = Alignment(horizontal='center', vertical='center')
            
            for idx, date in enumerate(sorted_dates, start=9):
                status = emp_data['dates'].get(date, '')
                cell = ws_matrix.cell(row=data_row, column=idx)
                
                if status == 'present':
                    cell.value = 'P'
                    cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                    cell.font = Font(bold=True, color="155724")
                elif status == 'absent':
                    cell.value = 'A'
                    cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                    cell.font = Font(bold=True, color="721c24")
                elif status == 'leave':
                    cell.value = 'L'
                    cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
                    cell.font = Font(bold=True, color="856404")
                elif status == 'sick':
                    cell.value = 'S'
                    cell.fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
                    cell.font = Font(bold=True, color="0c5460")
                else:
                    cell.value = ''
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            for col in range(1, 9):
                cell = ws_matrix.cell(row=data_row, column=col)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            data_row += 1
        
        ws_matrix.column_dimensions['A'].width = 25
        ws_matrix.column_dimensions['B'].width = 15
        ws_matrix.column_dimensions['C'].width = 12
        ws_matrix.column_dimensions['D'].width = 20
        ws_matrix.column_dimensions['E'].width = 15
        ws_matrix.column_dimensions['F'].width = 15
        ws_matrix.column_dimensions['G'].width = 15
        ws_matrix.column_dimensions['H'].width = 8
        
        ws_matrix.freeze_panes = ws_matrix.cell(row=header_row + 2, column=9)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'تقرير_حضور_{department_name}_{start_date.strftime("%Y%m%d")}_إلى_{end_date.strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"خطأ في تصدير حضور الفترة: {str(e)}")
        import traceback
        traceback.print_exc()
        flash('حدث خطأ أثناء تصدير الملف', 'error')
        return redirect(url_for('attendance.department_attendance_view'))
