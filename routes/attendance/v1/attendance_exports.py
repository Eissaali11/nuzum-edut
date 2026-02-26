from flask import send_file, request
from io import BytesIO
from core.extensions import db
from sqlalchemy import or_

def export_department_data():
    """Export attendance rows to Excel for a department using optimized queries.

    Generates the file in-memory and streams it back. For very large exports
    we could switch to a temp file in `static/temp/` and stream from disk.
    """
    logger = __import__('logging').getLogger(__name__)
    try:
        # Import domain models
        from modules.employees.domain.models import Attendance, Employee, employee_departments
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from utils.date_converter import parse_date

        department_id = request.args.get('department_id', '')
        search_query = request.args.get('search_query', '').strip()
        status_filter = request.args.get('status_filter', '')
        start_date_str = request.args.get('start_date', '')
        end_date_str = request.args.get('end_date', '')

        if not start_date_str:
            from datetime import datetime, timedelta
            start_date = datetime.now().date() - timedelta(days=30)
        else:
            start_date = parse_date(start_date_str)

        if not end_date_str:
            from datetime import datetime
            end_date = datetime.now().date()
        else:
            end_date = parse_date(end_date_str)

        # Build query - indexing on (date, status) will be used by SQLite automatically
        q = db.session.query(Attendance).join(Employee).filter(
            Attendance.date >= start_date,
            Attendance.date <= end_date
        )

        if department_id:
            q = q.join(employee_departments).filter(employee_departments.c.department_id == int(department_id))

        if search_query:
            q = q.filter(
                or_(
                    Employee.name.ilike(f'%{search_query}%'),
                    Employee.employee_id.ilike(f'%{search_query}%'),
                    Employee.national_id.ilike(f'%{search_query}%')
                )
            )

        if status_filter:
            q = q.filter(Attendance.status == status_filter)

        q = q.order_by(Attendance.date.desc(), Employee.name)

        # Stream results into an Excel workbook in-memory
        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance'
        headers = ['date', 'employee_id', 'name', 'status', 'notes']
        for idx, h in enumerate(headers, start=1):
            ws[f'{get_column_letter(idx)}1'] = h

        row = 2
        for att in q.yield_per(200):
            emp = att.employee if hasattr(att, 'employee') else None
            ws[f'A{row}'] = att.date.isoformat() if att.date else ''
            ws[f'B{row}'] = emp.employee_id if emp else ''
            ws[f'C{row}'] = emp.name if emp else ''
            ws[f'D{row}'] = att.status
            ws[f'E{row}'] = getattr(att, 'notes', '')
            row += 1

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        logger.info('export_department_data: generated %d rows', row - 2)
        resp = send_file(bio, as_attachment=True, download_name='attendance_export.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp.headers['X-Attendance-Handler'] = 'MODULAR_v1'
        return resp

    except Exception as e:
        logger.exception('Error in export_department_data')
        from flask import abort
        abort(500, description=str(e))
