"""
Profitability dashboard/report/export routes.
"""

from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required

from models import Department
from services.accounting_service import AccountingService


profitability_bp = Blueprint('profitability', __name__, url_prefix='/accounting/profitability')


@profitability_bp.route('/')
@login_required
def dashboard():
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)
    dashboard_data = AccountingService.get_profitability_dashboard_data(month, year)

    return render_template(
        'accounting/profitability/dashboard.html',
        summary=dashboard_data['summary'],
        selected_month=dashboard_data['selected_month'],
        selected_year=dashboard_data['selected_year'],
        months=dashboard_data['months'],
        current_year=dashboard_data['current_year'],
    )


@profitability_bp.route('/report')
@login_required
def report():
    department_id = request.args.get('department_id', type=int)
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)

    if not department_id:
        flash('يرجى اختيار المشروع', 'warning')
        return redirect(url_for('profitability.dashboard', month=month, year=year))

    report_data = AccountingService.get_profitability_report_data(department_id, month, year)
    if not report_data:
        flash('المشروع غير موجود', 'danger')
        return redirect(url_for('profitability.dashboard', month=month, year=year))

    departments = Department.query.order_by(Department.name).all()

    return render_template(
        'accounting/profitability/report.html',
        result=report_data['result'],
        departments=departments,
        selected_month=report_data['selected_month'],
        selected_year=report_data['selected_year'],
        selected_department=report_data['selected_department'],
    )


@profitability_bp.route('/export-excel')
@login_required
def export_excel():
    department_id = request.args.get('department_id', type=int)
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)

    if department_id:
        report_data = AccountingService.get_profitability_report_data(department_id, month, year)
        if not report_data:
            flash('المشروع غير موجود', 'danger')
            return redirect(url_for('profitability.dashboard'))
        return _export_project_excel(report_data['result'])

    dashboard_data = AccountingService.get_profitability_dashboard_data(month, year)
    return _export_summary_excel(dashboard_data['summary'])


def _export_project_excel(result):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'تقرير الربحية'
    ws.sheet_view.rightToLeft = True

    navy = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    teal = PatternFill(start_color='0D7377', end_color='0D7377', fill_type='solid')
    light_gray = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    green_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    red_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6'),
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('A1:O1')
    ws['A1'] = f"تقرير ربحية المشروع — {result['department']['name']}"
    ws['A1'].font = Font(color='FFFFFF', bold=True, size=16)
    ws['A1'].fill = navy
    ws['A1'].alignment = center

    ws.merge_cells('A2:O2')
    ws['A2'] = f"العميل: {result['contract']['client_name']} | الفترة: {result['period']['month_name']} {result['period']['year']} | النوع: {result['contract']['contract_type_ar']}"
    ws['A2'].font = Font(color='FFFFFF', size=11)
    ws['A2'].fill = teal
    ws['A2'].alignment = center

    kpi_row = 4
    kpis = [
        ('إجمالي الإيرادات', f"{result['totals']['revenue']:,.2f} ريال"),
        ('إجمالي التكاليف', f"{result['totals']['total_cost']:,.2f} ريال"),
        ('صافي الربح', f"{result['totals']['profit']:,.2f} ريال"),
        ('هامش الربح', f"{result['overall_margin']:.1f}%"),
        ('إقامة + تأمين', f"{result['totals'].get('iqama_insurance', 0):,.0f} ريال"),
    ]
    for i, (label, value) in enumerate(kpis):
        col = 1 + i * 3
        ws.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col + 2)
        cell = ws.cell(row=kpi_row, column=col, value=f"{label}: {value}")
        cell.font = bold_font
        cell.alignment = center
        cell.fill = light_gray
        cell.border = thin_border

    header_row = 6
    headers = [
        '#', 'الموظف', 'الرقم الوظيفي', 'المسمى', 'سعر الفوترة',
        'الإيراد', 'الراتب', 'GOSI', 'تكلفة السيارة',
        'بدل السكن', 'إيجار السكن', 'فواتير خدمات',
        'إقامة/تأمين', 'مصاريف أخرى',
        'إجمالي التكلفة', 'صافي الربح', 'الهامش %'
    ]

    col_widths = [5, 22, 14, 16, 14, 14, 14, 10, 14, 12, 12, 12, 12, 12, 14, 14, 10]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = navy
        cell.alignment = center
        cell.border = thin_border

    num_cols = len(headers)

    for idx, emp in enumerate(result['employees'], 1):
        row = header_row + idx
        values = [
            idx, emp['employee_name'], emp['employee_code'], emp['job_title'],
            emp['billing_rate'],
            emp['revenue'], emp['salary_cost'], emp['gosi_cost'],
            emp['vehicle_cost'], emp['housing'],
            emp.get('housing_rent', 0), emp.get('utility_cost', 0),
            emp.get('iqama_cost', 0) + emp.get('insurance_cost', 0),
            emp['overhead'],
            emp['total_cost'], emp['net_profit'], f"{emp['margin_pct']:.1f}%"
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = center
            if idx % 2 == 0:
                cell.fill = light_gray

        profit_cell = ws.cell(row=row, column=num_cols - 1)
        if emp['net_profit'] >= 0:
            profit_cell.fill = green_fill
        else:
            profit_cell.fill = red_fill

    totals_row = header_row + len(result['employees']) + 1
    for col in range(1, num_cols + 1):
        c = ws.cell(row=totals_row, column=col)
        c.fill = teal
        c.border = thin_border
        c.font = Font(color='FFFFFF', bold=True)
        c.alignment = center

    ws.cell(row=totals_row, column=2, value='الإجمالي')
    ws.cell(row=totals_row, column=6, value=result['totals']['revenue'])
    ws.cell(row=totals_row, column=7, value=result['totals']['salary_cost'])
    ws.cell(row=totals_row, column=8, value=result['totals']['gosi_cost'])
    ws.cell(row=totals_row, column=9, value=result['totals']['vehicle_cost'])
    ws.cell(row=totals_row, column=10, value=result['totals']['overhead'])
    ws.cell(row=totals_row, column=11, value=result['totals'].get('housing_rent', 0))
    ws.cell(row=totals_row, column=12, value=result['totals'].get('utility_cost', 0))
    ws.cell(row=totals_row, column=13, value=result['totals'].get('iqama_insurance', 0))
    ws.cell(row=totals_row, column=15, value=result['totals']['total_cost'])
    ws.cell(row=totals_row, column=16, value=result['totals']['profit'])
    ws.cell(row=totals_row, column=17, value=f"{result['overall_margin']:.1f}%")

    sig_row = totals_row + 3
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=5)
    ws.cell(row=sig_row, column=1, value='المدير المالي: _______________').font = bold_font

    ws.merge_cells(start_row=sig_row, start_column=10, end_row=sig_row, end_column=15)
    ws.cell(row=sig_row, column=10, value='المدير العام: _______________').font = bold_font

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    dept_name = result['department']['name']
    period = f"{result['period']['month_name']}_{result['period']['year']}"
    filename = f'تقرير_ربحية_{dept_name}_{period}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def _export_summary_excel(summary):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'ملخص الربحية'
    ws.sheet_view.rightToLeft = True

    navy = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    teal = PatternFill(start_color='0D7377', end_color='0D7377', fill_type='solid')
    light_gray = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6'),
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('A1:H1')
    ws['A1'] = f"ملخص ربحية المشاريع — {summary['period']['month_name']} {summary['period']['year']}"
    ws['A1'].font = Font(color='FFFFFF', bold=True, size=16)
    ws['A1'].fill = navy
    ws['A1'].alignment = center

    headers = ['#', 'المشروع', 'العميل', 'عدد الموظفين', 'الإيرادات', 'التكاليف', 'صافي الربح', 'الهامش %']
    col_widths = [5, 22, 22, 14, 16, 16, 16, 12]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = navy
        cell.alignment = center
        cell.border = thin_border

    for idx, proj in enumerate(summary['projects'], 1):
        row = 3 + idx
        values = [
            idx, proj['department_name'], proj['client_name'], proj['employee_count'],
            proj['revenue'], proj['total_cost'], proj['profit'], f"{proj['margin']:.1f}%"
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = center
            if idx % 2 == 0:
                cell.fill = light_gray

    totals_row = 3 + len(summary['projects']) + 1
    ws.cell(row=totals_row, column=2, value='الإجمالي').font = Font(color='FFFFFF', bold=True)
    for col in range(1, 9):
        ws.cell(row=totals_row, column=col).fill = teal
        ws.cell(row=totals_row, column=col).border = thin_border
        ws.cell(row=totals_row, column=col).alignment = center

    ws.cell(row=totals_row, column=4, value=summary['totals']['employees']).font = Font(color='FFFFFF', bold=True)
    ws.cell(row=totals_row, column=5, value=summary['totals']['revenue']).font = Font(color='FFFFFF', bold=True)
    ws.cell(row=totals_row, column=6, value=summary['totals']['total_cost']).font = Font(color='FFFFFF', bold=True)
    ws.cell(row=totals_row, column=7, value=summary['totals']['profit']).font = Font(color='FFFFFF', bold=True)
    ws.cell(row=totals_row, column=8, value=f"{summary['totals']['margin']:.1f}%").font = Font(color='FFFFFF', bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    period = f"{summary['period']['month_name']}_{summary['period']['year']}"
    filename = f'ملخص_ربحية_المشاريع_{period}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
