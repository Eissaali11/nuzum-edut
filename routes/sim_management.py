from flask import Blueprint, render_template, request, flash, redirect, url_for, current_app, jsonify
from flask_login import login_required, current_user
from models import SimCard, ImportedPhoneNumber, Employee, Department, DeviceAssignment, MobileDevice, UserRole
from core.extensions import db
from datetime import datetime
import logging
from utils.audit_logger import log_activity
import pandas as pd
from flask import send_file
import os
import tempfile
from werkzeug.utils import secure_filename
from sqlalchemy import or_

sim_management_bp = Blueprint('sim_management', __name__)

@sim_management_bp.route('/')
@login_required
def index():
    """صفحة عرض جميع بطاقات SIM"""
    try:
        # فلاتر البحث
        department_filter = request.args.get('department_id', '')
        carrier_filter = request.args.get('carrier', '')
        status_filter = request.args.get('status', '')
        search_term = request.args.get('search', '')
        
        # الحصول على الأرقام المربوطة حالياً في DeviceAssignment
        # استعلام SQL مباشر للحصول على أرقام الهواتف المربوطة بالأجهزة
        device_assigned_sql = db.session.execute(db.text("""
            SELECT DISTINCT sc.phone_number 
            FROM device_assignments da 
            JOIN sim_cards sc ON da.sim_card_id = sc.id 
            WHERE da.is_active = true AND da.sim_card_id IS NOT NULL
        """))
        
        device_assigned_phone_numbers = set()
        for row in device_assigned_sql:
            device_assigned_phone_numbers.add(row.phone_number)
        
        # استعلام أساسي - استخدم SimCard
        query = SimCard.query
        
        # تطبيق فلاتر
        if search_term:
            query = query.filter(SimCard.phone_number.contains(search_term))
        
        if carrier_filter:
            query = query.filter(SimCard.carrier == carrier_filter)
            
        if status_filter == 'available':
            # الأرقام غير المربوطة (لا في employee_id ولا في DeviceAssignment)
            if device_assigned_phone_numbers:
                query = query.filter(
                    SimCard.employee_id.is_(None),
                    ~SimCard.phone_number.in_(device_assigned_phone_numbers)
                )
            else:
                query = query.filter(SimCard.employee_id.is_(None))
        elif status_filter == 'assigned':
            # الأرقام المربوطة (إما employee_id أو في DeviceAssignment)
            if device_assigned_phone_numbers:
                query = query.filter(
                    db.or_(
                        SimCard.employee_id.isnot(None),
                        SimCard.phone_number.in_(device_assigned_phone_numbers)
                    )
                )
            else:
                query = query.filter(SimCard.employee_id.isnot(None))
        
        # ترتيب البيانات
        sim_cards = query.order_by(SimCard.id.desc()).all()
        
        # إضافة معلومة هل الرقم مربوط في DeviceAssignment وجلب اسم الموظف
        for sim in sim_cards:
            sim.is_device_assigned = sim.phone_number in device_assigned_phone_numbers
            sim.device_employee_name = None
            
            # البحث عن الموظف المربوط بالجهاز إذا كان الرقم مربوط بجهاز
            if sim.is_device_assigned:
                # استعلام SQL مباشر للحصول على اسم الموظف
                employee_sql = db.session.execute(db.text("""
                    SELECT e.name 
                    FROM device_assignments da 
                    JOIN employee e ON da.employee_id = e.id 
                    WHERE da.sim_card_id = :sim_id AND da.is_active = true
                """), {'sim_id': sim.id})
                
                employee_row = employee_sql.fetchone()
                if employee_row:
                    sim.device_employee_name = employee_row.name
        
        # الحصول على قائمة الأقسام للفلترة
        departments = Department.query.order_by(Department.name).all()
        
        # إحصائيات محدثة تأخذ في الاعتبار DeviceAssignment
        total_sims = SimCard.query.count()
        sims_with_employee = SimCard.query.filter(SimCard.employee_id.isnot(None)).count()
        sims_in_device_assignment = len(device_assigned_phone_numbers)
        
        # الأرقام المتاحة = الأرقام بدون employee_id وليست في DeviceAssignment
        if device_assigned_phone_numbers:
            available_sims_count = SimCard.query.filter(
                SimCard.employee_id.is_(None),
                ~SimCard.phone_number.in_(device_assigned_phone_numbers)
            ).count()
        else:
            available_sims_count = SimCard.query.filter(SimCard.employee_id.is_(None)).count()
        
        stats = {
            'total_sims': total_sims,
            'available_sims': available_sims_count,
            'assigned_sims': sims_with_employee,
            'device_assigned_sims': sims_in_device_assignment,
            'stc_count': SimCard.query.filter(SimCard.carrier == 'STC').count(),
            'mobily_count': SimCard.query.filter(SimCard.carrier == 'موبايلي').count(),
            'zain_count': SimCard.query.filter(SimCard.carrier == 'زين').count(),
        }
        
        return render_template('sim_management/index.html', 
                             sim_cards=sim_cards, 
                             departments=departments,
                             stats=stats,
                             department_filter=department_filter,
                             carrier_filter=carrier_filter,
                             status_filter=status_filter,
                             search_term=search_term)
    
    except Exception as e:
        current_app.logger.error(f"Error in sim_management index: {str(e)}")
        flash('حدث خطأ أثناء تحميل البيانات', 'danger')
        return render_template('sim_management/index.html', 
                             sim_cards=[], 
                             departments=[],
                             stats={},
                             department_filter='',
                             carrier_filter='',
                             status_filter='',
                             search_term='')

@sim_management_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    """إضافة رقم جديد"""
    if request.method == 'POST':
        try:
            # الحصول على البيانات من النموذج
            phone_number = request.form.get('phone_number', '').strip()
            carrier = request.form.get('carrier', '').strip()
            plan_type = request.form.get('plan_type', '').strip()
            monthly_cost = request.form.get('monthly_cost')
            description = request.form.get('description', '').strip()
            
            # التحقق من البيانات المطلوبة
            if not phone_number:
                flash('رقم الهاتف مطلوب', 'danger')
                return render_template('sim_management/create.html')
            
            if not carrier:
                flash('شركة الاتصالات مطلوبة', 'danger')
                return render_template('sim_management/create.html')
            
            # التحقق من عدم وجود الرقم مسبقاً
            existing_sim = SimCard.query.filter_by(phone_number=phone_number).first()
            if existing_sim:
                flash('هذا الرقم موجود بالفعل في النظام', 'danger')
                return render_template('sim_management/create.html')
            
            # تحويل التكلفة الشهرية
            try:
                monthly_cost = float(monthly_cost) if monthly_cost else 0.0
            except ValueError:
                monthly_cost = 0.0
            
            # إنشاء رقم جديد
            new_sim = SimCard(
                phone_number=phone_number,
                carrier=carrier,
                plan_type=plan_type if plan_type else None,
                monthly_cost=monthly_cost,
                description=description if description else None
            )
            
            db.session.add(new_sim)
            db.session.commit()
            
            # تسجيل العملية
            log_activity(
                action="create",
                entity_type="SIM",
                entity_id=new_sim.id,
                details=f"إضافة رقم SIM جديد: {phone_number} - الشركة: {carrier}, النوع: {plan_type or 'غير محدد'}, التكلفة: {monthly_cost} ريال"
            )
            
            flash(f'تم إضافة الرقم {phone_number} بنجاح', 'success')
            return redirect(url_for('sim_management.index'))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error creating SIM card: {str(e)}")
            flash('حدث خطأ أثناء إضافة الرقم', 'danger')
            return render_template('sim_management/create.html')
    
    return render_template('sim_management/create.html')

@sim_management_bp.route('/edit/<int:sim_id>', methods=['GET', 'POST'])
@login_required
def edit(sim_id):
    """تعديل بيانات رقم"""
    sim_card = SimCard.query.get_or_404(sim_id)
    
    if request.method == 'POST':
        try:
            # الحصول على البيانات من النموذج
            phone_number = request.form.get('phone_number', '').strip()
            carrier = request.form.get('carrier', '').strip()
            plan_type = request.form.get('plan_type', '').strip()
            monthly_cost = request.form.get('monthly_cost')
            description = request.form.get('description', '').strip()
            
            # التحقق من البيانات المطلوبة
            if not phone_number:
                flash('رقم الهاتف مطلوب', 'danger')
                return render_template('sim_management/edit.html', sim_card=sim_card)
            
            if not carrier:
                flash('شركة الاتصالات مطلوبة', 'danger')
                return render_template('sim_management/edit.html', sim_card=sim_card)
            
            # التحقق من عدم تكرار الرقم (باستثناء الرقم الحالي)
            existing_sim = SimCard.query.filter(
                SimCard.phone_number == phone_number,
                SimCard.id != sim_id
            ).first()
            if existing_sim:
                flash('هذا الرقم موجود بالفعل في النظام', 'danger')
                return render_template('sim_management/edit.html', sim_card=sim_card)
            
            # تحويل التكلفة الشهرية
            try:
                monthly_cost = float(monthly_cost) if monthly_cost else 0.0
            except ValueError:
                monthly_cost = 0.0
            
            # تحديث البيانات
            sim_card.phone_number = phone_number
            sim_card.carrier = carrier
            sim_card.plan_type = plan_type if plan_type else None
            sim_card.monthly_cost = monthly_cost
            sim_card.description = description if description else None
            
            db.session.commit()
            
            # تسجيل العملية
            log_activity(
                action="update",
                entity_type="SIM",
                entity_id=sim_card.id,
                details=f"تعديل رقم SIM: {phone_number} - الشركة: {carrier}, النوع: {plan_type or 'غير محدد'}, التكلفة: {monthly_cost} ريال"
            )
            
            flash(f'تم تحديث الرقم {phone_number} بنجاح', 'success')
            return redirect(url_for('sim_management.index'))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error updating SIM card: {str(e)}")
            flash('حدث خطأ أثناء تحديث الرقم', 'danger')
            return render_template('sim_management/edit.html', sim_card=sim_card)
    
    return render_template('sim_management/edit.html', sim_card=sim_card)

@sim_management_bp.route('/assign/<int:sim_id>', methods=['GET', 'POST'])
@login_required
def assign(sim_id):
    """صفحة ربط رقم SIM بموظف"""
    sim_card = SimCard.query.get_or_404(sim_id)
    
    # التحقق من أن الرقم متاح للربط
    if sim_card.employee_id:
        flash('هذا الرقم مربوط بموظف بالفعل', 'warning')
        return redirect(url_for('sim_management.index'))
    
    if request.method == 'POST':
        try:
            employee_id = request.form.get('employee_id')
            
            if not employee_id:
                flash('يرجى اختيار موظف', 'danger')
                return render_template('sim_management/assign.html', 
                                     sim_card=sim_card, 
                                     employees=Employee.query.all())
            
            employee = Employee.query.get(employee_id)
            if not employee:
                flash('الموظف المختار غير موجود', 'danger')
                return render_template('sim_management/assign.html', 
                                     sim_card=sim_card, 
                                     employees=Employee.query.all())
            
            # التحقق من أن الموظف ليس مربوط برقم آخر
            existing_sim = SimCard.query.filter_by(employee_id=employee_id).first()
            if existing_sim:
                flash(f'الموظف {employee.name} مربوط بالفعل برقم {existing_sim.phone_number}. لا يمكن ربط موظف بأكثر من رقم واحد', 'warning')
                return render_template('sim_management/assign.html', 
                                     sim_card=sim_card, 
                                     employees=Employee.query.all(),
                                     departments=Department.query.all())
            
            # ربط الرقم بالموظف
            sim_card.employee_id = employee_id
            sim_card.assignment_date = datetime.now()
            db.session.commit()
            
            # تسجيل العملية
            log_activity(
                action="assign",
                entity_type="SIM",
                entity_id=sim_card.id,
                details=f"ربط رقم SIM {sim_card.phone_number} بالموظف {employee.name}"
            )
            
            flash(f'تم ربط الرقم {sim_card.phone_number} بالموظف {employee.name} بنجاح', 'success')
            return redirect(url_for('sim_management.index'))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error assigning SIM card: {str(e)}")
            flash('حدث خطأ أثناء ربط الرقم', 'danger')
    
    # جلب الموظفين النشطين فقط مع معلومات أرقام SIM المرتبطة
    employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
    departments = Department.query.all()
    
    # إضافة معلومات أرقام SIM للموظفين
    for employee in employees:
        employee.current_sim = SimCard.query.filter_by(employee_id=employee.id).first()
    
    return render_template('sim_management/assign.html', 
                         sim_card=sim_card, 
                         employees=employees,
                         departments=departments)

@sim_management_bp.route('/delete/<int:sim_id>', methods=['POST'])
@login_required
def delete(sim_id):
    """حذف رقم"""
    try:
        sim_card = SimCard.query.get_or_404(sim_id)
        phone_number = sim_card.phone_number
        
        # التحقق من عدم ربط الرقم بموظف
        if sim_card.employee_id:
            flash('لا يمكن حذف الرقم لأنه مرتبط بموظف', 'danger')
            return redirect(url_for('sim_management.index'))
        
        db.session.delete(sim_card)
        db.session.commit()
        
        # تسجيل العملية
        log_activity(
            action="delete",
            entity_type="SIM",
            entity_id=sim_id,
            details=f"حذف رقم SIM: {phone_number}"
        )
        
        flash(f'تم حذف الرقم {phone_number} بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting SIM card: {str(e)}")
        flash('حدث خطأ أثناء حذف الرقم', 'danger')
    
    return redirect(url_for('sim_management.index'))

@sim_management_bp.route('/assign/<int:sim_id>', methods=['POST'])
@login_required
def assign_to_employee(sim_id):
    """ربط رقم بموظف"""
    try:
        sim_card = SimCard.query.get_or_404(sim_id)
        employee_id = request.form.get('employee_id')
        
        if not employee_id:
            flash('يرجى اختيار موظف', 'danger')
            return redirect(url_for('sim_management.index'))
        
        employee = Employee.query.get_or_404(employee_id)
        
        # التحقق من عدم ربط الرقم بموظف آخر
        if sim_card.employee_id:
            flash('هذا الرقم مربوط بموظف آخر بالفعل', 'danger')
            return redirect(url_for('sim_management.index'))
        
        # ربط الرقم بالموظف
        sim_card.employee_id = employee_id
        sim_card.assignment_date = datetime.now()
        sim_card.status = 'مربوط'
        
        db.session.commit()
        
        # تسجيل العملية
        log_activity(
            action="update",
            entity_type="SIM",
            entity_id=sim_id,
            details=f"ربط رقم SIM بموظف - الرقم: {sim_card.phone_number}, الموظف: {employee.name} ({employee.employee_id})"
        )
        
        flash(f'تم ربط الرقم {sim_card.phone_number} بالموظف {employee.name} بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error assigning SIM card: {str(e)}")
        flash('حدث خطأ أثناء ربط الرقم', 'danger')
    
    return redirect(url_for('sim_management.index'))

@sim_management_bp.route('/unassign/<int:sim_id>', methods=['POST'])
@sim_management_bp.route('/unassign-from-employee/<int:sim_id>', methods=['POST'])
@login_required
def unassign_from_employee(sim_id):
    """فك ربط رقم من موظف"""
    try:
        sim_card = SimCard.query.get(sim_id)
        if not sim_card:
            flash(f'لم يتم العثور على بطاقة SIM برقم المعرف {sim_id}', 'danger')
            return redirect(url_for('sim_management.index'))
        
        if not sim_card.employee_id:
            flash('هذا الرقم غير مربوط بأي موظف', 'danger')
            return redirect(url_for('sim_management.index'))
        
        # الحصول على معلومات الموظف قبل فك الربط
        employee = Employee.query.get(sim_card.employee_id)
        employee_name = employee.name if employee else "غير معروف"
        employee_id_num = employee.employee_id if employee else "غير معروف"
        
        # فك الربط
        sim_card.employee_id = None
        sim_card.assignment_date = None
        sim_card.status = 'متاح'
        
        db.session.commit()
        
        # تسجيل العملية
        log_activity(
            action="update",
            entity_type="SIM",
            entity_id=sim_id,
            details=f"فك ربط رقم SIM من موظف - الرقم: {sim_card.phone_number}, الموظف السابق: {employee_name} ({employee_id_num})"
        )
        
        flash(f'تم فك ربط الرقم {sim_card.phone_number} من الموظف {employee_name} بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error unassigning SIM card: {str(e)}")
        flash('حدث خطأ أثناء فك ربط الرقم', 'danger')
    
    return redirect(url_for('sim_management.index'))

@sim_management_bp.route('/api/available', methods=['GET'])
@login_required
def api_available_sims():
    """API للحصول على الأرقام المتاحة"""
    try:
        available_sims = SimCard.query.filter(
            SimCard.employee_id.is_(None)
        ).order_by(SimCard.phone_number).all()
        
        sims_data = []
        for sim in available_sims:
            sims_data.append({
                'id': sim.id,
                'phone_number': sim.phone_number,
                'carrier': sim.carrier,
                'plan_type': sim.plan_type,
                'monthly_cost': sim.monthly_cost,
                'status_class': 'success'
            })
        
        return jsonify({
            'success': True,
            'sims': sims_data
        })
        
    except Exception as e:
        current_app.logger.error(f"Error getting available SIMs: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'حدث خطأ أثناء جلب البيانات'
        }), 500

@sim_management_bp.route('/export-excel')
@login_required
def export_excel():
    """تصدير أرقام SIM إلى ملف Excel مع دعم الفلاتر"""
    try:
        # إنشاء استعلام أساسي
        query = db.session.query(SimCard, Employee).outerjoin(
            Employee, SimCard.employee_id == Employee.id
        )
        
        # تطبيق الفلاتر من المعاملات
        search = request.args.get('search', '').strip()
        carrier_filter = request.args.get('carrier', '').strip()
        status_filter = request.args.get('status', '').strip()
        employee_filter = request.args.get('employee', '').strip()
        
        if search:
            query = query.filter(
                or_(
                    SimCard.phone_number.contains(search),
                    SimCard.carrier.contains(search),
                    SimCard.plan_type.contains(search),
                    SimCard.description.contains(search)
                )
            )
        
        if carrier_filter:
            query = query.filter(SimCard.carrier == carrier_filter)
            
        if status_filter:
            query = query.filter(SimCard.status == status_filter)
            
        if employee_filter == 'assigned':
            query = query.filter(SimCard.employee_id.isnot(None))
        elif employee_filter == 'unassigned':
            query = query.filter(SimCard.employee_id.is_(None))
        
        sim_cards = query.order_by(SimCard.id).all()
        
        # الحصول على الأرقام المربوطة حالياً في DeviceAssignment
        assigned_device_sims = db.session.query(DeviceAssignment.sim_card_id).filter(
            DeviceAssignment.is_active.is_(True),
            DeviceAssignment.sim_card_id.isnot(None)
        ).distinct().all()
        assigned_device_sim_ids = [id[0] for id in assigned_device_sims]
        
        # الحصول على أرقام SimCard المربوطة مع أرقام ImportedPhoneNumber المقابلة
        device_assigned_phone_numbers = set()
        if assigned_device_sim_ids:
            imported_numbers_in_devices = ImportedPhoneNumber.query.filter(
                ImportedPhoneNumber.id.in_(assigned_device_sim_ids)
            ).all()
            device_assigned_phone_numbers = {sim.phone_number for sim in imported_numbers_in_devices}

        # إعداد البيانات للتصدير
        data = []
        for sim_card, employee in sim_cards:
            # تحديد حالة الربط
            is_assigned_to_employee = bool(sim_card.employee_id and employee)
            is_assigned_to_device = sim_card.phone_number in device_assigned_phone_numbers
            
            # تحديد الحالة والموظف
            if is_assigned_to_employee:
                status = 'مربوط (موظف)'
                employee_name = employee.name
                employee_id_display = employee.employee_id if employee else ''
                departments_text = ', '.join([dept.name for dept in employee.departments]) if employee and employee.departments else ''
            elif is_assigned_to_device:
                status = 'مربوط (جهاز)'
                # البحث عن الموظف المربوط عبر الجهاز
                imported_num = ImportedPhoneNumber.query.filter_by(phone_number=sim_card.phone_number).first()
                if imported_num:
                    device_assignment = DeviceAssignment.query.filter_by(
                        sim_card_id=imported_num.id, is_active=True
                    ).first()
                    if device_assignment and device_assignment.employee_id:
                        device_employee = Employee.query.get(device_assignment.employee_id)
                        if device_employee:
                            employee_name = f"{device_employee.name} (عبر الجهاز)"
                            employee_id_display = device_employee.employee_id
                            departments_text = ', '.join([dept.name for dept in device_employee.departments]) if device_employee.departments else ''
                        else:
                            employee_name = 'مربوط بجهاز'
                            employee_id_display = ''
                            departments_text = ''
                    else:
                        employee_name = 'مربوط بجهاز'
                        employee_id_display = ''
                        departments_text = ''
                else:
                    employee_name = 'مربوط بجهاز'
                    employee_id_display = ''
                    departments_text = ''
            else:
                status = 'متاح'
                employee_name = ''
                employee_id_display = ''
                departments_text = ''

            data.append({
                'رقم الهاتف': sim_card.phone_number,
                'شركة الاتصالات': sim_card.carrier,
                'نوع الخطة': sim_card.plan_type or '',
                'التكلفة الشهرية': sim_card.monthly_cost or 0,
                'الحالة': status,
                'اسم الموظف': employee_name,
                'رقم الموظف': employee_id_display,
                'القسم': departments_text,
                'تاريخ الإنشاء': sim_card.created_at.strftime('%Y-%m-%d') if sim_card.created_at else '',
                'تاريخ الربط': sim_card.assigned_date.strftime('%Y-%m-%d') if sim_card.assigned_date else '',
                'الملاحظات': getattr(sim_card, 'notes', None) or getattr(sim_card, 'description', None) or ''
            })
        
        # إنشاء DataFrame
        df = pd.DataFrame(data)
        
        # حفظ في مجلد دائم
        export_folder = os.path.join(UPLOAD_FOLDER, 'exports', 'sim')
        os.makedirs(export_folder, exist_ok=True)
        
        temp_filename = os.path.join(export_folder, f"sim_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        # كتابة البيانات إلى Excel بشكل مبسط وآمن
        try:
            # حفظ بسيط أولاً
            df.to_excel(temp_filename, index=False, sheet_name='SIM_Cards', engine='openpyxl')
            
            # التحقق من أن الملف تم إنشاؤه بنجاح
            if not os.path.exists(temp_filename) or os.path.getsize(temp_filename) == 0:
                raise Exception("فشل في إنشاء الملف")
                
            current_app.logger.info(f"تم إنشاء ملف Excel بحجم: {os.path.getsize(temp_filename)} بايت")
            
        except Exception as excel_error:
            current_app.logger.error(f"خطأ في كتابة Excel: {str(excel_error)}")
            # محاولة أخيرة مع CSV
            csv_temp = temp_filename.replace('.xlsx', '.csv')
            df.to_csv(csv_temp, index=False, encoding='utf-8-sig')
            temp_filename = csv_temp
        
        # تسجيل العملية
        log_activity(
            action="export",
            entity_type="SIM",
            details=f"تصدير {len(data)} رقم SIM إلى Excel"
        )
        
        # تحديد نوع الملف ومحاولة التحميل الآمن
        is_csv = temp_filename.endswith('.csv')
        file_extension = '.csv' if is_csv else '.xlsx'
        download_name = f'sim_cards_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}{"_filtered" if search or carrier_filter or status_filter or employee_filter else ""}{file_extension}'
        mimetype = 'text/csv' if is_csv else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        try:
            # إرسال الملف
            response = send_file(
                temp_filename,
                as_attachment=True,
                download_name=download_name,
                mimetype=mimetype
            )
            
            # تنظيف الملف المؤقت بعد الإرسال
            def cleanup_temp_file():
                try:
                    if os.path.exists(temp_filename):
                        os.unlink(temp_filename)
                except:
                    pass
            
            # تسجيل نجاح العملية
            current_app.logger.info(f"تم تصدير {len(data)} سجل SIM بنجاح كـ {file_extension}")
            
            # إضافة cleanup للملف بعد الإرسال
            response.call_on_close(cleanup_temp_file)
            
            return response
            
        except Exception as send_error:
            current_app.logger.error(f"خطأ في إرسال الملف: {str(send_error)}")
            # تنظيف الملف في حالة الخطأ
            try:
                if os.path.exists(temp_filename):
                    os.unlink(temp_filename)
            except:
                pass
            raise send_error
        
    except Exception as e:
        current_app.logger.error(f"Error exporting SIM cards: {str(e)}")
        flash('حدث خطأ أثناء تصدير البيانات', 'danger')
        return redirect(url_for('sim_management.index'))

@sim_management_bp.route('/import-excel', methods=['GET', 'POST'])
@login_required
def import_excel():
    """استيراد أرقام SIM من ملف Excel"""
    if request.method == 'POST':
        try:
            # التحقق من وجود الملف
            if 'excel_file' not in request.files:
                flash('يرجى اختيار ملف Excel', 'danger')
                return render_template('sim_management/import_excel.html')
            
            file = request.files['excel_file']
            if file.filename == '':
                flash('يرجى اختيار ملف', 'danger')
                return render_template('sim_management/import_excel.html')
            
            # التحقق من نوع الملف
            if not file.filename.lower().endswith(('.xlsx', '.xls')):
                flash('يرجى اختيار ملف Excel صحيح (.xlsx أو .xls)', 'danger')
                return render_template('sim_management/import_excel.html')
            
            # قراءة الملف
            df = pd.read_excel(file)
            
            # التحقق من وجود الأعمدة المطلوبة
            required_columns = ['رقم الهاتف', 'شركة الاتصالات']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                flash(f'الأعمدة التالية مفقودة: {", ".join(missing_columns)}', 'danger')
                return render_template('sim_management/import_excel.html')
            
            imported_count = 0
            skipped_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    phone_number = str(row['رقم الهاتف']).strip()
                    carrier = str(row['شركة الاتصالات']).strip()
                    
                    # تخطي الصفوف الفارغة
                    if not phone_number or phone_number == 'nan':
                        continue
                    
                    # التحقق من وجود الرقم
                    existing_sim = SimCard.query.filter_by(phone_number=phone_number).first()
                    if existing_sim:
                        skipped_count += 1
                        continue
                    
                    # إنشاء رقم جديد
                    new_sim = SimCard(
                        phone_number=phone_number,
                        carrier=carrier,
                        plan_type=str(row.get('نوع الخطة', '')).strip() if pd.notna(row.get('نوع الخطة')) else None,
                        monthly_cost=float(row.get('التكلفة الشهرية', 0)) if pd.notna(row.get('التكلفة الشهرية')) else 0.0,
                        description=str(row.get('الوصف', '')).strip() if pd.notna(row.get('الوصف')) else None
                    )
                    
                    db.session.add(new_sim)
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"صف {index + 2}: {str(e)}")
                    continue
            
            # حفظ التغييرات
            db.session.commit()
            
            # تسجيل العملية
            log_activity(
                action="import",
                entity_type="SIM",
                details=f"استيراد {imported_count} رقم SIM من Excel، تم تخطي {skipped_count} رقم موجود مسبقاً"
            )
            
            # إعداد رسالة النتيجة
            if imported_count > 0:
                flash(f'تم استيراد {imported_count} رقم بنجاح', 'success')
            if skipped_count > 0:
                flash(f'تم تخطي {skipped_count} رقم موجود مسبقاً', 'info')
            if errors:
                flash(f'حدثت أخطاء في {len(errors)} صف', 'warning')
            
            return redirect(url_for('sim_management.index'))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error importing SIM cards: {str(e)}")
            flash('حدث خطأ أثناء استيراد البيانات', 'danger')
            return render_template('sim_management/import_excel.html')
    
    return render_template('sim_management/import_excel.html')

@sim_management_bp.route('/details/<int:sim_id>')
@login_required
def details(sim_id):
    """صفحة عرض تفاصيل بطاقة SIM"""
    try:
        sim = SimCard.query.get_or_404(sim_id)
        
        # البحث عن ImportedPhoneNumber المقابل لهذا الرقم
        imported_number = ImportedPhoneNumber.query.filter_by(phone_number=sim.phone_number).first()
        device_assignment = None
        employee = None
        device = None
        
        if imported_number:
            # البحث عن DeviceAssignment باستخدام sim_card_id
            device_assignment = DeviceAssignment.query.filter_by(
                sim_card_id=imported_number.id,
                is_active=True
            ).first()
            
            if device_assignment:
                if device_assignment.employee_id:
                    employee = Employee.query.get(device_assignment.employee_id)
                if device_assignment.device_id:
                    device = MobileDevice.query.get(device_assignment.device_id)
        
        # البحث عن الموظف المربوط مباشرة في SimCard إذا لم نجد في DeviceAssignment
        if not employee and sim.employee_id:
            employee = Employee.query.get(sim.employee_id)
        
        # البحث عن الجهاز المربوط مباشرة في SimCard
        if not device and hasattr(sim, 'device_id') and sim.device_id:
            device = MobileDevice.query.get(sim.device_id)
        
        # البحث عن جميع الأجهزة المربوطة بهذا الرقم
        if not device and imported_number:
            all_assignments = DeviceAssignment.query.filter_by(
                sim_card_id=imported_number.id
            ).all()
            
            for assignment in all_assignments:
                if assignment.device_id and assignment.is_active:
                    device = MobileDevice.query.get(assignment.device_id)
                    if device:
                        device_assignment = assignment
                        break
        
        return render_template('sim_management/details.html', 
                             sim=sim, 
                             device_assignment=device_assignment,
                             employee=employee, 
                             device=device)
                             
    except Exception as e:
        current_app.logger.error(f"Error loading SIM details: {str(e)}")
        flash('حدث خطأ أثناء تحميل التفاصيل', 'danger')
        return redirect(url_for('sim_management.index'))

@sim_management_bp.route('/export-details-excel/<int:sim_id>')
@login_required
def export_sim_details_excel(sim_id):
    """تصدير تفاصيل SIM محددة إلى ملف Excel منسق"""
    try:
        # جلب تفاصيل SIM
        sim = SimCard.query.get_or_404(sim_id)
        
        # البحث عن ImportedPhoneNumber المقابل لهذا الرقم
        imported_number = ImportedPhoneNumber.query.filter_by(phone_number=sim.phone_number).first()
        device_assignment = None
        employee = None
        device = None
        
        if imported_number:
            # البحث عن DeviceAssignment باستخدام sim_card_id
            device_assignment = DeviceAssignment.query.filter_by(
                sim_card_id=imported_number.id,
                is_active=True
            ).first()
            
            if device_assignment:
                if device_assignment.employee_id:
                    employee = Employee.query.get(device_assignment.employee_id)
                if device_assignment.device_id:
                    device = MobileDevice.query.get(device_assignment.device_id)
        
        # البحث عن الموظف المربوط مباشرة في SimCard إذا لم نجد في DeviceAssignment
        if not employee and sim.employee_id:
            employee = Employee.query.get(sim.employee_id)
        
        # إعداد البيانات للتصدير
        data = []
        
        # معلومات أساسية
        data.append(['القسم', 'المعلومات'])
        data.append(['', ''])
        data.append(['المعلومات الأساسية', ''])
        data.append(['رقم الهاتف', sim.phone_number])
        data.append(['شركة الاتصالات', sim.carrier or 'غير محدد'])
        data.append(['نوع الخطة', sim.plan_type or 'غير محدد'])
        data.append(['التكلفة الشهرية', f'{sim.monthly_cost} ريال' if sim.monthly_cost else 'غير محدد'])
        data.append(['تاريخ التفعيل', sim.activation_date.strftime('%Y-%m-%d') if sim.activation_date else 'غير محدد'])
        data.append(['تاريخ الإنشاء', sim.created_at.strftime('%Y-%m-%d') if sim.created_at else 'غير محدد'])
        data.append(['تاريخ الربط', sim.assigned_date.strftime('%Y-%m-%d') if sim.assigned_date else 'غير محدد'])
        data.append(['الوصف', sim.description or 'لا يوجد'])
        data.append(['', ''])
        
        # معلومات الموظف المربوط
        if employee:
            data.append(['معلومات الموظف المربوط', ''])
            data.append(['اسم الموظف', employee.name])
            data.append(['رقم الموظف', employee.employee_id or 'غير محدد'])
            data.append(['المنصب', employee.job_title or 'غير محدد'])
            data.append(['الأقسام', ', '.join([dept.name for dept in employee.departments]) if employee.departments else 'غير محدد'])
            data.append(['هاتف العمل', employee.mobile or 'غير محدد'])
            data.append(['الهاتف الشخصي', employee.mobilePersonal or 'غير محدد'])
            data.append(['البريد الإلكتروني', employee.email or 'غير محدد'])
            data.append(['تاريخ التوظيف', employee.hire_date.strftime('%Y-%m-%d') if hasattr(employee, 'hire_date') and employee.hire_date else 'غير محدد'])
            data.append(['', ''])
        
        # معلومات الجهاز المربوط
        if device:
            data.append(['معلومات الجهاز المربوط', ''])
            data.append(['الماركة', device.device_brand or 'غير محدد'])
            data.append(['الطراز', device.model or 'غير محدد'])
            data.append(['رقم IMEI', device.imei or 'غير محدد'])
            data.append(['رقم المسلسل', device.serial_number or 'غير محدد'])
            data.append(['حالة الجهاز', device.status or 'غير محدد'])
            data.append(['تاريخ الشراء', device.purchase_date.strftime('%Y-%m-%d') if hasattr(device, 'purchase_date') and device.purchase_date else 'غير محدد'])
            data.append(['تاريخ انتهاء الضمان', device.warranty_expiry.strftime('%Y-%m-%d') if hasattr(device, 'warranty_expiry') and device.warranty_expiry else 'غير محدد'])
            data.append(['ملاحظات الجهاز', device.notes or 'لا توجد'])
            data.append(['', ''])
        
        # معلومات الربط
        if device_assignment:
            data.append(['معلومات الربط', ''])
            data.append(['تاريخ الربط', device_assignment.assignment_date.strftime('%Y-%m-%d %H:%M') if device_assignment.assignment_date else 'غير محدد'])
            data.append(['الحالة', 'نشط' if device_assignment.is_active else 'غير نشط'])
            data.append(['ملاحظات الربط', device_assignment.notes or 'لا توجد'])
            data.append(['', ''])
        
        # إحصائيات
        data.append(['إحصائيات', ''])
        data.append(['حالة الرقم', 'مربوط' if (employee or device) else 'متاح'])
        data.append(['نوع الربط', 'موظف وجهاز' if (employee and device) else ('موظف فقط' if employee else ('جهاز فقط' if device else 'غير مربوط'))])
        data.append(['تاريخ التصدير', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        
        # إنشاء DataFrame
        df = pd.DataFrame(data, columns=['البيان', 'القيمة'])
        
        # حفظ في مجلد دائم
        export_folder = os.path.join(UPLOAD_FOLDER, 'exports', 'sim_details')
        os.makedirs(export_folder, exist_ok=True)
        
        temp_filename = os.path.join(export_folder, f"sim_detail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        # كتابة البيانات إلى Excel مع تنسيق محسن
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = f'SIM {sim.phone_number}'
        
        # إضافة البيانات إلى الورقة
        for row_data in data:
            ws.append(row_data)
            
        # تعديل عرض الأعمدة
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        
        # تنسيق رؤوس الأقسام
        section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        section_font = Font(color='FFFFFF', bold=True, size=12)
        
        # تنسيق البيانات العادية
        data_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_num, row in enumerate(ws.iter_rows(min_row=1), start=1):
            cell_value = row[0].value
            if cell_value and len(row) > 1 and (not row[1].value or row[1].value == ''):
                # هذا رأس قسم
                for cell in row:
                    cell.fill = section_fill
                    cell.font = section_font
                    cell.alignment = Alignment(horizontal='center')
            else:
                # بيانات عادية
                for cell in row:
                    cell.font = data_font
                    cell.border = border
                    if cell.column == 1:  # عمود البيان
                        cell.alignment = Alignment(horizontal='right')
                    else:  # عمود القيمة
                        cell.alignment = Alignment(horizontal='right')
        
        # حفظ الملف
        wb.save(temp_filename)
        
        # تسجيل العملية
        log_activity(
            action="export_details",
            entity_type="SIM",
            entity_id=sim_id,
            details=f"تصدير تفاصيل SIM {sim.phone_number} إلى Excel"
        )
        
        # إرسال الملف
        filename = f'sim_details_{sim.phone_number.replace("+", "").replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            temp_filename,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        current_app.logger.error(f"Error exporting SIM details: {str(e)}")
        flash('حدث خطأ أثناء تصدير التفاصيل', 'danger')
        return redirect(url_for('sim_management.details', sim_id=sim_id))