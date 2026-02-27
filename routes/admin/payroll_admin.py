"""
Payroll Admin Routes
مسارات إدارة الرواتب
"""
from flask import Blueprint, render_template, request, jsonify, send_file, redirect, url_for, flash, current_app, render_template_string, make_response
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from decimal import Decimal
from sqlalchemy import func, and_, or_
from sqlalchemy.orm import joinedload
from io import BytesIO, StringIO
import csv
import threading
import os
from urllib.parse import quote_plus
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from core.extensions import db
from models import Employee, User, Module, Permission, Department, Document, VehicleHandover, DeviceAssignment, Attendance
from modules.payroll.domain.models import PayrollRecord, PayrollConfiguration, BankTransferFile
from modules.payroll.application.payroll_processor import PayrollProcessor
from modules.payroll.application.bank_transfer_generator import BankTransferGenerator
from services.finance_bridge import ERPNextClient, ERPNextBridgeError
from services.finance_bridge_app_service import FinanceBridgeSettingsService


payroll_bp = Blueprint('payroll', __name__)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def _ensure_default_departments():
    """التأكد من وجود الأقسام الافتراضية"""
    if Department.query.count() > 0:
        return  # الأقسام موجودة بالفعل
    
    default_departments = [
        {'name': 'الموارد البشرية', 'code': 'HR'},
        {'name': 'المبيعات', 'code': 'SALES'},
        {'name': 'التسويق', 'code': 'MARKETING'},
        {'name': 'تكنولوجيا المعلومات', 'code': 'IT'},
        {'name': 'العمليات', 'code': 'OPS'},
        {'name': 'التمويل', 'code': 'FINANCE'},
        {'name': 'الإدارة', 'code': 'ADMIN'},
        {'name': 'خدمة العملاء', 'code': 'CS'},
    ]
    
    try:
        for dept_data in default_departments:
            dept = Department(
                name=dept_data['name'],
                code=dept_data['code'],
                status='active'
            )
            db.session.add(dept)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        # لا نرفع الخطأ، نترك التطبيق يعمل


# ═══════════════════════════════════════════════════════════════════════════════
# MIDDLEWARE & PERMISSIONS
# ═══════════════════════════════════════════════════════════════════════════════

def check_payroll_access():
    """التحقق من صلاحيات الوصول لقسم الرواتب"""
    if not current_user.is_authenticated:
        return False

    role_value = getattr(current_user.role, 'value', current_user.role)
    if getattr(current_user, 'is_admin', False) or str(role_value).lower() == 'admin':
        return True
    
    # التحقق من الصلاحية
    payroll_module = Module.query.filter_by(name='payroll').first()
    if not payroll_module:
        return False
    
    permission = Permission.query.filter(
        and_(
            Permission.user_id == current_user.id,
            Permission.module_id == payroll_module.id
        )
    ).first()
    
    return permission is not None


def _parse_period_inputs():
    month = request.values.get('month', datetime.now().month, type=int)
    year = request.values.get('year', datetime.now().year, type=int)
    department_id = request.values.get('department_id', type=int)
    return month, year, department_id


def _load_compliance_rows(days_ahead=90):
    today = date.today()
    threshold = today + timedelta(days=days_ahead)

    rows = (
        db.session.query(Document, Employee)
        .join(Employee, Document.employee_id == Employee.id)
        .filter(Employee.status == 'active')
        .filter(Document.expiry_date.isnot(None))
        .filter(Document.expiry_date >= today)
        .filter(Document.expiry_date <= threshold)
        .order_by(Document.expiry_date.asc())
        .all()
    )

    payload = []
    for doc, employee in rows:
        days_left = (doc.expiry_date - today).days
        if days_left <= 30:
            severity = 'danger'
        elif days_left <= 60:
            severity = 'warning'
        else:
            severity = 'info'

        payload.append({
            'document_id': doc.id,
            'employee_id': employee.id,
            'employee_name': employee.name,
            'employee_mobile': employee.mobile,
            'employee_email': employee.email,
            'document_type': doc.document_type,
            'document_number': doc.document_number,
            'expiry_date': doc.expiry_date,
            'days_left': days_left,
            'severity': severity,
        })

    return payload


def _build_wps_csv(records, payment_date_value):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'RecordType',
        'EmployeeID',
        'EmployeeName',
        'IBAN',
        'NetAmount',
        'Currency',
        'PaymentDate',
        'Reference',
    ])

    for payroll in records:
        employee = payroll.employee
        writer.writerow([
            'SALARY',
            employee.employee_id,
            employee.name,
            getattr(employee, 'bank_iban', '') or '',
            f"{float(payroll.net_payable or 0):.2f}",
            'SAR',
            payment_date_value.strftime('%Y-%m-%d'),
            f"NUZUM-{payroll.pay_period_year}-{str(payroll.pay_period_month).zfill(2)}-{payroll.id}",
        ])

    return output.getvalue().encode('utf-8-sig')


def _create_erp_payroll_journal(month, year, expense_account, payable_account, cost_center=None):
    settings_data = FinanceBridgeSettingsService.load_settings()
    client = ERPNextClient(config_overrides=settings_data)
    if not client.is_configured():
        raise ERPNextBridgeError('بيانات ربط ERPNext غير مكتملة (Base URL / API Key / API Secret).')

    payroll_records = PayrollRecord.query.options(
        joinedload(PayrollRecord.employee)
    ).filter(
        PayrollRecord.pay_period_month == month,
        PayrollRecord.pay_period_year == year,
        PayrollRecord.payment_status.in_(['approved', 'pending', 'paid'])
    ).all()

    total_net = sum(Decimal(str(record.net_payable or 0)) for record in payroll_records)
    if total_net <= 0:
        raise ERPNextBridgeError('لا توجد رواتب مستحقة للترحيل في الفترة المحددة.')

    posting_date = date(year, month, 1)
    line_debit = {
        'account': expense_account,
        'debit_in_account_currency': float(total_net),
        'credit_in_account_currency': 0,
        'user_remark': f'Payroll expense {month}/{year}',
    }
    line_credit = {
        'account': payable_account,
        'debit_in_account_currency': 0,
        'credit_in_account_currency': float(total_net),
        'user_remark': f'Payroll payable {month}/{year}',
    }

    if cost_center:
        line_debit['cost_center'] = cost_center
        line_credit['cost_center'] = cost_center

    payload = {
        'posting_date': posting_date.isoformat(),
        'voucher_type': 'Journal Entry',
        'title': f'Payroll {month}/{year}',
        'user_remark': f'NUZUM payroll automation sync for {month}/{year}',
        'accounts': [line_debit, line_credit],
    }

    created = client.post_request('Journal Entry', payload)
    return {
        'journal_name': (created or {}).get('name') or (created or {}).get('journal_entry') or 'N/A',
        'total_net': float(total_net),
        'records_count': len(payroll_records),
    }


def _get_notification_service():
    try:
        from services.notification_service import NotificationService
        return NotificationService()
    except Exception:
        class _NoopNotificationService:
            def send_sms(self, *_args, **_kwargs):
                return False, 'twilio_not_configured'

        return _NoopNotificationService()


def _parse_month_year(month_year: str):
    raw = str(month_year or '').strip()
    if not raw:
        return None, None

    candidates = ['%Y-%m', '%m-%Y', '%Y/%m', '%m/%Y']
    for fmt in candidates:
        try:
            parsed = datetime.strptime(raw, fmt)
            return parsed.month, parsed.year
        except ValueError:
            continue
    return None, None


def _employee_in_user_scope(user, employee):
    if user._is_admin_role():
        return True

    assigned_department_id = getattr(user, 'assigned_department_id', None)
    if assigned_department_id:
        dept_ids = set()
        if getattr(employee, 'department_id', None):
            dept_ids.add(employee.department_id)
        if hasattr(employee, 'departments') and employee.departments:
            dept_ids.update([dep.id for dep in employee.departments])
        if assigned_department_id not in dept_ids:
            return False

    user_employee = getattr(user, 'employee', None)
    user_project = (getattr(user_employee, 'project', None) or '').strip().lower() if user_employee else ''
    employee_project = (getattr(employee, 'project', None) or '').strip().lower()
    if user_project and employee_project and user_project != employee_project:
        return False

    return True


def _normalize_identity(value):
    return ''.join(ch for ch in str(value or '') if ch.isdigit())


def _employee_identity_value(employee):
    for attr_name in ['iqama_number', 'residency_number', 'national_id', 'id_number', 'identity_number']:
        candidate = _normalize_identity(getattr(employee, attr_name, None))
        if candidate:
            return candidate
    return _normalize_identity(getattr(employee, 'employee_id', ''))


def _get_payslip_serializer():
    secret_key = (
        current_app.config.get('SECRET_KEY')
        or os.getenv('SESSION_SECRET')
        or 'dev-secret'
    )
    return URLSafeTimedSerializer(secret_key=secret_key, salt='nuzum-payslip-secure-link')


def _build_secure_payslip_url(employee, month, year):
    identity_value = _employee_identity_value(employee)
    payload = {
        'employee_id': int(employee.id),
        'month': int(month),
        'year': int(year),
        'id_tail': identity_value[-4:] if identity_value else '',
    }
    token = _get_payslip_serializer().dumps(payload)
    return url_for('payroll.generate_secure_payslip', token=token, _external=True)


def _small_number_to_words_en(number_value: int):
    ones = [
        'zero', 'one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine',
        'ten', 'eleven', 'twelve', 'thirteen', 'fourteen', 'fifteen', 'sixteen',
        'seventeen', 'eighteen', 'nineteen'
    ]
    tens = ['', '', 'twenty', 'thirty', 'forty', 'fifty', 'sixty', 'seventy', 'eighty', 'ninety']

    if number_value < 20:
        return ones[number_value]
    if number_value < 100:
        return tens[number_value // 10] + ('' if number_value % 10 == 0 else f"-{ones[number_value % 10]}")
    if number_value < 1000:
        rem = number_value % 100
        return ones[number_value // 100] + ' hundred' + ('' if rem == 0 else f' {_small_number_to_words_en(rem)}')
    if number_value < 1000000:
        rem = number_value % 1000
        return _small_number_to_words_en(number_value // 1000) + ' thousand' + ('' if rem == 0 else f' {_small_number_to_words_en(rem)}')
    rem = number_value % 1000000
    return _small_number_to_words_en(number_value // 1000000) + ' million' + ('' if rem == 0 else f' {_small_number_to_words_en(rem)}')


def _small_number_to_words_ar(number_value: int):
    mapping = {
        0: 'صفر', 1: 'واحد', 2: 'اثنان', 3: 'ثلاثة', 4: 'أربعة', 5: 'خمسة',
        6: 'ستة', 7: 'سبعة', 8: 'ثمانية', 9: 'تسعة', 10: 'عشرة'
    }
    if number_value in mapping:
        return mapping[number_value]
    if number_value < 20:
        return f"{mapping[number_value - 10]} عشر"
    if number_value < 100:
        tens_map = {
            20: 'عشرون', 30: 'ثلاثون', 40: 'أربعون', 50: 'خمسون',
            60: 'ستون', 70: 'سبعون', 80: 'ثمانون', 90: 'تسعون'
        }
        tens_val = (number_value // 10) * 10
        ones_val = number_value % 10
        if ones_val == 0:
            return tens_map[tens_val]
        return f"{mapping[ones_val]} و{tens_map[tens_val]}"
    if number_value < 1000:
        hundreds_map = {
            1: 'مائة', 2: 'مائتان', 3: 'ثلاثمائة', 4: 'أربعمائة', 5: 'خمسمائة',
            6: 'ستمائة', 7: 'سبعمائة', 8: 'ثمانمائة', 9: 'تسعمائة'
        }
        hundreds_val = number_value // 100
        rem = number_value % 100
        base = hundreds_map.get(hundreds_val, '')
        return base if rem == 0 else f"{base} و{_small_number_to_words_ar(rem)}"
    if number_value < 1000000:
        thousands = number_value // 1000
        rem = number_value % 1000
        if thousands == 1:
            base = 'ألف'
        elif thousands == 2:
            base = 'ألفان'
        else:
            base = f"{_small_number_to_words_ar(thousands)} ألف"
        return base if rem == 0 else f"{base} و{_small_number_to_words_ar(rem)}"
    millions = number_value // 1000000
    rem = number_value % 1000000
    if millions == 1:
        base = 'مليون'
    elif millions == 2:
        base = 'مليونان'
    else:
        base = f"{_small_number_to_words_ar(millions)} مليون"
    return base if rem == 0 else f"{base} و{_small_number_to_words_ar(rem)}"


def _tafqeet_sar(amount_value):
    amount_decimal = Decimal(str(amount_value or 0)).quantize(Decimal('0.01'))
    integer_part = int(amount_decimal)
    halala_part = int((amount_decimal - Decimal(integer_part)) * 100)

    en_words = _small_number_to_words_en(integer_part)
    ar_words = _small_number_to_words_ar(integer_part)

    if halala_part > 0:
        en_words += f" and {_small_number_to_words_en(halala_part)} halalas"
        ar_words += f" و{_small_number_to_words_ar(halala_part)} هللة"

    return {
        'en': f"{en_words} Saudi Riyals only",
        'ar': f"{ar_words} ريال سعودي فقط لا غير",
    }


def _build_payslip_payload(employee, payroll_record, month, year):
    attendance_rows = Attendance.query.filter(
        Attendance.employee_id == employee.id,
        func.extract('month', Attendance.date) == month,
        func.extract('year', Attendance.date) == year,
    ).all()

    attendance_stats = {
        'present': sum(1 for row in attendance_rows if row.status == 'present'),
        'absent': sum(1 for row in attendance_rows if row.status == 'absent'),
        'leave': sum(1 for row in attendance_rows if row.status == 'leave'),
        'sick_leave': sum(1 for row in attendance_rows if row.status == 'sick_leave'),
        'unpaid_leave': sum(1 for row in attendance_rows if row.status == 'unpaid_leave'),
    }

    overtime_amount = float((payroll_record.hourly_rate or 0) * (payroll_record.overtime_hours or 0))

    earnings = [
        {'key': 'basic', 'label_ar': 'الراتب الأساسي', 'label_en': 'Basic Salary', 'amount': float(payroll_record.basic_salary or 0)},
        {'key': 'housing', 'label_ar': 'بدل السكن', 'label_en': 'Housing Allowance', 'amount': float(payroll_record.housing_allowance or 0)},
        {'key': 'transportation', 'label_ar': 'بدل النقل', 'label_en': 'Transportation Allowance', 'amount': float(payroll_record.transportation or 0)},
        {'key': 'meal', 'label_ar': 'بدل الطعام', 'label_en': 'Meal Allowance', 'amount': float(payroll_record.meal_allowance or 0)},
        {'key': 'performance_bonus', 'label_ar': 'حافز الأداء', 'label_en': 'Performance Bonus', 'amount': float(payroll_record.performance_bonus or 0)},
        {'key': 'attendance_bonus', 'label_ar': 'حافز الحضور', 'label_en': 'Attendance Bonus', 'amount': float(payroll_record.attendance_bonus or 0)},
        {'key': 'other_allowances', 'label_ar': 'بدلات أخرى', 'label_en': 'Other Allowances', 'amount': float(payroll_record.other_allowances or 0)},
        {'key': 'overtime', 'label_ar': 'العمل الإضافي', 'label_en': 'Overtime', 'amount': overtime_amount},
    ]

    deductions = [
        {'key': 'absence', 'label_ar': 'خصم الغياب', 'label_en': 'Absence Deduction', 'amount': float(payroll_record.absence_deduction or 0)},
        {'key': 'late', 'label_ar': 'خصم التأخير', 'label_en': 'Late Deduction', 'amount': float(payroll_record.late_deduction or 0)},
        {'key': 'early_leave', 'label_ar': 'خصم الانصراف المبكر', 'label_en': 'Early Leave Deduction', 'amount': float(payroll_record.early_leave_deduction or 0)},
        {'key': 'loan', 'label_ar': 'خصم السلف', 'label_en': 'Loan Deduction', 'amount': float(payroll_record.loan_deduction or 0)},
        {'key': 'advance', 'label_ar': 'راتب مقدم', 'label_en': 'Advance Salary', 'amount': float(payroll_record.advance_salary or 0)},
        {'key': 'other', 'label_ar': 'خصومات أخرى', 'label_en': 'Other Deductions', 'amount': float(payroll_record.other_deductions or 0)},
    ]

    display_earnings_total = sum(float(item.get('amount') or 0) for item in earnings)
    display_deductions_total = sum(float(item.get('amount') or 0) for item in deductions)
    display_gross_total = display_earnings_total
    display_net_total = max(display_gross_total - display_deductions_total, 0)

    settings_data = FinanceBridgeSettingsService.load_settings()
    profile = FinanceBridgeSettingsService.get_invoice_profile(settings_data)
    logo_url = (profile.get('logo_url') or '').strip()
    company_name_ar = (profile.get('company_name_ar') or 'شركة راس السعودية المحدودة').strip()
    company_name_en = (profile.get('company_name_en') or 'RAS SAUDI COMPANY LTD').strip()
    company_tax_number = (profile.get('vat_no_ar') or profile.get('vat_no_en') or '310089997600003').strip()

    net_amount = display_net_total
    tafqeet = _tafqeet_sar(net_amount)
    qr_payload = f"EMP:{employee.name}|EMPID:{employee.employee_id}|DATE:{year}-{str(month).zfill(2)}|NET:{net_amount:.2f}|CUR:SAR"
    qr_url = f"https://quickchart.io/qr?size=140&text={quote_plus(qr_payload)}"

    return {
        'employee': employee,
        'payroll': payroll_record,
        'month': month,
        'year': year,
        'generated_at': datetime.utcnow(),
        'earnings': earnings,
        'deductions': deductions,
        'attendance_stats': attendance_stats,
        'totals': {
            'gross': display_gross_total,
            'deductions': display_deductions_total,
            'net': net_amount,
        },
        'tafqeet': tafqeet,
        'company_logo_url': logo_url,
        'company_name_ar': company_name_ar,
        'company_name_en': company_name_en,
        'company_tax_number': company_tax_number,
        'qr_payload': qr_payload,
        'qr_url': qr_url,
    }


_payslip_jobs = {}
_payslip_jobs_lock = threading.Lock()


def _set_payslip_job(job_id, **updates):
    with _payslip_jobs_lock:
        snapshot = _payslip_jobs.get(job_id, {})
        snapshot.update(updates)
        snapshot['updated_at'] = datetime.utcnow().isoformat()
        _payslip_jobs[job_id] = snapshot
        return dict(snapshot)


def _get_payslip_job(job_id):
    with _payslip_jobs_lock:
        row = _payslip_jobs.get(job_id)
        return dict(row) if row else None


# ═══════════════════════════════════════════════════════════════════════════════
# PAYROLL MANAGEMENT ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@payroll_bp.route('/dashboard')
@login_required
def dashboard():
    """لوحة معلومات الرواتب"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية للوصول لقسم الرواتب', 'danger')
        return redirect(url_for('dashboard.index'))
    
    # التأكد من وجود الأقسام الافتراضية
    _ensure_default_departments()
    
    # إحصائيات عامة
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    total_employees = Employee.query.filter_by(status='active').count()
    
    # آخر معالجة رواتب
    latest_payroll = PayrollRecord.query.filter_by(
        pay_period_month=current_month,
        pay_period_year=current_year
    ).order_by(PayrollRecord.calculated_at.desc()).first()
    
    # ✅ FIXED: Use joinedload to prevent N+1 queries on employee relationships
    month_payrolls = PayrollRecord.query.options(
        db.joinedload(PayrollRecord.employee).joinedload(Employee.departments)
    ).filter_by(
        pay_period_month=current_month,
        pay_period_year=current_year
    ).order_by(PayrollRecord.calculated_at.desc()).all()

    # الإحصائيات حسب الحالة
    pending_count = sum(1 for payroll in month_payrolls if payroll.payment_status == 'pending')
    approved_count = sum(1 for payroll in month_payrolls if payroll.payment_status == 'approved')
    rejected_count = sum(1 for payroll in month_payrolls if payroll.payment_status == 'rejected')
    paid_count = sum(1 for payroll in month_payrolls if payroll.payment_status == 'paid')

    # مجاميع المبالغ (None-safe)
    total_gross_salary = sum((payroll.gross_salary or Decimal('0')) for payroll in month_payrolls)
    total_deductions = sum((payroll.total_deductions or Decimal('0')) for payroll in month_payrolls)
    total_net_payable = sum((payroll.net_payable or Decimal('0')) for payroll in month_payrolls)
    total_gosi_company = sum((payroll.gosi_company or Decimal('0')) for payroll in month_payrolls)

    dept_salary_map = {}
    for p in month_payrolls:
        dept_name = 'غير محدد'
        if p.employee and p.employee.department:
            dept_name = p.employee.department.name
        if dept_name not in dept_salary_map:
            dept_salary_map[dept_name] = {'gross': Decimal('0'), 'net': Decimal('0'), 'count': 0}
        dept_salary_map[dept_name]['gross'] += (p.gross_salary or Decimal('0'))
        dept_salary_map[dept_name]['net'] += (p.net_payable or Decimal('0'))
        dept_salary_map[dept_name]['count'] += 1

    dept_labels = list(dept_salary_map.keys())
    dept_net_values = [float(dept_salary_map[k]['net']) for k in dept_labels]
    dept_counts = [dept_salary_map[k]['count'] for k in dept_labels]

    return render_template('payroll/dashboard.html',
        total_employees=total_employees,
        latest_payroll=latest_payroll,
        pending_count=pending_count,
        approved_count=approved_count,
        rejected_count=rejected_count,
        paid_count=paid_count,
        total_gross_salary=float(total_gross_salary),
        total_deductions=float(total_deductions),
        total_net_payable=float(total_net_payable),
        total_gosi_company=float(total_gosi_company),
        recent_payrolls=month_payrolls,
        current_month=current_month,
        current_year=current_year,
        dept_labels=dept_labels,
        dept_net_values=dept_net_values,
        dept_counts=dept_counts,
    )


@payroll_bp.route('/process', methods=['GET', 'POST'])
@login_required
def process():
    """معالجة رواتب الموظفين"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    if request.method == 'POST':
        try:
            month_value = (request.form.get('month') or '').strip()
            year_value = request.form.get('year') or request.form.get('fiscal_year')

            if month_value and '-' in month_value:
                parsed_year, parsed_month = month_value.split('-', 1)
                year = int(parsed_year)
                month = int(parsed_month)
            else:
                month = int(month_value) if month_value else datetime.now().month
                year = int(year_value) if year_value else datetime.now().year
            
            processor = PayrollProcessor(year, month)
            payrolls = processor.process_all_employees()

            total_gross = sum((p.gross_salary or Decimal('0')) for p in payrolls)
            total_deductions = sum((p.total_deductions or Decimal('0')) for p in payrolls)
            total_net_payable = sum((p.net_payable or Decimal('0')) for p in payrolls)

            wants_json = (
                request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                or 'application/json' in request.headers.get('Accept', '')
            )
            if wants_json:
                return jsonify({
                    'success': True,
                    'processed_count': len(payrolls),
                    'total_gross_salary': float(total_gross),
                    'total_deductions': float(total_deductions),
                    'total_net_payable': float(total_net_payable)
                })
            
            flash(f'تم معالجة رواتب {len(payrolls)} موظف بنجاح', 'success')
            return redirect(url_for('payroll.review', month=month, year=year))
        
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': str(e)}), 400
            flash(f'خطأ: {str(e)}', 'danger')
    
    return render_template('payroll/process.html',
        current_month=datetime.now().month,
        current_year=datetime.now().year
    )


@payroll_bp.route('/review')
@login_required
def review():
    """مراجعة رواتب الموظفين قبل الموافقة"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    # التأكد من وجود الأقسام الافتراضية
    _ensure_default_departments()
    
    month = request.args.get('month', datetime.now().month, type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    status = request.args.get('status', 'pending')
    department_id = request.args.get('department_id', type=int)
    
    # جلب الرواتب
    query = PayrollRecord.query.join(Employee, PayrollRecord.employee_id == Employee.id).filter(
        and_(
            PayrollRecord.pay_period_month == month,
            PayrollRecord.pay_period_year == year,
            PayrollRecord.payment_status == status
        )
    )

    if department_id:
        query = query.filter(Employee.department_id == department_id)
    
    payrolls = query.all()
    
    # الإحصائيات
    total_gross = sum(p.gross_salary for p in payrolls)
    total_deductions = sum(p.total_deductions for p in payrolls)
    total_net = sum(p.net_payable for p in payrolls)
    departments = Department.query.order_by(Department.name.asc()).all()
    
    return render_template('payroll/review.html',
        payrolls=payrolls,
        month=month,
        year=year,
        status=status,
        departments=departments,
        selected_department_id=department_id,
        total_gross=float(total_gross),
        total_deductions=float(total_deductions),
        total_net=float(total_net),
        payroll_count=len(payrolls)
    )


@payroll_bp.route('/<int:payroll_id>/details')
@payroll_bp.route('/payroll/<int:payroll_id>/details')
@login_required
def payroll_details(payroll_id):
    """عرض تفاصيل الراتب الكامل"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    payroll = PayrollRecord.query.get_or_404(payroll_id)
    
    return render_template('payroll/details.html',
        payroll=payroll
    )


@payroll_bp.route('/payroll/<int:payroll_id>/details/legacy')
@login_required
def payroll_details_legacy(payroll_id):
    return redirect(url_for('payroll.payroll_details', payroll_id=payroll_id))


@payroll_bp.route('/<int:payroll_id>/approve', methods=['POST'])
@payroll_bp.route('/payroll/<int:payroll_id>/approve', methods=['POST'])
@login_required
def approve_payroll(payroll_id):
    """الموافقة على الراتب"""
    if not check_payroll_access():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'}), 403
    
    try:
        payroll = PayrollRecord.query.get_or_404(payroll_id)
        processor = PayrollProcessor(payroll.pay_period_year, payroll.pay_period_month)
        
        notes = request.form.get('notes', '')
        processor.approve_payroll(payroll_id, current_user.id, notes)
        
        return jsonify({'success': True, 'message': 'تمت الموافقة على الراتب'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@payroll_bp.route('/<int:payroll_id>/reject', methods=['POST'])
@payroll_bp.route('/payroll/<int:payroll_id>/reject', methods=['POST'])
@login_required
def reject_payroll(payroll_id):
    """رفض الراتب"""
    if not check_payroll_access():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'}), 403
    
    try:
        reason = request.form.get('reason', 'لم يتم تحديد السبب')
        processor = PayrollProcessor(None, None)
        processor.reject_payroll(payroll_id, reason)
        
        return jsonify({'success': True, 'message': 'تم رفض الراتب'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@payroll_bp.route('/payroll/<int:payroll_id>/approve/legacy', methods=['POST'])
@login_required
def approve_payroll_legacy(payroll_id):
    return redirect(url_for('payroll.approve_payroll', payroll_id=payroll_id), code=307)


@payroll_bp.route('/payroll/<int:payroll_id>/reject/legacy', methods=['POST'])
@login_required
def reject_payroll_legacy(payroll_id):
    return redirect(url_for('payroll.reject_payroll', payroll_id=payroll_id), code=307)


@payroll_bp.route('/batch-approve', methods=['POST'])
@login_required
def batch_approve():
    """الموافقة على مجموعة من الرواتب"""
    if not check_payroll_access():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'}), 403
    
    try:
        payroll_ids = request.json.get('payroll_ids', [])
        processor = PayrollProcessor(None, None)
        
        for payroll_id in payroll_ids:
            processor.approve_payroll(payroll_id, current_user.id)
        
        return jsonify({
            'success': True,
            'message': f'تمت الموافقة على {len(payroll_ids)} راتب'
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


# ═══════════════════════════════════════════════════════════════════════════════
# BANK TRANSFER & EXPORT ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@payroll_bp.route('/bank-transfer/create', methods=['POST'])
@login_required
def create_bank_transfer():
    """إنشاء ملف التحويل البنكي"""
    if not check_payroll_access():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'}), 403
    
    try:
        month = int(request.form.get('month'))
        year = int(request.form.get('year'))
        bank_code = request.form.get('bank_code', 'NCB')
        file_format = request.form.get('format', 'txt')
        
        generator = BankTransferGenerator(year, month, bank_code)
        transfer_file = generator.create_transfer_file(file_format)
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء ملف التحويل بنجاح',
            'file_id': transfer_file.id,
            'file_name': transfer_file.file_name
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@payroll_bp.route('/bank-transfer/<int:file_id>/download')
@login_required
def download_bank_transfer(file_id):
    """تنزيل ملف التحويل البنكي"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    transfer_file = BankTransferFile.query.get_or_404(file_id)
    
    try:
        return send_file(
            transfer_file.file_path,
            as_attachment=True,
            download_name=transfer_file.file_name
        )
    except Exception as e:
        flash(f'خطأ في التحميل: {str(e)}', 'danger')
        return redirect(url_for('payroll.dashboard'))


@payroll_bp.route('/export/excel')
@login_required
def export_to_excel():
    """تصدير الرواتب إلى Excel"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        month = request.args.get('month', datetime.now().month, type=int)
        year = request.args.get('year', datetime.now().year, type=int)
        
        # جلب الرواتب
        payrolls = PayrollRecord.query.filter_by(
            pay_period_month=month,
            pay_period_year=year,
            payment_status='approved'
        ).all()
        
        # إنشاء Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Payroll_Details'
        
        # الرؤوس
        headers = ['رقم الموظف', 'اسم الموظف', 'الراتب الأساسي', 'المزايا', 'الراتب الإجمالي', 
                   'الخصومات', 'الراتب الصافي', 'الحالة']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # البيانات
        row = 2
        for payroll in payrolls:
            employee = payroll.employee
            
            ws.cell(row=row, column=1).value = employee.employee_id
            ws.cell(row=row, column=2).value = employee.name
            ws.cell(row=row, column=3).value = float(payroll.basic_salary)
            ws.cell(row=row, column=4).value = float(payroll.housing_allowance + payroll.transportation + payroll.meal_allowance)
            ws.cell(row=row, column=5).value = float(payroll.gross_salary)
            ws.cell(row=row, column=6).value = float(payroll.total_deductions)
            ws.cell(row=row, column=7).value = float(payroll.net_payable)
            ws.cell(row=row, column=8).value = payroll.payment_status
            
            row += 1
        
        # حفظ الملف
        file_path = f'static/payroll_exports/payroll_{month:02d}_{year}.xlsx'
        wb.save(file_path)
        
        return send_file(file_path, as_attachment=True)
    
    except Exception as e:
        flash(f'خطأ: {str(e)}', 'danger')
        return redirect(url_for('payroll.dashboard'))


# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@payroll_bp.route('/configuration')
@login_required
def configuration():
    """إعدادات الرواتب"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    config = PayrollConfiguration.query.filter_by(is_active=True).first()
    if not config:
        config = PayrollConfiguration()
    
    return render_template('payroll/configuration.html', config=config)


@payroll_bp.route('/configuration/save', methods=['POST'])
@login_required
def save_configuration():
    """حفظ إعدادات الرواتب"""
    if not check_payroll_access():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'}), 403
    
    try:
        config = PayrollConfiguration.query.filter_by(is_active=True).first()
        if not config:
            config = PayrollConfiguration()
        
        config.gosi_employee_percentage = Decimal(request.form.get('gosi_employee', 10))
        config.gosi_company_percentage = Decimal(request.form.get('gosi_company', 13))
        config.working_days_per_month = int(request.form.get('working_days', 22))
        config.minimum_wage = Decimal(request.form.get('minimum_wage', 2000))
        
        db.session.add(config)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'تم حفظ الإعدادات بنجاح'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@payroll_bp.route('/review/export-excel')
@login_required
def export_review_excel():
    """تصدير كشف مراجعة الرواتب إلى Excel احترافي"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    month = request.args.get('month', datetime.now().month, type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    status = request.args.get('status', 'pending')
    department_id = request.args.get('department_id', type=int)

    query = PayrollRecord.query.join(Employee, PayrollRecord.employee_id == Employee.id).options(
        joinedload(PayrollRecord.employee).joinedload(Employee.departments)
    ).filter(
        and_(
            PayrollRecord.pay_period_month == month,
            PayrollRecord.pay_period_year == year,
            PayrollRecord.payment_status == status
        )
    )
    if department_id:
        query = query.filter(Employee.department_id == department_id)

    payrolls = query.order_by(Employee.name).all()

    STATUS_AR = {
        'pending': 'قيد الانتظار',
        'approved': 'معتمدة',
        'rejected': 'مرفوضة',
        'paid': 'مدفوعة',
    }

    MONTH_AR = [
        '', 'يناير', 'فبراير', 'مارس', 'إبريل', 'مايو', 'يونيو',
        'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "كشف الرواتب"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.rightToLeft = True

    navy = "1A237E"
    navy_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
    white_font_xl = Font(bold=True, color="FFFFFF", size=18)
    white_font_md = Font(bold=True, color="FFFFFF", size=12)
    white_font_sm = Font(bold=True, color="FFFFFF", size=10)

    teal = "18B2B0"
    teal_fill = PatternFill(start_color=teal, end_color=teal, fill_type="solid")

    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    orange_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    blue_fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    green_font = Font(bold=True, color="155724", size=11)
    red_font = Font(bold=True, color="721C24", size=11)
    blue_font = Font(bold=True, color="0C5460", size=11)
    orange_font = Font(bold=True, color="856404", size=11)

    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

    total_gross = sum(float(p.gross_salary or 0) for p in payrolls)
    total_deductions = sum(float(p.total_deductions or 0) for p in payrolls)
    total_net = sum(float(p.net_payable or 0) for p in payrolls)

    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = "كشف مراجعة الرواتب"
    title_cell.font = white_font_xl
    title_cell.fill = navy_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:N2')
    subtitle_cell = ws['A2']
    status_label = STATUS_AR.get(status, status)
    month_label = MONTH_AR[month] if 1 <= month <= 12 else str(month)
    dept_obj = Department.query.get(department_id) if department_id else None
    dept_label = dept_obj.name if dept_obj else 'جميع الأقسام'
    subtitle_cell.value = f"الفترة: {month_label} {year}  |  الحالة: {status_label}  |  القسم: {dept_label}  |  تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    subtitle_cell.font = Font(bold=True, color="FFFFFF", size=11)
    subtitle_cell.fill = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
    subtitle_cell.alignment = center
    ws.row_dimensions[2].height = 28

    kpi_row = 4
    kpi_data = [
        ('A', 'C', 'عدد الموظفين', str(len(payrolls)), blue_fill, blue_font),
        ('D', 'F', 'إجمالي الراتب', f'{total_gross:,.2f} ر.س', green_fill, green_font),
        ('G', 'I', 'إجمالي الخصومات', f'{total_deductions:,.2f} ر.س', red_fill, red_font),
        ('J', 'L', 'صافي المستحقات', f'{total_net:,.2f} ر.س', orange_fill, orange_font),
    ]
    for start_col, end_col, label, value, fill, font in kpi_data:
        ws.merge_cells(f'{start_col}{kpi_row}:{end_col}{kpi_row}')
        label_cell = ws[f'{start_col}{kpi_row}']
        label_cell.value = label
        label_cell.font = Font(bold=True, size=10, color="555555")
        label_cell.alignment = center
        label_cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")

        val_row = kpi_row + 1
        ws.merge_cells(f'{start_col}{val_row}:{end_col}{val_row}')
        val_cell = ws[f'{start_col}{val_row}']
        val_cell.value = value
        val_cell.font = font
        val_cell.alignment = center
        val_cell.fill = fill

    ws.row_dimensions[kpi_row].height = 22
    ws.row_dimensions[kpi_row + 1].height = 30

    header_row = 7
    headers = [
        ('#', 5),
        ('رقم الموظف', 14),
        ('اسم الموظف', 28),
        ('القسم', 18),
        ('الراتب الأساسي', 16),
        ('بدل السكن', 14),
        ('بدل المواصلات', 14),
        ('بدل الطعام', 13),
        ('الراتب الإجمالي', 16),
        ('خصم الغياب', 13),
        ('خصم التأخير', 13),
        ('GOSI', 12),
        ('إجمالي الخصومات', 16),
        ('الراتب الصافي', 16),
    ]

    for col_idx, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header_text
        cell.font = white_font_sm
        cell.fill = teal_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 30

    row_num = header_row + 1
    for idx, p in enumerate(payrolls, 1):
        emp = p.employee
        emp_name = emp.name if emp else 'غير متاح'
        dept_name = ''
        if emp:
            if hasattr(emp, 'departments') and emp.departments:
                dept_name = ', '.join(d.name for d in emp.departments)
            elif hasattr(emp, 'department') and emp.department:
                dept_name = emp.department.name

        row_data = [
            idx,
            emp.employee_id if emp else '-',
            emp_name,
            dept_name or '-',
            float(p.basic_salary or 0),
            float(p.housing_allowance or 0),
            float(p.transportation or 0),
            float(p.meal_allowance or 0),
            float(p.gross_salary or 0),
            float(p.absence_deduction or 0),
            float(p.late_deduction or 0),
            float(p.gosi_employee or 0),
            float(p.total_deductions or 0),
            float(p.net_payable or 0),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = value
            cell.alignment = center
            cell.border = thin_border

            if idx % 2 == 0:
                cell.fill = light_gray_fill

            if col_idx == 3:
                cell.font = Font(bold=True, size=10)
                cell.alignment = right_align
            elif col_idx >= 5:
                cell.number_format = '#,##0.00'
                if col_idx == 9:
                    cell.font = green_font
                elif col_idx in (10, 11, 12, 13):
                    cell.font = red_font
                elif col_idx == 14:
                    cell.font = blue_font

        ws.row_dimensions[row_num].height = 24
        row_num += 1

    if payrolls:
        total_row = row_num
        ws.merge_cells(f'A{total_row}:D{total_row}')
        total_label = ws.cell(row=total_row, column=1)
        total_label.value = "الإجمالي"
        total_label.font = Font(bold=True, color="FFFFFF", size=12)
        total_label.fill = navy_fill
        total_label.alignment = center

        for col in range(2, 5):
            c = ws.cell(row=total_row, column=col)
            c.fill = navy_fill

        total_cols = {
            5: sum(float(p.basic_salary or 0) for p in payrolls),
            6: sum(float(p.housing_allowance or 0) for p in payrolls),
            7: sum(float(p.transportation or 0) for p in payrolls),
            8: sum(float(p.meal_allowance or 0) for p in payrolls),
            9: total_gross,
            10: sum(float(p.absence_deduction or 0) for p in payrolls),
            11: sum(float(p.late_deduction or 0) for p in payrolls),
            12: sum(float(p.gosi_employee or 0) for p in payrolls),
            13: total_deductions,
            14: total_net,
        }

        for col_idx, total_val in total_cols.items():
            cell = ws.cell(row=total_row, column=col_idx)
            cell.value = total_val
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = navy_fill
            cell.alignment = center
            cell.border = thin_border

        ws.row_dimensions[total_row].height = 30
        row_num = total_row + 1

    sign_row = row_num + 3
    ws.merge_cells(f'A{sign_row}:D{sign_row}')
    ws.cell(row=sign_row, column=1).value = "المحاسب: ________________"
    ws.cell(row=sign_row, column=1).font = Font(size=11, bold=True)
    ws.cell(row=sign_row, column=1).alignment = right_align

    ws.merge_cells(f'F{sign_row}:I{sign_row}')
    ws.cell(row=sign_row, column=6).value = "مدير الموارد البشرية: ________________"
    ws.cell(row=sign_row, column=6).font = Font(size=11, bold=True)
    ws.cell(row=sign_row, column=6).alignment = right_align

    ws.merge_cells(f'K{sign_row}:N{sign_row}')
    ws.cell(row=sign_row, column=11).value = "المدير العام: ________________"
    ws.cell(row=sign_row, column=11).font = Font(size=11, bold=True)
    ws.cell(row=sign_row, column=11).alignment = right_align

    ws.row_dimensions[sign_row].height = 35

    date_row = sign_row + 1
    ws.merge_cells(f'A{date_row}:D{date_row}')
    ws.cell(row=date_row, column=1).value = f"التاريخ: ____/____/________"
    ws.cell(row=date_row, column=1).font = Font(size=9, color="777777")
    ws.cell(row=date_row, column=1).alignment = right_align

    ws.merge_cells(f'F{date_row}:I{date_row}')
    ws.cell(row=date_row, column=6).value = f"التاريخ: ____/____/________"
    ws.cell(row=date_row, column=6).font = Font(size=9, color="777777")
    ws.cell(row=date_row, column=6).alignment = right_align

    ws.merge_cells(f'K{date_row}:N{date_row}')
    ws.cell(row=date_row, column=11).value = f"التاريخ: ____/____/________"
    ws.cell(row=date_row, column=11).font = Font(size=9, color="777777")
    ws.cell(row=date_row, column=11).alignment = right_align

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    status_en = status.capitalize()
    filename = f"Payroll_Review_{status_en}_{month}_{year}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@payroll_bp.route('/hr-command-center', methods=['GET'])
@login_required
def hr_command_center():
    """لوحة قيادة HR للأتمتة (Payroll + Compliance + Assets)."""
    if not check_payroll_access():
        flash('ليس لديك صلاحية للوصول إلى مركز أتمتة الموارد البشرية', 'danger')
        return redirect(url_for('dashboard.index'))

    month, year, _ = _parse_period_inputs()
    project_name = (request.args.get('project') or '').strip()
    days_ahead = request.args.get('days', 90, type=int)

    payroll_query = PayrollRecord.query.options(
        joinedload(PayrollRecord.employee)
    ).filter_by(
        pay_period_month=month,
        pay_period_year=year,
    )
    if project_name:
        payroll_query = payroll_query.join(Employee, PayrollRecord.employee_id == Employee.id).filter(
            func.lower(func.coalesce(Employee.project, '')).like(f"%{project_name.lower()}%")
        )

    payroll_records = payroll_query.all()
    payroll_rows = [
        row for row in payroll_records
        if row.employee and _employee_in_user_scope(current_user, row.employee)
    ]
    total_net = sum(Decimal(str(item.net_payable or 0)) for item in payroll_rows)

    compliance_rows = _load_compliance_rows(days_ahead=days_ahead)
    active_vehicle_handover_count = db.session.query(func.count(VehicleHandover.id)).scalar() or 0
    active_device_assignment_count = DeviceAssignment.query.filter_by(is_active=True).count()

    recent_vehicle_handovers = (
        VehicleHandover.query
        .order_by(VehicleHandover.created_at.desc())
        .limit(15)
        .all()
    )
    recent_device_assignments = (
        DeviceAssignment.query
        .filter_by(is_active=True)
        .order_by(DeviceAssignment.assignment_date.desc())
        .limit(15)
        .all()
    )

    return render_template(
        'payroll/hr_command_center.html',
        month=month,
        year=year,
        project=project_name,
        payroll_count=len(payroll_rows),
        total_net=float(total_net),
        payroll_rows=payroll_rows,
        compliance_rows=compliance_rows,
        days_ahead=days_ahead,
        active_vehicle_handover_count=active_vehicle_handover_count,
        active_device_assignment_count=active_device_assignment_count,
        recent_vehicle_handovers=recent_vehicle_handovers,
        recent_device_assignments=recent_device_assignments,
    )


@payroll_bp.route('/automation/payroll-draft', methods=['POST'])
@login_required
def automation_generate_payroll_draft():
    """إنشاء/تحديث مسودة رواتب تلقائياً من الحضور."""
    if not check_payroll_access():
        return jsonify({'ok': False, 'message': 'ليس لديك صلاحية'}), 403

    month, year, department_id = _parse_period_inputs()
    processor = PayrollProcessor(year, month)

    query = Employee.query.filter(Employee.status == 'active')
    if department_id:
        query = query.filter(
            or_(
                Employee.department_id == department_id,
                Employee.departments.any(Department.id == department_id),
            )
        )

    employees = query.all()
    created_or_updated = 0
    failures = []

    for employee in employees:
        try:
            payroll_record = processor.process_employee_payroll(employee.id)
            db.session.add(payroll_record)
            created_or_updated += 1
        except Exception as exc:
            failures.append(f'{employee.name}: {exc}')

    db.session.commit()

    response_payload = {
        'ok': True,
        'processed': created_or_updated,
        'failed': len(failures),
        'errors': failures[:10],
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
        return jsonify(response_payload)

    flash(f"تم تحديث مسودة الرواتب: {created_or_updated} سجل", 'success')
    if failures:
        flash(f"تعذر معالجة {len(failures)} موظف", 'warning')
    return redirect(url_for('payroll.hr_command_center', month=month, year=year))


@payroll_bp.route('/automation/sync-erp-journal', methods=['POST'])
@login_required
def automation_sync_erp_journal():
    """ترحيل قيد رواتب إجمالي إلى ERPNext."""
    if not check_payroll_access():
        return jsonify({'ok': False, 'message': 'ليس لديك صلاحية'}), 403

    month, year, _ = _parse_period_inputs()
    expense_account = (request.values.get('expense_account') or '').strip()
    payable_account = (request.values.get('payable_account') or '').strip()
    cost_center = (request.values.get('cost_center') or '').strip()

    if not expense_account or not payable_account:
        return jsonify({'ok': False, 'message': 'يرجى إدخال حساب المصروف وحساب الالتزام'}), 400

    try:
        result = _create_erp_payroll_journal(
            month=month,
            year=year,
            expense_account=expense_account,
            payable_account=payable_account,
            cost_center=cost_center or None,
        )
        return jsonify({
            'ok': True,
            'message': 'تم ترحيل قيد الرواتب إلى ERPNext بنجاح',
            **result,
        })
    except ERPNextBridgeError as exc:
        return jsonify({'ok': False, 'message': str(exc)}), 400
    except Exception as exc:
        return jsonify({'ok': False, 'message': f'فشل ترحيل قيد ERP: {exc}'}), 500


@payroll_bp.route('/automation/wps-export', methods=['GET'])
@login_required
def automation_wps_export():
    """تصدير ملف WPS (CSV) للفترة المحددة."""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))

    month, year, department_id = _parse_period_inputs()
    payment_date_raw = request.args.get('payment_date', '')

    if payment_date_raw:
        try:
            payment_date_value = datetime.strptime(payment_date_raw, '%Y-%m-%d').date()
        except ValueError:
            payment_date_value = date.today()
    else:
        payment_date_value = date.today()

    query = PayrollRecord.query.options(joinedload(PayrollRecord.employee)).filter(
        PayrollRecord.pay_period_month == month,
        PayrollRecord.pay_period_year == year,
        PayrollRecord.payment_status.in_(['approved', 'paid', 'pending'])
    )

    if department_id:
        query = query.join(Employee, PayrollRecord.employee_id == Employee.id).filter(
            or_(
                Employee.department_id == department_id,
                Employee.departments.any(Department.id == department_id),
            )
        )

    records = query.all()
    if not records:
        flash('لا توجد رواتب للتصدير في هذه الفترة', 'warning')
        return redirect(url_for('payroll.hr_command_center', month=month, year=year))

    csv_bytes = _build_wps_csv(records, payment_date_value)
    filename = f'NUZUM_WPS_{year}_{str(month).zfill(2)}.csv'
    return send_file(
        BytesIO(csv_bytes),
        as_attachment=True,
        download_name=filename,
        mimetype='text/csv',
    )


@payroll_bp.route('/automation/compliance/escalate', methods=['POST'])
@login_required
def automation_compliance_escalate():
    """إرسال تصعيد تنبيهات انتهاء الوثائق عبر SMS/WhatsApp عند توفر الإعدادات."""
    if not check_payroll_access():
        return jsonify({'ok': False, 'message': 'ليس لديك صلاحية'}), 403

    days_ahead = request.values.get('days', 90, type=int)
    channel = (request.values.get('channel') or 'sms').strip().lower()
    rows = _load_compliance_rows(days_ahead=days_ahead)

    notifier = _get_notification_service()
    sent = 0
    failed = 0
    errors = []

    for row in rows:
        phone = (row.get('employee_mobile') or '').strip()
        if not phone:
            failed += 1
            errors.append(f"{row['employee_name']}: رقم الجوال غير متوفر")
            continue

        message = (
            f"تنبيه صلاحية وثيقة - نُظم\n"
            f"الموظف: {row['employee_name']}\n"
            f"الوثيقة: {row['document_type']} ({row['document_number']})\n"
            f"تنتهي خلال {row['days_left']} يوم\n"
            f"تاريخ الانتهاء: {row['expiry_date']}"
        )

        if channel == 'whatsapp':
            try:
                from app import whatsapp_service
                normalized_phone = phone if phone.startswith('+') else f'+966{phone.lstrip("0")}'
                response = whatsapp_service.send_text_message(normalized_phone, message) if whatsapp_service else None
                if response:
                    sent += 1
                else:
                    failed += 1
                    errors.append(f"{row['employee_name']}: فشل إرسال WhatsApp")
            except Exception as exc:
                failed += 1
                errors.append(f"{row['employee_name']}: {exc}")
        else:
            ok, err = notifier.send_sms(phone, message)
            if ok:
                sent += 1
            else:
                failed += 1
                errors.append(f"{row['employee_name']}: {err}")

    return jsonify({
        'ok': True,
        'channel': channel,
        'total': len(rows),
        'sent': sent,
        'failed': failed,
        'errors': errors[:20],
    })


@payroll_bp.route('/generate-payslip/<int:employee_id>/<string:month_year>', methods=['GET'])
@login_required
def generate_payslip(employee_id, month_year):
    """عرض قسيمة راتب احترافية قابلة للطباعة/PDF."""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))

    month, year = _parse_month_year(month_year)
    if not month or not year:
        flash('صيغة الفترة غير صحيحة. استخدم YYYY-MM', 'danger')
        return redirect(url_for('payroll.hr_command_center'))

    employee = Employee.query.get_or_404(employee_id)
    if not _employee_in_user_scope(current_user, employee):
        flash('غير مصرح لك بعرض هذه القسيمة', 'danger')
        return redirect(url_for('payroll.hr_command_center', month=month, year=year))

    payroll_record = PayrollRecord.query.filter_by(
        employee_id=employee_id,
        pay_period_month=month,
        pay_period_year=year,
    ).first()

    if not payroll_record:
        flash('لا توجد مسودة/سجل راتب لهذه الفترة', 'warning')
        return redirect(url_for('payroll.hr_command_center', month=month, year=year))

    payload = _build_payslip_payload(employee, payroll_record, month, year)
    return render_template('payroll/payslip_pdf.html', **payload)


@payroll_bp.route('/payslips/review', methods=['GET'])
@login_required
def payslips_review():
    """صفحة استعراض القسائم قبل الإرسال."""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))

    month, year, department_id = _parse_period_inputs()
    project_name = (request.args.get('project') or '').strip()

    query = PayrollRecord.query.options(joinedload(PayrollRecord.employee)).filter_by(
        pay_period_month=month,
        pay_period_year=year,
    )
    if department_id or project_name:
        query = query.join(Employee, PayrollRecord.employee_id == Employee.id)
    if department_id:
        query = query.filter(
            or_(
                Employee.department_id == department_id,
                Employee.departments.any(Department.id == department_id),
            )
        )
    if project_name:
        query = query.filter(func.lower(func.coalesce(Employee.project, '')).like(f"%{project_name.lower()}%"))

    rows = []
    for item in query.all():
        employee = item.employee
        if not employee or not _employee_in_user_scope(current_user, employee):
            continue
        secure_url = _build_secure_payslip_url(employee, month, year)
        phone_raw = (getattr(employee, 'mobile', None) or getattr(employee, 'phone', None) or '').strip()
        normalized_phone = ''.join(ch for ch in phone_raw if ch.isdigit())
        if normalized_phone.startswith('0'):
            normalized_phone = '966' + normalized_phone[1:]
        elif normalized_phone.startswith('5') and len(normalized_phone) == 9:
            normalized_phone = '966' + normalized_phone
        wa_text = f"عزيزي {employee.name}، تم صدور قسيمة راتبك لشهر {str(month).zfill(2)}/{year}. يمكنك عرضها عبر الرابط التالي: {secure_url}"
        wa_link = f"https://wa.me/{normalized_phone}?text={quote_plus(wa_text)}" if normalized_phone else ''
        rows.append({
            'employee': employee,
            'payroll': item,
            'secure_url': secure_url,
            'wa_link': wa_link,
        })

    return render_template(
        'payroll/payslips_review.html',
        month=month,
        year=year,
        project=project_name,
        rows=rows,
    )


@payroll_bp.route('/secure-payslip/<string:token>', methods=['GET'])
def generate_secure_payslip(token):
    """رابط قسيمة آمن: يتطلب رمز صالح + رقم إقامة/هوية للتأكيد."""
    try:
        payload = _get_payslip_serializer().loads(token, max_age=60 * 60 * 24 * 31)
    except SignatureExpired:
        return "الرابط منتهي الصلاحية", 410
    except BadSignature:
        return "الرابط غير صالح", 403

    employee_id = int(payload.get('employee_id') or 0)
    month = int(payload.get('month') or 0)
    year = int(payload.get('year') or 0)
    id_tail = str(payload.get('id_tail') or '')

    employee = Employee.query.get_or_404(employee_id)
    payroll_record = PayrollRecord.query.filter_by(
        employee_id=employee_id,
        pay_period_month=month,
        pay_period_year=year,
    ).first_or_404()

    provided_id = _normalize_identity(request.args.get('id_number') or request.args.get('iqama') or '')
    expected_id = _employee_identity_value(employee)

    if not provided_id:
        html = render_template_string(
            """
            <html lang="ar" dir="rtl"><head><meta charset="utf-8"><title>تحقق الهوية</title></head>
            <body style="font-family:Tahoma,Arial,sans-serif;background:#f7f8fa;padding:24px;">
                <div style="max-width:460px;margin:20px auto;background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:20px;">
                    <h3 style="margin-top:0;color:#001f3f;">تحقق عرض قسيمة الراتب</h3>
                    <p>أدخل رقم الإقامة/الهوية لعرض القسيمة.</p>
                    <form method="get">
                        <input type="text" name="id_number" placeholder="رقم الإقامة" style="width:100%;padding:10px;border:1px solid #d1d5db;border-radius:6px;" required>
                        <button type="submit" style="margin-top:12px;width:100%;background:#001f3f;color:#fff;border:none;padding:10px;border-radius:6px;cursor:pointer;">عرض القسيمة</button>
                    </form>
                    <small style="display:block;margin-top:8px;color:#6b7280;">آخر 4 أرقام المسجلة: {{ id_tail or '****' }}</small>
                </div>
            </body></html>
            """,
            id_tail=id_tail,
        )
        response = make_response(html)
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response

    if not expected_id or provided_id != expected_id:
        return "فشل تحقق الهوية", 403

    payslip_payload = _build_payslip_payload(employee, payroll_record, month, year)
    response = make_response(render_template('payroll/payslip_secure.html', **payslip_payload))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@payroll_bp.route('/automation/payslips/generate-bulk', methods=['POST'])
@login_required
def automation_generate_bulk_payslips():
    """تجهيز جماعي لقسائم الرواتب (تحقق البيانات وتوفير الروابط)."""
    if not check_payroll_access():
        return jsonify({'ok': False, 'message': 'ليس لديك صلاحية'}), 403

    month, year, department_id = _parse_period_inputs()
    project_name = (request.values.get('project') or '').strip()
    query = PayrollRecord.query.options(joinedload(PayrollRecord.employee)).filter_by(
        pay_period_month=month,
        pay_period_year=year,
    )
    if department_id or project_name:
        query = query.join(Employee, PayrollRecord.employee_id == Employee.id)
    if department_id:
        query = query.filter(
            or_(
                Employee.department_id == department_id,
                Employee.departments.any(Department.id == department_id),
            )
        )
    if project_name:
        query = query.filter(func.lower(func.coalesce(Employee.project, '')).like(f"%{project_name.lower()}%"))

    records = query.all()
    allowed_records = [
        item for item in records
        if item.employee and _employee_in_user_scope(current_user, item.employee)
    ]

    links = [
        url_for('payroll.generate_payslip', employee_id=item.employee_id, month_year=f"{year}-{str(month).zfill(2)}", _external=True)
        for item in allowed_records[:200]
    ]

    return jsonify({
        'ok': True,
        'period': f'{year}-{str(month).zfill(2)}',
        'project': project_name,
        'count': len(allowed_records),
        'sample_links': links[:20],
    })


def _send_payslips_email_job(job_id, month, year, department_id, project_name, channel, finance_cc_email, triggered_by_user_id):
    try:
        from services.email_service import EmailService

        _set_payslip_job(job_id, status='running', sent=0, failed=0, total=0, message='جاري تجهيز القسائم...')

        query = PayrollRecord.query.options(joinedload(PayrollRecord.employee)).filter_by(
            pay_period_month=month,
            pay_period_year=year,
        )
        if department_id or project_name:
            query = query.join(Employee, PayrollRecord.employee_id == Employee.id)
        if department_id:
            query = query.filter(
                or_(
                    Employee.department_id == department_id,
                    Employee.departments.any(Department.id == department_id),
                )
            )
        if project_name:
            query = query.filter(func.lower(func.coalesce(Employee.project, '')).like(f"%{project_name.lower()}%"))

        records = query.all()

        triggered_user = User.query.get(triggered_by_user_id) if triggered_by_user_id else None
        if triggered_user:
            records = [
                item for item in records
                if item.employee and _employee_in_user_scope(triggered_user, item.employee)
            ]

        _set_payslip_job(
            job_id,
            total=len(records),
            channel=channel,
            project=project_name,
            finance_cc_email=finance_cc_email,
        )

        email_service = EmailService() if channel == 'email' else None
        sent = 0
        failed = 0
        errors = []
        whatsapp_links = []

        for record in records:
            employee = record.employee
            secure_payslip_url = _build_secure_payslip_url(employee, month, year)

            if channel == 'whatsapp':
                phone_number = (getattr(employee, 'mobile', None) or getattr(employee, 'phone', None) or '').strip()
                if not phone_number:
                    failed += 1
                    errors.append(f"{getattr(employee, 'name', '-')}: phone_not_found")
                    continue
                try:
                    normalized_phone = ''.join(ch for ch in phone_number if ch.isdigit())
                    if normalized_phone.startswith('0'):
                        normalized_phone = '966' + normalized_phone[1:]
                    elif normalized_phone.startswith('5') and len(normalized_phone) == 9:
                        normalized_phone = '966' + normalized_phone

                    message_text = (
                        f"عزيزي {employee.name}، تم صدور قسيمة راتبك لشهر {str(month).zfill(2)}/{year}. "
                        f"يمكنك عرضها عبر الرابط التالي: {secure_payslip_url}"
                    )
                    wa_link = f"https://wa.me/{normalized_phone}?text={quote_plus(message_text)}"
                    whatsapp_links.append({
                        'employee': employee.name,
                        'employee_id': employee.employee_id,
                        'phone': normalized_phone,
                        'link': wa_link,
                    })
                    sent += 1
                except Exception as wa_exc:
                    failed += 1
                    errors.append(f"{employee.name}: {wa_exc}")
            else:
                if not employee or not employee.email:
                    failed += 1
                    errors.append(f"{getattr(employee, 'name', '-')}: email_not_found")
                    continue

                subject = f"قسيمة راتب - {employee.name} - {year}/{str(month).zfill(2)}"
                html_body = (
                    f"<div dir='rtl' style='font-family:Cairo,Arial,sans-serif'>"
                    f"<h3>قسيمة راتب نُظم</h3>"
                    f"<p>الموظف: <strong>{employee.name}</strong></p>"
                    f"<p>الفترة: <strong>{year}/{str(month).zfill(2)}</strong></p>"
                    f"<p>رابط القسيمة الآمن: <a href='{secure_payslip_url}'>{secure_payslip_url}</a></p>"
                    f"</div>"
                )

                result = email_service.send_simple_email(employee.email, subject, html_body)
                if result.get('success'):
                    sent += 1
                    if finance_cc_email:
                        email_service.send_simple_email(
                            finance_cc_email,
                            f"[Archive Copy] {subject}",
                            html_body,
                        )
                else:
                    failed += 1
                    errors.append(f"{employee.name}: {result.get('message', 'send_failed')}")

            _set_payslip_job(
                job_id,
                status='running',
                sent=sent,
                failed=failed,
                whatsapp_links=whatsapp_links[:50],
                message=f'جاري الإرسال... {sent + failed}/{len(records)}',
                errors=errors[:30],
            )

        _set_payslip_job(
            job_id,
            status='done',
            sent=sent,
            failed=failed,
            whatsapp_links=whatsapp_links[:50],
            message='اكتمل إرسال القسائم' if channel == 'email' else 'اكتمل تجهيز روابط واتساب',
            errors=errors[:30],
        )
    except Exception as exc:
        _set_payslip_job(job_id, status='failed', message=f'فشل الإرسال الجماعي: {exc}')


@payroll_bp.route('/automation/payslips/email-all', methods=['POST'])
@login_required
def automation_email_all_payslips():
    """تشغيل إرسال جماعي لقسائم الرواتب في خلفية Thread."""
    if not check_payroll_access():
        return jsonify({'ok': False, 'message': 'ليس لديك صلاحية'}), 403

    month, year, department_id = _parse_period_inputs()
    project_name = (request.values.get('project') or '').strip()
    channel = (request.values.get('channel') or 'email').strip().lower()
    finance_cc_email = (request.values.get('finance_cc_email') or '').strip()
    if channel not in {'email', 'whatsapp'}:
        channel = 'email'
    trigger_user_id = getattr(current_user, 'id', None)
    job_id = f"payslip-{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}"
    _set_payslip_job(
        job_id,
        id=job_id,
        status='queued',
        sent=0,
        failed=0,
        total=0,
        channel=channel,
        project=project_name,
        finance_cc_email=finance_cc_email,
        message='تم جدولة المهمة'
    )

    from flask import current_app
    app_obj = current_app._get_current_object()

    def _runner():
        with app_obj.app_context():
            _send_payslips_email_job(
                job_id,
                month,
                year,
                department_id,
                project_name,
                channel,
                finance_cc_email,
                trigger_user_id,
            )

    thread = threading.Thread(target=_runner, daemon=True, name=f'payslip-email-{job_id[-6:]}')
    thread.start()

    channel_label = 'واتساب' if channel == 'whatsapp' else 'البريد الإلكتروني'
    return jsonify({'ok': True, 'job_id': job_id, 'message': f'تم تشغيل الإرسال الجماعي عبر {channel_label} في الخلفية'})


@payroll_bp.route('/automation/approve-payroll', methods=['POST'])
@login_required
def automation_approve_payroll_batch():
    """اعتماد دفعة الرواتب مع ترحيل قيد ERPNext بإعدادات افتراضية لمشروع القصيم."""
    if not check_payroll_access():
        return jsonify({'ok': False, 'message': 'ليس لديك صلاحية'}), 403

    month, year, _ = _parse_period_inputs()
    expense_account = (request.values.get('expense_account') or 'Payroll Expenses').strip()
    payable_account = (request.values.get('payable_account') or 'Salaries Payable').strip()
    cost_center = (request.values.get('cost_center') or 'Al Qassim Project').strip()

    try:
        result = _create_erp_payroll_journal(
            month=month,
            year=year,
            expense_account=expense_account,
            payable_account=payable_account,
            cost_center=cost_center,
        )
        PayrollRecord.query.filter_by(
            pay_period_month=month,
            pay_period_year=year,
        ).update({'payment_status': 'approved'}, synchronize_session=False)
        db.session.commit()

        return jsonify({
            'ok': True,
            'message': 'تم اعتماد الدفعة وترحيل القيد إلى ERPNext',
            **result,
        })
    except ERPNextBridgeError as exc:
        return jsonify({'ok': False, 'message': str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'فشل اعتماد الرواتب: {exc}'}), 500


@payroll_bp.route('/automation/payslips/jobs/<string:job_id>', methods=['GET'])
@login_required
def automation_payslip_job_status(job_id):
    """جلب حالة مهمة إرسال القسائم."""
    if not check_payroll_access():
        return jsonify({'ok': False, 'message': 'ليس لديك صلاحية'}), 403

    payload = _get_payslip_job(job_id)
    if not payload:
        return jsonify({'ok': False, 'message': 'job_not_found'}), 404

    return jsonify({'ok': True, 'job': payload})
