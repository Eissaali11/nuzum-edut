from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response, send_file
from werkzeug.utils import secure_filename
from sqlalchemy import func, or_
from sqlalchemy.orm import selectinload
from datetime import datetime, timedelta
import io
from io import BytesIO
import csv
import xlsxwriter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask_login import current_user, login_required
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.platypus import PageBreak
from core.extensions import db
from models import Document, Employee, Department, SystemAudit
from utils.excel import parse_document_excel
from utils.date_converter import parse_date, format_date_hijri, format_date_gregorian
from utils.audit_logger import log_activity
import json

documents_bp = Blueprint('documents', __name__)


def create_expiry_notification(user_id, document_type, employee_name, days_until_expiry, document_id):
    """إشعار انتهاء صلاحية وثيقة"""
    from models import Notification, User
    
    if days_until_expiry < 0:
        title = f'وثيقة منتهية - {document_type}'
        description = f'انتهت صلاحية {document_type} للموظف {employee_name} منذ {abs(days_until_expiry)} يوم'
        priority = 'critical'
    elif days_until_expiry <= 7:
        title = f'تنبيه عاجل: وثيقة تنتهي قريباً'
        description = f'{document_type} للموظف {employee_name} تنتهي خلال {days_until_expiry} أيام'
        priority = 'critical'
    elif days_until_expiry <= 30:
        title = f'تذكير: وثيقة تنتهي خلال شهر'
        description = f'{document_type} للموظف {employee_name} تنتهي خلال {days_until_expiry} يوماً'
        priority = 'high'
    else:
        title = f'تذكير: وثيقة قريبة من الانتهاء'
        description = f'{document_type} للموظف {employee_name} تنتهي خلال {days_until_expiry} يوماً'
        priority = 'normal'
    
    notification = Notification(
        user_id=user_id,
        notification_type='document_expiry',
        title=title,
        description=description,
        related_entity_type='document',
        related_entity_id=document_id,
        priority=priority,
        action_url=url_for('documents.dashboard')
    )
    db.session.add(notification)
    return notification


@documents_bp.route('/test-notifications', methods=['GET', 'POST'])
def test_expiry_notifications():
    """اختبار إنشاء إشعارات انتهاء الصلاحيات لجميع المستخدمين"""
    try:
        from models import Notification, User
        
        current_date = datetime.now().date()
        warning_date = current_date + timedelta(days=30)
        
        # الحصول على وثائق قريبة من الانتهاء أو منتهية
        expiring_docs = Document.query.join(Employee)\
            .filter(Document.expiry_date <= warning_date)\
            .order_by(Document.expiry_date)\
            .limit(5).all()
        
        if not expiring_docs:
            return jsonify({'success': False, 'message': 'لا توجد وثائق منتهية أو قريبة من الانتهاء'}), 404
        
        all_users = User.query.all()
        
        notification_count = 0
        for doc in expiring_docs:
            days_until_expiry = (doc.expiry_date - current_date).days if doc.expiry_date else -999
            
            for user in all_users:
                try:
                    create_expiry_notification(
                        user_id=user.id,
                        document_type=doc.document_type or 'وثيقة',
                        employee_name=doc.employee.name if doc.employee else 'غير محدد',
                        days_until_expiry=days_until_expiry,
                        document_id=doc.id
                    )
                    notification_count += 1
                except Exception as e:
                    pass
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء {notification_count} إشعار لـ {len(expiring_docs)} وثيقة',
            'documents_count': len(expiring_docs),
            'users_count': len(all_users)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@documents_bp.route('/dashboard')
@login_required
def dashboard():
    """داش بورد شامل لإحصائيات الوثائق"""
    current_date = datetime.now().date()
    
    # حساب تواريخ الفترات
    expiring_date = current_date + timedelta(days=60)
    warning_date = current_date + timedelta(days=30)
    
    # إحصائيات الوثائق
    total_documents = Document.query.count()
    expired_documents = Document.query.filter(Document.expiry_date < current_date).count()
    expiring_soon = Document.query.filter(
        Document.expiry_date >= current_date,
        Document.expiry_date <= warning_date
    ).count()
    expiring_later = Document.query.filter(
        Document.expiry_date > warning_date,
        Document.expiry_date <= expiring_date
    ).count()
    valid_documents = Document.query.filter(Document.expiry_date > expiring_date).count()
    
    # الوثائق المنتهية (آخر 10)
    expired_docs = Document.query.join(Employee)\
        .filter(Document.expiry_date < current_date)\
        .order_by(Document.expiry_date.desc())\
        .limit(10).all()
    
    # الوثائق القريبة من الانتهاء (30 يوم)
    expiring_docs = Document.query.join(Employee)\
        .filter(Document.expiry_date >= current_date, Document.expiry_date <= warning_date)\
        .order_by(Document.expiry_date)\
        .limit(10).all()
    
    # إحصائيات حسب نوع الوثيقة
    document_types_stats = db.session.query(
        Document.document_type,
        func.count(Document.id).label('count')
    ).group_by(Document.document_type).all()
    
    # إحصائيات حسب القسم
    department_stats = db.session.query(
        Department.name,
        func.count(Document.id).label('count')
    ).select_from(Department)\
     .join(Employee, Employee.department_id == Department.id)\
     .join(Document, Document.employee_id == Employee.id)\
     .group_by(Department.name)\
     .order_by(func.count(Document.id).desc())\
     .limit(5).all()
    
    return render_template('documents/dashboard.html',
                         total_documents=total_documents,
                         expired_documents=expired_documents,
                         expiring_soon=expiring_soon,
                         expiring_later=expiring_later,
                         valid_documents=valid_documents,
                         expired_docs=expired_docs,
                         expiring_docs=expiring_docs,
                         document_types_stats=document_types_stats,
                         department_stats=department_stats,
                         current_date=current_date)

@documents_bp.route('/update_expiry_date/<int:document_id>', methods=['POST'])
@login_required
def update_expiry_date(document_id):
    """تحديث تاريخ انتهاء الوثيقة"""
    try:
        document = Document.query.get_or_404(document_id)
        
        # الحصول على التاريخ الجديد من الطلب
        new_expiry_date = request.form.get('new_expiry_date')
        
        if not new_expiry_date:
            flash('يجب إدخال تاريخ الانتهاء الجديد', 'error')
            return redirect(request.referrer or url_for('documents.expiring'))
        
        # تحويل التاريخ من string إلى date object
        try:
            new_date = datetime.strptime(new_expiry_date, '%Y-%m-%d').date()
        except ValueError:
            flash('تنسيق التاريخ غير صحيح', 'error')
            return redirect(request.referrer or url_for('documents.expiring'))
        
        # حفظ التاريخ القديم للسجل
        old_expiry_date = document.expiry_date
        
        # تحديث التاريخ
        document.expiry_date = new_date
        document.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        # تسجيل العملية في سجل النشاط
        log_activity(
            action='update',
            entity_type='document',
            entity_id=document_id,
            details=f'تم تحديث تاريخ انتهاء الوثيقة من {old_expiry_date} إلى {new_date} للموظف {document.employee.name}'
        )
        
        flash('تم تحديث تاريخ انتهاء الوثيقة بنجاح', 'success')
        return redirect(request.referrer or url_for('documents.expiring'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء تحديث تاريخ الانتهاء: {str(e)}', 'error')
        return redirect(request.referrer or url_for('documents.expiring'))

@documents_bp.route('/template/pdf')
@login_required
def document_template_pdf():
    """إنشاء نموذج PDF فارغ للوثائق"""
    try:
        # إنشاء ملف PDF في الذاكرة
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # تسجيل الخط العربي
        try:
            pdfmetrics.registerFont(TTFont('Cairo', 'Cairo.ttf'))
            arabic_font = 'Cairo'
        except:
            arabic_font = 'Helvetica'
        
        # إنشاء الأنماط
        styles = getSampleStyleSheet()
        
        # نمط العنوان الرئيسي
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontName=arabic_font,
            fontSize=20,
            spaceAfter=30,
            alignment=1,  # وسط
            textColor=colors.darkblue
        )
        
        # نمط العنوان الفرعي
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontName=arabic_font,
            fontSize=16,
            spaceAfter=20,
            alignment=1,
            textColor=colors.blue
        )
        
        # نمط النص العادي
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=arabic_font,
            fontSize=12,
            spaceAfter=12,
            alignment=2,  # يمين
            leading=18
        )
        
        # نمط الجدول
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        
        # محتوى الوثيقة
        story = []
        
        # العنوان الرئيسي
        title_text = arabic_reshaper.reshape("نموذج إدارة الوثائق")
        title_text = get_display(title_text)
        story.append(Paragraph(title_text, title_style))
        
        # معلومات الشركة
        company_text = arabic_reshaper.reshape("شركة نُظم لإدارة الموارد البشرية")
        company_text = get_display(company_text)
        story.append(Paragraph(company_text, subtitle_style))
        
        story.append(Spacer(1, 20))
        
        # جدول معلومات الموظف
        employee_title = arabic_reshaper.reshape("معلومات الموظف")
        employee_title = get_display(employee_title)
        story.append(Paragraph(employee_title, subtitle_style))
        
        employee_data = [
            [get_display(arabic_reshaper.reshape("البيان")), get_display(arabic_reshaper.reshape("القيمة"))],
            [get_display(arabic_reshaper.reshape("اسم الموظف")), "________________________"],
            [get_display(arabic_reshaper.reshape("رقم الموظف")), "________________________"],
            [get_display(arabic_reshaper.reshape("رقم الهوية الوطنية")), "________________________"],
            [get_display(arabic_reshaper.reshape("القسم")), "________________________"],
            [get_display(arabic_reshaper.reshape("المنصب")), "________________________"]
        ]
        
        employee_table = Table(employee_data, colWidths=[8*cm, 8*cm])
        employee_table.setStyle(table_style)
        story.append(employee_table)
        
        story.append(Spacer(1, 30))
        
        # جدول معلومات الوثيقة
        document_title = arabic_reshaper.reshape("معلومات الوثيقة")
        document_title = get_display(document_title)
        story.append(Paragraph(document_title, subtitle_style))
        
        document_data = [
            [get_display(arabic_reshaper.reshape("البيان")), get_display(arabic_reshaper.reshape("القيمة"))],
            [get_display(arabic_reshaper.reshape("نوع الوثيقة")), "________________________"],
            [get_display(arabic_reshaper.reshape("رقم الوثيقة")), "________________________"],
            [get_display(arabic_reshaper.reshape("تاريخ الإصدار")), "________________________"],
            [get_display(arabic_reshaper.reshape("تاريخ الانتهاء")), "________________________"],
            [get_display(arabic_reshaper.reshape("الجهة المصدرة")), "________________________"]
        ]
        
        document_table = Table(document_data, colWidths=[8*cm, 8*cm])
        document_table.setStyle(table_style)
        story.append(document_table)
        
        story.append(Spacer(1, 30))
        
        # حقل الملاحظات
        notes_title = arabic_reshaper.reshape("الملاحظات")
        notes_title = get_display(notes_title)
        story.append(Paragraph(notes_title, subtitle_style))
        
        # مساحة فارغة للملاحظات
        notes_lines = []
        for i in range(5):
            notes_lines.append("_________________________________________________")
        
        for line in notes_lines:
            story.append(Paragraph(line, normal_style))
        
        story.append(Spacer(1, 40))
        
        # التوقيعات
        signature_data = [
            [get_display(arabic_reshaper.reshape("توقيع الموظف")), get_display(arabic_reshaper.reshape("توقيع المسؤول"))],
            ["", ""],
            ["", ""],
            [get_display(arabic_reshaper.reshape("التاريخ: ___________")), get_display(arabic_reshaper.reshape("التاريخ: ___________"))]
        ]
        
        signature_table = Table(signature_data, colWidths=[8*cm, 8*cm], rowHeights=[None, 2*cm, None, None])
        signature_table.setStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, 2), 1, colors.black)
        ])
        
        story.append(signature_table)
        
        # بناء الوثيقة
        doc.build(story)
        
        # إرجاع PDF
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename=document_template.pdf'
        
        return response
        
    except Exception as e:
        flash(f'حدث خطأ في إنشاء نموذج PDF: {str(e)}', 'error')
        return redirect(url_for('documents.expiring'))

@documents_bp.route('/delete/<int:document_id>', methods=['GET', 'POST'])
@login_required
def delete_document(document_id):
    """صفحة تأكيد وتنفيذ حذف الوثيقة"""
    document = Document.query.get_or_404(document_id)
    
    if request.method == 'POST':
        try:
            # حفظ معلومات الوثيقة للسجل
            employee_name = document.employee.name if document.employee else 'غير محدد'
            document_type = document.document_type
            
            # 💾 الملف يبقى محفوظاً - نحذف فقط المرجع من قاعدة البيانات
            # لا حذف للملفات الفعلية - الاحتفاظ بجميع الوثائق للأمان
            
            # حذف الوثيقة من قاعدة البيانات
            db.session.delete(document)
            db.session.commit()
            
            # تسجيل العملية في سجل النشاط
            log_activity(
                action='delete',
                entity_type='document',
                entity_id=document_id,
                details=f'تم حذف وثيقة {document_type} للموظف {employee_name}'
            )
            
            flash(f'تم حذف وثيقة {document_type} للموظف {employee_name} بنجاح', 'success')
            return redirect(url_for('documents.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء حذف الوثيقة: {str(e)}', 'error')
            print(f"خطأ في حذف الوثيقة {document_id}: {e}")
            return redirect(url_for('documents.index'))
    
    # GET request - عرض صفحة التأكيد
    return render_template('documents/confirm_delete.html', document=document)

@documents_bp.route('/get_sponsorship_employees', methods=['POST'])
def get_sponsorship_employees():
    """Get employees filtered by sponsorship status"""
    try:
        sponsorship_filter = request.form.get('sponsorship_filter')
        
        if not sponsorship_filter:
            return jsonify({'success': False, 'message': 'يرجى تحديد نوع الكفالة'})
        
        # Query employees based on sponsorship status
        if sponsorship_filter == 'on_sponsorship':
            employees = Employee.query.filter(Employee.sponsorship_status == 'على الكفالة').all()
        elif sponsorship_filter == 'off_sponsorship':
            employees = Employee.query.filter(Employee.sponsorship_status == 'خارج الكفالة').all()
        else:
            return jsonify({'success': False, 'message': 'نوع الكفالة غير صحيح'})
        
        # Format employee data
        employees_data = []
        for emp in employees:
            dept_names = ', '.join([dept.name for dept in emp.departments]) if emp.departments else 'غير محدد'
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id,
                'national_id': emp.national_id,
                'sponsorship_status': emp.sponsorship_status,
                'department_names': dept_names
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@documents_bp.route('/department-bulk-create', methods=['GET', 'POST'])
@login_required
def department_bulk_create():
    """صفحة إنشاء وثائق القسم الكامل"""
    try:
        # جلب البيانات الأساسية
        departments = Department.query.all()
        employees = Employee.query.options(selectinload(Employee.departments)).all()
        document_types = [
            'national_id', 'passport', 'health_certificate', 
            'work_permit', 'education_certificate'
        ]
        
        if request.method == 'POST':
            # معالجة طلب الحفظ
            save_type = request.form.get('save_type')
            department_id = request.form.get('department_id')
            document_type = request.form.get('document_type')
            
            if save_type == 'individual':
                # حفظ موظف واحد
                employee_id = request.form.get('employee_id')
                document_number = request.form.get('document_number')
                issue_date = request.form.get('issue_date')
                expiry_date = request.form.get('expiry_date')
                notes = request.form.get('notes', '')
                
                if not all([employee_id, document_type, document_number]):
                    return jsonify({
                        'success': False, 
                        'message': 'يرجى إدخال جميع البيانات المطلوبة'
                    })
                
                # إنشاء الوثيقة
                document = Document(
                    employee_id=employee_id,
                    document_type=document_type,
                    document_number=document_number,
                    issue_date=parse_date(issue_date) if issue_date else None,
                    expiry_date=parse_date(expiry_date) if expiry_date else None,
                    notes=notes
                )
                
                db.session.add(document)
                
                # تسجيل العملية
                employee = Employee.query.get(employee_id)
                audit = SystemAudit(
                    action='create',
                    entity_type='document',
                    entity_id=document.id,
                    details=f'تم إضافة وثيقة {document_type} للموظف {employee.name if employee else "غير محدد"}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'message': 'تم حفظ الوثيقة بنجاح'
                })
            
            elif save_type == 'bulk':
                # حفظ جماعي
                import json
                employees_data = json.loads(request.form.get('employees_data', '[]'))
                
                if not employees_data:
                    return jsonify({
                        'success': False,
                        'message': 'لا توجد بيانات للحفظ'
                    })
                
                saved_count = 0
                for emp_data in employees_data:
                    if emp_data.get('document_number'):
                        document = Document(
                            employee_id=emp_data['employee_id'],
                            document_type=document_type,
                            document_number=emp_data['document_number'],
                            issue_date=parse_date(emp_data['issue_date']) if emp_data.get('issue_date') else None,
                            expiry_date=parse_date(emp_data['expiry_date']) if emp_data.get('expiry_date') else None,
                            notes=emp_data.get('notes', '')
                        )
                        db.session.add(document)
                        saved_count += 1
                
                # تسجيل العملية
                department = Department.query.get(department_id)
                audit = SystemAudit(
                    action='bulk_create',
                    entity_type='document',
                    entity_id=department_id,
                    details=f'تم إضافة {saved_count} وثيقة من نوع {document_type} لقسم {department.name if department else "غير محدد"}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'message': f'تم حفظ {saved_count} وثيقة بنجاح',
                    'redirect': url_for('documents.index')
                })
        
        return render_template('documents/department_bulk_create.html',
                             departments=departments,
                             employees=employees,
                             document_types=document_types)
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@documents_bp.route('/get_employees_by_sponsorship', methods=['POST'])
def get_employees_by_sponsorship():
    """Get employees filtered by sponsorship status"""
    try:
        data = request.get_json()
        sponsorship_type = data.get('sponsorship_type')  # 'internal' or 'external'
        
        # Query employees based on sponsorship status
        if sponsorship_type == 'internal':
            employees = Employee.query.filter(Employee.sponsorship_status == 'على الكفالة').all()
        elif sponsorship_type == 'external':
            employees = Employee.query.filter(Employee.sponsorship_status == 'خارج الكفالة').all()
        else:
            return jsonify({'success': False, 'message': 'نوع الكفالة غير صحيح'})
        
        # Format employee data
        employees_data = []
        for emp in employees:
            dept_names = ', '.join([dept.name for dept in emp.departments]) if emp.departments else 'غير محدد'
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id,
                'national_id': emp.national_id,
                'department_names': dept_names
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@documents_bp.route('/get_employees_by_department_and_sponsorship', methods=['POST'])
def get_employees_by_department_and_sponsorship():
    """Get employees filtered by department and sponsorship status"""
    try:
        data = request.get_json()
        department_id = data.get('department_id')
        sponsorship_type = data.get('sponsorship_type')
        
        # Build base query
        query = Employee.query.options(selectinload(Employee.departments))
        
        # Filter by department
        if department_id:
            query = query.filter(Employee.departments.any(Department.id == department_id))
        
        # Filter by sponsorship status
        if sponsorship_type == 'on_sponsorship':
            query = query.filter(Employee.sponsorship_status == 'على الكفالة')
        elif sponsorship_type == 'off_sponsorship':
            query = query.filter(Employee.sponsorship_status == 'خارج الكفالة')
        
        employees = query.all()
        
        # Format employee data
        employees_data = []
        for emp in employees:
            dept_names = ', '.join([dept.name for dept in emp.departments]) if emp.departments else 'غير محدد'
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id,
                'national_id': emp.national_id,
                'department_names': dept_names,
                'sponsorship_status': sponsorship_type
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@documents_bp.route('/save_individual_document', methods=['POST'])
def save_individual_document():
    """Save individual document for sponsorship-based addition"""
    try:
        data = request.get_json()
        
        # Create new document
        document = Document(
            employee_id=data['employee_id'],
            document_type=data['document_type'],
            document_number=data['document_number'],
            issue_date=datetime.strptime(data['issue_date'], '%Y-%m-%d').date() if data['issue_date'] else None,
            expiry_date=datetime.strptime(data['expiry_date'], '%Y-%m-%d').date() if data['expiry_date'] else None,
            notes=data.get('notes', ''),
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        db.session.add(document)
        db.session.commit()
        
        # Log the activity
        log_activity(
            action='create',
            entity_type='document',
            entity_id=document.id,
            details=f'تم إضافة وثيقة {data["document_type"]} للموظف {data["employee_id"]} فردياً'
        )
        
        return jsonify({
            'success': True,
            'message': 'تم حفظ الوثيقة بنجاح'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'حدث خطأ في حفظ الوثيقة: {str(e)}'
        })

@documents_bp.route('/save_bulk_documents', methods=['POST'])
def save_bulk_documents():
    """Save bulk documents for advanced filtering"""
    try:
        data = request.get_json()
        document_type = data['document_type']
        documents_data = data['documents']
        
        saved_count = 0
        
        for doc_data in documents_data:
            if doc_data.get('document_number') or doc_data.get('issue_date') or doc_data.get('expiry_date'):
                document = Document(
                    employee_id=doc_data['employee_id'],
                    document_type=document_type,
                    document_number=doc_data.get('document_number', ''),
                    issue_date=datetime.strptime(doc_data['issue_date'], '%Y-%m-%d').date() if doc_data.get('issue_date') else None,
                    expiry_date=datetime.strptime(doc_data['expiry_date'], '%Y-%m-%d').date() if doc_data.get('expiry_date') else None,
                    notes=doc_data.get('notes', ''),
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                
                db.session.add(document)
                saved_count += 1
        
        db.session.commit()
        
        # Log the activity
        log_activity(
            action='create',
            entity_type='document',
            entity_id=0,
            details=f'تم إضافة {saved_count} وثيقة من نوع {document_type} بشكل جماعي'
        )
        
        return jsonify({
            'success': True,
            'message': f'تم حفظ {saved_count} وثيقة بنجاح',
            'saved_count': saved_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'حدث خطأ في حفظ الوثائق: {str(e)}'
        })

# Duplicate route code removed - using the one above

@documents_bp.route('/')
def index():
    """List document records with filtering options"""
    # Get filter parameters
    document_type = request.args.get('document_type', '')
    employee_id = request.args.get('employee_id', '')
    department_id = request.args.get('department_id', '')
    sponsorship_status = request.args.get('sponsorship_status', '')
    status_filter = request.args.get('expiring', '')  # Fixed parameter name
    show_all = request.args.get('show_all', 'false')
    search_query = request.args.get('search_query', '').strip()  # حقل البحث الجديد
    
    # Build query
    query = Document.query
    
    # Apply filters
    if document_type:
        query = query.filter(Document.document_type == document_type)
    
    if employee_id and employee_id.isdigit():
        query = query.filter(Document.employee_id == int(employee_id))
    
    # تصفية حسب القسم والكفالة (نحتاج للـ join مع Employee)
    if department_id and department_id.isdigit():
        # فلترة الوثائق للموظفين في قسم محدد
        dept_employees = Employee.query.join(Employee.departments).filter_by(id=int(department_id)).all()
        dept_employee_ids = [emp.id for emp in dept_employees]
        if dept_employee_ids:
            query = query.filter(Document.employee_id.in_(dept_employee_ids))
        else:
            # لا توجد موظفين في هذا القسم
            query = query.filter(False)
    
    if sponsorship_status:
        query = query.join(Employee).filter(Employee.sponsorship_status == sponsorship_status)
    
    # إضافة البحث بالاسم والرقم الوظيفي ورقم الهوية
    if search_query:
        # نضمن أن هناك join مع Employee للبحث
        if not sponsorship_status:
            query = query.join(Employee)
            
        # البحث في اسم الموظف أو رقم الموظف أو رقم الهوية الوطنية مع معالجة القيم الفارغة
        search_conditions = []
        
        # البحث في الاسم
        search_conditions.append(Employee.name.ilike(f'%{search_query}%'))
        
        # البحث في رقم الموظف (مع التعامل مع القيم الفارغة)
        search_conditions.append(
            func.coalesce(Employee.employee_id, '').ilike(f'%{search_query}%')
        )
        
        # البحث في رقم الهوية (مع التعامل مع القيم الفارغة)
        search_conditions.append(
            func.coalesce(Employee.national_id, '').ilike(f'%{search_query}%')
        )
        
        query = query.filter(or_(*search_conditions))
    
    # تطبيق فلتر حالة الصلاحية
    today = datetime.now().date()
    
    if status_filter == 'expired':
        # الوثائق المنتهية فقط
        query = query.filter(
            Document.expiry_date.isnot(None),
            Document.expiry_date < today
        )
    elif status_filter == 'expiring_30':
        # الوثائق التي تنتهي خلال 30 يوم
        future_date = today + timedelta(days=30)
        query = query.filter(
            Document.expiry_date.isnot(None),
            Document.expiry_date >= today,
            Document.expiry_date <= future_date
        )
    elif status_filter == 'expiring_60':
        # الوثائق التي تنتهي خلال 60 يوم
        future_date = today + timedelta(days=60)
        query = query.filter(
            Document.expiry_date.isnot(None),
            Document.expiry_date >= today,
            Document.expiry_date <= future_date
        )
    elif status_filter == 'expiring_90':
        # الوثائق التي تنتهي خلال 90 يوم
        future_date = today + timedelta(days=90)
        query = query.filter(
            Document.expiry_date.isnot(None),
            Document.expiry_date >= today,
            Document.expiry_date <= future_date
        )
    elif status_filter == 'valid':
        # الوثائق السارية (أكثر من 30 يوم للانتهاء)
        future_date = today + timedelta(days=30)
        query = query.filter(
            or_(
                Document.expiry_date.is_(None),  # الوثائق بدون تاريخ انتهاء
                Document.expiry_date > future_date  # الوثائق التي تنتهي بعد أكثر من 30 يوم
            )
        )
    elif show_all.lower() != 'true' and not search_query:
        # العرض الافتراضي فقط إذا لم يكن هناك بحث: الوثائق المنتهية أو القريبة من الانتهاء (خلال 30 يوم)
        future_date_30_days = today + timedelta(days=30)
        query = query.filter(
            Document.expiry_date.isnot(None),
            Document.expiry_date <= future_date_30_days
        )
    
    # Execute query with eager loading للموظف
    documents = query.options(selectinload(Document.employee)).all()
    
    # احسب عدد الوثائق الكلي والمنتهية والقريبة من الانتهاء
    total_docs = Document.query.count()
    # حساب عدد الوثائق المنتهية (يجب أن يكون لها تاريخ انتهاء حتى تعتبر منتهية)
    expired_docs = Document.query.filter(
        Document.expiry_date.isnot(None),
        Document.expiry_date < today
    ).count()
    # حساب عدد الوثائق التي ستنتهي قريباً
    expiring_soon = Document.query.filter(
        Document.expiry_date.isnot(None),
        Document.expiry_date <= today + timedelta(days=30),
        Document.expiry_date >= today
    ).count()
    # عدد الوثائق الآمنة والتي لها تاريخ انتهاء
    docs_with_expiry = Document.query.filter(Document.expiry_date.isnot(None)).count()
    safe_docs = docs_with_expiry - expired_docs - expiring_soon
    
    # Get all employees for filter dropdown
    employees = Employee.query.all()
    
    # Get all departments for filter dropdown
    departments = Department.query.all()
    
    # Get document types for filter dropdown
    document_types = [
        'national_id', 'passport', 'health_certificate', 
        'work_permit', 'education_certificate', 'driving_license',
        'annual_leave', 'other'
    ]
    
    return render_template('documents/index.html',
                          documents=documents,
                          employees=employees,
                          departments=departments,
                          document_types=document_types,
                          selected_type=document_type,
                          selected_employee=employee_id,
                          selected_department=department_id,
                          selected_sponsorship=sponsorship_status,
                          selected_status_filter=status_filter,
                          search_query=search_query,
                          show_all=show_all.lower() == 'true',
                          total_docs=total_docs,
                          expired_docs=expired_docs,
                          expiring_soon=expiring_soon,
                          safe_docs=safe_docs,
                          valid_docs=safe_docs,
                          status_filter=status_filter,
                          today=today,
                          now=datetime.now())

@documents_bp.route('/create', methods=['GET', 'POST'])
def create():
    """Create a new document record"""
    if request.method == 'POST':
        try:
            # تحقق من وجود CSRF token
            if 'csrf_token' not in request.form:
                flash('خطأ في التحقق من الأمان. يرجى المحاولة مرة أخرى.', 'danger')
                return redirect(url_for('documents.create'))
                
            document_type = request.form['document_type']
            document_number = request.form.get('document_number', '')
            issue_date_str = request.form.get('issue_date', '')
            expiry_date_str = request.form.get('expiry_date', '')
            notes = request.form.get('notes', '')
            add_type = request.form.get('add_type', 'single')
            
            # Parse dates (فقط إذا تم إدخالها)
            issue_date = parse_date(issue_date_str) if issue_date_str else None
            expiry_date = parse_date(expiry_date_str) if expiry_date_str else None
            
            # تحديد ما إذا كان الإضافة لموظف واحد أو لقسم كامل
            if add_type == 'single':
                # إضافة وثيقة لموظف واحد
                employee_id = request.form.get('employee_id')
                if not employee_id:
                    flash('يرجى اختيار الموظف', 'danger')
                    return redirect(url_for('documents.create'))
                
                # Create new document record
                document = Document(
                    employee_id=employee_id,
                    document_type=document_type,
                    document_number=document_number,
                    issue_date=issue_date,
                    expiry_date=expiry_date,
                    notes=notes
                )
                
                db.session.add(document)
                
                # Log the action
                employee = Employee.query.get(employee_id)
                audit = SystemAudit(
                    action='create',
                    entity_type='document',
                    entity_id=employee_id,
                    details=f'تم إضافة وثيقة جديدة من نوع {document_type} للموظف: {employee.name}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                flash('تم إضافة الوثيقة بنجاح', 'success')
            
            elif add_type == 'sponsorship_single':
                # حفظ موظف واحد من قائمة الكفالة
                employee_id = request.form.get('employee_id')
                
                if not employee_id:
                    return jsonify({'success': False, 'message': 'يرجى اختيار الموظف'})
                
                # Create new document record
                document = Document(
                    employee_id=employee_id,
                    document_type=document_type,
                    document_number=document_number,
                    issue_date=issue_date,
                    expiry_date=expiry_date,
                    notes=notes
                )
                
                db.session.add(document)
                
                # Log the action
                employee = Employee.query.get(employee_id)
                audit = SystemAudit(
                    action='create',
                    entity_type='document',
                    entity_id=employee_id,
                    details=f'تم إضافة وثيقة {document_type} للموظف: {employee.name} (حفظ فردي من قائمة الكفالة)',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({'success': True, 'message': 'تم حفظ الوثيقة بنجاح'})
            
            elif add_type == 'department_bulk':
                # حفظ جميع وثائق القسم
                import json
                employees_data = json.loads(request.form.get('employees_data', '[]'))
                department_id = request.form.get('department_id')
                
                if not employees_data:
                    return jsonify({'success': False, 'message': 'لا توجد بيانات للحفظ'})
                
                saved_count = 0
                for emp_data in employees_data:
                    if emp_data.get('document_number'):
                        document = Document(
                            employee_id=emp_data['employee_id'],
                            document_type=document_type,
                            document_number=emp_data['document_number'],
                            issue_date=parse_date(emp_data['issue_date']) if emp_data.get('issue_date') else None,
                            expiry_date=parse_date(emp_data['expiry_date']) if emp_data.get('expiry_date') else None,
                            notes=emp_data.get('notes', '')
                        )
                        
                        db.session.add(document)
                        saved_count += 1
                
                # Log the action
                department = Department.query.get(department_id)
                audit = SystemAudit(
                    action='bulk_create',
                    entity_type='document',
                    entity_id=department_id,
                    details=f'تم إضافة {saved_count} وثيقة من نوع {document_type} لقسم {department.name if department else "غير محدد"}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({'success': True, 'message': f'تم حفظ {saved_count} وثيقة بنجاح'})
            
            elif add_type == 'sponsorship_bulk':
                # حفظ جميع بيانات الكفالة
                import json
                employees_data = json.loads(request.form.get('employees', '[]'))
                
                if not employees_data:
                    return jsonify({'success': False, 'message': 'لا توجد بيانات للحفظ'})
                
                saved_count = 0
                for emp_data in employees_data:
                    # تحقق من وجود بيانات للحفظ
                    if emp_data.get('documentNumber'):
                        document = Document(
                            employee_id=emp_data['id'],
                            document_type=document_type,
                            document_number=emp_data['documentNumber'],
                            issue_date=parse_date(emp_data['issueDate']) if emp_data.get('issueDate') else None,
                            expiry_date=parse_date(emp_data['expiryDate']) if emp_data.get('expiryDate') else None,
                            notes=emp_data.get('notes', '')
                        )
                        
                        db.session.add(document)
                        saved_count += 1
                
                # Log the action
                audit = SystemAudit(
                    action='bulk_create',
                    entity_type='document',
                    entity_id=None,
                    details=f'تم إضافة {saved_count} وثيقة من نوع {document_type} (حفظ جماعي من قائمة الكفالة)',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({'success': True, 'message': f'تم حفظ {saved_count} وثيقة بنجاح'})
            
            elif add_type == 'sponsorship_individual':
                # حفظ موظف واحد من الإضافة الجماعية
                employee_id = request.form.get('employee_id')
                sponsorship_type = request.form.get('sponsorship_type')
                
                if not employee_id:
                    return jsonify({'success': False, 'message': 'يرجى اختيار الموظف'})
                
                # Create new document record
                document = Document(
                    employee_id=employee_id,
                    document_type=document_type,
                    document_number=document_number,
                    issue_date=issue_date,
                    expiry_date=expiry_date,
                    notes=notes
                )
                
                db.session.add(document)
                
                # Log the action
                employee = Employee.query.get(employee_id)
                audit = SystemAudit(
                    action='create',
                    entity_type='document',
                    entity_id=employee_id,
                    details=f'تم إضافة وثيقة {document_type} (كفالة: {sponsorship_type}) للموظف: {employee.name}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                return jsonify({'success': True, 'message': 'تم حفظ الوثيقة بنجاح'})
            
            else:
                # إضافة وثيقة لقسم كامل
                department_id = request.form.get('department_id')
                if not department_id:
                    flash('يرجى اختيار القسم', 'danger')
                    return redirect(url_for('documents.create'))
                
                # الحصول على جميع موظفي القسم
                department = Department.query.get(department_id)
                employees = department.employees
                
                if not employees:
                    flash(f'لا يوجد موظفين في قسم "{department.name}"', 'warning')
                    return redirect(url_for('documents.create'))
                
                # إنشاء وثيقة لكل موظف في القسم
                document_count = 0
                for employee in employees:
                    # التحقق من عدم وجود وثيقة من نفس النوع للموظف (منع التكرار)
                    existing_document = Document.query.filter_by(
                        employee_id=employee.id,
                        document_type=document_type
                    ).first()
                    
                    # تخطي الموظف إذا كان لديه وثيقة من نفس النوع بالفعل
                    if existing_document:
                        continue
                        
                    # استخدام رقم الهوية الوطنية للموظف إذا كان متوفراً
                    # أو استخدام رقم الموظف التسلسلي إذا كان رقم الهوية غير متوفر
                    national_id = None
                    
                    # ابحث عن وثيقة هوية وطنية مسجلة للموظف
                    existing_national_id = Document.query.filter_by(
                        employee_id=employee.id,
                        document_type='national_id'
                    ).first()
                    
                    if existing_national_id:
                        national_id = existing_national_id.document_number
                    
                    # إذا لم نجد رقم هوية، نستخدم الرقم الوظيفي (المسلسل) للموظف
                    document_number_to_use = national_id if national_id else f"ID-{employee.employee_id}"
                    
                    document = Document(
                        employee_id=employee.id,
                        document_type=document_type,
                        document_number=document_number_to_use,
                        issue_date=issue_date,
                        expiry_date=expiry_date,
                        notes=notes
                    )
                    
                    db.session.add(document)
                    document_count += 1
                
                # إنشاء سجل تدقيق للعملية
                audit = SystemAudit(
                    action='create_bulk',
                    entity_type='document',
                    entity_id=department_id,
                    details=f'تم إضافة {document_count} وثيقة من نوع {document_type} لقسم: {department.name}',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                flash(f'تم إضافة {document_count} وثائق بنجاح لجميع موظفي قسم "{department.name}"', 'success')
            
            return redirect(url_for('documents.index'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # Get all employees for dropdown with their departments loaded
    employees = Employee.query.options(db.selectinload(Employee.departments)).all()
    
    # Get all departments for dropdown
    departments = Department.query.all()
    
    # Get document types for dropdown
    document_types = [
        'national_id', 'passport', 'health_certificate', 
        'work_permit', 'education_certificate', 'driving_license',
        'annual_leave', 'other'
    ]
    
    # Default dates
    today = datetime.now().date()
    hijri_today = format_date_hijri(today)
    gregorian_today = format_date_gregorian(today)
    
    return render_template('documents/create.html',
                          employees=employees,
                          departments=departments,
                          document_types=document_types,
                          today=today,
                          hijri_today=hijri_today,
                          gregorian_today=gregorian_today)

@documents_bp.route('/<int:id>/confirm-delete')
def confirm_delete(id):
    """صفحة تأكيد حذف وثيقة"""
    document = Document.query.get_or_404(id)
    
    # تحويل أنواع الوثائق إلى عربي للعرض
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'health_certificate': 'الشهادة الصحية',
        'work_permit': 'تصريح العمل',
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'residency_permit': 'تصريح الإقامة',
        'visa': 'تأشيرة',
        'insurance': 'التأمين',
        'contract': 'العقد',
        'certification': 'شهادة مهنية',
        'training_certificate': 'شهادة تدريب',
        'other': 'أخرى'
    }
    
    # الحصول على اسم نوع الوثيقة بالعربية
    document_type_ar = document_types_map.get(document.document_type, document.document_type)
    
    return render_template('documents/confirm_delete.html', 
                          document=document, 
                          document_type_ar=document_type_ar)

@documents_bp.route('/<int:id>/delete', methods=['POST'])
def delete(id):
    """Delete a document record"""
    # تحقق من وجود CSRF token
    if 'csrf_token' not in request.form:
        flash('خطأ في التحقق من الأمان. يرجى المحاولة مرة أخرى.', 'danger')
        return redirect(url_for('documents.index'))
    
    document = Document.query.get_or_404(id)
    employee_name = document.employee.name
    document_type = document.document_type
    
    try:
        db.session.delete(document)
        
        # Log the action
        audit = SystemAudit(
            action='delete',
            entity_type='document',
            entity_id=id,
            details=f'تم حذف وثيقة من نوع {document_type} للموظف: {employee_name}',
            user_id=current_user.id if current_user.is_authenticated else None
        )
        db.session.add(audit)
        db.session.commit()
        
        flash('تم حذف الوثيقة بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الوثيقة: {str(e)}', 'danger')
    
    return redirect(url_for('documents.index'))

@documents_bp.route('/<int:id>/update_expiry', methods=['GET', 'POST'])
def update_expiry(id):
    """تحديث تاريخ انتهاء الوثيقة"""
    document = Document.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # تحقق من وجود CSRF token
            if 'csrf_token' not in request.form:
                flash('خطأ في التحقق من الأمان. يرجى المحاولة مرة أخرى.', 'danger')
                return redirect(url_for('documents.update_expiry', id=id))
            
            expiry_date_str = request.form['expiry_date']
            # تحليل التاريخ
            expiry_date = parse_date(expiry_date_str)
            
            # حفظ التاريخ القديم للسجل
            old_expiry_date = document.expiry_date
            
            # تحديث تاريخ الانتهاء
            document.expiry_date = expiry_date
            
            # إضافة سجل للتدقيق
            audit = SystemAudit(
                action='update',
                entity_type='document',
                entity_id=id,
                details=f'تم تحديث تاريخ انتهاء وثيقة {document.document_type} للموظف: {document.employee.name} من {old_expiry_date} إلى {expiry_date}',
                user_id=current_user.id if current_user.is_authenticated else None
            )
            db.session.add(audit)
            db.session.commit()
            
            flash('تم تحديث تاريخ انتهاء الوثيقة بنجاح', 'success')
            return redirect(url_for('documents.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث تاريخ انتهاء الوثيقة: {str(e)}', 'danger')
            return redirect(url_for('documents.update_expiry', id=id))
    
    # Get document types for dropdown
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'health_certificate': 'الشهادة الصحية',
        'work_permit': 'تصريح العمل',
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية',
        'other': 'أخرى'
    }
    
    # احصل على اسم نوع الوثيقة بالعربية
    doc_type_ar = document_types_map.get(document.document_type, document.document_type)
    
    # Default dates
    today = datetime.now().date()
    hijri_today = format_date_hijri(today)
    gregorian_today = format_date_gregorian(today)
    
    return render_template('documents/update_expiry.html',
                          document=document,
                          document_type_ar=doc_type_ar,
                          today=today,
                          hijri_today=hijri_today,
                          gregorian_today=gregorian_today)

@documents_bp.route('/import', methods=['GET', 'POST'])
def import_excel():
    """Import document records from Excel file"""
    if request.method == 'POST':
        # تحقق من وجود CSRF token
        if 'csrf_token' not in request.form:
            flash('خطأ في التحقق من الأمان. يرجى المحاولة مرة أخرى.', 'danger')
            return redirect(request.url)
            
        if 'file' not in request.files:
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                # Parse Excel file
                documents_data = parse_document_excel(file)
                success_count = 0
                error_count = 0
                
                for data in documents_data:
                    try:
                        document = Document(**data)
                        db.session.add(document)
                        db.session.commit()
                        success_count += 1
                    except Exception:
                        db.session.rollback()
                        error_count += 1
                
                # Log the import
                audit = SystemAudit(
                    action='import',
                    entity_type='document',
                    entity_id=0,
                    details=f'تم استيراد {success_count} وثيقة بنجاح و {error_count} فشل',
                    user_id=current_user.id if current_user.is_authenticated else None
                )
                db.session.add(audit)
                db.session.commit()
                
                if error_count > 0:
                    flash(f'تم استيراد {success_count} وثيقة بنجاح و {error_count} فشل', 'warning')
                else:
                    flash(f'تم استيراد {success_count} وثيقة بنجاح', 'success')
                return redirect(url_for('documents.index'))
            except Exception as e:
                flash(f'حدث خطأ أثناء استيراد الملف: {str(e)}', 'danger')
        else:
            flash('الملف يجب أن يكون بصيغة Excel (.xlsx, .xls)', 'danger')
    
    return render_template('documents/import.html')

@documents_bp.route('/expiring')
def expiring():
    """Show documents that are about to expire or already expired"""
    days = int(request.args.get('days', '30'))
    document_type = request.args.get('document_type', '')
    status = request.args.get('status', 'expiring')  # 'expiring' or 'expired'
    employee_id = request.args.get('employee_id', '')
    department_id = request.args.get('department_id', '')
    sponsorship_status = request.args.get('sponsorship_status', '')
    
    # Calculate expiry date range
    today = datetime.now().date()
    future_date = today + timedelta(days=days)
    
    # Build query for documents based on status
    query = Document.query.filter(Document.expiry_date.isnot(None))  # فقط الوثائق التي لها تاريخ انتهاء
    
    if status == 'expired':
        # Get documents that have already expired
        query = query.filter(Document.expiry_date < today)
    else:
        # Get documents that are about to expire
        query = query.filter(
            Document.expiry_date <= future_date,
            Document.expiry_date >= today
        )
    
    # Apply document type filter if provided
    if document_type:
        query = query.filter(Document.document_type == document_type)
    
    # Apply employee filter if provided
    if employee_id and employee_id.isdigit():
        query = query.filter(Document.employee_id == int(employee_id))
    
    # Apply filters that require Employee join
    needs_employee_join = department_id or sponsorship_status
    
    if needs_employee_join:
        query = query.join(Employee)
        
        if department_id and department_id.isdigit():
            query = query.filter(Employee.department_id == int(department_id))
        
        if sponsorship_status:
            query = query.filter(Employee.sponsorship_status == sponsorship_status)
    
    # Execute query
    documents = query.all()
    
    # Calculate days to expiry for each document
    today = datetime.now().date()
    for doc in documents:
        if doc.expiry_date:
            doc.days_to_expiry = (doc.expiry_date - today).days
        else:
            doc.days_to_expiry = None
    
    # Get document types for filter dropdown
    document_types = [
        'national_id', 'passport', 'health_certificate', 
        'work_permit', 'education_certificate', 'driving_license',
        'annual_leave', 'other'
    ]
    
    # Get all employees and departments for filter dropdowns
    employees = Employee.query.all()
    departments = Department.query.all()
    
    return render_template('documents/expiring.html',
                          documents=documents,
                          days=days,
                          document_types=document_types,
                          employees=employees,
                          departments=departments,
                          selected_type=document_type,
                          selected_employee=employee_id,
                          selected_department=department_id,
                          selected_sponsorship=sponsorship_status,
                          status=status)

@documents_bp.route('/expiry_stats')
def expiry_stats():
    """Get document expiry statistics"""
    # Calculate expiry date ranges
    today = datetime.now().date()
    thirty_days = today + timedelta(days=30)
    sixty_days = today + timedelta(days=60)
    ninety_days = today + timedelta(days=90)
    
    # استبعاد الوثائق التي ليس لها تاريخ انتهاء
    base_query = Document.query.filter(Document.expiry_date.isnot(None))
    
    # Get count of documents expiring in different periods
    expiring_30 = base_query.filter(
        Document.expiry_date <= thirty_days,
        Document.expiry_date >= today
    ).count()
    
    expiring_60 = base_query.filter(
        Document.expiry_date <= sixty_days,
        Document.expiry_date > thirty_days
    ).count()
    
    expiring_90 = base_query.filter(
        Document.expiry_date <= ninety_days,
        Document.expiry_date > sixty_days
    ).count()
    
    expired = base_query.filter(
        Document.expiry_date < today
    ).count()
    
    # Get document counts by type
    type_counts = db.session.query(
        Document.document_type,
        func.count(Document.id).label('count')
    ).group_by(Document.document_type).all()
    
    # Format for response
    type_stats = {}
    for doc_type, count in type_counts:
        type_stats[doc_type] = count
    
    return jsonify({
        'expiring_30': expiring_30,
        'expiring_60': expiring_60,
        'expiring_90': expiring_90,
        'expired': expired,
        'type_stats': type_stats
    })

@documents_bp.route('/employee/<int:employee_id>/export_pdf')
def export_employee_documents_pdf(employee_id):
    """Export employee documents to PDF"""
    employee = Employee.query.get_or_404(employee_id)
    documents = Document.query.filter_by(employee_id=employee_id).all()
    
    # إنشاء ملف PDF
    buffer = BytesIO()
    
    # تسجيل الخط العربي - استخدام خط Amiri لأنه يدعم الربط العربي بشكل ممتاز
    try:
        pdfmetrics.registerFont(TTFont('ArabicFont', 'static/fonts/Amiri-Regular.ttf'))
        pdfmetrics.registerFont(TTFont('ArabicFontBold', 'static/fonts/Amiri-Bold.ttf'))
        arabic_font_name = 'ArabicFont'
        arabic_font_bold = 'ArabicFontBold'
    except Exception as e:
        # إذا كان هناك خطأ، نستخدم خط بديل
        try:
            pdfmetrics.registerFont(TTFont('ArabicFont', 'static/fonts/Cairo.ttf'))
            arabic_font_name = 'ArabicFont'
            arabic_font_bold = 'ArabicFont'
        except:
            # آخر خيار: استخدام الخط الافتراضي
            arabic_font_name = 'Helvetica'
            arabic_font_bold = 'Helvetica-Bold'
    
    # تعيين أبعاد الصفحة واتجاهها
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # إعداد الأنماط
    styles = getSampleStyleSheet()
    # إنشاء نمط للنص العربي
    arabic_style = ParagraphStyle(
        name='Arabic',
        parent=styles['Normal'],
        fontName=arabic_font_name,
        fontSize=10,
        alignment=2, # يمين (RTL)
        textColor=colors.black
    )
    
    # إنشاء نمط للعناوين
    title_style = ParagraphStyle(
        name='Title',
        parent=styles['Title'],
        fontName=arabic_font_bold,
        fontSize=18,
        alignment=1, # وسط
        textColor=colors.HexColor('#1e3a5c'),
        spaceAfter=12
    )
    
    # إنشاء نمط للعناوين الفرعية
    subtitle_style = ParagraphStyle(
        name='Subtitle',
        parent=styles['Heading2'],
        fontName=arabic_font_bold,
        fontSize=14,
        alignment=2, # يمين (RTL)
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=8
    )
    
    # إعداد المحتوى
    elements = []
    
    # إضافة العنوان
    title = f"وثائق الموظف: {employee.name}"
    # تهيئة النص العربي للعرض في PDF
    title = get_display(arabic_reshaper.reshape(title))
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 20))
    
    # إضافة بيانات الموظف في جدول
    employee_data = [
        [get_display(arabic_reshaper.reshape("بيانات الموظف")), "", get_display(arabic_reshaper.reshape("معلومات العمل")), ""],
        [
            get_display(arabic_reshaper.reshape("الاسم:")), 
            get_display(arabic_reshaper.reshape(employee.name)), 
            get_display(arabic_reshaper.reshape("المسمى الوظيفي:")), 
            get_display(arabic_reshaper.reshape(employee.job_title))
        ],
        [
            get_display(arabic_reshaper.reshape("الرقم الوظيفي:")), 
            employee.employee_id, 
            get_display(arabic_reshaper.reshape("القسم:")), 
            get_display(arabic_reshaper.reshape(', '.join([dept.name for dept in employee.departments]) if employee.departments else '-'))
        ],
        [
            get_display(arabic_reshaper.reshape("رقم الهوية:")), 
            employee.national_id, 
            get_display(arabic_reshaper.reshape("الحالة:")), 
            get_display(arabic_reshaper.reshape(employee.status))
        ],
        [
            get_display(arabic_reshaper.reshape("رقم الجوال:")), 
            employee.mobile, 
            get_display(arabic_reshaper.reshape("الموقع:")), 
            get_display(arabic_reshaper.reshape(employee.location or '-'))
        ]
    ]
    
    # إنشاء جدول بيانات الموظف
    employee_table = Table(employee_data, colWidths=[3*cm, 5*cm, 3*cm, 5*cm])
    employee_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#93c5fd')),
        ('BACKGROUND', (2, 0), (3, 0), colors.HexColor('#93c5fd')),
        ('TEXTCOLOR', (0, 0), (3, 0), colors.black),
        ('FONTNAME', (0, 0), (3, 0), arabic_font_bold),
        ('FONTNAME', (0, 1), (3, 4), arabic_font_name),
        ('FONTSIZE', (0, 0), (3, 0), 11),
        ('FONTSIZE', (0, 1), (3, 4), 10),
        ('SPAN', (0, 0), (1, 0)),
        ('SPAN', (2, 0), (3, 0)),
        ('ALIGN', (0, 0), (3, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, 4), 'RIGHT'),
        ('ALIGN', (2, 1), (2, 4), 'RIGHT'),
        ('VALIGN', (0, 0), (3, 4), 'MIDDLE'),
        ('GRID', (0, 0), (3, 4), 1, colors.grey),
        ('BOX', (0, 0), (3, 4), 2, colors.HexColor('#1e3a5c')),
        ('LEFTPADDING', (0, 0), (3, 4), 8),
        ('RIGHTPADDING', (0, 0), (3, 4), 8),
        ('TOPPADDING', (0, 0), (3, 4), 6),
        ('BOTTOMPADDING', (0, 0), (3, 4), 6),
    ]))
    elements.append(employee_table)
    elements.append(Spacer(1, 20))
    
    # إضافة عنوان قائمة الوثائق
    subtitle = get_display(arabic_reshaper.reshape("قائمة الوثائق"))
    elements.append(Paragraph(subtitle, subtitle_style))
    elements.append(Spacer(1, 10))
    
    # إنشاء جدول الوثائق
    headers = [
        get_display(arabic_reshaper.reshape("نوع الوثيقة")),
        get_display(arabic_reshaper.reshape("رقم الوثيقة")),
        get_display(arabic_reshaper.reshape("تاريخ الإصدار")),
        get_display(arabic_reshaper.reshape("تاريخ الانتهاء")),
        get_display(arabic_reshaper.reshape("الحالة")),
        get_display(arabic_reshaper.reshape("ملاحظات"))
    ]
    
    data = [headers]
    
    # إضافة صفوف الوثائق
    today = datetime.now().date()
    
    # ترجمة أنواع الوثائق
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'health_certificate': 'الشهادة الصحية',
        'work_permit': 'تصريح العمل',
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية'
    }
    
    for doc_item in documents:
        # الحصول على نوع الوثيقة بالعربية
        doc_type_ar = document_types_map.get(doc_item.document_type, doc_item.document_type)
        
        # التحقق من حالة انتهاء الصلاحية
        days_to_expiry = (doc_item.expiry_date - today).days
        if days_to_expiry < 0:
            status_text = "منتهية"
        elif days_to_expiry < 30:
            status_text = f"تنتهي خلال {days_to_expiry} يوم"
        else:
            status_text = "سارية"
        
        # إضافة صف للجدول
        row = [
            get_display(arabic_reshaper.reshape(doc_type_ar)),
            doc_item.document_number,
            format_date_gregorian(doc_item.issue_date),
            format_date_gregorian(doc_item.expiry_date),
            get_display(arabic_reshaper.reshape(status_text)),
            get_display(arabic_reshaper.reshape(doc_item.notes or '-'))
        ]
        data.append(row)
    
    # إنشاء جدول الوثائق إذا كان هناك وثائق
    if len(data) > 1:
        # حساب عرض الأعمدة بناءً على عرض الصفحة
        table_width = A4[0] - 4*cm  # العرض الإجمالي ناقص الهوامش
        col_widths = [3.5*cm, 3*cm, 2.5*cm, 2.5*cm, 3*cm, 3*cm]
        documents_table = Table(data, colWidths=col_widths)
        
        # تنسيق الجدول
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#93c5fd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), arabic_font_bold),
            ('FONTNAME', (0, 1), (-1, -1), arabic_font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a5c')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        # تطبيق التناوب في ألوان الصفوف
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
            
            # إضافة ألوان حالة انتهاء الصلاحية
            days_to_expiry = (documents[i-1].expiry_date - today).days
            if days_to_expiry < 0:
                table_style.add('TEXTCOLOR', (4, i), (4, i), colors.red)
                table_style.add('FONTSIZE', (4, i), (4, i), 10)
            elif days_to_expiry < 30:
                table_style.add('TEXTCOLOR', (4, i), (4, i), colors.orange)
        
        documents_table.setStyle(table_style)
        elements.append(documents_table)
    else:
        # إذا لم تكن هناك وثائق
        no_data_text = get_display(arabic_reshaper.reshape("لا توجد وثائق مسجلة لهذا الموظف"))
        elements.append(Paragraph(no_data_text, arabic_style))
    
    # إضافة معلومات التقرير في أسفل الصفحة
    elements.append(Spacer(1, 30))
    footer_text = f"تم إنشاء هذا التقرير بتاريخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    footer_text = get_display(arabic_reshaper.reshape(footer_text))
    elements.append(Paragraph(footer_text, arabic_style))
    
    # بناء المستند
    doc.build(elements)
    
    # إعادة المؤشر إلى بداية البايت
    buffer.seek(0)
    
    # إنشاء استجابة تحميل
    buffer.seek(0)
    return make_response(send_file(
        buffer,
        as_attachment=True,
        download_name=f'employee_{employee_id}_documents.pdf',
        mimetype='application/pdf'
    ))

@documents_bp.route('/employee/<int:employee_id>/export_excel')
def export_employee_documents_excel(employee_id):
    """Export employee documents to Excel"""
    employee = Employee.query.get_or_404(employee_id)
    documents = Document.query.filter_by(employee_id=employee_id).all()
    
    # Create Excel in memory
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet("الوثائق")
    
    # Add formatting
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#D3E0EA',
        'border': 1,
        'font_size': 13
    })
    
    # RTL format for workbook
    worksheet.right_to_left()
    
    # Add cell formats
    cell_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'font_size': 11
    })
    
    date_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'font_size': 11,
        'num_format': 'dd/mm/yyyy'
    })
    
    # Write headers
    headers = ['نوع الوثيقة', 'رقم الوثيقة', 'تاريخ الإصدار', 'تاريخ الانتهاء', 'ملاحظات']
    for col_num, data in enumerate(headers):
        worksheet.write(0, col_num, data, header_format)
    
    # Adjust column widths
    worksheet.set_column(0, 0, 20)  # نوع الوثيقة
    worksheet.set_column(1, 1, 20)  # رقم الوثيقة
    worksheet.set_column(2, 2, 15)  # تاريخ الإصدار
    worksheet.set_column(3, 3, 15)  # تاريخ الانتهاء
    worksheet.set_column(4, 4, 30)  # ملاحظات
    
    # Map for document types
    document_types_map = {
        'national_id': 'الهوية الوطنية',
        'passport': 'جواز السفر',
        'health_certificate': 'الشهادة الصحية',
        'work_permit': 'تصريح العمل',
        'education_certificate': 'الشهادة الدراسية',
        'driving_license': 'رخصة القيادة',
        'annual_leave': 'الإجازة السنوية'
    }
    
    # Write data
    for row_num, doc in enumerate(documents, 1):
        # Get document type in Arabic
        doc_type_ar = document_types_map.get(doc.document_type, doc.document_type)
            
        worksheet.write(row_num, 0, doc_type_ar, cell_format)
        worksheet.write(row_num, 1, doc.document_number, cell_format)
        worksheet.write_datetime(row_num, 2, doc.issue_date, date_format)
        worksheet.write_datetime(row_num, 3, doc.expiry_date, date_format)
        worksheet.write(row_num, 4, doc.notes or '', cell_format)
    
    # Add title with employee info
    info_worksheet = workbook.add_worksheet("معلومات الموظف")
    info_worksheet.right_to_left()
    
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#B8D9EB',
        'border': 2
    })
    
    info_worksheet.merge_range('A1:B1', f'بيانات الموظف: {employee.name}', title_format)
    info_worksheet.set_column(0, 0, 20)
    info_worksheet.set_column(1, 1, 30)
    
    field_format = workbook.add_format({
        'bold': True,
        'align': 'right',
        'valign': 'vcenter',
        'bg_color': '#F0F0F0',
        'border': 1
    })
    
    info_fields = [
        ['الاسم', employee.name],
        ['الرقم الوظيفي', employee.employee_id],
        ['رقم الهوية', employee.national_id],
        ['رقم الجوال', employee.mobile],
        ['القسم', ', '.join([dept.name for dept in employee.departments]) if employee.departments else ''],
        ['المسمى الوظيفي', employee.job_title],
        ['الحالة', employee.status],
        ['الموقع', employee.location or '']
    ]
    
    for row_num, (field, value) in enumerate(info_fields):
        info_worksheet.write(row_num + 1, 0, field, field_format)
        info_worksheet.write(row_num + 1, 1, value, cell_format)
    
    # Close workbook
    workbook.close()
    
    # Create response
    output.seek(0)
    return make_response(send_file(
        output,
        as_attachment=True,
        download_name=f'employee_{employee_id}_documents.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ))

@documents_bp.route('/export_excel')
@login_required
def export_excel():
    """تصدير الوثائق إلى ملف Excel حسب الفلاتر المطبقة"""
    try:
        # Get filter parameters (same as index route)
        document_type = request.args.get('document_type', '')
        employee_id = request.args.get('employee_id', '')
        department_id = request.args.get('department_id', '')
        sponsorship_status = request.args.get('sponsorship_status', '')
        status_filter = request.args.get('expiring', '')
        show_all = request.args.get('show_all', 'false')
        
        # Build query (same logic as index route)
        query = Document.query.options(
            selectinload(Document.employee).selectinload(Employee.departments)
        )
        
        # Apply filters
        if document_type:
            query = query.filter(Document.document_type == document_type)
        
        if employee_id and employee_id.isdigit():
            query = query.filter(Document.employee_id == int(employee_id))
        
        if department_id and department_id.isdigit():
            dept_employees = Employee.query.join(Employee.departments).filter_by(id=int(department_id)).all()
            dept_employee_ids = [emp.id for emp in dept_employees]
            if dept_employee_ids:
                query = query.filter(Document.employee_id.in_(dept_employee_ids))
            else:
                query = query.filter(False)
        
        if sponsorship_status:
            query = query.join(Employee).filter(Employee.sponsorship_status == sponsorship_status)
        
        # تطبيق فلتر حالة الصلاحية
        today = datetime.now().date()
        
        if status_filter == 'expired':
            query = query.filter(
                Document.expiry_date.isnot(None),
                Document.expiry_date < today
            )
        elif status_filter == 'expiring_30':
            future_date = today + timedelta(days=30)
            query = query.filter(
                Document.expiry_date.isnot(None),
                Document.expiry_date >= today,
                Document.expiry_date <= future_date
            )
        elif status_filter == 'expiring_60':
            future_date = today + timedelta(days=60)
            query = query.filter(
                Document.expiry_date.isnot(None),
                Document.expiry_date >= today,
                Document.expiry_date <= future_date
            )
        elif status_filter == 'expiring_90':
            future_date = today + timedelta(days=90)
            query = query.filter(
                Document.expiry_date.isnot(None),
                Document.expiry_date >= today,
                Document.expiry_date <= future_date
            )
        elif status_filter == 'valid':
            future_date = today + timedelta(days=30)
            query = query.filter(
                or_(
                    Document.expiry_date.is_(None),
                    Document.expiry_date > future_date
                )
            )
        elif show_all.lower() != 'true':
            future_date_30_days = today + timedelta(days=30)
            query = query.filter(
                Document.expiry_date.isnot(None),
                Document.expiry_date <= future_date_30_days
            )
        
        # Execute query
        documents = query.all()
        
        # تحضير البيانات للتصدير
        data = []
        
        # خريطة أنواع الوثائق
        document_types_map = {
            'national_id': 'الهوية الوطنية',
            'passport': 'جواز السفر',
            'health_certificate': 'الشهادة الصحية',
            'work_permit': 'تصريح العمل',
            'education_certificate': 'الشهادة الدراسية',
            'driving_license': 'رخصة القيادة',
            'annual_leave': 'الإجازة السنوية',
            'other': 'أخرى'
        }
        
        for doc in documents:
            # حساب حالة الوثيقة
            status = 'غير محدد'
            days_remaining = ''
            
            if doc.expiry_date:
                diff = (doc.expiry_date - today).days
                if diff < 0:
                    status = 'منتهية الصلاحية'
                    days_remaining = f'منتهية منذ {abs(diff)} يوم'
                elif diff <= 30:
                    status = 'تنتهي قريباً'
                    days_remaining = f'{diff} يوم متبقي'
                else:
                    status = 'سارية'
                    days_remaining = f'{diff} يوم متبقي'
            else:
                status = 'بدون تاريخ انتهاء'
                days_remaining = 'غير محدد'
            
            # معلومات الأقسام
            departments_list = ', '.join([dept.name for dept in doc.employee.departments]) if doc.employee.departments else 'غير محدد'
            
            row = {
                'نوع الوثيقة': document_types_map.get(doc.document_type, doc.document_type),
                'رقم الوثيقة': doc.document_number or '',
                'اسم الموظف': doc.employee.name if doc.employee else '',
                'رقم الموظف': doc.employee.employee_id if doc.employee else '',
                'رقم الهوية': doc.employee.national_id if doc.employee else '',
                'الأقسام': departments_list,
                'الجوال': doc.employee.mobile if doc.employee else '',
                'المنصب': doc.employee.job_title if doc.employee else '',
                'الحالة الوظيفية': doc.employee.status if doc.employee else '',
                'حالة الكفالة': doc.employee.sponsorship_status if doc.employee else '',
                'تاريخ الإصدار': doc.issue_date.strftime('%Y-%m-%d') if doc.issue_date else '',
                'تاريخ الانتهاء': doc.expiry_date.strftime('%Y-%m-%d') if doc.expiry_date else '',
                'حالة الوثيقة': status,
                'الأيام المتبقية': days_remaining,
                'ملاحظات': doc.notes or '',
                'تاريخ الإنشاء': doc.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(doc, 'created_at') and doc.created_at else '',
                'آخر تحديث': doc.updated_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(doc, 'updated_at') and doc.updated_at else ''
            }
            data.append(row)
        
        # إنشاء DataFrame
        df = pd.DataFrame(data)
        
        if df.empty:
            flash('لا توجد وثائق لتصديرها حسب الفلاتر المحددة', 'warning')
            return redirect(request.referrer or url_for('documents.index'))
        
        # إنشاء ملف Excel
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='الوثائق', index=False, startrow=2)
            
            workbook = writer.book
            worksheet = writer.sheets['الوثائق']
            
            # تنسيق الرأس
            header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # تنسيق العنوان الرئيسي
            title_font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
            title_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
            title_alignment = Alignment(horizontal='center', vertical='center')
            
            # إضافة العنوان الرئيسي
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = f'تقرير الوثائق - تم التصدير في {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = title_alignment
            
            # تطبيق التنسيق على رأس الجدول
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
                # تعديل عرض العمود
                column_width = max(len(str(column)), 15)
                if column_width > 50:
                    column_width = 50
                worksheet.column_dimensions[cell.column_letter].width = column_width
            
            # تطبيق الحدود على الجدول
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # تطبيق الحدود على كامل البيانات
            for row in range(1, len(df) + 4):  # +4 للعنوان والرأس
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).border = thin_border
            
            # تطبيق التنسيق على البيانات
            data_alignment = Alignment(horizontal='center', vertical='center')
            for row in range(4, len(df) + 4):  # البيانات تبدأ من الصف الرابع
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = data_alignment
                    
                    # تلوين الصفوف بالتناوب
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # إعداد اسم الملف
        filter_parts = []
        if document_type:
            filter_parts.append(f'نوع_{document_type}')
        if status_filter:
            filter_parts.append(f'حالة_{status_filter}')
        if department_id:
            dept = Department.query.get(department_id)
            if dept:
                filter_parts.append(f'قسم_{dept.name}')
        
        filename_suffix = '_'.join(filter_parts) if filter_parts else 'جميع_الوثائق'
        filename = f'تقرير_الوثائق_{filename_suffix}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        # تسجيل العملية
        audit = SystemAudit(
            action='export_excel',
            entity_type='document',
            entity_id=0,
            details=f'تم تصدير {len(documents)} وثيقة إلى ملف Excel - الفلاتر: {", ".join(filter_parts) if filter_parts else "بدون فلاتر"}'
        )
        db.session.add(audit)
        db.session.commit()
        
        output.seek(0)
        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تصدير البيانات: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('documents.index'))

@documents_bp.route('/excel-dashboard', methods=['GET', 'POST'])
@login_required
def excel_dashboard():
    """صفحة داش بورد تفاعلي لبيانات Excel"""
    
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)
        
        file = request.files['excel_file']
        
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(request.url)
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('يرجى رفع ملف Excel فقط (.xlsx أو .xls)', 'danger')
            return redirect(request.url)
        
        try:
            # قراءة ملف Excel
            df = pd.read_excel(file)
            
            # تحليل البيانات
            stats = analyze_excel_data(df)
            
            # حفظ البيانات في session للعرض
            preview_data = df.head(50).to_dict('records')
            columns = df.columns.tolist()
            
            return render_template('documents/excel_dashboard.html',
                                 stats=stats,
                                 preview_data=preview_data,
                                 columns=columns,
                                 total_rows=len(df),
                                 uploaded=True)
        
        except Exception as e:
            flash(f'حدث خطأ أثناء قراءة الملف: {str(e)}', 'danger')
            return redirect(request.url)
    
    return render_template('documents/excel_dashboard.html', uploaded=False)


def analyze_excel_data(df):
    """تحليل بيانات Excel واستخراج الإحصائيات"""
    stats = {
        'total_rows': len(df),
        'total_columns': len(df.columns),
        'columns': df.columns.tolist()
    }
    
    # محاولة اكتشاف أعمدة الحالة تلقائياً
    status_columns = [col for col in df.columns if any(word in str(col).lower() for word in ['status', 'حالة', 'state'])]
    
    if status_columns:
        status_col = status_columns[0]
        status_counts = df[status_col].value_counts().to_dict()
        stats['status_data'] = status_counts
        stats['status_column'] = status_col
    else:
        stats['status_data'] = {}
        stats['status_column'] = None
    
    # محاولة اكتشاف أعمدة التاريخ
    date_columns = [col for col in df.columns if any(word in str(col).lower() for word in ['date', 'تاريخ', 'expiry', 'انتهاء'])]
    
    if date_columns:
        stats['date_columns'] = date_columns
        
        # حساب الإحصائيات الزمنية للعمود الأول
        try:
            df[date_columns[0]] = pd.to_datetime(df[date_columns[0]], errors='coerce')
            current_date = datetime.now()
            
            stats['expired_count'] = len(df[df[date_columns[0]] < current_date])
            stats['valid_count'] = len(df[df[date_columns[0]] >= current_date])
            stats['expiring_soon'] = len(df[(df[date_columns[0]] >= current_date) & (df[date_columns[0]] <= current_date + timedelta(days=60))])
        except:
            stats['expired_count'] = 0
            stats['valid_count'] = 0
            stats['expiring_soon'] = 0
    else:
        stats['date_columns'] = []
        stats['expired_count'] = 0
        stats['valid_count'] = 0
        stats['expiring_soon'] = 0
    
    # إحصائيات عددية عامة
    numeric_columns = df.select_dtypes(include=['int64', 'float64']).columns.tolist()
    if numeric_columns:
        stats['numeric_summary'] = {}
        for col in numeric_columns[:3]:  # أول 3 أعمدة عددية
            stats['numeric_summary'][col] = {
                'sum': float(df[col].sum()),
                'mean': float(df[col].mean()),
                'max': float(df[col].max()),
                'min': float(df[col].min())
            }
    
    # حساب الاتجاه الشهري (محاكاة)
    stats['monthly_trend'] = generate_monthly_trend(len(df))
    
    return stats


def generate_monthly_trend(total):
    """إنشاء بيانات اتجاه شهري محاكاة"""
    import random
    months = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو']
    base = max(10, total // 6)
    trend = []
    
    for i in range(6):
        value = base + random.randint(-5, 15) * (i + 1)
        trend.append(min(value, total))
    
    return {
        'labels': months,
        'data': trend
    }
