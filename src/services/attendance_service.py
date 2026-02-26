"""
Attendance Service Layer
========================
Business logic for attendance management - Pure Python, Zero Flask dependencies

Author: Auto-Refactored from _attendance_main.py (3,403 lines)
Date: 2026-02-20
Architecture: Service Layer Pattern

Key Principles:
- Pure Python business logic
- Zero Flask dependencies (no render_template, request, etc.)
- Fully testable with unit tests
- Reusable across multiple interfaces (Web, API, CLI)
"""

from datetime import datetime, date, time, timedelta
from sqlalchemy import func, extract, or_
from src.core.extensions import db
from models import (
    Attendance, Employee, Department, SystemAudit, 
    employee_departments, GeofenceSession, EmployeeLocation
)
from src.utils.date_converter import parse_date, format_date_hijri, format_date_gregorian
from src.utils.audit_logger import log_activity
from src.services.attendance_analytics import AttendanceAnalytics
from src.services.attendance_engine import AttendanceEngine
import logging
from typing import Dict, List, Tuple, Optional, Any
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class AttendanceService:
    """
    Centralized attendance business logic
    All methods are static for easy testing and reuse
    """
    
    # ==================== Helper Methods ====================
    
    @staticmethod
    def format_time_12h_ar(dt: Optional[datetime]) -> str:
        """تحويل الوقت من 24 ساعة إلى 12 ساعة بصيغة عربية (صباح/مساء)"""
        if not dt:
            return '-'
        hour = dt.hour
        minute = dt.minute
        second = dt.second
        
        period = 'ص' if hour < 12 else 'م'
        
        if hour > 12:
            hour = hour - 12
        elif hour == 0:
            hour = 12
        
        return f'{hour}:{minute:02d}:{second:02d} {period}'
    
    @staticmethod
    def format_time_12h_ar_short(dt: Optional[datetime]) -> str:
        """تحويل الوقت من 24 ساعة إلى 12 ساعة بصيغة قصيرة (بدون ثوانٍ)"""
        if not dt:
            return '-'
        hour = dt.hour
        minute = dt.minute
        
        period = 'ص' if hour < 12 else 'م'
        
        if hour > 12:
            hour = hour - 12
        elif hour == 0:
            hour = 12
        
        return f'{hour}:{minute:02d} {period}'
    
    @staticmethod
    def log_attendance_activity(action: str, attendance_data: Dict[str, Any]) -> None:
        """تسجيل نشاط الحضور في audit log"""
        try:
            details = f"Attendance {action}: {attendance_data}"
            SystemAudit.create_audit_record(
                user_id=attendance_data.get('user_id'),
                action=action,
                entity_type='attendance',
                entity_id=attendance_data.get('id'),
                entity_name=attendance_data.get('employee_name', 'Unknown'),
                details=details
            )
            db.session.commit()
        except Exception as e:
            logger.error(f"Failed to log attendance activity: {str(e)}")
    
    # ==================== Data Retrieval Methods ====================
    
    @staticmethod
    def get_unified_attendance_list(
        att_date: date,
        department_id: Optional[int] = None,
        status_filter: Optional[str] = None,
        user_accessible_departments: Optional[List[Department]] = None
    ) -> List[Dict[str, Any]]:
        """
        Get unified attendance list with filtering
        Uses AttendanceEngine for optimized queries
        """
        try:
            unified_attendances = AttendanceEngine.get_unified_attendance_list(
                att_date=att_date,
                department_id=department_id,
                status_filter=status_filter
            )
            
            # Apply user department filtering if provided
            if user_accessible_departments:
                accessible_dept_ids = [d.id for d in user_accessible_departments]
                unified_attendances = [
                    att for att in unified_attendances 
                    if att.get('department_id') in accessible_dept_ids
                ]
            
            return unified_attendances
        except Exception as e:
            logger.error(f"Error in get_unified_attendance_list: {str(e)}", exc_info=True)
            return []
    
    @staticmethod
    def calculate_stats_from_attendances(attendances: List[Dict[str, Any]]) -> Dict[str, int]:
        """Calculate statistics from attendance list"""
        return {
            'present': sum(1 for rec in attendances if rec['status'] == 'present'),
            'absent': sum(1 for rec in attendances if rec['status'] == 'absent'),
            'leave': sum(1 for rec in attendances if rec['status'] == 'leave'),
            'sick': sum(1 for rec in attendances if rec['status'] == 'sick'),
            'total': len(attendances)
        }
    
    @staticmethod
    def get_active_employees(
        department_id: Optional[int] = None,
        user_role: Optional[str] = None,
        user_assigned_department_id: Optional[int] = None
    ) -> List[Employee]:
        """
        Get active employees based on user permissions
        """
        try:
            # Admin/Manager/Supervisor see all employees
            if user_role in ['ADMIN', 'MANAGER', 'SUPERVISOR']:
                employees = Employee.query.filter(
                    ~Employee.status.in_(['terminated', 'inactive'])
                ).order_by(Employee.name).all()
            
            # Department-specific users see only their department
            elif user_assigned_department_id:
                dept = Department.query.get(user_assigned_department_id)
                if dept:
                    employees = [
                        emp for emp in dept.employees 
                        if emp.status not in ['terminated', 'inactive']
                    ]
                    employees.sort(key=lambda x: x.name)
                else:
                    employees = []
            
            # Department filter
            elif department_id:
                dept = Department.query.get(department_id)
                if dept:
                    employees = [
                        emp for emp in dept.employees 
                        if emp.status not in ['terminated', 'inactive']
                    ]
                    employees.sort(key=lambda x: x.name)
                else:
                    employees = []
            
            # Fallback: all employees
            else:
                employees = Employee.query.filter(
                    ~Employee.status.in_(['terminated', 'inactive'])
                ).order_by(Employee.name).all()
            
            return employees
        
        except Exception as e:
            logger.error(f"Error getting active employees: {str(e)}", exc_info=True)
            return []
    
    # ==================== CRUD Operations ====================
    
    @staticmethod
    def record_attendance(
        employee_id: int,
        att_date: date,
        status: str,
        check_in: Optional[time] = None,
        check_out: Optional[time] = None,
        notes: Optional[str] = ''
    ) -> Tuple[Optional[Attendance], bool, str]:
        """
        Record attendance for a single employee
        Returns: (attendance_obj, is_new, message)
        """
        try:
            return AttendanceEngine.record_attendance(
                employee_id=employee_id,
                att_date=att_date,
                status=status,
                check_in=check_in,
                check_out=check_out,
                notes=notes
            )
        except Exception as e:
            logger.error(f"Error recording attendance: {str(e)}", exc_info=True)
            return None, False, f"خطأ: {str(e)}"
    
    @staticmethod
    def bulk_record_department(
        department_id: int,
        att_date: date,
        status: str
    ) -> Tuple[int, str]:
        """
        Record attendance for entire department
        Returns: (count, message)
        """
        try:
            return AttendanceEngine.bulk_record_department(
                department_id=department_id,
                att_date=att_date,
                status=status
            )
        except Exception as e:
            logger.error(f"Error bulk recording department: {str(e)}", exc_info=True)
            return 0, f"خطأ: {str(e)}"
    
    @staticmethod
    def bulk_record_for_period(
        department_id: int,
        start_date: date,
        end_date: date,
        status: str,
        skip_weekends: bool = False,
        overwrite_existing: bool = False
    ) -> Dict[str, int]:
        """
        Record attendance for department across date range
        Returns: {'created': count, 'updated': count, 'skipped': count}
        """
        try:
            # Validate date range
            if start_date > end_date:
                return {'created': 0, 'updated': 0, 'skipped': 0, 'error': 'Invalid date range'}
            
            if (end_date - start_date).days > 90:
                return {'created': 0, 'updated': 0, 'skipped': 0, 'error': 'Period > 90 days'}
            
            # Get department employees
            department = Department.query.get(department_id)
            if not department:
                return {'created': 0, 'updated': 0, 'skipped': 0, 'error': 'Department not found'}
            
            employees = [
                emp for emp in department.employees 
                if emp.status not in ['terminated', 'inactive']
            ]
            
            if not employees:
                return {'created': 0, 'updated': 0, 'skipped': 0, 'error': 'No active employees'}
            
            # Build date list
            date_list = []
            current_date = start_date
            while current_date <= end_date:
                if skip_weekends and current_date.weekday() in [4, 5]:  # Friday, Saturday
                    current_date += timedelta(days=1)
                    continue
                date_list.append(current_date)
                current_date += timedelta(days=1)
            
            # Process attendance records
            created_count = 0
            updated_count = 0
            skipped_count = 0
            
            for employee in employees:
                for attendance_date in date_list:
                    existing = Attendance.query.filter_by(
                        employee_id=employee.id,
                        date=attendance_date
                    ).first()
                    
                    if existing:
                        if overwrite_existing:
                            existing.status = status
                            if status != 'present':
                                existing.check_in = None
                                existing.check_out = None
                            updated_count += 1
                        else:
                            skipped_count += 1
                    else:
                        new_attendance = Attendance(
                            employee_id=employee.id,
                            date=attendance_date,
                            status=status
                        )
                        db.session.add(new_attendance)
                        created_count += 1
            
            db.session.commit()
            
            return {
                'created': created_count,
                'updated': updated_count,
                'skipped': skipped_count
            }
        
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error bulk recording for period: {str(e)}", exc_info=True)
            return {'created': 0, 'updated': 0, 'skipped': 0, 'error': str(e)}
    
    @staticmethod
    def delete_attendance(attendance_id: int) -> Tuple[bool, str]:
        """
        Delete attendance record
        Returns: (success, message)
        """
        try:
            attendance = Attendance.query.get(attendance_id)
            if not attendance:
                return False, 'السجل غير موجود'
            
            employee = Employee.query.get(attendance.employee_id)
            entity_name = employee.name if employee else f'ID: {attendance_id}'
            
            db.session.delete(attendance)
            
            SystemAudit.create_audit_record(
                user_id=None,
                action='delete',
                entity_type='attendance',
                entity_id=attendance_id,
                entity_name=entity_name,
                details=f'تم حذف سجل حضور للموظف: {employee.name if employee else "غير معروف"} بتاريخ {attendance.date}'
            )
            
            db.session.commit()
            return True, 'تم حذف سجل الحضور بنجاح'
        
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error deleting attendance: {str(e)}", exc_info=True)
            return False, f'حدث خطأ: {str(e)}'
    
    @staticmethod
    def bulk_delete_attendance(attendance_ids: List[int]) -> Dict[str, Any]:
        """
        Delete multiple attendance records
        Returns: {'deleted_count': int, 'errors': List[str]}
        """
        try:
            if not attendance_ids:
                return {'deleted_count': 0, 'errors': ['لا توجد سجلات محددة للحذف']}
            
            deleted_count = 0
            errors = []
            
            for attendance_id in attendance_ids:
                try:
                    attendance = Attendance.query.get(attendance_id)
                    if attendance:
                        employee = Employee.query.get(attendance.employee_id)
                        entity_name = employee.name if employee else f'ID: {attendance_id}'
                        
                        db.session.delete(attendance)
                        
                        SystemAudit.create_audit_record(
                            user_id=None,
                            action='delete',
                            entity_type='attendance',
                            entity_id=attendance_id,
                            entity_name=entity_name,
                            details=f'حذف جماعي - تم حذف سجل حضور للموظف: {employee.name if employee else "غير معروف"} بتاريخ {attendance.date}'
                        )
                        
                        deleted_count += 1
                    else:
                        errors.append(f'السجل {attendance_id} غير موجود')
                
                except Exception as e:
                    errors.append(f'خطأ في حذف السجل {attendance_id}: {str(e)}')
            
            db.session.commit()
            
            return {'deleted_count': deleted_count, 'errors': errors}
        
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error bulk deleting: {str(e)}", exc_info=True)
            return {'deleted_count': 0, 'errors': [f'حدث خطأ: {str(e)}']}
    
    # ==================== Statistics & Analytics ====================
    
    @staticmethod
    def get_stats_for_period(
        start_date: date,
        end_date: date,
        department_id: Optional[int] = None
    ) -> Dict[str, int]:
        """Get attendance statistics for date range"""
        try:
            query = db.session.query(
                Attendance.status,
                func.count(Attendance.id).label('count')
            ).filter(
                Attendance.date >= start_date,
                Attendance.date <= end_date
            )
            
            if department_id:
                query = query.join(Employee).filter(Employee.department_id == department_id)
            
            stats = query.group_by(Attendance.status).all()
            
            result = {'present': 0, 'absent': 0, 'leave': 0, 'sick': 0}
            for status, count in stats:
                result[status] = count
            
            return result
        
        except Exception as e:
            logger.error(f"Error getting stats: {str(e)}", exc_info=True)
            return {'present': 0, 'absent': 0, 'leave': 0, 'sick': 0}
    
    @staticmethod
    def get_dashboard_data(
        current_date: Optional[date] = None,
        project_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Get comprehensive dashboard data
        Includes daily, weekly, monthly stats with charts
        """
        try:
            today = current_date or datetime.now().date()
            
            # Date ranges
            start_of_week = today - timedelta(days=today.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            start_of_month = today.replace(day=1)
            
            # Get last day of month
            if today.month == 12:
                end_of_month = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_of_month = date(today.year, today.month + 1, 1) - timedelta(days=1)
            
            # Daily stats
            daily_stats = AttendanceService.get_stats_for_period(today, today)
            
            # Weekly stats
            weekly_stats = AttendanceService.get_stats_for_period(start_of_week, end_of_week)
            
            # Monthly stats
            monthly_stats = AttendanceService.get_stats_for_period(start_of_month, end_of_month)
            
            # Daily attendance data (for chart)
            daily_attendance_data = []
            current_check_date = start_of_month
            while current_check_date <= today:
                day_stats = AttendanceService.get_stats_for_period(current_check_date, current_check_date)
                daily_attendance_data.append({
                    'day': current_check_date.strftime('%d'),
                    'present': day_stats['present'],
                    'absent': day_stats['absent']
                })
                current_check_date += timedelta(days=1)
            
            # Calculate attendance rates
            total_daily = sum(daily_stats.values())
            daily_attendance_rate = round((daily_stats['present'] / max(total_daily, 1)) * 100)
            
            total_weekly = sum(weekly_stats.values())
            weekly_attendance_rate = round((weekly_stats['present'] / max(total_weekly, 1)) * 100)
            
            total_monthly = sum(monthly_stats.values())
            monthly_attendance_rate = round((monthly_stats['present'] / max(total_monthly, 1)) * 100)
            
            # Get department summaries
            daily_summary = AttendanceAnalytics.get_department_summary(
                start_date=today,
                end_date=today,
                project_name=project_name
            )
            
            monthly_summary = AttendanceAnalytics.get_department_summary(
                start_date=start_of_month,
                end_date=end_of_month,
                project_name=project_name
            )
            
            # Active employees count
            active_employees_count = Employee.query.filter(
                ~Employee.status.in_(['terminated', 'inactive'])
            ).count()
            
            return {
                'today': today,
                'start_of_week': start_of_week,
                'end_of_week': end_of_week,
                'start_of_month': start_of_month,
                'end_of_month': end_of_month,
                'daily_stats': daily_stats,
                'weekly_stats': weekly_stats,
                'monthly_stats': monthly_stats,
                'daily_attendance_rate': daily_attendance_rate,
                'weekly_attendance_rate': weekly_attendance_rate,
                'monthly_attendance_rate': monthly_attendance_rate,
                'daily_attendance_data': daily_attendance_data,
                'daily_summary': daily_summary,
                'monthly_summary': monthly_summary,
                'active_employees_count': active_employees_count
            }
        
        except Exception as e:
            logger.error(f"Error getting dashboard data: {str(e)}", exc_info=True)
            return {}
    
    # ==================== Export Methods ====================
    
    @staticmethod
    def export_to_excel(
        start_date: date,
        end_date: date,
        department_id: Optional[int] = None,
        status_filter: Optional[str] = None,
        search_query: Optional[str] = None
    ) -> BytesIO:
        """
        Export attendance data to Excel with P/A/S format
        Returns BytesIO object ready for sending
        """
        try:
            # Build query
            query = Attendance.query.join(Employee).filter(
                Attendance.date >= start_date,
                Attendance.date <= end_date
            )
            
            if department_id:
                query = query.join(employee_departments).filter(
                    employee_departments.c.department_id == int(department_id)
                )
            
            if search_query:
                query = query.filter(
                    or_(
                        Employee.name.ilike(f'%{search_query}%'),
                        Employee.employee_id.ilike(f'%{search_query}%'),
                        Employee.national_id.ilike(f'%{search_query}%')
                    )
                )
            
            if status_filter:
                query = query.filter(Attendance.status == status_filter)
            
            attendances = query.order_by(Attendance.date.desc(), Employee.name).all()
            
            # Calculate stats
            total_count = len(attendances)
            present_count = sum(1 for a in attendances if a.status == 'present')
            absent_count = sum(1 for a in attendances if a.status == 'absent')
            leave_count = sum(1 for a in attendances if a.status == 'leave')
            sick_count = sum(1 for a in attendances if a.status == 'sick')
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "بيانات الحضور"
            
            # Get unique dates and employees
            all_dates = sorted(set(att.date for att in attendances if att.date))
            employees_dict = {}
            for att in attendances:
                if att.employee.id not in employees_dict:
                    employees_dict[att.employee.id] = att.employee
            
            sorted_employees = sorted(employees_dict.values(), key=lambda e: e.name)
            
            # Build attendance map: (employee_id, date) -> status
            attendance_map = {}
            for att in attendances:
                key = (att.employee.id, att.date)
                attendance_map[key] = att.status
            
            # Set column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 16
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 22
            
            for col_idx in range(len(all_dates)):
                col_letter = get_column_letter(col_idx + 5)
                ws.column_dimensions[col_letter].width = 4
            
            # Create header
            header_row = ['الموظف', 'الرقم الوظيفي', 'رقم الهوية', 'القسم'] + [
                d.strftime('%b %d') for d in all_dates
            ]
            ws.append(header_row)
            
            # Format header
            header_fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            thin_border = Border(
                left=Side(style='thin', color='FFFFFF'),
                right=Side(style='thin', color='FFFFFF'),
                top=Side(style='thin', color='FFFFFF'),
                bottom=Side(style='thin', color='FFFFFF')
            )
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            ws.row_dimensions[1].height = 28
            
            # Add employee data
            row_num = 2
            for emp in sorted_employees:
                department_name = ', '.join([d.name for d in emp.departments]) if emp.departments else '-'
                
                row_data = [
                    emp.name,
                    emp.employee_id or '-',
                    emp.national_id if hasattr(emp, 'national_id') and emp.national_id else '-',
                    department_name
                ]
                
                for date in all_dates:
                    status = attendance_map.get((emp.id, date), '')
                    
                    if status == 'present':
                        row_data.append('P')
                    elif status == 'absent':
                        row_data.append('A')
                    elif status in ['leave', 'sick']:
                        row_data.append('S')
                    else:
                        row_data.append('')
                
                ws.append(row_data)
                
                # Format rows
                for col_idx in range(1, len(row_data) + 1):
                    cell_obj = ws.cell(row=row_num, column=col_idx)
                    cell_obj.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell_obj.border = Border(
                        left=Side(style='thin', color='E0E0E0'),
                        right=Side(style='thin', color='E0E0E0'),
                        top=Side(style='thin', color='E0E0E0'),
                        bottom=Side(style='thin', color='E0E0E0')
                    )
                    
                    if col_idx <= 4:
                        cell_obj.font = Font(bold=True if col_idx == 1 else False)
                        if row_num % 2 == 0:
                            cell_obj.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                    else:
                        value = cell_obj.value
                        if value == 'P':
                            cell_obj.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                            cell_obj.font = Font(bold=True, color="155724", size=11)
                        elif value == 'A':
                            cell_obj.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                            cell_obj.font = Font(bold=True, color="721C24", size=11)
                        elif value == 'S':
                            cell_obj.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                            cell_obj.font = Font(bold=True, color="856404", size=11)
                        else:
                            if row_num % 2 == 0:
                                cell_obj.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                
                ws.row_dimensions[row_num].height = 20
                row_num += 1
            
            # Add statistics sheet
            stats_ws = wb.create_sheet("الإحصائيات")
            
            stats_ws.merge_cells('A1:D1')
            title_cell = stats_ws['A1']
            title_cell.value = "ملخص إحصائيات الحضور"
            title_cell.font = Font(bold=True, size=16, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            stats_ws.row_dimensions[1].height = 32
            
            stats_ws.column_dimensions['A'].width = 25
            stats_ws.column_dimensions['B'].width = 18
            stats_ws.column_dimensions['C'].width = 18
            stats_ws.column_dimensions['D'].width = 18
            
            stats_data = [
                ['', 'العدد', 'النسبة %', 'الحالة'],
                ['إجمالي السجلات', total_count, '100%', 'total'],
                ['موظفون حاضرون', present_count, f'{int((present_count/max(total_count,1))*100)}%', 'present'],
                ['موظفون غائبون', absent_count, f'{int((absent_count/max(total_count,1))*100)}%', 'absent'],
                ['إجازات', leave_count, f'{int((leave_count/max(total_count,1))*100)}%', 'leave'],
                ['مرضي', sick_count, f'{int((sick_count/max(total_count,1))*100)}%', 'sick']
            ]
            
            for idx, row_data in enumerate(stats_data, 2):
                for col_idx, value in enumerate(row_data[:3], 1):
                    cell = stats_ws.cell(row=idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if idx == 2:
                        cell.font = Font(bold=True, color="FFFFFF", size=12)
                        cell.fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
                    elif row_data[3] == 'total':
                        cell.font = Font(bold=True, size=12)
                        cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                    elif row_data[3] == 'present':
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        cell.font = Font(bold=True, color="155724")
                    elif row_data[3] == 'absent':
                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                        cell.font = Font(bold=True, color="721C24")
                    elif row_data[3] == 'leave':
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                        cell.font = Font(bold=True, color="856404")
                    elif row_data[3] == 'sick':
                        cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
                        cell.font = Font(bold=True, color="0C5460")
                    
                    stats_ws.row_dimensions[idx].height = 26
            
            # Add pie chart
            if total_count > 0:
                pie = PieChart()
                pie.title = "توزيع حالات الحضور"
                pie.style = 10
                labels = Reference(stats_ws, min_col=1, min_row=3, max_row=6)
                data = Reference(stats_ws, min_col=2, min_row=2, max_row=6)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.height = 12
                pie.width = 16
                stats_ws.add_chart(pie, "A9")
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
        
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}", exc_info=True)
            raise
    
    @staticmethod
    def get_employees_by_department(department_id: int) -> List[Dict[str, Any]]:
        """Get all employees in a department (for API)"""
        try:
            department = Department.query.get(department_id)
            if not department:
                return []
            
            employees = [
                emp for emp in department.employees 
                if emp.status not in ['terminated', 'inactive']
            ]
            
            return [
                {
                    'id': emp.id,
                    'name': emp.name,
                    'employee_id': emp.employee_id,
                    'national_id': emp.national_id,
                    'mobile': emp.mobile,
                    'job_title': emp.job_title
                }
                for emp in employees
            ]
        
        except Exception as e:
            logger.error(f"Error getting employees by department: {str(e)}", exc_info=True)
            return []
    
    # ==================== Geo Circle Methods ====================
    
    @staticmethod
    def get_circle_accessed_employees(
        department_id: int,
        circle_name: str,
        selected_date: date
    ) -> List[Dict[str, Any]]:
        """Get employees who accessed a specific geo circle"""
        try:
            saudi_tz = timedelta(hours=3)
            
            dept = Department.query.get(department_id)
            if not dept:
                return []
            
            # Get employees in circle
            active_employees = [
                emp for emp in dept.employees 
                if emp.status == 'active' and emp.location and emp.location.strip() == circle_name
            ]
            
            start_date = selected_date
            end_date = datetime.now().date()
            
            employees_accessed = []
            
            for emp in active_employees:
                # Get geofence sessions
                geo_sessions = db.session.query(GeofenceSession).filter(
                    GeofenceSession.employee_id == emp.id,
                    GeofenceSession.entry_time >= datetime.combine(start_date, time(0, 0, 0)),
                    GeofenceSession.entry_time <= datetime.combine(end_date, time(23, 59, 59))
                ).order_by(GeofenceSession.entry_time.asc()).all()
                
                if geo_sessions:
                    latest_geo = geo_sessions[-1]
                    
                    emp_location = db.session.query(EmployeeLocation).filter(
                        EmployeeLocation.employee_id == emp.id,
                        EmployeeLocation.recorded_at >= datetime.combine(start_date, time(0, 0, 0)),
                        EmployeeLocation.recorded_at <= datetime.combine(end_date, time(23, 59, 59))
                    ).order_by(EmployeeLocation.recorded_at.desc()).first()
                    
                    duration_minutes = latest_geo.duration_minutes if latest_geo.duration_minutes else 0
                    entry_time_sa = (latest_geo.entry_time + saudi_tz) if latest_geo.entry_time else None
                    exit_time_sa = (latest_geo.exit_time + saudi_tz) if latest_geo.exit_time else None
                    
                    employees_accessed.append({
                        'employee_id': emp.id,
                        'employee_name': emp.name,
                        'employee_number': emp.employee_id,
                        'national_id': emp.national_id,
                        'mobile': emp.mobile,
                        'job_title': emp.job_title,
                        'circle_entry_time': AttendanceService.format_time_12h_ar(entry_time_sa) if entry_time_sa else None,
                        'circle_exit_time': AttendanceService.format_time_12h_ar(exit_time_sa) if exit_time_sa else None,
                        'duration_minutes': duration_minutes,
                        'duration_display': f'{duration_minutes // 60}س {duration_minutes % 60}د' if duration_minutes > 0 else '-',
                        'gps_latitude': emp_location.latitude if emp_location else None,
                        'gps_longitude': emp_location.longitude if emp_location else None,
                        'profile_image': emp.profile_image,
                    })
            
            return employees_accessed
        
        except Exception as e:
            logger.error(f"Error getting circle accessed employees: {str(e)}", exc_info=True)
            return []
    
    @staticmethod
    def mark_circle_employees_attendance(
        department_id: int,
        circle_name: str,
        selected_date: date
    ) -> Tuple[int, str]:
        """Mark employees in circle as present with check times"""
        try:
            saudi_tz = timedelta(hours=3)
            
            dept = Department.query.get(department_id)
            if not dept:
                return 0, 'القسم غير موجود'
            
            active_employees = [
                emp for emp in dept.employees 
                if emp.status == 'active' and emp.location and emp.location.strip() == circle_name
            ]
            
            start_date = selected_date
            end_date = datetime.now().date()
            
            marked_count = 0
            
            for emp in active_employees:
                geo_sessions = db.session.query(GeofenceSession).filter(
                    GeofenceSession.employee_id == emp.id,
                    GeofenceSession.entry_time >= datetime.combine(start_date, time(0, 0, 0)),
                    GeofenceSession.entry_time <= datetime.combine(end_date, time(23, 59, 59))
                ).order_by(GeofenceSession.entry_time.asc()).all()
                
                if geo_sessions:
                    # Extract morning entry and evening exit
                    morning_check_in = None
                    evening_check_out = None
                    
                    for geo in geo_sessions:
                        if geo.entry_time and not morning_check_in:
                            entry_sa = geo.entry_time + saudi_tz
                            if entry_sa.hour < 12:
                                morning_check_in = entry_sa
                        
                        if geo.exit_time:
                            exit_sa = geo.exit_time + saudi_tz
                            if exit_sa.hour >= 12:
                                evening_check_out = exit_sa
                    
                    if not morning_check_in and geo_sessions:
                        morning_check_in = geo_sessions[0].entry_time + saudi_tz
                    
                    if not evening_check_out and geo_sessions:
                        evening_check_out = geo_sessions[-1].exit_time + saudi_tz if geo_sessions[-1].exit_time else None
                    
                    check_in_time = morning_check_in.time() if isinstance(morning_check_in, datetime) else morning_check_in
                    check_out_time = evening_check_out.time() if isinstance(evening_check_out, datetime) else evening_check_out
                    
                    # Update or create attendance record
                    existing_attendance = Attendance.query.filter(
                        Attendance.employee_id == emp.id,
                        Attendance.date == selected_date
                    ).first()
                    
                    if existing_attendance:
                        if check_in_time:
                            existing_attendance.check_in = check_in_time
                        if check_out_time:
                            existing_attendance.check_out = check_out_time
                        existing_attendance.status = 'present'
                        existing_attendance.updated_at = datetime.utcnow()
                    else:
                        attendance = Attendance(
                            employee_id=emp.id,
                            date=selected_date,
                            status='present',
                            check_in=check_in_time,
                            check_out=check_out_time,
                        )
                        db.session.add(attendance)
                    
                    marked_count += 1
            
            db.session.commit()
            
            return marked_count, f'تم تسجيل {marked_count} موظف كحاضرين'
        
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error marking circle employees: {str(e)}", exc_info=True)
            return 0, f'حدث خطأ: {str(e)}'
