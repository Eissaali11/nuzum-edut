"""
خدمة تقارير الحضور - Attendance Reports Service
يتعامل مع جميع عمليات تصدير Excel والتقارير المعقدة
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file
from models import Attendance, Employee, Department, employee_departments
from src.core.extensions import db
from src.services.attendance_analytics import AttendanceAnalytics
from sqlalchemy import func, or_


class AttendanceReportService:
    """خدمة تصدير تقارير الحضور إلى Excel"""
    
    @staticmethod
    def export_dashboard_summary(selected_department=None, selected_project=None):
        """تصدير لوحة المعلومات إلى Excel مع تصميم داش بورد خيالي ومبهر"""
        try:
            today = datetime.now().date()
            start_date = today.replace(day=1)
            end_date = today
            
            # استخدام خدمة analytics لجلب البيانات التفصيلية
            summary = AttendanceAnalytics.get_department_summary(
                start_date=start_date,
                end_date=end_date,
                project_name=selected_project
            )
            
            # تحويل البيانات للصيغة المطلوبة
            department_data = []
            for dept in summary['departments']:
                if selected_department and dept['name'] != selected_department:
                    continue
                
                department_data.append({
                    'name': dept['name'],
                    'employees': dept['total_employees'],
                    'present': dept['present'],
                    'absent': dept['absent'],
                    'leave': dept['leave'],
                    'sick': dept['sick'],
                    'total': dept['total_records'],
                    'rate': dept['attendance_rate'],
                    'absentees': dept.get('absentees', []),
                    'on_leave': dept.get('on_leave', []),
                    'sick_employees': dept.get('sick_employees', [])
                })
            
            # الإجماليات
            total_employees = summary['total_employees']
            total_present = summary['total_present']
            total_absent = summary['total_absent']
            total_leave = summary['total_leave']
            total_sick = summary['total_sick']
            total_records = summary['total_records']
            
            if not department_data:
                wb = Workbook()
                ws = wb.active
                ws.title = "لا توجد بيانات"
                ws['A1'] = "لا توجد بيانات للعرض"
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                return {
                    'buffer': output,
                    'filename': f'تقرير_الحضور_{today.strftime("%Y%m%d")}.xlsx',
                    'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                }
            
            # إنشاء ملف Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "📊 لوحة المعلومات"
            
            # === العنوان الرئيسي ===
            ws.merge_cells('A1:M3')
            title_cell = ws['A1']
            title_cell.value = f"📊 لوحة معلومات الحضور والغياب\n{start_date.strftime('%Y/%m/%d')} - {end_date.strftime('%Y/%m/%d')}"
            title_cell.font = Font(size=24, bold=True, color="FFFFFF", name="Arial")
            title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 30
            
            # === بطاقات KPI ===
            kpi_row = 5
            ws.row_dimensions[kpi_row].height = 35
            ws.row_dimensions[kpi_row + 1].height = 30
            
            thick_border = Border(
                left=Side(style='thick', color='FFFFFF'),
                right=Side(style='thick', color='FFFFFF'),
                top=Side(style='thick', color='FFFFFF'),
                bottom=Side(style='thick', color='FFFFFF')
            )
            
            kpis = [
                (f'A{kpi_row}', f'C{kpi_row+1}', f"👥 إجمالي الموظفين\n{total_employees}", "5B9BD5"),
                (f'D{kpi_row}', f'F{kpi_row+1}', f"✅ الحضور\n{total_present}", "70AD47"),
                (f'G{kpi_row}', f'I{kpi_row+1}', f"❌ الغياب\n{total_absent}", "E74C3C"),
                (f'J{kpi_row}', f'L{kpi_row+1}', f"🏖️ الإجازات\n{total_leave}", "F39C12"),
                (f'M{kpi_row}', f'O{kpi_row+1}', f"🏥 المرضي\n{total_sick}", "3498DB"),
            ]
            
            for start_cell, end_cell, value, color in kpis:
                ws.merge_cells(f'{start_cell}:{end_cell}')
                kpi_cell = ws[start_cell]
                kpi_cell.value = value
                kpi_cell.font = Font(size=16, bold=True, color="FFFFFF")
                kpi_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                kpi_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                kpi_cell.border = thick_border
            
            # === جدول البيانات ===
            table_start_row = kpi_row + 3
            
            ws.merge_cells(f'A{table_start_row}:H{table_start_row}')
            table_title = ws[f'A{table_start_row}']
            table_title.value = "📋 تفاصيل الأقسام"
            table_title.font = Font(size=14, bold=True, color="FFFFFF")
            table_title.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            table_title.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[table_start_row].height = 25
            
            # رؤوس الأعمدة
            headers_row = table_start_row + 1
            headers = ['القسم', 'عدد الموظفين', 'حاضر', 'غائب', 'إجازة', 'مرضي', 'إجمالي السجلات', 'معدل الحضور %']
            header_colors = ['34495E', '5B9BD5', '70AD47', 'E74C3C', 'F39C12', '3498DB', '95A5A6', '16A085']
            
            for col_idx, (header, color) in enumerate(zip(headers, header_colors), 1):
                cell = ws.cell(row=headers_row, column=col_idx)
                cell.value = header
                cell.font = Font(size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            ws.row_dimensions[headers_row].height = 30
            
            # البيانات
            data_start_row = headers_row + 1
            for row_idx, dept in enumerate(department_data, data_start_row):
                values = [dept['name'], dept['employees'], dept['present'], dept['absent'], 
                         dept['leave'], dept['sick'], dept['total'], dept['rate']]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                    
                    if col_idx == 3:
                        cell.fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                        cell.font = Font(bold=True, color="27AE60")
                    elif col_idx == 4:
                        cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                        cell.font = Font(bold=True, color="C0392B")
                    elif col_idx == 5:
                        cell.fill = PatternFill(start_color="FEF5E7", end_color="FEF5E7", fill_type="solid")
                        cell.font = Font(bold=True, color="D68910")
                    elif col_idx == 6:
                        cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                        cell.font = Font(bold=True, color="2874A6")
                    elif col_idx == 8:
                        cell.font = Font(bold=True, size=11)
                        if value >= 90:
                            cell.fill = PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type="solid")
                        elif value >= 75:
                            cell.fill = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
            
            # === مخطط عمودي ===
            chart_row = data_start_row + len(department_data) + 2
            
            bar_chart = BarChart()
            bar_chart.title = "مقارنة حالات الحضور بين الأقسام"
            bar_chart.style = 13
            bar_chart.height = 12
            bar_chart.width = 20
            
            data_ref = Reference(ws, min_col=3, min_row=headers_row, max_row=data_start_row + len(department_data) - 1, max_col=6)
            cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_start_row + len(department_data) - 1)
            
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats_ref)
            ws.add_chart(bar_chart, f"A{chart_row}")
            
            # === مخطط دائري ===
            pie_chart = PieChart()
            pie_chart.title = "نسب حالات الحضور الإجمالية"
            pie_chart.height = 12
            pie_chart.width = 15
            
            summary_row = data_start_row + len(department_data) + 15
            ws[f'K{summary_row}'] = 'حاضر'
            ws[f'L{summary_row}'] = total_present
            ws[f'K{summary_row+1}'] = 'غائب'
            ws[f'L{summary_row+1}'] = total_absent
            ws[f'K{summary_row+2}'] = 'إجازة'
            ws[f'L{summary_row+2}'] = total_leave
            ws[f'K{summary_row+3}'] = 'مرضي'
            ws[f'L{summary_row+3}'] = total_sick
            
            pie_data = Reference(ws, min_col=12, min_row=summary_row, max_row=summary_row+3)
            pie_labels = Reference(ws, min_col=11, min_row=summary_row, max_row=summary_row+3)
            pie_chart.add_data(pie_data, titles_from_data=False)
            pie_chart.set_categories(pie_labels)
            ws.add_chart(pie_chart, f"K{chart_row}")
            
            # تعيين عرض الأعمدة
            column_widths = [20, 15, 12, 12, 12, 12, 18, 18]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # حفظ الملف
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'📊_لوحة_معلومات_الحضور_{today.strftime("%Y%m%d")}.xlsx'
            if selected_department:
                filename = f'📊_{selected_department}_{today.strftime("%Y%m%d")}.xlsx'
            
            # إرجاع البيانات بدل استدعاء send_file (سيتم التعامل معها في الـ route)
            return {
                'buffer': output,
                'filename': filename,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
        except Exception as e:
            raise Exception(f"خطأ في تصدير لوحة المعلومات: {str(e)}")
    
    
    @staticmethod
    def export_department_details(department_name, selected_project=None):
        """تصدير تفاصيل القسم إلى Excel مع لوحة معلومات تفصيلية مميزة"""
        try:
            # جلب القسم
            department = Department.query.filter_by(name=department_name).first()
            if not department:
                raise ValueError('القسم غير موجود')
            
            today = datetime.now().date()
            start_date = today.replace(day=1)
            end_date = today
            
            # إنشاء قائمة بجميع أيام الشهر حتى اليوم
            date_range = []
            current = start_date
            while current <= end_date:
                date_range.append(current)
                current += timedelta(days=1)
            
            # جلب الموظفين والبيانات - استخدام علاقة many-to-many
            all_employees = [emp for emp in department.employees if emp.status not in ['terminated', 'inactive']]
            
            if selected_project and selected_project != 'None' and selected_project.strip():
                employees = [emp for emp in all_employees if emp.project == selected_project]
            else:
                employees = all_employees
            
            # إنشاء ملف Excel
            wb = Workbook()
            
            # صفحة لوحة المعلومات الرئيسية
            ws_dashboard = wb.active
            ws_dashboard.title = "لوحة المعلومات"
            
            # العنوان الرئيسي
            title = f"لوحة معلومات قسم {department.name}"
            if selected_project and selected_project != 'None':
                title += f" - مشروع {selected_project}"
            
            ws_dashboard['A1'] = title
            ws_dashboard['A1'].font = Font(size=18, bold=True, color="FFFFFF")
            ws_dashboard['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_dashboard.merge_cells('A1:H3')
            
            # معلومات الفترة
            period_info = f"الفترة: من {start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')}"
            ws_dashboard['A4'] = period_info
            ws_dashboard['A4'].font = Font(size=12, bold=True)
            ws_dashboard['A4'].alignment = Alignment(horizontal='center')
            ws_dashboard.merge_cells('A4:H4')
            
            # جمع بيانات الموظفين والحضور
            employee_data = []
            total_stats = {
                'total_employees': len(employees),
                'present': 0,
                'absent': 0,
                'leave': 0,
                'sick': 0,
                'total_records': 0
            }
            
            for employee in employees:
                attendance_records = Attendance.query.filter(
                    Attendance.employee_id == employee.id,
                    Attendance.date >= start_date,
                    Attendance.date <= end_date
                ).all()
                
                stats = {
                    'present': sum(1 for r in attendance_records if r.status == 'present'),
                    'absent': sum(1 for r in attendance_records if r.status == 'absent'),
                    'leave': sum(1 for r in attendance_records if r.status == 'leave'),
                    'sick': sum(1 for r in attendance_records if r.status == 'sick')
                }
                
                total_records = sum(stats.values())
                attendance_rate = (stats['present'] / total_records * 100) if total_records > 0 else 0
                
                employee_data.append({
                    'الاسم': employee.name,
                    'الهوية': employee.employee_id or 'غير محدد',
                    'حاضر': stats['present'],
                    'غائب': stats['absent'],
                    'إجازة': stats['leave'],
                    'مرضي': stats['sick'],
                    'الإجمالي': total_records,
                    'المعدل %': round(attendance_rate, 1)
                })
                
                # إضافة للإحصائيات الإجمالية
                total_stats['present'] += stats['present']
                total_stats['absent'] += stats['absent']
                total_stats['leave'] += stats['leave']
                total_stats['sick'] += stats['sick']
                total_stats['total_records'] += total_records
            
            # حساب معدل الحضور الإجمالي
            overall_rate = (total_stats['present'] / total_stats['total_records'] * 100) if total_stats['total_records'] > 0 else 0
            
            # الإحصائيات الإجمالية في لوحة المعلومات
            stats_row = 6
            
            # عناوين الإحصائيات
            stats_headers = ['إجمالي الموظفين', 'إجمالي الحضور', 'إجمالي الغياب', 'إجمالي الإجازات', 'إجمالي المرضي', 'معدل الحضور %', 'أيام العمل']
            stats_values = [total_stats['total_employees'], total_stats['present'], total_stats['absent'], 
                           total_stats['leave'], total_stats['sick'], round(overall_rate, 1), len(date_range)]
            
            for col, (header, value) in enumerate(zip(stats_headers, stats_values), 1):
                # العنوان
                header_cell = ws_dashboard.cell(row=stats_row, column=col, value=header)
                header_cell.font = Font(bold=True, color="FFFFFF")
                header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # القيمة
                value_cell = ws_dashboard.cell(row=stats_row + 1, column=col, value=value)
                value_cell.font = Font(bold=True, size=14)
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # تلوين حسب النوع
                if 'حضور' in header:
                    value_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif 'غياب' in header:
                    value_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif 'إجازة' in header:
                    value_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif 'مرضي' in header:
                    value_cell.fill = PatternFill(start_color="9CC3F7", end_color="9CC3F7", fill_type="solid")
                else:
                    value_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            # جدول تفاصيل الموظفين
            table_start_row = stats_row + 4
            
            # عنوان الجدول
            ws_dashboard[f'A{table_start_row}'] = "تفاصيل حضور الموظفين"
            ws_dashboard[f'A{table_start_row}'].font = Font(size=14, bold=True)
            ws_dashboard.merge_cells(f'A{table_start_row}:H{table_start_row}')
            
            # رؤوس الجدول
            headers = list(employee_data[0].keys()) if employee_data else []
            header_row = table_start_row + 1
            
            for col_num, header in enumerate(headers, 1):
                cell = ws_dashboard.cell(row=header_row, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # بيانات الموظفين
            for row_num, emp_data in enumerate(employee_data, header_row + 1):
                for col_num, value in enumerate(emp_data.values(), 1):
                    cell = ws_dashboard.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # تلوين صفوف متناوبة
                    if row_num % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # تعديل عرض الأعمدة
            ws_dashboard.column_dimensions['A'].width = 25
            ws_dashboard.column_dimensions['B'].width = 15
            ws_dashboard.column_dimensions['C'].width = 12
            ws_dashboard.column_dimensions['D'].width = 12
            ws_dashboard.column_dimensions['E'].width = 12
            ws_dashboard.column_dimensions['F'].width = 12
            ws_dashboard.column_dimensions['G'].width = 12
            ws_dashboard.column_dimensions['H'].width = 12
            
            # صفحة التفاصيل اليومية
            ws_daily = wb.create_sheet("التفاصيل اليومية")
            
            # عنوان صفحة التفاصيل اليومية
            ws_daily['A1'] = f"التفاصيل اليومية - قسم {department.name}"
            ws_daily['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws_daily['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws_daily['A1'].alignment = Alignment(horizontal='center')
            ws_daily.merge_cells('A1:AF3')
            
            # رؤوس الأعمدة للتفاصيل اليومية
            ws_daily['A5'] = 'اسم الموظف'
            ws_daily['B5'] = 'رقم الهوية'
            
            # تواريخ الشهر
            for col_num, date in enumerate(date_range, 3):
                cell = ws_daily.cell(row=5, column=col_num, value=date.strftime('%d-%m'))
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', textRotation=90)
            
            # بيانات الحضور اليومي
            for row_num, employee in enumerate(employees, 6):
                ws_daily.cell(row=row_num, column=1, value=employee.name)
                ws_daily.cell(row=row_num, column=2, value=employee.employee_id or 'غير محدد')
                
                # جلب سجلات الحضور للموظف
                attendance_records = Attendance.query.filter(
                    Attendance.employee_id == employee.id,
                    Attendance.date >= start_date,
                    Attendance.date <= end_date
                ).all()
                
                attendance_dict = {record.date: record.status for record in attendance_records}
                
                for col_num, date in enumerate(date_range, 3):
                    status = attendance_dict.get(date, '-')
                    cell = ws_daily.cell(row=row_num, column=col_num, value=status)
                    
                    # تلوين الخلايا حسب الحالة
                    if status == 'present':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.value = "✓"
                    elif status == 'absent':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.value = "✗"
                    elif status == 'leave':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.value = "إ"
                    elif status == 'sick':
                        cell.fill = PatternFill(start_color="9CC3F7", end_color="9CC3F7", fill_type="solid")
                        cell.value = "م"
                    
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # تنسيق أعمدة التفاصيل اليومية
            ws_daily.column_dimensions['A'].width = 30
            ws_daily.column_dimensions['B'].width = 15
            
            # تنسيق أعمدة التواريخ
            try:
                for col_num in range(3, min(3 + len(date_range), 26)):
                    if col_num <= 26:
                        col_letter = chr(64 + col_num)
                        ws_daily.column_dimensions[col_letter].width = 4
            except Exception:
                pass
            
            # حفظ الملف
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'تفاصيل_قسم_{department.name}_{today.strftime("%Y%m%d")}.xlsx'
            
            # إرجاع البيانات بدل استدعاء send_file (سيتم التعامل معها في الـ route)
            return {
                'buffer': output,
                'filename': filename,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
        except Exception as e:
            raise Exception(f"خطأ في تصدير تفاصيل القسم: {str(e)}")
