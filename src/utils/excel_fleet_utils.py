from .excel_base import ExcelStyles
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, timedelta
from src.utils.date_converter import parse_date, format_date_gregorian, format_date_hijri
from calendar import monthrange
import xlsxwriter

def generate_vehicles_excel(vehicles, output=None):
    """
    Generate Professional Excel file from vehicles data with Dashboard
    
    Args:
        vehicles: List of Vehicle objects
        output: BytesIO object to write to (optional)
        
    Returns:
        BytesIO object containing the Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from collections import Counter
        
        if output is None:
            output = BytesIO()
        
        workbook = Workbook()
        
        # ===== ورقة Dashboard الاحترافية =====
        dashboard = workbook.active
        dashboard.title = "Dashboard"
        
        # حساب الإحصائيات
        total_vehicles = len(vehicles)
        
        # إحصائيات الحالة
        status_map = {
            'available': 'متاحة',
            'rented': 'مؤجرة',
            'in_workshop': 'في الورشة',
            'in_project': 'في المشروع',
            'accident': 'حادث',
            'sold': 'مباعة'
        }
        status_counter = Counter()
        for v in vehicles:
            status_ar = status_map.get(v.status, v.status)
            status_counter[status_ar] += 1
        
        # إحصائيات الشركة المصنعة
        make_counter = Counter(v.make for v in vehicles if v.make)
        
        # إحصائيات الموديل
        model_counter = Counter(v.model for v in vehicles if v.model)
        
        # إحصائيات السنوات
        year_counter = Counter(str(v.year) for v in vehicles if v.year)
        
        # تنسيقات Dashboard
        title_font = Font(bold=True, color="FFFFFF", size=18, name='Calibri')
        title_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
        stat_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # عنوان Dashboard
        dashboard.merge_cells('A1:F1')
        title_cell = dashboard['A1']
        title_cell.value = f"تقرير المركبات - لوحة التحكم | {datetime.now().strftime('%Y-%m-%d')}"
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = title_fill
        
        # كارت الإحصائيات الرئيسية
        dashboard.merge_cells('A3:B3')
        dashboard['A3'].value = "🚗 إحصائيات عامة"
        dashboard['A3'].font = header_font
        dashboard['A3'].fill = header_fill
        dashboard['A3'].alignment = center_align
        
        # الإحصائيات حسب الحالة
        row = 4
        dashboard.cell(row=row, column=1).value = "إجمالي المركبات"
        dashboard.cell(row=row, column=1).font = Font(bold=True, size=11)
        dashboard.cell(row=row, column=1).alignment = Alignment(horizontal='right')
        dashboard.cell(row=row, column=2).value = total_vehicles
        dashboard.cell(row=row, column=2).font = Font(bold=True, size=14, color="4472C4")
        dashboard.cell(row=row, column=2).alignment = center_align
        dashboard.cell(row=row, column=2).fill = stat_fill
        row += 1
        
        for status_ar, count in status_counter.most_common():
            dashboard.cell(row=row, column=1).value = status_ar
            dashboard.cell(row=row, column=1).font = Font(bold=True, size=11)
            dashboard.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            dashboard.cell(row=row, column=2).value = count
            dashboard.cell(row=row, column=2).font = Font(bold=True, size=12)
            dashboard.cell(row=row, column=2).alignment = center_align
            dashboard.cell(row=row, column=2).fill = stat_fill
            row += 1
        
        # إحصائيات الشركات المصنعة
        dashboard.merge_cells('D3:F3')
        dashboard['D3'].value = "🏭 توزيع المركبات حسب الشركة المصنعة"
        dashboard['D3'].font = header_font
        dashboard['D3'].fill = header_fill
        dashboard['D3'].alignment = center_align
        
        row = 4
        for make, count in make_counter.most_common(10):
            dashboard.cell(row=row, column=4).value = make
            dashboard.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            dashboard.cell(row=row, column=5).value = count
            dashboard.cell(row=row, column=5).alignment = center_align
            dashboard.cell(row=row, column=5).fill = stat_fill
            
            # نسبة مئوية
            percentage = f"{(count/total_vehicles*100):.1f}%"
            dashboard.cell(row=row, column=6).value = percentage
            dashboard.cell(row=row, column=6).alignment = center_align
            dashboard.cell(row=row, column=6).font = Font(bold=True, color="4472C4")
            row += 1
        
        # أكثر الموديلات شيوعاً
        dashboard.merge_cells('A12:B12')
        dashboard['A12'].value = "🚙 أكثر الموديلات شيوعاً"
        dashboard['A12'].font = header_font
        dashboard['A12'].fill = header_fill
        dashboard['A12'].alignment = center_align
        
        row = 13
        for model, count in model_counter.most_common(8):
            dashboard.cell(row=row, column=1).value = model
            dashboard.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            dashboard.cell(row=row, column=2).value = count
            dashboard.cell(row=row, column=2).alignment = center_align
            dashboard.cell(row=row, column=2).fill = stat_fill
            row += 1
        
        # توزيع حسب سنة الصنع
        dashboard.merge_cells('D12:F12')
        dashboard['D12'].value = "📅 توزيع حسب سنة الصنع"
        dashboard['D12'].font = header_font
        dashboard['D12'].fill = header_fill
        dashboard['D12'].alignment = center_align
        
        row = 13
        for year, count in year_counter.most_common(8):
            dashboard.cell(row=row, column=4).value = year
            dashboard.cell(row=row, column=4).alignment = center_align
            dashboard.cell(row=row, column=5).value = count
            dashboard.cell(row=row, column=5).alignment = center_align
            dashboard.cell(row=row, column=5).fill = stat_fill
            
            percentage = f"{(count/total_vehicles*100):.1f}%"
            dashboard.cell(row=row, column=6).value = percentage
            dashboard.cell(row=row, column=6).alignment = center_align
            dashboard.cell(row=row, column=6).font = Font(bold=True, color="70AD47")
            row += 1
        
        # ضبط عرض الأعمدة في Dashboard
        dashboard.column_dimensions['A'].width = 25
        dashboard.column_dimensions['B'].width = 15
        dashboard.column_dimensions['C'].width = 3
        dashboard.column_dimensions['D'].width = 30
        dashboard.column_dimensions['E'].width = 12
        dashboard.column_dimensions['F'].width = 12
        
        # ===== ورقة بيانات المركبات الاحترافية =====
        vehicles_sheet = workbook.create_sheet(title="Vehicle Data")
        
        # العناوين الأساسية
        headers = [
            "رقم اللوحة", "الشركة المصنعة", "الموديل", "اللون",
            "سنة الصنع", "الحالة", "اسم السائق", "نوع السيارة",
            "تاريخ انتهاء الفحص", "تاريخ انتهاء الاستمارة"
        ]
        
        # كتابة العناوين
        for col_idx, header in enumerate(headers, start=1):
            cell = vehicles_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = center_align
            cell.border = border
        
        # كتابة البيانات
        for row_idx, vehicle in enumerate(vehicles, start=2):
            # البيانات الأساسية
            status_ar = status_map.get(vehicle.status, vehicle.status)
            data_row = [
                vehicle.plate_number or "",
                vehicle.make or "",
                vehicle.model or "",
                vehicle.color or "",
                str(vehicle.year) if vehicle.year else "",
                status_ar,
                vehicle.driver_name or "",
                vehicle.type_of_car or "",
                vehicle.inspection_expiry_date.strftime('%Y-%m-%d') if vehicle.inspection_expiry_date else "",
                vehicle.registration_expiry_date.strftime('%Y-%m-%d') if vehicle.registration_expiry_date else ""
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = vehicles_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = center_align
                cell.border = border
                
                # تلوين الصفوف بالتناوب
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # تلوين الحالة
                if col_idx == 6:  # عمود الحالة
                    if value == 'متاحة':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif value == 'في الورشة':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")
                    elif value in ['مؤجرة', 'في المشروع']:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True, color="9C6500")
        
        # ضبط عرض الأعمدة
        column_widths = [18, 20, 18, 15, 12, 15, 20, 22, 20, 20]
        for idx, width in enumerate(column_widths, start=1):
            vehicles_sheet.column_dimensions[get_column_letter(idx)].width = width
        
        # ===== ورقة البيانات الكاملة =====
        full_data_sheet = workbook.create_sheet(title="Complete Data")
        
        # جميع البيانات
        all_headers = [
            'رقم اللوحة', 'الشركة المصنعة', 'الموديل', 'اللون',
            'سنة الصنع', 'الحالة', 'اسم السائق', 'نوع السيارة',
            'المشروع', 'تاريخ انتهاء الفحص', 'تاريخ انتهاء الاستمارة',
            'تاريخ انتهاء التفويض', 'رابط Google Drive', 'الملاحظات', 'تاريخ الإضافة'
        ]
        
        for col_idx, header in enumerate(all_headers, start=1):
            cell = full_data_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = center_align
            cell.border = border
        
        for row_idx, vehicle in enumerate(vehicles, start=2):
            status_ar = status_map.get(vehicle.status, vehicle.status)
            all_data = [
                vehicle.plate_number or "",
                vehicle.make or "",
                vehicle.model or "",
                vehicle.color or "",
                str(vehicle.year) if vehicle.year else "",
                status_ar,
                vehicle.driver_name or "",
                vehicle.type_of_car or "",
                vehicle.project or "",
                vehicle.inspection_expiry_date.strftime('%Y-%m-%d') if vehicle.inspection_expiry_date else "",
                vehicle.registration_expiry_date.strftime('%Y-%m-%d') if vehicle.registration_expiry_date else "",
                vehicle.authorization_expiry_date.strftime('%Y-%m-%d') if vehicle.authorization_expiry_date else "",
                vehicle.drive_folder_link or "",
                vehicle.notes or "",
                vehicle.created_at.strftime('%Y-%m-%d') if vehicle.created_at else ""
            ]
            
            for col_idx, value in enumerate(all_data, start=1):
                cell = full_data_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = center_align
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # ضبط عرض الأعمدة
        for idx in range(1, len(all_headers) + 1):
            full_data_sheet.column_dimensions[get_column_letter(idx)].width = 20
        
        # ===== ورقة English Format =====
        english_sheet = workbook.create_sheet(title="English Format")
        
        # العناوين بالإنجليزية
        english_headers = [
            '#', 'Name', 'ID Num', 'EMP', 'Private Num', 'Work Num', 
            'Plate Num.', 'Owned By', 'TYPE OF VEHICLE', 'Project', 
            'Location', 'Start Date'
        ]
        
        # كتابة العناوين
        for col_idx, header in enumerate(english_headers, start=1):
            cell = english_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
            cell.fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # كتابة البيانات
        for row_idx, vehicle in enumerate(vehicles, start=2):
            # الحصول على بيانات السائق والموظف المرتبط
            driver_name = vehicle.driver_name or ""
            employee_id_num = ""
            employee_num = ""
            private_num = ""
            work_num = ""
            project = vehicle.project or ""
            location = ""
            start_date = ""
            owner = ""
            
            # محاولة الحصول على بيانات السائق من نموذج Employee بناءً على الاسم
            if driver_name:
                from models import Employee
                driver = Employee.query.filter_by(name=driver_name).first()
                if driver:
                    employee_id_num = driver.national_id or ""
                    employee_num = driver.employee_id or ""
                    private_num = driver.mobilePersonal or ""
                    work_num = driver.mobile or ""
            
            # الحصول على المنطقة من حقل region في المركبة
            location = vehicle.region or ""
            
            # محاولة الحصول على تاريخ البدء من المشروع
            if vehicle.project:
                from models import VehicleProject
                project_obj = VehicleProject.query.filter_by(
                    project_name=vehicle.project
                ).first()
                if project_obj:
                    if project_obj.start_date:
                        start_date = project_obj.start_date.strftime('%Y-%m-%d')
            
            # الحصول على الشركة المالكة من حقل owned_by أو من سجلات الإيجار
            owner = vehicle.owned_by or ""
            if not owner:
                # إذا لم يتم تحديد الشركة المالكة، جرب من سجلات الإيجار
                from models import VehicleRental
                rental = VehicleRental.query.filter_by(
                    vehicle_id=vehicle.id, 
                    is_active=True
                ).first()
                if rental:
                    owner = rental.lessor_name or ""
            
            # البيانات
            data_row = [
                row_idx - 1,  # الرقم التسلسلي
                driver_name,
                employee_id_num,
                employee_num,
                private_num,
                work_num,
                vehicle.plate_number or "",
                owner,  # شركة الإيجار أو المالك
                f"{vehicle.make or ''} - {vehicle.model or ''}".strip(' -'),
                project,
                location,
                start_date
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = english_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # تلوين الصفوف بالتناوب
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # ضبط عرض الأعمدة
        column_widths = [8, 30, 15, 12, 15, 15, 15, 20, 25, 18, 15, 15]
        for idx, width in enumerate(column_widths, start=1):
            english_sheet.column_dimensions[get_column_letter(idx)].width = width
        
        # حفظ الملف
        workbook.save(output)
        output.seek(0)
        return output
    
    except Exception as e:
        print(f"خطأ في إنشاء ملف Excel للمركبات: {str(e)}")
        import traceback
        traceback.print_exc()
        raise Exception(f"Error generating vehicles Excel file: {str(e)}")

