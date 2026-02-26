"""
أدوات مساعدة للتعامل مع ملفات Excel
"""

from openpyxl.utils import get_column_letter


def adjust_column_width(sheet):
    """
    ضبط عرض الأعمدة بشكل تلقائي في ورقة عمل Excel
    
    Args:
        sheet: ورقة العمل المراد ضبط عرض أعمدتها
    """
    # أعلى عدد من الأعمدة يمكن أن يكون في الورقة
    column_count = sheet.max_column
    
    # نضبط كل عمود على حدة
    for col_idx in range(1, column_count + 1):
        # الحصول على حرف العمود
        column_letter = get_column_letter(col_idx)
        
        # الحد الأقصى لطول القيم في العمود
        max_length = 0
        
        # تصفح جميع الخلايا في العمود
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                try:
                    # حساب طول القيمة كنص
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
                except:
                    pass
        
        # ضبط العرض مع مراعاة الحدود الدنيا والقصوى
        if max_length > 0:
            adjusted_width = min(max(max_length + 2, 10) * 1.2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        else:
            # عرض قياسي للأعمدة الفارغة
            sheet.column_dimensions[column_letter].width = 10