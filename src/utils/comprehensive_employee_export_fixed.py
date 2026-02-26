from datetime import date, datetime
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_comprehensive_employee_excel(employees):
    """إنشاء ملف Excel شامل لبيانات الموظفين مع جميع التفاصيل"""
    try:
        # إنشاء workbook
        workbook = Workbook()
        
        # حذف الورقة الافتراضية
        workbook.remove(workbook.active)
        
        # إنشاء ورقة البيانات الشاملة
        main_sheet = workbook.create_sheet("البيانات الشاملة")
        
        # إعداد البيانات الشاملة لكل موظف
        comprehensive_data = []
        
        for employee in employees:
            try:
                # البيانات الأساسية
                base_data = {
                    'رقم الموظف': str(employee.employee_id or ''),
                    'الاسم': str(employee.name or ''),
                    'رقم الهوية': str(employee.national_id or ''),
                    'رقم الجوال الأول': str(employee.mobile or ''),
                    'رقم الجوال الثاني': str(employee.mobile2 or ''),
                    'الإيميل': str(employee.email or ''),
                    'المنصب': str(employee.position or ''),
                    'الحالة الوظيفية': str(employee.status or ''),
                    'الموقع': str(employee.location or ''),
                    'المشروع': str(employee.project or ''),
                    'تاريخ الانضمام': employee.join_date.strftime('%Y-%m-%d') if employee.join_date else '',
                    'تاريخ الميلاد': employee.birth_date.strftime('%Y-%m-%d') if employee.birth_date else '',
                    'الجنسية': str(employee.nationality or ''),
                    'نوع العقد': 'سعودي' if employee.contract_type == 'saudi' else 'وافد',
                    'الراتب الأساسي': float(employee.basic_salary or 0),
                    'حالة العقد': str(employee.contract_status or ''),
                    'حالة الرخصة': str(employee.license_status or ''),
                    'ملاحظات': str(employee.notes or ''),
                    'تاريخ الإنشاء': employee.created_at.strftime('%Y-%m-%d %H:%M') if employee.created_at else '',
                    'تاريخ آخر تحديث': employee.updated_at.strftime('%Y-%m-%d %H:%M') if employee.updated_at else '',
                }
                
                # معلومات الأقسام
                departments_list = []
                if hasattr(employee, 'departments') and employee.departments:
                    departments_list = [str(dept.name) for dept in employee.departments]
                base_data['الأقسام'] = ' | '.join(departments_list) if departments_list else 'بدون قسم'
                base_data['عدد الأقسام'] = len(departments_list)
                
                # معلومات الجنسية المفصلة
                if hasattr(employee, 'nationality_rel') and employee.nationality_rel:
                    base_data['الجنسية التفصيلية'] = str(employee.nationality_rel.name_ar or '')
                    base_data['رمز الجنسية'] = str(employee.nationality_rel.code or '')
                else:
                    base_data['الجنسية التفصيلية'] = ''
                    base_data['رمز الجنسية'] = ''
                
                # معلومات نوع الموظف والعُهد
                base_data['نوع الموظف'] = 'سائق' if employee.employee_type == 'driver' else 'عادي'
                base_data['لديه عهدة جوال'] = 'نعم' if employee.has_mobile_custody else 'لا'
                base_data['نوع الجوال'] = str(employee.mobile_type or '')
                base_data['رقم IMEI'] = str(employee.mobile_imei or '')
                
                # معلومات الكفالة
                base_data['حالة الكفالة'] = 'على الكفالة' if employee.sponsorship_status == 'inside' else 'خارج الكفالة'
                base_data['اسم الكفيل الحالي'] = str(employee.current_sponsor_name or '')
                
                # المعلومات البنكية
                base_data['رقم الإيبان'] = str(employee.bank_iban or '')
                base_data['لديه صورة إيبان'] = 'نعم' if employee.bank_iban_image else 'لا'
                
                # معلومات الصور
                base_data['لديه صورة شخصية'] = 'نعم' if employee.profile_image else 'لا'
                base_data['لديه صورة هوية'] = 'نعم' if employee.national_id_image else 'لا'
                base_data['لديه صورة رخصة'] = 'نعم' if employee.license_image else 'لا'
                
                # إحصائيات الوثائق
                documents_count = 0
                expired_docs = 0
                valid_docs = 0
                doc_types = []
                
                if hasattr(employee, 'documents') and employee.documents:
                    documents_count = len(employee.documents)
                    for doc in employee.documents:
                        if hasattr(doc, 'document_type'):
                            doc_types.append(str(doc.document_type))
                        if hasattr(doc, 'expiry_date') and doc.expiry_date:
                            if doc.expiry_date < date.today():
                                expired_docs += 1
                            else:
                                valid_docs += 1
                
                base_data['عدد الوثائق'] = documents_count
                base_data['أنواع الوثائق'] = ' | '.join(set(doc_types)) if doc_types else ''
                base_data['الوثائق المنتهية'] = expired_docs
                base_data['الوثائق السارية'] = valid_docs
                
                # إحصائيات الرواتب
                salary_count = 0
                total_salaries = 0
                max_salary = 0
                min_salary = 0
                
                if hasattr(employee, 'salaries') and employee.salaries:
                    salary_records = [s for s in employee.salaries if hasattr(s, 'total_salary') and s.total_salary]
                    salary_count = len(salary_records)
                    
                    if salary_records:
                        salary_amounts = [float(s.total_salary) for s in salary_records]
                        total_salaries = sum(salary_amounts)
                        max_salary = max(salary_amounts)
                        min_salary = min(salary_amounts)
                
                base_data['عدد سجلات الراتب'] = salary_count
                base_data['متوسط الراتب'] = round(total_salaries / salary_count, 2) if salary_count > 0 else 0
                base_data['أعلى راتب'] = max_salary
                base_data['أدنى راتب'] = min_salary
                
                # إحصائيات الحضور
                attendance_count = 0
                present_days = 0
                absent_days = 0
                
                if hasattr(employee, 'attendances') and employee.attendances:
                    attendance_count = len(employee.attendances)
                    present_days = len([a for a in employee.attendances if hasattr(a, 'status') and a.status == 'present'])
                    absent_days = len([a for a in employee.attendances if hasattr(a, 'status') and a.status == 'absent'])
                
                base_data['عدد سجلات الحضور'] = attendance_count
                base_data['أيام الحضور'] = present_days
                base_data['أيام الغياب'] = absent_days
                base_data['نسبة الحضور'] = round((present_days / attendance_count) * 100, 2) if attendance_count > 0 else 0
                
                comprehensive_data.append(base_data)
                
            except Exception as e:
                print(f"خطأ في معالجة بيانات الموظف {employee.name}: {str(e)}")
                continue
        
        # تحويل البيانات إلى DataFrame
        if comprehensive_data:
            df = pd.DataFrame(comprehensive_data)
            
            # إضافة البيانات إلى الورقة
            for r in dataframe_to_rows(df, index=False, header=True):
                main_sheet.append(r)
            
            # تنسيق الرأس
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in main_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        
        # إنشاء ورقة الإحصائيات العامة
        stats_sheet = workbook.create_sheet("الإحصائيات العامة")
        
        # حساب الإحصائيات العامة
        total_employees = len(employees)
        if total_employees > 0:
            active_employees = len([e for e in employees if e.status == 'active'])
            inactive_employees = total_employees - active_employees
            mobile_custody_count = len([e for e in employees if e.has_mobile_custody])
            inside_sponsorship = len([e for e in employees if e.sponsorship_status == 'inside'])
            outside_sponsorship = len([e for e in employees if e.sponsorship_status == 'outside'])
            saudi_count = len([e for e in employees if e.contract_type == 'saudi'])
            foreign_count = len([e for e in employees if e.contract_type == 'foreign'])
            driver_count = len([e for e in employees if e.employee_type == 'driver'])
            regular_count = len([e for e in employees if e.employee_type == 'regular'])
            with_profile_image = len([e for e in employees if e.profile_image])
            with_id_image = len([e for e in employees if e.national_id_image])
            with_license_image = len([e for e in employees if e.license_image])
            with_iban_image = len([e for e in employees if e.bank_iban_image])
            with_iban = len([e for e in employees if e.bank_iban])
            
            # بيانات الإحصائيات
            stats_data = [
                ['إجمالي الموظفين', total_employees],
                ['الموظفون النشطون', active_employees],
                ['الموظفون غير النشطين', inactive_employees],
                ['نسبة النشطين %', round((active_employees / total_employees) * 100, 2)],
                ['', ''],
                ['الموظفون السعوديون', saudi_count],
                ['الموظفون الوافدون', foreign_count],
                ['نسبة السعوديين %', round((saudi_count / total_employees) * 100, 2)],
                ['', ''],
                ['السائقون', driver_count],
                ['الموظفون العاديون', regular_count],
                ['نسبة السائقين %', round((driver_count / total_employees) * 100, 2)],
                ['', ''],
                ['لديهم عهدة جوال', mobile_custody_count],
                ['نسبة العُهد %', round((mobile_custody_count / total_employees) * 100, 2)],
                ['', ''],
                ['على الكفالة', inside_sponsorship],
                ['خارج الكفالة', outside_sponsorship],
                ['نسبة على الكفالة %', round((inside_sponsorship / total_employees) * 100, 2)],
                ['', ''],
                ['لديهم صورة شخصية', with_profile_image],
                ['لديهم صورة هوية', with_id_image],
                ['لديهم صورة رخصة', with_license_image],
                ['لديهم صورة إيبان', with_iban_image],
                ['لديهم رقم إيبان', with_iban],
                ['', ''],
                ['نسبة اكتمال الصور الشخصية %', round((with_profile_image / total_employees) * 100, 2)],
                ['نسبة اكتمال صور الهوية %', round((with_id_image / total_employees) * 100, 2)],
                ['نسبة اكتمال المعلومات البنكية %', round((with_iban / total_employees) * 100, 2)],
                ['', ''],
                ['تاريخ التصدير', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]
            
            # إضافة رأس الإحصائيات
            stats_sheet.append(['البيان', 'القيمة'])
            
            # إضافة بيانات الإحصائيات
            for row in stats_data:
                stats_sheet.append(row)
            
            # تنسيق ورقة الإحصائيات
            for cell in stats_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        
        # حفظ الملف في memory
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        print(f"خطأ في إنشاء ملف Excel: {str(e)}")
        raise Exception(f"Error generating Excel file: {str(e)}")