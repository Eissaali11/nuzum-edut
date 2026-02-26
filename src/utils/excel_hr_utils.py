from .excel_base import ExcelStyles
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, timedelta
from src.utils.date_converter import parse_date, format_date_gregorian, format_date_hijri
from calendar import monthrange
import xlsxwriter

def parse_employee_excel(file):
    """
    Parse Excel file containing employee data
    
    Args:
        file: The uploaded Excel file
        
    Returns:
        List of dictionaries containing employee data
    """
    try:
        # Reset file pointer to beginning
        file.seek(0)
        
        # Read the Excel file explicitly using openpyxl engine
        df = pd.read_excel(file, engine='openpyxl')
        
        # Debug: Print column names
        print(f"Excel columns: {df.columns.tolist()}")
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        # Check if DataFrame is empty
        if df.empty:
            raise ValueError("Excel file is empty or has no data")
        
        # Create a more flexible column mapping
        column_mappings = {
            'name': ['name', 'الاسم الكامل', 'اسم', 'الاسم', 'full name', 'employee name', 'Name'],
            'employee_id': ['رقم الموظف', 'employee_id', 'emp_id', 'emp id', 'Emp .N', 'Emp.N', 'EmpN'],
            'national_id': ['رقم الهوية الوطنية', 'national_id', 'id', 'ID .N', 'ID Number', 'هوية'],
            'mobile': ['رقم الجوال', 'mobile', 'phone', 'هاتف', 'جوال', 'No.Mobile', 'Mobil'],
            'job_title': ['المسمى الوظيفي', 'job_title', 'position', 'title', 'Job Title', 'وظيفة'],
            'status': ['الحالة الوظيفية', 'status', 'حالة', 'Status'],
            'location': ['الموقع', 'location', 'موقع', 'Location'],
            'project': ['المشروع', 'project', 'مشروع', 'Project'],
            'email': ['البريد الإلكتروني', 'email', 'بريد', 'Email'],
            'department': ['الأقسام', 'department', 'قسم', 'Department'],
            'join_date': ['تاريخ الانضمام', 'join_date', 'hire_date', 'انضمام'],
            'license_end_date': ['تاريخ انتهاء الإقامة', 'license_end_date', 'انتهاء الإقامة'],
            'contract_status': ['حالة العقد', 'contract_status', 'عقد'],
            'license_status': ['حالة الرخصة', 'license_status', 'رخصة'],
            'nationality': ['الجنسية', 'nationality', 'جنسية'],
            'notes': ['ملاحظات', 'notes', 'remarks', 'comments'],
            'mobilePersonal': ['الجوال الشخصي', 'mobile_personal', 'جوال شخصي']
        }
        
        # Map columns to their field names
        detected_columns = {}
        for col in df.columns:
            if isinstance(col, datetime):
                continue
                
            col_str = str(col).strip()
            
            # Check for matches in column mappings
            for field, variations in column_mappings.items():
                if col_str in variations:
                    detected_columns[field] = col
                    print(f"Detected '{field}' column: {col}")
                    break
        
        # If no columns detected, try to guess from position and content
        if not detected_columns:
            print("No columns detected by name, trying to guess from position...")
            columns_list = [col for col in df.columns if not isinstance(col, datetime)]
            
            # If we have enough columns, try to guess based on position
            if len(columns_list) >= 3:
                # Basic required fields
                detected_columns['name'] = columns_list[0]
                detected_columns['employee_id'] = columns_list[1] if len(columns_list) > 1 else None
                detected_columns['national_id'] = columns_list[2] if len(columns_list) > 2 else None
                
                # Optional fields
                if len(columns_list) > 3:
                    detected_columns['mobile'] = columns_list[3]
                if len(columns_list) > 4:
                    detected_columns['job_title'] = columns_list[4]
                
                print(f"Guessed columns: {detected_columns}")
        
        # Check for minimum required columns
        required_fields = ['name']
        missing_required = [field for field in required_fields if field not in detected_columns]
        
        if missing_required:
            raise ValueError(f"Required columns missing: {', '.join(missing_required)}. Available columns: {[c for c in df.columns if not isinstance(c, datetime)]}")
        
        # Process each row
        employees = []
        for idx, row in df.iterrows():
            try:
                # Skip completely empty rows
                if row.isnull().all():
                    continue
                
                # Check if name is present
                name_col = detected_columns.get('name')
                if name_col and pd.isna(row[name_col]):
                    continue
                
                # Create employee dictionary
                employee = {}
                
                # Add name (required)
                if name_col:
                    employee['name'] = str(row[name_col]).strip()
                
                # Add employee_id (auto-generate if missing)
                emp_id_col = detected_columns.get('employee_id')
                if emp_id_col and not pd.isna(row[emp_id_col]):
                    employee['employee_id'] = str(row[emp_id_col]).strip()
                else:
                    employee['employee_id'] = f"EMP{idx+1000}"
                
                # Add national_id (auto-generate if missing)
                national_id_col = detected_columns.get('national_id')
                if national_id_col and not pd.isna(row[national_id_col]):
                    employee['national_id'] = str(row[national_id_col]).strip()
                else:
                    employee['national_id'] = f"N{idx+5000:07d}"
                
                # Add mobile (auto-generate if missing)
                mobile_col = detected_columns.get('mobile')
                if mobile_col and not pd.isna(row[mobile_col]):
                    employee['mobile'] = str(row[mobile_col]).strip()
                else:
                    employee['mobile'] = f"05xxxxxxxx"
                
                # Add job_title (default if missing)
                job_title_col = detected_columns.get('job_title')
                if job_title_col and not pd.isna(row[job_title_col]):
                    employee['job_title'] = str(row[job_title_col]).strip()
                else:
                    employee['job_title'] = "موظف"
                
                # Add status (default to active)
                status_col = detected_columns.get('status')
                if status_col and not pd.isna(row[status_col]):
                    status_value = str(row[status_col]).lower().strip()
                    if status_value in ['active', 'نشط', 'فعال']:
                        employee['status'] = 'active'
                    elif status_value in ['inactive', 'غير نشط', 'غير فعال']:
                        employee['status'] = 'inactive'
                    elif status_value in ['on_leave', 'on leave', 'leave', 'إجازة', 'في إجازة']:
                        employee['status'] = 'on_leave'
                    else:
                        employee['status'] = 'active'
                else:
                    employee['status'] = 'active'
                
                # Add optional fields (excluding department which is handled separately)
                optional_fields = ['location', 'project', 'email', 'join_date', 
                                 'license_end_date', 'contract_status', 'license_status', 
                                 'nationality', 'notes', 'mobilePersonal']
                
                for field in optional_fields:
                    col = detected_columns.get(field)
                    if col and not pd.isna(row[col]):
                        employee[field] = str(row[col]).strip()
                
                # Handle department separately
                dept_col = detected_columns.get('department')
                if dept_col and not pd.isna(row[dept_col]):
                    employee['department'] = str(row[dept_col]).strip()
                
                # Debug: Print processed employee
                print(f"Processed employee {idx+1}: {employee.get('name', 'Unknown')}")
                
                employees.append(employee)
                
            except Exception as e:
                print(f"Error processing row {idx+1}: {str(e)}")
                continue
        
        if not employees:
            raise ValueError("No valid employee records found in the Excel file")
            
        print(f"Successfully parsed {len(employees)} employee records")
        return employees
    
    except Exception as e:
        import traceback
        print(f"Error parsing Excel: {str(e)}")
        print(traceback.format_exc())
        raise Exception(f"Error parsing Excel file: {str(e)}")

def export_employees_to_excel(employees, output=None):
    """
    Export employees to Excel file
    
    Args:
        employees: List of Employee objects
        output: BytesIO object to write to (optional)
        
    Returns:
        BytesIO object containing the Excel file
    """
    return generate_employee_excel(employees, output)
    
def generate_employee_excel(employees, output=None):
    """
    Generate Professional Excel file from employee data with Dashboard
    
    Args:
        employees: List of Employee objects
        output: BytesIO object to write to (optional)
        
    Returns:
        BytesIO object containing the Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter
        from collections import Counter
        
        if output is None:
            output = BytesIO()
        
        workbook = Workbook()
        
        # ===== ورقة Dashboard الاحترافية =====
        dashboard = workbook.active
        dashboard.title = "Dashboard"
        
        # حساب الإحصائيات
        total_employees = len(employees)
        active_employees = sum(1 for e in employees if e.status == 'active')
        inactive_employees = sum(1 for e in employees if e.status == 'inactive')
        on_leave_employees = sum(1 for e in employees if e.status == 'on_leave')
        
        # إحصائيات الأقسام
        dept_counter = Counter()
        for emp in employees:
            if emp.departments:
                for dept in emp.departments:
                    dept_counter[dept.name] += 1
            else:
                dept_counter['بدون قسم'] += 1
        
        # إحصائيات الوظائف
        job_counter = Counter(e.job_title for e in employees if e.job_title)
        
        # تنسيقات Dashboard
        title_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=16, name='Calibri')
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
        stat_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # عنوان Dashboard
        dashboard.merge_cells('A1:F1')
        title_cell = dashboard['A1']
        title_cell.value = f"تقرير الموظفين - لوحة التحكم | {datetime.now().strftime('%Y-%m-%d')}"
        title_cell.font = title_fill = Font(bold=True, color="FFFFFF", size=18, name='Calibri')
        title_cell.alignment = center_align
        title_cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        
        # كارت الإحصائيات الرئيسية
        dashboard.merge_cells('A3:B3')
        dashboard['A3'].value = "📊 إحصائيات عامة"
        dashboard['A3'].font = header_font
        dashboard['A3'].fill = header_fill
        dashboard['A3'].alignment = center_align
        
        stats_data = [
            ("إجمالي الموظفين", total_employees, "4472C4"),
            ("الموظفون النشطون", active_employees, "70AD47"),
            ("الموظفون غير النشطين", inactive_employees, "FFC000"),
            ("في إجازة", on_leave_employees, "ED7D31")
        ]
        
        row = 4
        for label, value, color in stats_data:
            dashboard.cell(row=row, column=1).value = label
            dashboard.cell(row=row, column=1).font = Font(bold=True, size=11)
            dashboard.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            
            dashboard.cell(row=row, column=2).value = value
            dashboard.cell(row=row, column=2).font = Font(bold=True, size=14, color=color)
            dashboard.cell(row=row, column=2).alignment = center_align
            dashboard.cell(row=row, column=2).fill = stat_fill
            row += 1
        
        # إحصائيات الأقسام
        dashboard.merge_cells('D3:F3')
        dashboard['D3'].value = "🏢 توزيع الموظفين حسب الأقسام"
        dashboard['D3'].font = header_font
        dashboard['D3'].fill = header_fill
        dashboard['D3'].alignment = center_align
        
        row = 4
        for dept_name, count in dept_counter.most_common(10):
            dashboard.cell(row=row, column=4).value = dept_name
            dashboard.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            dashboard.cell(row=row, column=5).value = count
            dashboard.cell(row=row, column=5).alignment = center_align
            dashboard.cell(row=row, column=5).fill = stat_fill
            
            # نسبة مئوية
            percentage = f"{(count/total_employees*100):.1f}%"
            dashboard.cell(row=row, column=6).value = percentage
            dashboard.cell(row=row, column=6).alignment = center_align
            dashboard.cell(row=row, column=6).font = Font(bold=True, color="4472C4")
            row += 1
        
        # أكثر الوظائف
        dashboard.merge_cells('A10:B10')
        dashboard['A10'].value = "💼 أكثر الوظائف شيوعاً"
        dashboard['A10'].font = header_font
        dashboard['A10'].fill = header_fill
        dashboard['A10'].alignment = center_align
        
        row = 11
        for job_title, count in job_counter.most_common(8):
            dashboard.cell(row=row, column=1).value = job_title
            dashboard.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            dashboard.cell(row=row, column=2).value = count
            dashboard.cell(row=row, column=2).alignment = center_align
            dashboard.cell(row=row, column=2).fill = stat_fill
            row += 1
        
        # ضبط عرض الأعمدة في Dashboard
        dashboard.column_dimensions['A'].width = 25
        dashboard.column_dimensions['B'].width = 15
        dashboard.column_dimensions['C'].width = 3
        dashboard.column_dimensions['D'].width = 25
        dashboard.column_dimensions['E'].width = 12
        dashboard.column_dimensions['F'].width = 12
        
        # ===== ورقة بيانات الموظفين الاحترافية =====
        employees_sheet = workbook.create_sheet(title="Employee Data")
        
        # العناوين الأساسية
        headers = [
            "الاسم الكامل", "رقم الموظف", "رقم الهوية", "رقم الجوال", 
            "المسمى الوظيفي", "الحالة", "الموقع", "المشروع", 
            "الأقسام", "البريد الإلكتروني", "تاريخ الانضمام"
        ]
        
        # كتابة العناوين
        for col_idx, header in enumerate(headers, start=1):
            cell = employees_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = center_align
            cell.border = border
        
        # كتابة البيانات
        for row_idx, employee in enumerate(employees, start=2):
            # البيانات الأساسية
            data_row = [
                employee.name,
                employee.employee_id,
                employee.national_id or "",
                employee.mobile or "",
                employee.job_title or "",
                employee.status or "",
                employee.location or "",
                employee.project or "",
                ', '.join([dept.name for dept in employee.departments]) if employee.departments else "",
                employee.email or "",
                employee.join_date.strftime('%Y-%m-%d') if employee.join_date else ""
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = employees_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = center_align
                cell.border = border
                
                # تلوين الصفوف بالتناوب
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # تلوين الحالة
                if col_idx == 6:  # عمود الحالة
                    if value == 'active':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif value == 'inactive':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")
                    elif value == 'on_leave':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True, color="9C6500")
        
        # ضبط عرض الأعمدة
        column_widths = [30, 15, 18, 16, 20, 12, 18, 18, 25, 25, 18]
        for idx, width in enumerate(column_widths, start=1):
            employees_sheet.column_dimensions[get_column_letter(idx)].width = width
        
        # ===== ورقة البيانات الكاملة =====
        full_data_sheet = workbook.create_sheet(title="Complete Data")
        
        # جميع البيانات بالترتيب المطلوب
        all_headers = [
            'الاسم الكامل', 'رقم الهوية الوطنية', 'رقم الموظف', 'الجنسية',
            'الجوال الشخصي', 'مقاس البنطلون', 'مقاس التيشرت', 'المسمى الوظيفي',
            'الحالة الوظيفية', 'نوع الجوال', 'رقم IMEI', 'رقم الجوال',
            'السيارة الحاليه', 'الموقع', 'المشروع', 'البريد الإلكتروني',
            'الأقسام', 'تاريخ الانضمام', 'تاريخ الميلاد', 'نوع الموظف',
            'نوع العقد', 'الراتب الأساسي', 'حالة العقد', 'حالة الرخصة',
            'حالة الكفالة', 'اسم الكفيل', 'رقم الإيبان', 'تفاصيل السكن', 'رابط موقع السكن',
            'ملف العرض الوظيفي', 'صورة الجواز', 'شهادة العنوان الوطني',
            'رابط العرض الوظيفي الخارجي', 'رابط صورة الجواز الخارجي', 'رابط شهادة العنوان الخارجي'
        ]
        
        for col_idx, header in enumerate(all_headers, start=1):
            cell = full_data_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = center_align
            cell.border = border
        
        for row_idx, employee in enumerate(employees, start=2):
            # جلب السيارة الحالية (المسلمة للموظف) من نظام إدارة السيارات
            current_vehicle = ""
            try:
                from models import VehicleHandover
                # البحث عن آخر سجل تسليم (delivery) للموظف
                latest_delivery = VehicleHandover.query.filter_by(
                    employee_id=employee.id
                ).filter(
                    VehicleHandover.handover_type.in_(['delivery', 'تسليم', 'handover'])
                ).order_by(VehicleHandover.handover_date.desc()).first()
                
                if latest_delivery and latest_delivery.vehicle:
                    current_vehicle = f"{latest_delivery.vehicle.plate_number}"
            except:
                pass
            
            # جلب بيانات الجهاز المحمول من نظام إدارة الأجهزة المحمولة
            mobile_type = ""
            mobile_imei = ""
            mobile_number = ""
            try:
                from models import DeviceAssignment, MobileDevice, SimCard
                
                # البحث عن التعيين النشط للموظف
                active_assignment = DeviceAssignment.query.filter_by(
                    employee_id=employee.id,
                    is_active=True
                ).first()
                
                print(f"🔍 الموظف {employee.name} (ID: {employee.id}): التعيين النشط = {active_assignment}")
                
                if active_assignment:
                    # جلب معلومات الجهاز
                    if active_assignment.device_id:
                        device = MobileDevice.query.get(active_assignment.device_id)
                        print(f"   📱 الجهاز: {device}")
                        if device:
                            mobile_type = f"{device.device_brand or ''} {device.device_model or ''}".strip()
                            mobile_imei = device.imei or ""
                            print(f"   ✅ نوع الجوال: {mobile_type}, IMEI: {mobile_imei}")
                    
                    # جلب رقم الهاتف من SIM Card
                    if active_assignment.sim_card_id:
                        sim = SimCard.query.get(active_assignment.sim_card_id)
                        print(f"   📞 SIM Card: {sim}")
                        if sim:
                            mobile_number = sim.phone_number or ""
                            print(f"   ✅ رقم الجوال: {mobile_number}")
                else:
                    print(f"   ❌ لا يوجد تعيين نشط للموظف")
                    
            except Exception as e:
                import traceback
                print(f"❌ خطأ في جلب بيانات الجهاز للموظف {employee.name} ({employee.id}): {str(e)}")
                print(traceback.format_exc())
            
            # إضافة روابط الملفات الجديدة
            job_offer_link = f"https://nuzum.site/static/{getattr(employee, 'job_offer_file', '')}" if getattr(employee, 'job_offer_file', '') else '-'
            passport_link = f"https://nuzum.site/static/{getattr(employee, 'passport_image_file', '')}" if getattr(employee, 'passport_image_file', '') else '-'
            national_address_link = f"https://nuzum.site/static/{getattr(employee, 'national_address_file', '')}" if getattr(employee, 'national_address_file', '') else '-'
            
            # إضافة الروابط الخارجية
            job_offer_external = getattr(employee, 'job_offer_link', '') or '-'
            passport_external = getattr(employee, 'passport_image_link', '') or '-'
            national_address_external = getattr(employee, 'national_address_link', '') or '-'
            
            all_data = [
                employee.name,  # 1. الاسم الكامل
                employee.national_id or "",  # 2. رقم الهوية الوطنية
                employee.employee_id,  # 3. رقم الموظف
                employee.nationality_rel.name_ar if hasattr(employee, 'nationality_rel') and employee.nationality_rel else (employee.nationality if hasattr(employee, 'nationality') else ""),  # 4. الجنسية
                getattr(employee, 'mobilePersonal', '') or '',  # 5. الجوال الشخصي
                getattr(employee, 'pants_size', '') or '',  # 6. مقاس البنطلون
                getattr(employee, 'shirt_size', '') or '',  # 7. مقاس التيشرت
                employee.job_title or "",  # 8. المسمى الوظيفي
                employee.status or "",  # 9. الحالة الوظيفية
                mobile_type,  # 10. نوع الجوال (من نظام إدارة الأجهزة)
                mobile_imei,  # 11. رقم IMEI (من نظام إدارة الأجهزة)
                mobile_number,  # 12. رقم الجوال (من نظام إدارة الأجهزة)
                current_vehicle,  # 13. السيارة الحالية (من نظام إدارة السيارات)
                employee.location or "",  # 14. الموقع
                employee.project or "",  # 15. المشروع
                employee.email or "",  # 16. البريد الإلكتروني
                ', '.join([dept.name for dept in employee.departments]) if employee.departments else "",  # 17. الأقسام
                employee.join_date.strftime('%Y-%m-%d') if employee.join_date else "",  # 18. تاريخ الانضمام
                employee.birth_date.strftime('%Y-%m-%d') if employee.birth_date else "",  # 19. تاريخ الميلاد
                getattr(employee, 'employee_type', '') or '',  # 20. نوع الموظف
                getattr(employee, 'contract_type', '') or '',  # 21. نوع العقد
                str(getattr(employee, 'basic_salary', '') or ''),  # 22. الراتب الأساسي
                getattr(employee, 'contract_status', '') or '',  # 23. حالة العقد
                getattr(employee, 'license_status', '') or '',  # 24. حالة الرخصة
                getattr(employee, 'sponsorship_status', '') or '',  # 25. حالة الكفالة
                getattr(employee, 'current_sponsor_name', '') or '',  # 26. اسم الكفيل
                getattr(employee, 'bank_iban', '') or '',  # 27. رقم الإيبان
                getattr(employee, 'residence_details', '') or '',  # 28. تفاصيل السكن
                getattr(employee, 'residence_location_url', '') or '',  # 29. رابط موقع السكن
                job_offer_link,  # 30. ملف العرض الوظيفي
                passport_link,  # 31. صورة الجواز
                national_address_link,  # 32. شهادة العنوان الوطني
                job_offer_external,  # 33. رابط العرض الوظيفي الخارجي
                passport_external,  # 34. رابط صورة الجواز الخارجي
                national_address_external  # 35. رابط شهادة العنوان الخارجي
            ]
            
            for col_idx, value in enumerate(all_data, start=1):
                cell = full_data_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = center_align
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # ضبط عرض الأعمدة
        for idx in range(1, len(all_headers) + 1):
            full_data_sheet.column_dimensions[get_column_letter(idx)].width = 18
        
        # حفظ الملف
        workbook.save(output)
        output.seek(0)
        return output
    
    except Exception as e:
        print(f"خطأ في إنشاء ملف Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise Exception(f"Error generating Excel file: {str(e)}")

def parse_document_excel(file):
    """
    Parse Excel file containing document data
    
    Args:
        file: The uploaded Excel file
        
    Returns:
        List of dictionaries containing document data
    """
    try:
        # Reset file pointer to beginning
        file.seek(0)
        
        # Read the Excel file explicitly using openpyxl engine
        # Try to detect if this is a report-style export with header rows
        try:
            df = pd.read_excel(file, engine='openpyxl')
            
            # Check if first row looks like a report title
            if len(df.columns) > 0 and str(df.columns[0]).startswith('تقرير'):
                # This is a report-style export, try reading from row 2 or find actual header
                df_test = pd.read_excel(file, engine='openpyxl', header=None)
                
                # Look for a row that contains document field names
                header_row = None
                for i in range(min(10, len(df_test))):  # Check first 10 rows
                    row_values = [str(val).lower() for val in df_test.iloc[i].values if pd.notna(val)]
                    if any('موظف' in val or 'employee' in val for val in row_values):
                        header_row = i
                        break
                
                if header_row is not None:
                    df = pd.read_excel(file, engine='openpyxl', header=header_row)
                else:
                    # Try standard document import format
                    df = pd.read_excel(file, engine='openpyxl', header=0)
            
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            # Fallback to standard reading
            df = pd.read_excel(file, engine='openpyxl')
        
        # Print column names for debugging
        print(f"Document Excel columns: {df.columns.tolist()}")
        
        # Create a mapping for column detection - include export format columns
        column_mappings = {
            'employee_id': ['employee_id', 'employee id', 'emp id', 'employee number', 'emp no', 'emp.id', 'emp.no', 'رقم الموظف', 'معرف الموظف', 'الرقم الوظيفي', 'موظف', 'اسم الموظف'],
            'document_type': ['document_type', 'document type', 'type', 'doc type', 'نوع الوثيقة', 'نوع المستند', 'النوع', 'نوع الوثيقة'],
            'document_number': ['document_number', 'document no', 'doc number', 'doc no', 'رقم الوثيقة', 'رقم المستند'],
            'issue_date': ['issue_date', 'issue date', 'start date', 'تاريخ الإصدار', 'تاريخ البدء'],
            'expiry_date': ['expiry_date', 'expiry date', 'end date', 'valid until', 'تاريخ الانتهاء', 'صالح حتى'],
            'notes': ['notes', 'comments', 'remarks', 'ملاحظات', 'تعليقات']
        }
        
        # Map columns to their normalized names
        detected_columns = {}
        for col in df.columns:
            if isinstance(col, datetime):
                continue
                
            col_str = str(col).lower().strip()
            
            # Check for exact column name or common variations
            for field, variations in column_mappings.items():
                if col_str in variations or any(var in col_str for var in variations):
                    detected_columns[field] = col
                    print(f"Detected '{field}' column: {col}")
                    break
        
        # Handle special case for Excel files with specific column names
        explicit_mappings = {
            'Employee ID': 'employee_id',
            'Document Type': 'document_type',
            'Document Number': 'document_number',
            'Issue Date': 'issue_date',
            'Expiry Date': 'expiry_date',
            'Notes': 'notes'
        }
        
        for excel_col, field in explicit_mappings.items():
            if excel_col in df.columns:
                detected_columns[field] = excel_col
                print(f"Explicitly mapped '{excel_col}' to '{field}'")
        
        # Print final column mapping
        print(f"Final document column mapping: {detected_columns}")
        
        # تقسيم الحقول المطلوبة إلى أساسية وغير أساسية
        essential_fields = ['employee_id', 'document_type', 'document_number']  # الحقول الأساسية التي يجب توفرها
        other_fields = ['issue_date', 'expiry_date', 'notes']  # الحقول التي يمكن إضافة قيم افتراضية لها
        
        # التحقق من الحقول الأساسية
        missing_essential = [field for field in essential_fields if field not in detected_columns]
        if missing_essential:
            missing_str = ", ".join(missing_essential)
            raise ValueError(f"Required columns missing: {missing_str}. Available columns: {[c for c in df.columns if not isinstance(c, datetime)]}")
        
        # بالنسبة للحقول غير الأساسية المفقودة، سننشئ أعمدة وهمية تحتوي على قيم افتراضية
        for field in other_fields:
            if field not in detected_columns:
                print(f"Warning: Creating default column for: {field}")
                dummy_column_name = f"__{field}__default"
                # إذا كان الحقل هو تاريخ، نضيف تاريخ افتراضي (اليوم للإصدار، وبعد سنة للانتهاء)
                if field == 'issue_date':
                    default_value = datetime.now()
                elif field == 'expiry_date':
                    default_value = datetime.now() + timedelta(days=365)
                else:
                    default_value = ''  # للملاحظات
                
                df[dummy_column_name] = default_value
                detected_columns[field] = dummy_column_name  # تعيين العمود الوهمي للحقل
        
        # Process each row
        documents = []
        for idx, row in df.iterrows():
            try:
                # Skip completely empty rows
                if row.isnull().all():
                    continue
                
                # Get employee_id field
                emp_id_col = detected_columns['employee_id']
                emp_id = row[emp_id_col]
                
                # Skip rows with missing employee_id
                if pd.isna(emp_id):
                    print(f"Skipping row {idx+1} due to missing employee ID")
                    continue
                
                # Try to convert employee_id to integer
                try:
                    employee_id = int(emp_id)
                except (ValueError, TypeError):
                    # If not convertible to int, use as string (could be employee code)
                    employee_id = str(emp_id).strip()
                
                # Get document type and number
                doc_type_col = detected_columns['document_type']
                doc_type = row[doc_type_col]
                
                doc_number_col = detected_columns['document_number']
                doc_number = row[doc_number_col]
                
                # Skip rows with missing document type or number
                if pd.isna(doc_type) or pd.isna(doc_number):
                    print(f"Skipping row {idx+1} due to missing document type or number")
                    continue
                
                # Get dates and parse them
                issue_date_col = detected_columns['issue_date']
                expiry_date_col = detected_columns['expiry_date']
                
                # تعامل مع التواريخ المفقودة - استخدام تاريخ اليوم للإصدار وبعد سنة للانتهاء
                if pd.isna(row[issue_date_col]):
                    print(f"Row {idx+1}: Using today's date for missing issue date")
                    issue_date_val = datetime.now()
                else:
                    issue_date_val = row[issue_date_col]
                    
                if pd.isna(row[expiry_date_col]):
                    print(f"Row {idx+1}: Using date one year from today for missing expiry date")
                    expiry_date_val = datetime.now() + timedelta(days=365)
                else:
                    expiry_date_val = row[expiry_date_col]
                
                try:
                    # Handle different date formats and convert to datetime
                    if isinstance(issue_date_val, datetime):
                        issue_date = issue_date_val
                    else:
                        issue_date = parse_date(str(issue_date_val))
                        
                    if isinstance(expiry_date_val, datetime):
                        expiry_date = expiry_date_val
                    else:
                        expiry_date = parse_date(str(expiry_date_val))
                    
                    # استخدام تواريخ افتراضية في حالة فشل تحليل التواريخ
                    if not issue_date:
                        print(f"Row {idx+1}: Using today's date due to invalid issue date format")
                        issue_date = datetime.now()
                        
                    if not expiry_date:
                        print(f"Row {idx+1}: Using date one year from today due to invalid expiry date format")
                        expiry_date = datetime.now() + timedelta(days=365)
                        
                except (ValueError, TypeError) as e:
                    print(f"Row {idx+1}: Date parsing error: {str(e)}, using default dates")
                    issue_date = datetime.now()
                    expiry_date = datetime.now() + timedelta(days=365)
                
                # Create document dictionary
                document = {
                    'employee_id': employee_id,
                    'document_type': str(doc_type).strip(),
                    'document_number': str(doc_number).strip(),
                    'issue_date': issue_date,
                    'expiry_date': expiry_date
                }
                
                # Add notes if present
                if 'notes' in detected_columns and not pd.isna(row[detected_columns['notes']]):
                    document['notes'] = str(row[detected_columns['notes']])
                
                print(f"Processed document for employee ID: {employee_id}, type: {document['document_type']}")
                documents.append(document)
                
            except Exception as e:
                print(f"Error processing document row {idx+1}: {str(e)}")
                # Continue to next row instead of failing the entire import
                continue
        
        if not documents:
            raise ValueError("No valid document records found in the Excel file")
            
        return documents
    
    except Exception as e:
        import traceback
        print(f"Error parsing document Excel: {str(e)}")
        print(traceback.format_exc())
        raise Exception(f"Error parsing document Excel file: {str(e)}")

