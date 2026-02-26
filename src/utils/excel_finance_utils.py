from .excel_base import ExcelStyles
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, timedelta
from src.utils.date_converter import parse_date, format_date_gregorian, format_date_hijri
from src.utils.salary_calculator import compute_net_salary
from calendar import monthrange
import xlsxwriter


def _get_gosi_amount_for_salary_record(salary_record):
    """استخراج قيمة GOSI لسجل راتب باستخدام الحاسبة الموحدة."""
    breakdown = compute_net_salary(
        basic_salary=getattr(salary_record, 'basic_salary', 0),
        allowances=0,
        bonus=0,
        deductions=0,
        nationality=getattr(getattr(salary_record, 'employee', None), 'nationality', None),
        contract_type=getattr(getattr(salary_record, 'employee', None), 'contract_type', None),
        return_breakdown=True,
    )
    return float(breakdown.get('gosi_deduction', 0) or 0)


def _calculate_column_width(dataframe, column_name, padding=4):
    """حساب عرض العمود بشكل آمن حتى مع القيم الفارغة أو غير النصية."""
    if dataframe is None or column_name not in dataframe.columns or len(dataframe) == 0:
        return len(str(column_name)) + padding

    max_value_len = dataframe[column_name].map(
        lambda value: len(str(value)) if pd.notna(value) else 0
    ).max()

    return max(max_value_len, len(str(column_name))) + padding


def parse_salary_excel(file, month, year):
    """
    Parse Excel file containing salary data
    
    Args:
        file: The uploaded Excel file
        month: The month for these salaries
        year: The year for these salaries
        
    Returns:
        List of dictionaries containing salary data
    """
    try:
        from models import Employee

        # Reset file pointer to beginning
        file.seek(0)
        
        # Read the Excel file explicitly using openpyxl engine
        df = pd.read_excel(file, engine='openpyxl')
        
        # Print column names for debugging
        print(f"Salary Excel columns: {df.columns.tolist()}")
        
        # Create a mapping for column detection
        column_mappings = {
            'employee_id': ['employee_id', 'employee id', 'emp id', 'employee number', 'emp no', 'emp.id', 'emp.no', 'emp .n', 'رقم الموظف', 'معرف الموظف', 'الرقم الوظيفي'],
            'basic_salary': ['basic_salary', 'basic salary', 'salary', 'راتب', 'الراتب', 'الراتب الأساسي'],
            'allowances': ['allowances', 'بدل', 'بدلات', 'البدلات'],
            'deductions': ['deductions', 'خصم', 'خصومات', 'الخصومات'],
            'bonus': ['bonus', 'مكافأة', 'علاوة', 'مكافآت'],
            'notes': ['notes', 'ملاحظات']
        }
        
        # Map columns to their normalized names
        detected_columns = {}
        for col in df.columns:
            if isinstance(col, datetime):
                continue
                
            col_str = str(col).lower().strip()
            
            # Check for exact column name or common variations
            for field, variations in column_mappings.items():
                if col_str in variations or any(var in col_str for var in variations):
                    detected_columns[field] = col
                    print(f"Detected '{field}' column: {col}")
                    break
        
        # Handle special case for Excel files with specific column names
        explicit_mappings = {
            'Employee ID': 'employee_id',
            'Emp .N': 'employee_id',  # شكل آخر لرقم الموظف
            'Basic Salary': 'basic_salary',
            'Allowances': 'allowances',
            'Deductions': 'deductions',
            'Bonus': 'bonus',
            'Notes': 'notes'
        }
        
        for excel_col, field in explicit_mappings.items():
            if excel_col in df.columns:
                detected_columns[field] = excel_col
                print(f"Explicitly mapped '{excel_col}' to '{field}'")
        
        # Print final column mapping
        print(f"Final salary column mapping: {detected_columns}")
        
        # تقسيم الحقول المطلوبة إلى أساسية وغير أساسية
        essential_fields = ['employee_id']  # رقم الموظف هو الأساسي الوحيد الضروري دائماً
        other_fields = ['basic_salary', 'allowances', 'deductions', 'bonus']  # يمكن وضع قيم افتراضية لها
        
        # التحقق من الحقول الأساسية
        missing_essential = [field for field in essential_fields if field not in detected_columns]
        if missing_essential:
            missing_str = ", ".join(missing_essential)
            raise ValueError(f"Required columns missing: {missing_str}. Available columns: {[c for c in df.columns if not isinstance(c, datetime)]}")
        
        # بالنسبة للحقول غير الأساسية المفقودة، سننشئ أعمدة وهمية تحتوي على قيم افتراضية
        for field in other_fields:
            if field not in detected_columns:
                print(f"Warning: Creating default column for: {field}")
                dummy_column_name = f"__{field}__default"
                df[dummy_column_name] = 0  # إنشاء عمود فارغ (0 للقيم المالية)
                detected_columns[field] = dummy_column_name  # تعيين العمود الوهمي للحقل
        
        # Process each row
        salaries = []
        for idx, row in df.iterrows():
            try:
                # Skip completely empty rows
                if row.isnull().all():
                    continue
                
                # Get employee_id field
                emp_id_col = detected_columns['employee_id']
                emp_id = row[emp_id_col]
                
                # Skip rows with missing employee_id
                if pd.isna(emp_id):
                    print(f"Skipping row {idx+1} due to missing employee ID")
                    continue
                
                # Try to convert employee_id to integer
                try:
                    employee_id = int(emp_id)
                except (ValueError, TypeError):
                    # If not convertible to int, use as string (could be employee code)
                    employee_id = str(emp_id).strip()
                
                # Get basic_salary field
                basic_salary_col = detected_columns['basic_salary']
                basic_salary_val = row[basic_salary_col]
                
                # تعامل مع القيم المفقودة أو غير الرقمية للراتب الأساسي
                if pd.isna(basic_salary_val) or not isinstance(basic_salary_val, (int, float)):
                    print(f"Row {idx+1}: Using default value of 0 for basic salary")
                    basic_salary_val = 0
                
                basic_salary = float(basic_salary_val)
                
                # Get optional fields with default values
                allowances = 0.0
                deductions = 0.0
                bonus = 0.0
                notes = ''
                
                # Extract allowances if column exists
                if 'allowances' in detected_columns and not pd.isna(row[detected_columns['allowances']]):
                    try:
                        allowances = float(row[detected_columns['allowances']])
                    except (ValueError, TypeError):
                        allowances = 0.0
                
                # Extract deductions if column exists
                if 'deductions' in detected_columns and not pd.isna(row[detected_columns['deductions']]):
                    try:
                        deductions = float(row[detected_columns['deductions']])
                    except (ValueError, TypeError):
                        deductions = 0.0
                
                # Extract bonus if column exists
                if 'bonus' in detected_columns and not pd.isna(row[detected_columns['bonus']]):
                    try:
                        bonus = float(row[detected_columns['bonus']])
                    except (ValueError, TypeError):
                        bonus = 0.0
                
                # Extract notes if column exists
                if 'notes' in detected_columns and not pd.isna(row[detected_columns['notes']]):
                    notes = str(row[detected_columns['notes']])
                
                employee = None
                if isinstance(employee_id, int):
                    employee = Employee.query.get(employee_id)
                if not employee:
                    employee = Employee.query.filter_by(employee_id=str(employee_id)).first()

                # Calculate net salary (shared formula + nationality-based GOSI)
                salary_breakdown = compute_net_salary(
                    basic_salary=basic_salary,
                    allowances=allowances,
                    bonus=bonus,
                    deductions=deductions,
                    nationality=getattr(employee, 'nationality', None),
                    contract_type=getattr(employee, 'contract_type', None),
                    return_breakdown=True
                )
                net_salary = salary_breakdown['net_salary']
                total_deductions = salary_breakdown['total_deductions']
                
                # Create salary dictionary
                salary = {
                    'employee_id': employee_id,
                    'month': month,
                    'year': year,
                    'basic_salary': basic_salary,
                    'allowances': allowances,
                    'other_deductions': deductions,
                    'deductions': total_deductions,
                    'gosi_deduction': salary_breakdown.get('gosi_deduction', 0),
                    'bonus': bonus,
                    'net_salary': net_salary
                }
                
                if notes:
                    salary['notes'] = notes
                
                print(f"Processed salary for employee ID: {employee_id}")
                salaries.append(salary)
                
            except Exception as e:
                print(f"Error processing salary row {idx+1}: {str(e)}")
                # Continue to next row instead of failing the entire import
                continue
        
        if not salaries:
            raise ValueError("No valid salary records found in the Excel file")
            
        return salaries
    
    except Exception as e:
        import traceback
        print(f"Error parsing salary Excel: {str(e)}")
        print(traceback.format_exc())
        raise Exception(f"Error parsing salary Excel file: {str(e)}")

def generate_comprehensive_employee_report(db_session, department_id=None, employee_id=None, month=None, year=None):
    """
    إنشاء تقرير شامل للموظفين مع كامل تفاصيل الرواتب والبيانات
    
    Args:
        db_session: جلسة قاعدة البيانات
        department_id: معرف القسم (اختياري للتصفية)
        employee_id: معرف الموظف (اختياري للتصفية)
        month: الشهر (اختياري للتصفية)
        year: السنة (اختياري للتصفية)
        
    Returns:
        كائن BytesIO يحتوي على ملف Excel
    """
    try:
        from models import Employee, Department, Salary, Attendance, Document
        from sqlalchemy import func
        from datetime import datetime, timedelta
        from dateutil.relativedelta import relativedelta
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Color
        from openpyxl.utils import get_column_letter
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting.rule import Rule
        from openpyxl.chart import BarChart, Reference, Series
        from openpyxl.chart.marker import DataPoint
        from openpyxl.drawing.image import Image
        
        # استعلام الموظفين مع التصفية المطلوبة
        query = db_session.query(Employee)
        
        if department_id:
            query = query.filter(Employee.department_id == department_id)
        if employee_id:
            query = query.filter(Employee.id == employee_id)
            
        # الحصول على كل الموظفين المطلوبين
        employees = query.all()
        
        # البحث عن الرواتب المرتبطة بهذه الفترة
        salary_query = db_session.query(Salary).filter(Salary.employee_id.in_([e.id for e in employees]))
        if month:
            salary_query = salary_query.filter(Salary.month == month)
        if year:
            salary_query = salary_query.filter(Salary.year == year)
            
        salaries = salary_query.all()
        
        # تحديد الألوان والتنسيقات
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        total_row_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        total_row_font = Font(name="Arial", size=12, bold=True)
        
        normal_font = Font(name="Arial", size=11)
        highlight_font = Font(name="Arial", size=11, bold=True)
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        title_alignment = Alignment(horizontal='center', vertical='center')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='center', vertical='center')
        text_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # تنسيقات للخلايا المالية
        money_format = '#,##0.00 "ر.س"'
        percentage_format = '0.00%'
        date_format = 'yyyy-mm-dd'
        
        # إعداد البيانات المجمعة للموظفين
        employees_data = []
        salaries_by_employee = {}
        
        # تجميع الرواتب حسب الموظف
        for salary in salaries:
            if salary.employee_id not in salaries_by_employee:
                salaries_by_employee[salary.employee_id] = []
            salaries_by_employee[salary.employee_id].append(salary)
        
        # تجميع بيانات الموظفين مع الرواتب
        for employee in employees:
            emp_salaries = salaries_by_employee.get(employee.id, [])
            
            # حساب متوسط الراتب وأعلى وأدنى راتب
            basic_salaries = [s.basic_salary for s in emp_salaries] if emp_salaries else [0]
            net_salaries = [s.net_salary for s in emp_salaries] if emp_salaries else [0]
            
            avg_basic = sum(basic_salaries) / len(basic_salaries) if basic_salaries else 0
            avg_net = sum(net_salaries) / len(net_salaries) if net_salaries else 0
            max_net = max(net_salaries) if net_salaries else 0
            min_net = min(net_salaries) if net_salaries else 0
            
            # تجميع معلومات الموظف
            emp_data = {
                'معرف': employee.id,
                'رقم الموظف': employee.employee_id,
                'الاسم': employee.name,
                'القسم': ', '.join([dept.name for dept in employee.departments]) if employee.departments else 'بدون قسم',
                'الوظيفة': employee.job_title or '',
                'تاريخ التعيين': getattr(employee, 'hire_date', None) or getattr(employee, 'join_date', None),
                'الجنسية': employee.nationality or '',
                'الهاتف': getattr(employee, 'phone', None) or getattr(employee, 'mobile', None) or '',
                'البريد الإلكتروني': employee.email or '',
                'الرقم الوطني/الإقامة': employee.national_id or '',
                'الحالة': employee.status or '',
                'متوسط الراتب الأساسي': avg_basic,
                'متوسط صافي الراتب': avg_net,
                'أعلى راتب': max_net,
                'أدنى راتب': min_net,
                'عدد الرواتب المسجلة': len(emp_salaries),
                'الملاحظات': getattr(employee, 'notes', None) or ''
            }
            
            # إضافة تفاصيل آخر راتب
            if emp_salaries:
                # ترتيب الرواتب حسب السنة والشهر (تنازلياً)
                sorted_salaries = sorted(emp_salaries, key=lambda s: (s.year, s.month), reverse=True)
                latest_salary = sorted_salaries[0]
                
                emp_data.update({
                    'آخر راتب - الشهر': latest_salary.month,
                    'آخر راتب - السنة': latest_salary.year,
                    'آخر راتب - الأساسي': latest_salary.basic_salary,
                    'آخر راتب - البدلات': latest_salary.allowances,
                    'آخر راتب - الخصومات': latest_salary.deductions,
                    'آخر راتب - التأمينات (GOSI)': _get_gosi_amount_for_salary_record(latest_salary),
                    'آخر راتب - المكافآت': latest_salary.bonus,
                    'آخر راتب - الصافي': latest_salary.net_salary
                })
            
            employees_data.append(emp_data)
        
        # إنشاء ملف Excel باستخدام openpyxl
        output = BytesIO()
        with pd.ExcelWriter(path=output, engine='openpyxl') as writer:
            # ======= ورقة ملخص الموظفين =======
            emp_df = pd.DataFrame(employees_data)
            
            # ترتيب الأعمدة للتقرير الشامل
            columns_order = [
                'معرف', 'رقم الموظف', 'الاسم', 'القسم', 'الوظيفة', 'تاريخ التعيين', 
                'الجنسية', 'الرقم الوطني/الإقامة', 'الهاتف', 'البريد الإلكتروني', 'الحالة',
                'متوسط الراتب الأساسي', 'متوسط صافي الراتب', 'أعلى راتب', 'أدنى راتب', 
                'عدد الرواتب المسجلة',
                'آخر راتب - الشهر', 'آخر راتب - السنة', 'آخر راتب - الأساسي', 
                'آخر راتب - البدلات', 'آخر راتب - الخصومات', 'آخر راتب - التأمينات (GOSI)', 'آخر راتب - المكافآت', 
                'آخر راتب - الصافي', 'الملاحظات'
            ]
            
            # استبعاد الأعمدة غير الموجودة
            actual_columns = [col for col in columns_order if col in emp_df.columns]
            emp_df = emp_df[actual_columns]
            
            # كتابة البيانات إلى الملف
            emp_df.to_excel(writer, sheet_name='ملخص الموظفين', index=False, startrow=2)
            
            # الحصول على ورقة العمل وتنسيقها
            summary_sheet = writer.sheets['ملخص الموظفين']
            ExcelStyles.apply_rtl(summary_sheet, engine="openpyxl")
            
            # إضافة عنوان للتقرير
            summary_sheet.merge_cells(f'A1:{get_column_letter(len(actual_columns))}1')
            title_cell = summary_sheet.cell(1, 1)
            title_cell.value = "التقرير الشامل للموظفين مع تفاصيل الرواتب"
            title_cell.font = Font(name="Arial", size=16, bold=True, color="1F4E78")
            title_cell.alignment = title_alignment
            
            # تنسيق العناوين
            for col_idx, column_name in enumerate(actual_columns, 1):
                cell = summary_sheet.cell(3, col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
                # ضبط عرض العمود
                column_width = _calculate_column_width(emp_df, column_name, padding=4)
                column_letter = get_column_letter(col_idx)
                summary_sheet.column_dimensions[column_letter].width = column_width
            
            # تنسيق البيانات
            for row_idx, (_, row) in enumerate(emp_df.iterrows(), 1):
                for col_idx, column_name in enumerate(actual_columns, 1):
                    cell = summary_sheet.cell(row_idx + 3, col_idx)
                    value = row[column_name]
                    cell.value = value
                    
                    # تنسيق خاص لأنواع البيانات المختلفة
                    if 'راتب' in column_name:
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
                    elif 'تاريخ' in column_name and value:
                        cell.number_format = date_format
                        cell.alignment = cell_alignment
                    else:
                        cell.alignment = text_alignment
                    
                    # تنسيق صفوف بديلة
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    
                    cell.font = normal_font
                    cell.border = thin_border
            
            # ======= ورقة تفاصيل الرواتب لكل موظف =======
            # تحضير بيانات الرواتب مفصلة حسب الموظف
            all_salary_data = []
            
            for employee in employees:
                emp_salaries = salaries_by_employee.get(employee.id, [])
                
                for salary in emp_salaries:
                    all_salary_data.append({
                        'معرف الموظف': employee.id,
                        'رقم الموظف': employee.employee_id,
                        'اسم الموظف': employee.name,
                        'القسم': ', '.join([dept.name for dept in employee.departments]) if employee.departments else 'بدون قسم',
                        'الشهر': salary.month,
                        'السنة': salary.year,
                        'الراتب الأساسي': salary.basic_salary,
                        'البدلات': salary.allowances,
                        'الخصومات': salary.deductions,
                        'التأمينات (GOSI)': _get_gosi_amount_for_salary_record(salary),
                        'المكافآت': salary.bonus,
                        'صافي الراتب': salary.net_salary,
                        'ملاحظات': salary.notes or ''
                    })
            
            if all_salary_data:
                # إنشاء DataFrame للرواتب
                salary_df = pd.DataFrame(all_salary_data)
                
                # ترتيب البيانات حسب القسم، الموظف، السنة، الشهر
                salary_df = salary_df.sort_values(by=['القسم', 'اسم الموظف', 'السنة', 'الشهر'], ascending=[True, True, False, False])
                
                # كتابة البيانات إلى ورقة عمل جديدة
                salary_df.to_excel(writer, sheet_name='تفاصيل الرواتب', index=False, startrow=2)
                
                # تنسيق ورقة تفاصيل الرواتب
                salary_sheet = writer.sheets['تفاصيل الرواتب']
                ExcelStyles.apply_rtl(salary_sheet, engine="openpyxl")
                
                # إضافة عنوان
                salary_sheet.merge_cells(f'A1:{get_column_letter(len(salary_df.columns))}1')
                title_cell = salary_sheet.cell(1, 1)
                title_cell.value = "تفاصيل رواتب الموظفين"
                title_cell.font = Font(name="Arial", size=16, bold=True, color="1F4E78")
                title_cell.alignment = title_alignment
                
                # تنسيق العناوين
                for col_idx, column_name in enumerate(salary_df.columns, 1):
                    cell = salary_sheet.cell(3, col_idx)
                    cell.value = column_name
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    
                    # ضبط عرض العمود
                    column_width = _calculate_column_width(salary_df, column_name, padding=4)
                    column_letter = get_column_letter(col_idx)
                    salary_sheet.column_dimensions[column_letter].width = column_width
                
                # تجميع الصفوف حسب الموظف بألوان مختلفة
                current_employee = None
                color_index = 0
                colors = ["E6F2FF", "F2F2F2"]  # ألوان التناوب
                
                # تنسيق البيانات
                for row_idx, (_, row) in enumerate(salary_df.iterrows(), 1):
                    # تغيير اللون عند تغيير الموظف
                    if current_employee != row['اسم الموظف']:
                        current_employee = row['اسم الموظف']
                        color_index = (color_index + 1) % 2
                    
                    row_fill = PatternFill(start_color=colors[color_index], end_color=colors[color_index], fill_type="solid")
                    
                    for col_idx, column_name in enumerate(salary_df.columns, 1):
                        cell = salary_sheet.cell(row_idx + 3, col_idx)
                        cell.value = row[column_name]
                        
                        # تنسيق خاص لأنواع البيانات المختلفة
                        if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'التأمينات (GOSI)', 'المكافآت', 'صافي الراتب']:
                            cell.number_format = money_format
                            cell.alignment = cell_alignment
                        else:
                            cell.alignment = text_alignment
                        
                        cell.fill = row_fill
                        cell.font = normal_font
                        cell.border = thin_border
                
                # إضافة صف للمجاميع في نهاية الجدول
                total_row_idx = len(salary_df) + 4
                salary_sheet.cell(total_row_idx, 1).value = "المجموع الكلي"
                salary_sheet.cell(total_row_idx, 1).font = total_row_font
                salary_sheet.cell(total_row_idx, 1).alignment = text_alignment
                salary_sheet.cell(total_row_idx, 1).fill = total_row_fill
                salary_sheet.cell(total_row_idx, 1).border = thick_border
                
                # دمج خلايا المجموع
                merge_cols = 6  # دمج الخلايا الأولى للمجموع
                salary_sheet.merge_cells(f'A{total_row_idx}:{get_column_letter(merge_cols)}{total_row_idx}')
                
                # تنسيق وحساب المجاميع
                for col_idx, column_name in enumerate(salary_df.columns, 1):
                    cell = salary_sheet.cell(total_row_idx, col_idx)
                    cell.font = total_row_font
                    cell.fill = total_row_fill
                    cell.border = thick_border
                    
                    if col_idx <= merge_cols:
                        continue  # تخطي الخلايا المدمجة
                    
                    # حساب المجاميع للأعمدة المالية
                    if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'التأمينات (GOSI)', 'المكافآت', 'صافي الراتب']:
                        col_letter = get_column_letter(col_idx)
                        cell.value = f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})"
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
                
                # إضافة رسم بياني للرواتب حسب القسم
                try:
                    chart_sheet = writer.book.create_sheet(title="الرسوم البيانية")
                    ExcelStyles.apply_rtl(chart_sheet, engine="openpyxl")
                    
                    # إعداد بيانات الرسم البياني - متوسط الراتب حسب القسم
                    dept_avg_salary = salary_df.groupby('القسم')['صافي الراتب'].mean().reset_index()
                    dept_avg_salary.to_excel(writer, sheet_name="الرسوم البيانية", startrow=1, startcol=1, index=False)
                    
                    chart_sheet.cell(1, 1).value = "متوسط الرواتب حسب القسم"
                    chart_sheet.cell(1, 1).font = Font(name="Arial", size=14, bold=True)
                    
                    # إنشاء الرسم البياني
                    chart = BarChart()
                    chart.title = "متوسط الرواتب حسب القسم"
                    chart.y_axis.title = "متوسط الراتب (ر.س)"
                    chart.x_axis.title = "القسم"
                    
                    # تحديد نطاق البيانات
                    data = Reference(chart_sheet, min_col=3, min_row=2, max_row=2+len(dept_avg_salary))
                    cats = Reference(chart_sheet, min_col=2, min_row=3, max_row=2+len(dept_avg_salary))
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    
                    # إضافة الرسم البياني إلى الورقة
                    chart_sheet.add_chart(chart, "E5")
                    
                except Exception as chart_error:
                    print(f"حدث خطأ أثناء إنشاء الرسم البياني: {chart_error}")
            
            # ======= ورقة معلومات التقرير =======
            # إنشاء ورقة معلومات التقرير
            info_data = []
            
            # إضافة معلومات عامة
            info_data.append({
                'المعلومة': 'تاريخ التصدير',
                'القيمة': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # إضافة معلومات حول التصفية
            if department_id:
                dept = db_session.query(Department).get(department_id)
                info_data.append({
                    'المعلومة': 'تصفية حسب القسم',
                    'القيمة': dept.name if dept else department_id
                })
            
            if employee_id:
                emp = db_session.query(Employee).get(employee_id)
                info_data.append({
                    'المعلومة': 'تصفية حسب الموظف',
                    'القيمة': emp.name if emp else employee_id
                })
            
            if month:
                info_data.append({
                    'المعلومة': 'تصفية حسب الشهر',
                    'القيمة': month
                })
            
            if year:
                info_data.append({
                    'المعلومة': 'تصفية حسب السنة',
                    'القيمة': year
                })
            
            # إضافة إحصائيات عامة
            info_data.append({
                'المعلومة': 'إجمالي عدد الموظفين',
                'القيمة': len(employees)
            })
            
            info_data.append({
                'المعلومة': 'إجمالي عدد الرواتب المسجلة',
                'القيمة': len(salaries)
            })
            
            # حساب متوسطات الرواتب
            if salaries:
                avg_basic = sum(s.basic_salary for s in salaries) / len(salaries)
                avg_net = sum(s.net_salary for s in salaries) / len(salaries)
                
                info_data.append({
                    'المعلومة': 'متوسط الراتب الأساسي',
                    'القيمة': avg_basic
                })
                
                info_data.append({
                    'المعلومة': 'متوسط صافي الراتب',
                    'القيمة': avg_net
                })
                
                info_data.append({
                    'المعلومة': 'إجمالي مصاريف الرواتب',
                    'القيمة': sum(s.net_salary for s in salaries)
                })

                info_data.append({
                    'المعلومة': 'إجمالي التأمينات (GOSI)',
                    'القيمة': sum(_get_gosi_amount_for_salary_record(s) for s in salaries)
                })
            
            # إنشاء DataFrame للمعلومات
            info_df = pd.DataFrame(info_data)
            info_df.to_excel(writer, sheet_name='معلومات التقرير', index=False, startrow=2)
            
            # تنسيق ورقة المعلومات
            info_sheet = writer.sheets['معلومات التقرير']
            ExcelStyles.apply_rtl(info_sheet, engine="openpyxl")
            
            # إضافة عنوان للورقة
            info_sheet.merge_cells('A1:B1')
            info_sheet.cell(1, 1).value = "معلومات التقرير الشامل"
            info_sheet.cell(1, 1).font = Font(name="Arial", size=16, bold=True, color="1F4E78")
            info_sheet.cell(1, 1).alignment = title_alignment
            
            # تنسيق العناوين
            for col_idx, col_name in enumerate(info_df.columns, 1):
                info_sheet.cell(3, col_idx).value = col_name
                info_sheet.cell(3, col_idx).font = header_font
                info_sheet.cell(3, col_idx).fill = header_fill
                info_sheet.cell(3, col_idx).alignment = header_alignment
                info_sheet.cell(3, col_idx).border = thin_border
                
                # ضبط عرض العمود
                column_width = _calculate_column_width(info_df, col_name, padding=4)
                column_letter = get_column_letter(col_idx)
                info_sheet.column_dimensions[column_letter].width = column_width
            
            # تنسيق البيانات
            for row_idx, (_, row) in enumerate(info_df.iterrows(), 1):
                for col_idx, col_name in enumerate(info_df.columns, 1):
                    cell = info_sheet.cell(row_idx + 3, col_idx)
                    cell.value = row[col_name]
                    
                    # تنسيق خاص للقيم المالية
                    if 'متوسط' in row['المعلومة'] or 'إجمالي مصاريف' in row['المعلومة'] or 'التأمينات (GOSI)' in row['المعلومة']:
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
                    else:
                        cell.alignment = text_alignment
                    
                    # تنسيق صفوف بديلة
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    
                    cell.font = normal_font
                    cell.border = thin_border
            
            # تعيين الصفحة الأولى كصفحة نشطة
            writer.book.active = writer.book.worksheets[0]
        
        output.seek(0)
        return output
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise Exception(f"خطأ في إنشاء التقرير الشامل: {str(e)}")

def generate_employee_salary_simple_excel(db_session, month=None, year=None, department_id=None):
    """
    إنشاء ملف Excel بسيط وواضح لبيانات الموظفين مع تفاصيل الرواتب
    مع تنسيق احترافي للجداول وألوان متناوبة للصفوف وتنسيق مالي للأرقام
    
    Args:
        db_session: جلسة قاعدة البيانات
        month: الشهر المطلوب (اختياري)
        year: السنة المطلوبة (اختياري)
        department_id: معرّف القسم (اختياري)
        
    Returns:
        كائن BytesIO يحتوي على ملف Excel
    """
    try:
        from models import Employee, Department, Salary
        from datetime import datetime
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # تحديد الألوان والتنسيقات
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        normal_font = Font(name="Arial", size=11)
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # تنسيقات للخلايا المالية
        money_format = '#,##0.00 "ر.س"'
        
        # استعلام الرواتب مع بيانات الموظفين بتحديد صريح للعلاقات
        query = db_session.query(Salary).\
            join(Employee, Salary.employee_id == Employee.id).\
            join(Department, Employee.department_id == Department.id)
        
        # تطبيق الفلاتر
        if department_id:
            query = query.filter(Department.id == department_id)
        if month:
            query = query.filter(Salary.month == month)
        if year:
            query = query.filter(Salary.year == year)
            
        # ترتيب البيانات حسب القسم، اسم الموظف، السنة والشهر
        query = query.order_by(Department.name, Employee.name, Salary.year.desc(), Salary.month.desc())
        
        # الحصول على النتائج
        results = query.all()
        
        # جمع البيانات في قائمة
        employee_data = []
        
        for salary in results:
            employee = salary.employee
            departments_list = employee.departments if employee.departments else []
            department_name = ', '.join([dept.name for dept in departments_list]) if departments_list else 'بدون قسم'
            
            data = {
                'اسم الموظف': employee.name,
                'رقم الموظف': employee.employee_id,
                'رقم الهوية': employee.national_id or '',
                'القسم': department_name,
                'الشهر': salary.month,
                'السنة': salary.year,
                'الراتب الأساسي': salary.basic_salary,
                'البدلات': salary.allowances,
                'الخصومات': salary.deductions,
                'المكافآت': salary.bonus,
                'صافي الراتب': salary.net_salary,
                'ملاحظات': salary.notes or ''
            }
            
            employee_data.append(data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        
        with pd.ExcelWriter(path=output, engine='openpyxl') as writer:
            # إنشاء DataFrame
            df = pd.DataFrame(employee_data)
            
            # ترتيب الأعمدة بالشكل المطلوب
            columns_order = [
                'اسم الموظف', 'رقم الموظف', 'رقم الهوية', 'القسم',
                'الشهر', 'السنة', 'الراتب الأساسي', 'البدلات',
                'الخصومات', 'المكافآت', 'صافي الراتب', 'ملاحظات'
            ]
            
            # ترتيب الأعمدة حسب الترتيب المحدد
            df = df[columns_order]
            
            # كتابة البيانات إلى الملف
            df.to_excel(writer, sheet_name='بيانات الموظفين والرواتب', index=False)
            
            # الحصول على ورقة العمل وتنسيقها
            sheet = writer.sheets['بيانات الموظفين والرواتب']
            ExcelStyles.apply_rtl(sheet, engine="openpyxl")
            
            # تنسيق العناوين
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = sheet.cell(1, col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # ضبط عرض العمود
                column_letter = get_column_letter(col_idx)
                max_length = _calculate_column_width(df, col_name, padding=2)
                sheet.column_dimensions[column_letter].width = max_length
            
            # تنسيق البيانات
            for row_idx, _ in enumerate(df.iterrows(), 2):  # بدء من الصف 2 (بعد العناوين)
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = sheet.cell(row_idx, col_idx)
                    
                    # تنسيق الخلايا المالية
                    if col_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الراتب']:
                        cell.number_format = money_format
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = right_alignment
                    
                    # تنسيق عام
                    cell.font = normal_font
                    cell.border = thin_border
                    
                    # تلوين الصفوف بالتناوب
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # إضافة صف المجموع في النهاية
            total_row = len(df) + 2
            
            # إضافة نص "المجموع" في أول خلية
            sheet.cell(total_row, 1).value = "المجموع"
            sheet.cell(total_row, 1).font = Font(name="Arial", size=12, bold=True)
            sheet.cell(total_row, 1).alignment = right_alignment
            sheet.cell(total_row, 1).border = thin_border
            sheet.cell(total_row, 1).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # دمج الخلايا من العمود 1 إلى العمود 6
            for col_idx in range(2, 7):
                cell = sheet.cell(total_row, col_idx)
                cell.border = thin_border
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
            
            # حساب المجاميع للأعمدة المالية
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الراتب']:
                    col_letter = get_column_letter(col_idx)
                    cell = sheet.cell(total_row, col_idx)
                    cell.value = f"=SUM({col_letter}2:{col_letter}{total_row-1})"
                    cell.font = Font(name="Arial", size=12, bold=True)
                    cell.number_format = money_format
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                elif col_idx > 6:
                    # تنسيق باقي الخلايا في صف المجموع
                    cell = sheet.cell(total_row, col_idx)
                    cell.border = thin_border
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # إضافة معلومات الفلترة في أعلى الورقة
            info_row = sheet.max_row + 2
            
            # إضافة عنوان المعلومات
            sheet.cell(info_row, 1).value = "معلومات التقرير:"
            sheet.cell(info_row, 1).font = Font(name="Arial", size=12, bold=True)
            
            # إضافة تفاصيل الفلترة
            info_row += 1
            sheet.cell(info_row, 1).value = "تاريخ التصدير:"
            sheet.cell(info_row, 2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            info_row += 1
            filter_text = []
            if month:
                filter_text.append(f"الشهر: {month}")
            if year:
                filter_text.append(f"السنة: {year}")
            if department_id:
                dept = db_session.query(Department).get(department_id)
                if dept:
                    filter_text.append(f"القسم: {dept.name}")
            
            sheet.cell(info_row, 1).value = "الفلاتر المطبقة:"
            sheet.cell(info_row, 2).value = " | ".join(filter_text) if filter_text else "كافة البيانات"
            
            info_row += 1
            sheet.cell(info_row, 1).value = "عدد السجلات:"
            sheet.cell(info_row, 2).value = len(df)
        
        output.seek(0)
        return output
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise Exception(f"خطأ في إنشاء ملف Excel: {str(e)}")

def generate_salary_excel(salaries, filter_description=None):
    """
    إنشاء ملف Excel من بيانات الرواتب مع تنظيم وتجميع حسب القسم وتنسيق ممتاز
    
    Args:
        salaries: قائمة كائنات Salary 
        filter_description: وصف مرشحات البحث المستخدمة (اختياري)
        
    Returns:
        كائن BytesIO يحتوي على ملف Excel
    """
    try:
        from datetime import datetime
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Color
        from openpyxl.utils import get_column_letter
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting.rule import Rule
        from openpyxl.drawing.image import Image
        
        # تحديد الألوان والتنسيقات
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        total_row_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        total_row_font = Font(name="Arial", size=12, bold=True)
        
        normal_font = Font(name="Arial", size=11)
        highlight_font = Font(name="Arial", size=11, bold=True)
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        title_alignment = Alignment(horizontal='center', vertical='center')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='center', vertical='center')
        text_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # تنسيقات للخلايا المالية
        money_format = '#,##0.00 "ر.س"'
        percentage_format = '0.00%'
        
        # تجميع البيانات حسب القسم
        departments_data = {}
        for salary in salaries:
            dept_name = salary.employee.department.name if salary.employee.department else 'بدون قسم'
            if dept_name not in departments_data:
                departments_data[dept_name] = []
            
            # إضافة بيانات الراتب إلى القسم المناسب
            departments_data[dept_name].append({
                'معرف': salary.id,
                'اسم الموظف': salary.employee.name,
                'رقم الموظف': salary.employee.employee_id,
                'رقم الهوية': salary.employee.national_id or '',
                'الوظيفة': salary.employee.job_title or '',
                'القسم': dept_name,
                'الشهر': salary.month,
                'السنة': salary.year,
                'الراتب الأساسي': salary.basic_salary,
                'البدلات': salary.allowances,
                'الخصومات': salary.deductions,
                'التأمينات (GOSI)': _get_gosi_amount_for_salary_record(salary),
                'المكافآت': salary.bonus,
                'أيام الحضور': salary.present_days if salary.attendance_calculated else '-',
                'أيام الغياب': salary.absent_days if salary.attendance_calculated else '-',
                'خصم الغياب': salary.attendance_deduction if salary.attendance_calculated else 0,
                'صافي الراتب': salary.net_salary,
                'ملاحظات': salary.notes or ''
            })
        
        # إنشاء ملف Excel باستخدام openpyxl
        output = BytesIO()
        with pd.ExcelWriter(path=output, engine='openpyxl') as writer:
            # بيانات الملخص
            summary_data = []
            total_salaries = 0
            total_basic = 0
            total_allowances = 0
            total_deductions = 0
            total_gosi = 0
            total_bonus = 0
            total_net = 0
            
            # حساب المجاميع لكل قسم
            for dept_name, dept_salaries in departments_data.items():
                dept_count = len(dept_salaries)
                dept_basic_sum = sum(s['الراتب الأساسي'] for s in dept_salaries)
                dept_allowances_sum = sum(s['البدلات'] for s in dept_salaries)
                dept_deductions_sum = sum(s['الخصومات'] for s in dept_salaries)
                dept_gosi_sum = sum(s['التأمينات (GOSI)'] for s in dept_salaries)
                dept_bonus_sum = sum(s['المكافآت'] for s in dept_salaries)
                dept_net_sum = sum(s['صافي الراتب'] for s in dept_salaries)
                
                summary_data.append({
                    'القسم': dept_name,
                    'عدد الموظفين': dept_count,
                    'إجمالي الرواتب الأساسية': dept_basic_sum,
                    'إجمالي البدلات': dept_allowances_sum,
                    'إجمالي الخصومات': dept_deductions_sum,
                    'إجمالي التأمينات (GOSI)': dept_gosi_sum,
                    'إجمالي المكافآت': dept_bonus_sum,
                    'إجمالي صافي الرواتب': dept_net_sum
                })
                
                # تحديث المجاميع الكلية
                total_salaries += dept_count
                total_basic += dept_basic_sum
                total_allowances += dept_allowances_sum
                total_deductions += dept_deductions_sum
                total_gosi += dept_gosi_sum
                total_bonus += dept_bonus_sum
                total_net += dept_net_sum
            
            # إضافة صف المجموع الكلي
            summary_data.append({
                'القسم': 'الإجمالي',
                'عدد الموظفين': total_salaries,
                'إجمالي الرواتب الأساسية': total_basic,
                'إجمالي البدلات': total_allowances,
                'إجمالي الخصومات': total_deductions,
                'إجمالي التأمينات (GOSI)': total_gosi,
                'إجمالي المكافآت': total_bonus,
                'إجمالي صافي الرواتب': total_net
            })
            
            # إنشاء DataFrame للملخص
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='ملخص الرواتب', index=False)
            
            # تنسيق ورقة الملخص
            summary_sheet = writer.sheets['ملخص الرواتب']
            ExcelStyles.apply_rtl(summary_sheet, engine="openpyxl")
            
            # إضافة عنوان للتقرير
            summary_sheet.merge_cells(f"A1:{get_column_letter(len(summary_df.columns))}1")
            summary_sheet.cell(1, 1).value = "تقرير ملخص الرواتب"
            summary_sheet.cell(1, 1).font = Font(name="Arial", size=16, bold=True, color="1F4E78")
            summary_sheet.cell(1, 1).alignment = title_alignment
            
            # إضافة معلومات الفلترة تحت العنوان
            if filter_description:
                summary_sheet.merge_cells(f"A2:{get_column_letter(len(summary_df.columns))}2")
                summary_sheet.cell(2, 1).value = "مرشحات البحث: " + " - ".join(filter_description)
                summary_sheet.cell(2, 1).font = Font(name="Arial", size=12, italic=True)
                summary_sheet.cell(2, 1).alignment = title_alignment
                
                # ضبط العنوان ليبدأ من الصف الثالث
                title_row_offset = 3
            else:
                # ضبط العنوان ليبدأ من الصف الثاني
                title_row_offset = 2
            
            # الحصول على عدد الصفوف والأعمدة في البيانات
            num_rows = len(summary_data) + 1  # +1 للعنوان
            num_cols = len(summary_df.columns)
            
            # تنسيق الترويسات
            for col_idx, column_name in enumerate(summary_df.columns, 1):
                cell = summary_sheet.cell(title_row_offset, col_idx)
                cell.value = column_name
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
                # ضبط عرض العمود
                column_width = _calculate_column_width(summary_df, column_name, padding=4)
                column_letter = get_column_letter(col_idx)
                summary_sheet.column_dimensions[column_letter].width = column_width
            
            # تنسيق البيانات
            for row_idx, row in enumerate(summary_data, 1):
                is_total_row = row_idx == len(summary_data)
                
                for col_idx, (column_name, value) in enumerate(zip(summary_df.columns, row.values()), 1):
                    cell = summary_sheet.cell(title_row_offset + row_idx, col_idx)
                    cell.value = value
                    
                    # تنسيق خاص للصف الأخير (الإجمالي)
                    if is_total_row:
                        cell.font = total_row_font
                        cell.fill = total_row_fill
                        cell.border = thick_border
                    else:
                        cell.font = normal_font
                        cell.border = thin_border
                    
                    # تنسيق خاص للأعمدة المالية
                    if 'إجمالي' in column_name or column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'التأمينات (GOSI)', 'المكافآت', 'صافي الراتب']:
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
                    else:
                        cell.alignment = text_alignment
            
            # إنشاء ورقة عمل مفصلة لكل قسم
            for dept_name, dept_salaries in departments_data.items():
                dept_df = pd.DataFrame(dept_salaries)
                
                # ترتيب الأعمدة بشكل منطقي
                ordered_columns = [
                    'معرف', 'اسم الموظف', 'رقم الموظف', 'رقم الهوية', 'الوظيفة', 'القسم',
                    'الشهر', 'السنة', 'الراتب الأساسي', 'البدلات', 'الخصومات', 'التأمينات (GOSI)',
                    'المكافآت', 'أيام الحضور', 'أيام الغياب', 'خصم الغياب', 'صافي الراتب', 'ملاحظات'
                ]
                
                # إعادة ترتيب الأعمدة واستبعاد الأعمدة غير الموجودة
                actual_columns = [col for col in ordered_columns if col in dept_df.columns]
                dept_df = dept_df[actual_columns]
                
                # إضافة ورقة العمل للقسم
                sheet_name = dept_name[:31]  # تقليص اسم الورقة إلى 31 حرف (أقصى طول مسموح في Excel)
                dept_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)  # بدء من الصف الثالث لإتاحة مساحة للعنوان
                
                # الحصول على sheet وتنسيق عنوان القسم
                dept_sheet = writer.sheets[sheet_name]
                ExcelStyles.apply_rtl(dept_sheet, engine="openpyxl")
                
                # دمج الخلايا للعنوان
                num_dept_cols = len(actual_columns)
                dept_sheet.merge_cells(f'A1:{get_column_letter(num_dept_cols)}1')
                title_cell = dept_sheet.cell(1, 1)
                title_cell.value = f"تفاصيل رواتب قسم {dept_name}"
                title_cell.font = Font(name="Arial", size=16, bold=True, color="1F4E78")
                title_cell.alignment = title_alignment
                
                # تنسيق الترويسات
                for col_idx, column_name in enumerate(actual_columns, 1):
                    cell = dept_sheet.cell(3, col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    
                    # ضبط عرض العمود
                    column_width = _calculate_column_width(dept_df, column_name, padding=4)
                    column_letter = get_column_letter(col_idx)
                    dept_sheet.column_dimensions[column_letter].width = column_width
                
                # تنسيق بيانات الموظفين
                for row_idx in range(len(dept_df)):
                    for col_idx, column_name in enumerate(actual_columns, 1):
                        cell = dept_sheet.cell(row_idx + 4, col_idx)  # +4 للترويسة والعنوان
                        
                        # تنسيق خاص للخلايا المالية
                        if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'التأمينات (GOSI)', 'المكافآت', 'خصم الغياب', 'صافي الراتب']:
                            cell.number_format = money_format
                            cell.alignment = cell_alignment
                        elif column_name in ['أيام الحضور', 'أيام الغياب']:
                            cell.alignment = cell_alignment
                        else:
                            cell.alignment = text_alignment
                        
                        cell.font = normal_font
                        cell.border = thin_border
                
                # إضافة صف للمجاميع في نهاية جدول القسم
                total_row_idx = len(dept_df) + 4
                dept_sheet.cell(total_row_idx, 1).value = "المجموع"
                dept_sheet.cell(total_row_idx, 1).font = total_row_font
                dept_sheet.cell(total_row_idx, 1).alignment = text_alignment
                
                # تنسيق صف المجموع وحساب المجاميع للأعمدة المالية
                for col_idx, column_name in enumerate(actual_columns, 1):
                    cell = dept_sheet.cell(total_row_idx, col_idx)
                    cell.font = total_row_font
                    cell.fill = total_row_fill
                    cell.border = thick_border
                    
                    # حساب المجاميع للأعمدة المالية
                    if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'التأمينات (GOSI)', 'المكافآت', 'صافي الراتب']:
                        col_letter = get_column_letter(col_idx)
                        cell.value = f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})"
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
                
                # إضافة قواعد تنسيق شرطية للخلايا
                # تلوين الخلايا ذات القيم السالبة باللون الأحمر
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                red_font = Font(color="9C0006")
                
                # تلوين الخلايا ذات القيم الموجبة باللون الأخضر
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                green_font = Font(color="006100")
                
                # تطبيق التنسيق الشرطي على أعمدة المال
                for col_idx, column_name in enumerate(actual_columns, 1):
                    if column_name in ['الراتب الأساسي', 'البدلات', 'المكافآت', 'صافي الراتب']:
                        col_letter = get_column_letter(col_idx)
                        
                        # قاعدة للقيم العالية (أعلى من المتوسط)
                        high_rule = Rule(
                            type="cellIs",
                            operator="greaterThan",
                            formula=[f"AVERAGE({col_letter}4:{col_letter}{total_row_idx-1})"],
                            dxf=DifferentialStyle(fill=green_fill, font=green_font)
                        )
                        dept_sheet.conditional_formatting.add(f"{col_letter}4:{col_letter}{total_row_idx-1}", high_rule)
                    
                    elif column_name == 'الخصومات':
                        col_letter = get_column_letter(col_idx)
                        
                        # قاعدة للخصومات العالية
                        high_deduction_rule = Rule(
                            type="cellIs",
                            operator="greaterThan",
                            formula=[f"AVERAGE({col_letter}4:{col_letter}{total_row_idx-1})"],
                            dxf=DifferentialStyle(fill=red_fill, font=red_font)
                        )
                        dept_sheet.conditional_formatting.add(f"{col_letter}4:{col_letter}{total_row_idx-1}", high_deduction_rule)
            
            # إنشاء ورقة لجميع الرواتب
            all_data = []
            for dept_salaries in departments_data.values():
                all_data.extend(dept_salaries)
            
            if all_data:
                all_df = pd.DataFrame(all_data)
                
                # ترتيب الأعمدة بشكل منطقي
                all_columns = [col for col in ordered_columns if col in all_df.columns]
                all_df = all_df[all_columns]
                
                # إضافة ورقة كل الرواتب
                all_df.to_excel(writer, sheet_name='جميع الرواتب', index=False, startrow=2)
                
                # تنسيق ورقة كل الرواتب
                all_sheet = writer.sheets['جميع الرواتب']
                ExcelStyles.apply_rtl(all_sheet, engine="openpyxl")
                
                # دمج الخلايا للعنوان
                num_all_cols = len(all_columns)
                all_sheet.merge_cells(f'A1:{get_column_letter(num_all_cols)}1')
                title_cell = all_sheet.cell(1, 1)
                title_cell.value = "قائمة كاملة بالرواتب"
                title_cell.font = Font(name="Arial", size=16, bold=True, color="1F4E78")
                title_cell.alignment = title_alignment
                
                # تنسيق الترويسات
                for col_idx, column_name in enumerate(all_columns, 1):
                    cell = all_sheet.cell(3, col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    
                    # ضبط عرض العمود
                    column_width = _calculate_column_width(all_df, column_name, padding=4)
                    column_letter = get_column_letter(col_idx)
                    all_sheet.column_dimensions[column_letter].width = column_width
                
                # تنسيق كافة البيانات
                for row_idx in range(len(all_df)):
                    for col_idx, column_name in enumerate(all_columns, 1):
                        cell = all_sheet.cell(row_idx + 4, col_idx)  # +4 للترويسة والعنوان
                        
                        # تنسيق خاص للخلايا المالية
                        if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'التأمينات (GOSI)', 'المكافآت', 'خصم الغياب', 'صافي الراتب']:
                            cell.number_format = money_format
                            cell.alignment = cell_alignment
                        elif column_name in ['أيام الحضور', 'أيام الغياب']:
                            cell.alignment = cell_alignment
                        else:
                            cell.alignment = text_alignment
                        
                        # تمييز الصفوف بألوان متناوبة
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        
                        cell.font = normal_font
                        cell.border = thin_border
                
                # إضافة صف للمجاميع في نهاية الجدول
                total_row_idx = len(all_df) + 4
                all_sheet.cell(total_row_idx, 1).value = "المجموع الكلي"
                all_sheet.cell(total_row_idx, 1).font = total_row_font
                all_sheet.cell(total_row_idx, 1).alignment = text_alignment
                
                # تنسيق صف المجموع وحساب المجاميع للأعمدة المالية
                for col_idx, column_name in enumerate(all_columns, 1):
                    cell = all_sheet.cell(total_row_idx, col_idx)
                    cell.font = total_row_font
                    cell.fill = total_row_fill
                    cell.border = thick_border
                    
                    # حساب المجاميع للأعمدة المالية
                    if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'التأمينات (GOSI)', 'المكافآت', 'صافي الراتب']:
                        col_letter = get_column_letter(col_idx)
                        cell.value = f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})"
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
            
            # إنشاء داش بورد احترافي لمعلومات التقرير
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.chart.label import DataLabelList
            
            info_sheet = writer.book.create_sheet('معلومات التقرير')
            ExcelStyles.apply_rtl(info_sheet, engine="openpyxl")
            
            # ==== العنوان الرئيسي ====
            info_sheet.merge_cells('A1:J1')
            title_cell = info_sheet.cell(1, 1)
            title_cell.value = "📊 داش بورد تحليل الرواتب"
            title_cell.font = Font(name="Arial", size=22, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            info_sheet.row_dimensions[1].height = 35
            
            # ==== معلومات التصفية ====
            info_sheet.merge_cells('A2:J2')
            filter_cell = info_sheet.cell(2, 1)
            filter_text = ' | '.join(filter_description) if filter_description else 'جميع البيانات'
            filter_cell.value = f"المرشحات: {filter_text}"
            filter_cell.font = Font(name="Arial", size=11, italic=True, color="666666")
            filter_cell.alignment = Alignment(horizontal='center', vertical='center')
            info_sheet.row_dimensions[2].height = 20
            
            # ==== بطاقات المؤشرات الرئيسية (KPI Cards) ====
            kpi_row = 4
            kpis = [
                {'title': 'إجمالي الموظفين', 'value': total_salaries, 'color': '4472C4', 'icon': '👥'},
                {'title': 'عدد الأقسام', 'value': len(departments_data), 'color': '70AD47', 'icon': '🏢'},
                {'title': 'متوسط الراتب', 'value': f'{(total_net / total_salaries if total_salaries > 0 else 0):,.0f} ر.س', 'color': 'FFC000', 'icon': '💰'},
                {'title': 'إجمالي المصاريف', 'value': f'{total_net:,.0f} ر.س', 'color': 'E74C3C', 'icon': '💵'}
            ]
            
            col_start = 1
            for idx, kpi in enumerate(kpis):
                col = col_start + (idx * 3)
                
                # دمج خلايا البطاقة
                info_sheet.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row+2, end_column=col+1)
                
                # الأيقونة والعنوان
                header_cell = info_sheet.cell(kpi_row, col)
                header_cell.value = f"{kpi['icon']} {kpi['title']}"
                header_cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
                header_cell.fill = PatternFill(start_color=kpi['color'], end_color=kpi['color'], fill_type="solid")
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.border = Border(
                    left=Side(style='medium', color='000000'),
                    right=Side(style='medium', color='000000'),
                    top=Side(style='medium', color='000000'),
                    bottom=Side(style='medium', color='000000')
                )
                
                # القيمة
                value_cell = info_sheet.cell(kpi_row+3, col)
                info_sheet.merge_cells(start_row=kpi_row+3, start_column=col, end_row=kpi_row+3, end_column=col+1)
                value_cell.value = kpi['value']
                value_cell.font = Font(name="Arial", size=18, bold=True, color=kpi['color'])
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                value_cell.border = Border(
                    left=Side(style='medium', color='000000'),
                    right=Side(style='medium', color='000000'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='medium', color='000000')
                )
                
                info_sheet.row_dimensions[kpi_row+3].height = 40
            
            # ==== جدول تفصيل الأقسام ====
            table_row = kpi_row + 6
            info_sheet.cell(table_row, 1).value = "📋 تفصيل الأقسام"
            info_sheet.cell(table_row, 1).font = Font(name="Arial", size=14, bold=True, color="1F4E78")
            
            # رؤوس الجدول
            headers = ['القسم', 'عدد الموظفين', 'الرواتب الأساسية', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الرواتب']
            for col_idx, header in enumerate(headers, 1):
                cell = info_sheet.cell(table_row + 1, col_idx)
                cell.value = header
                cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # بيانات الأقسام
            data_row = table_row + 2
            for dept_name, dept_salaries in departments_data.items():
                dept_count = len(dept_salaries)
                dept_basic = sum(s['الراتب الأساسي'] for s in dept_salaries)
                dept_allowances = sum(s['البدلات'] for s in dept_salaries)
                dept_deductions = sum(s['الخصومات'] for s in dept_salaries)
                dept_bonus = sum(s['المكافآت'] for s in dept_salaries)
                dept_net = sum(s['صافي الراتب'] for s in dept_salaries)
                
                row_data = [dept_name, dept_count, dept_basic, dept_allowances, dept_deductions, dept_bonus, dept_net]
                for col_idx, value in enumerate(row_data, 1):
                    cell = info_sheet.cell(data_row, col_idx)
                    cell.value = value
                    if col_idx > 2:
                        cell.number_format = '#,##0 "ر.س"'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.fill = PatternFill(start_color="E7E6E6" if data_row % 2 == 0 else "FFFFFF", end_color="E7E6E6" if data_row % 2 == 0 else "FFFFFF", fill_type="solid")
                data_row += 1
            
            # ==== مخطط دائري لتوزيع الموظفين ====
            pie_chart = PieChart()
            pie_chart.title = "توزيع الموظفين حسب القسم"
            pie_chart.style = 10
            
            data = Reference(info_sheet, min_col=2, min_row=table_row+1, max_row=data_row-1)
            cats = Reference(info_sheet, min_col=1, min_row=table_row+2, max_row=data_row-1)
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(cats)
            pie_chart.height = 10
            pie_chart.width = 15
            
            info_sheet.add_chart(pie_chart, f"A{data_row + 2}")
            
            # ==== مخطط عمودي للرواتب ====
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.title = "مقارنة الرواتب حسب القسم"
            bar_chart.y_axis.title = 'المبلغ (ر.س)'
            bar_chart.x_axis.title = 'القسم'
            bar_chart.style = 11
            
            data = Reference(info_sheet, min_col=7, min_row=table_row+1, max_row=data_row-1)
            cats = Reference(info_sheet, min_col=1, min_row=table_row+2, max_row=data_row-1)
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)
            bar_chart.height = 10
            bar_chart.width = 15
            
            info_sheet.add_chart(bar_chart, f"H{data_row + 2}")
            
            # ==== معلومات إضافية ====
            info_row = data_row + 20
            info_sheet.cell(info_row, 1).value = "📅 تاريخ التصدير:"
            info_sheet.cell(info_row, 1).font = Font(name="Arial", size=10, bold=True)
            info_sheet.cell(info_row, 2).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            info_sheet.cell(info_row, 2).font = Font(name="Arial", size=10)
            
            # ضبط عرض الأعمدة
            for col in range(1, 11):
                info_sheet.column_dimensions[get_column_letter(col)].width = 15
            
            # تعيين الصفحة الأولى كصفحة نشطة
            writer.book.active = writer.book.worksheets[0]
        
        output.seek(0)
        return output
    
    except Exception as e:
        raise Exception(f"خطأ في إنشاء ملف Excel: {str(e)}")

