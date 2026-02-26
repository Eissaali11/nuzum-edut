"""
مولد التقرير الشامل للموظف
يقوم بتجميع كل بيانات الموظف في تقرير واحد
"""

import os
from datetime import datetime
from fpdf import FPDF
import arabic_reshaper
from bidi.algorithm import get_display
import pandas as pd
from io import BytesIO
from models import Employee, Attendance, Salary, Vehicle, Department

# تحميل الخط العربي - استخدام مسار نسبي من الملف الحالي
FONT_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'fonts')

class EmployeeComprehensivePDF(FPDF):
    """فئة لإنشاء تقرير PDF شامل للموظف"""
    
    def __init__(self, orientation='P', unit='mm', format='A4'):
        super().__init__(orientation, unit, format)
        self.add_font('Amiri', '', os.path.join(FONT_PATH, 'Amiri-Regular.ttf'), uni=True)
        self.add_font('Amiri', 'B', os.path.join(FONT_PATH, 'Amiri-Bold.ttf'), uni=True)
        
    def header(self):   
        """ترويسة الصفحة - تظهر في كل صفحة"""
        # إضافة الشعار
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'images', 'logo.png')
        if os.path.exists(logo_path):
            self.image(logo_path, 10, 8, 33)
        
        # إضافة عنوان الصفحة
        self.set_font('Amiri', 'B', 16)
        self.cell(0, 10, self.arabic_text('التقرير الشامل للموظف'), 0, 1, 'C')
        
        # إضافة التاريخ الحالي
        self.set_font('Amiri', '', 10)
        self.cell(0, 10, self.arabic_text(f'تاريخ التقرير: {datetime.now().strftime("%Y-%m-%d")}'), 0, 1, 'R')
        
        # خط تحت الترويسة
        self.line(10, 30, 200, 30)
        self.ln(15)
        
    def footer(self):
        """تذييل الصفحة - يظهر في كل صفحة"""
        self.set_y(-15)
        self.set_font('Amiri', '', 8)
        self.cell(0, 10, self.arabic_text('صفحة ') + str(self.page_no()) + '/{nb}', 0, 0, 'C')
        
    def arabic_text(self, text):
        """معالجة النص العربي للعرض الصحيح"""
        if not text:
            return ''
        reshaped_text = arabic_reshaper.reshape(str(text))
        bidi_text = get_display(reshaped_text)
        return bidi_text
    
    def add_section_title(self, title):
        """إضافة عنوان قسم"""
        self.set_fill_color(220, 220, 220)
        self.set_font('Amiri', 'B', 14)
        self.cell(0, 10, self.arabic_text(title), 1, 1, 'R', True)
        self.ln(5)
    
    def add_info_row(self, label, value):
        """إضافة صف معلومات"""
        self.set_font('Amiri', 'B', 12)
        self.cell(100, 8, self.arabic_text(str(value) if value else '-'), 1, 0, 'R')
        self.set_font('Amiri', '', 12)
        self.cell(90, 8, self.arabic_text(label), 1, 1, 'R', True)
    
    def add_table_header(self, headers, widths):
        """إضافة ترويسة جدول"""
        self.set_font('Amiri', 'B', 12)
        self.set_fill_color(200, 200, 200)
        
        for i, header in enumerate(headers):
            self.cell(widths[i], 10, self.arabic_text(header), 1, 0, 'C', True)
        self.ln()
    
    def add_table_row(self, data, widths, fill=False):
        """إضافة صف جدول"""
        self.set_font('Amiri', '', 10)
        self.set_fill_color(240, 240, 240)
        
        for i, value in enumerate(data):
            self.cell(widths[i], 8, self.arabic_text(str(value) if value is not None else '-'), 1, 0, 'C', fill)
        self.ln()


def generate_employee_comprehensive_pdf(employee_id):
    """
    إنشاء تقرير PDF شامل للموظف
    
    :param employee_id: معرف الموظف
    :return: مخرجات ملف PDF
    """
    # استعلام بيانات الموظف
    employee = Employee.query.get(employee_id)
    if not employee:
        return None
    
    # استعلام بيانات الحضور
    attendances = Attendance.query.filter_by(employee_id=employee_id).order_by(Attendance.date.desc()).all()
    
    # استعلام بيانات الرواتب
    salaries = Salary.query.filter_by(employee_id=employee_id).order_by(Salary.year.desc(), Salary.month.desc()).all()
    
    # لا نستعلم عن المركبات في هذه النسخة لعدم وجود علاقة مباشرة
    vehicles = []
    
    # إنشاء ملف PDF
    pdf = EmployeeComprehensivePDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    
    # معلومات الموظف الأساسية
    pdf.add_section_title('المعلومات الأساسية')
    
    pdf.add_info_row('اسم الموظف', employee.name)
    pdf.add_info_row('الرقم الوظيفي', employee.employee_id)
    pdf.add_info_row('رقم الهوية', employee.national_id)
    pdf.add_info_row('تاريخ الالتحاق', employee.join_date.strftime('%Y-%m-%d') if employee.join_date else '-')
    pdf.add_info_row('الجنسية', employee.nationality)
    # الحصول على أسماء الأقسام (many-to-many relationship)
    department_names = ', '.join([dept.name for dept in employee.departments]) if employee.departments else '-'
    pdf.add_info_row('القسم', department_names)
    pdf.add_info_row('المشروع', employee.project)
    pdf.add_info_row('الموقع', employee.location)
    pdf.add_info_row('البريد الإلكتروني', employee.email)
    pdf.add_info_row('رقم الجوال', employee.mobile)
    
    # معلومات الكفالة
    pdf.ln(5)
    pdf.add_section_title('معلومات الكفالة')
    sponsorship_status_text = 'على الكفالة' if employee.sponsorship_status != 'outside' else 'خارج الكفالة'
    pdf.add_info_row('حالة الكفالة', sponsorship_status_text)
    
    if employee.sponsorship_status != 'outside' and employee.current_sponsor_name:
        pdf.add_info_row('اسم الكفيل الحالي', employee.current_sponsor_name)
    elif employee.sponsorship_status != 'outside':
        pdf.add_info_row('اسم الكفيل الحالي', 'غير محدد')
    
    # معلومات الوثائق
    pdf.ln(5)
    pdf.add_section_title('معلومات الوثائق')
    
    if employee.documents:
        # عرض الوثائق مع تواريخ الانتهاء
        for doc in employee.documents:
            doc_type_map = {
                'national_id': 'الهوية الوطنية',
                'passport': 'جواز السفر',
                'visa': 'التأشيرة',
                'health_certificate': 'الشهادة الصحية',
                'driving_license': 'رخصة القيادة',
                'contract': 'عقد العمل',
                'other': 'أخرى'
            }
            doc_type_name = doc_type_map.get(doc.document_type, doc.document_type)
            expiry_date = doc.expiry_date.strftime('%Y-%m-%d') if doc.expiry_date else '-'
            pdf.add_info_row(f'{doc_type_name} - تاريخ الانتهاء', expiry_date)
    else:
        pdf.set_font('Amiri', '', 12)
        pdf.cell(0, 10, pdf.arabic_text('لا توجد وثائق مسجلة'), 0, 1, 'C')
    
    # بيانات المركبات (قسم اختياري، يتم عرضه فقط إذا كان هناك مركبات)
    if vehicles:
        pdf.ln(5)
        pdf.add_section_title('المركبات المسجلة')
        
        headers = ['رقم اللوحة', 'الطراز', 'الموديل', 'اللون', 'سنة الصنع']
        widths = [40, 40, 40, 40, 30]
        pdf.add_table_header(headers, widths)
        
        for i, vehicle in enumerate(vehicles):
            row_data = [
                vehicle.plate_number,
                vehicle.make,
                vehicle.model,
                vehicle.color,
                vehicle.year
            ]
            pdf.add_table_row(row_data, widths, fill=(i % 2 == 0))
    
    # بيانات الحضور والغياب
    pdf.ln(5)
    pdf.add_section_title('سجل الحضور والغياب')
    
    if attendances:
        headers = ['التاريخ', 'الحالة', 'الملاحظات']
        widths = [60, 60, 70]
        pdf.add_table_header(headers, widths)
        
        # عرض آخر 15 سجل حضور
        attendance_limit = min(15, len(attendances))
        for i, attendance in enumerate(attendances[:attendance_limit]):
            status_map = {
                'present': 'حاضر',
                'absent': 'غائب',
                'leave': 'إجازة',
                'sick': 'مرضي'
            }
            status_text = status_map.get(attendance.status, attendance.status)
            
            row_data = [
                attendance.date.strftime('%Y-%m-%d'),
                status_text,
                attendance.notes or '-'
            ]
            pdf.add_table_row(row_data, widths, fill=(i % 2 == 0))
        
        # إضافة ملخص إحصائيات الحضور
        pdf.ln(5)
        present_count = sum(1 for a in attendances if a.status == 'present')
        absent_count = sum(1 for a in attendances if a.status == 'absent')
        leave_count = sum(1 for a in attendances if a.status == 'leave')
        sick_count = sum(1 for a in attendances if a.status == 'sick')
        
        pdf.set_font('Amiri', 'B', 12)
        pdf.cell(0, 10, pdf.arabic_text(f'ملخص الحضور: حاضر: {present_count}, غائب: {absent_count}, إجازة: {leave_count}, مرضي: {sick_count}'), 0, 1, 'R')
    else:
        pdf.set_font('Amiri', '', 12)
        pdf.cell(0, 10, pdf.arabic_text('لا توجد سجلات حضور مسجلة'), 0, 1, 'C')
    
    # بيانات الرواتب
    pdf.ln(5)
    pdf.add_section_title('سجل الرواتب')
    
    if salaries:
        headers = ['الشهر', 'السنة', 'الراتب الأساسي', 'البدلات', 'الخصومات', 'صافي الراتب']
        widths = [30, 30, 35, 35, 35, 35]
        pdf.add_table_header(headers, widths)
        
        # عرض آخر 10 رواتب
        salary_limit = min(10, len(salaries))
        for i, salary in enumerate(salaries[:salary_limit]):
            row_data = [
                salary.month,
                salary.year,
                salary.basic_salary,
                salary.allowances,
                salary.deductions,
                salary.net_salary
            ]
            pdf.add_table_row(row_data, widths, fill=(i % 2 == 0))
    else:
        pdf.set_font('Amiri', '', 12)
        pdf.cell(0, 10, pdf.arabic_text('لا توجد سجلات رواتب مسجلة'), 0, 1, 'C')
    
    # طباعة بعض التفاصيل للتشخيص
    print(f"مسار الخط: {FONT_PATH}")
    print(f"ملفات الخطوط الموجودة: {os.listdir(FONT_PATH)}")
    
    # استخدام طريقة أبسط لإنشاء المخرج
    try:
        # إنشاء مخرجات الملف
        output = BytesIO()
        # حفظ الملف مباشرة إلى BytesIO
        pdf_content = pdf.output(dest='S')
        
        # تحويل المحتوى إلى بيانات ثنائية
        if isinstance(pdf_content, str):
            pdf_content = pdf_content.encode('latin1')
            
        # كتابة المحتوى إلى البافر
        output.write(pdf_content)
        output.seek(0)
        
        print(f"تم إنشاء ملف PDF بحجم: {len(output.getvalue())} بايت")
        return output
    except Exception as e:
        print(f"حدث خطأ أثناء إنشاء ملف PDF: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return None


def generate_employee_comprehensive_excel(employee_id):
    """
    إنشاء تقرير Excel شامل واحترافي للموظف بـ 29 عمود
    
    :param employee_id: معرف الموظف
    :return: مخرجات ملف Excel
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # استعلام بيانات الموظف
        employee = Employee.query.get(employee_id)
        if not employee:
            return None
        
        output = BytesIO()
        workbook = Workbook()
        
        # ===== ورقة معلومات الموظف الكاملة =====
        sheet = workbook.active
        sheet.title = "معلومات الموظف الكاملة"
        
        # التنسيقات
        title_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=16, name='Calibri')
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # عنوان التقرير
        sheet.merge_cells('A1:B1')
        title_cell = sheet['A1']
        title_cell.value = f"تقرير شامل - {employee.name}"
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = title_fill
        
        # التاريخ
        sheet.merge_cells('A2:B2')
        date_cell = sheet['A2']
        date_cell.value = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        date_cell.font = Font(bold=True, size=11, name='Calibri')
        date_cell.alignment = center_align
        date_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # جلب السيارة الحالية
        current_vehicle = ""
        try:
            from models import VehicleHandover
            latest_delivery = VehicleHandover.query.filter_by(
                employee_id=employee.id
            ).filter(
                VehicleHandover.handover_type.in_(['delivery', 'تسليم', 'handover'])
            ).order_by(VehicleHandover.handover_date.desc()).first()
            
            if latest_delivery and latest_delivery.vehicle:
                current_vehicle = f"{latest_delivery.vehicle.plate_number}"
        except:
            pass
        
        # جلب بيانات الجهاز المحمول
        mobile_type = ""
        mobile_imei = ""
        mobile_number = ""
        try:
            from models import DeviceAssignment, MobileDevice, SimCard
            active_assignment = DeviceAssignment.query.filter_by(
                employee_id=employee.id,
                is_active=True
            ).first()
            
            if active_assignment:
                if active_assignment.device_id:
                    device = MobileDevice.query.get(active_assignment.device_id)
                    if device:
                        mobile_type = f"{device.device_brand or ''} {device.device_model or ''}".strip()
                        mobile_imei = device.imei or ""
                
                if active_assignment.sim_card_id:
                    sim = SimCard.query.get(active_assignment.sim_card_id)
                    if sim:
                        mobile_number = sim.phone_number or ""
        except:
            pass
        
        # إضافة روابط الوثائق
        job_offer_file = f"https://nuzum.site/static/{getattr(employee, 'job_offer_file', '')}" if getattr(employee, 'job_offer_file', '') else '-'
        passport_file = f"https://nuzum.site/static/{getattr(employee, 'passport_image_file', '')}" if getattr(employee, 'passport_image_file', '') else '-'
        national_address_file = f"https://nuzum.site/static/{getattr(employee, 'national_address_file', '')}" if getattr(employee, 'national_address_file', '') else '-'
        
        # الروابط الخارجية
        job_offer_link = getattr(employee, 'job_offer_link', '') or '-'
        passport_link = getattr(employee, 'passport_image_link', '') or '-'
        national_address_link = getattr(employee, 'national_address_link', '') or '-'
        
        # البيانات الكاملة بالترتيب المطلوب (35 حقل)
        employee_full_data = [
            ("الاسم الكامل", employee.name),
            ("رقم الهوية الوطنية", employee.national_id or ""),
            ("رقم الموظف", employee.employee_id),
            ("الجنسية", employee.nationality_rel.name_ar if hasattr(employee, 'nationality_rel') and employee.nationality_rel else (employee.nationality if hasattr(employee, 'nationality') else "")),
            ("الجوال الشخصي", getattr(employee, 'mobilePersonal', '') or ''),
            ("مقاس البنطلون", getattr(employee, 'pants_size', '') or ''),
            ("مقاس التيشرت", getattr(employee, 'shirt_size', '') or ''),
            ("المسمى الوظيفي", employee.job_title or ""),
            ("الحالة الوظيفية", employee.status or ""),
            ("نوع الجوال", mobile_type),
            ("رقم IMEI", mobile_imei),
            ("رقم الجوال", mobile_number),
            ("السيارة الحالية", current_vehicle),
            ("الموقع", employee.location or ""),
            ("المشروع", employee.project or ""),
            ("البريد الإلكتروني", employee.email or ""),
            ("الأقسام", ', '.join([dept.name for dept in employee.departments]) if employee.departments else ""),
            ("تاريخ الانضمام", employee.join_date.strftime('%Y-%m-%d') if employee.join_date else ""),
            ("تاريخ الميلاد", employee.birth_date.strftime('%Y-%m-%d') if employee.birth_date else ""),
            ("نوع الموظف", getattr(employee, 'employee_type', '') or ''),
            ("نوع العقد", getattr(employee, 'contract_type', '') or ''),
            ("الراتب الأساسي", f"{getattr(employee, 'basic_salary', 0):,.2f} ريال" if getattr(employee, 'basic_salary', 0) else ""),
            ("حالة العقد", getattr(employee, 'contract_status', '') or ''),
            ("حالة الرخصة", getattr(employee, 'license_status', '') or ''),
            ("حالة الكفالة", getattr(employee, 'sponsorship_status', '') or ''),
            ("اسم الكفيل", getattr(employee, 'current_sponsor_name', '') or ''),
            ("رقم الإيبان", getattr(employee, 'bank_iban', '') or ''),
            ("تفاصيل السكن", getattr(employee, 'residence_details', '') or ''),
            ("رابط موقع السكن", getattr(employee, 'residence_location_url', '') or ''),
            ("ملف العرض الوظيفي", job_offer_file),
            ("صورة الجواز", passport_file),
            ("شهادة العنوان الوطني", national_address_file),
            ("رابط العرض الوظيفي الخارجي", job_offer_link),
            ("رابط صورة الجواز الخارجي", passport_link),
            ("رابط شهادة العنوان الخارجي", national_address_link)
        ]
        
        # كتابة البيانات
        row_num = 4
        for label, value in employee_full_data:
            # العمود الأول: التسمية
            label_cell = sheet.cell(row=row_num, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True, size=11, name='Calibri')
            label_cell.fill = header_fill
            label_cell.alignment = right_align
            label_cell.border = border
            
            # العمود الثاني: القيمة
            value_cell = sheet.cell(row=row_num, column=2)
            value_cell.value = value if value else "-"
            value_cell.font = Font(size=11, name='Calibri')
            value_cell.alignment = right_align
            value_cell.border = border
            
            # تلوين الصفوف بالتناوب
            if row_num % 2 == 0:
                value_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            row_num += 1
        
        # ضبط عرض الأعمدة
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 50
        
        # حفظ الملف
        workbook.save(output)
        output.seek(0)
        return output
    
    except Exception as e:
        print(f"خطأ في إنشاء ملف Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return None