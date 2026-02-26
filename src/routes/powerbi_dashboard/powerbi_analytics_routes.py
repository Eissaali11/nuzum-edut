"""
مسارات التحليلات المتقدمة لـ Power BI
جميع مسارات API للإحصائيات والتقارير التفصيلية
"""

from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from src.core.extensions import db
from models import Employee, Attendance, Document, Vehicle, Department
from sqlalchemy import func
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

powerbi_analytics_bp = Blueprint('powerbi_analytics', __name__, url_prefix='/powerbi/api')


@powerbi_analytics_bp.route('/attendance-summary')
@login_required
def attendance_summary():
    """ملخص الحضور مع تحليلات متقدمة"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    department_id = request.args.get('department_id')
    
    try:
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = datetime.now().date() - timedelta(days=30)
        
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = datetime.now().date()
        
        query = Attendance.query.filter(
            Attendance.date >= date_from,
            Attendance.date <= date_to
        )
        
        if department_id:
            employee_ids = [e.id for e in Employee.query.filter_by(department_id=department_id).all()]
            query = query.filter(Attendance.employee_id.in_(employee_ids))
        
        attendance_records = query.all()
        
        present = sum(1 for a in attendance_records if a.status == 'present')
        absent = sum(1 for a in attendance_records if a.status in ['absent', 'غائب'])
        leave = sum(1 for a in attendance_records if a.status == 'leave')
        sick = sum(1 for a in attendance_records if a.status == 'sick')
        total = len(attendance_records)
        
        attendance_rate = round((present / total) * 100, 1) if total > 0 else 0
        absence_rate = round((absent / total) * 100, 1) if total > 0 else 0
        
        prev_date_from = date_from - timedelta(days=30)
        prev_query = Attendance.query.filter(
            Attendance.date >= prev_date_from,
            Attendance.date < date_from
        )
        if department_id:
            prev_query = prev_query.filter(Attendance.employee_id.in_(employee_ids))
        
        prev_records = prev_query.all()
        prev_present = sum(1 for a in prev_records if a.status == 'present')
        prev_total = len(prev_records)
        prev_rate = round((prev_present / prev_total) * 100, 1) if prev_total > 0 else 0
        
        trend = round(attendance_rate - prev_rate, 1)
        trend_direction = 'up' if trend > 0 else 'down' if trend < 0 else 'stable'
        
        return jsonify({
            'success': True,
            'data': {
                'present': present,
                'absent': absent,
                'leave': leave,
                'sick': sick,
                'total': total,
                'attendance_rate': attendance_rate,
                'absence_rate': absence_rate,
                'trend': {
                    'value': abs(trend),
                    'direction': trend_direction,
                    'previous_rate': prev_rate
                },
                'date_range': {
                    'from': date_from.strftime('%Y-%m-%d'),
                    'to': date_to.strftime('%Y-%m-%d')
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@powerbi_analytics_bp.route('/attendance-by-department')
@login_required
def attendance_by_department():
    """الحضور حسب القسم مع تحليل مفصل"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    try:
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = datetime.now().date() - timedelta(days=30)
        
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = datetime.now().date()
        
        departments_data = []
        departments = Department.query.all()
        
        for dept in departments:
            employees = Employee.query.filter_by(department_id=dept.id).all()
            employee_ids = [e.id for e in employees]
            
            if not employee_ids:
                continue
            
            attendance_records = Attendance.query.filter(
                Attendance.date >= date_from,
                Attendance.date <= date_to,
                Attendance.employee_id.in_(employee_ids)
            ).all()
            
            present = sum(1 for a in attendance_records if a.status == 'present')
            absent = sum(1 for a in attendance_records if a.status in ['absent', 'غائب'])
            leave = sum(1 for a in attendance_records if a.status == 'leave')
            sick = sum(1 for a in attendance_records if a.status == 'sick')
            total = len(attendance_records)
            
            attendance_rate = round((present / total) * 100, 1) if total > 0 else 0
            
            performance = 'excellent' if attendance_rate >= 90 else 'good' if attendance_rate >= 75 else 'average' if attendance_rate >= 60 else 'poor'
            
            departments_data.append({
                'department': dept.name,
                'department_id': dept.id,
                'employee_count': len(employees),
                'present': present,
                'absent': absent,
                'leave': leave,
                'sick': sick,
                'total': total,
                'attendance_rate': attendance_rate,
                'performance': performance
            })
        
        departments_data.sort(key=lambda x: x['attendance_rate'], reverse=True)
        
        return jsonify({
            'success': True,
            'data': departments_data,
            'summary': {
                'total_departments': len(departments_data),
                'avg_attendance_rate': round(sum(d['attendance_rate'] for d in departments_data) / len(departments_data), 1) if departments_data else 0,
                'best_department': departments_data[0]['department'] if departments_data else None,
                'worst_department': departments_data[-1]['department'] if departments_data else None
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@powerbi_analytics_bp.route('/documents-status')
@login_required
def documents_status():
    """حالة الوثائق المطلوبة مع تحليل شامل"""
    try:
        required_docs = [
            {'type': 'الهوية الوطنية', 'priority': 'high'},
            {'type': 'جواز السفر', 'priority': 'high'},
            {'type': 'رخصة القيادة', 'priority': 'medium'},
            {'type': 'وثيقة التأمين', 'priority': 'medium'}
        ]
        
        total_employees = Employee.query.count()
        documents_summary = []
        
        today = datetime.now().date()
        thirty_days_later = today + timedelta(days=30)
        
        for doc_info in required_docs:
            doc_type = doc_info['type']
            
            docs = Document.query.filter(
                Document.document_type.ilike(f'%{doc_type}%')
            ).all()
            
            available = len(docs)
            missing = total_employees - available
            
            expiring_soon = 0
            expired = 0
            valid = 0
            
            for doc in docs:
                if hasattr(doc, 'expiry_date') and doc.expiry_date:
                    if doc.expiry_date < today:
                        expired += 1
                    elif doc.expiry_date <= thirty_days_later:
                        expiring_soon += 1
                    else:
                        valid += 1
                else:
                    valid += 1
            
            documents_summary.append({
                'type': doc_type,
                'priority': doc_info['priority'],
                'available': available,
                'missing': missing,
                'valid': valid,
                'expiring_soon': expiring_soon,
                'expired': expired,
                'completion_rate': round((available / total_employees * 100), 1) if total_employees > 0 else 0,
                'health_score': round(((valid) / available * 100), 1) if available > 0 else 0
            })
        
        total_available = sum(d['available'] for d in documents_summary)
        total_required = total_employees * len(required_docs)
        overall_completion = round((total_available / total_required * 100), 1) if total_required > 0 else 0
        
        return jsonify({
            'success': True,
            'data': documents_summary,
            'total_employees': total_employees,
            'overall_completion': overall_completion,
            'total_expiring_soon': sum(d['expiring_soon'] for d in documents_summary),
            'total_expired': sum(d['expired'] for d in documents_summary),
            'total_missing': sum(d['missing'] for d in documents_summary)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@powerbi_analytics_bp.route('/employee-documents')
@login_required
def employee_documents():
    """قائمة الموظفين والوثائق الناقصة مع تفاصيل"""
    department_id = request.args.get('department_id')
    limit = request.args.get('limit', 50, type=int)
    
    try:
        query = Employee.query
        
        if department_id:
            query = query.filter_by(department_id=department_id)
        
        employees = query.limit(limit).all()
        
        required_docs = ['الهوية الوطنية', 'جواز السفر', 'رخصة القيادة', 'وثيقة التأمين']
        employees_data = []
        
        for emp in employees:
            docs = Document.query.filter_by(employee_id=emp.id).all()
            doc_types = [d.document_type for d in docs]
            
            missing_docs = []
            for req_doc in required_docs:
                if not any(req_doc in dt for dt in doc_types):
                    missing_docs.append(req_doc)
            
            completion_rate = round(((len(required_docs) - len(missing_docs)) / len(required_docs)) * 100, 0)
            
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'employee_id': emp.employee_id or '-',
                'department': emp.department.name if emp.department else 'بدون قسم',
                'total_docs': len(docs),
                'missing_docs': missing_docs,
                'missing_count': len(missing_docs),
                'documents_complete': len(missing_docs) == 0,
                'completion_rate': completion_rate
            })
        
        employees_data.sort(key=lambda x: x['missing_count'], reverse=True)
        
        complete_count = sum(1 for e in employees_data if e['documents_complete'])
        incomplete_count = len(employees_data) - complete_count
        
        return jsonify({
            'success': True,
            'data': employees_data,
            'summary': {
                'total': len(employees_data),
                'complete': complete_count,
                'incomplete': incomplete_count,
                'completion_rate': round((complete_count / len(employees_data)) * 100, 1) if employees_data else 0
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@powerbi_analytics_bp.route('/vehicles-summary')
@login_required
def vehicles_summary():
    """ملخص حالة السيارات مع تحليل الأسطول"""
    try:
        vehicles = Vehicle.query.all()
        
        statuses = {}
        brands = {}
        years = {}
        
        for vehicle in vehicles:
            status = vehicle.status or 'unknown'
            statuses[status] = statuses.get(status, 0) + 1
            
            if hasattr(vehicle, 'make') and vehicle.make:
                brands[vehicle.make] = brands.get(vehicle.make, 0) + 1
            
            if hasattr(vehicle, 'year') and vehicle.year:
                years[str(vehicle.year)] = years.get(str(vehicle.year), 0) + 1
        
        vehicles_by_status = []
        status_labels = {
            'in_project': 'في المشروع',
            'in_workshop': 'في الورشة',
            'out_of_service': 'خارج الخدمة',
            'accident': 'حادث',
            'unknown': 'غير محدد'
        }
        
        for status, count in statuses.items():
            vehicles_by_status.append({
                'status': status_labels.get(status, status),
                'status_key': status,
                'count': count,
                'percentage': round((count / len(vehicles) * 100), 1) if vehicles else 0
            })
        
        vehicles_by_brand = [{'brand': b, 'count': c} for b, c in sorted(brands.items(), key=lambda x: x[1], reverse=True)]
        
        total = len(vehicles)
        in_project = statuses.get('in_project', 0)
        in_workshop = statuses.get('in_workshop', 0)
        out_of_service = statuses.get('out_of_service', 0)
        accident = statuses.get('accident', 0)
        
        fleet_health = 'excellent' if in_project >= total * 0.8 else 'good' if in_project >= total * 0.6 else 'average' if in_project >= total * 0.4 else 'poor'
        
        return jsonify({
            'success': True,
            'data': {
                'by_status': vehicles_by_status,
                'by_brand': vehicles_by_brand[:5],
                'total_vehicles': total,
                'in_project': in_project,
                'in_workshop': in_workshop,
                'out_of_service': out_of_service,
                'accident': accident,
                'utilization_rate': round((in_project / total) * 100, 1) if total > 0 else 0,
                'fleet_health': fleet_health
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@powerbi_analytics_bp.route('/vehicle-operations-summary')
@login_required
def vehicle_operations_summary():
    """ملخص عمليات السيارات"""
    try:
        vehicles = Vehicle.query.all()
        
        handovers = sum(1 for v in vehicles if v.handover_records)
        in_workshop = sum(1 for v in vehicles if v.status == 'in_workshop')
        in_project = sum(1 for v in vehicles if v.status == 'in_project')
        out_of_service = sum(1 for v in vehicles if v.status == 'out_of_service')
        accident = sum(1 for v in vehicles if v.status == 'accident')
        
        operations_data = [
            {'type': 'في المشروع', 'count': in_project, 'color': '#38ef7d'},
            {'type': 'مستلمة', 'count': handovers, 'color': '#667eea'},
            {'type': 'في الورشة', 'count': in_workshop, 'color': '#fbbf24'},
            {'type': 'خارج الخدمة', 'count': out_of_service, 'color': '#ef4444'},
            {'type': 'حادث', 'count': accident, 'color': '#ff6b6b'}
        ]
        
        return jsonify({
            'success': True,
            'data': {
                'by_type': operations_data,
                'total_vehicles': len(vehicles),
                'summary': {
                    'active_percentage': round((in_project / len(vehicles)) * 100, 1) if vehicles else 0,
                    'handover_percentage': round((handovers / len(vehicles)) * 100, 1) if vehicles else 0
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@powerbi_analytics_bp.route('/export-data')
@login_required
def export_data():
    """تصدير البيانات بصيغة Excel احترافية"""
    data_type = request.args.get('type', 'all')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "تصدير البيانات"
        
        # إضافة رأس الجدول
        headers = ['الموظف', 'القسم', 'الحالة', 'التاريخ', 'الملاحظات']
        ws.append(headers)
        
        # تنسيق الرأس
        header_fill = PatternFill(start_color="00D4AA", end_color="00D4AA", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="000000")
        
        # إضافة البيانات
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            except:
                date_from = None
        
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            except:
                date_to = None
        
        # جلب سجلات الحضور
        query = Attendance.query
        if date_from and date_to:
            query = query.filter(Attendance.date >= date_from, Attendance.date <= date_to)
        
        records = query.all()
        for record in records:
            employee = Employee.query.get(record.employee_id)
            if employee:
                ws.append([
                    employee.name,
                    employee.department.name if employee.department else '-',
                    record.status,
                    record.date.strftime('%Y-%m-%d') if record.date else '-',
                    ''
                ])
        
        # إعداد الملف
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'powerbi_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@powerbi_analytics_bp.route('/dashboard-stats')
@login_required
def dashboard_stats():
    """إحصائيات شاملة للوحة المعلومات"""
    try:
        total_employees = Employee.query.count()
        total_vehicles = Vehicle.query.count()
        total_documents = Document.query.count()
        total_departments = Department.query.count()
        
        today = datetime.now().date()
        today_attendance = Attendance.query.filter(Attendance.date == today).all()
        today_present = sum(1 for a in today_attendance if a.status == 'present')
        
        working_vehicles = Vehicle.query.filter_by(status='working').count()
        
        return jsonify({
            'success': True,
            'data': {
                'employees': {
                    'total': total_employees,
                    'present_today': today_present
                },
                'vehicles': {
                    'total': total_vehicles,
                    'working': working_vehicles
                },
                'documents': {
                    'total': total_documents
                },
                'departments': {
                    'total': total_departments
                },
                'last_updated': datetime.now().isoformat()
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
