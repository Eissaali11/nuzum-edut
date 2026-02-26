from datetime import datetime, date, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from sqlalchemy import func, or_, and_
import os
import uuid
from PIL import Image
import pillow_heif
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from src.core.extensions import db
from models import RentalProperty, PropertyImage, PropertyPayment, PropertyFurnishing, User, Employee, Department
from src.forms.property_forms import (
    RentalPropertyForm, PropertyImagesForm, PropertyPaymentForm, PropertyFurnishingForm
)
from src.utils.audit_logger import log_activity

properties_bp = Blueprint('properties', __name__)

# إعدادات رفع الملفات
UPLOAD_FOLDER = 'static/uploads/properties'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'heic', 'webp'}


def allowed_file(filename):
    """التحقق من امتداد الملف المسموح به"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def process_and_save_image(file, property_id):
    """معالجة وحفظ الصورة مع دعم HEIC"""
    try:
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        unique_filename = f"{uuid.uuid4()}.{file_ext}"
        
        # إنشاء مجلد التخزين
        property_folder = os.path.join(UPLOAD_FOLDER, str(property_id))
        os.makedirs(property_folder, exist_ok=True)
        filepath = os.path.join(property_folder, unique_filename)
        
        # معالجة صور HEIC
        if file_ext == 'heic':
            heif_file = pillow_heif.read_heif(file)
            image = Image.frombytes(
                heif_file.mode,
                heif_file.size,
                heif_file.data,
                "raw",
            )
            # حفظ كـ JPG
            unique_filename = f"{uuid.uuid4()}.jpg"
            filepath = os.path.join(property_folder, unique_filename)
            image.save(filepath, "JPEG", quality=85)
        else:
            file.save(filepath)
        
        # إرجاع المسار النسبي (يجب أن يبدأ بـ static/)
        return filepath
    except Exception as e:
        print(f"خطأ في معالجة الصورة: {e}")
        return None


def process_and_save_contract(file, property_id):
    """معالجة وحفظ ملف العقد (PDF أو صورة)"""
    try:
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        unique_filename = f"contract_{uuid.uuid4()}.{file_ext}"
        
        # إنشاء مجلد التخزين
        property_folder = os.path.join(UPLOAD_FOLDER, str(property_id))
        os.makedirs(property_folder, exist_ok=True)
        filepath = os.path.join(property_folder, unique_filename)
        
        file.save(filepath)
        
        # إرجاع المسار النسبي (يجب أن يبدأ بـ static/)
        return filepath
    except Exception as e:
        print(f"خطأ في معالجة ملف العقد: {e}")
        return None


@properties_bp.route('/dashboard')
@login_required
def dashboard():
    """لوحة التحكم الرئيسية للعقارات المستأجرة"""
    
    # الحصول على الفلتر من الـ URL
    filter_type = request.args.get('filter', 'all')
    
    # إحصائيات العقارات
    total_properties = RentalProperty.query.filter_by(is_active=True).count()
    active_properties = RentalProperty.query.filter_by(status='active', is_active=True).count()
    
    # العقود المنتهية
    expired_properties_count = RentalProperty.query.filter(
        RentalProperty.contract_end_date < date.today(),
        RentalProperty.is_active == True
    ).count()
    
    # العقود القريبة من الانتهاء (60 يوم)
    expiring_soon_date = date.today() + timedelta(days=60)
    expiring_soon = RentalProperty.query.filter(
        RentalProperty.contract_end_date.between(date.today(), expiring_soon_date),
        RentalProperty.is_active == True
    ).count()
    
    # إجمالي الإيجار السنوي
    total_annual_rent = db.session.query(
        func.sum(RentalProperty.annual_rent_amount)
    ).filter_by(is_active=True, status='active').scalar() or 0
    
    # الدفعات المعلقة
    pending_payments = PropertyPayment.query.filter_by(status='pending').count()
    
    # الدفعات المتأخرة
    overdue_payments = PropertyPayment.query.filter(
        PropertyPayment.status == 'pending',
        PropertyPayment.payment_date < date.today()
    ).count()
    
    # إجمالي المدفوعات (المدفوعة فقط)
    total_paid = db.session.query(
        func.sum(PropertyPayment.amount)
    ).filter_by(status='paid').scalar() or 0
    
    # تطبيق الفلتر على قائمة العقارات
    query = RentalProperty.query.filter_by(is_active=True)
    
    if filter_type == 'active':
        # العقود النشطة (غير منتهية وليست قريبة من الانتهاء)
        query = query.filter(
            RentalProperty.status == 'active',
            RentalProperty.contract_end_date > expiring_soon_date
        )
    elif filter_type == 'expiring':
        # العقود القريبة من الانتهاء (خلال 60 يوم)
        query = query.filter(
            RentalProperty.contract_end_date.between(date.today(), expiring_soon_date)
        )
    elif filter_type == 'expired':
        # العقود المنتهية
        query = query.filter(
            RentalProperty.contract_end_date < date.today()
        )
    
    properties = query.order_by(RentalProperty.created_at.desc()).all()
    
    # قائمة العقارات القريبة من الانتهاء (للعمود الجانبي)
    expiring_properties = RentalProperty.query.filter(
        RentalProperty.contract_end_date.between(date.today(), expiring_soon_date),
        RentalProperty.is_active == True
    ).order_by(RentalProperty.contract_end_date).all()
    
    # الدفعات القادمة (خلال 30 يوم)
    upcoming_payments_date = date.today() + timedelta(days=30)
    upcoming_payments = PropertyPayment.query.filter(
        PropertyPayment.status == 'pending',
        PropertyPayment.payment_date.between(date.today(), upcoming_payments_date)
    ).order_by(PropertyPayment.payment_date).all()
    
    # قائمة الدفعات المعلقة
    pending_payments_list = PropertyPayment.query.filter_by(status='pending').order_by(
        PropertyPayment.payment_date
    ).all()
    
    # قائمة الدفعات المتأخرة
    overdue_payments_list = PropertyPayment.query.filter(
        PropertyPayment.status == 'pending',
        PropertyPayment.payment_date < date.today()
    ).order_by(PropertyPayment.payment_date).all()
    
    return render_template('properties/dashboard.html',
                         total_properties=total_properties,
                         active_properties=active_properties,
                         expired_properties=expired_properties_count,
                         expiring_soon=expiring_soon,
                         total_annual_rent=total_annual_rent,
                         pending_payments=pending_payments,
                         overdue_payments=overdue_payments,
                         total_paid=total_paid,
                         properties=properties,
                         expiring_properties=expiring_properties,
                         upcoming_payments=upcoming_payments,
                         pending_payments_list=pending_payments_list,
                         overdue_payments_list=overdue_payments_list,
                         today=date.today(),
                         current_filter=filter_type)


@properties_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    """إضافة عقار جديد"""
    form = RentalPropertyForm()
    
    if form.validate_on_submit():
        try:
            # إنشاء العقار
            property = RentalProperty(
                city=form.name.data,
                address=form.address.data,
                map_link='',
                location_link=form.location_link.data or None,
                contract_number=form.contract_number.data or None,  # استخدام None بدلاً من قيمة فارغة
                owner_name=form.landlord_name.data,
                owner_id=form.property_type.data,  # استخدام owner_id لحفظ نوع العقار مؤقتاً
                contract_start_date=form.contract_start_date.data,
                contract_end_date=form.contract_end_date.data,
                annual_rent_amount=form.monthly_rent.data * 12,
                includes_utilities=False,
                payment_method=form.payment_method.data,
                status='active',
                notes=form.notes.data,
                created_by=current_user.id
            )
            
            db.session.add(property)
            db.session.commit()
            
            # معالجة ملف العقد إن وجد
            if form.contract_file.data:
                contract_file = request.files.get('contract_file')
                if contract_file and contract_file.filename:
                    contract_path = process_and_save_contract(contract_file, property.id)
                    if contract_path:
                        property.contract_file = contract_path
                        db.session.commit()
            
            # معالجة الصور إن وجدت
            if form.images.data:
                files = request.files.getlist('images')
                for file in files:
                    if file and allowed_file(file.filename):
                        filepath = process_and_save_image(file, property.id)
                        if filepath:
                            image = PropertyImage(
                                property_id=property.id,
                                image_path=filepath
                            )
                            db.session.add(image)
                db.session.commit()
            
            # تسجيل النشاط
            log_activity(
                action='إضافة عقار مستأجر',
                entity_type='RentalProperty',
                entity_id=property.id,
                details=f'تم إضافة عقار جديد: {property.contract_number} - {property.city}'
            )
            
            flash('تم إضافة العقار بنجاح!', 'success')
            return redirect(url_for('properties.view', property_id=property.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء إضافة العقار: {str(e)}', 'danger')
    
    return render_template('properties/create.html', form=form)


@properties_bp.route('/<int:property_id>')
@login_required
def view(property_id):
    """عرض تفاصيل العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # جلب الصور
    images = PropertyImage.query.filter_by(property_id=property_id).order_by(PropertyImage.uploaded_at.desc()).all()
    
    # جلب الدفعات
    payments = PropertyPayment.query.filter_by(property_id=property_id).order_by(PropertyPayment.payment_date.desc()).all()
    
    # جلب التجهيزات
    furnishing = PropertyFurnishing.query.filter_by(property_id=property_id).first()
    
    # جلب الأقسام والموظفين لإدارة القاطنين
    departments = Department.query.order_by(Department.name).all()
    employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
    
    return render_template('properties/view.html',
                         property=property,
                         images=images,
                         payments=payments,
                         furnishing=furnishing,
                         departments=departments,
                         employees=employees)


@properties_bp.route('/<int:property_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(property_id):
    """تعديل بيانات العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # ملء النموذج بالبيانات الحالية
    form = RentalPropertyForm()
    if request.method == 'GET':
        form.name.data = property.city
        form.property_type.data = property.owner_id
        form.address.data = property.address
        form.contract_number.data = property.contract_number
        form.landlord_name.data = property.owner_name
        form.landlord_phone.data = ''
        form.contract_start_date.data = property.contract_start_date
        form.contract_end_date.data = property.contract_end_date
        form.monthly_rent.data = property.annual_rent_amount / 12
        form.payment_method.data = property.payment_method
        form.location_link.data = property.location_link
        form.notes.data = property.notes
    
    if form.validate_on_submit():
        try:
            property.city = form.name.data
            property.address = form.address.data
            property.map_link = ''
            property.location_link = form.location_link.data or None
            property.contract_number = form.contract_number.data or None  # استخدام None بدلاً من قيمة فارغة
            property.owner_name = form.landlord_name.data
            property.owner_id = form.property_type.data  # استخدام owner_id لحفظ نوع العقار مؤقتاً
            property.contract_start_date = form.contract_start_date.data
            property.contract_end_date = form.contract_end_date.data
            property.annual_rent_amount = form.monthly_rent.data * 12
            property.includes_utilities = False
            property.payment_method = form.payment_method.data
            property.status = 'active'
            property.notes = form.notes.data
            
            db.session.commit()
            
            # معالجة ملف العقد الجديد إن وجد
            if form.contract_file.data:
                contract_file = request.files.get('contract_file')
                if contract_file and contract_file.filename:
                    contract_path = process_and_save_contract(contract_file, property.id)
                    if contract_path:
                        property.contract_file = contract_path
                        db.session.commit()
            
            # معالجة الصور الجديدة إن وجدت
            if form.images.data:
                files = request.files.getlist('images')
                for file in files:
                    if file and allowed_file(file.filename):
                        filepath = process_and_save_image(file, property.id)
                        if filepath:
                            image = PropertyImage(
                                property_id=property.id,
                                image_path=filepath
                            )
                            db.session.add(image)
                db.session.commit()
            
            # تسجيل النشاط
            log_activity(
                action='تعديل عقار مستأجر',
                entity_type='RentalProperty',
                entity_id=property.id,
                details=f'تم تعديل العقار: {property.contract_number}'
            )
            
            flash('تم تحديث بيانات العقار بنجاح!', 'success')
            return redirect(url_for('properties.view', property_id=property.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث البيانات: {str(e)}', 'danger')
    
    return render_template('properties/edit.html', form=form, property=property)


@properties_bp.route('/<int:property_id>/delete', methods=['POST'])
@login_required
def delete(property_id):
    """حذف العقار (حذف منطقي)"""
    property = RentalProperty.query.get_or_404(property_id)
    
    try:
        property.is_active = False
        db.session.commit()
        
        # تسجيل النشاط
        log_activity(
            action='حذف عقار مستأجر',
            entity_type='RentalProperty',
            entity_id=property.id,
            details=f'تم حذف العقار: {property.contract_number}'
        )
        
        flash('تم حذف العقار بنجاح!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف العقار: {str(e)}', 'danger')
    
    return redirect(url_for('properties.dashboard'))


@properties_bp.route('/<int:property_id>/images/upload', methods=['POST'])
@login_required
def upload_images(property_id):
    """رفع صور العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    
    image_type = request.form.get('image_type', 'أخرى')
    description = request.form.get('description', '')
    
    files = request.files.getlist('images')
    
    if not files:
        flash('الرجاء اختيار صور للرفع', 'warning')
        return redirect(url_for('properties.view', property_id=property_id))
    
    uploaded_count = 0
    for file in files:
        if file and allowed_file(file.filename):
            filepath = process_and_save_image(file, property_id)
            if filepath:
                image = PropertyImage(
                    property_id=property_id,
                    image_path=filepath,
                    image_type=image_type,
                    description=description
                )
                db.session.add(image)
                uploaded_count += 1
    
    try:
        db.session.commit()
        flash(f'تم رفع {uploaded_count} صورة بنجاح!', 'success')
        
        # تسجيل النشاط
        log_activity(
            action='رفع صور عقار',
            entity_type='RentalProperty',
            entity_id=property.id,
            details=f'تم رفع {uploaded_count} صورة للعقار: {property.contract_number}'
        )
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء رفع الصور: {str(e)}', 'danger')
    
    return redirect(url_for('properties.view', property_id=property_id))


@properties_bp.route('/images/<int:image_id>/delete', methods=['POST'])
@login_required
def delete_image(image_id):
    """حذف صورة"""
    image = PropertyImage.query.get_or_404(image_id)
    property_id = image.property_id
    
    try:
        # 💾 الملف يبقى محفوظاً - نحذف فقط المرجع من قاعدة البيانات
        db.session.delete(image)
        db.session.commit()
        
        flash('تم حذف الصورة (الملف محفوظ بشكل آمن)!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الصورة: {str(e)}', 'danger')
    
    return redirect(url_for('properties.view', property_id=property_id))


@properties_bp.route('/<int:property_id>/payments/add', methods=['GET', 'POST'])
@login_required
def add_payment(property_id):
    """إضافة دفعة جديدة"""
    property = RentalProperty.query.get_or_404(property_id)
    form = PropertyPaymentForm()
    
    if form.validate_on_submit():
        try:
            payment = PropertyPayment(
                property_id=property_id,
                payment_date=form.payment_date.data,
                amount=form.amount.data,
                status=form.status.data,
                actual_payment_date=form.actual_payment_date.data,
                payment_method=form.payment_method.data,
                reference_number=form.reference_number.data,
                notes=form.notes.data
            )
            
            db.session.add(payment)
            db.session.commit()
            
            # تسجيل النشاط
            log_activity(
                action='إضافة دفعة إيجار',
                entity_type='PropertyPayment',
                entity_id=payment.id,
                details=f'تم إضافة دفعة بقيمة {payment.amount} ريال للعقار: {property.contract_number}'
            )
            
            flash('تم إضافة الدفعة بنجاح!', 'success')
            return redirect(url_for('properties.view', property_id=property_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء إضافة الدفعة: {str(e)}', 'danger')
    
    return render_template('properties/add_payment.html', form=form, property=property)


@properties_bp.route('/payments/<int:payment_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_payment(payment_id):
    """تعديل دفعة"""
    payment = PropertyPayment.query.get_or_404(payment_id)
    property = payment.rental_property
    form = PropertyPaymentForm(obj=payment)
    
    if form.validate_on_submit():
        try:
            payment.payment_date = form.payment_date.data
            payment.amount = form.amount.data
            payment.status = form.status.data
            payment.actual_payment_date = form.actual_payment_date.data
            payment.payment_method = form.payment_method.data
            payment.reference_number = form.reference_number.data
            payment.notes = form.notes.data
            
            db.session.commit()
            
            # تسجيل النشاط
            log_activity(
                action='تعديل دفعة إيجار',
                entity_type='PropertyPayment',
                entity_id=payment.id,
                details=f'تم تعديل دفعة بقيمة {payment.amount} ريال للعقار: {property.contract_number}'
            )
            
            flash('تم تحديث الدفعة بنجاح!', 'success')
            return redirect(url_for('properties.view', property_id=property.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث الدفعة: {str(e)}', 'danger')
    
    return render_template('properties/edit_payment.html', form=form, payment=payment, property=property)


@properties_bp.route('/<int:property_id>/payments')
@login_required
def payments_list(property_id):
    """عرض جميع دفعات العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    payments = PropertyPayment.query.filter_by(property_id=property_id).order_by(
        PropertyPayment.payment_date.desc()
    ).all()
    
    # حساب الإحصائيات
    paid_count = sum(1 for p in payments if p.status == 'paid')
    pending_count = sum(1 for p in payments if p.status == 'pending')
    overdue_count = sum(1 for p in payments if p.status == 'overdue')
    
    return render_template('properties/payments_list.html', 
                         property=property,
                         payments=payments,
                         paid_count=paid_count,
                         pending_count=pending_count,
                         overdue_count=overdue_count)


@properties_bp.route('/<int:property_id>/furnishing-view')
@login_required
def furnishing_view(property_id):
    """عرض تجهيزات العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    furnishing = PropertyFurnishing.query.filter_by(property_id=property_id).first()
    
    return render_template('properties/furnishing_view.html',
                         property=property,
                         furnishing=furnishing)


@properties_bp.route('/payments/<int:payment_id>/delete', methods=['POST'])
@login_required
def delete_payment(payment_id):
    """حذف دفعة"""
    payment = PropertyPayment.query.get_or_404(payment_id)
    property_id = payment.property_id
    
    try:
        db.session.delete(payment)
        db.session.commit()
        
        flash('تم حذف الدفعة بنجاح!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الدفعة: {str(e)}', 'danger')
    
    return redirect(url_for('properties.view', property_id=property_id))


@properties_bp.route('/<int:property_id>/furnishing', methods=['GET', 'POST'])
@login_required
def manage_furnishing(property_id):
    """إدارة تجهيزات العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    furnishing = PropertyFurnishing.query.filter_by(property_id=property_id).first()
    
    if not furnishing:
        furnishing = PropertyFurnishing(property_id=property_id)
    
    form = PropertyFurnishingForm(obj=furnishing)
    
    if form.validate_on_submit():
        try:
            furnishing.gas_cylinder = form.gas_cylinder.data or 0
            furnishing.stoves = form.stoves.data or 0
            furnishing.beds = form.beds.data or 0
            furnishing.blankets = form.blankets.data or 0
            furnishing.pillows = form.pillows.data or 0
            furnishing.other_items = form.other_items.data
            furnishing.notes = form.notes.data
            
            if not furnishing.id:
                db.session.add(furnishing)
            
            db.session.commit()
            
            # تسجيل النشاط
            log_activity(
                action='تحديث تجهيزات عقار',
                entity_type='PropertyFurnishing',
                entity_id=furnishing.id,
                details=f'تم تحديث تجهيزات العقار: {property.contract_number}'
            )
            
            flash('تم تحديث التجهيزات بنجاح!', 'success')
            return redirect(url_for('properties.view', property_id=property_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تحديث التجهيزات: {str(e)}', 'danger')
    
    return render_template('properties/furnishing.html', form=form, property=property, furnishing=furnishing)


@properties_bp.route('/<int:property_id>/export-excel')
@login_required
def export_excel(property_id):
    """تصدير بيانات العقار إلى Excel"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # جلب البيانات المرتبطة
    payments = PropertyPayment.query.filter_by(property_id=property_id).order_by(PropertyPayment.payment_date).all()
    furnishing = PropertyFurnishing.query.filter_by(property_id=property_id).first()
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات العقار"
    
    # تعريف الألوان والتنسيقات
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان الرئيسي
    ws.merge_cells('A1:D1')
    ws['A1'] = f"تقرير العقار: {property.city}"
    ws['A1'].font = Font(bold=True, size=16, color="1F4788")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # معلومات العقار الأساسية
    ws['A3'] = "معلومات العقار"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:B3')
    
    property_data = [
        ('اسم العقار', property.city),
        ('نوع العقار', {'apartment': 'شقة', 'villa': 'فيلا', 'building': 'عمارة', 
                       'full_floor': 'دور كامل', 'office': 'مكتب', 'warehouse': 'مستودع'}.get(property.owner_id, '-')),
        ('العنوان', property.address),
        ('رقم العقد', property.contract_number or '-'),
        ('اسم المالك', property.owner_name),
        ('تاريخ البداية', property.contract_start_date.strftime('%Y-%m-%d')),
        ('تاريخ الانتهاء', property.contract_end_date.strftime('%Y-%m-%d')),
        ('الإيجار السنوي', f"{property.annual_rent_amount:,.0f} ريال"),
        ('طريقة الدفع', {'monthly': 'شهري', 'quarterly': 'ربع سنوي', 
                        'semi_annually': 'نصف سنوي', 'annually': 'سنوي'}.get(property.payment_method, '-')),
    ]
    
    row = 4
    for label, value in property_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # التجهيزات
    if furnishing:
        row += 1
        ws[f'A{row}'] = "التجهيزات"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        furnishing_data = [
            ('جرات الغاز', furnishing.gas_cylinder),
            ('الطباخات', furnishing.stoves),
            ('الأسرّة', furnishing.beds),
            ('البطانيات', furnishing.blankets),
            ('المخدات', furnishing.pillows),
        ]
        
        for label, value in furnishing_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
    
    # ========== تفاصيل الدفعات بتصميم مميز ==========
    row += 2
    
    # عنوان رئيسي للدفعات مع تصميم مميز
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "📊 تفاصيل الدفعات"
    cell.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 35
    row += 1
    
    # إنشاء تنسيقات مميزة بألوان جميلة
    upcoming_fill = PatternFill(start_color='17A2B8', end_color='17A2B8', fill_type='solid')  # أزرق سماوي
    upcoming_light = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')
    pending_fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')  # أصفر ذهبي
    pending_light = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    overdue_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')  # أحمر
    overdue_light = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    paid_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')  # أخضر
    paid_light = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    
    # 1. الدفعات المستحقة (30 يوم قادمة)
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "🔵 الدفعات المستحقة (30 يوم قادمة)"
    cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    cell.fill = upcoming_fill
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    upcoming_payments_date = date.today() + timedelta(days=30)
    upcoming_payments = [p for p in payments if p.status == 'pending' and date.today() <= p.payment_date <= upcoming_payments_date]
    
    if upcoming_payments:
        # عناوين
        headers = ['التاريخ', 'المبلغ (ريال)', 'الأيام المتبقية', 'طريقة الدفع', 'الحالة']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='0C5460')
            cell.fill = upcoming_light
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for payment in upcoming_payments:
            days_left = (payment.payment_date - date.today()).days
            ws.cell(row=row, column=1, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=2, value=f"{payment.amount:,.0f}").border = border
            cell = ws.cell(row=row, column=3, value=f"⏰ {days_left} يوم")
            cell.border = border
            cell.font = Font(bold=True, color='0C5460')
            ws.cell(row=row, column=4, value=payment.payment_method or '-').border = border
            ws.cell(row=row, column=5, value='معلق').border = border
            row += 1
    else:
        ws[f'A{row}'] = "✅ لا توجد دفعات مستحقة خلال 30 يوم"
        ws[f'A{row}'].font = Font(italic=True, color='155724')
        row += 1
    
    # 2. الدفعات المعلقة
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "🟡 الدفعات المعلقة"
    cell.font = Font(name='Arial', size=12, bold=True, color='000000')
    cell.fill = pending_fill
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    pending_payments = [p for p in payments if p.status == 'pending']
    
    if pending_payments:
        # عناوين
        headers = ['التاريخ المتوقع', 'المبلغ (ريال)', 'طريقة الدفع', 'الرقم المرجعي', 'ملاحظات']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='856404')
            cell.fill = pending_light
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for payment in pending_payments:
            ws.cell(row=row, column=1, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=2, value=f"{payment.amount:,.0f}").border = border
            ws.cell(row=row, column=3, value=payment.payment_method or '-').border = border
            ws.cell(row=row, column=4, value=payment.reference_number or '-').border = border
            ws.cell(row=row, column=5, value=payment.notes or '-').border = border
            row += 1
    else:
        ws[f'A{row}'] = "✅ لا توجد دفعات معلقة"
        ws[f'A{row}'].font = Font(italic=True, color='155724')
        row += 1
    
    # 3. الدفعات المتأخرة
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "🔴 الدفعات المتأخرة"
    cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    cell.fill = overdue_fill
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    overdue_payments = [p for p in payments if p.status == 'pending' and p.payment_date < date.today()]
    
    if overdue_payments:
        # عناوين
        headers = ['التاريخ المتوقع', 'المبلغ (ريال)', 'أيام التأخير', 'طريقة الدفع', 'حالة خطرة']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='721C24')
            cell.fill = overdue_light
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for payment in overdue_payments:
            days_overdue = (date.today() - payment.payment_date).days
            ws.cell(row=row, column=1, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=2, value=f"{payment.amount:,.0f}").border = border
            cell = ws.cell(row=row, column=3, value=f"⚠️ متأخر {days_overdue} يوم")
            cell.border = border
            cell.font = Font(bold=True, color='721C24')
            ws.cell(row=row, column=4, value=payment.payment_method or '-').border = border
            ws.cell(row=row, column=5, value='يتطلب متابعة عاجلة').border = border
            row += 1
    else:
        ws[f'A{row}'] = "✅ لا توجد دفعات متأخرة"
        ws[f'A{row}'].font = Font(italic=True, color='155724')
        row += 1
    
    # 4. الدفعات المدفوعة
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "🟢 الدفعات المدفوعة"
    cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    cell.fill = paid_fill
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    paid_payments = [p for p in payments if p.status == 'paid']
    
    if paid_payments:
        # عناوين
        headers = ['تاريخ الدفع الفعلي', 'المبلغ (ريال)', 'طريقة الدفع', 'الرقم المرجعي', 'ملاحظات']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='155724')
            cell.fill = paid_light
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for payment in paid_payments:
            ws.cell(row=row, column=1, value=payment.actual_payment_date.strftime('%Y-%m-%d') if payment.actual_payment_date else '-').border = border
            ws.cell(row=row, column=2, value=f"{payment.amount:,.0f}").border = border
            ws.cell(row=row, column=3, value=payment.payment_method or '-').border = border
            ws.cell(row=row, column=4, value=payment.reference_number or '-').border = border
            ws.cell(row=row, column=5, value=payment.notes or '-').border = border
            row += 1
    else:
        ws[f'A{row}'] = "لم يتم دفع أي دفعات بعد"
        ws[f'A{row}'].font = Font(italic=True)
        row += 1
    
    # ملخص إحصائي مميز
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "📈 ملخص إحصائي للدفعات"
    cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    total_payments = len(payments)
    total_amount = sum(p.amount for p in payments)
    paid_amount = sum(p.amount for p in paid_payments)
    pending_amount = sum(p.amount for p in pending_payments)
    
    stats = [
        ['إجمالي الدفعات', total_payments],
        ['دفعات مدفوعة', len(paid_payments)],
        ['دفعات معلقة', len(pending_payments)],
        ['دفعات متأخرة', len(overdue_payments)],
        ['المبلغ الإجمالي', f"{total_amount:,.0f} ريال"],
        ['المبلغ المدفوع', f"{paid_amount:,.0f} ريال"],
        ['المبلغ المعلق', f"{pending_amount:,.0f} ريال"],
    ]
    
    for stat in stats:
        ws[f'A{row}'] = stat[0]
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
        ws[f'A{row}'].border = border
        
        ws[f'B{row}'] = stat[1]
        ws[f'B{row}'].border = border
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        row += 1
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    
    # حفظ في الذاكرة
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # اسم الملف
    filename = f"عقار_{property.city}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@properties_bp.route('/<int:property_id>/export-residents-excel')
@login_required
def export_residents_excel(property_id):
    """تصدير بيانات الموظفين القاطنين إلى Excel"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # جلب الموظفين القاطنين
    residents = property.residents
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "الموظفون القاطنون"
    
    # تعريف الألوان والتنسيقات
    header_fill = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان الرئيسي
    ws.merge_cells('A1:H1')
    ws['A1'] = f"الموظفون القاطنون في العقار: {property.city}"
    ws['A1'].font = Font(bold=True, size=16, color="1F4788")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # معلومات العقار
    property_types = {'apartment': 'شقة', 'villa': 'فيلا', 'building': 'عمارة', 'full_floor': 'دور كامل', 'office': 'مكتب', 'warehouse': 'مستودع'}
    ws['A3'] = f"نوع العقار: {property_types.get(property.owner_id, '-')}"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A4'] = f"العنوان: {property.address or '-'}"
    ws['A4'].font = Font(size=10)
    ws['A5'] = f"رقم العقد: {property.contract_number or '-'}"
    ws['A5'].font = Font(size=10)
    
    # عنوان القائمة
    row = 7
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = f"👥 قائمة الموظفين القاطنين ({len(residents)} موظف)"
    cell.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30
    row += 1
    
    # عناوين الأعمدة
    headers = ['#', 'اسم الموظف', 'رقم الموظف', 'المسمى الوظيفي', 'القسم', 'الجنسية', 'رقم الهوية', 'الحالة']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    # بيانات الموظفين
    for idx, resident in enumerate(residents, start=1):
        # الرقم التسلسلي
        cell = ws.cell(row=row, column=1, value=idx)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # اسم الموظف
        cell = ws.cell(row=row, column=2, value=resident.name)
        cell.border = border
        cell.font = Font(bold=True)
        
        # رقم الموظف
        cell = ws.cell(row=row, column=3, value=resident.employee_id)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # المسمى الوظيفي
        cell = ws.cell(row=row, column=4, value=resident.job_title or '-')
        cell.border = border
        
        # القسم
        department_name = resident.departments[0].name if resident.departments else 'بدون قسم'
        cell = ws.cell(row=row, column=5, value=department_name)
        cell.border = border
        
        # الجنسية
        nationality = resident.nationality_rel.name_ar if resident.nationality_rel else '-'
        cell = ws.cell(row=row, column=6, value=nationality)
        cell.border = border
        
        # رقم الهوية
        cell = ws.cell(row=row, column=7, value=resident.national_id or '-')
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # الحالة
        status_text = 'نشط' if resident.status == 'active' else 'غير نشط'
        cell = ws.cell(row=row, column=8, value=status_text)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
        # تلوين الصف بالتبادل
        if idx % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        row += 1
    
    # إضافة ملخص إحصائي
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = "📊 ملخص إحصائي"
    cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # حساب الإحصائيات
    active_count = sum(1 for r in residents if r.status == 'active')
    departments_count = len(set(r.departments[0].name for r in residents if r.departments))
    
    stats = [
        ['إجمالي الموظفين القاطنين', len(residents)],
        ['موظفين نشطين', active_count],
        ['موظفين غير نشطين', len(residents) - active_count],
        ['عدد الأقسام المختلفة', departments_count],
    ]
    
    for stat in stats:
        ws[f'A{row}'] = stat[0]
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
        ws[f'A{row}'].border = border
        
        ws[f'B{row}'] = stat[1]
        ws[f'B{row}'].border = border
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        ws[f'B{row}'].font = Font(bold=True, color='667EEA')
        row += 1
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    # حفظ في الذاكرة
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # اسم الملف
    filename = f"الموظفون_القاطنون_{property.city}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@properties_bp.route('/export-all-excel')
@login_required
def export_all_properties_excel():
    """تصدير جميع بيانات العقارات إلى Excel"""
    
    # جلب جميع العقارات النشطة
    properties = RentalProperty.query.filter_by(is_active=True).order_by(
        RentalProperty.created_at.desc()
    ).all()
    
    # إنشاء ملف Excel
    wb = Workbook()
    
    # ========== ورقة 1: لوحة المعلومات ==========
    ws_dashboard = wb.active
    ws_dashboard.title = "لوحة المعلومات"
    
    # تنسيقات
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    title_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    info_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان الرئيسي
    ws_dashboard.merge_cells('A1:F1')
    cell = ws_dashboard['A1']
    cell.value = "تقرير العقارات المستأجرة - لوحة المعلومات"
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_dashboard.row_dimensions[1].height = 30
    
    # تاريخ التقرير
    ws_dashboard.merge_cells('A2:F2')
    cell = ws_dashboard['A2']
    cell.value = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    cell.font = Font(name='Arial', size=10, italic=True)
    cell.alignment = Alignment(horizontal='center')
    
    # الإحصائيات الرئيسية
    row = 4
    ws_dashboard.merge_cells(f'A{row}:F{row}')
    cell = ws_dashboard[f'A{row}']
    cell.value = "الإحصائيات الرئيسية"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    
    # حساب الإحصائيات
    total_properties = len(properties)
    active_properties = sum(1 for p in properties if p.status == 'active')
    expired_properties = sum(1 for p in properties if p.is_expired)
    expiring_soon = sum(1 for p in properties if p.is_expiring_soon and not p.is_expired)
    total_annual_rent = sum(p.annual_rent_amount for p in properties if p.status == 'active')
    
    # عرض الإحصائيات
    stats_data = [
        ['إجمالي العقارات', total_properties],
        ['عقود نشطة', active_properties],
        ['عقود منتهية', expired_properties],
        ['قريبة من الانتهاء (60 يوم)', expiring_soon],
        ['إجمالي الإيجار السنوي', f"{total_annual_rent:,.0f} ريال"],
    ]
    
    row = 5
    for stat in stats_data:
        ws_dashboard[f'A{row}'] = stat[0]
        ws_dashboard[f'A{row}'].font = subheader_font
        ws_dashboard[f'A{row}'].fill = info_fill
        ws_dashboard[f'A{row}'].border = border
        
        ws_dashboard[f'B{row}'] = stat[1]
        ws_dashboard[f'B{row}'].font = normal_font
        ws_dashboard[f'B{row}'].border = border
        ws_dashboard[f'B{row}'].alignment = Alignment(horizontal='center')
        row += 1
    
    # ملخص حسب النوع
    row += 1
    ws_dashboard.merge_cells(f'A{row}:F{row}')
    cell = ws_dashboard[f'A{row}']
    cell.value = "ملخص حسب نوع العقار"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    
    row += 1
    # عناوين
    headers = ['نوع العقار', 'العدد', 'إجمالي الإيجار السنوي']
    for col, header in enumerate(headers, start=1):
        cell = ws_dashboard.cell(row=row, column=col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = info_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # تجميع حسب النوع
    property_types = {}
    for prop in properties:
        ptype = prop.owner_id or 'غير محدد'  # owner_id يحتوي نوع العقار مؤقتاً
        if ptype not in property_types:
            property_types[ptype] = {'count': 0, 'total_rent': 0}
        property_types[ptype]['count'] += 1
        if prop.status == 'active':
            property_types[ptype]['total_rent'] += prop.annual_rent_amount
    
    row += 1
    for ptype, data in property_types.items():
        ws_dashboard.cell(row=row, column=1, value=ptype).border = border
        ws_dashboard.cell(row=row, column=2, value=data['count']).border = border
        ws_dashboard.cell(row=row, column=3, value=f"{data['total_rent']:,.0f} ريال").border = border
        row += 1
    
    # ضبط عرض الأعمدة
    ws_dashboard.column_dimensions['A'].width = 30
    ws_dashboard.column_dimensions['B'].width = 20
    ws_dashboard.column_dimensions['C'].width = 25
    ws_dashboard.column_dimensions['D'].width = 20
    ws_dashboard.column_dimensions['E'].width = 20
    ws_dashboard.column_dimensions['F'].width = 20
    
    # ========== ورقة 2: قائمة العقارات ==========
    ws_properties = wb.create_sheet(title="قائمة العقارات")
    
    # العنوان
    ws_properties.merge_cells('A1:L1')
    cell = ws_properties['A1']
    cell.value = "قائمة العقارات المستأجرة"
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_properties.row_dimensions[1].height = 30
    
    # عناوين الأعمدة
    row = 3
    headers = [
        'رقم العقد', 'اسم العقار', 'نوع العقار', 'العنوان', 'المالك',
        'تاريخ البداية', 'تاريخ الانتهاء', 'الأيام المتبقية',
        'الإيجار السنوي', 'الإيجار الشهري', 'طريقة الدفع', 'الحالة'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_properties.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # بيانات العقارات
    row = 4
    for prop in properties:
        ws_properties.cell(row=row, column=1, value=prop.contract_number or '-').border = border
        ws_properties.cell(row=row, column=2, value=prop.city).border = border
        ws_properties.cell(row=row, column=3, value=prop.owner_id or 'غير محدد').border = border
        ws_properties.cell(row=row, column=4, value=prop.address).border = border
        ws_properties.cell(row=row, column=5, value=prop.owner_name).border = border
        ws_properties.cell(row=row, column=6, value=prop.contract_start_date.strftime('%Y-%m-%d')).border = border
        ws_properties.cell(row=row, column=7, value=prop.contract_end_date.strftime('%Y-%m-%d')).border = border
        ws_properties.cell(row=row, column=8, value=prop.remaining_days).border = border
        ws_properties.cell(row=row, column=9, value=f"{prop.annual_rent_amount:,.0f}").border = border
        ws_properties.cell(row=row, column=10, value=f"{prop.annual_rent_amount/12:,.0f}").border = border
        ws_properties.cell(row=row, column=11, value=prop.payment_method or '-').border = border
        
        # تحديد الحالة
        if prop.is_expired:
            status = 'منتهي'
            status_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif prop.is_expiring_soon:
            status = f'قريب من الانتهاء ({prop.remaining_days} يوم)'
            status_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        else:
            status = f'نشط ({prop.remaining_days} يوم)'
            status_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        cell = ws_properties.cell(row=row, column=12, value=status)
        cell.border = border
        cell.fill = status_fill
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
    
    # ضبط عرض الأعمدة
    column_widths = [15, 20, 15, 35, 20, 15, 15, 15, 18, 18, 15, 25]
    for i, width in enumerate(column_widths, start=1):
        ws_properties.column_dimensions[get_column_letter(i)].width = width
    
    # ========== ورقة 3: التجهيزات ==========
    ws_furnishing = wb.create_sheet(title="التجهيزات")
    
    # العنوان
    ws_furnishing.merge_cells('A1:G1')
    cell = ws_furnishing['A1']
    cell.value = "تجهيزات العقارات المستأجرة"
    cell.font = title_font
    cell.fill = PatternFill(start_color='F093FB', end_color='F093FB', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_furnishing.row_dimensions[1].height = 30
    
    # عناوين الأعمدة
    row = 3
    headers = ['رقم العقد', 'اسم العقار', 'جرات الغاز', 'الطباخات', 'الأسرّة', 'البطانيات', 'المخدات']
    furnishing_fill = PatternFill(start_color='F5576C', end_color='F5576C', fill_type='solid')
    
    for col, header in enumerate(headers, start=1):
        cell = ws_furnishing.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = furnishing_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # بيانات التجهيزات
    row = 4
    for prop in properties:
        furnishing = PropertyFurnishing.query.filter_by(property_id=prop.id).first()
        
        ws_furnishing.cell(row=row, column=1, value=prop.contract_number or '-').border = border
        ws_furnishing.cell(row=row, column=2, value=prop.city).border = border
        
        if furnishing:
            ws_furnishing.cell(row=row, column=3, value=furnishing.gas_cylinder or 0).border = border
            ws_furnishing.cell(row=row, column=4, value=furnishing.stoves or 0).border = border
            ws_furnishing.cell(row=row, column=5, value=furnishing.beds or 0).border = border
            ws_furnishing.cell(row=row, column=6, value=furnishing.blankets or 0).border = border
            ws_furnishing.cell(row=row, column=7, value=furnishing.pillows or 0).border = border
        else:
            for col in range(3, 8):
                ws_furnishing.cell(row=row, column=col, value=0).border = border
        
        row += 1
    
    # تجهيزات إضافية
    row += 2
    ws_furnishing.merge_cells(f'A{row}:G{row}')
    cell = ws_furnishing[f'A{row}']
    cell.value = "تجهيزات إضافية وملاحظات"
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.fill = furnishing_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    for prop in properties:
        furnishing = PropertyFurnishing.query.filter_by(property_id=prop.id).first()
        if furnishing and (furnishing.other_items or furnishing.notes):
            ws_furnishing[f'A{row}'] = f"{prop.contract_number or prop.city}:"
            ws_furnishing[f'A{row}'].font = Font(bold=True)
            row += 1
            
            if furnishing.other_items:
                ws_furnishing[f'A{row}'] = f"تجهيزات أخرى: {furnishing.other_items}"
                row += 1
            
            if furnishing.notes:
                ws_furnishing[f'A{row}'] = f"ملاحظات: {furnishing.notes}"
                row += 1
            
            row += 1
    
    # ضبط عرض الأعمدة
    column_widths = [15, 25, 15, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws_furnishing.column_dimensions[get_column_letter(i)].width = width
    
    # ========== ورقة 4: الدفعات ==========
    ws_payments = wb.create_sheet(title="الدفعات")
    
    # العنوان
    ws_payments.merge_cells('A1:H1')
    cell = ws_payments['A1']
    cell.value = "دفعات الإيجار للعقارات"
    cell.font = title_font
    cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_payments.row_dimensions[1].height = 30
    
    # عناوين الأعمدة
    row = 3
    headers = ['رقم العقد', 'اسم العقار', 'التاريخ المتوقع', 'المبلغ', 'الحالة', 'تاريخ الدفع الفعلي', 'طريقة الدفع', 'الرقم المرجعي']
    payment_fill = PatternFill(start_color='764BA2', end_color='764BA2', fill_type='solid')
    
    for col, header in enumerate(headers, start=1):
        cell = ws_payments.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = payment_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # بيانات الدفعات
    row = 4
    for prop in properties:
        payments = PropertyPayment.query.filter_by(property_id=prop.id).order_by(PropertyPayment.payment_date).all()
        
        for payment in payments:
            ws_payments.cell(row=row, column=1, value=prop.contract_number or '-').border = border
            ws_payments.cell(row=row, column=2, value=prop.city).border = border
            ws_payments.cell(row=row, column=3, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
            ws_payments.cell(row=row, column=4, value=f"{payment.amount:,.0f}").border = border
            
            # الحالة مع تلوين
            status_text = {'pending': 'معلق', 'paid': 'مدفوع', 'overdue': 'متأخر'}.get(payment.status, '-')
            cell = ws_payments.cell(row=row, column=5, value=status_text)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            if payment.status == 'paid':
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif payment.status == 'overdue':
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            
            ws_payments.cell(row=row, column=6, value=payment.actual_payment_date.strftime('%Y-%m-%d') if payment.actual_payment_date else '-').border = border
            ws_payments.cell(row=row, column=7, value=payment.payment_method or '-').border = border
            ws_payments.cell(row=row, column=8, value=payment.reference_number or '-').border = border
            
            row += 1
    
    # إضافة تفاصيل الدفعات
    row += 2
    
    # الدفعات المستحقة (القادمة خلال 30 يوم)
    ws_payments.merge_cells(f'A{row}:H{row}')
    cell = ws_payments[f'A{row}']
    cell.value = "الدفعات المستحقة (30 يوم قادمة)"
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='17A2B8', end_color='17A2B8', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    upcoming_payments_date = date.today() + timedelta(days=30)
    upcoming_payments = PropertyPayment.query.filter(
        PropertyPayment.status == 'pending',
        PropertyPayment.payment_date.between(date.today(), upcoming_payments_date)
    ).order_by(PropertyPayment.payment_date).all()
    
    if upcoming_payments:
        for payment in upcoming_payments:
            prop = payment.rental_property
            ws_payments.cell(row=row, column=1, value=prop.contract_number or '-').border = border
            ws_payments.cell(row=row, column=2, value=prop.city).border = border
            ws_payments.cell(row=row, column=3, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
            ws_payments.cell(row=row, column=4, value=f"{payment.amount:,.0f} ريال").border = border
            days_left = (payment.payment_date - date.today()).days
            cell = ws_payments.cell(row=row, column=5, value=f"بعد {days_left} يوم")
            cell.border = border
            cell.fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')
            row += 1
    else:
        ws_payments[f'A{row}'] = "لا توجد دفعات مستحقة خلال 30 يوم"
        ws_payments[f'A{row}'].font = Font(italic=True)
        row += 1
    
    # الدفعات المعلقة
    row += 1
    ws_payments.merge_cells(f'A{row}:H{row}')
    cell = ws_payments[f'A{row}']
    cell.value = "الدفعات المعلقة"
    cell.font = Font(name='Arial', size=11, bold=True, color='000000')
    cell.fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    pending_payments_list = PropertyPayment.query.filter_by(status='pending').order_by(
        PropertyPayment.payment_date
    ).all()
    
    if pending_payments_list:
        for payment in pending_payments_list:
            prop = payment.rental_property
            ws_payments.cell(row=row, column=1, value=prop.contract_number or '-').border = border
            ws_payments.cell(row=row, column=2, value=prop.city).border = border
            ws_payments.cell(row=row, column=3, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
            ws_payments.cell(row=row, column=4, value=f"{payment.amount:,.0f} ريال").border = border
            cell = ws_payments.cell(row=row, column=5, value="معلق")
            cell.border = border
            cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            row += 1
    else:
        ws_payments[f'A{row}'] = "لا توجد دفعات معلقة"
        ws_payments[f'A{row}'].font = Font(italic=True)
        row += 1
    
    # الدفعات المتأخرة
    row += 1
    ws_payments.merge_cells(f'A{row}:H{row}')
    cell = ws_payments[f'A{row}']
    cell.value = "الدفعات المتأخرة"
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    overdue_payments_list = PropertyPayment.query.filter(
        PropertyPayment.status == 'pending',
        PropertyPayment.payment_date < date.today()
    ).order_by(PropertyPayment.payment_date).all()
    
    if overdue_payments_list:
        for payment in overdue_payments_list:
            prop = payment.rental_property
            ws_payments.cell(row=row, column=1, value=prop.contract_number or '-').border = border
            ws_payments.cell(row=row, column=2, value=prop.city).border = border
            ws_payments.cell(row=row, column=3, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
            ws_payments.cell(row=row, column=4, value=f"{payment.amount:,.0f} ريال").border = border
            days_overdue = (date.today() - payment.payment_date).days
            cell = ws_payments.cell(row=row, column=5, value=f"متأخر {days_overdue} يوم")
            cell.border = border
            cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            row += 1
    else:
        ws_payments[f'A{row}'] = "لا توجد دفعات متأخرة"
        ws_payments[f'A{row}'].font = Font(italic=True)
        row += 1
    
    # إضافة صف إحصائيات
    row += 1
    ws_payments.merge_cells(f'A{row}:H{row}')
    cell = ws_payments[f'A{row}']
    cell.value = "إحصائيات الدفعات"
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.fill = payment_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # حساب الإحصائيات
    all_payments = PropertyPayment.query.all()
    total_payments = len(all_payments)
    paid_payments = sum(1 for p in all_payments if p.status == 'paid')
    pending_payments_count = sum(1 for p in all_payments if p.status == 'pending')
    overdue_payments_count = sum(1 for p in all_payments if p.status == 'pending' and p.payment_date < date.today())
    total_amount = sum(p.amount for p in all_payments if p.status == 'paid')
    
    stats = [
        ['إجمالي الدفعات', total_payments],
        ['دفعات مدفوعة', paid_payments],
        ['دفعات معلقة', pending_payments_count],
        ['دفعات متأخرة', overdue_payments_count],
        ['إجمالي المبالغ المدفوعة', f"{total_amount:,.0f} ريال"]
    ]
    
    for stat in stats:
        ws_payments[f'A{row}'] = stat[0]
        ws_payments[f'A{row}'].font = Font(bold=True)
        ws_payments[f'A{row}'].fill = info_fill
        ws_payments[f'A{row}'].border = border
        
        ws_payments[f'B{row}'] = stat[1]
        ws_payments[f'B{row}'].border = border
        ws_payments[f'B{row}'].alignment = Alignment(horizontal='center')
        row += 1
    
    # ضبط عرض الأعمدة
    column_widths = [15, 25, 18, 18, 15, 18, 18, 18]
    for i, width in enumerate(column_widths, start=1):
        ws_payments.column_dimensions[get_column_letter(i)].width = width
    
    # حفظ في الذاكرة
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # اسم الملف
    filename = f"تقرير_العقارات_الشامل_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    # تسجيل النشاط
    log_activity(
        action='تصدير تقرير العقارات',
        entity_type='RentalProperty',
        entity_id=0,
        details=f'تم تصدير تقرير شامل لجميع العقارات ({total_properties} عقار)'
    )
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@properties_bp.route('/<int:property_id>/manage-residents')
@login_required
def manage_residents(property_id):
    """صفحة إدارة الموظفين القاطنين في العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # جلب جميع الأقسام
    departments = Department.query.order_by(Department.name).all()
    
    # جلب جميع الموظفين مع أقسامهم
    employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
    
    # تنظيم الموظفين حسب القسم
    employees_by_department = {}
    for dept in departments:
        dept_employees = [emp for emp in employees if emp.department_id == dept.id]
        if dept_employees:
            employees_by_department[dept.name] = dept_employees
    
    # موظفون بدون قسم
    no_dept_employees = [emp for emp in employees if emp.department_id is None]
    if no_dept_employees:
        employees_by_department['بدون قسم'] = no_dept_employees
    
    # الموظفين الحاليين في العقار
    current_residents = property.residents
    
    return render_template('properties/manage_residents.html',
                         property=property,
                         employees_by_department=employees_by_department,
                         current_residents=current_residents,
                         departments=departments)


@properties_bp.route('/<int:property_id>/add-resident-page')
@login_required
def add_resident_page(property_id):
    """صفحة إضافة موظف قاطن"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # جلب جميع الأقسام
    departments = Department.query.order_by(Department.name).all()
    
    # جلب جميع الموظفين النشطين
    employees = Employee.query.filter_by(status='active').order_by(Employee.name).all()
    
    # الموظفين المضافين مسبقاً
    current_resident_ids = [r.id for r in property.residents]
    
    # تنظيم الموظفين حسب الأقسام باستخدام العلاقة many-to-many
    employees_by_dept = {}
    for dept in departments:
        dept_employees = [emp for emp in employees if dept in emp.departments]
        if dept_employees:
            employees_by_dept[dept.id] = dept_employees
    
    # الموظفين بدون قسم
    no_dept_employees = [emp for emp in employees if not emp.departments]
    
    return render_template('properties/add_resident.html',
                         property=property,
                         departments=departments,
                         employees=employees,
                         current_resident_ids=current_resident_ids,
                         employees_by_dept=employees_by_dept,
                         no_dept_employees=no_dept_employees)


@properties_bp.route('/<int:property_id>/add-resident', methods=['POST'])
@login_required
def add_resident(property_id):
    """إضافة موظفين إلى العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # قبول عدة موظفين
    employee_ids = request.form.getlist('employee_ids')
    
    if employee_ids:
        added_count = 0
        employee_names = []
        
        for employee_id in employee_ids:
            employee = Employee.query.get_or_404(employee_id)
            
            # التحقق من عدم وجود الموظف مسبقاً
            if employee not in property.residents:
                property.residents.append(employee)
                added_count += 1
                employee_names.append(employee.name)
        
        db.session.commit()
        
        if added_count > 0:
            log_activity(
                action='إضافة موظفين للعقار',
                entity_type='RentalProperty',
                entity_id=property_id,
                details=f'تم إضافة {added_count} موظف للعقار {property.city}: {", ".join(employee_names[:5])}'
            )
            
            flash(f'تم إضافة {added_count} موظف للعقار بنجاح', 'success')
        else:
            flash('جميع الموظفين المحددين مضافين مسبقاً للعقار', 'warning')
    else:
        flash('الرجاء تحديد موظف واحد على الأقل', 'warning')
    
    return redirect(url_for('properties.view', property_id=property_id))


@properties_bp.route('/<int:property_id>/add-department-page')
@login_required
def add_department_page(property_id):
    """صفحة إضافة قسم كامل"""
    property = RentalProperty.query.get_or_404(property_id)
    
    # جلب جميع الأقسام
    departments = Department.query.order_by(Department.name).all()
    
    # جلب جميع الموظفين النشطين
    employees = Employee.query.filter_by(status='active').all()
    
    # حساب عدد الموظفين لكل قسم (باستخدام العلاقة many-to-many)
    dept_employee_counts = {}
    for dept in departments:
        # حساب الموظفين النشطين المرتبطين بهذا القسم عبر employee_departments
        count = Employee.query.join(
            Employee.departments
        ).filter(
            Department.id == dept.id,
            Employee.status == 'active'
        ).count()
        dept_employee_counts[dept.id] = count
    
    return render_template('properties/add_department.html',
                         property=property,
                         departments=departments,
                         employees=employees,
                         dept_employee_counts=dept_employee_counts)


@properties_bp.route('/<int:property_id>/add-department-residents', methods=['POST'])
@login_required
def add_department_residents(property_id):
    """إضافة جميع موظفي قسم للعقار"""
    property = RentalProperty.query.get_or_404(property_id)
    
    department_id = request.form.get('department_id')
    
    if department_id:
        department = Department.query.get_or_404(department_id)
        
        # جلب جميع موظفي القسم النشطين (باستخدام العلاقة many-to-many)
        dept_employees = Employee.query.join(
            Employee.departments
        ).filter(
            Department.id == department_id,
            Employee.status == 'active'
        ).all()
        
        added_count = 0
        for employee in dept_employees:
            if employee not in property.residents:
                property.residents.append(employee)
                added_count += 1
        
        db.session.commit()
        
        log_activity(
            action='إضافة قسم كامل للعقار',
            entity_type='RentalProperty',
            entity_id=property_id,
            details=f'تم إضافة {added_count} موظف من قسم {department.name} للعقار {property.city}'
        )
        
        flash(f'تم إضافة {added_count} موظف من قسم {department.name} للعقار', 'success')
    
    return redirect(url_for('properties.view', property_id=property_id))


@properties_bp.route('/<int:property_id>/remove-resident/<int:employee_id>', methods=['POST'])
@login_required
def remove_resident(property_id, employee_id):
    """إزالة موظف من العقار"""
    property = RentalProperty.query.get_or_404(property_id)
    employee = Employee.query.get_or_404(employee_id)
    
    if employee in property.residents:
        property.residents.remove(employee)
        db.session.commit()
        
        log_activity(
            action='إزالة موظف من العقار',
            entity_type='RentalProperty',
            entity_id=property_id,
            details=f'تم إزالة الموظف {employee.name} من العقار {property.city}'
        )
        
        flash(f'تم إزالة الموظف {employee.name} من العقار', 'success')
    else:
        flash('الموظف غير موجود في العقار', 'warning')
    
    return redirect(url_for('properties.view', property_id=property_id))
