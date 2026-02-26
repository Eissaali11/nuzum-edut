from flask import Blueprint, render_template, request, flash, redirect, url_for, current_app, jsonify
from flask_login import login_required, current_user
from models import MobileDevice, SimCard, Employee, Department, UserRole, DeviceAssignment
from src.core.extensions import db
from datetime import datetime
import logging
from src.utils.audit_logger import log_activity

device_assignment_bp = Blueprint('device_assignment', __name__)

@device_assignment_bp.route('/departments')
@login_required
def departments_view():
    """صفحة عرض الأجهزة والأرقام حسب الأقسام"""
    try:
        # جلب جميع الأقسام
        departments = Department.query.order_by(Department.name).all()
        
        # إضافة إحصائيات لكل قسم
        for department in departments:
            # عدد الموظفين في القسم
            department.employee_count = len(department.employees)
            
            # حساب عدد الأجهزة والأرقام لموظفي القسم
            device_count = 0
            sim_count = 0
            
            for employee in department.employees:
                # عدد الأجهزة المحمولة
                devices = MobileDevice.query.filter_by(employee_id=employee.id).all()
                device_count += len(devices)
                
                # عدد أرقام SIM من نظام إدارة SIM
                sims = SimCard.query.filter_by(employee_id=employee.id).all()
                sim_count += len(sims)
                
                # إضافة الأجهزة والأرقام للموظف للعرض
                employee.mobile_devices = devices
                employee.imported_phone_numbers = sims
            
            department.device_count = device_count
            department.sim_count = sim_count
            
            # حساب نسب التغطية
            if department.employee_count > 0:
                department.device_coverage = round((device_count / department.employee_count) * 100, 1)
                department.sim_coverage = round((sim_count / department.employee_count) * 100, 1)
            else:
                department.device_coverage = 0
                department.sim_coverage = 0
        
        # إحصائيات عامة
        stats = {
            'total_departments': len(departments),
            'total_devices': MobileDevice.query.count(),
            'total_sims': SimCard.query.count(),
            'total_employees': Employee.query.count(),
        }
        
        return render_template('device_assignment/departments_view.html',
                             departments=departments,
                             stats=stats)
        
    except Exception as e:
        flash(f'حدث خطأ أثناء تحميل البيانات: {str(e)}', 'danger')
        return redirect(url_for('device_assignment.index'))

@device_assignment_bp.route('/')
@login_required
def index():
    """صفحة ربط الأجهزة والأرقام بالموظفين"""
    try:
        # فلاتر البحث
        department_filter = request.args.get('department_id', '')
        device_status = request.args.get('device_status', '')
        sim_status = request.args.get('sim_status', '')
        employee_search = request.args.get('employee_search', '')
        phone_search = request.args.get('phone_search', '')
        
        # الحصول على البيانات الأساسية
        departments = Department.query.order_by(Department.name).all()
        
        # فلترة الموظفين
        employees_query = Employee.query
        if department_filter:
            employees_query = employees_query.join(Employee.departments).filter(Department.id == department_filter)
        if employee_search:
            employees_query = employees_query.filter(
                Employee.name.contains(employee_search) | 
                Employee.employee_id.contains(employee_search)
            )
        employees = employees_query.order_by(Employee.name).all()
        
        # الأجهزة المتاحة (غير مرتبطة في DeviceAssignment نشط)
        # جلب IDs الأجهزة المربوطة حالياً
        assigned_device_ids = db.session.query(DeviceAssignment.device_id).filter(
            DeviceAssignment.is_active == True,
            DeviceAssignment.device_id.isnot(None)
        ).distinct().all()
        assigned_device_ids = [id[0] for id in assigned_device_ids]
        
        available_devices_query = MobileDevice.query
        if device_status == 'available':
            # الأجهزة غير المربوطة
            if assigned_device_ids:
                available_devices_query = available_devices_query.filter(~MobileDevice.id.in_(assigned_device_ids))
        elif device_status == 'assigned':
            # الأجهزة المربوطة
            if assigned_device_ids:
                available_devices_query = available_devices_query.filter(MobileDevice.id.in_(assigned_device_ids))
            else:
                available_devices_query = available_devices_query.filter(False)  # لا توجد أجهزة مربوطة
        else:
            # عرض جميع الأجهزة مع تمييز المربوطة
            pass
            
        available_devices = available_devices_query.order_by(MobileDevice.device_brand, MobileDevice.device_model).all()
        
        # إضافة معلومة هل الجهاز مربوط أم لا
        for device in available_devices:
            device.is_assigned = device.id in assigned_device_ids
        
        # الأرقام المتاحة - استخدام SimCard من نظام إدارة بطاقات SIM فقط
        from models import SimCard
        
        # فلترة الأرقام حسب الحالة
        if sim_status == 'assigned':
            # عرض الأرقام المربوطة من نظام SIM
            available_sims_query = SimCard.query.filter(
                SimCard.employee_id.isnot(None)
            )
        elif sim_status == 'available':
            # عرض الأرقام المتاحة (غير مربوطة) فقط
            available_sims_query = SimCard.query.filter(
                SimCard.employee_id.is_(None)
            )
        else:
            # الافتراضي: عرض الأرقام المتاحة فقط (كما في صفحة إدارة SIM)
            available_sims_query = SimCard.query.filter(
                SimCard.employee_id.is_(None)
            )
        
        # فلترة حسب رقم الهاتف
        if phone_search:
            available_sims_query = available_sims_query.filter(
                SimCard.phone_number.contains(phone_search)
            )
            
        available_sims = available_sims_query.order_by(SimCard.phone_number).all()
        
        # إضافة معلومة هل الرقم مربوط أم لا
        for sim in available_sims:
            sim.is_assigned = sim.employee_id is not None
        
        # عمليات الربط النشطة
        active_assignments = DeviceAssignment.query.filter(
            DeviceAssignment.is_active == True
        ).order_by(DeviceAssignment.assignment_date.desc()).limit(10).all()
        
        # إحصائيات محدثة تعتمد على نظام إدارة بطاقات SIM فقط
        total_sims = SimCard.query.count()
        assigned_sims_count = SimCard.query.filter(SimCard.employee_id.isnot(None)).count()
        available_sims_count = total_sims - assigned_sims_count
        
        stats = {
            'total_devices': MobileDevice.query.count(),
            'available_devices': MobileDevice.query.count() - len(assigned_device_ids),
            'assigned_devices': len(assigned_device_ids),
            'total_sims': total_sims,
            'available_sims': available_sims_count,
            'assigned_sims': assigned_sims_count,
        }
        
        return render_template('device_assignment/index.html',
                             departments=departments,
                             employees=employees,
                             available_devices=available_devices,
                             available_sims=available_sims,
                             active_assignments=active_assignments,
                             stats=stats,
                             department_filter=department_filter,
                             device_status=device_status,
                             sim_status=sim_status,
                             employee_search=employee_search,
                             phone_search=phone_search)
    
    except Exception as e:
        current_app.logger.error(f"Error in device_assignment index: {str(e)}")
        flash('حدث خطأ أثناء تحميل البيانات', 'danger')
        return render_template('device_assignment/index.html',
                             departments=[],
                             employees=[],
                             available_devices=[],
                             available_sims=[],
                             active_assignments=[],
                             stats={},
                             department_filter='',
                             device_status='',
                             sim_status='',
                             employee_search='',
                             phone_search='')

@device_assignment_bp.route('/assign', methods=['POST'])
@login_required
def assign():
    """ربط جهاز و/أو رقم بموظف"""
    try:
        assignment_type = request.form.get('assignment_type')
        employee_id = request.form.get('employee_id')
        device_id = request.form.get('device_id') if request.form.get('device_id') else None
        sim_id = request.form.get('sim_id') if request.form.get('sim_id') else None
        notes = request.form.get('notes', '').strip()
        
        # التحقق من البيانات المطلوبة
        if not assignment_type or not employee_id:
            flash('يرجى اختيار نوع الربط والموظف', 'danger')
            return redirect(url_for('device_assignment.index'))
        
        employee = Employee.query.get_or_404(employee_id)
        
        # التحقق من حالة الموظف أولاً
        if employee.status not in ['نشط', 'active']:
            flash(f'لا يمكن ربط أجهزة أو أرقام بالموظف {employee.name} لأن حالته غير نشطة', 'warning')
            return redirect(url_for('device_assignment.index'))
        
        # التحقق من نوع الربط والمعلومات المطلوبة
        if assignment_type == 'device_only' and not device_id:
            flash('يرجى اختيار جهاز للربط', 'danger')
            return redirect(url_for('device_assignment.index'))
        elif assignment_type == 'sim_only' and not sim_id:
            flash('يرجى اختيار رقم للربط', 'danger')
            return redirect(url_for('device_assignment.index'))
        elif assignment_type == 'device_and_sim' and (not device_id or not sim_id):
            flash('يرجى اختيار جهاز ورقم للربط', 'danger')
            return redirect(url_for('device_assignment.index'))
        
        device = None
        sim_card = None
        
        # التحقق من الجهاز إذا تم تحديده
        if device_id:
            device = MobileDevice.query.get_or_404(device_id)
            # فحص إذا كان الجهاز مربوط حالياً في DeviceAssignment
            existing_assignment = DeviceAssignment.query.filter_by(
                device_id=device_id, 
                is_active=True
            ).first()
            if existing_assignment:
                flash('هذا الجهاز مربوط بموظف آخر بالفعل', 'danger')
                return redirect(url_for('device_assignment.index'))
        
        # التحقق من الرقم إذا تم تحديده
        if sim_id:
            sim_card = SimCard.query.get_or_404(sim_id)
            if sim_card.employee_id:
                flash('هذا الرقم مربوط بموظف آخر بالفعل', 'danger')
                return redirect(url_for('device_assignment.index'))
        
        # إنشاء سجل الربط
        assignment = DeviceAssignment(
            employee_id=employee_id,
            device_id=device_id,
            sim_card_id=sim_id,
            assignment_date=datetime.now(),
            assignment_type=assignment_type,
            notes=notes,
            assigned_by=current_user.id,
            is_active=True
        )
        
        db.session.add(assignment)
        
        # تحديث حالة الجهاز (للتوافق مع النظام القديم)
        if device:
            device.employee_id = employee_id
            device.status = 'مربوط'
        
        # تحديث حالة الرقم
        if sim_card:
            sim_card.employee_id = employee_id
            sim_card.assignment_date = datetime.now()
            sim_card.status = 'مربوط'
        
        db.session.commit()
        
        # تسجيل العملية
        action_details = []
        if device:
            action_details.append(f"جهاز: {device.imei}")
        if sim_card:
            action_details.append(f"رقم: {sim_card.phone_number}")
        
        log_activity(
            action="create",
            entity_type="DeviceAssignment",
            entity_id=assignment.id,
            details=f"ربط {assignment_type} بموظف - الموظف: {employee.name} ({employee.employee_id}), {', '.join(action_details)}"
        )
        
        success_message = f"تم ربط "
        if device and sim_card:
            success_message += f"الجهاز {device.imei} والرقم {sim_card.phone_number}"
        elif device:
            success_message += f"الجهاز {device.imei}"
        elif sim_card:
            success_message += f"الرقم {sim_card.phone_number}"
        success_message += f" بالموظف {employee.name} بنجاح"
        
        flash(success_message, 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in device assignment: {str(e)}")
        flash('حدث خطأ أثناء عملية الربط', 'danger')
    
    return redirect(url_for('device_assignment.index'))

@device_assignment_bp.route('/unassign/<int:assignment_id>', methods=['POST'])
@login_required
def unassign(assignment_id):
    """فك ربط جهاز و/أو رقم من موظف"""
    try:
        assignment = DeviceAssignment.query.get_or_404(assignment_id)
        reason = request.form.get('reason', 'فك ربط يدوي')
        
        if not assignment.is_active:
            flash('هذا الربط غير نشط بالفعل', 'danger')
            return redirect(url_for('device_assignment.index'))
        
        # الحصول على معلومات العملية قبل فك الربط
        employee = assignment.employee
        device = assignment.device
        sim_card = assignment.sim_card
        
        # فك ربط الجهاز (عكس التحديثات السابقة)
        if device:
            device.employee_id = None
            device.status = 'متاح'
        
        # فك ربط الرقم
        if sim_card:
            sim_card.employee_id = None
            sim_card.assignment_date = None
            sim_card.status = 'متاح'
        
        # تحديث سجل الربط
        assignment.is_active = False
        assignment.unassignment_date = datetime.now()
        assignment.unassignment_reason = reason
        assignment.unassigned_by = current_user.id
        
        db.session.commit()
        
        # تسجيل العملية
        action_details = []
        if device:
            action_details.append(f"جهاز: {device.imei}")
        if sim_card:
            action_details.append(f"رقم: {sim_card.phone_number}")
        
        log_activity(
            action="update",
            entity_type="DeviceAssignment",
            entity_id=assignment_id,
            details=f"فك ربط من موظف - الموظف: {employee.name} ({employee.employee_id}), {', '.join(action_details)}, السبب: {reason}"
        )
        
        flash(f'تم فك الربط من الموظف {employee.name} بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in device unassignment: {str(e)}")
        flash('حدث خطأ أثناء فك الربط', 'danger')
    
    return redirect(url_for('device_assignment.index'))

@device_assignment_bp.route('/confirm-unassign/<int:assignment_id>')
@login_required
def confirm_unassign(assignment_id):
    """صفحة تأكيد فك الربط"""
    try:
        assignment = DeviceAssignment.query.get_or_404(assignment_id)
        
        # التأكد من أن الربط نشط
        if not assignment.is_active:
            flash('هذا الربط غير نشط بالفعل', 'warning')
            return redirect(url_for('device_assignment.index'))
            
        return render_template('device_assignment/confirm_unassign.html', assignment=assignment)
        
    except Exception as e:
        current_app.logger.error(f"Error in confirm_unassign: {str(e)}")
        flash('حدث خطأ أثناء تحميل صفحة التأكيد', 'danger')
        return redirect(url_for('device_assignment.index'))

@device_assignment_bp.route('/history')
@login_required
def history():
    """تاريخ عمليات الربط"""
    try:
        # فلاتر البحث
        employee_filter = request.args.get('employee_id', '')
        status_filter = request.args.get('status', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # استعلام أساسي
        query = DeviceAssignment.query
        
        # تطبيق فلاتر
        if employee_filter:
            query = query.filter(DeviceAssignment.employee_id == employee_filter)
        
        if status_filter == 'active':
            query = query.filter(DeviceAssignment.is_active == True)
        elif status_filter == 'inactive':
            query = query.filter(DeviceAssignment.is_active == False)
        
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                query = query.filter(DeviceAssignment.assignment_date >= date_from_obj)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                query = query.filter(DeviceAssignment.assignment_date <= date_to_obj)
            except ValueError:
                pass
        
        # ترتيب البيانات
        assignments = query.order_by(DeviceAssignment.assignment_date.desc()).all()
        
        # قائمة الموظفين للفلترة
        employees_with_assignments = Employee.query.join(DeviceAssignment).distinct().order_by(Employee.name).all()
        
        # إحصائيات
        stats = {
            'total_assignments': DeviceAssignment.query.count(),
            'active_assignments': DeviceAssignment.query.filter(DeviceAssignment.is_active == True).count(),
            'completed_assignments': DeviceAssignment.query.filter(DeviceAssignment.is_active == False).count(),
        }
        
        return render_template('device_assignment/history.html',
                             assignments=assignments,
                             employees=employees_with_assignments,
                             stats=stats,
                             employee_filter=employee_filter,
                             status_filter=status_filter,
                             date_from=date_from,
                             date_to=date_to)
    
    except Exception as e:
        current_app.logger.error(f"Error in assignment history: {str(e)}")
        flash('حدث خطأ أثناء تحميل التاريخ', 'danger')
        return render_template('device_assignment/history.html',
                             assignments=[],
                             employees=[],
                             stats={},
                             employee_filter='',
                             status_filter='',
                             date_from='',
                             date_to='')

@device_assignment_bp.route('/api/employee/<int:employee_id>')
@login_required
def api_employee_assignments(employee_id):
    """API للحصول على ربطات الموظف"""
    try:
        employee = Employee.query.get_or_404(employee_id)
        
        # الربطات النشطة
        active_assignments = DeviceAssignment.query.filter(
            DeviceAssignment.employee_id == employee_id,
            DeviceAssignment.is_active == True
        ).all()
        
        # الأجهزة المربوطة
        devices = []
        for assignment in active_assignments:
            if assignment.device:
                devices.append({
                    'id': assignment.device.id,
                    'phone_number': assignment.device.phone_number,
                    'imei': assignment.device.imei,
                    'brand': assignment.device.device_brand,
                    'model': assignment.device.device_model,
                    'assignment_date': assignment.assignment_date.strftime('%Y-%m-%d') if assignment.assignment_date else None
                })
        
        # الأرقام المربوطة
        sim_cards = []
        for assignment in active_assignments:
            if assignment.sim_card:
                sim_cards.append({
                    'id': assignment.sim_card.id,
                    'phone_number': assignment.sim_card.phone_number,
                    'carrier': assignment.sim_card.carrier,
                    'plan_type': assignment.sim_card.plan_type,
                    'assignment_date': assignment.assignment_date.strftime('%Y-%m-%d') if assignment.assignment_date else None
                })
        
        return jsonify({
            'success': True,
            'employee': {
                'id': employee.id,
                'name': employee.name,
                'employee_id': employee.employee_id,
                'mobile': employee.mobile
            },
            'devices': devices,
            'sim_cards': sim_cards
        })
        
    except Exception as e:
        current_app.logger.error(f"Error getting employee assignments: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'حدث خطأ أثناء جلب البيانات'
        }), 500

@device_assignment_bp.route('/department-management')
@login_required
def department_management():
    """صفحة إدارة ربط الأجهزة والأرقام بالأقسام مباشرة"""
    try:
        departments = Department.query.all()
        
        # جلب الأجهزة المتاحة (غير مربوطة بموظف أو قسم)
        available_devices = MobileDevice.query.filter(
            MobileDevice.employee_id.is_(None),
            MobileDevice.department_id.is_(None)
        ).all()
        
        # جلب الأرقام المتاحة (غير مربوطة بموظف) من نظام إدارة SIM
        available_sims = SimCard.query.filter_by(employee_id=None).all()
        
        return render_template('device_assignment/department_management.html',
                             departments=departments,
                             available_devices=available_devices,
                             available_sims=available_sims)
    
    except Exception as e:
        current_app.logger.error(f"Error in department management: {str(e)}")
        flash('حدث خطأ في تحميل البيانات', 'danger')
        return redirect(url_for('device_assignment.index'))

@device_assignment_bp.route('/assign-device/<int:device_id>')
@login_required
def assign_device(device_id):
    """صفحة ربط جهاز بموظف"""
    try:
        # جلب الجهاز
        device = MobileDevice.query.get_or_404(device_id)
        
        # التحقق من أن الجهاز غير مربوط
        if device.is_assigned:
            flash('هذا الجهاز مربوط بالفعل بموظف آخر', 'warning')
            return redirect(url_for('device_management.index'))
        
        # جلب جميع الأقسام للفلترة
        departments = Department.query.order_by(Department.name).all()
        
        # جلب الموظفين النشطين
        employees = Employee.query.order_by(Employee.name).all()
        
        # جلب الأرقام المتاحة (غير مربوطة) من نظام إدارة SIM
        available_sims = SimCard.query.filter(
            SimCard.employee_id.is_(None)
        ).all()
        
        return render_template('device_assignment/assign_device.html',
                             device=device,
                             departments=departments,
                             employees=employees,
                             available_sims=available_sims)
                             
    except Exception as e:
        current_app.logger.error(f"Error in assign device: {str(e)}")
        flash('حدث خطأ في تحميل الصفحة', 'danger')
        return redirect(url_for('device_management.index'))

@device_assignment_bp.route('/assign-device/<int:device_id>', methods=['POST'])
@login_required
def process_assign_device(device_id):
    """معالجة ربط الجهاز بموظف"""
    try:
        device = MobileDevice.query.get_or_404(device_id)
        
        # جلب بيانات النموذج
        employee_id = request.form.get('employee_id')
        assignment_type = request.form.get('assignment_type')
        sim_card_id = request.form.get('sim_card_id')
        notes = request.form.get('notes', '')
        
        if not employee_id:
            flash('يجب اختيار موظف', 'error')
            return redirect(url_for('device_assignment.assign_device', device_id=device_id))
        
        employee = Employee.query.get(employee_id)
        if not employee:
            flash('الموظف المحدد غير موجود', 'error')
            return redirect(url_for('device_assignment.assign_device', device_id=device_id))
        
        # التحقق من حالة الموظف قبل الربط
        print(f"DEBUG: محاولة ربط الجهاز {device_id} بالموظف {employee.name} - حالة الموظف: {employee.status}")
        if employee.status not in ['نشط', 'active']:
            print(f"DEBUG: تم رفض الربط - الموظف {employee.name} غير نشط")
            flash(f'لا يمكن ربط الجهاز بالموظف {employee.name} لأن حالته غير نشطة', 'warning')
            return redirect(url_for('device_assignment.assign_device', device_id=device_id))
        
        # إنشاء عملية ربط جديدة
        assignment = DeviceAssignment()
        assignment.employee_id = employee_id
        assignment.device_id = device_id
        assignment.assignment_type = assignment_type
        assignment.assignment_date = datetime.utcnow()
        assignment.is_active = True
        assignment.notes = notes
        assignment.assigned_by = current_user.id
        
        # ربط الرقم إذا تم اختياره
        if sim_card_id and assignment_type in ['sim_only', 'device_and_sim']:
            sim_card = SimCard.query.get(sim_card_id)
            if sim_card and not sim_card.employee_id:
                assignment.sim_card_id = sim_card_id
                sim_card.employee_id = employee_id
                sim_card.status = 'مربوطة'
        
        # تحديث بيانات الجهاز
        device.employee_id = employee_id
        device.status = 'مربوط'
        device.updated_at = datetime.utcnow()
        
        db.session.add(assignment)
        db.session.commit()
        
        # تسجيل العملية
        log_activity(
            action='device_assigned',
            entity_type='DeviceAssignment',
            entity_id=assignment.id,
            details=f'تم ربط الجهاز {device.device_brand} {device.device_model} بالموظف {employee.name}'
        )
        
        flash(f'تم ربط الجهاز بالموظف {employee.name} بنجاح', 'success')
        return redirect(url_for('device_management.index'))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error processing device assignment: {str(e)}")
        flash(f'حدث خطأ في ربط الجهاز: {str(e)}', 'error')
        return redirect(url_for('device_assignment.assign_device', device_id=device_id))

@device_assignment_bp.route('/assign-to-department', methods=['POST'])
@login_required
def assign_to_department():
    """ربط جهاز و/أو رقم بقسم مباشرة"""
    try:
        assignment_type = request.form.get('assignment_type')
        department_id = request.form.get('department_id')
        device_id = request.form.get('device_id') if request.form.get('device_id') else None
        sim_id = request.form.get('sim_id') if request.form.get('sim_id') else None
        notes = request.form.get('notes', '').strip()
        
        # التحقق من البيانات المطلوبة
        if not assignment_type or not department_id:
            flash('يرجى اختيار نوع الربط والقسم', 'danger')
            return redirect(url_for('device_assignment.department_management'))
        
        department = Department.query.get_or_404(department_id)
        
        # التحقق من نوع الربط والمعلومات المطلوبة
        if assignment_type == 'device_only' and not device_id:
            flash('يرجى اختيار جهاز للربط', 'danger')
            return redirect(url_for('device_assignment.department_management'))
        elif assignment_type == 'sim_only' and not sim_id:
            flash('يرجى اختيار رقم للربط', 'danger')
            return redirect(url_for('device_assignment.department_management'))
        elif assignment_type == 'device_and_sim' and (not device_id or not sim_id):
            flash('يرجى اختيار جهاز ورقم للربط', 'danger')
            return redirect(url_for('device_assignment.department_management'))
        
        # ربط الجهاز بالقسم
        if device_id:
            device = MobileDevice.query.get_or_404(device_id)
            if device.employee_id or device.department_id:
                flash('هذا الجهاز مربوط بالفعل', 'danger')
                return redirect(url_for('device_assignment.department_management'))
            
            device.department_id = department_id
            device.department_assignment_date = datetime.now()
            device.assignment_type = 'department'
            device.status = 'مرتبط'
            if notes:
                device.notes = notes
        
        # ربط الرقم (نحتاج لإنشاء DeviceAssignment)
        if sim_id:
            sim_card = SimCard.query.get_or_404(sim_id)
            if sim_card.employee_id:
                flash('هذا الرقم مربوط بموظف آخر بالفعل', 'danger')
                return redirect(url_for('device_assignment.department_management'))
            
            # إنشاء سجل ربط للقسم
            assignment = DeviceAssignment(
                department_id=department_id,
                device_id=device_id,
                sim_card_id=sim_id,
                assignment_date=datetime.now(),
                assignment_type=assignment_type,
                notes=notes,
                assigned_by=current_user.id,
                is_active=True
            )
            db.session.add(assignment)
            
            sim_card.status = 'مربوط'
        
        db.session.commit()
        
        # رسالة النجاح
        success_message = f"تم ربط "
        if device_id and sim_id:
            success_message += f"الجهاز والرقم"
        elif device_id:
            success_message += f"الجهاز"
        elif sim_id:
            success_message += f"الرقم"
        success_message += f" بقسم {department.name} بنجاح"
        
        flash(success_message, 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in department assignment: {str(e)}")
        flash('حدث خطأ أثناء عملية الربط', 'danger')
    
    return redirect(url_for('device_assignment.department_management'))

@device_assignment_bp.route('/unassign-from-department/<int:device_id>', methods=['POST'])
@login_required
def unassign_from_department(device_id):
    """فك ربط جهاز من قسم"""
    try:
        device = MobileDevice.query.get_or_404(device_id)
        
        if not device.department_id:
            flash('هذا الجهاز غير مربوط بأي قسم', 'danger')
            return redirect(url_for('device_assignment.department_management'))
        
        department_name = device.department.name if device.department else "غير محدد"
        
        # فك الربط
        device.department_id = None
        device.department_assignment_date = None
        device.assignment_type = 'employee'
        device.status = 'متاح'
        
        # فك ربط أي سجلات DeviceAssignment مرتبطة
        assignments = DeviceAssignment.query.filter_by(device_id=device_id, is_active=True).all()
        for assignment in assignments:
            if assignment.department_id:
                assignment.is_active = False
                if assignment.sim_card:
                    assignment.sim_card.status = 'متاح'
        
        db.session.commit()
        
        flash(f'تم فك ربط الجهاز من قسم {department_name} بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in department unassignment: {str(e)}")
        flash('حدث خطأ أثناء فك الربط', 'danger')
    
    return redirect(url_for('device_assignment.department_management'))

# مسارات التقارير والتصدير
@device_assignment_bp.route('/reports')
@login_required
def reports():
    """صفحة التقارير والإحصائيات"""
    try:
        # إحصائيات الأجهزة
        total_devices = MobileDevice.query.count()
        assigned_devices_count = DeviceAssignment.query.filter_by(is_active=True).filter(DeviceAssignment.device_id.isnot(None)).count()
        available_devices_count = total_devices - assigned_devices_count
        
        # إحصائيات الأرقام
        total_sims = SimCard.query.count()
        assigned_sims_count = SimCard.query.filter(SimCard.employee_id.isnot(None)).count()
        available_sims_count = total_sims - assigned_sims_count
        
        # قوائم الأجهزة المربوطة والمتاحة
        assigned_device_ids = db.session.query(DeviceAssignment.device_id).filter_by(is_active=True).subquery()
        assigned_devices_list = db.session.query(MobileDevice, DeviceAssignment, Employee).join(
            DeviceAssignment, MobileDevice.id == DeviceAssignment.device_id
        ).outerjoin(Employee, DeviceAssignment.employee_id == Employee.id).filter(
            DeviceAssignment.is_active == True
        ).all()
        
        available_devices_list = MobileDevice.query.filter(~MobileDevice.id.in_(assigned_device_ids)).all()
        
        # قوائم الأرقام المربوطة والمتاحة
        assigned_sims_list = db.session.query(SimCard, Employee).join(
            Employee, SimCard.employee_id == Employee.id
        ).all()
        
        available_sims_list = SimCard.query.filter(SimCard.employee_id.is_(None)).all()
        
        # إحصائيات الأقسام
        departments = Department.query.all()
        department_stats = []
        
        for dept in departments:
            # أجهزة مربوطة بموظفين في هذا القسم
            dept_devices = db.session.query(DeviceAssignment).join(Employee).filter(
                DeviceAssignment.is_active == True,
                DeviceAssignment.device_id.isnot(None),
                Employee.departments.any(Department.id == dept.id)
            ).count()
            
            # أرقام مربوطة بموظفين في هذا القسم
            dept_sims = db.session.query(SimCard).join(Employee).filter(
                SimCard.employee_id.isnot(None),
                Employee.departments.any(Department.id == dept.id)
            ).count()
            
            # أجهزة مربوطة بالقسم مباشرة
            direct_devices = MobileDevice.query.filter_by(department_id=dept.id).count()
            
            department_stats.append({
                'name': dept.name,
                'devices': dept_devices + direct_devices,
                'sims': dept_sims
            })
        
        return render_template('device_assignment/reports.html',
                             total_devices=total_devices,
                             assigned_devices=assigned_devices_count,
                             available_devices=available_devices_count,
                             total_sims=total_sims,
                             assigned_sims=assigned_sims_count,
                             available_sims=available_sims_count,
                             assigned_devices_list=assigned_devices_list,
                             available_devices_list=available_devices_list,
                             assigned_sims_list=assigned_sims_list,
                             available_sims_list=available_sims_list,
                             department_stats=department_stats)
    
    except Exception as e:
        current_app.logger.error(f"Error in reports: {str(e)}")
        flash('حدث خطأ في تحميل التقارير', 'danger')
        return redirect(url_for('device_assignment.index'))

@device_assignment_bp.route('/export/available-sims')
@login_required
def export_available_sims():
    """تصدير الأرقام المتاحة إلى Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import make_response
        import io
        
        # جلب الأرقام المتاحة
        available_sims = SimCard.query.filter_by(employee_id=None).all()
        
        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "الأرقام المتاحة"
        
        # إعداد العناوين
        headers = ['رقم الهاتف', 'مزود الخدمة', 'نوع الباقة', 'الحالة', 'تاريخ الإضافة', 'المرتبط بموظف']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, name='Arial')
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', name='Arial')
        
        # إضافة البيانات
        for row, sim in enumerate(available_sims, 2):
            ws.cell(row=row, column=1, value=sim.phone_number)
            ws.cell(row=row, column=2, value=sim.carrier or 'غير محدد')
            ws.cell(row=row, column=3, value=sim.plan_type or 'غير محدد')
            ws.cell(row=row, column=4, value=sim.status or 'متاحة')
            ws.cell(row=row, column=5, value=sim.created_at.strftime('%Y-%m-%d') if sim.created_at else '')
            ws.cell(row=row, column=6, value=sim.employee.name if sim.employee else 'غير مرتبطة')
        
        # حفظ الملف
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=available_sims.xlsx'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Error exporting available sims: {str(e)}")
        flash('حدث خطأ أثناء تصدير الأرقام المتاحة', 'danger')
        return redirect(url_for('device_assignment.reports'))

@device_assignment_bp.route('/export/available-devices')
@login_required
def export_available_devices():
    """تصدير الأجهزة المتاحة إلى Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import make_response
        import io
        
        # جلب الأجهزة المتاحة (غير مربوطة)
        assigned_device_ids = db.session.query(DeviceAssignment.device_id).filter_by(is_active=True).subquery()
        available_devices = MobileDevice.query.filter(~MobileDevice.id.in_(assigned_device_ids)).all()
        
        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "الأجهزة المتاحة"
        
        # إعداد العناوين
        headers = ['IMEI', 'ماركة الجهاز', 'موديل الجهاز', 'حالة الجهاز', 'تاريخ الإضافة', 'ملاحظات']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, name='Arial')
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', name='Arial')
        
        # إضافة البيانات
        for row, device in enumerate(available_devices, 2):
            ws.cell(row=row, column=1, value=device.imei)
            ws.cell(row=row, column=2, value=device.device_brand or 'غير محدد')
            ws.cell(row=row, column=3, value=device.device_model or 'غير محدد')
            ws.cell(row=row, column=4, value=device.status or 'متاح')
            ws.cell(row=row, column=5, value=device.created_at.strftime('%Y-%m-%d') if device.created_at else '')
            ws.cell(row=row, column=6, value=device.notes or '')
        
        # حفظ الملف
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=available_devices.xlsx'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Error exporting available devices: {str(e)}")
        flash('حدث خطأ أثناء تصدير الأجهزة المتاحة', 'danger')
        return redirect(url_for('device_assignment.reports'))

@device_assignment_bp.route('/export/available-combined')
@login_required
def export_available_combined():
    """تصدير الأجهزة والأرقام المتاحة معاً إلى Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import make_response
        import io
        
        # جلب البيانات المتاحة
        assigned_device_ids = db.session.query(DeviceAssignment.device_id).filter_by(is_active=True).subquery()
        available_devices = MobileDevice.query.filter(~MobileDevice.id.in_(assigned_device_ids)).all()
        available_sims = SimCard.query.filter_by(employee_id=None).all()
        
        # إنشاء ملف Excel
        wb = Workbook()
        
        # ورقة الأجهزة المتاحة
        ws1 = wb.active
        ws1.title = "الأجهزة المتاحة"
        
        # عناوين الأجهزة
        device_headers = ['IMEI', 'ماركة الجهاز', 'موديل الجهاز', 'حالة الجهاز', 'تاريخ الإضافة', 'ملاحظات']
        for col, header in enumerate(device_headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, name='Arial')
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', name='Arial')
        
        # بيانات الأجهزة
        for row, device in enumerate(available_devices, 2):
            ws1.cell(row=row, column=1, value=device.imei)
            ws1.cell(row=row, column=2, value=device.device_brand or 'غير محدد')
            ws1.cell(row=row, column=3, value=device.device_model or 'غير محدد')
            ws1.cell(row=row, column=4, value=device.status or 'متاح')
            ws1.cell(row=row, column=5, value=device.created_at.strftime('%Y-%m-%d') if device.created_at else '')
            ws1.cell(row=row, column=6, value=device.notes or '')
        
        # ورقة الأرقام المتاحة
        ws2 = wb.create_sheet(title="الأرقام المتاحة")
        
        # عناوين الأرقام
        sim_headers = ['رقم الهاتف', 'مزود الخدمة', 'نوع الباقة', 'الحالة', 'تاريخ الإضافة', 'المرتبط بموظف']
        for col, header in enumerate(sim_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, name='Arial')
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', name='Arial')
        
        # بيانات الأرقام
        for row, sim in enumerate(available_sims, 2):
            ws2.cell(row=row, column=1, value=sim.phone_number)
            ws2.cell(row=row, column=2, value=sim.carrier or 'غير محدد')
            ws2.cell(row=row, column=3, value=sim.plan_type or 'غير محدد')
            ws2.cell(row=row, column=4, value=sim.status or 'متاحة')
            ws2.cell(row=row, column=5, value=sim.created_at.strftime('%Y-%m-%d') if sim.created_at else '')
            ws2.cell(row=row, column=6, value=sim.employee.name if sim.employee else 'غير مرتبطة')
        
        # ورقة الإحصائيات
        ws3 = wb.create_sheet(title="ملخص الإحصائيات")
        
        # إحصائيات عامة
        stats_data = [
            ['البيان', 'العدد'],
            ['إجمالي الأجهزة المتاحة', len(available_devices)],
            ['إجمالي الأرقام المتاحة', len(available_sims)],
            ['تاريخ التقرير', datetime.now().strftime('%Y-%m-%d %H:%M')]
        ]
        
        for row, data in enumerate(stats_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws3.cell(row=row, column=col, value=value)
                if row == 1:  # رأس الجدول
                    cell.font = Font(bold=True, name='Arial')
                    cell.fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF', name='Arial')
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')
        
        # حفظ الملف
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=available_devices_and_sims.xlsx'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Error exporting combined data: {str(e)}")
        flash('حدث خطأ أثناء تصدير البيانات المجمعة', 'danger')
        return redirect(url_for('device_assignment.reports'))