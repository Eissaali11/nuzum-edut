"""
Payroll Admin Routes
مسارات إدارة الرواتب
"""
from flask import Blueprint, render_template, request, jsonify, send_file, redirect, url_for, flash
from flask_login import login_required, current_user
from datetime import datetime, date
from decimal import Decimal
from sqlalchemy import func, and_
from src.core.extensions import db
from models import Employee, User, Module, Permission, Department
from src.modules.payroll.domain.models import PayrollRecord, PayrollConfiguration, BankTransferFile
from src.modules.payroll.application.payroll_processor import PayrollProcessor
from src.modules.payroll.application.bank_transfer_generator import BankTransferGenerator


payroll_bp = Blueprint('payroll', __name__)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def _ensure_default_departments():
    """التأكد من وجود الأقسام الافتراضية"""
    if Department.query.count() > 0:
        return  # الأقسام موجودة بالفعل
    
    default_departments = [
        {'name': 'الموارد البشرية', 'code': 'HR'},
        {'name': 'المبيعات', 'code': 'SALES'},
        {'name': 'التسويق', 'code': 'MARKETING'},
        {'name': 'تكنولوجيا المعلومات', 'code': 'IT'},
        {'name': 'العمليات', 'code': 'OPS'},
        {'name': 'التمويل', 'code': 'FINANCE'},
        {'name': 'الإدارة', 'code': 'ADMIN'},
        {'name': 'خدمة العملاء', 'code': 'CS'},
    ]
    
    try:
        for dept_data in default_departments:
            dept = Department(
                name=dept_data['name'],
                code=dept_data['code'],
                status='active'
            )
            db.session.add(dept)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        # لا نرفع الخطأ، نترك التطبيق يعمل


# ═══════════════════════════════════════════════════════════════════════════════
# MIDDLEWARE & PERMISSIONS
# ═══════════════════════════════════════════════════════════════════════════════

def check_payroll_access():
    """التحقق من صلاحيات الوصول لقسم الرواتب"""
    if not current_user.is_authenticated:
        return False

    role_value = getattr(current_user.role, 'value', current_user.role)
    if getattr(current_user, 'is_admin', False) or str(role_value).lower() == 'admin':
        return True
    
    # التحقق من الصلاحية
    payroll_module = Module.query.filter_by(name='payroll').first()
    if not payroll_module:
        return False
    
    permission = Permission.query.filter(
        and_(
            Permission.user_id == current_user.id,
            Permission.module_id == payroll_module.id
        )
    ).first()
    
    return permission is not None


# ═══════════════════════════════════════════════════════════════════════════════
# PAYROLL MANAGEMENT ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@payroll_bp.route('/dashboard')
@login_required
def dashboard():
    """لوحة معلومات الرواتب"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية للوصول لقسم الرواتب', 'danger')
        return redirect(url_for('dashboard.index'))
    
    # التأكد من وجود الأقسام الافتراضية
    _ensure_default_departments()
    
    # إحصائيات عامة
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    total_employees = Employee.query.filter_by(status='active').count()
    
    # آخر معالجة رواتب
    latest_payroll = PayrollRecord.query.filter_by(
        pay_period_month=current_month,
        pay_period_year=current_year
    ).order_by(PayrollRecord.calculated_at.desc()).first()
    
    # ✅ FIXED: Use joinedload to prevent N+1 queries on employee relationships
    month_payrolls = PayrollRecord.query.options(
        db.joinedload(PayrollRecord.employee).joinedload(Employee.departments)
    ).filter_by(
        pay_period_month=current_month,
        pay_period_year=current_year
    ).order_by(PayrollRecord.calculated_at.desc()).all()

    # الإحصائيات حسب الحالة
    pending_count = sum(1 for payroll in month_payrolls if payroll.payment_status == 'pending')
    approved_count = sum(1 for payroll in month_payrolls if payroll.payment_status == 'approved')
    rejected_count = sum(1 for payroll in month_payrolls if payroll.payment_status == 'rejected')
    paid_count = sum(1 for payroll in month_payrolls if payroll.payment_status == 'paid')

    # مجاميع المبالغ (None-safe)
    total_gross_salary = sum((payroll.gross_salary or Decimal('0')) for payroll in month_payrolls)
    total_deductions = sum((payroll.total_deductions or Decimal('0')) for payroll in month_payrolls)
    total_net_payable = sum((payroll.net_payable or Decimal('0')) for payroll in month_payrolls)
    total_gosi_company = sum((payroll.gosi_company or Decimal('0')) for payroll in month_payrolls)

    return render_template('payroll/dashboard.html',
        total_employees=total_employees,
        latest_payroll=latest_payroll,
        pending_count=pending_count,
        approved_count=approved_count,
        rejected_count=rejected_count,
        paid_count=paid_count,
        total_gross_salary=float(total_gross_salary),
        total_deductions=float(total_deductions),
        total_net_payable=float(total_net_payable),
        total_gosi_company=float(total_gosi_company),
        recent_payrolls=month_payrolls,
        current_month=current_month,
        current_year=current_year,
        pending_payrolls=pending_count,
        approved_payrolls=approved_count,
        paid_payrolls=paid_count,
        total_payroll=float(total_net_payable)
    )


@payroll_bp.route('/process', methods=['GET', 'POST'])
@login_required
def process():
    """معالجة رواتب الموظفين"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    if request.method == 'POST':
        try:
            month_value = (request.form.get('month') or '').strip()
            year_value = request.form.get('year') or request.form.get('fiscal_year')

            if month_value and '-' in month_value:
                parsed_year, parsed_month = month_value.split('-', 1)
                year = int(parsed_year)
                month = int(parsed_month)
            else:
                month = int(month_value) if month_value else datetime.now().month
                year = int(year_value) if year_value else datetime.now().year
            
            processor = PayrollProcessor(year, month)
            payrolls = processor.process_all_employees()

            total_gross = sum((p.gross_salary or Decimal('0')) for p in payrolls)
            total_deductions = sum((p.total_deductions or Decimal('0')) for p in payrolls)
            total_net_payable = sum((p.net_payable or Decimal('0')) for p in payrolls)

            wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            if wants_json:
                return jsonify({
                    'success': True,
                    'processed_count': len(payrolls),
                    'total_gross_salary': float(total_gross),
                    'total_deductions': float(total_deductions),
                    'total_net_payable': float(total_net_payable)
                })
            
            flash(f'تم معالجة رواتب {len(payrolls)} موظف بنجاح', 'success')
            return redirect(url_for('payroll.review', month=month, year=year))
        
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': str(e)}), 400
            flash(f'خطأ: {str(e)}', 'danger')
    
    return render_template('payroll/process.html',
        current_month=datetime.now().month,
        current_year=datetime.now().year
    )


@payroll_bp.route('/review')
@login_required
def review():
    """مراجعة رواتب الموظفين قبل الموافقة"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    # التأكد من وجود الأقسام الافتراضية
    _ensure_default_departments()
    
    month = request.args.get('month', datetime.now().month, type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    status = request.args.get('status', 'pending')
    department_id = request.args.get('department_id', type=int)
    
    # جلب الرواتب
    query = PayrollRecord.query.join(Employee, PayrollRecord.employee_id == Employee.id).filter(
        and_(
            PayrollRecord.pay_period_month == month,
            PayrollRecord.pay_period_year == year,
            PayrollRecord.payment_status == status
        )
    )

    if department_id:
        query = query.filter(Employee.department_id == department_id)
    
    payrolls = query.all()
    
    # الإحصائيات
    total_gross = sum(p.gross_salary for p in payrolls)
    total_deductions = sum(p.total_deductions for p in payrolls)
    total_net = sum(p.net_payable for p in payrolls)
    departments = Department.query.order_by(Department.name.asc()).all()
    
    return render_template('payroll/review.html',
        payrolls=payrolls,
        month=month,
        year=year,
        status=status,
        departments=departments,
        selected_department_id=department_id,
        total_gross=float(total_gross),
        total_deductions=float(total_deductions),
        total_net=float(total_net),
        payroll_count=len(payrolls)
    )


@payroll_bp.route('/<int:payroll_id>/details')
@payroll_bp.route('/payroll/<int:payroll_id>/details')
@login_required
def payroll_details(payroll_id):
    """عرض تفاصيل الراتب الكامل"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    payroll = PayrollRecord.query.get_or_404(payroll_id)
    
    return render_template('payroll/details.html',
        payroll=payroll
    )


@payroll_bp.route('/payroll/<int:payroll_id>/details/legacy')
@login_required
def payroll_details_legacy(payroll_id):
    return redirect(url_for('payroll.payroll_details', payroll_id=payroll_id))


@payroll_bp.route('/<int:payroll_id>/approve', methods=['POST'])
@payroll_bp.route('/payroll/<int:payroll_id>/approve', methods=['POST'])
@login_required
def approve_payroll(payroll_id):
    """الموافقة على الراتب"""
    if not check_payroll_access():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'}), 403
    
    try:
        payroll = PayrollRecord.query.get_or_404(payroll_id)
        processor = PayrollProcessor(payroll.pay_period_year, payroll.pay_period_month)
        
        notes = request.form.get('notes', '')
        processor.approve_payroll(payroll_id, current_user.id, notes)
        
        return jsonify({'success': True, 'message': 'تمت الموافقة على الراتب'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@payroll_bp.route('/<int:payroll_id>/reject', methods=['POST'])
@payroll_bp.route('/payroll/<int:payroll_id>/reject', methods=['POST'])
@login_required
def reject_payroll(payroll_id):
    """رفض الراتب"""
    if not check_payroll_access():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'}), 403
    
    try:
        reason = request.form.get('reason', 'لم يتم تحديد السبب')
        processor = PayrollProcessor(None, None)
        processor.reject_payroll(payroll_id, reason)
        
        return jsonify({'success': True, 'message': 'تم رفض الراتب'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@payroll_bp.route('/payroll/<int:payroll_id>/approve/legacy', methods=['POST'])
@login_required
def approve_payroll_legacy(payroll_id):
    return redirect(url_for('payroll.approve_payroll', payroll_id=payroll_id), code=307)


@payroll_bp.route('/payroll/<int:payroll_id>/reject/legacy', methods=['POST'])
@login_required
def reject_payroll_legacy(payroll_id):
    return redirect(url_for('payroll.reject_payroll', payroll_id=payroll_id), code=307)


@payroll_bp.route('/batch-approve', methods=['POST'])
@login_required
def batch_approve():
    """الموافقة على مجموعة من الرواتب"""
    if not check_payroll_access():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'}), 403
    
    try:
        payroll_ids = request.json.get('payroll_ids', [])
        processor = PayrollProcessor(None, None)
        
        for payroll_id in payroll_ids:
            processor.approve_payroll(payroll_id, current_user.id)
        
        return jsonify({
            'success': True,
            'message': f'تمت الموافقة على {len(payroll_ids)} راتب'
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


# ═══════════════════════════════════════════════════════════════════════════════
# BANK TRANSFER & EXPORT ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@payroll_bp.route('/bank-transfer/create', methods=['POST'])
@login_required
def create_bank_transfer():
    """إنشاء ملف التحويل البنكي"""
    if not check_payroll_access():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'}), 403
    
    try:
        month = int(request.form.get('month'))
        year = int(request.form.get('year'))
        bank_code = request.form.get('bank_code', 'NCB')
        file_format = request.form.get('format', 'txt')
        
        generator = BankTransferGenerator(year, month, bank_code)
        transfer_file = generator.create_transfer_file(file_format)
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء ملف التحويل بنجاح',
            'file_id': transfer_file.id,
            'file_name': transfer_file.file_name
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@payroll_bp.route('/bank-transfer/<int:file_id>/download')
@login_required
def download_bank_transfer(file_id):
    """تنزيل ملف التحويل البنكي"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    transfer_file = BankTransferFile.query.get_or_404(file_id)
    
    try:
        return send_file(
            transfer_file.file_path,
            as_attachment=True,
            download_name=transfer_file.file_name
        )
    except Exception as e:
        flash(f'خطأ في التحميل: {str(e)}', 'danger')
        return redirect(url_for('payroll.dashboard'))


@payroll_bp.route('/export/excel')
@login_required
def export_to_excel():
    """تصدير الرواتب إلى Excel"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        month = request.args.get('month', datetime.now().month, type=int)
        year = request.args.get('year', datetime.now().year, type=int)
        
        # جلب الرواتب
        payrolls = PayrollRecord.query.filter_by(
            pay_period_month=month,
            pay_period_year=year,
            payment_status='approved'
        ).all()
        
        # إنشاء Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Payroll_Details'
        
        # الرؤوس
        headers = ['رقم الموظف', 'اسم الموظف', 'الراتب الأساسي', 'المزايا', 'الراتب الإجمالي', 
                   'الخصومات', 'الراتب الصافي', 'الحالة']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # البيانات
        row = 2
        for payroll in payrolls:
            employee = payroll.employee
            
            ws.cell(row=row, column=1).value = employee.employee_id
            ws.cell(row=row, column=2).value = employee.name
            ws.cell(row=row, column=3).value = float(payroll.basic_salary)
            ws.cell(row=row, column=4).value = float(payroll.housing_allowance + payroll.transportation + payroll.meal_allowance)
            ws.cell(row=row, column=5).value = float(payroll.gross_salary)
            ws.cell(row=row, column=6).value = float(payroll.total_deductions)
            ws.cell(row=row, column=7).value = float(payroll.net_payable)
            ws.cell(row=row, column=8).value = payroll.payment_status
            
            row += 1
        
        # حفظ الملف
        file_path = f'static/payroll_exports/payroll_{month:02d}_{year}.xlsx'
        wb.save(file_path)
        
        return send_file(file_path, as_attachment=True)
    
    except Exception as e:
        flash(f'خطأ: {str(e)}', 'danger')
        return redirect(url_for('payroll.dashboard'))


# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@payroll_bp.route('/configuration')
@login_required
def configuration():
    """إعدادات الرواتب"""
    if not check_payroll_access():
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))
    
    config = PayrollConfiguration.query.filter_by(is_active=True).first()
    if not config:
        config = PayrollConfiguration()
    
    return render_template('payroll/configuration.html', config=config)


@payroll_bp.route('/configuration/save', methods=['POST'])
@login_required
def save_configuration():
    """حفظ إعدادات الرواتب"""
    if not check_payroll_access():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية'}), 403
    
    try:
        config = PayrollConfiguration.query.filter_by(is_active=True).first()
        if not config:
            config = PayrollConfiguration()
        
        config.gosi_employee_percentage = Decimal(request.form.get('gosi_employee', 10))
        config.gosi_company_percentage = Decimal(request.form.get('gosi_company', 13))
        config.working_days_per_month = int(request.form.get('working_days', 22))
        config.minimum_wage = Decimal(request.form.get('minimum_wage', 2000))
        
        db.session.add(config)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'تم حفظ الإعدادات بنجاح'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400
