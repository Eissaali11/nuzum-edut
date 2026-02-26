from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, current_app, send_file
from werkzeug.utils import secure_filename
import os
import uuid
from datetime import datetime
from PIL import Image
from pillow_heif import register_heif_opener

# تسجيل plugin الـ HEIC/HEIF للتعامل مع صور الآيفون
register_heif_opener()
from models import VehicleExternalSafetyCheck, VehicleSafetyImage, Vehicle, Employee, User, UserRole, VehicleHandover
from src.core.extensions import db
from src.utils.audit_logger import log_audit
from src.utils.storage_helper import upload_image, delete_image
from src.utils.vehicle_drive_uploader import VehicleDriveUploader
from flask_login import current_user, login_required
from sqlalchemy import func, select
from sqlalchemy.orm import aliased, contains_eager

from dotenv import load_dotenv
import resend

# تم نقل whatsapp_service إلى app.py لاستخدامه بشكل مركزي

# دوال الإشعارات المحلية
def create_safety_check_notification(user_id, vehicle_plate, supervisor_name, check_status, check_id):
    """إشعار فحص السلامة الخارجية"""
    from models import Notification
    
    status_labels = {
        'pending': 'قيد الانتظار',
        'under_review': 'قيد المراجعة',
        'approved': 'موافق عليه',
        'rejected': 'مرفوض'
    }
    
    priority_map = {
        'pending': 'high',
        'under_review': 'normal',
        'approved': 'normal',
        'rejected': 'critical'
    }
    
    status_label = status_labels.get(check_status, check_status)
    
    notification = Notification(
        user_id=user_id,
        notification_type='safety_check',
        title=f'فحص السلامة - السيارة {vehicle_plate}',
        description=f'طلب فحص السلامة الخارجية للسيارة {vehicle_plate} من قبل {supervisor_name} - الحالة: {status_label}',
        related_entity_type='safety_check',
        related_entity_id=check_id,
        priority=priority_map.get(check_status, 'normal'),
        action_url=url_for('external_safety.admin_external_safety_checks')
    )
    db.session.add(notification)
    db.session.commit()
    return notification

def create_safety_check_review_notification(user_id, vehicle_plate, action, reviewer_name, check_id):
    """إشعار بمراجعة/موافقة فحص السلامة"""
    from models import Notification
    
    action_labels = {
        'approved': 'تمت الموافقة على',
        'rejected': 'تم رفض',
        'under_review': 'قيد المراجعة'
    }
    
    priority = 'high' if action in ['rejected'] else 'normal'
    action_label = action_labels.get(action, action)
    
    notification = Notification(
        user_id=user_id,
        notification_type='safety_check_review',
        title=f'{action_label} فحص السلامة - {vehicle_plate}',
        description=f'تمت مراجعة فحص السلامة للسيارة {vehicle_plate} بواسطة {reviewer_name}: {action_label}',
        related_entity_type='safety_check',
        related_entity_id=check_id,
        priority=priority,
        action_url=url_for('external_safety.admin_external_safety_checks')
    )
    db.session.add(notification)
    db.session.commit()
    return notification


# قم بتحميل المتغيرات من ملف .env
load_dotenv()
# قم بإعداد مفتاح Resend مرة واحدة عند بدء التطبيق
resend.api_key = os.environ.get("RESEND_API_KEY")
supervisor_email = os.environ.get("SAFETY_CHECK_SUPERVISOR_EMAIL")
company_name = os.environ.get("COMPANY_NAME")
external_safety_bp = Blueprint('external_safety', __name__)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'heic', 'heif'}


@external_safety_bp.route('/test-notifications', methods=['GET', 'POST'])
def test_safety_check_notifications():
    """اختبار إنشاء إشعارات فحص السلامة لجميع المستخدمين"""
    try:
        # الحصول على آخر فحص سلامة
        last_check = VehicleExternalSafetyCheck.query.order_by(VehicleExternalSafetyCheck.id.desc()).first()
        
        if not last_check:
            return jsonify({'success': False, 'message': 'لا توجد فحوصات سلامة'}), 404
        
        all_users = User.query.all()
        
        notification_count = 0
        for user in all_users:
            try:
                create_safety_check_notification(
                    user_id=user.id,
                    vehicle_plate=last_check.vehicle_plate_number or 'غير محدد',
                    supervisor_name=last_check.driver_name or 'غير محدد',
                    check_status=last_check.approval_status or 'pending',
                    check_id=last_check.id
                )
                notification_count += 1
            except Exception as e:
                current_app.logger.error(f'خطأ في إنشاء إشعار للمستخدم {user.id}: {str(e)}')
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء {notification_count} إشعار لفحص السلامة {last_check.id}',
            'check_id': last_check.id,
            'users_count': len(all_users)
        })
    except Exception as e:
        current_app.logger.error(f'خطأ في اختبار الإشعارات: {str(e)}')
        return jsonify({'success': False, 'message': str(e)}), 500

def get_all_current_driversWithEmil():
    """
    تسترجع قاموساً يحتوي على معلومات السائق الحالي لكل مركبة.
    المفتاح هو ID المركبة، والقيمة هي قاموس يحتوي على (name, email, mobile).
    """
    # 1. نحدد أنواع التسليم
    delivery_handover_types = ['delivery', 'تسليم', 'handover']
    
    # 2. إنشاء استعلام فرعي لتحديد أحدث سجل تسليم لكل مركبة
    # (نفس منطق Window Function السابق)
    subq = select(
        VehicleHandover.id,
        func.row_number().over(
            partition_by=VehicleHandover.vehicle_id,
            order_by=VehicleHandover.handover_date.desc()
        ).label('row_num')
    ).where(
        VehicleHandover.handover_type.in_(delivery_handover_types)
    ).subquery()

    # 3. إنشاء الاستعلام الرئيسي
    # سنربط (JOIN) بين السجلات الأحدث والموظفين المرتبطين بها
    # ونستخدم `contains_eager` لجلب بيانات الموظف بكفاءة عالية في نفس الاستعلام
    stmt = select(VehicleHandover).join(
        subq, VehicleHandover.id == subq.c.id
    ).join(
        Employee, VehicleHandover.employee_id == Employee.id  # الربط باستخدام جدول Employee
    ).where(subq.c.row_num == 1)

    # 4. تنفيذ الاستعلام وجلب النتائج
    latest_handovers_with_drivers = db.session.execute(stmt).scalars().all()
    
    # 5. تحويل النتائج إلى القاموس (dictionary) بالصيغة الجديدة
    current_drivers_map = {
        record.vehicle_id: {
            'name': record.driver_employee.name,
            'email': record.driver_employee.email,
            'mobile': record.driver_employee.mobile,
            'phone' : record.driver_employee.mobile,
            'national_id': record.driver_employee.national_id
        }
        for record in latest_handovers_with_drivers if record.driver_employee # نتأكد من وجود سائق
    }
    
    return current_drivers_map



# في نفس ملف الراوت external_safety_bp




# قم بإنشاء نسخة واحدة من الكلاس على مستوى الـ Blueprint أو التطبيق
# من الأفضل وضعها في __init__.py الخاص بالتطبيق و استيرادها
# استيراد whatsapp_service من app بدلاً من إنشاء نسخة جديدة
try:
    from src.app import whatsapp_service
except ImportError:
    whatsapp_service = None  # التطبيق قد لا يوفر خدمة WhatsApp

# ----- أضف هذه الدالة الجديدة بجانب دالة send_vehicle_email -----

@external_safety_bp.route('/api/send-whatsapp', methods=['POST'])
def send_vehicle_whatsapp():
    """
    نقطة نهاية (API endpoint) لإرسال طلب فحص المركبة عبر واتساب.
    """
    # 1. استلام البيانات من الطلب (نفس بيانات البريد الإلكتروني)
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'الطلب فارغ'}), 400

    # استخراج البيانات. لاحظ أننا نحتاج رقم هاتف السائق بدلاً من البريد
    driver_phone = data.get('driver_phone') # <-- أهم معلومة جديدة
    driver_name = data.get('driver_name', 'زميلنا العزيز')
    plate_number = data.get('plate_number')
    vehicle_model = data.get('vehicle_model')
    form_url = data.get('form_url')

    # التحقق من وجود البيانات الضرورية
    if not all([driver_phone, plate_number, vehicle_model, form_url]):
        error_message = "بيانات ناقصة. تأكد من إرسال: driver_phone, driver_name, plate_number, vehicle_model, form_url."
        return jsonify({'success': False, 'error': error_message}), 400

    # 2. تجهيز مكونات قالب واتساب
    template_name = "vehicle_safety_check_request" # <-- اسم القالب الذي وافقت عليه Meta

    # ترتيب المتغيرات مهم جداً ويجب أن يطابق ترتيبها في القالب
    components = [
        {
            "type": "body",
            "parameters": [
                {"type": "text", "text": driver_name},     # يحل محل {{1}}
                {"type": "text", "text": plate_number},    # يحل محل {{2}}
                {"type": "text", "text": vehicle_model},   # يحل محل {{3}}
                {"type": "text", "text": form_url},        # يحل محل {{4}} في الجسم
            ]
         }
        #  ,
        # # إذا كان زر الرابط ديناميكياً، نضيف له مكوناً أيضاً
        # {
        #     "type": "button",
        #     "sub_type": "url",
        #     "index": "0",  # رقم الزر (يبدأ من 0)
        #     "parameters": [
        #         {"type": "text", "text": form_url.split('/')[-1]} # الرابط يجب أن يكون الجزء الأخير بعد /
        #                                                         # مثال: "external_safety_check/vehicle_id"
        #     ]
        # }
    ]

    # ملاحظة على رابط الزر: واتساب يتطلب أن يكون الجزء المتغير من الرابط
    # فقط. الرابط الأساسي (e.g. https://nuzum.sa) تضعه عند تصميم القالب.

    # 3. استدعاء خدمة واتساب للإرسال
    if not whatsapp_service:
        return jsonify({'success': False, 'error': 'خدمة واتساب غير متاحة. يرجى إعداد WHATSAPP_ACCESS_TOKEN و WHATSAPP_PHONE_NUMBER_ID في ملف .env'}), 503
    
    try:
        response = whatsapp_service.send_template_message(
            recipient_number=driver_phone, # رقم السائق مع رمز الدولة
            template_name=template_name,
            language_code="ar",
            components=components
        )
        
        if response:
            return jsonify({'success': True, 'message': f"تم إرسال رسالة واتساب بنجاح إلى {driver_name}"})
        else:
            # إذا فشلت دالتنا في الإرسال (مثلاً خطأ في الاتصال)
            return jsonify({'success': False, 'error': 'فشل إرسال رسالة واتساب من الخادم'}), 500

    except Exception as e:
        # لأي خطأ آخر غير متوقع
        print(f"Error sending WhatsApp message: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

def get_all_current_drivers():
    """
    تسترجع قاموساً يحتوي على السائق الحالي لكل مركبة. (بصيغة حديثة)
    المفتاح هو ID المركبة، والقيمة هي اسم السائق.
    """
    
    # 1. نحدد أنواع التسليم
    delivery_handover_types = ['delivery', 'تسليم', 'handover']
    
    # 2. إنشاء استعلام فرعي (Subquery) باستخدام Window Function
    subq = select(
        VehicleHandover.id,
        func.row_number().over(
            partition_by=VehicleHandover.vehicle_id,
            order_by=VehicleHandover.handover_date.desc()
        ).label('row_num')
    ).where(
        VehicleHandover.handover_type.in_(delivery_handover_types)
    ).subquery()

    # 3. الآن، نحصل على أحدث السجلات عن طريق اختيار التي لها row_num = 1
    stmt = select(VehicleHandover).join(
        subq, VehicleHandover.id == subq.c.id
    ).where(subq.c.row_num == 1)

    # 4. تنفيذ الاستعلام وجلب النتائج
    # .scalars() تجلب الكائنات (objects) مباشرة بدلاً من الصفوف (rows)
    latest_handovers = db.session.execute(stmt).scalars().all()
    
    # 5. تحويل النتائج إلى قاموس (dictionary) سهل الاستخدام
    current_drivers_map = {
        record.vehicle_id: record.person_name for record in latest_handovers
    }
    
    return current_drivers_map


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def compress_image(image_path, max_size=1200, quality=85):
    """ضغط الصورة لتقليل حجمها مع دعم HEIC من الآيفون"""
    try:
        with Image.open(image_path) as img:
            # تحويل RGBA أو أي تنسيق آخر إلى RGB
            if img.mode in ('RGBA', 'LA', 'P'):
                img = img.convert('RGB')
            
            # تحويل HEIC إلى RGB وضغطها
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # تغيير حجم الصورة إذا كانت أكبر من الحد المسموح
            if img.width > max_size or img.height > max_size:
                img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
            
            # حفظ الصورة المضغوطة
            img.save(image_path, 'JPEG', quality=quality, optimize=True)
            return True
    except Exception as e:
        current_app.logger.error(f"خطأ في ضغط الصورة: {str(e)}")
        return False


def send_supervisor_notification_email(safety_check):
    """
    تقوم ببناء وإرسال بريد إلكتروني لإشعار المشرف بوجود طلب فحص جديد.
    """
    # --- !! هام: قم بتعديل هذه المتغيرات !! ---
    # الخيار الأفضل: اقرأ هذا من إعدادات التطبيق (config file)
    # SUPERVISOR_EMAIL = current_app.config.get('SAFETY_CHECK_SUPERVISOR_EMAIL')
    # supervisor_email = "ferasswed2022@gmail.com"  # <--- ضع بريد المشرف هنا
    # company_name = "نُــظــم لإدارة الأساطيل"
    # ----------------------------------------------
    
    # توليد الرابط الخاص بمراجعة الطلب في لوحة التحكم
    # تأكد من أن اسم الـ blueprint والنقطة النهائية صحيحين. قد يكون 'admin.view_check' أو ما شابه
    logo_url = "https://i.postimg.cc/LXzD6b0N/logo.png" # <--- رابط الشعار العام
    try:
        review_url = url_for('external_safety_bp.admin_view_safety_check', # <--- تأكد من هذا المسار
                             check_id=safety_check.id, 
                             _external=True)
    except Exception as e:
        # حل احتياطي في حال حدوث خطأ، لكن يجب إصلاح المسار أعلاه
        review_url = f"http://127.0.0.1:4032//admin/external-safety-check/{safety_check.id}"
        current_app.logger.error(f"Failed to generate review URL, using fallback. Error: {e}")

    # بناء قالب HTML احترافي للإشعار

    email_html_content = f"""
    <!DOCTYPE html>
    <html lang="ar" dir="rtl">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;700&display=swap');
            body {{
                margin: 0;
                padding: 0;
                background-color: #e9ecef; /* خلفية رمادية فاتحة جدا */
                font-family: 'Tajawal', sans-serif;
            }}
            .email-wrapper {{
                max-width: 680px;
                margin: 40px auto;
                background-color: #ffffff;
                border-radius: 12px;
                box-shadow: 0 8px 30px rgba(0,0,0,0.07);
            }}
            .email-header {{
                background: linear-gradient(135deg, #182551 10%, #425359 100%);
                border-radius: 12px 12px 0 0;
                padding: 25px;
                text-align: center;
            }}
            .email-header img {{
                max-height: 50px; /* حجم مناسب للشعار */
                margin-bottom: 15px;
            }}
            .email-header h1 {{
                margin: 0;
                color: #ffffff;
                font-size: 26px;
                font-weight: 700;
            }}
            .email-body {{
                padding: 25px 35px;
                text-align: right;
            }}
            .greeting {{
                font-size: 20px;
                color: #2c3e50;
                font-weight: 700;
                margin-bottom: 10px;
            }}
            .main-message {{
                font-size: 16px;
                color: #555;
                line-height: 1.7;
            }}
            .details-card {{
                background-color: #f8f9fa;
                border: 1px dashed #ced4da;
                border-radius: 8px;
                padding: 20px;
                margin: 25px 0;
            }}
            .details-card h3 {{
                margin-top: 0;
                color: #343a40;
                border-bottom: 2px solid #dee2e6;
                padding-bottom: 10px;
                margin-bottom: 15px;
            }}
            .details-card p {{
                margin: 8px 0;
                font-size: 16px;
            }}
            .details-card p strong {{
                color: #495057;
                display: inline-block;
                width: 110px;
            }}
            .action-button-container {{
                text-align: center;
                margin: 30px 0;
            }}
            .action-button {{
                background: linear-gradient(135deg, #0d6efd 0%, #0a58ca 100%);
                color: #ffffff !important;
                padding: 14px 40px;
                text-decoration: none;
                border-radius: 50px;
                font-weight: 700;
                font-size: 18px;
                box-shadow: 0 5px 15px rgba(52,152,219,0.3);
                transition: all 0.3s ease;
            }}
            .action-button:hover {{
                transform: translateY(-2px);
                box-shadow: 0 8px 20px rgba(52,152,219,0.4);
            }}
            .email-footer {{
                padding: 20px;
                text-align: center;
                font-size: 13px;
                color: #888;
                background-color: #f8f9fa;
                border-top: 1px solid #dee2e6;
            }}
        </style>
    </head>
    <body>
        <div class="email-wrapper">
            <div class="email-header">
                <img src="{logo_url}" alt="{company_name} Logo">
                <h1>إشعار بفحص جديد</h1>
            </div>
            <div class="email-body">
                <p class="greeting">مرحباً أيها المشرف،</p>
                <p class="main-message">تم استلام طلب فحص سلامة جديد وهو في انتظار مراجعتكم لاتخاذ الإجراء المناسب. يرجى الاطلاع على التفاصيل أدناه.</p>
                
                <div class="details-card">
                    <h3>تفاصيل الطلب</h3>
                    <p><strong>رقم الطلب:</strong> #{safety_check.id}</p>
                    <p><strong>المركبة:</strong> {safety_check.vehicle_plate_number} ({safety_check.vehicle_make_model})</p>
                    <p><strong>السائق:</strong> {safety_check.driver_name}</p>
                    <p><strong>التاريخ:</strong> {safety_check.inspection_date.strftime('%d-%m-%Y %I:%M %p')}</p>
                </div>

                <p class="main-message">اضغط على الزر أدناه للانتقال مباشرة إلى صفحة المراجعة:</p>

                <div class="action-button-container">
                    <a href="{review_url}" class="action-button">مراجعة طلب الفحص</a>
                </div>
            </div>
            <div class="email-footer">
                <p>© {datetime.now().year} {company_name}. جميع الحقوق محفوظة.</p>
            </div>
        </div>
    </body>
    </html>
    """

    

    # إرسال البريد الإلكتروني عبر Resend
    try:
        params = {
            "from": f"{company_name} <notifications@resend.dev>",
            "to": [supervisor_email],
            "subject": f"طلب فحص جديد للمركبة {safety_check.vehicle_plate_number} بحاجة لمراجعة",
            "html": email_html_content,
        }
        resend.Emails.send(params)
        current_app.logger.info(f"تم إرسال إشعار للمشرف بنجاح بخصوص فحص ID: {safety_check.id}")
    except Exception as e:
        current_app.logger.error(f"فشل إرسال إشعار للمشرف بخصوص فحص ID: {safety_check.id}. الخطأ: {e}")





@external_safety_bp.route('/external-safety-check/<int:vehicle_id>', methods=['GET', 'POST'])
def external_safety_check_form(vehicle_id):
    """عرض نموذج فحص السلامة الخارجي للسيارة أو معالجة البيانات المرسلة"""
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    
    if request.method == 'POST':
        return handle_safety_check_submission(vehicle)
    
    return render_template('external_safety_check.html', vehicle=vehicle)



def handle_safety_check_submission(vehicle):
    """معالجة إرسال بيانات فحص السلامة"""
    try:
        # الحصول على البيانات من النموذج
        driver_name = request.form.get('driver_name')
        driver_national_id = request.form.get('driver_national_id')
        driver_department = request.form.get('driver_department')
        driver_city = request.form.get('driver_city')
        vehicle_plate_number = request.form.get('vehicle_plate_number', vehicle.plate_number)
        vehicle_make_model = request.form.get('vehicle_make_model', f"{vehicle.make} {vehicle.model}")
        current_delegate = request.form.get('current_delegate')
        notes = request.form.get('notes')
        
        # التحقق من البيانات المطلوبة
        if not all([driver_name, driver_national_id, driver_department, driver_city]):
            return jsonify({'error': 'يرجى ملء جميع الحقول المطلوبة'}), 400
        
        # إنشاء سجل فحص السلامة
        safety_check = VehicleExternalSafetyCheck()
        safety_check.vehicle_id = vehicle.id
        safety_check.driver_name = driver_name
        safety_check.driver_national_id = driver_national_id
        safety_check.driver_department = driver_department
        safety_check.driver_city = driver_city
        safety_check.vehicle_plate_number = vehicle_plate_number
        safety_check.vehicle_make_model = vehicle_make_model
        safety_check.current_delegate = current_delegate
        safety_check.notes = notes
        safety_check.inspection_date = datetime.now()
        safety_check.approval_status = 'pending'
        
        db.session.add(safety_check)
        db.session.flush()  # للحصول على ID الجديد
        
        # معالجة الصور من الملفات (للمدراء)
        uploaded_files = request.files.getlist('file_images')
        if uploaded_files and uploaded_files[0].filename:
            for file in uploaded_files:
                if file and file.filename:
                    try:
                        # توليد اسم ملف آمن
                        ext = secure_filename(file.filename).rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
                        filename = f"{uuid.uuid4()}.{ext}"
                        
                        # قراءة بيانات الملف
                        file_data = file.read()
                        file.seek(0)  # إعادة المؤشر للبداية
                        
                        # حفظ مؤقتاً للضغط في مجلد آمن بدلاً من /tmp
                        temp_dir = os.path.join(current_app.static_folder, '.temp')
                        os.makedirs(temp_dir, exist_ok=True)
                        temp_path = os.path.join(temp_dir, filename)
                        with open(temp_path, 'wb') as f:
                            f.write(file_data)
                        
                        # ضغط الصورة
                        compress_image(temp_path)
                        
                        # قراءة الصورة المضغوطة
                        with open(temp_path, 'rb') as f:
                            compressed_data = f.read()
                        
                        # رفع إلى Object Storage
                        object_key = upload_image(compressed_data, 'safety_checks', filename)
                        
                        # 💾 الملف المؤقت يبقى محفوظاً في static/.temp/ بشكل دائم
                        
                        # حفظ في قاعدة البيانات
                        safety_image = VehicleSafetyImage()
                        safety_image.safety_check_id = safety_check.id
                        safety_image.image_path = object_key
                        safety_image.image_description = f"تم رفعها من قبل المدير"
                        db.session.add(safety_image)
                        
                    except Exception as e:
                        current_app.logger.error(f"خطأ في معالجة الملف {file.filename}: {str(e)}")
                        continue
        
        # معالجة الصور من الكاميرا (للسائقين)
        camera_images = request.form.get('camera_images', '')
        image_notes = request.form.get('image_notes', '')
        
        if camera_images:
            import base64
            
            # تقسيم الصور والملاحظات
            image_list = camera_images.split('|||') if camera_images else []
            notes_list = image_notes.split('|||') if image_notes else []
            
            for i, image_data in enumerate(image_list):
                if image_data and image_data.startswith('data:image'):
                    try:
                        # استخراج البيانات من base64
                        header, data = image_data.split(',', 1)
                        image_bytes = base64.b64decode(data)
                        
                        # تحديد التنسيق المصدر وتحويل إلى JPEG للتوافق
                        source_format = 'jpg'  # افتراضي
                        if 'png' in header:
                            source_format = 'png'
                        elif 'jpeg' in header or 'jpg' in header:
                            source_format = 'jpg'
                        elif 'heic' in header or 'heif' in header:
                            source_format = 'heic'
                        elif 'webp' in header:
                            source_format = 'webp'
                        
                        # دائماً حفظ كـ JPEG للتوافق مع جميع المتصفحات
                        ext = 'jpg'
                        
                        # إنشاء اسم ملف آمن
                        filename = f"{uuid.uuid4()}.{ext}"
                        
                        # حفظ مؤقتاً في مجلد آمن بدلاً من /tmp
                        temp_dir = os.path.join(current_app.static_folder, '.temp')
                        os.makedirs(temp_dir, exist_ok=True)
                        temp_path = os.path.join(temp_dir, filename)
                        
                        with open(temp_path, 'wb') as f:
                            f.write(image_bytes)
                        
                        # ضغط الصورة
                        success = compress_image(temp_path)
                        if not success:
                            current_app.logger.warning(f"فشل ضغط الصورة {filename}")
                        else:
                            current_app.logger.info(f"تم تحويل صورة {source_format} إلى JPEG: {filename}")
                        
                        # قراءة الصورة المضغوطة
                        with open(temp_path, 'rb') as f:
                            compressed_data = f.read()
                        
                        # رفع إلى Object Storage
                        object_key = upload_image(compressed_data, 'safety_checks', filename)
                        
                        # 💾 الملف المؤقت يبقى محفوظاً في static/.temp/ بشكل دائم
                        
                        # حفظ معلومات الصورة في قاعدة البيانات
                        description = notes_list[i] if i < len(notes_list) else None
                        
                        safety_image = VehicleSafetyImage()
                        safety_image.safety_check_id = safety_check.id
                        safety_image.image_path = object_key
                        safety_image.image_description = description
                        
                        db.session.add(safety_image)
                        
                    except Exception as e:
                        current_app.logger.error(f"خطأ في معالجة الصورة {i}: {str(e)}")
                        continue
        
        # حفظ جميع التغييرات
        db.session.commit()
        
        # إنشاء إشعارات للمسؤولين عند إنشاء فحص جديد
        try:
            # الحصول على جميع المستخدمين (الموديل مستورد في أعلى الملف)
            all_users = User.query.all()
            current_app.logger.info(f"Found {len(all_users)} users for safety check notifications")
            
            for user in all_users:
                try:
                    # استخدام الدالة المحلية create_safety_check_notification
                    create_safety_check_notification(
                        user_id=user.id,
                        vehicle_plate=safety_check.vehicle_plate_number,
                        supervisor_name=safety_check.driver_name,
                        check_status='pending',
                        check_id=safety_check.id
                    )
                    current_app.logger.info(f"Created safety check notification for user {user.id}")
                except Exception as e:
                    current_app.logger.error(f'خطأ في إنشاء إشعار للمستخدم {user.id}: {str(e)}')
        except Exception as e:
            current_app.logger.error(f'خطأ في إنشاء إشعارات فحص السلامة: {str(e)}')

        # رفع تلقائي إلى Google Drive
        try:
            uploader = VehicleDriveUploader()
            uploader.upload_safety_check(safety_check.id)
        except Exception as e:
            current_app.logger.error(f'خطأ في الرفع التلقائي إلى Google Drive: {str(e)}')
            # لا نوقف العملية إذا فشل الرفع

        send_supervisor_notification_email(safety_check)

        
        # تسجيل العملية في سجل المراجعة
        log_audit(
            user_id=current_user.id if current_user.is_authenticated else None,
            action='create',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم إنشاء طلب فحص السلامة الخارجي للسيارة {vehicle.plate_number} بواسطة {safety_check.driver_name}'
        )
        
        current_app.logger.info(f'تم إنشاء طلب فحص السلامة بنجاح: ID={safety_check.id}, Vehicle={vehicle.plate_number}')
        
        # توجيه المستخدم لصفحة التأكيد المميزة
        return redirect(url_for('external_safety.success_page', safety_check_id=safety_check.id))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'خطأ في معالجة طلب فحص السلامة: {str(e)}')
        flash('حدث خطأ أثناء معالجة الطلب', 'danger')
        return redirect(url_for('external_safety.external_safety_check_form', vehicle_id=vehicle.id))

@external_safety_bp.route('/success/<int:safety_check_id>')
def success_page(safety_check_id):
    """صفحة تأكيد إرسال طلب فحص السلامة"""
    safety_check = VehicleExternalSafetyCheck.query.get_or_404(safety_check_id)
    return render_template('external_safety_success.html', safety_check=safety_check)

@external_safety_bp.route('/status/<int:safety_check_id>')
def check_status(safety_check_id):
    """التحقق من حالة طلب فحص السلامة - للإشعارات"""
    safety_check = VehicleExternalSafetyCheck.query.get_or_404(safety_check_id)
    
    # إنشاء رسالة حسب الحالة
    if safety_check.approval_status == 'approved':
        message = {
            'type': 'success',
            'title': 'تم اعتماد الطلب',
            'text': 'تم اعتماد طلب فحص السلامة بنجاح من قبل الإدارة.'
        }
    elif safety_check.approval_status == 'rejected':
        message = {
            'type': 'error',
            'title': 'تم رفض الطلب',
            'text': f'نرجو المحاولة مرة أخرى. تم رفض الطلب.\nسبب الرفض: {safety_check.rejection_reason or "لم يتم تحديد السبب"}'
        }
    else:
        message = {
            'type': 'pending',
            'title': 'قيد المراجعة',
            'text': 'طلبك قيد المراجعة من قبل الإدارة المختصة.'
        }
    
    return jsonify(message)





@external_safety_bp.route('/')
def external_safety_index():
    """الصفحة الرئيسية لنظام فحص السلامة الخارجية"""
    return redirect(url_for('external_safety.share_links'))

@external_safety_bp.route('/share-links')
def share_links():
    """صفحة مشاركة روابط النماذج الخارجية لجميع السيارات مع الفلاتر"""
    # الحصول على معاملات الفلترة من الطلب
    status_filter = request.args.get('status', '')
    make_filter = request.args.get('make', '')
    search_plate = request.args.get('search_plate', '')
    project_filter = request.args.get('project', '')

    
    # قاعدة الاستعلام الأساسية
    query = Vehicle.query
    
    # فلترة المركبات حسب القسم المحدد للمستخدم الحالي
    from flask_login import current_user
    from models import employee_departments, Department, Employee, VehicleHandover
    if current_user.is_authenticated and hasattr(current_user, 'assigned_department_id') and current_user.assigned_department_id:
        # الحصول على معرفات الموظفين في القسم المحدد
        dept_employee_ids = db.session.query(Employee.id).join(
            employee_departments
        ).join(Department).filter(
            Department.id == current_user.assigned_department_id
        ).all()
        dept_employee_ids = [emp.id for emp in dept_employee_ids]
        
        if dept_employee_ids:
            # فلترة المركبات التي لها تسليم لموظف في القسم المحدد
            vehicle_ids_with_handovers = db.session.query(
                VehicleHandover.vehicle_id
            ).filter(
                VehicleHandover.handover_type == 'delivery',
                VehicleHandover.employee_id.in_(dept_employee_ids)
            ).distinct().all()
            
            vehicle_ids = [h.vehicle_id for h in vehicle_ids_with_handovers]
            if vehicle_ids:
                query = query.filter(Vehicle.id.in_(vehicle_ids))
            else:
                query = query.filter(Vehicle.id == -1)  # قائمة فارغة
        else:
            query = query.filter(Vehicle.id == -1)  # قائمة فارغة
    
    # إضافة التصفية حسب الحالة إذا تم تحديدها
    if status_filter:
        query = query.filter(Vehicle.status == status_filter)
    
    # إضافة التصفية حسب الشركة المصنعة إذا تم تحديدها
    if make_filter:
        query = query.filter(Vehicle.make == make_filter)
    
    # إضافة التصفية حسب المشروع إذا تم تحديده
    if project_filter:
        query = query.filter(Vehicle.project == project_filter)
    
    # إضافة البحث برقم السيارة إذا تم تحديده
    if search_plate:
        query = query.filter(Vehicle.plate_number.contains(search_plate))
    
    # الحصول على قائمة بالشركات المصنعة لقائمة التصفية
    makes = db.session.query(Vehicle.make).distinct().all()
    makes = [make[0] for make in makes]
    
    # الحصول على قائمة بالمشاريع لقائمة التصفية
    projects = db.session.query(Vehicle.project).filter(Vehicle.project.isnot(None)).distinct().all()
    projects = [project[0] for project in projects]
    
    # الحصول على قائمة السيارات
    vehicles = query.order_by(Vehicle.status, Vehicle.plate_number).all()
    all_current_drivers = get_all_current_drivers()
    all_current_drivers_with_emil = get_all_current_driversWithEmil()
    
    # قائمة حالات السيارات
    statuses = ['available', 'rented', 'in_project', 'in_workshop', 'accident']
    
    return render_template('external_safety_share_links.html', 
                           vehicles=vehicles,
                           status_filter=status_filter,
                           make_filter=make_filter,
                           search_plate=search_plate,
                           project_filter=project_filter,
                           makes=makes,
                           projects=projects,
                           statuses=statuses,
                           all_current_drivers=all_current_drivers,
                           all_current_drivers_with_emil=all_current_drivers_with_emil
                           )

@external_safety_bp.route('/share-links/export-excel')
def export_share_links_excel():
    """تصدير روابط الفحص الخارجي كملف Excel مع معلومات السيارة والسائق"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    # الحصول على معاملات الفلترة من الطلب (نفس الفلاتر المستخدمة في share_links)
    status_filter = request.args.get('status', '')
    make_filter = request.args.get('make', '')
    search_plate = request.args.get('search_plate', '')
    project_filter = request.args.get('project', '')
    
    # قاعدة الاستعلام الأساسية
    query = Vehicle.query
    
    # فلترة المركبات حسب القسم المحدد للمستخدم الحالي
    from models import employee_departments, Department, Employee, VehicleHandover
    if current_user.is_authenticated and hasattr(current_user, 'assigned_department_id') and current_user.assigned_department_id:
        dept_employee_ids = db.session.query(Employee.id).join(
            employee_departments
        ).join(Department).filter(
            Department.id == current_user.assigned_department_id
        ).all()
        dept_employee_ids = [emp.id for emp in dept_employee_ids]
        
        if dept_employee_ids:
            vehicle_ids_with_handovers = db.session.query(
                VehicleHandover.vehicle_id
            ).filter(
                VehicleHandover.handover_type == 'delivery',
                VehicleHandover.employee_id.in_(dept_employee_ids)
            ).distinct().all()
            
            vehicle_ids = [h.vehicle_id for h in vehicle_ids_with_handovers]
            if vehicle_ids:
                query = query.filter(Vehicle.id.in_(vehicle_ids))
            else:
                query = query.filter(Vehicle.id == -1)
        else:
            query = query.filter(Vehicle.id == -1)
    
    # إضافة التصفية
    if status_filter:
        query = query.filter(Vehicle.status == status_filter)
    if make_filter:
        query = query.filter(Vehicle.make == make_filter)
    if project_filter:
        query = query.filter(Vehicle.project == project_filter)
    if search_plate:
        query = query.filter(Vehicle.plate_number.contains(search_plate))
    
    # الحصول على السيارات
    vehicles = query.order_by(Vehicle.status, Vehicle.plate_number).all()
    all_current_drivers_with_details = get_all_current_driversWithEmil()
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "روابط الفحص الخارجي"
    ws.sheet_view.rightToLeft = True
    
    # تنسيقات الألوان والخطوط
    header_fill = PatternFill(start_color="171E3F", end_color="171E3F", fill_type="solid")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    cell_font = Font(name='Arial', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العناوين
    headers = ['#', 'رقم اللوحة', 'الشركة المصنعة', 'الموديل', 'السنة', 'الحالة', 'السائق الحالي', 'رقم الهوية', 'رابط الفحص الخارجي']
    ws.append(headers)
    
    # تنسيق العناوين
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # إضافة البيانات
    for idx, vehicle in enumerate(vehicles, start=1):
        driver_info = all_current_drivers_with_details.get(vehicle.id)
        driver_name = driver_info['name'] if driver_info else '-'
        national_id = driver_info['national_id'] if driver_info else '-'
        
        # ترجمة الحالة
        status_map = {
            'available': 'متاحة',
            'rented': 'مؤجرة',
            'in_project': 'في مشروع',
            'in_workshop': 'في الورشة',
            'accident': 'حادث'
        }
        status_ar = status_map.get(vehicle.status, vehicle.status)
        
        # إنشاء رابط الفحص الخارجي
        form_url = url_for('external_safety.external_safety_check_form', vehicle_id=vehicle.id, _external=True)
        
        row_data = [
            idx,
            vehicle.plate_number or '-',
            vehicle.make or '-',
            vehicle.model or '-',
            vehicle.year or '-',
            status_ar,
            driver_name,
            national_id,
            form_url
        ]
        
        ws.append(row_data)
        
        # تنسيق الخلايا
        for cell in ws[idx + 1]:
            cell.font = cell_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    # ضبط عرض الأعمدة
    column_widths = [8, 20, 18, 18, 12, 15, 25, 20, 60]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    
    # حفظ الملف في الذاكرة
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # إنشاء اسم الملف مع التاريخ
    filename = f'external_safety_links_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# في ملف الراوت الخاص بك (e.g., external_safety_bp.py)

@external_safety_bp.route('/api/send-email', methods=['POST'])
def send_vehicle_email():
    """
    نقطة نهاية (API endpoint) متكاملة لتلقي طلب إرسال بريد إلكتروني
    احترافي ومصمم لفحص المركبة عبر Resend.
    """
    # 1. استلام البيانات من الطلب القادم من JavaScript
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'الطلب فارغ أو ليس بصيغة JSON'}), 400

    driver_email = data.get('driver_email')
    driver_name = data.get('driver_name', 'زميلنا العزيز') # اسم افتراضي
    plate_number = data.get('plate_number')
    vehicle_model = data.get('vehicle_model')
    form_url = data.get('form_url')

    # التحقق من وجود جميع البيانات الضرورية
    if not all([driver_email, plate_number, vehicle_model, form_url]):
        error_message = "بيانات ناقصة في الطلب. تأكد من إرسال كل من: driver_email, plate_number, vehicle_model, form_url."
        return jsonify({'success': False, 'error': error_message}), 400

    # 2. إعداد المتغيرات الخاصة بالرسالة (الشعار والاسم)
    # ===== تم تطبيق الإصلاحات هنا =====
    company_name = os.environ.get("COMPANY_NAME", "نُــظــم لإدارة الأساطيل")
    logo_url = "https://i.postimg.cc/LXzD6b0N/logo.png" # رابط ثابت وآمن للشعار

    # 3. بناء قالب HTML الكامل للبريد الإلكتروني
    email_html_content = f"""
    <!DOCTYPE html>
    <html lang="ar" dir="rtl">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;700&display=swap');
            body {{ margin: 0; padding: 0; background-color: #f4f7f6; font-family: 'Tajawal', sans-serif; }}
            .email-container {{ max-width: 600px; margin: 20px auto; background-color: #ffffff; border: 1px solid #e0e0e0; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.05); }}
            .email-header {{ background-color: #171e3f; color: #ffffff; padding: 20px; text-align: center; }}
            .email-header img {{ max-width: 150px; margin-bottom: 10px; }}
            .email-body {{ padding: 30px; color: #333333; line-height: 1.6; text-align: right; }}
            .email-body h2 {{ color: #2c3e50; font-size: 22px; }}
            .vehicle-info {{ background-color: #f8f9fa; border: 1px solid #dee2e6; border-radius: 5px; padding: 15px; margin: 20px 0; }}
            .button-container {{ text-align: center; margin: 30px 0; }}
            .button {{ background: linear-gradient(135deg, #3498db, #2980b9); color: #ffffff !important; padding: 12px 30px; text-decoration: none; border-radius: 25px; font-weight: bold; display: inline-block; font-size: 16px; transition: transform 0.2s ease; }}
            .button:hover {{ transform: translateY(-2px); }}
            .instructions-section {{ margin-top: 25px; border-top: 1px solid #eeeeee; padding-top: 20px; }}
            .instructions-section h3 {{ color: #e67e22; font-size: 18px; }}
            .instructions-section ul {{ padding-right: 20px; list-style-type: '✔️  '; }}
            .email-footer {{ background-color: #2c3e50; color: #bdc3c7; padding: 20px; text-align: center; font-size: 12px; }}
        </style>
    </head>
    <body>
        <div class="email-container">
            <div class="email-header">
                <img src="{logo_url}" alt="{company_name} Logo">
                <h1>{company_name}</h1>
            </div>
            <div class="email-body">
                <h2>إجراء مطلوب: فحص السلامة الخارجي للمركبة</h2>
                <p>مرحباً <strong>{driver_name}</strong> 👋،</p>
                <p>نرجو منك تعبئة نموذج فحص السلامة الخارجي للمركبة التالية بدقة وعناية.</p>
                <div class="vehicle-info">
                    🚗 <strong>المركبة:</strong> {plate_number} ({vehicle_model})
                </div>
                <p><strong>👇 الرابط المباشر للنموذج:</strong></p>
                <div class="button-container">
                    <a href="{form_url}" class="button">فتح نموذج الفحص</a>
                </div>
                <div class="instructions-section">
                    <h3>📋 التعليمات المطلوبة (مهم جدًا):</h3>
                    <h4>1️⃣ الصور الأساسية (إلزامية):</h4>
                    <ul>
                        <li>صورة من <strong>الأمام</strong> (تظهر كامل واجهة المركبة).</li>
                        <li>صورة من <strong>الخلف</strong> (تظهر كامل خلفية المركبة).</li>
                        <li>صورة من <strong>الجانب الأيمن والأيسر</strong> (بشكل واضح وكامل).</li>
                        <li>صورة <strong>لسقف</strong> المركبة.</li>
                        <li>صورة لـ <strong>أسفل المركبة من الأمام</strong>.</li>
                    </ul>
                    <h4>2️⃣ صور الملاحظات (إن وجدت):</h4>
                    <ul>
                        <li>إذا وجدت أي خدوش، صدمات، أو عيوب، قم بتصويرها عن قرب.</li>
                        <li><strong>هام:</strong> قم بالإشارة بإصبعك إلى مكان الملاحظة في الصورة.</li>
                        <li>اكتب وصفاً لكل ملاحظة أسفل الصورة المرفوعة.</li>
                    </ul>
                </div>
                <div class="instructions-section">
                    <h3>✅ ما بعد إرسال النموذج:</h3>
                    <ul>
                        <li><strong>في حال القبول:</strong> سيتم إعلامك وتفعيل إجراءات الوقود.</li>
                        <li><strong>في حال الرفض:</strong> ستصلك رسالة بالسبب. يرجى الدخول على نفس الرابط مجدداً وتصحيح الملاحظات.</li>
                    </ul>
                </div>
                <p>شكرًا لتعاونكم وحرصكم على السلامة.</p>
            </div>
            <div class="email-footer">
                <p>هذه رسالة آلية من {company_name}.</p>
                <p>© {datetime.now().year} جميع الحقوق محفوظة.</p>
            </div>
        </div>
    </body>
    </html>
    """

    # 4. بناء طلب الإرسال واستدعاء Resend API
    try:
        params = {
            "from": f"{company_name} <onboarding@resend.dev>",
            "to": [driver_email],
            "subject": f"إجراء مطلوب: فحص السلامة للمركبة {plate_number}",
            "html": email_html_content,
        }
        sent_email = resend.Emails.send(params)

        # يمكنك تفعيل السطر التالي للتشخيص إذا لزم الأمر
        # current_app.logger.info(f"Email sent successfully. ID: {sent_email['id']}")

        return jsonify({'success': True, 'message': f"تم إرسال البريد الإلكتروني بنجاح إلى {driver_email}"})

    except Exception as e:
        # تسجيل الخطأ بالتفصيل في سجلات الخادم للمساعدة في التشخيص
        current_app.logger.error(f"Error sending email with Resend: {e}")
        # إرجاع رسالة خطأ واضحة
        return jsonify({'success': False, 'error': f"فشل في إرسال البريد عبر الخدمة الخارجية: {str(e)}"}), 500

# # # ----- أضف هذه الدالة الجديدة لمشروعك -----
# @external_safety_bp.route('/api/send-email', methods=['POST'])
# def send_vehicle_email():
#     """
#     نقطة نهاية (API endpoint) متكاملة لتلقي طلب إرسال بريد إلكتروني
#     احترافي ومصمم لفحص المركبة عبر Resend.
#     """
#     # 1. استلام البيانات من الطلب القادم من JavaScript
#     data = request.get_json()
#     if not data:
#         return jsonify({'success': False, 'error': 'الطلب فارغ أو ليس بصيغة JSON'}), 400

#     driver_email = data.get('driver_email')
#     driver_name = data.get('driver_name', 'زميلنا العزيز') # اسم افتراضي
#     plate_number = data.get('plate_number')
#     vehicle_model = data.get('vehicle_model')
#     form_url = data.get('form_url')

#     # التحقق من وجود جميع البيانات الضرورية
#     if not all([driver_email, plate_number, vehicle_model, form_url]):
#         error_message = "بيانات ناقصة في الطلب. تأكد من إرسال كل من: driver_email, plate_number, vehicle_model, form_url."
#         return jsonify({'success': False, 'error': error_message}), 400

#     # 2. إعداد المتغيرات الخاصة بالرسالة (الشعار والاسم)
#     # company_name = "شركة رأس السعودية المحدوده"  # <--- يمكنك تغيير هذا
#     # تأكد من أن مسار الشعار صحيح. _external=True ضروري لتوليد رابط كامل.
#     logo_path = 'images/logo.png' # <--- يمكنك تغيير هذا
#     try:
#         logo_url = url_for('static', filename=logo_path, _external=True)

#     except RuntimeError:
#         # هذا الحل الاحتياطي يعمل إذا تم استدعاء الدالة خارج سياق الطلب
#         # (على الرغم من أنه في حالتك لن يحدث ذلك مع استدعاء API)
#         logo_url = "https://your-fallback-domain.com" + url_for('static', filename=logo_path)


#     # 3. بناء قالب HTML الكامل للبريد الإلكتروني
#     email_html_content = f"""
#     <!DOCTYPE html>
#     <html lang="ar" dir="rtl">
#     <head>
#         <meta charset="UTF-8">
#         <meta name="viewport" content="width=device-width, initial-scale=1.0">
#         <style>
#             body {{ margin: 0; padding: 0; background-color: #f4f7f6; font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; }}
#             .email-container {{ max-width: 600px; margin: 20px auto; background-color: #ffffff; border: 1px solid #e0e0e0; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.05); }}
#             .email-header {{ background-color: #171e3f; color: #ffffff; padding: 20px; text-align: center; }}
#             .email-header img {{ max-width: 150px; margin-bottom: 10px; }}
#             .email-body {{ padding: 30px; color: #333333; line-height: 1.6; text-align: right; }}
#             .email-body h2 {{ color: #2c3e50; font-size: 22px; }}
#             .vehicle-info {{ background-color: #f8f9fa; border: 1px solid #dee2e6; border-radius: 5px; padding: 15px; margin: 20px 0; }}
#             .button-container {{ text-align: center; margin: 30px 0; }}
#             .button {{ background-color: #3498db; color: #ffffff !important; padding: 12px 30px; text-decoration: none; border-radius: 25px; font-weight: bold; display: inline-block; font-size: 16px; }}
#             .instructions-section {{ margin-top: 25px; border-top: 1px solid #eeeeee; padding-top: 20px; }}
#             .instructions-section h3 {{ color: #e67e22; font-size: 18px; }}
#             .instructions-section ul {{ padding-right: 20px; list-style-type: '✔️ '; }}
#             .email-footer {{ background-color: #2c3e50; color: #bdc3c7; padding: 20px; text-align: center; font-size: 12px; }}
#         </style>
#     </head>
#     <body>
#         <div class="email-container">
#             <div class="email-header">
#                 <img src="https://i.postimg.cc/LXzD6b0N/logo.png" alt="نُــظــم  للحلول البرمجية">
#                 <h1>{company_name}</h1>
#             </div>
#             <div class="email-body">
#                 <h2>إجراء مطلوب: فحص السلامة الخارجي للمركبة</h2>
#                 <p>مرحباً <strong>{driver_name}</strong> 👋،</p>
#                 <p>نرجو منك تعبئة نموذج فحص السلامة الخارجي للمركبة التالية بدقة وعناية.</p>
#                 <div class="vehicle-info">
#                     🚗 <strong>المركبة:</strong> {plate_number} ({vehicle_model})
#                 </div>
#                 <p><strong>👇 الرابط المباشر للنموذج:</strong></p>
#                 <div class="button-container">
#                     <a href="{form_url}" class="button">فتح نموذج الفحص</a>
#                 </div>
#                 <div class="instructions-section">
#                     <h3>📋 التعليمات المطلوبة (مهم جدًا):</h3>
#                     <h4>1️⃣ الصور الأساسية (إلزامية):</h4>
#                     <ul>
#                         <li>صورة من <strong>الأمام</strong> (تظهر كامل واجهة المركبة).</li>
#                         <li>صورة من <strong>الخلف</strong> (تظهر كامل خلفية المركبة).</li>
#                         <li>صورة من <strong>الجانب الأيمن والأيسر</strong> (من الزاوية).</li>
#                         <li>صورة <strong>لسقف</strong> المركبة.</li>
#                         <li>صورة لـ <strong>أسفل المركبة من الأمام</strong>.</li>
#                     </ul>
#                     <h4>2️⃣ صور الملاحظات (إن وجدت):</h4>
#                     <ul>
#                         <li>إذا وجدت أي خدوش، صدمات، أو عيوب، قم بتصويرها عن قرب.</li>
#                         <li><strong>هام:</strong> قم بالإشارة بإصبعك إلى مكان الملاحظة في الصورة.</li>
#                         <li>اكتب وصفاً لكل ملاحظة أسفل الصورة المرفوعة.</li>
#                     </ul>
#                 </div>
#                 <div class="instructions-section">
#                     <h3>✅ ما بعد إرسال النموذج:</h3>
#                     <ul>
#                         <li><strong>في حال القبول:</strong> سيتم إعلامك وتفعيل إجراءات الوقود.</li>
#                         <li><strong>في حال الرفض:</strong> ستصلك رسالة بالسبب. يرجى الدخول على نفس الرابط مجدداً وتصحيح الملاحظات.</li>
#                     </ul>
#                 </div>
#                 <p>شكرًا لتعاونكم وحرصكم على السلامة.</p>
#             </div>
#             <div class="email-footer">
#                 <p>هذه رسالة آلية من {company_name}.</p>
#                 <p>© {datetime.now().year} جميع الحقوق محفوظة.</p>
#             </div>
#         </div>
#     </body>
#     </html>
#     """

#     # 4. بناء طلب الإرسال واستدعاء Resend API
#     try:
#         params = {
#             "from": f"{company_name} <onboarding@resend.dev>",
#             "to": [driver_email],
#             "subject": f"إجراء مطلوب: فحص السلامة للمركبة {plate_number}",
#             "html": email_html_content,
#         }
#         sent_email = resend.Emails.send(params)
        
#         # يمكنك تفعيل هذه للتشخيص
#         # print(f"Email sent successfully. ID: {sent_email['id']}")
        
#         return jsonify({'success': True, 'message': f"تم إرسال البريد الإلكتروني بنجاح إلى {driver_email}"})

#     except Exception as e:
#         # في حال حدوث خطأ من Resend أو غيره
#         print(f"Error sending email with Resend: {e}")
#         return jsonify({'success': False, 'error': str(e)}), 500



@external_safety_bp.route('/api/verify-employee/<national_id>')
def verify_employee(national_id):
    """التحقق من الموظف بواسطة رقم الهوية"""
    try:
        # البحث عن الموظف بواسطة رقم الهوية
        employee = Employee.query.filter_by(national_id=national_id).first()
        
        if not employee:
            return jsonify({'success': False, 'message': 'الموظف غير موجود'}), 404
        
        # الحصول على أسماء الأقسام
        department_names = [dept.name for dept in employee.departments] if employee.departments else []
        
        return jsonify({
            'success': True,
            'employee': {
                'id': employee.id,
                'name': employee.name,
                'department': ', '.join(department_names) if department_names else 'غير محدد',
                'city': employee.city if hasattr(employee, 'city') else 'الرياض',
                'national_id': employee.national_id
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"خطأ في التحقق من الموظف: {str(e)}")
        return jsonify({'success': False, 'message': 'حدث خطأ في التحقق من الموظف'}), 500

@external_safety_bp.route('/external-safety-check/success')
def external_safety_success():
    """صفحة نجاح إرسال طلب فحص السلامة"""
    return render_template('external_safety_success.html')

@external_safety_bp.route('/admin/external-safety-checks')
def admin_external_safety_checks():
    """عرض جميع طلبات فحص السلامة للإدارة مع الفلاتر"""
    from flask_login import current_user
    from models import employee_departments, Department, Employee, VehicleHandover, Vehicle
    
    # التحقق من تسجيل الدخول
    if not current_user.is_authenticated:
        flash('يرجى تسجيل الدخول أولاً', 'error')
        return redirect('/login')
    
    # الحصول على معايير الفلترة من request
    vehicle_filter = request.args.get('vehicle_filter', '').strip()
    vehicle_search = request.args.get('vehicle_search', '').strip()
    department_filter = request.args.get('department_filter', '').strip()
    status_filter = request.args.get('status_filter', '').strip()
    
    # بناء الاستعلام مع الفلاتر
    query = VehicleExternalSafetyCheck.query
    
    # فلترة فحوصات السلامة حسب القسم المحدد للمستخدم الحالي
    if current_user.is_authenticated and hasattr(current_user, 'assigned_department_id') and current_user.assigned_department_id:
        # الحصول على معرفات الموظفين في القسم المحدد
        dept_employee_ids = db.session.query(Employee.id).join(
            employee_departments
        ).join(Department).filter(
            Department.id == current_user.assigned_department_id
        ).all()
        dept_employee_ids = [emp.id for emp in dept_employee_ids]
        
        if dept_employee_ids:
            # فلترة فحوصات السلامة للمركبات المسلمة لموظفي القسم المحدد
            dept_vehicle_plates = db.session.query(Vehicle.plate_number).join(
                VehicleHandover, Vehicle.id == VehicleHandover.vehicle_id
            ).filter(
                VehicleHandover.handover_type == 'delivery',
                VehicleHandover.employee_id.in_(dept_employee_ids)
            ).distinct().all()
            dept_vehicle_plates = [v.plate_number for v in dept_vehicle_plates]
            if dept_vehicle_plates:
                query = query.filter(VehicleExternalSafetyCheck.vehicle_plate_number.in_(dept_vehicle_plates))
            else:
                query = query.filter(VehicleExternalSafetyCheck.id == -1)  # قائمة فارغة
        else:
            query = query.filter(VehicleExternalSafetyCheck.id == -1)  # قائمة فارغة
    
    # فلترة حسب رقم السيارة (من القائمة المنسدلة)
    if vehicle_filter:
        query = query.filter(VehicleExternalSafetyCheck.vehicle_plate_number.contains(vehicle_filter))
    
    # البحث في السيارة (من حقل البحث)
    if vehicle_search:
        query = query.filter(VehicleExternalSafetyCheck.vehicle_plate_number.contains(vehicle_search))
    
    # فلترة حسب القسم
    if department_filter:
        query = query.filter(VehicleExternalSafetyCheck.driver_department.contains(department_filter))
    
    # فلترة حسب الحالة
    if status_filter:
        query = query.filter(VehicleExternalSafetyCheck.approval_status == status_filter)
    
    # جلب النتائج مرتبة حسب التاريخ
    safety_checks = query.order_by(VehicleExternalSafetyCheck.created_at.desc()).all()
    
    # إحصائيات للفلاتر
    total_checks = VehicleExternalSafetyCheck.query.count()
    pending_checks = VehicleExternalSafetyCheck.query.filter_by(approval_status='pending').count()
    approved_checks = VehicleExternalSafetyCheck.query.filter_by(approval_status='approved').count()
    rejected_checks = VehicleExternalSafetyCheck.query.filter_by(approval_status='rejected').count()
    
    # جلب قائمة السيارات والأقسام للفلاتر
    vehicles_list = db.session.query(VehicleExternalSafetyCheck.vehicle_plate_number).distinct().all()
    vehicles_list = [v[0] for v in vehicles_list if v[0]]
    
    departments_list = db.session.query(VehicleExternalSafetyCheck.driver_department).distinct().all()
    departments_list = [d[0] for d in departments_list if d[0]]
    
    # جلب جميع السيارات لنموذج إنشاء الفحص مع مراعاة فلترة القسم
    vehicles_query = Vehicle.query
    
    # فلترة السيارات حسب القسم المحدد للمستخدم الحالي
    if current_user.is_authenticated and hasattr(current_user, 'assigned_department_id') and current_user.assigned_department_id:
        # الحصول على معرفات الموظفين في القسم المحدد
        dept_employee_ids = db.session.query(Employee.id).join(
            employee_departments
        ).join(Department).filter(
            Department.id == current_user.assigned_department_id
        ).all()
        dept_employee_ids = [emp.id for emp in dept_employee_ids]
        
        if dept_employee_ids:
            # فلترة المركبات التي لها تسليم لموظف في القسم المحدد
            vehicle_ids_with_handovers = db.session.query(
                VehicleHandover.vehicle_id
            ).filter(
                VehicleHandover.handover_type == 'delivery',
                VehicleHandover.employee_id.in_(dept_employee_ids)
            ).distinct().all()
            
            vehicle_ids = [h.vehicle_id for h in vehicle_ids_with_handovers]
            if vehicle_ids:
                vehicles_query = vehicles_query.filter(Vehicle.id.in_(vehicle_ids))
            else:
                vehicles_query = vehicles_query.filter(Vehicle.id == -1)  # قائمة فارغة
        else:
            vehicles_query = vehicles_query.filter(Vehicle.id == -1)  # قائمة فارغة
    
    all_vehicles = vehicles_query.order_by(Vehicle.plate_number).all()
    
    return render_template('admin_external_safety_checks.html', 
                         safety_checks=safety_checks,
                         vehicle_filter=vehicle_filter,
                         vehicle_search=vehicle_search,
                         department_filter=department_filter,
                         status_filter=status_filter,
                         vehicles_list=vehicles_list,
                         departments_list=departments_list,
                         total_checks=total_checks,
                         pending_checks=pending_checks,
                         approved_checks=approved_checks,
                         rejected_checks=rejected_checks,
                         all_vehicles=all_vehicles)

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>')
def admin_view_safety_check(check_id):
    """عرض تفاصيل طلب فحص السلامة"""
    if not current_user.is_authenticated:
        flash('يرجى تسجيل الدخول أولاً', 'error')
        return redirect('/login')
    
    # استخدام العلاقة المحددة مسبقاً لجلب الصور مع فحص السلامة
    safety_check = VehicleExternalSafetyCheck.query.options(
        db.selectinload(VehicleExternalSafetyCheck.safety_images)
    ).get_or_404(check_id)
    
    current_app.logger.info(f'تم جلب فحص السلامة ID={check_id} مع {len(safety_check.safety_images)} صور')
    
    return render_template('admin_view_safety_check.html', safety_check=safety_check)

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/reject', methods=['GET', 'POST'])
def reject_safety_check_page(check_id):
    """صفحة رفض طلب فحص السلامة"""
    if not current_user.is_authenticated:
        flash('يرجى تسجيل الدخول أولاً', 'error')
        return redirect('/login')
    
    safety_check = VehicleExternalSafetyCheck.query.get_or_404(check_id)
    
    if request.method == 'POST':
        # معالجة رفض الطلب
        rejection_reason = request.form.get('rejection_reason')
        
        if not rejection_reason or not rejection_reason.strip():
            flash('يرجى كتابة سبب الرفض', 'error')
            return render_template('admin_reject_safety_check.html', safety_check=safety_check)
        
        # تحديث حالة الطلب
        safety_check.approval_status = 'rejected'
        safety_check.rejection_reason = rejection_reason.strip()
        safety_check.approved_by = current_user.id
        safety_check.approved_at = datetime.now()
        
        db.session.commit()
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='reject',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم رفض طلب فحص السلامة للسيارة {safety_check.vehicle_plate_number}. السبب: {rejection_reason}'
        )
        
        current_app.logger.info(f'تم رفض طلب فحص السلامة ID={safety_check.id} بواسطة {current_user.name}')
        
        flash('تم رفض الطلب بنجاح', 'success')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))
    
    return render_template('admin_reject_safety_check.html', safety_check=safety_check)
    return render_template('admin_view_safety_check.html', safety_check=safety_check)

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/approve', methods=['POST'])
def approve_safety_check(check_id):
    """اعتماد طلب فحص السلامة"""
    if not current_user.is_authenticated:
        return jsonify({'error': 'غير مصرح لك'}), 403
    
    try:
        safety_check = VehicleExternalSafetyCheck.query.get_or_404(check_id)
        
        safety_check.approval_status = 'approved'
        safety_check.approved_by = current_user.id
        safety_check.approved_at = datetime.now()
        
        db.session.commit()
        
        # إنشاء إشعار عند الموافقة
        try:
            if current_user.is_authenticated:
                create_safety_check_review_notification(
                    user_id=current_user.id,
                    vehicle_plate=safety_check.vehicle_plate_number,
                    action='approved',
                    reviewer_name=current_user.name,
                    check_id=safety_check.id
                )
        except Exception as e:
            current_app.logger.error(f'خطأ في إنشاء إشعار الموافقة: {str(e)}')
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='approve',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم اعتماد طلب فحص السلامة للسيارة {safety_check.vehicle_plate_number}'
        )
        
        flash('تم اعتماد طلب فحص السلامة بنجاح', 'success')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في اعتماد طلب فحص السلامة: {str(e)}")
        flash('حدث خطأ في اعتماد الطلب', 'error')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/reject', methods=['POST'])
def reject_safety_check(check_id):
    """رفض طلب فحص السلامة"""
    if not current_user.is_authenticated:
        return jsonify({'error': 'غير مصرح لك'}), 403
    
    try:
        safety_check = VehicleExternalSafetyCheck.query.get_or_404(check_id)
        
        safety_check.approval_status = 'rejected'
        safety_check.approved_by = current_user.id
        safety_check.approved_at = datetime.now()
        safety_check.rejection_reason = request.form.get('rejection_reason', '')
        
        db.session.commit()
        
        # إنشاء إشعار عند الرفض
        try:
            if current_user.is_authenticated:
                create_safety_check_review_notification(
                    user_id=current_user.id,
                    vehicle_plate=safety_check.vehicle_plate_number,
                    action='rejected',
                    reviewer_name=current_user.name,
                    check_id=safety_check.id
                )
        except Exception as e:
            current_app.logger.error(f'خطأ في إنشاء إشعار الرفض: {str(e)}')
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='reject',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم رفض طلب فحص السلامة للسيارة {safety_check.vehicle_plate_number}. السبب: {safety_check.rejection_reason}'
        )
        
        flash('تم رفض طلب فحص السلامة', 'success')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في رفض طلب فحص السلامة: {str(e)}")
        flash('حدث خطأ في رفض الطلب', 'error')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/delete', methods=['GET', 'POST'])
def delete_external_safety_check(check_id):
    """حذف طلب فحص السلامة"""
    if not current_user.is_authenticated:
        flash('يرجى تسجيل الدخول أولاً', 'error')
        return redirect('/login')
    
    safety_check = VehicleExternalSafetyCheck.query.get_or_404(check_id)
    
    if request.method == 'GET':
        # عرض صفحة تأكيد الحذف
        return render_template('admin_delete_safety_check.html', safety_check=safety_check)
    
    # POST method
    try:
        # 💾 الصور تبقى محفوظة - لا نحذف الملفات الفعلية
        # نحذف فقط المراجع من قاعدة البيانات للأمان
        current_app.logger.info(f"💾 الصور محفوظة للأمان ({len(safety_check.safety_images)} صورة)")
        
        # تسجيل العملية قبل الحذف
        log_audit(
            user_id=current_user.id,
            action='delete',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم حذف طلب فحص السلامة للسيارة {safety_check.vehicle_plate_number} - السائق: {safety_check.driver_name}'
        )
        
        # حذف السجل من قاعدة البيانات
        db.session.delete(safety_check)
        db.session.commit()
        
        flash('تم حذف طلب فحص السلامة بنجاح', 'success')
        return redirect(url_for('external_safety.admin_external_safety_checks'))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في حذف طلب فحص السلامة: {str(e)}")
        flash('حدث خطأ في حذف الطلب', 'error')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/edit', methods=['GET', 'POST'])
def edit_safety_check(check_id):
    """تعديل طلب فحص السلامة"""
    if not current_user.is_authenticated:
        flash('يرجى تسجيل الدخول أولاً', 'error')
        return redirect(url_for('external_safety.admin_external_safety_checks'))
    
    safety_check = VehicleExternalSafetyCheck.query.get_or_404(check_id)
    
    if request.method == 'POST':
        try:
            # تحديث البيانات
            safety_check.current_delegate = request.form.get('current_delegate', '')
            inspection_date_str = request.form.get('inspection_date')
            safety_check.inspection_date = datetime.fromisoformat(inspection_date_str) if inspection_date_str else datetime.now()
            safety_check.driver_name = request.form.get('driver_name', '')
            safety_check.driver_national_id = request.form.get('driver_national_id', '')
            safety_check.driver_department = request.form.get('driver_department', '')
            safety_check.driver_city = request.form.get('driver_city', '')
            safety_check.notes = request.form.get('notes', '')
            
            # معالجة رفع ملف PDF - بترتيب آمن
            if 'pdf_file' in request.files:
                pdf_file = request.files['pdf_file']
                if pdf_file and pdf_file.filename:
                    # التحقق من نوع الملف
                    if pdf_file.filename.lower().endswith('.pdf'):
                        # 1️⃣ حفظ الملف الجديد أولاً
                        filename = f"{uuid.uuid4()}_{pdf_file.filename}"
                        pdf_path = upload_image(pdf_file, 'safety_checks_pdfs', filename)
                        
                        if pdf_path:
                            # 2️⃣ حفظ مسار القديم للحذف لاحقاً
                            old_pdf_path = safety_check.pdf_file_path
                            # 3️⃣ تحديث DB
                            safety_check.pdf_file_path = pdf_path
                            current_app.logger.info(f'تم رفع ملف PDF للفحص {safety_check.id}: {pdf_path}')
                            
                            # 4️⃣ حذف القديم بعد نجاح الحفظ
                            if old_pdf_path:
                                try:
                                    delete_image(old_pdf_path)
                                    current_app.logger.info(f'✅ تم حذف PDF القديم: {old_pdf_path}')
                                except Exception as del_err:
                                    current_app.logger.warning(f'⚠️ لم يتم حذف PDF القديم: {del_err}')
                        else:
                            flash('فشل في رفع ملف PDF', 'error')
                            return render_template('admin_edit_safety_check.html', safety_check=safety_check)
                    else:
                        flash('يرجى رفع ملف بصيغة PDF فقط', 'error')
                        return render_template('admin_edit_safety_check.html', safety_check=safety_check)
            
            # معالجة الرابط الخارجي (Google Drive, Dropbox, إلخ)
            drive_pdf_link = request.form.get('drive_pdf_link', '').strip()
            if drive_pdf_link:
                # التحقق من صحة الرابط
                if drive_pdf_link.startswith(('http://', 'https://')):
                    safety_check.drive_pdf_link = drive_pdf_link
                    current_app.logger.info(f'تم تحديث رابط خارجي للفحص {safety_check.id}: {drive_pdf_link[:50]}...')
                else:
                    flash('الرابط يجب أن يبدأ بـ http:// أو https://', 'warning')
            elif 'drive_pdf_link' in request.form:
                # إذا كان الحقل موجود لكن فارغ، نحذف الرابط
                safety_check.drive_pdf_link = None
            
            # معالجة رفع صور جديدة
            if 'new_images' in request.files:
                new_images = request.files.getlist('new_images')
                uploaded_count = 0
                
                for image_file in new_images:
                    if image_file and image_file.filename:
                        try:
                            # إنشاء اسم ملف فريد
                            filename = f"{uuid.uuid4()}_{secure_filename(image_file.filename)}"
                            
                            # رفع الصورة باستخدام storage_helper
                            image_path = upload_image(image_file, 'safety_checks', filename)
                            
                            # إنشاء سجل جديد للصورة
                            new_image = VehicleSafetyImage(
                                safety_check_id=safety_check.id,
                                image_path=image_path,
                                image_description=''
                            )
                            db.session.add(new_image)
                            uploaded_count += 1
                            
                            current_app.logger.info(f'تم رفع صورة جديدة للفحص {safety_check.id}: {image_path}')
                        except Exception as e:
                            current_app.logger.error(f'خطأ في رفع صورة: {str(e)}')
                            continue
                
                if uploaded_count > 0:
                    flash(f'تم رفع {uploaded_count} صورة جديدة بنجاح', 'success')
            
            # تحديث أوصاف الصور
            for image in safety_check.safety_images:
                description_field = f'image_description_{image.id}'
                if description_field in request.form:
                    image.image_description = request.form.get(description_field, '')
            
            # تحديث تاريخ التعديل
            safety_check.updated_at = datetime.now()
            
            db.session.commit()
            
            # تسجيل العملية
            log_audit(
                user_id=current_user.id,
                action='update',
                entity_type='VehicleExternalSafetyCheck',
                entity_id=safety_check.id,
                details=f'تم تحديث طلب فحص السلامة للسيارة {safety_check.vehicle_plate_number}'
            )
            
            flash('تم تحديث طلب فحص السلامة بنجاح', 'success')
            return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"خطأ في تحديث طلب فحص السلامة: {str(e)}")
            flash('حدث خطأ في تحديث الطلب', 'error')
    
    return render_template('admin_edit_safety_check.html', safety_check=safety_check)

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/delete-images', methods=['POST'])
def delete_safety_check_images(check_id):
    """حذف صور محددة من طلب فحص السلامة"""
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'message': 'غير مصرح لك'}), 403
    
    try:
        safety_check = VehicleExternalSafetyCheck.query.get_or_404(check_id)
        
        # الحصول على معرفات الصور المراد حذفها
        data = request.get_json()
        image_ids = data.get('image_ids', [])
        
        if not image_ids:
            return jsonify({'success': False, 'message': 'لم يتم تحديد أي صور للحذف'}), 400
        
        # حذف الصور المحددة
        deleted_count = 0
        for image_id in image_ids:
            image = VehicleSafetyImage.query.filter_by(
                id=image_id,
                safety_check_id=check_id
            ).first()
            
            if image:
                try:
                    # حذف الصورة من التخزين
                    if image.image_path:
                        delete_image(image.image_path)
                    
                    # حذف السجل من قاعدة البيانات
                    db.session.delete(image)
                    deleted_count += 1
                except Exception as e:
                    current_app.logger.error(f"خطأ في حذف الصورة {image_id}: {str(e)}")
                    continue
        
        db.session.commit()
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='delete_images',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم حذف {deleted_count} صورة من طلب فحص السلامة للسيارة {safety_check.vehicle_plate_number}'
        )
        
        return jsonify({
            'success': True,
            'message': f'تم حذف {deleted_count} صورة بنجاح'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في حذف الصور: {str(e)}")
        return jsonify({'success': False, 'message': 'حدث خطأ في حذف الصور'}), 500

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/delete', methods=['POST'])
def delete_safety_check(check_id):
    """حذف طلب فحص السلامة"""
    if not current_user.is_authenticated:
        return jsonify({'error': 'غير مصرح لك'}), 403
    
    try:
        safety_check = VehicleExternalSafetyCheck.query.get_or_404(check_id)
        
        # 💾 الصور تبقى محفوظة - لا نحذف الملفات الفعلية
        current_app.logger.info(f"💾 الصور محفوظة للأمان ({len(safety_check.safety_images)} صورة)")
        
        # تسجيل العملية قبل الحذف
        log_audit(
            user_id=current_user.id,
            action='delete',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم حذف طلب فحص السلامة للسيارة {safety_check.vehicle_plate_number}'
        )
        
        db.session.delete(safety_check)
        db.session.commit()
        
        flash('تم حذف طلب فحص السلامة بنجاح', 'success')
        return redirect(url_for('external_safety.admin_external_safety_checks'))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في حذف طلب فحص السلامة: {str(e)}")
        flash('حدث خطأ في حذف الطلب', 'error')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/pdf')
def view_safety_check_pdf(check_id):
    """عرض صفحة جميلة لفحص السلامة مع زر تحميل PDF"""
    if not current_user.is_authenticated:
        flash('يرجى تسجيل الدخول أولاً', 'error')
        return redirect(url_for('external_safety.admin_external_safety_checks'))
    
    try:
        safety_check = VehicleExternalSafetyCheck.query.options(
            db.selectinload(VehicleExternalSafetyCheck.safety_images)
        ).get_or_404(check_id)
        
        return render_template('external_safety_check_view.html', safety_check=safety_check)
        
    except Exception as e:
        current_app.logger.error(f"خطأ في عرض صفحة فحص السلامة: {str(e)}")
        flash('حدث خطأ في عرض الطلب', 'error')
        return redirect(url_for('external_safety.admin_external_safety_checks'))

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/pdf/download')
def export_safety_check_pdf(check_id):
    """تصدير طلب فحص السلامة كملف PDF بتصميم احترافي عصري"""
    if not current_user.is_authenticated:
        flash('يرجى تسجيل الدخول أولاً', 'error')
        return redirect(url_for('external_safety.admin_external_safety_checks'))
    
    try:
        from weasyprint import HTML, CSS
        from io import BytesIO
        import base64
        
        safety_check = VehicleExternalSafetyCheck.query.options(
            db.selectinload(VehicleExternalSafetyCheck.safety_images)
        ).get_or_404(check_id)
        
        # تحويل الصور إلى base64
        for image in safety_check.safety_images:
            try:
                # محاولة عدة مسارات للملف
                possible_paths = [
                    os.path.join(current_app.root_path, 'static/uploads', image.image_path),
                    os.path.join(current_app.root_path, image.image_path),
                    os.path.join('static/uploads', image.image_path),
                    image.image_path
                ]
                
                img_file_path = None
                for path in possible_paths:
                    if os.path.exists(path):
                        img_file_path = path
                        break
                
                if img_file_path:
                    with open(img_file_path, 'rb') as img_file:
                        image_data = img_file.read()
                        image.image_base64 = base64.b64encode(image_data).decode()
                        current_app.logger.info(f"تم قراءة الصورة: {img_file_path}, حجم: {len(image_data)}")
                else:
                    current_app.logger.warning(f"لم يتم العثور على الصورة: {image.image_path}")
                    image.image_base64 = ''
            except Exception as e:
                current_app.logger.error(f"خطأ في تحويل الصورة: {str(e)}")
                image.image_base64 = ''
        
        # الحصول على مسار الشعار RASSCO
        logo_path = os.path.join(current_app.root_path, 'static/img/rassco_logo.png')
        logo_uri = f"file://{logo_path}"
        
        # إنشاء HTML للـ PDF بتصميم احترافي
        html_content = render_template('safety_check_pdf_template.html', 
                                      safety_check=safety_check, 
                                      now=datetime.now(),
                                      logo_path=logo_uri)
        
        # تحويل HTML إلى PDF
        pdf_buffer = BytesIO()
        HTML(string=html_content).write_pdf(pdf_buffer)
        pdf_buffer.seek(0)
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='export_pdf',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم تصدير طلب فحص السلامة للسيارة {safety_check.vehicle_plate_number} كملف PDF'
        )
        
        # تنسيق اسم الملف برقم اللوحة واسم السائق والتاريخ
        inspection_date = safety_check.inspection_date.strftime('%Y-%m-%d') if safety_check.inspection_date else 'بدون_تاريخ'
        plate_number = safety_check.vehicle_plate_number.replace(' ', '_')
        driver_name = safety_check.driver_name.replace(' ', '_') if safety_check.driver_name else 'غير_محدد'
        
        # إرسال الـ PDF
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f'{plate_number}_{driver_name}_{inspection_date}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        current_app.logger.error(f"خطأ في تصدير طلب فحص السلامة كـ PDF: {str(e)}")
        flash('حدث خطأ في تصدير الطلب', 'error')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/image/<int:image_id>')
def view_safety_check_image(check_id, image_id):
    """عرض صورة فحص السلامة في صفحة منفصلة مع الملاحظات"""
    if not current_user.is_authenticated:
        return redirect(url_for('auth.login'))
    
    try:
        safety_check = VehicleExternalSafetyCheck.query.get_or_404(check_id)
        image = VehicleSafetyImage.query.filter_by(id=image_id, safety_check_id=check_id).first_or_404()
        
        # حساب عدد الصورة من إجمالي الصور
        all_images = VehicleSafetyImage.query.filter_by(safety_check_id=check_id).all()
        image_number = next((i+1 for i, img in enumerate(all_images) if img.id == image_id), 1)
        total_images = len(all_images)
        
        # البحث عن الصور السابقة واللاحقة
        previous_image = None
        next_image = None
        for i, img in enumerate(all_images):
            if img.id == image_id:
                if i > 0:
                    previous_image = all_images[i-1]
                if i < len(all_images) - 1:
                    next_image = all_images[i+1]
                break
        
        return render_template('safety_check_image_view.html', 
                             safety_check=safety_check, 
                             image=image,
                             image_number=image_number,
                             total_images=total_images,
                             previous_image=previous_image,
                             next_image=next_image)
    except Exception as e:
        current_app.logger.error(f"خطأ في عرض صورة فحص السلامة: {str(e)}")
        flash('حدث خطأ في عرض الصورة', 'error')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/pdf/english')
def export_safety_check_pdf_english(check_id):
    """تصدير طلب فحص السلامة كملف PDF إنجليزي احترافي مع الشعار"""
    if not current_user.is_authenticated:
        flash('يرجى تسجيل الدخول أولاً', 'error')
        return redirect(url_for('external_safety.admin_external_safety_checks'))
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
        from io import BytesIO
        import os
        
        safety_check = VehicleExternalSafetyCheck.query.options(
            db.selectinload(VehicleExternalSafetyCheck.safety_images)
        ).get_or_404(check_id)
        
        # إنشاء buffer للـ PDF
        buffer = BytesIO()
        
        # إنشاء PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                              rightMargin=2*cm, leftMargin=2*cm,
                              topMargin=2.5*cm, bottomMargin=2*cm)
        
        # العناصر
        elements = []
        
        # الأنماط
        styles = getSampleStyleSheet()
        
        # إضافة الشعار
        logo_path = 'static/images/logo.png'
        if os.path.exists(logo_path):
            try:
                logo = RLImage(logo_path, width=4*cm, height=4*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.5*cm))
            except:
                pass
        
        # عنوان التقرير
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=22,
            textColor=colors.HexColor('#18B2B0'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        title = Paragraph("<b>NUZUM VEHICLE SAFETY INSPECTION REPORT</b>", title_style)
        elements.append(title)
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#666666'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica'
        )
        
        subtitle = Paragraph(f"External Safety Check Report", subtitle_style)
        elements.append(subtitle)
        elements.append(Spacer(1, 1*cm))
        
        # دالة لتنظيف النص من الأحرف العربية وغير القابلة للطباعة
        def clean_text_for_english_pdf(text):
            if not text:
                return 'N/A'
            # إزالة الأحرف العربية والاحتفاظ بالأحرف الإنجليزية والأرقام فقط
            import re
            # الاحتفاظ بالأحرف الإنجليزية والأرقام والمسافات والرموز الأساسية
            cleaned = re.sub(r'[^\x00-\x7F]+', ' ', str(text))
            cleaned = ' '.join(cleaned.split())  # إزالة المسافات الزائدة
            return cleaned if cleaned.strip() else 'N/A'
        
        # بيانات المركبة والسائق (تنظيف النصوص)
        info_data = [
            ['Vehicle Information', ''],
            ['Plate Number:', clean_text_for_english_pdf(safety_check.vehicle_plate_number)],
            ['Make & Model:', clean_text_for_english_pdf(safety_check.vehicle_make_model)],
            ['Driver Name:', clean_text_for_english_pdf(safety_check.driver_name)],
            ['Department:', clean_text_for_english_pdf(safety_check.driver_department)],
            ['City:', clean_text_for_english_pdf(safety_check.driver_city)],
            ['National ID:', clean_text_for_english_pdf(safety_check.driver_national_id)],
            ['', ''],
            ['Inspection Details', ''],
            ['Inspection Date:', safety_check.inspection_date.strftime('%Y-%m-%d') if safety_check.inspection_date else 'N/A'],
            ['Status:', 'Approved' if safety_check.approval_status == 'approved' else 'Pending'],
            ['Notes:', 'See safety inspection images below' if safety_check.notes else 'No notes']
        ]
        
        info_table = Table(info_data, colWidths=[7*cm, 10*cm])
        info_table.setStyle(TableStyle([
            # رؤوس الأقسام
            ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#18B2B0')),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (1, 0), 14),
            ('ALIGN', (0, 0), (1, 0), 'CENTER'),
            ('SPAN', (0, 0), (1, 0)),
            
            ('BACKGROUND', (0, 8), (1, 8), colors.HexColor('#18B2B0')),
            ('TEXTCOLOR', (0, 8), (1, 8), colors.whitesmoke),
            ('FONTNAME', (0, 8), (1, 8), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 8), (1, 8), 14),
            ('ALIGN', (0, 8), (1, 8), 'CENTER'),
            ('SPAN', (0, 8), (1, 8)),
            
            # البيانات
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            
            # الصفوف المتناوبة
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            
            # الحدود
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#18B2B0')),
            
            # المسافات
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 1*cm))
        
        # إضافة الصور إذا وجدت
        if safety_check.safety_images:
            images_title_style = ParagraphStyle(
                'ImagesTitle',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#18B2B0'),
                spaceAfter=12,
                alignment=TA_LEFT,
                fontName='Helvetica-Bold'
            )
            
            images_title = Paragraph("<b>Vehicle Safety Images</b>", images_title_style)
            elements.append(images_title)
            elements.append(Spacer(1, 0.5*cm))
            
            # عرض الصور في شبكة
            images_per_row = 2
            image_width = 7*cm
            image_height = 5*cm
            
            image_elements = []
            for idx, safety_image in enumerate(safety_check.safety_images):
                try:
                    # محاولة الحصول على الصورة من Object Storage
                    from src.utils.object_storage import get_image_url
                    image_url = get_image_url(safety_image.image_path)
                    
                    if image_url:
                        # تحميل الصورة
                        import requests
                        response = requests.get(image_url, timeout=10)
                        if response.status_code == 200:
                            from PIL import Image as PILImage
                            img_buffer = BytesIO(response.content)
                            # فتح الصورة بواسطة PIL للتأكد من صحتها
                            pil_img = PILImage.open(img_buffer)
                            # تحويل إلى RGB إذا كانت RGBA
                            if pil_img.mode == 'RGBA':
                                pil_img = pil_img.convert('RGB')
                            # حفظ في buffer جديد
                            final_buffer = BytesIO()
                            pil_img.save(final_buffer, format='JPEG', quality=85)
                            final_buffer.seek(0)
                            # إنشاء صورة ReportLab
                            img = RLImage(final_buffer, width=image_width, height=image_height)
                            image_elements.append(img)
                            current_app.logger.info(f"تم تحميل الصورة بنجاح: {safety_image.image_path}")
                except Exception as e:
                    current_app.logger.error(f"خطأ في تحميل الصورة {safety_image.image_path}: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            # ترتيب الصور في جدول
            if image_elements:
                image_rows = []
                for i in range(0, len(image_elements), images_per_row):
                    row = image_elements[i:i+images_per_row]
                    # إضافة خلايا فارغة إذا كان الصف غير مكتمل
                    while len(row) < images_per_row:
                        row.append('')
                    image_rows.append(row)
                
                image_table = Table(image_rows, colWidths=[8*cm, 8*cm])
                image_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                
                elements.append(image_table)
        
        # التذييل
        elements.append(Spacer(1, 2*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#999999'),
            alignment=TA_CENTER,
            fontName='Helvetica-Oblique'
        )
        
        footer_text = f"Generated by NUZUM Fleet Management System | https://nuzum.site<br/>Report ID: {safety_check.id} | Date: {safety_check.created_at.strftime('%Y-%m-%d %H:%M')}"
        footer = Paragraph(footer_text, footer_style)
        elements.append(footer)
        
        # بناء الـ PDF
        doc.build(elements)
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='export_pdf_english',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم تصدير تقرير فحص السلامة (إنجليزي) للسيارة {safety_check.vehicle_plate_number}'
        )
        
        # إعادة المؤشر
        buffer.seek(0)
        
        # تنسيق اسم الملف
        inspection_date = safety_check.inspection_date.strftime('%Y-%m-%d') if safety_check.inspection_date else 'no_date'
        plate_number = safety_check.vehicle_plate_number.replace(' ', '_')
        driver_name = safety_check.driver_name.replace(' ', '_') if safety_check.driver_name else 'unknown'
        
        # إرسال الملف
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'Safety_Check_{plate_number}_{driver_name}_{inspection_date}.pdf'
        )
        
    except Exception as e:
        current_app.logger.error(f"خطأ في تصدير تقرير فحص السلامة (إنجليزي): {str(e)}")
        import traceback
        traceback.print_exc()
        flash('حدث خطأ في تصدير التقرير', 'error')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))

@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/upload-file', methods=['POST'])
def upload_safety_check_file(check_id):
    """رفع ملف PDF أو إضافة رابط خارجي لفحص السلامة"""
    if not current_user.is_authenticated:
        flash('يرجى تسجيل الدخول أولاً', 'error')
        return redirect('/login')
    
    try:
        from werkzeug.utils import secure_filename
        import uuid
        
        safety_check = VehicleExternalSafetyCheck.query.get_or_404(check_id)
        
        # معالجة ملف PDF المرفوع
        pdf_file = request.files.get('pdf_file')
        if pdf_file and pdf_file.filename:
            # التحقق من نوع الملف
            if not pdf_file.filename.lower().endswith('.pdf'):
                flash('يجب أن يكون الملف من نوع PDF', 'error')
                return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))
            
            # إنشاء اسم فريد للملف
            original_filename = secure_filename(pdf_file.filename)
            unique_filename = f"safety_check_{check_id}_{uuid.uuid4().hex[:8]}_{original_filename}"
            
            # حفظ الملف
            upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'safety_check_pdfs')
            os.makedirs(upload_folder, exist_ok=True)
            
            file_path = os.path.join(upload_folder, unique_filename)
            pdf_file.save(file_path)
            
            # تحديث مسار الملف في قاعدة البيانات
            safety_check.pdf_file_path = f"safety_check_pdfs/{unique_filename}"
            current_app.logger.info(f"تم رفع ملف PDF: {unique_filename}")
        
        # معالجة الرابط الخارجي
        external_link = request.form.get('external_link', '').strip()
        if external_link:
            # التحقق من صحة الرابط
            if not external_link.startswith(('http://', 'https://')):
                flash('الرابط يجب أن يبدأ بـ http:// أو https://', 'error')
                return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))
            
            safety_check.drive_pdf_link = external_link
            current_app.logger.info(f"تم إضافة رابط خارجي: {external_link[:50]}...")
        
        # حفظ التغييرات
        db.session.commit()
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='upload_file',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم تحديث ملف/رابط فحص السلامة للسيارة {safety_check.vehicle_plate_number}'
        )
        
        flash('تم حفظ الملف/الرابط بنجاح', 'success')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في رفع الملف: {str(e)}")
        flash('حدث خطأ في حفظ الملف', 'error')
        return redirect(url_for('external_safety.admin_view_safety_check', check_id=check_id))

@external_safety_bp.route('/admin/bulk-delete-safety-checks', methods=['POST'])
def bulk_delete_safety_checks():
    """حذف عدة طلبات فحص سلامة جماعياً"""
    if not current_user.is_authenticated:
        flash('يرجى تسجيل الدخول أولاً', 'error')
        return redirect(url_for('external_safety.admin_external_safety_checks'))
    
    try:
        # الحصول على معرفات الطلبات المحددة
        check_ids = request.form.getlist('check_ids')
        
        if not check_ids:
            flash('لم يتم تحديد أي طلبات للحذف', 'warning')
            return redirect(url_for('external_safety.admin_external_safety_checks'))
        
        # تحويل المعرفات إلى أرقام صحيحة
        try:
            check_ids = [int(check_id) for check_id in check_ids]
        except ValueError:
            flash('معرفات الطلبات غير صحيحة', 'error')
            return redirect(url_for('external_safety.admin_external_safety_checks'))
        
        # جلب جميع الطلبات المحددة
        safety_checks = VehicleExternalSafetyCheck.query.filter(
            VehicleExternalSafetyCheck.id.in_(check_ids)
        ).all()
        
        if not safety_checks:
            flash('لم يتم العثور على الطلبات المحددة', 'warning')
            return redirect(url_for('external_safety.admin_external_safety_checks'))
        
        deleted_count = 0
        deleted_plates = []
        
        # حذف كل طلب مع صوره
        for safety_check in safety_checks:
            try:
                # 💾 الصور تبقى محفوظة - لا نحذف الملفات الفعلية
                images_count = len(safety_check.safety_images)
                current_app.logger.info(f"💾 الصور محفوظة للأمان ({images_count} صورة)")
                
                # تسجيل العملية قبل الحذف
                log_audit(
                    user_id=current_user.id,
                    action='bulk_delete',
                    entity_type='VehicleExternalSafetyCheck',
                    entity_id=safety_check.id,
                    details=f'تم حذف طلب فحص السلامة للسيارة {safety_check.vehicle_plate_number} - السائق: {safety_check.driver_name} (ضمن حذف جماعي لـ {len(check_ids)} طلب)'
                )
                
                # حذف السجل من قاعدة البيانات
                plate_number = safety_check.vehicle_plate_number
                deleted_plates.append(plate_number)
                db.session.delete(safety_check)
                deleted_count += 1
                
                current_app.logger.info(f"تم حذف طلب فحص السلامة رقم {safety_check.id} للسيارة {plate_number} مع {images_deleted} صورة")
                
            except Exception as e:
                current_app.logger.error(f"خطأ في حذف طلب فحص السلامة رقم {safety_check.id}: {str(e)}")
                continue
        
        # حفظ التغييرات
        db.session.commit()
        
        # تسجيل العملية الجماعية
        log_audit(
            user_id=current_user.id,
            action='bulk_delete_completed',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=0,  # للحذف الجماعي
            details=f'تم حذف {deleted_count} طلب فحص سلامة بنجاح من أصل {len(check_ids)} طلب محدد. السيارات: {", ".join(deleted_plates[:5])}{"..." if len(deleted_plates) > 5 else ""}'
        )
        
        if deleted_count > 0:
            flash(f'تم حذف {deleted_count} طلب فحص سلامة بنجاح مع جميع الصور المرفقة', 'success')
        else:
            flash('لم يتم حذف أي طلبات. قد تكون هناك مشكلة في البيانات المحددة', 'warning')
        
        return redirect(url_for('external_safety.admin_external_safety_checks'))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في الحذف الجماعي لطلبات فحص السلامة: {str(e)}")
        flash('حدث خطأ في عملية الحذف الجماعي. يرجى المحاولة مرة أخرى', 'error')
        return redirect(url_for('external_safety.admin_external_safety_checks'))

@external_safety_bp.route('/admin/create-check-from-images', methods=['POST'])
def admin_create_check_from_images():
    """إنشاء فحص سلامة جديد من خلال رفع صور السيارة مباشرة"""
    from flask_login import login_required, current_user
    from models import Vehicle, VehicleSafetyImage
    from datetime import datetime
    from src.utils.audit_logger import log_audit
    
    # التحقق من تسجيل الدخول
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'error': 'يرجى تسجيل الدخول'}), 401
    
    try:
        # الحصول على البيانات
        vehicle_id = request.form.get('vehicle_id')
        notes = request.form.get('notes', '')
        
        if not vehicle_id:
            return jsonify({'success': False, 'error': 'يرجى اختيار السيارة'}), 400
        
        # التحقق من وجود السيارة
        vehicle = Vehicle.query.get_or_404(vehicle_id)
        
        # التحقق من وجود صور
        images = request.files.getlist('images')
        if not images or len(images) == 0:
            return jsonify({'success': False, 'error': 'يرجى رفع صورة واحدة على الأقل'}), 400
        
        # إنشاء سجل فحص السلامة
        safety_check = VehicleExternalSafetyCheck()
        safety_check.vehicle_id = vehicle.id
        safety_check.vehicle_plate_number = vehicle.plate_number
        safety_check.vehicle_make_model = f"{vehicle.make} {vehicle.model}"
        safety_check.driver_name = current_user.username
        safety_check.driver_department = current_user.assigned_department.name if hasattr(current_user, 'assigned_department') and current_user.assigned_department else 'الإدارة'
        safety_check.driver_city = 'الرياض'
        safety_check.driver_national_id = ''
        safety_check.inspection_date = datetime.now()
        safety_check.notes = f"تم إنشاء الفحص من الصور المرفوعة. {notes}"
        safety_check.approval_status = 'approved'  # موافق عليه مباشرة
        safety_check.created_at = datetime.now()
        
        db.session.add(safety_check)
        db.session.flush()  # للحصول على ID
        
        # حفظ الصور
        saved_images_count = 0
        for image_file in images:
            if image_file and image_file.filename and allowed_file(image_file.filename):
                # توليد اسم ملف آمن
                original_filename = secure_filename(image_file.filename)
                file_ext = original_filename.rsplit('.', 1)[1].lower()
                unique_filename = f"safety_check_{safety_check.id}_{uuid.uuid4().hex}.{file_ext}"
                
                # حفظ مؤقتاً للضغط في مجلد آمن بدلاً من /tmp
                temp_dir = os.path.join(current_app.static_folder, '.temp')
                os.makedirs(temp_dir, exist_ok=True)
                temp_path = os.path.join(temp_dir, unique_filename)
                image_file.save(temp_path)
                
                # ضغط الصورة
                compress_image(temp_path)
                
                # قراءة الصورة المضغوطة
                with open(temp_path, 'rb') as f:
                    compressed_data = f.read()
                
                # رفع إلى Object Storage
                object_key = upload_image(compressed_data, 'safety_checks', unique_filename)
                
                # 💾 الملف المؤقت يبقى محفوظاً في static/.temp/ بشكل دائم
                
                # إنشاء سجل الصورة
                image_record = VehicleSafetyImage()
                image_record.safety_check_id = safety_check.id
                image_record.image_path = object_key
                image_record.uploaded_at = datetime.now()
                
                db.session.add(image_record)
                saved_images_count += 1
        
        db.session.commit()
        
        # تسجيل النشاط
        log_audit(
            user_id=current_user.id,
            action='create',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم إنشاء فحص سلامة للمركبة {vehicle.plate_number} مع {saved_images_count} صورة'
        )
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء الفحص بنجاح مع {saved_images_count} صورة',
            'check_id': safety_check.id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في إنشاء فحص السلامة من الصور: {str(e)}")
        return jsonify({'success': False, 'error': 'حدث خطأ في إنشاء الفحص'}), 500


@external_safety_bp.route('/admin/external-safety-check/<int:check_id>/upload-to-drive', methods=['POST'])
@login_required
def upload_safety_check_to_drive(check_id):
    """رفع فحص السلامة على Google Drive"""
    try:
        safety_check = VehicleExternalSafetyCheck.query.get_or_404(check_id)
        
        # التحقق من الصلاحيات
        if not current_user.is_authenticated:
            return jsonify({
                'success': False,
                'message': 'ليس لديك صلاحية لهذا الإجراء'
            }), 403
        
        # استخدام VehicleDriveUploader لرفع الفحص
        uploader = VehicleDriveUploader()
        uploader.upload_safety_check(safety_check)
        
        # حفظ التغييرات في قاعدة البيانات
        db.session.commit()
        
        current_app.logger.info(f"تم رفع فحص السلامة {check_id} على Google Drive بنجاح")
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='upload_to_drive',
            entity_type='VehicleExternalSafetyCheck',
            entity_id=safety_check.id,
            details=f'تم رفع فحص السلامة للسيارة {safety_check.vehicle_plate_number} على Google Drive'
        )
        
        folder_url = ''
        if hasattr(safety_check, 'drive_folder_id') and safety_check.drive_folder_id:
            folder_url = f"https://drive.google.com/drive/folders/{safety_check.drive_folder_id}"
        
        return jsonify({
            'success': True,
            'message': 'تم رفع فحص السلامة على Google Drive بنجاح',
            'folder_url': folder_url,
            'folder_id': safety_check.drive_folder_id if hasattr(safety_check, 'drive_folder_id') else None
        }), 200
    
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في رفع فحص السلامة على Google Drive: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'حدث خطأ في الرفع: {str(e)}'
        }), 500
