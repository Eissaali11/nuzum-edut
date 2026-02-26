"""
REFACTORED: Operations routes now organized in operations/ subpackage

All routes have been refactored into specialized modules:
- operations_core_routes.py: Dashboard, list, view, delete (30 lines)
- operations_workflow_routes.py: Approve, reject, review, notifications (180 lines)
- operations_export_routes.py: Export to Excel (230 lines)
- operations_sharing_routes.py: Share via email, Outlook, ZIP packages (440 lines)
- operations_accidents_routes.py: Accident reports management (150 lines)
- operations_helpers.py: Shared utilities (helper functions)
- __init__.py: Registrar aggregating all operations blueprints

Total refactored from 2398 lines → 5 focused modules + 1 helpers module

This file remains for backward compatibility only.
"""

from flask import Blueprint

# Create main operations blueprint
operations_bp = Blueprint('operations', __name__, url_prefix='/operations')

# Register all operations sub-routes
from src.routes.operations import register_operations_routes
register_operations_routes(operations_bp)

__all__ = ['operations_bp']




from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, send_file
from flask_login import login_required, current_user
from flask_wtf import FlaskForm
from werkzeug.utils import secure_filename
from sqlalchemy import extract, func, or_
from src.forms.vehicle_forms import VehicleAccidentForm, VehicleDocumentsForm
import os
import uuid
import io
import urllib.parse
import pandas as pd
from fpdf import FPDF
import base64
import uuid
import zipfile
import shutil

from src.core.extensions import db
from models import (
        Vehicle, VehicleRental, VehicleWorkshop, VehicleWorkshopImage, 
        VehicleProject, VehicleHandover, VehicleHandoverImage, SystemAudit,
        VehiclePeriodicInspection, VehicleSafetyCheck, VehicleAccident, Employee,
        Department, ExternalAuthorization, Module, Permission, UserRole,
        VehicleExternalSafetyCheck, OperationRequest
)
from src.utils.audit_logger import log_activity
from src.utils.audit_logger import log_audit
from src.utils.whatsapp_message_generator import generate_whatsapp_url
from src.utils.vehicles_export import export_vehicle_pdf, export_workshop_records_pdf, export_vehicle_excel, export_workshop_records_excel
from src.utils.simple_pdf_generator import create_vehicle_handover_pdf as generate_complete_vehicle_report
from src.utils.vehicle_excel_report import generate_complete_vehicle_excel_report
from src.services.email_service import EmailService
from src.utils.unified_storage_service import unified_storage
# from src.utils.workshop_report import generate_workshop_report_pdf
# from src.utils.html_to_pdf import generate_pdf_from_template
# from src.utils.fpdf_arabic_report import generate_workshop_report_pdf_fpdf
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
import arabic_reshaper
from bidi.algorithm import get_display
# from src.utils.fpdf_handover_pdf import generate_handover_report_pdf
# ============ تأكد من وجود هذه الاستيرادات في أعلى الملف ============
from datetime import date
# =================================================================


operations_bp = Blueprint('operations', __name__, url_prefix='/operations')



def update_vehicle_state(vehicle_id):
    """
    الدالة المركزية الذكية لتحديد وتحديث الحالة النهائية للمركبة وسائقها.
    (نسخة معدلة لا تعتمد على حقل is_approved).
    تعتمد على حالة OperationRequest المرتبط لتحديد السجلات الرسمية.
    """
    try:
        vehicle = Vehicle.query.get(vehicle_id)
        if not vehicle:
            current_app.logger.warning(f"محاولة تحديث حالة لمركبة غير موجودة: ID={vehicle_id}")
            return

        # 1. فحص الحالات ذات الأولوية القصوى (تبقى كما هي)
        if vehicle.status == 'out_of_service':
            return

        active_accident = VehicleAccident.query.filter(VehicleAccident.vehicle_id == vehicle_id, VehicleAccident.accident_status != 'مغلق').first()
        in_workshop = VehicleWorkshop.query.filter(VehicleWorkshop.vehicle_id == vehicle_id, VehicleWorkshop.exit_date.is_(None)).first()

        # نحدد ما إذا كانت السيارة في حالة حرجة
        is_critical_state = bool(active_accident or in_workshop)

        if active_accident:
            vehicle.status = 'accident'
        elif in_workshop:
            vehicle.status = 'in_workshop'

        # 2. التحقق من الإيجار النشط
        active_rental = VehicleRental.query.filter_by(vehicle_id=vehicle_id, is_active=True).first()

        # ================== بداية المنطق الجديد لتحديد السجلات الرسمية ==================

        # 3. إنشاء استعلام فرعي لجلب ID لكل سجل handover له طلب موافقة.
        approved_handover_ids_subquery = db.session.query(
            OperationRequest.related_record_id
        ).filter(
            OperationRequest.operation_type == 'handover',
            OperationRequest.status == 'approved',
            OperationRequest.vehicle_id == vehicle_id
        ).subquery()

        # 4. إنشاء استعلام فرعي لجلب ID لكل سجل handover له طلب (بغض النظر عن حالته).
        all_handover_request_ids_subquery = db.session.query(
            OperationRequest.related_record_id
        ).filter(
            OperationRequest.operation_type == 'handover',
            OperationRequest.vehicle_id == vehicle_id
        ).subquery()

        # 5. بناء الاستعلام الأساسي الذي يختار السجلات "الرسمية" فقط.
        # السجل يعتبر رسمياً إذا تمت الموافقة عليه، أو إذا كان قديماً (ليس له طلب موافقة أصلاً).
        base_official_query = VehicleHandover.query.filter(
            VehicleHandover.vehicle_id == vehicle_id
        ).filter(
            or_(
                VehicleHandover.id.in_(approved_handover_ids_subquery),
                ~VehicleHandover.id.in_(all_handover_request_ids_subquery)
            )
        )

        # 6. الآن نستخدم هذا الاستعلام الرسمي للحصول على آخر عملية تسليم واستلام
        latest_delivery = base_official_query.filter(
            VehicleHandover.handover_type.in_(['delivery', 'تسليم'])
        ).order_by(VehicleHandover.created_at.desc()).first()

        latest_return = base_official_query.filter(
            VehicleHandover.handover_type.in_(['return', 'استلام', 'receive'])
        ).order_by(VehicleHandover.created_at.desc()).first()

        # =================== نهاية المنطق الجديد لتحديد السجلات الرسمية ===================

        # 7. تطبيق التحديثات على السائق والحالة بناءً على السجلات الرسمية فقط
        is_currently_handed_out = False
        if latest_delivery:
            if not latest_return or latest_delivery.created_at > latest_return.created_at:
                is_currently_handed_out = True

        if is_currently_handed_out:
            # السيناريو (أ): السيارة مسلّمة حالياً (بناءً على سجل معتمد)
            vehicle.driver_name = latest_delivery.person_name
            # تحديث الحالة فقط إذا لم تكن السيارة في حالة حرجة (ورشة/حادث)
            if not is_critical_state:
                vehicle.status = 'rented' if active_rental else 'in_project'
        else:
            # السيناريو (ب): السيارة متاحة (بناءً على سجل معتمد)
            vehicle.driver_name = None
            # تحديث الحالة فقط إذا لم تكن السيارة في حالة حرجة
            if not is_critical_state:
                vehicle.status = 'rented' if active_rental else 'available'

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في دالة update_vehicle_state لـ vehicle_id {vehicle_id}: {str(e)}")



# def update_vehicle_state(vehicle_id):
#     """
#     الدالة المركزية الذكية لتحديد وتحديث الحالة النهائية للمركبة وسائقها
#     بناءً على هرم أولويات الحالات (ورشة > إيجار > تسليم > متاحة).
#     """
#     try:
#         vehicle = Vehicle.query.get(vehicle_id)
#         if not vehicle:
#             return

#         # -- هرم أولوية الحالات (من الأعلى إلى الأدنى) --

#         # 1. حالة "خارج الخدمة": لها أعلى أولوية ولا تتغير تلقائياً
#         if vehicle.status == 'out_of_service':
#             # لا تفعل شيئاً، هذه الحالة لا تتغير إلا يدوياً
#             return

#         # 2. حالة "الحادث"
#         # يجب تعديل منطق الحادث بحيث تبقى الحالة accident حتى يتم إغلاق السجل
#         active_accident = VehicleAccident.query.filter(
#             VehicleAccident.vehicle_id == vehicle_id,
#             VehicleAccident.accident_status != 'مغلق' # نفترض أن 'مغلق' هي الحالة النهائية
#         ).first()
#         if active_accident:
#             vehicle.status = 'accident'
#             # (منطق السائق يبقى كما هو أدناه لأنه قد يكون هناك سائق وقت الحادث)

#         # 3. حالة "الورشة"
#         in_workshop = VehicleWorkshop.query.filter(
#             VehicleWorkshop.vehicle_id == vehicle_id,
#             VehicleWorkshop.exit_date.is_(None) # لا يزال في الورشة
#         ).first()
#         if in_workshop:
#             vehicle.status = 'in_workshop'
#             db.session.commit() # نحفظ الحالة وننهي لأنها ذات أولوية
#             return # إنهاء الدالة، لأن الورشة لها الأسبقية على الإيجار والتسليم

#         # --- إذا لم تكن السيارة في ورشة، ننتقل للحالات التشغيلية ---

#         # 4. حالة "مؤجرة"
#         active_rental = VehicleRental.query.filter(
#             VehicleRental.vehicle_id == vehicle_id,
#             VehicleRental.is_active == True
#         ).first()
#         if active_rental:
#             vehicle.status = 'rented'
#             # لا ننهي هنا، سنكمل لتحديد السائق

#         # 5. حالة "التسليم" و "متاحة" (نفس منطق الدالة السابقة)
#         latest_delivery = VehicleHandover.query.filter(
#             VehicleHandover.vehicle_id == vehicle_id,
#             VehicleHandover.handover_type.in_(['delivery', 'تسليم'])
#         ).order_by(VehicleHandover.handover_date.desc(), VehicleHandover.id.desc()).first()

#         latest_return = VehicleHandover.query.filter(
#             VehicleHandover.vehicle_id == vehicle_id,
#             VehicleHandover.handover_type.in_(['return', 'استلام', 'receive'])
#         ).order_by(VehicleHandover.handover_date.desc(), VehicleHandover.id.desc()).first()

#         is_currently_handed_out = False
#         if latest_delivery:
#             if not latest_return or latest_delivery.created_at > latest_return.created_at:
#                  is_currently_handed_out = True

#         if is_currently_handed_out:
#             # مسلمة لسائق
#             vehicle.driver_name = latest_delivery.person_name
#             # إذا لم تكن مؤجرة، فستكون في مشروع
#             if not active_rental: 
#                 vehicle.status = 'in_project'
#         else:
#             # تم استلامها أو لم تسلم أبداً
#             vehicle.driver_name = None
#             # إذا لم تكن مؤجرة، فستكون متاحة
#             if not active_rental:
#                 vehicle.status = 'available'

#         db.session.commit()

#     except Exception as e:
#         db.session.rollback()
#         current_app.logger.error(f"خطأ في تحديث حالة المركبة {vehicle_id}: {e}")



# في ملف operations_bp.py

@operations_bp.route('/')
@login_required
def operations_dashboard():
    """لوحة إدارة العمليات الرئيسية مع إمكانية البحث"""

    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'danger')
        return redirect(url_for('dashboard.index'))

    # === بداية التعديلات ===

    # 1. استلام قيمة البحث من رابط URL
    search_plate = request.args.get('search_plate', '').strip()

    # 2. بناء استعلام العمليات المعلقة
    pending_query = OperationRequest.query.filter_by(status='pending')

    # 3. تطبيق فلتر البحث (إذا تم إدخال قيمة)
    if search_plate:
        # للانضمام مع جدول المركبات والبحث في حقل رقم اللوحة
        pending_query = pending_query.join(Vehicle).filter(Vehicle.plate_number.ilike(f"%{search_plate}%"))

    # 4. تنفيذ الاستعلام النهائي
    pending_requests = pending_query.order_by(
        OperationRequest.priority.desc(),
        OperationRequest.requested_at.desc()
    ).limit(10).all()

    # === نهاية التعديلات ===

    # إحصائيات العمليات (تبقى كما هي)
    stats = {
        'pending': OperationRequest.query.filter_by(status='pending').count(),
        'under_review': OperationRequest.query.filter_by(status='under_review').count(),
        'approved': OperationRequest.query.filter_by(status='approved').count(),
        'rejected': OperationRequest.query.filter_by(status='rejected').count(),
        'unread_notifications': OperationNotification.query.filter_by(user_id=current_user.id, is_read=False).count(),
        'pending_accidents': VehicleAccident.query.filter_by(review_status='pending').count()
    }

    return render_template('operations/dashboard.html', 
                         stats=stats, 
                         pending_requests=pending_requests)
@operations_bp.route('/list')
@login_required
def operations_list():
    """قائمة جميع العمليات مع فلترة"""
    
    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'danger')
        return redirect(url_for('dashboard.index'))
    
    # فلترة العمليات
    status_filter = request.args.get('status', 'all')
    operation_type_filter = request.args.get('operation_type', 'all')
    priority_filter = request.args.get('priority', 'all')
    vehicle_search = request.args.get('vehicle_search', '').strip()
    
    query = OperationRequest.query
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if operation_type_filter != 'all':
        query = query.filter_by(operation_type=operation_type_filter)
    
    if priority_filter != 'all':
        query = query.filter_by(priority=priority_filter)
    
    # البحث حسب السيارة
    if vehicle_search:
        # البحث في عنوان العملية أو الوصف حيث يتم ذكر رقم السيارة عادة
        query = query.filter(
            OperationRequest.title.contains(vehicle_search) |
            OperationRequest.description.contains(vehicle_search)
        )
    
    # ترتيب العمليات
    operations = query.order_by(
        OperationRequest.priority.desc(),
        OperationRequest.requested_at.desc()
    ).all()
    
    return render_template('operations/list.html', 
                         operations=operations,
                         status_filter=status_filter,
                         operation_type_filter=operation_type_filter,
                         priority_filter=priority_filter,
                         vehicle_search=vehicle_search)

@operations_bp.route('/<int:operation_id>')
@login_required
def view_operation(operation_id):
    """عرض تفاصيل العملية"""
    
    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'danger')
        return redirect(url_for('dashboard.index'))
    
    operation = OperationRequest.query.get_or_404(operation_id)
    related_record = operation.get_related_record()
    
    # توجيه عمليات الورشة إلى صفحة منفصلة
    if operation.operation_type == 'workshop_record':
        # جلب معلومات السائق الحالي من آخر تسليم
        driver_employee = None
        current_driver_info = None
        
        # البحث عن آخر تسليم للمركبة للحصول على السائق الحالي
        last_handover = VehicleHandover.query.filter_by(
            vehicle_id=operation.vehicle_id,
            handover_type='delivery'
        ).order_by(VehicleHandover.handover_date.desc()).first()
        
        if last_handover:
            current_driver_info = {
                'name': last_handover.person_name,
                'phone': last_handover.driver_phone_number,
                'residency_number': last_handover.driver_residency_number
            }
            
            # محاولة العثور على بيانات الموظف في النظام
            if last_handover.driver_residency_number:
                driver_employee = Employee.query.filter_by(
                    national_id=last_handover.driver_residency_number
                ).first()
            
            if not driver_employee and last_handover.person_name:
                driver_employee = Employee.query.filter_by(
                    name=last_handover.person_name
                ).first()
        
        # جلب الصور من قاعدة البيانات مباشرة إذا كانت العلاقة لا تعمل
        from sqlalchemy import text
        workshop_images = []
        if related_record and hasattr(related_record, 'id'):
            try:
                # استعلام مباشر لجلب الصور
                result = db.session.execute(
                    text("SELECT id, image_type, image_path, notes, uploaded_at FROM vehicle_workshop_images WHERE workshop_record_id = :workshop_id ORDER BY uploaded_at DESC"),
                    {'workshop_id': related_record.id}
                )
                workshop_images = [
                    {
                        'id': row[0],
                        'image_type': row[1], 
                        'image_path': row[2],
                        'notes': row[3],
                        'uploaded_at': row[4]
                    } 
                    for row in result
                ]
                current_app.logger.debug(f"تم جلب {len(workshop_images)} صور لسجل الورشة {related_record.id}")
            except Exception as e:
                current_app.logger.error(f"خطأ في جلب صور الورشة: {str(e)}")
        
        return render_template('operations/view_workshop.html', 
                             operation=operation,
                             related_record=related_record,
                             workshop_images=workshop_images,
                             driver_employee=driver_employee,
                             current_driver_info=current_driver_info,
)
    
    # جلب بيانات الموظف إذا كانت متاحة
    employee = None
    if related_record:
        # محاولة العثور على الموظف من خلال رقم الهوية أو الاسم
        if hasattr(related_record, 'driver_residency_number') and related_record.driver_residency_number:
            employee = Employee.query.filter_by(national_id=related_record.driver_residency_number).first()
        if not employee and hasattr(related_record, 'person_name') and related_record.person_name:
            employee = Employee.query.filter_by(name=related_record.person_name).first()
        if not employee and hasattr(related_record, 'driver_name') and related_record.driver_name:
            employee = Employee.query.filter_by(name=related_record.driver_name).first()
    
    # تحميل معلومات إضافية للطالب إذا كان موظفاً
    if operation.requester and hasattr(operation.requester, 'id'):
        try:
            # تحميل العلاقات بشكل صريح
            requester_employee = Employee.query.options(db.joinedload(Employee.departments)).get(operation.requester.id)
            if requester_employee:
                operation.requester = requester_employee
        except Exception as e:
            print(f"خطأ في تحميل بيانات الموظف الطالب: {e}")
    
    # البحث عن جهاز الجوال المخصص للمستخدم
    from models import MobileDevice
    mobile_device = MobileDevice.query.filter_by(employee_id=operation.requester.id).first()
    
    return render_template('operations/view.html', 
                         operation=operation,
                         related_record=related_record,
                         employee=employee,
                         mobile_device=mobile_device,
)



@operations_bp.route('/<int:operation_id>/approve', methods=['POST'])
@login_required
def approve_operation(operation_id):
    """الموافقة على العملية مع تفعيل السجل المرتبط وتحديث حالة السيارة."""

    if current_user.role != UserRole.ADMIN:
        return jsonify({'success': False, 'message': 'غير مسموح لك بالقيام بهذا الإجراء'})

    operation = OperationRequest.query.get_or_404(operation_id)
    review_notes = request.form.get('review_notes', '').strip()

    try:
        # 1. تحديث حالة الطلب نفسه
        operation.status = 'approved'
        operation.reviewed_by = current_user.id
        operation.reviewed_at = datetime.utcnow()
        operation.review_notes = review_notes

        # 2. البحث عن السجل المرتبط وتفعيله إذا كان من نوع handover
        if operation.operation_type == 'handover':
            handover_record = VehicleHandover.query.get(operation.related_record_id)
            if handover_record:
                # === الخطوة الأهم: تحويل المسودة إلى سجل رسمي ===
                # handover_record.is_approved = True
                db.session.commit() # نحفظ التفعيل أولاً

                # === بعد التفعيل، نقوم بتحديث حالة السيارة الآن ===
                if operation.vehicle_id:
                    update_vehicle_state(operation.vehicle_id)
                    log_audit(
                        'update', 'vehicle_state', operation.vehicle_id,
                        f'تم تحديث حالة السيارة والسائق بعد الموافقة على الطلب #{operation.id}'
                    )
            else:
                current_app.logger.warning(f"لم يتم العثور على سجل Handover رقم {operation.related_record_id} للموافقة عليه.")

        # (يمكن إضافة منطق لأنواع أخرى من العمليات هنا مستقبلاً)

        # 3. إنشاء إشعار للطالب
        create_notification(
            operation_id=operation.id,
            user_id=operation.requested_by,
            notification_type='status_change',
            title=f'✅ تمت الموافقة على طلبك: {operation.title}',
            message=f'تمت الموافقة على طلبك من قبل {current_user.username}.'
        )

        db.session.commit()

        log_activity(
            action='approve',
            entity_type='operation_request',
            entity_id=operation.id,
            details=f'تمت الموافقة على العملية: {operation.title}'
        )
        flash('تمت الموافقة على العملية وتحديث حالة المركبة بنجاح', 'success')
        return jsonify({'success': True, 'message': 'تمت الموافقة بنجاح'})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في الموافقة على العملية #{operation_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})



# @operations_bp.route('/<int:operation_id>/approve', methods=['POST'])
# @login_required
# def approve_operation(operation_id):
#     """الموافقة على العملية"""
    
#     if current_user.role != UserRole.ADMIN:
#         return jsonify({'success': False, 'message': 'غير مسموح'})
    
#     operation = OperationRequest.query.get_or_404(operation_id)
#     review_notes = request.form.get('review_notes', '').strip()
    
#     try:
#         # تحديث حالة العملية
#         operation.status = 'approved'
#         operation.reviewed_by = current_user.id
#         operation.reviewed_at = datetime.utcnow()
#         operation.review_notes = review_notes
        
#         # إنشاء إشعار للطالب
#         create_notification(
#             operation_id=operation.id,
#             user_id=operation.requested_by,
#             notification_type='status_change',
#             title=f'تمت الموافقة على العملية: {operation.title}',
#             message=f'تمت الموافقة على العملية من قبل {current_user.username}.\n{review_notes if review_notes else ""}'
#         )
        
#         db.session.commit()
        
#         # تحديث السائق الحالي وحالة السيارة إذا كانت العملية من نوع handover
#         if operation.operation_type == 'handover' and operation.vehicle_id:
#             try:
#                 from src.utils.vehicle_driver_utils import update_vehicle_driver_approved
#                 update_vehicle_driver_approved(operation.vehicle_id)
                
#                 # تحديث حالة السيارة بناءً على نوع التسليم/الاستلام
#                 handover_record = VehicleHandover.query.get(operation.related_record_id)
#                 if handover_record:
#                     vehicle = Vehicle.query.get(operation.vehicle_id)
#                     if vehicle and handover_record.handover_type:
#                         old_status = vehicle.status
                        
#                         # تحديد الحالة الجديدة بناءً على نوع العملية
#                         if handover_record.handover_type == 'return':  # استلام السيارة
#                             new_status = 'متاحة'
#                         elif handover_record.handover_type == 'delivery':  # تسليم السيارة
#                             new_status = 'في المشروع'
#                         else:
#                             new_status = None
                        
#                         if new_status and old_status != new_status:
#                             vehicle.status = new_status
#                             db.session.add(vehicle)
                            
#                             # تسجيل تغيير الحالة
#                             action_type = 'الاستلام' if handover_record.handover_type == 'return' else 'التسليم'
#                             log_audit(current_user.id, 'update', 'vehicle_status', vehicle.id,
#                                      f'تم تحديث حالة السيارة {vehicle.plate_number} إلى "{new_status}" بعد الموافقة على عملية {action_type}')
                            
#                             print(f"تم تحديث حالة السيارة {vehicle.plate_number} من '{old_status}' إلى '{new_status}' بعد الموافقة على عملية {action_type}")
                        
#             except Exception as e:
#                 print(f"خطأ في تحديث السائق وحالة السيارة بعد الموافقة: {e}")
#                 import traceback
#                 traceback.print_exc()
        
#         # تسجيل العملية
#         log_audit(current_user.id, 'approve', 'operation_request', operation.id, 
#                  f'تمت الموافقة على العملية: {operation.title}')
        
#         flash('تمت الموافقة على العملية بنجاح', 'success')
#         return jsonify({'success': True, 'message': 'تمت الموافقة بنجاح'})
        
#     except Exception as e:
#         db.session.rollback()
#         return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})


# @operations_bp.route('/<int:operation_id>/reject', methods=['POST'])
# @login_required
# def reject_operation(operation_id):
#     """رفض العملية مع حذف السجل المؤقت المرتبط بها."""

#     if current_user.role != UserRole.ADMIN:
#         return jsonify({'success': False, 'message': 'غير مسموح لك بالقيام بهذا الإجراء'})

#     operation = OperationRequest.query.get_or_404(operation_id)
#     review_notes = request.form.get('review_notes', '').strip()

#     if not review_notes:
#         return jsonify({'success': False, 'message': 'يجب إدخال سبب الرفض'})

#     try:
#         # 1. تحديث حالة الطلب
#         operation.status = 'rejected'
#         operation.reviewed_by = current_user.id
#         operation.reviewed_at = datetime.utcnow()
#         operation.review_notes = review_notes

#         # 2. البحث عن السجل المرتبط وحذفه (إذا كان handover)
#         record_to_delete = None
#         if operation.operation_type == 'handover':
#             record_to_delete = VehicleHandover.query.get(operation.related_record_id)

#         if record_to_delete:
#             db.session.delete(record_to_delete)
#             log_audit(
#                 'delete_on_rejection', 'VehicleHandover', record_to_delete.id,
#                 f"تم حذف سجل Handover رقم {record_to_delete.id} تلقائياً بسبب رفض الطلب #{operation.id}"
#             )
#         else:
#             current_app.logger.warning(f"لم يتم العثور على سجل Handover رقم {operation.related_record_id} لحذفه بعد الرفض.")

#         # 3. إنشاء إشعار
#         create_notification(
#             operation_id=operation.id, user_id=operation.requested_by, notification_type='status_change',
#             title=f'❌ تم رفض طلبك: {operation.title}',
#             message=f'تم رفض طلبك من قبل {current_user.username}. السبب: {review_notes}'
#         )

#         db.session.commit()

#         log_audit('reject', 'operation_request', operation.id, f'تم رفض العملية: {operation.title}')
#         flash('تم رفض العملية وحذف السجل المؤقت.', 'warning')
#         return jsonify({'success': True, 'message': 'تم رفض العملية'})

#     except Exception as e:
#         db.session.rollback()
#         current_app.logger.error(f"خطأ في رفض العملية #{operation_id}: {str(e)}")
#         return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})


@operations_bp.route('/<int:operation_id>/reject', methods=['POST'])
@login_required
def reject_operation(operation_id):
    """رفض العملية"""
    
    if current_user.role != UserRole.ADMIN:
        return jsonify({'success': False, 'message': 'غير مسموح'})
    
    operation = OperationRequest.query.get_or_404(operation_id)
    review_notes = request.form.get('review_notes', '').strip()
    
    if not review_notes:
        return jsonify({'success': False, 'message': 'يجب إدخال سبب الرفض'})
    
    try:
        # تحديث حالة العملية
        operation.status = 'rejected'
        operation.reviewed_by = current_user.id
        operation.reviewed_at = datetime.utcnow()
        operation.review_notes = review_notes
        
        # إنشاء إشعار للطالب
        create_notification(
            operation_id=operation.id,
            user_id=operation.requested_by,
            notification_type='status_change',
            title=f'تم رفض العملية: {operation.title}',
            message=f'تم رفض العملية من قبل {current_user.username}.\nالسبب: {review_notes}'
        )


        record_to_delete = None
        if operation.operation_type == 'handover':
            record_to_delete = VehicleHandover.query.get(operation.related_record_id)

        if record_to_delete:
            db.session.delete(record_to_delete)
            log_audit(
                'delete_on_rejection', 'VehicleHandover', record_to_delete.id,
                f"تم حذف سجل Handover رقم {record_to_delete.id} تلقائياً بسبب رفض الطلب #{operation.id}"
            )
        else:
            current_app.logger.warning(f"لم يتم العثور على سجل Handover رقم {operation.related_record_id} لحذفه بعد الرفض.")

        
        db.session.commit()
        
        # تسجيل العملية
        log_audit(current_user.id, 'reject', 'operation_request', operation.id, 
                 f'تم رفض العملية: {operation.title} - السبب: {review_notes}')
        
        flash('تم رفض العملية', 'warning')
        return jsonify({'success': True, 'message': 'تم رفض العملية'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})







@operations_bp.route('/<int:operation_id>/under-review', methods=['POST'])
@login_required  
def set_under_review(operation_id):
    """وضع العملية تحت المراجعة"""
    
    if current_user.role != UserRole.ADMIN:
        return jsonify({'success': False, 'message': 'غير مسموح'})
    
    operation = OperationRequest.query.get_or_404(operation_id)
    
    try:
        operation.status = 'under_review'
        operation.reviewed_by = current_user.id
        operation.reviewed_at = datetime.utcnow()
        
        # إنشاء إشعار للطالب
        create_notification(
            operation_id=operation.id,
            user_id=operation.requested_by,
            notification_type='status_change',
            title=f'العملية تحت المراجعة: {operation.title}',
            message=f'العملية قيد المراجعة من قبل {current_user.username}'
        )
        
        db.session.commit()
        
        log_audit(current_user.id, 'review', 'operation_request', operation.id, 
                 f'العملية تحت المراجعة: {operation.title}')
        
        return jsonify({'success': True, 'message': 'تم وضع العملية تحت المراجعة'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})



@operations_bp.route('/<int:id>/delete', methods=['GET', 'POST'])
@login_required
def delete_operation(id):
    """حذف العملية"""
    
    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بحذف العمليات', 'error')
        return redirect(url_for('operations.operations_list'))
    
    operation = OperationRequest.query.get_or_404(id)
    
    if request.method == 'GET':
        # عرض صفحة تأكيد الحذف
        return render_template('operations/delete.html', operation=operation)
    
    # معالجة الحذف عند POST
    try:
        # تسجيل العملية قبل الحذف
        operation_title = operation.title
        operation_type = operation.operation_type
        
        # حذف الإشعارات المرتبطة بالعملية أولاً
        OperationNotification.query.filter_by(operation_request_id=id).delete()
        
        # حذف العملية
        db.session.delete(operation)
        db.session.commit()
        
        # تسجيل عملية الحذف
        log_audit(current_user.id, 'delete', 'operation_request', id, f'تم حذف العملية: {operation_title} من النوع {operation_type}')
        
        flash('تم حذف العملية بنجاح', 'success')
        return redirect(url_for('operations.operations_list'))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في حذف العملية #{id}: {str(e)}")
        flash(f'حدث خطأ أثناء حذف العملية: {str(e)}', 'error')
        return redirect(url_for('operations.view_operation', operation_id=id))

@operations_bp.route('/notifications')
@login_required
def notifications():
    """عرض الإشعارات"""
    
    user_notifications = OperationNotification.query.filter_by(
        user_id=current_user.id
    ).order_by(OperationNotification.created_at.desc()).all()
    
    # تحديد الإشعارات كمقروءة
    OperationNotification.query.filter_by(
        user_id=current_user.id, 
        is_read=False
    ).update({'is_read': True, 'read_at': datetime.utcnow()})
    
    db.session.commit()
    
    return render_template('operations/notifications.html', 
                         notifications=user_notifications)

def create_operation_request(operation_type, related_record_id, vehicle_id, 
                           title, description, requested_by, priority='normal'):
    """إنشاء طلب عملية جديد"""
    
    operation = OperationRequest()
    operation.operation_type = operation_type
    operation.related_record_id = related_record_id
    operation.vehicle_id = vehicle_id
    operation.title = title
    operation.description = description
    operation.requested_by = requested_by
    operation.requested_at = datetime.utcnow()
    operation.priority = priority
    operation.status = 'pending'
    
    try:
        db.session.add(operation)
        db.session.flush()  # للحصول على ID
        
        # إنشاء إشعارات للمديرين (النظام القديم)
        admins = User.query.filter_by(role=UserRole.ADMIN).all()
        for admin in admins:
            create_notification(
                operation_id=operation.id,
                user_id=admin.id,
                notification_type='new_operation',
                title=f'عملية جديدة تحتاج موافقة: {title}',
                message=f'عملية جديدة من نوع {get_operation_type_name(operation_type)} تحتاج للمراجعة والموافقة.'
            )
        
        # إنشاء إشعارات لجميع المستخدمين (نظام الإشعارات الجديد)
        try:
            from models import Notification
            all_users = User.query.all()
            
            vehicle_info = ""
            if vehicle_id:
                vehicle = Vehicle.query.get(vehicle_id)
                if vehicle:
                    vehicle_info = f" - المركبة {vehicle.plate_number}"
            
            for user in all_users:
                try:
                    new_notification = Notification(
                        user_id=user.id,
                        notification_type='operations',
                        title=f'عملية جديدة{vehicle_info}',
                        description=description,
                        related_entity_type='operation',
                        related_entity_id=operation.id,
                        priority='high' if priority == 'urgent' else 'normal',
                        action_url=url_for('operations.view_operation', operation_id=operation.id)
                    )
                    db.session.add(new_notification)
                except Exception as e:
                    current_app.logger.error(f'خطأ في إنشاء إشعار للمستخدم {user.id}: {str(e)}')
        except Exception as e:
            current_app.logger.error(f'خطأ في إنشاء إشعارات العمليات الجديدة: {str(e)}')
        
        # لا نحفظ هنا، الدالة المستدعية مسؤولة عن الحفظ
        return operation
    except Exception as e:
        print(f"خطأ في create_operation_request: {str(e)}")
        import traceback
        traceback.print_exc()
        raise e

def create_notification(operation_id, user_id, notification_type, title, message):
    """إنشاء إشعار جديد"""
    
    notification = OperationNotification()
    notification.operation_request_id = operation_id
    notification.user_id = user_id
    notification.notification_type = notification_type
    notification.title = title
    notification.message = message
    
    db.session.add(notification)
    return notification

def get_operation_type_name(operation_type):
    """الحصول على اسم نوع العملية بالعربية"""
    
    type_names = {
        'handover': 'تسليم/استلام مركبة',
        'workshop': 'عملية ورشة',
        'workshop_record': 'سجل ورشة',
        'external_authorization': 'تفويض خارجي',
        'safety_inspection': 'فحص سلامة'
    }
    
    return type_names.get(operation_type, operation_type)

# دالة مساعدة للحصول على عدد العمليات المعلقة للمدير
def get_pending_operations_count():
    """الحصول على عدد العمليات المعلقة"""
    return OperationRequest.query.filter_by(status='pending').count()

# دالة مساعدة للحصول على عدد الإشعارات غير المقروءة
def get_unread_notifications_count(user_id):
    """الحصول على عدد الإشعارات غير المقروءة للمستخدم"""
    return OperationNotification.query.filter_by(
        user_id=user_id, 
        is_read=False
    ).count()

@operations_bp.route('/test-notifications', methods=['GET', 'POST'])
def test_operations_notifications():
    """اختبار إنشاء إشعارات العمليات لجميع المستخدمين"""
    try:
        from models import Notification
        
        # الحصول على آخر عملية
        last_operation = OperationRequest.query.order_by(OperationRequest.id.desc()).first()
        
        if not last_operation:
            return jsonify({'success': False, 'message': 'لا توجد عمليات'}), 404
        
        all_users = User.query.all()
        
        vehicle_info = ""
        if last_operation.vehicle_id:
            vehicle = Vehicle.query.get(last_operation.vehicle_id)
            if vehicle:
                vehicle_info = f" - المركبة {vehicle.plate_number}"
        
        notification_count = 0
        for user in all_users:
            try:
                new_notification = Notification(
                    user_id=user.id,
                    notification_type='operations',
                    title=f'عملية جديدة{vehicle_info}',
                    description=last_operation.description or 'عملية تحتاج مراجعة',
                    related_entity_type='operation',
                    related_entity_id=last_operation.id,
                    priority='normal',
                    action_url=url_for('operations.view_operation', operation_id=last_operation.id)
                )
                db.session.add(new_notification)
                notification_count += 1
            except Exception as e:
                current_app.logger.error(f'خطأ في إنشاء إشعار للمستخدم {user.id}: {str(e)}')
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء {notification_count} إشعار للعملية {last_operation.id}',
            'operation_id': last_operation.id,
            'users_count': len(all_users)
        })
    except Exception as e:
        current_app.logger.error(f'خطأ في اختبار الإشعارات: {str(e)}')
        return jsonify({'success': False, 'message': str(e)}), 500

@operations_bp.route('/api/count')
@login_required
def api_operations_count():
    """API للحصول على إحصائيات العمليات"""
    
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'غير مسموح'})
    
    pending_count = OperationRequest.query.filter_by(status='pending').count()
    under_review_count = OperationRequest.query.filter_by(status='under_review').count()
    
    return jsonify({
        'pending': pending_count,
        'under_review': under_review_count,
        'unread_notifications': get_unread_notifications_count(current_user.id)
    })

@operations_bp.route('/<int:operation_id>/export-excel')
@login_required
def export_operation_excel(operation_id):
    """تصدير جميع بيانات العملية إلى Excel"""
    
    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'danger')
        return redirect(url_for('dashboard.index'))
    
    try:
        # جلب العملية والبيانات المرتبطة
        operation = OperationRequest.query.get_or_404(operation_id)
        related_record = operation.get_related_record()
        
        # إنشاء ملف Excel مع تنسيق محسن
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        
        # شيت 1: معلومات العملية الأساسية
        ws1 = wb.active
        ws1.title = 'معلومات العملية'
        
        # تنسيقات الألوان والخطوط
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2D5AA0', end_color='2D5AA0', fill_type='solid')
        data_font = Font(name='Arial', size=11)
        alignment = Alignment(horizontal='right', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # رؤوس الأعمدة
        operation_headers = ['رقم العملية', 'نوع العملية', 'العنوان', 'الوصف', 'الحالة', 
                           'الأولوية', 'تاريخ الطلب', 'تاريخ المراجعة', 'طالب العملية', 
                           'مراجع العملية', 'ملاحظات المراجعة']
        
        for col_num, header in enumerate(operation_headers, 1):
            cell = ws1.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border
            ws1.column_dimensions[cell.column_letter].width = 15
        
        # البيانات
        operation_values = [
            operation.id,
            get_operation_type_name(operation.operation_type),
            operation.title or '',
            operation.description or '',
            get_status_name(operation.status),
            get_priority_name(operation.priority),
            operation.requested_at.strftime('%Y-%m-%d %H:%M:%S') if operation.requested_at else '',
            operation.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if operation.reviewed_at else '',
            operation.requester.name if operation.requester else '',
            operation.reviewer.name if operation.reviewer else '',
            operation.review_notes or ''
        ]
        
        for col_num, value in enumerate(operation_values, 1):
            cell = ws1.cell(row=2, column=col_num, value=value)
            cell.font = data_font
            cell.alignment = alignment
            cell.border = border
            
        # شيت 2: بيانات المركبة
        if operation.vehicle:
            vehicle = operation.vehicle
            ws2 = wb.create_sheet('بيانات المركبة')
            
            vehicle_headers = ['رقم اللوحة', 'نوع المركبة', 'الماركة', 'الموديل', 'السنة', 
                              'اللون', 'الحالة', 'ملاحظات']
            
            for col_num, header in enumerate(vehicle_headers, 1):
                cell = ws2.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws2.column_dimensions[cell.column_letter].width = 18
            
            vehicle_values = [
                vehicle.plate_number or '',
                getattr(vehicle, 'type_of_car', '') or '',
                getattr(vehicle, 'make', '') or '',
                vehicle.model or '',
                str(vehicle.year) if vehicle.year else '',
                vehicle.color or '',
                vehicle.status or '',
                vehicle.notes or ''
            ]
            
            for col_num, value in enumerate(vehicle_values, 1):
                cell = ws2.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
            
        # شيت 3: بيانات السائق/الموظف
        employee = None
        current_driver_info = None
        
        # البحث عن السائق من العملية الحالية أولاً
        if operation.operation_type == 'handover' and related_record:
            current_driver_info = {
                'name': getattr(related_record, 'person_name', ''),
                'phone': getattr(related_record, 'driver_phone_number', ''),
                'residency_number': getattr(related_record, 'driver_residency_number', '')
            }
            
            # البحث عن الموظف في النظام بالهوية
            if related_record.driver_residency_number:
                employee = Employee.query.filter_by(
                    national_id=related_record.driver_residency_number
                ).first()
            
            # البحث بالاسم إذا لم نجد بالهوية
            if not employee and related_record.person_name:
                employee = Employee.query.filter_by(
                    name=related_record.person_name
                ).first()
        
        # إذا لم نجد من العملية الحالية، ابحث من آخر تسليم للمركبة
        if not employee and not current_driver_info and operation.vehicle_id:
            last_handover = VehicleHandover.query.filter_by(
                vehicle_id=operation.vehicle_id,
                handover_type='delivery'
            ).order_by(VehicleHandover.handover_date.desc()).first()
            
            if last_handover:
                current_driver_info = {
                    'name': getattr(last_handover, 'person_name', ''),
                    'phone': getattr(last_handover, 'driver_phone_number', ''),
                    'residency_number': getattr(last_handover, 'driver_residency_number', '')
                }
                
                # البحث عن الموظف في النظام
                if last_handover.driver_residency_number:
                    employee = Employee.query.filter_by(
                        national_id=last_handover.driver_residency_number
                    ).first()
                
                if not employee and last_handover.person_name:
                    employee = Employee.query.filter_by(
                        name=last_handover.person_name
                    ).first()
        
        if employee or current_driver_info:
            ws3 = wb.create_sheet('بيانات السائق')
            
            driver_headers = ['الاسم', 'الرقم الوظيفي', 'رقم الهوية', 'رقم الجوال', 'رقم جوال العمل',
                             'رقم IMEI', 'القسم', 'المنصب', 'تاريخ التوظيف', 'الحالة', 
                             'البريد الإلكتروني', 'العنوان', 'تاريخ الميلاد', 'الجنسية']
            
            for col_num, header in enumerate(driver_headers, 1):
                cell = ws3.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws3.column_dimensions[cell.column_letter].width = 16
            
            # ملء البيانات - أولوية للموظف المسجل في النظام
            if employee:
                # الحصول على بيانات الجوال للموظف
                from models import MobileDevice
                emp_mobile_device = MobileDevice.query.filter_by(employee_id=employee.id).first()
                
                driver_values = [
                    employee.name or '',
                    getattr(employee, 'employee_id', '') or '',
                    employee.national_id or '',
                    employee.mobile or '',
                    emp_mobile_device.phone_number if emp_mobile_device else '',
                    emp_mobile_device.imei if emp_mobile_device else '',
                    employee.departments[0].name if employee.departments else '',
                    employee.job_title or '',
                    employee.join_date.strftime('%Y-%m-%d') if employee.join_date else '',
                    employee.status or '',
                    employee.email or '',
                    getattr(employee, 'location', '') or '',
                    employee.birth_date.strftime('%Y-%m-%d') if employee.birth_date else '',
                    employee.nationality or ''
                ]
            elif current_driver_info:
                # البحث عن بيانات الجوال للسائق الحالي إذا كان موجود في النظام
                from models import MobileDevice
                driver_mobile = None
                if current_driver_info.get('residency_number'):
                    driver_employee = Employee.query.filter_by(
                        national_id=current_driver_info.get('residency_number')
                    ).first()
                    if driver_employee:
                        driver_mobile = MobileDevice.query.filter_by(employee_id=driver_employee.id).first()
                
                # استخدام بيانات السائق من نموذج التسليم مع القيم الافتراضية
                driver_values = [
                    current_driver_info.get('name', '') or operation.requester.username or '',
                    '', # الرقم الوظيفي - غير متوفر
                    current_driver_info.get('residency_number', '') or '',
                    current_driver_info.get('phone', '') or '',
                    driver_mobile.phone_number if driver_mobile else '', # رقم جوال العمل
                    driver_mobile.imei if driver_mobile else '', # رقم IMEI
                    '', # القسم - غير متوفر
                    '', # المنصب - غير متوفر
                    '', # تاريخ التوظيف - غير متوفر
                    '', # الحالة - غير متوفر
                    operation.requester.email or '', # البريد الإلكتروني
                    '', # العنوان - غير متوفر
                    '', # تاريخ الميلاد - غير متوفر
                    ''  # الجنسية - غير متوفر
                ]
            else:
                # البحث عن بيانات الجوال للمستخدم الطالب
                from models import MobileDevice
                requester_mobile = MobileDevice.query.filter_by(employee_id=operation.requester.id).first()
                
                # لا توجد بيانات سائق - استخدام بيانات المستخدم الطالب
                driver_values = [
                    operation.requester.username or '',
                    '', # الرقم الوظيفي - غير متوفر
                    '', # رقم الهوية
                    operation.requester.phone or '',
                    requester_mobile.phone_number if requester_mobile else '', # رقم جوال العمل
                    requester_mobile.imei if requester_mobile else '', # رقم IMEI
                    '', # القسم - غير متوفر
                    '', # المنصب - غير متوفر
                    '', # تاريخ التوظيف - غير متوفر
                    'نشط', # الحالة
                    operation.requester.email or '', # البريد الإلكتروني
                    '', # العنوان - غير متوفر
                    '', # تاريخ الميلاد - غير متوفر
                    ''  # الجنسية - غير متوفر
                ]
            
            # كتابة البيانات في الخلايا
            for col_num, value in enumerate(driver_values, 1):
                cell = ws3.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
            
            # إضافة قسم تفاصيل الجوال المُستلم
            ws3.cell(row=4, column=1, value='تفاصيل الجوال المُستلم').font = Font(bold=True, size=14, color='FF0000')
            
            mobile_details_headers = ['البيان', 'القيمة']
            for col_num, header in enumerate(mobile_details_headers, 1):
                cell = ws3.cell(row=5, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws3.column_dimensions[cell.column_letter].width = 20
            
            # الحصول على بيانات الجوال الحقيقية من قاعدة البيانات
            from models import MobileDevice
            mobile_device = MobileDevice.query.filter_by(employee_id=operation.requester.id).first()
            
            # بيانات تفاصيل الجوال المُستلم من البيانات الحقيقية
            mobile_details = [
                ['رقم الجوال', mobile_device.phone_number if mobile_device else (operation.requester.phone or 'غير متوفر')],
                ['رقم IMEI', mobile_device.imei if mobile_device else 'غير متوفر'],
                ['تاريخ الاستلام', mobile_device.assigned_date.strftime('%Y-%m-%d') if mobile_device and mobile_device.assigned_date else 'غير محدد'],
                ['حالة الجهاز', mobile_device.status if mobile_device else 'غير محدد']
            ]
            
            for row_num, (label, value) in enumerate(mobile_details, 6):
                cell1 = ws3.cell(row=row_num, column=1, value=label)
                cell1.font = Font(bold=True, size=10)
                cell1.alignment = alignment
                cell1.border = border
                cell1.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                
                cell2 = ws3.cell(row=row_num, column=2, value=value)
                cell2.font = data_font
                cell2.alignment = alignment
                cell2.border = border
        
        # شيت 4: تفاصيل نموذج التسليم/الاستلام
        if related_record and hasattr(related_record, 'handover_type'):
            ws4 = wb.create_sheet('نموذج التسليم-الاستلام')
            
            handover_headers = ['نوع العملية', 'تاريخ العملية', 'اسم الشخص', 'رقم الجوال', 
                               'رقم الهوية', 'قراءة العداد', 'مستوى الوقود', 'حالة المركبة',
                               'ملاحظات', 'الموقع', 'تاريخ الإنشاء', 'أنشئ بواسطة', 'مصدر الإنشاء', 'رابط النموذج']
            
            for col_num, header in enumerate(handover_headers, 1):
                cell = ws4.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws4.column_dimensions[cell.column_letter].width = 15
            
            # إنشاء رابط النموذج
            from flask import request
            base_url = request.host_url.rstrip('/')
            pdf_link = f"{base_url}/vehicles/handover/{related_record.id}/pdf/public"
            
            handover_values = [
                'تسليم' if related_record.handover_type == 'delivery' else 'استلام',
                related_record.handover_date.strftime('%Y-%m-%d %H:%M:%S') if related_record.handover_date else '',
                getattr(related_record, 'person_name', '') or '',
                getattr(related_record, 'driver_phone_number', '') or '',
                getattr(related_record, 'driver_residency_number', '') or '',
                str(getattr(related_record, 'mileage', '')) if getattr(related_record, 'mileage', None) else '',
                getattr(related_record, 'fuel_level', '') or '',
                getattr(related_record, 'vehicle_condition', '') or '',
                getattr(related_record, 'notes', '') or '',
                getattr(related_record, 'location', '') or '',
                related_record.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(related_record, 'created_at') and related_record.created_at else '',
                getattr(related_record, 'created_by_user', None).name if getattr(related_record, 'created_by_user', None) else '',
                'موبايل' if getattr(related_record, 'created_via_mobile', False) else 'ويب',
                pdf_link
            ]
            
            for col_num, value in enumerate(handover_values, 1):
                cell = ws4.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
            
        # شيت 5: سجلات الورشة (إذا كانت متوفرة)
        if operation.operation_type == 'workshop_record' and related_record:
            ws5 = wb.create_sheet('سجل الورشة')
            
            workshop_headers = ['نوع الخدمة', 'وصف المشكلة', 'الحل المطبق', 'التكلفة',
                               'تاريخ الدخول', 'تاريخ الخروج', 'الحالة', 'الفني المسؤول', 'ملاحظات']
            
            for col_num, header in enumerate(workshop_headers, 1):
                cell = ws5.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                ws5.column_dimensions[cell.column_letter].width = 16
            
            workshop_values = [
                getattr(related_record, 'service_type', '') or '',
                getattr(related_record, 'problem_description', '') or '',
                getattr(related_record, 'solution_applied', '') or '',
                str(getattr(related_record, 'cost', '')) if getattr(related_record, 'cost', None) else '',
                related_record.entry_date.strftime('%Y-%m-%d') if getattr(related_record, 'entry_date', None) else '',
                related_record.exit_date.strftime('%Y-%m-%d') if getattr(related_record, 'exit_date', None) else '',
                getattr(related_record, 'status', '') or '',
                getattr(related_record, 'technician_name', '') or '',
                getattr(related_record, 'notes', '') or ''
            ]
            
            for col_num, value in enumerate(workshop_values, 1):
                cell = ws5.cell(row=2, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = alignment
                cell.border = border
        
        # حفظ الملف
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # إنشاء اسم الملف
        filename = f"operation_details_{operation.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='export',
            entity_type='operation_request',
            details=f'تصدير تفاصيل العملية {operation_id} إلى Excel'
        )
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'خطأ في تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('operations.view_operation', operation_id=operation_id))

def get_status_name(status):
    """تحويل حالة العملية إلى النص العربي"""
    status_names = {
        'pending': 'معلقة',
        'under_review': 'تحت المراجعة',
        'approved': 'موافق عليها',
        'rejected': 'مرفوضة'
    }
    return status_names.get(status, status)

def get_priority_name(priority):
    """تحويل أولوية العملية إلى النص العربي"""
    priority_names = {
        'urgent': 'عاجل',
        'high': 'عالي',
        'normal': 'عادي',
        'low': 'منخفض'
    }
    return priority_names.get(priority, priority)


@operations_bp.route('/<int:operation_id>/send-email', methods=['POST'])
@login_required
def send_operation_email(operation_id):
    """إرسال ملفات العملية عبر الإيميل"""
    
    operation = OperationRequest.query.get_or_404(operation_id)
    
    try:
        data = request.get_json()
        to_email = data.get('email')
        to_name = data.get('name', '')
        include_excel = data.get('include_excel', True)
        include_pdf = data.get('include_pdf', True)
        
        if not to_email:
            return jsonify({'success': False, 'message': 'عنوان الإيميل مطلوب'})
        
        # التحقق من صحة الإيميل
        import re
        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_regex, to_email):
            return jsonify({'success': False, 'message': 'عنوان الإيميل غير صحيح'})
        
        # الحصول على معلومات المركبة والسائق
        vehicle = operation.vehicle
        if not vehicle:
            return jsonify({'success': False, 'message': 'لا توجد مركبة مرتبطة بهذه العملية'})
        
        vehicle_plate = vehicle.plate_number
        driver_name = getattr(vehicle, 'driver_name', None) or 'غير محدد'
        
        # مسارات الملفات المؤقتة
        excel_file_path = None
        pdf_file_path = None
        
        try:
            # إنشاء ملف Excel مؤقت
            if include_excel:
                excel_filename = f"operation_{operation_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                excel_file_path = os.path.join('/tmp', excel_filename)
                
                # توليد ملف Excel
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                wb = Workbook()
                ws = wb.active
                ws.title = 'تفاصيل العملية'
                
                # تنسيق الخلايا
                header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                data_font = Font(name='Arial', size=11)
                alignment = Alignment(horizontal='center', vertical='center')
                border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                
                # إضافة البيانات الأساسية
                headers = ['البيان', 'القيمة']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    cell.border = border
                    ws.column_dimensions[cell.column_letter].width = 25
                
                # البيانات
                data_rows = [
                    ('رقم العملية', f"#{operation.id}"),
                    ('عنوان العملية', operation.title),
                    ('نوع العملية', get_operation_type_name(operation.operation_type)),
                    ('حالة العملية', get_status_name(operation.status)),
                    ('الأولوية', get_priority_name(operation.priority)),
                    ('رقم لوحة المركبة', vehicle_plate),
                    ('السائق الحالي', driver_name),
                    ('تاريخ الطلب', operation.requested_at.strftime('%Y/%m/%d %H:%M') if operation.requested_at else operation.created_at.strftime('%Y/%m/%d %H:%M')),
                    ('طالب العملية', operation.requester.username if operation.requester else 'غير محدد'),
                    ('مراجع العملية', operation.reviewer.username if operation.reviewer else 'لم يتم المراجعة بعد'),
                    ('تاريخ المراجعة', operation.reviewed_at.strftime('%Y/%m/%d %H:%M') if operation.reviewed_at else 'لم يتم المراجعة بعد'),
                ]
                
                if operation.description:
                    data_rows.append(('الوصف', operation.description))
                if operation.review_notes:
                    data_rows.append(('ملاحظات المراجعة', operation.review_notes))
                
                for row_num, (label, value) in enumerate(data_rows, 2):
                    ws.cell(row=row_num, column=1, value=label).font = Font(name='Arial', size=11, bold=True)
                    ws.cell(row=row_num, column=1).alignment = alignment
                    ws.cell(row=row_num, column=1).border = border
                    
                    ws.cell(row=row_num, column=2, value=value).font = data_font
                    ws.cell(row=row_num, column=2).alignment = alignment
                    ws.cell(row=row_num, column=2).border = border
                
                wb.save(excel_file_path)
            
            # إنشاء ملف PDF مؤقت (إذا كان متوفراً)
            if include_pdf and operation.operation_type == 'handover' and operation.related_record_id:
                pdf_filename = f"operation_{operation_id}_report.pdf"
                pdf_file_path = os.path.join('/tmp', pdf_filename)
                
                # نسخ PDF الموجود إلى مجلد tmp
                try:
                    from src.utils.simple_pdf_generator import create_vehicle_handover_pdf
                    handover_record = VehicleHandover.query.get(operation.related_record_id)
                    if handover_record:
                        pdf_content = create_vehicle_handover_pdf(handover_record)
                        with open(pdf_file_path, 'wb') as f:
                            f.write(pdf_content.read())
                except Exception as pdf_error:
                    current_app.logger.warning(f"فشل في إنشاء PDF: {str(pdf_error)}")
                    pdf_file_path = None
            
            # إرسال الإيميل مع نظام احتياطي
            # إعداد بيانات العملية للقالب
            operation_data = {
                'title': operation.title,
                'operation_type': operation.operation_type,
                'status': operation.status,
                'priority': operation.priority,
                'requested_at': operation.requested_at.strftime('%Y/%m/%d %H:%M') if operation.requested_at else operation.created_at.strftime('%Y/%m/%d %H:%M'),
                'requester': operation.requester.username if operation.requester else 'غير محدد',
                'reviewer': operation.reviewer.username if operation.reviewer else 'لم يتم المراجعة بعد',
                'description': operation.description
            }
            
            # إعداد المرفقات
            attachments = []
            
            # إضافة ملف Excel
            if include_excel and excel_file_path and os.path.exists(excel_file_path):
                with open(excel_file_path, 'rb') as f:
                    excel_content = f.read()
                    attachments.append({
                        'content': excel_content,
                        'filename': f'operation_{operation.id}_details.xlsx',
                        'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    })
            
            # إضافة ملف PDF إن وجد
            if include_pdf and pdf_file_path and os.path.exists(pdf_file_path):
                with open(pdf_file_path, 'rb') as f:
                    pdf_content = f.read()
                    attachments.append({
                        'content': pdf_content,
                        'filename': f'operation_{operation.id}_document.pdf',
                        'content_type': 'application/pdf'
                    })
            
            # إرسال الإيميل باستخدام SendGrid
            try:
                from src.services.email_service import EmailService
                
                email_service = EmailService()
                
                # استخدام الدالة المحسّنة لعمليات التسليم/الاستلام
                if operation.operation_type == 'handover' and operation.related_record_id:
                    handover_record = VehicleHandover.query.get(operation.related_record_id)
                    if handover_record:
                        result = email_service.send_handover_operation_email(
                            to_email=to_email,
                            to_name=to_name or 'العميل',
                            handover_record=handover_record,
                            vehicle_plate=vehicle_plate,
                            driver_name=driver_name,
                            excel_file_path=excel_file_path if include_excel else None,
                            pdf_file_path=pdf_file_path if include_pdf else None
                        )
                    else:
                        # استخدام الدالة العادية إذا لم يتم العثور على سجل التسليم
                        result = email_service.send_vehicle_operation_files(
                            to_email=to_email,
                            to_name=to_name or 'العميل',
                            operation=operation,
                            vehicle_plate=vehicle_plate,
                            driver_name=driver_name,
                            excel_file_path=excel_file_path if include_excel else None,
                            pdf_file_path=pdf_file_path if include_pdf else None
                        )
                else:
                    # استخدام الدالة العادية للعمليات الأخرى
                    result = email_service.send_vehicle_operation_files(
                        to_email=to_email,
                        to_name=to_name or 'العميل',
                        operation=operation,
                        vehicle_plate=vehicle_plate,
                        driver_name=driver_name,
                        excel_file_path=excel_file_path if include_excel else None,
                        pdf_file_path=pdf_file_path if include_pdf else None
                    )
                
                if result.get('success'):
                    current_app.logger.info(f'تم إرسال الإيميل بنجاح عبر SendGrid إلى {to_email}')
                else:
                    # عرض رسالة تفصيلية عن المشكلة والحل
                    error_details = result.get('solution', 'يتطلب إعداد مُرسل مُتحقق في SendGrid')
                    current_app.logger.warning(f'فشل إرسال الإيميل عبر SendGrid: {result.get("message")}')
                    
                    raise Exception(f'SendGrid غير مُعد بشكل صحيح: {result.get("message", "خطأ غير معروف")}')
                    
            except Exception as sendgrid_error:
                current_app.logger.warning(f'فشل إرسال الإيميل عبر SendGrid: {sendgrid_error}')
                
                # استخدام النظام الاحتياطي
                try:
                    from src.services.fallback_email_service import FallbackEmailService, create_operation_email_template_simple
                    
                    fallback_service = FallbackEmailService()
                    html_content = create_operation_email_template_simple(
                        operation_data=operation_data,
                        vehicle_plate=vehicle_plate,
                        driver_name=driver_name
                    )
                    
                    result = fallback_service.send_email(
                        to_email=to_email,
                        subject=f'تفاصيل العملية: {operation.title} - مركبة {vehicle_plate}',
                        html_content=html_content,
                        attachments=attachments if attachments else None
                    )
                    
                    if result.get('success'):
                        current_app.logger.info(f'تم حفظ الإيميل محلياً - ID: {result.get("message_id")}')
                        result['message'] = f'تم حفظ الإيميل محلياً بنجاح. سيتم إرساله لاحقاً إلى {to_email}. يمكنك مراجعته في <a href="/email-queue/" target="_blank">قائمة الإيميلات المحفوظة</a>.'
                    else:
                        raise Exception(f'فشل النظام الاحتياطي أيضاً: {result.get("error")}')
                        
                except Exception as fallback_error:
                    current_app.logger.error(f'فشل النظام الاحتياطي: {fallback_error}')
                    result = {
                        'success': False,
                        'message': f'فشل في إرسال الإيميل: {str(fallback_error)}'
                    }
            
            # تسجيل العملية
            log_audit(
                user_id=current_user.id,
                action='send_email',
                entity_type='operation_request',
                entity_id=operation.id,
                details=f'إرسال ملفات العملية {operation_id} إلى {to_email}'
            )
            
            return jsonify(result)
            
        finally:
            # حذف الملفات المؤقتة
            if excel_file_path and os.path.exists(excel_file_path):
                try:
                    os.remove(excel_file_path)
                except:
                    pass
            
            if pdf_file_path and os.path.exists(pdf_file_path):
                try:
                    os.remove(pdf_file_path)
                except:
                    pass
                    
    except Exception as e:
        current_app.logger.error(f"خطأ في إرسال الإيميل للعملية {operation_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})


@operations_bp.route('/<int:operation_id>/share-outlook', methods=['GET'])
@login_required
def share_with_outlook(operation_id):
    """إنشاء ملف .eml لمشاركته مع Outlook"""
    
    operation = OperationRequest.query.get_or_404(operation_id)
    
    # التحقق من أن العملية من نوع handover
    if operation.operation_type != 'handover' or not operation.related_record_id:
        flash('هذه الميزة متاحة فقط لعمليات التسليم/الاستلام', 'warning')
        return redirect(url_for('operations.view_operation', operation_id=operation_id))
    
    # الحصول على سجل التسليم/الاستلام
    handover_record = VehicleHandover.query.get(operation.related_record_id)
    if not handover_record:
        flash('لم يتم العثور على سجل التسليم/الاستلام', 'danger')
        return redirect(url_for('operations.view_operation', operation_id=operation_id))
    
    # الحصول على معلومات المركبة
    vehicle = operation.vehicle
    if not vehicle:
        flash('لا توجد مركبة مرتبطة بهذه العملية', 'danger')
        return redirect(url_for('operations.view_operation', operation_id=operation_id))
    
    vehicle_plate = vehicle.plate_number
    driver_name = getattr(vehicle, 'driver_name', None) or 'غير محدد'
    
    # إنشاء الملفات المؤقتة (Excel و PDF)
    excel_file_path = None
    pdf_file_path = None
    
    try:
        # إنشاء ملف Excel
        excel_filename = f"handover_{operation_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_file_path = os.path.join('/tmp', excel_filename)
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'تفاصيل العملية'
        
        # تنسيق الخلايا
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        data_font = Font(name='Arial', size=11)
        alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        
        # إضافة البيانات
        headers = ['البيان', 'القيمة']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 25
        
        operation_type_text = "تسليم" if handover_record.handover_type == 'delivery' else "استلام"
        
        data_rows = [
            ('رقم العملية', f"#{operation.id}"),
            ('نوع العملية', operation_type_text),
            ('رقم لوحة المركبة', vehicle_plate),
            ('السائق', driver_name),
            ('اسم المستلم', handover_record.person_name),
            ('تاريخ العملية', handover_record.handover_date.strftime('%Y/%m/%d') if handover_record.handover_date else 'غير محدد'),
            ('المدينة', handover_record.city or 'غير محدد'),
        ]
        
        if handover_record.notes:
            data_rows.append(('ملاحظات', handover_record.notes))
        
        for row_num, (label, value) in enumerate(data_rows, 2):
            ws.cell(row=row_num, column=1, value=label).font = Font(name='Arial', size=11, bold=True)
            ws.cell(row=row_num, column=1).alignment = alignment
            ws.cell(row=row_num, column=1).border = border
            
            ws.cell(row=row_num, column=2, value=value).font = data_font
            ws.cell(row=row_num, column=2).alignment = alignment
            ws.cell(row=row_num, column=2).border = border
        
        wb.save(excel_file_path)
        
        # إنشاء ملف PDF
        try:
            from src.utils.simple_pdf_generator import create_vehicle_handover_pdf
            pdf_content = create_vehicle_handover_pdf(handover_record)
            pdf_filename = f"handover_{operation_id}_report.pdf"
            pdf_file_path = os.path.join('/tmp', pdf_filename)
            with open(pdf_file_path, 'wb') as f:
                f.write(pdf_content.read())
        except Exception as pdf_error:
            current_app.logger.warning(f"فشل في إنشاء PDF: {str(pdf_error)}")
            pdf_file_path = None
        
        # إنشاء ملف .eml مع معالجة الأخطاء
        try:
            from src.services.email_service import EmailService
            email_service = EmailService()
            
            eml_bytes, eml_filename = email_service.build_handover_eml(
                to_email="recipient@example.com",
                to_name="المستلم",
                handover_record=handover_record,
                vehicle_plate=vehicle_plate,
                driver_name=driver_name,
                excel_file_path=excel_file_path,
                pdf_file_path=pdf_file_path
            )
            
            if not eml_bytes:
                flash('فشل في إنشاء ملف البريد الإلكتروني. يرجى المحاولة مرة أخرى.', 'warning')
                return redirect(url_for('operations.view_operation', operation_id=operation_id))
                
        except Exception as eml_error:
            current_app.logger.error(f"خطأ في إنشاء ملف .eml: {str(eml_error)}")
            flash(f'فشل في إنشاء ملف البريد الإلكتروني: {str(eml_error)}', 'danger')
            return redirect(url_for('operations.view_operation', operation_id=operation_id))
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='share_outlook',
            entity_type='operation_request',
            entity_id=operation.id,
            details=f'إنشاء ملف .eml لمشاركة العملية {operation_id} مع Outlook'
        )
        
        # إرسال الملف للتنزيل
        return send_file(
            eml_bytes,
            mimetype='message/rfc822',
            as_attachment=True,
            download_name=eml_filename
        )
    
    except Exception as e:
        current_app.logger.error(f"خطأ في إنشاء ملف .eml للعملية {operation_id}: {str(e)}")
        flash(f'حدث خطأ: {str(e)}', 'danger')
        return redirect(url_for('operations.view_operation', operation_id=operation_id))
        
    finally:
        # حذف الملفات المؤقتة
        if excel_file_path and os.path.exists(excel_file_path):
            try:
                os.remove(excel_file_path)
            except:
                pass
        
        if pdf_file_path and os.path.exists(pdf_file_path):
            try:
                os.remove(pdf_file_path)
            except:
                pass


@operations_bp.route('/<int:operation_id>/share-data', methods=['GET'])
@login_required
def share_data(operation_id):
    """إرجاع بيانات للمشاركة عبر Web Share API"""
    
    try:
        operation = OperationRequest.query.get_or_404(operation_id)
        
        # بناء الرسالة المنسقة
        message_parts = ["السادة المعنيين، تحية طيبة وبعد،\n\n"]
        message_parts.append("مرفق لكم تفاصيل عملية استلام أو تسليم المركبة وفقاً للمعلومات التالية:\n\n")
        
        # نوع العملية
        if operation.operation_type == 'handover':
            handover = VehicleHandover.query.get(operation.related_record_id) if operation.related_record_id else None
            if handover:
                operation_title = "🔄 تسليم مركبة" if handover.handover_type == 'delivery' else "🔄 استلام مركبة"
            else:
                operation_title = "🔄 عملية تسليم/استلام"
        elif operation.operation_type == 'workshop':
            operation_title = "🔧 دخول ورشة"
        elif operation.operation_type == 'safety_inspection':
            operation_title = "✅ فحص سلامة"
        else:
            operation_title = "📋 عملية"
        
        message_parts.append(f"{operation_title}\n")
        
        # معلومات المركبة
        if operation.vehicle:
            vehicle_info = f"{operation.vehicle.plate_number} - {operation.vehicle.make} {operation.vehicle.model}"
            message_parts.append(f"• رقم السيارة: {vehicle_info}\n")
        
        # نوع العملية
        if operation.operation_type == 'handover':
            handover = VehicleHandover.query.get(operation.related_record_id) if operation.related_record_id else None
            if handover:
                operation_type_text = "تسليم" if handover.handover_type == 'delivery' else "استلام"
                message_parts.append(f"• نوع العملية: {operation_type_text}\n")
                message_parts.append(f"• تاريخ العملية: {handover.handover_date.strftime('%Y-%m-%d')}\n")
                
                # معلومات السائق
                if handover.person_name:
                    message_parts.append("\n👤 معلومات السائق\n")
                    message_parts.append(f"• اسم الموظف: {handover.person_name}\n")
                    
                    # جلب بيانات الموظف الكاملة
                    employee = None
                    
                    # محاولة الحصول على الموظف من العلاقة أولاً
                    if handover.driver_employee:
                        employee = handover.driver_employee
                    else:
                        # إذا لم يكن مرتبطاً، ابحث عن الموظف بالاسم
                        employee = Employee.query.filter_by(name=handover.person_name).first()
                    
                    if employee:
                        # رقم الإقامة
                        residency = handover.driver_residency_number or employee.national_id or "[غير متوفر]"
                        message_parts.append(f"• رقم الإقامة: {residency}\n")
                        
                        # رقم الموظف
                        emp_number = employee.employee_id or "[غير متوفر]"
                        message_parts.append(f"• رقم الموظف: {emp_number}\n")
                        
                        # القسم
                        department_name = employee.department.name if employee.department else "[غير متوفر]"
                        message_parts.append(f"• القسم: {department_name}\n")
                        
                        # تاريخ الميلاد
                        birth_date = employee.birth_date.strftime('%Y-%m-%d') if employee.birth_date else "[غير متوفر]"
                        message_parts.append(f"• تاريخ الميلاد: {birth_date}\n")
                        
                        # المدينة
                        city = handover.city or employee.location or "[غير متوفر]"
                        message_parts.append(f"• المدينة: {city}\n")
                    else:
                        # في حالة عدم العثور على الموظف، نستخدم البيانات المتاحة فقط
                        if handover.driver_residency_number:
                            message_parts.append(f"• رقم الإقامة: {handover.driver_residency_number}\n")
                        if handover.driver_phone_number:
                            message_parts.append(f"• الجوال: {handover.driver_phone_number}\n")
                
                # معلومات إضافية
                message_parts.append("\n📊 تفاصيل إضافية\n")
                message_parts.append(f"• المسافة: {handover.mileage:,} كم\n")
                
                # الملاحظات
                if handover.notes:
                    message_parts.append(f"\n💬 ملاحظات\n{handover.notes}\n")
        
        message_parts.append("\n📎 مرفقات:\n\n")
        message_parts.append("ملف Excel\n\n")
        message_parts.append("ملف PDF\n\n")
        message_parts.append("شاكرين لكم تعاونكم المستمر، وفي حال الحاجة لأي تفاصيل إضافية أو استفسارات، نحن في خدمتكم.\n\n")
        message_parts.append("مع خالص التحية،")
        
        message = ''.join(message_parts)
        
        # بناء URLs للملفات
        excel_url = url_for('operations.export_operation_excel', operation_id=operation_id, _external=True)
        
        # URL لملف PDF (نحتاج related_record_id)
        pdf_url = None
        if operation.operation_type == 'handover' and operation.related_record_id:
            pdf_url = url_for('vehicles.handover_pdf_public', id=operation.related_record_id, _external=True)
        
        return jsonify({
            'success': True,
            'message': message,
            'excelUrl': excel_url,
            'pdfUrl': pdf_url
        })
        
    except Exception as e:
        current_app.logger.error(f"خطأ في إنشاء بيانات المشاركة للعملية {operation_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@operations_bp.route('/<int:operation_id>/share-package', methods=['GET'])
@login_required
def share_package(operation_id):
    """إنشاء حزمة ZIP شاملة للمشاركة الخارجية"""
    
    operation = OperationRequest.query.get_or_404(operation_id)
    
    # مسار مجلد مؤقت لتجميع الملفات في مجلد آمن بدلاً من /tmp
    temp_dir = os.path.join(current_app.static_folder, '.temp', f'operation_{operation_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
    zip_path = None
    
    try:
        os.makedirs(temp_dir, exist_ok=True)
        
        # 1. إنشاء ملف نصي بالتفاصيل
        details_path = os.path.join(temp_dir, 'تفاصيل_العملية.txt')
        with open(details_path, 'w', encoding='utf-8') as f:
            f.write('═' * 50 + '\n')
            f.write(f'          تفاصيل العملية #{operation.id}\n')
            f.write('═' * 50 + '\n\n')
            
            # نوع العملية
            operation_types = {
                'handover': 'تسليم/استلام مركبة',
                'workshop': 'ورشة صيانة',
                'external_authorization': 'تفويض خارجي',
                'safety_inspection': 'فحص سلامة'
            }
            f.write(f'نوع العملية: {operation_types.get(operation.operation_type, operation.operation_type)}\n')
            f.write(f'الحالة: {operation.status}\n')
            f.write(f'التاريخ: {operation.created_at.strftime("%Y/%m/%d %H:%M")}\n\n')
            
            # معلومات المركبة
            if operation.vehicle:
                f.write('─' * 50 + '\n')
                f.write('معلومات المركبة:\n')
                f.write('─' * 50 + '\n')
                f.write(f'رقم اللوحة: {operation.vehicle.plate_number}\n')
                f.write(f'النوع: {operation.vehicle.make} {operation.vehicle.model}\n')
                
                # محاولة الحصول على السائق من آخر عملية تسليم
                last_handover = VehicleHandover.query.filter_by(
                    vehicle_id=operation.vehicle.id,
                    handover_type='delivery'
                ).order_by(VehicleHandover.handover_date.desc()).first()
                
                if last_handover and last_handover.person_name:
                    f.write(f'السائق الحالي: {last_handover.person_name}\n')
                f.write('\n')
            
            # تفاصيل خاصة بنوع العملية
            if operation.operation_type == 'handover' and operation.related_record_id:
                handover = VehicleHandover.query.get(operation.related_record_id)
                if handover:
                    f.write('─' * 50 + '\n')
                    f.write('تفاصيل التسليم/الاستلام:\n')
                    f.write('─' * 50 + '\n')
                    f.write(f'النوع: {"تسليم" if handover.handover_type == "delivery" else "استلام"}\n')
                    f.write(f'اسم المستلم: {handover.person_name}\n')
                    f.write(f'المسافة المقطوعة: {handover.mileage} كم\n')
                    if handover.city:
                        f.write(f'المدينة: {handover.city}\n')
                    if handover.project_name:
                        f.write(f'المشروع: {handover.project_name}\n')
                    if handover.notes:
                        f.write(f'\nملاحظات:\n{handover.notes}\n')
                    f.write('\n')
            
            elif operation.operation_type == 'workshop' and operation.related_record_id:
                workshop = VehicleWorkshop.query.get(operation.related_record_id)
                if workshop:
                    f.write('─' * 50 + '\n')
                    f.write('تفاصيل الورشة:\n')
                    f.write('─' * 50 + '\n')
                    f.write(f'السبب: {workshop.reason}\n')
                    f.write(f'تاريخ الدخول: {workshop.entry_date.strftime("%Y/%m/%d")}\n')
                    if workshop.exit_date:
                        f.write(f'تاريخ الخروج: {workshop.exit_date.strftime("%Y/%m/%d")}\n')
                    if workshop.notes:
                        f.write(f'\nملاحظات:\n{workshop.notes}\n')
                    f.write('\n')
            
            # معلومات إضافية
            if operation.description:
                f.write('─' * 50 + '\n')
                f.write('الوصف:\n')
                f.write('─' * 50 + '\n')
                f.write(f'{operation.description}\n\n')
            
            f.write('═' * 50 + '\n')
            f.write('تم إنشاء هذا الملف من نظام نُظم لإدارة المركبات\n')
            f.write('═' * 50 + '\n')
        
        # 2. إنشاء ملف Excel
        try:
            excel_path = os.path.join(temp_dir, f'بيانات_العملية_{operation_id}.xlsx')
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'تفاصيل العملية'
            
            # إضافة البيانات
            ws['A1'] = 'البيان'
            ws['B1'] = 'القيمة'
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            
            row = 2
            ws[f'A{row}'] = 'رقم العملية'
            ws[f'B{row}'] = f'#{operation.id}'
            row += 1
            
            ws[f'A{row}'] = 'نوع العملية'
            ws[f'B{row}'] = operation_types.get(operation.operation_type, operation.operation_type)
            row += 1
            
            if operation.vehicle:
                ws[f'A{row}'] = 'رقم اللوحة'
                ws[f'B{row}'] = operation.vehicle.plate_number
                row += 1
            
            wb.save(excel_path)
        except Exception as e:
            current_app.logger.warning(f'فشل في إنشاء Excel: {str(e)}')
            excel_path = None
        
        # 3. إنشاء ملف PDF
        pdf_path = None
        if operation.operation_type == 'handover' and operation.related_record_id:
            try:
                from src.utils.simple_pdf_generator import create_vehicle_handover_pdf
                pdf_content = create_vehicle_handover_pdf(operation.related_record_id)
                pdf_path = os.path.join(temp_dir, f'تقرير_العملية_{operation_id}.pdf')
                with open(pdf_path, 'wb') as f:
                    f.write(pdf_content)
            except Exception as e:
                current_app.logger.warning(f'فشل في إنشاء PDF: {str(e)}')
                pdf_path = None
        
        # 4. نسخ الصور المرفقة
        images_copied = 0
        if operation.operation_type == 'handover' and operation.related_record_id:
            handover_images = VehicleHandoverImage.query.filter_by(handover_record_id=operation.related_record_id).all()
            for img in handover_images:
                try:
                    src_path = os.path.join('static/uploads/handover_images', img.get_path())
                    if os.path.exists(src_path):
                        dst_path = os.path.join(temp_dir, f'صورة_{images_copied + 1}_{os.path.basename(img.get_path())}')
                        shutil.copy2(src_path, dst_path)
                        images_copied += 1
                except Exception as e:
                    current_app.logger.warning(f'فشل في نسخ الصورة: {str(e)}')
        
        elif operation.operation_type == 'workshop' and operation.related_record_id:
            workshop_images = VehicleWorkshopImage.query.filter_by(workshop_record_id=operation.related_record_id).all()
            for img in workshop_images:
                try:
                    src_path = os.path.join('static/uploads/workshop_images', img.image_path)
                    if os.path.exists(src_path):
                        dst_path = os.path.join(temp_dir, f'صورة_{images_copied + 1}_{os.path.basename(img.image_path)}')
                        shutil.copy2(src_path, dst_path)
                        images_copied += 1
                except Exception as e:
                    current_app.logger.warning(f'فشل في نسخ الصورة: {str(e)}')
        
        # 5. إنشاء ملف ZIP
        # حفظ ملف ZIP في مجلد آمن بدلاً من /tmp
        zip_dir = os.path.join(current_app.static_folder, '.temp')
        os.makedirs(zip_dir, exist_ok=True)
        zip_path = os.path.join(zip_dir, f'operation_{operation_id}_package.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zipf.write(file_path, arcname)
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='share_package',
            entity_type='operation_request',
            entity_id=operation.id,
            details=f'إنشاء حزمة مشاركة شاملة للعملية {operation_id}'
        )
        
        # إرسال الملف
        return send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'عملية_{operation_id}_شاملة.zip'
        )
    
    except Exception as e:
        current_app.logger.error(f"خطأ في إنشاء حزمة المشاركة للعملية {operation_id}: {str(e)}")
        flash(f'حدث خطأ: {str(e)}', 'danger')
        return redirect(url_for('operations.view_operation', operation_id=operation_id))
    
    finally:
        # 💾 جميع الملفات والمجلدات تبقى محفوظة بشكل دائم في static/.temp/
        # لا يتم حذف أي شيء - كل المعالجات آمنة
        pass


def get_operation_type_name(operation_type):
    """تحويل نوع العملية إلى النص العربي"""
    type_names = {
        'handover': 'تسليم/استلام',
        'workshop': 'ورشة صيانة',
        'external_authorization': 'تفويض خارجي',
        'safety_inspection': 'فحص سلامة'
    }
    return type_names.get(operation_type, operation_type)

# ============ تقارير الحوادث - مراجعة وموافقة ============

@operations_bp.route('/accident-reports', methods=['GET'])
@login_required
def list_accident_reports():
    """عرض قائمة تقارير الحوادث المعلقة"""
    from models import VehicleAccident, VehicleAccidentImage
    
    # الحصول على حالة الفلتر
    status_filter = request.args.get('status', 'pending')
    
    # بناء الاستعلام
    query = VehicleAccident.query
    
    if status_filter and status_filter != 'all':
        query = query.filter_by(review_status=status_filter)
    
    # ترتيب حسب التاريخ
    accidents = query.order_by(VehicleAccident.accident_date.desc(), VehicleAccident.created_at.desc()).all()
    
    # إحصائيات
    total_pending = VehicleAccident.query.filter_by(review_status='pending').count()
    total_approved = VehicleAccident.query.filter_by(review_status='approved').count()
    total_rejected = VehicleAccident.query.filter_by(review_status='rejected').count()
    total_under_review = VehicleAccident.query.filter_by(review_status='under_review').count()
    
    return render_template('operations/accident_reports_list.html',
                          accidents=accidents,
                          status_filter=status_filter,
                          stats={
                              'total_pending': total_pending,
                              'total_approved': total_approved,
                              'total_rejected': total_rejected,
                              'total_under_review': total_under_review
                          })


@operations_bp.route('/accident-reports/<int:accident_id>', methods=['GET'])
@login_required
def view_accident_report(accident_id):
    """عرض تفاصيل تقرير حادث"""
    from models import VehicleAccident, VehicleAccidentImage
    
    accident = VehicleAccident.query.get_or_404(accident_id)
    images = accident.images.all()
    
    return render_template('operations/accident_report_details.html',
                          accident=accident,
                          images=images)


@operations_bp.route('/accident-reports/<int:accident_id>/review', methods=['POST'])
@login_required
def review_accident_report(accident_id):
    """مراجعة والموافقة/رفض تقرير حادث"""
    from models import VehicleAccident
    
    accident = VehicleAccident.query.get_or_404(accident_id)
    
    action = request.form.get('action')  # approve, reject, under_review
    reviewer_notes = request.form.get('reviewer_notes', '')
    
    if action not in ['approve', 'reject', 'under_review']:
        flash('إجراء غير صالح', 'danger')
        return redirect(url_for('operations.view_accident_report', accident_id=accident_id))
    
    # تحديث حالة المراجعة
    if action == 'approve':
        accident.review_status = 'approved'
        accident.accident_status = 'معتمد'
        flash_message = 'تم اعتماد تقرير الحادث بنجاح'
        flash_type = 'success'
    elif action == 'reject':
        accident.review_status = 'rejected'
        flash_message = 'تم رفض تقرير الحادث'
        flash_type = 'warning'
    else:  # under_review
        accident.review_status = 'under_review'
        flash_message = 'تم تحديث حالة التقرير إلى "قيد المراجعة"'
        flash_type = 'info'
    
    accident.reviewed_by = current_user.id
    accident.reviewed_at = datetime.utcnow()
    accident.reviewer_notes = reviewer_notes
    
    # إذا كانت موافقة، يمكن إضافة بيانات مالية
    if action == 'approve':
        liability_percentage = request.form.get('liability_percentage')
        deduction_amount = request.form.get('deduction_amount')
        vehicle_condition = request.form.get('vehicle_condition')
        
        if liability_percentage:
            accident.liability_percentage = int(liability_percentage)
        if deduction_amount:
            accident.deduction_amount = float(deduction_amount)
        if vehicle_condition:
            accident.vehicle_condition = vehicle_condition
    
    try:
        db.session.commit()
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action=f'accident_report_{action}',
            entity_type='vehicle_accident',
            entity_id=accident.id,
            details=f'مراجعة تقرير الحادث: {flash_message}'
        )
        
        flash(flash_message, flash_type)
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('operations.view_accident_report', accident_id=accident_id))


@operations_bp.route('/accident-reports/delete', methods=['POST'])
@login_required
def delete_accident_reports():
    """حذف مجموعة من تقارير الحوادث"""
    from models import VehicleAccident, VehicleAccidentImage
    import os
    import shutil
    
    if current_user.role != UserRole.ADMIN:
        flash('غير مسموح لك بالوصول لهذه الصفحة', 'danger')
        return redirect(url_for('operations.list_accident_reports'))
    
    accident_ids = request.form.getlist('accident_ids')
    
    if not accident_ids:
        flash('الرجاء تحديد تقرير واحد على الأقل للحذف', 'warning')
        return redirect(url_for('operations.list_accident_reports'))
    
    try:
        deleted_count = 0
        for accident_id in accident_ids:
            accident = VehicleAccident.query.get(int(accident_id))
            if accident:
                # حذف الملفات من السيرفر
                accident_folder = os.path.join('static', 'uploads', 'accidents', str(accident.id))
                if os.path.exists(accident_folder):
                    try:
                        shutil.rmtree(accident_folder)
                        current_app.logger.info(f"تم حذف مجلد التقرير: {accident_folder}")
                    except Exception as e:
                        current_app.logger.error(f"خطأ في حذف مجلد التقرير {accident_folder}: {e}")
                
                # حذف صور الحادث من قاعدة البيانات
                VehicleAccidentImage.query.filter_by(accident_id=accident.id).delete()
                
                # حذف التقرير من قاعدة البيانات
                db.session.delete(accident)
                deleted_count += 1
                
                # تسجيل العملية
                log_audit(
                    user_id=current_user.id,
                    action='delete_accident_report',
                    entity_type='vehicle_accident',
                    entity_id=accident.id,
                    details=f'حذف تقرير حادث للمركبة {accident.vehicle.plate_number}'
                )
        
        db.session.commit()
        flash(f'تم حذف {deleted_count} تقرير بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"خطأ في حذف تقارير الحوادث: {e}")
        flash(f'حدث خطأ أثناء الحذف: {str(e)}', 'danger')
    
    return redirect(url_for('operations.list_accident_reports'))


@operations_bp.route('/accident-reports/<int:accident_id>/download-pdf', methods=['GET'])
@login_required
def download_accident_report_pdf(accident_id):
    """تنزيل تقرير الحادث بصيغة PDF"""
    from models import VehicleAccident
    from src.utils.accident_report_pdf import generate_accident_report_pdf
    import io
    from flask import make_response
    
    accident = VehicleAccident.query.get_or_404(accident_id)
    
    try:
        # إنشاء PDF
        pdf = generate_accident_report_pdf(accident)
        
        # إنشاء response (output returns bytearray in fpdf2, convert to bytes)
        pdf_output = bytes(pdf.output())
        
        response = make_response(pdf_output)
        response.headers['Content-Type'] = 'application/pdf'
        
        # اسم الملف
        filename = f'accident_report_{accident.id}_{accident.vehicle.plate_number}_{accident.accident_date.strftime("%Y%m%d")}.pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # تسجيل العملية
        log_audit(
            user_id=current_user.id,
            action='download_accident_pdf',
            entity_type='vehicle_accident',
            entity_id=accident.id,
            details=f'تنزيل تقرير PDF للحادث {accident.id}'
        )
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"خطأ في إنشاء PDF: {e}")
        flash(f'حدث خطأ أثناء إنشاء ملف PDF: {str(e)}', 'danger')
        return redirect(url_for('operations.view_accident_report', accident_id=accident_id))
