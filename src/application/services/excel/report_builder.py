"""
بناء التقارير الاحترافية
=======================
تجميع جميع مكونات التقرير في ملف احترافي
"""

from typing import Union, Tuple, Optional
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import pandas as pd

from .styles import ExcelStyles, TextFormatters, ColorPalette
from .data_extractor import DataExtractor
from .chart_generator import ChartGenerator


class SheetBuilder:
    """بناء الأوراق الفردية"""
    
    def __init__(self):
        """تهيئة بناء الأوراق"""
        self.styles = ExcelStyles()
        self.data_extractor = DataExtractor()
        self.chart_generator = ChartGenerator()
    
    def build_header(self, sheet, title: str, subtitle: str = '', row: int = 1):
        """
        بناء رأس الصفحة الاحترافي
        
        Args:
            sheet: ورقة العمل
            title: العنوان الرئيسي
            subtitle: العنوان الفرعي
            row: رقم الصف
        """
        # دمج الخلايا للعنوان
        sheet.merge_cells(f'A{row}:J{row}')
        title_cell = sheet[f'A{row}']
        title_cell.value = title
        title_cell.font = self.styles.title_font()
        title_cell.alignment = self.styles.center_alignment()
        title_cell.fill = PatternFill(
            start_color=ColorPalette.NAVY_BLUE,
            end_color=ColorPalette.NAVY_BLUE,
            fill_type='solid'
        )
        
        # العنوان الفرعي
        if subtitle:
            sheet.merge_cells(f'A{row+1}:J{row+1}')
            subtitle_cell = sheet[f'A{row+1}']
            subtitle_cell.value = subtitle
            subtitle_cell.font = self.styles.subheader_font()
            subtitle_cell.alignment = self.styles.center_alignment()
            subtitle_cell.fill = self.styles.subheader_fill()
        
        # ضبط ارتفاع الصفوف
        sheet.row_dimensions[row].height = 30
        if subtitle:
            sheet.row_dimensions[row+1].height = 25
    def build_metrics_section(self, sheet, metrics: dict, start_row: int = 1):
        """
        بناء قسم المقاييس الرئيسية
        
        Args:
            sheet: ورقة العمل
            metrics: قاموس المقاييس
            start_row: الصف البادئ
        """
        from openpyxl.styles import PatternFill
        
        metric_keys = ['total_records', 'today_records', 'growth_rate']
        metric_labels = ['إجمالي السجلات', 'سجلات اليوم', 'معدل النمو']
        
        col = 1
        for key, label in zip(metric_keys, metric_labels):
            if key in metrics:
                # خلية العنوان
                cell_title = sheet.cell(row=start_row, column=col)
                cell_title.value = label
                cell_title.font = self.styles.header_font()
                cell_title.fill = self.styles.header_fill()
                cell_title.alignment = self.styles.center_alignment()
                
                # خلية القيمة
                cell_value = sheet.cell(row=start_row+1, column=col)
                value = metrics[key]
                if key == 'growth_rate':
                    cell_value.value = f"{value:.2f}%"
                else:
                    cell_value.value = int(value)
                cell_value.font = self.styles.metric_font()
                cell_value.fill = self.styles.metric_fill()
                cell_value.alignment = self.styles.center_alignment()
                cell_value.border = self.styles.thick_border()
                
                sheet.column_dimensions[get_column_letter(col)].width = 20
                col += 1
    
    def build_table(self, sheet, data: pd.DataFrame, 
                   start_row: int = 1, start_col: int = 1) -> int:
        """
        بناء جدول البيانات الاحترافي
        
        Args:
            sheet: ورقة العمل
            data: بيانات الجدول
            start_row: الصف البادئ
            start_col: العمود البادئ
            
        Returns:
            آخر صف استخدم
        """
        if data.empty:
            return start_row
        
        # كتابة رؤوس الأعمدة
        for col_idx, column in enumerate(data.columns, 1):
            cell = sheet.cell(row=start_row, column=start_col + col_idx - 1)
            cell.value = str(column)
            cell_style = self.styles.header_style()
            cell.font = cell_style['font']
            cell.fill = cell_style['fill']
            cell.alignment = cell_style['alignment']
            cell.border = cell_style['border']
            
            # ضبط عرض العمود
            sheet.column_dimensions[get_column_letter(start_col + col_idx - 1)].width = 18
        
        # كتابة البيانات
        for row_idx, (_, row) in enumerate(data.iterrows(), 1):
            for col_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=start_row + row_idx, 
                                 column=start_col + col_idx - 1)
                cell.value = value
                cell_style = self.styles.data_cell_style()
                cell.font = cell_style['font']
                cell.alignment = cell_style['alignment']
                cell.border = cell_style['border']
        
        return start_row + len(data) + 1
    
    def add_footer(self, sheet, text: str, last_row: int):
        """
        إضافة تذييل الصفحة
        
        Args:
            sheet: ورقة العمل
            text: نص التذييل
            last_row: آخر صف
        """
        footer_row = last_row + 2
        sheet.merge_cells(f'A{footer_row}:J{footer_row}')
        footer_cell = sheet[f'A{footer_row}']
        footer_cell.value = text
        footer_cell.font = self.styles.normal_font()
        footer_cell.alignment = self.styles.center_alignment()
        footer_cell.fill = PatternFill(
            start_color=ColorPalette.LIGHT_GRAY,
            end_color=ColorPalette.LIGHT_GRAY,
            fill_type='solid'
        )

class ReportBuilder:
    """بناء التقرير الكامل"""
    
    def __init__(self):
        """تهيئة بناء التقرير"""
        self.sheet_builder = SheetBuilder()
        self.data_extractor = DataExtractor()
        self.chart_generator = ChartGenerator()
        self.styles = ExcelStyles()
    
    def create_dashboard_sheet(self, workbook: Workbook) -> None:
        """
        إنشاء ورقة لوحة الحكم
        
        Args:
            workbook: دفتر العمل
        """
        # إنشاء الورقة
        sheet = workbook.create_sheet('لوحة العمل', 0)
        
        # ضبط خصائص الورقة
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_margins.left = 0.5
        sheet.page_margins.right = 0.5
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        
        # بناء الرأس
        self.sheet_builder.build_header(
            sheet,
            'لوحة القيادة الشاملة',
            'تقرير شامل للأداء والإحصائيات',
            row=1
        )
        
        # استخراج البيانات
        metrics = self.data_extractor.get_dashboard_metrics()
        summary = self.data_extractor.get_summary_statistics()
        category_data = self.data_extractor.get_category_distribution()
        monthly_data = self.data_extractor.get_monthly_trends()
        performance_data = self.data_extractor.get_performance_data()
        
        # بناء قسم المقاييس
        self.sheet_builder.build_metrics_section(sheet, metrics, start_row=4)
        
        # بناء جدول الفئات
        current_row = 9
        sheet[f'A{current_row}'] = 'توزيع الفئات'
        sheet[f'A{current_row}'].font = self.styles.subheader_font()
        current_row += 1
        
        current_row = self.sheet_builder.build_table(
            sheet,
            category_data,
            start_row=current_row
        )
        
        # بناء جدول الاتجاهات الشهرية
        current_row += 2
        sheet[f'A{current_row}'] = 'الاتجاهات الشهرية'
        sheet[f'A{current_row}'].font = self.styles.subheader_font()
        current_row += 1
        
        current_row = self.sheet_builder.build_table(
            sheet,
            monthly_data,
            start_row=current_row
        )
        
        # إضافة التذييل
        self.sheet_builder.add_footer(
            sheet,
            f"تم إنشاء التقرير في {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            current_row
        )
    
    def create_details_sheet(self, workbook: Workbook) -> None:
        """
        إنشاء ورقة التفاصيل
        
        Args:
            workbook: دفتر العمل
        """
        sheet = workbook.create_sheet('التفاصيل', 1)
        
        # ضبط خصائص الورقة
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        
        # بناء الرأس
        self.sheet_builder.build_header(
            sheet,
            'تقرير التفاصيل',
            'بيانات الأداء والموارد البشرية',
            row=1
        )
        
        # استخراج البيانات
        performance_data = self.data_extractor.get_performance_data()
        top_performers = self.data_extractor.get_top_performers(10)
        
        # بناء جدول الأداء
        current_row = 5
        sheet[f'A{current_row}'] = 'بيانات الأداء'
        sheet[f'A{current_row}'].font = self.styles.subheader_font()
        current_row += 1
        
        current_row = self.sheet_builder.build_table(
            sheet,
            performance_data,
            start_row=current_row
        )
        
        # بناء جدول الأفضل أداءً
        current_row += 2
        sheet[f'A{current_row}'] = 'الموظفون الأفضل أداءً'
        sheet[f'A{current_row}'].font = self.styles.subheader_font()
        current_row += 1
        
        current_row = self.sheet_builder.build_table(
            sheet,
            top_performers,
            start_row=current_row
        )
        
        # إضافة التذييل
        self.sheet_builder.add_footer(
            sheet,
            f"تم إنشاء التقرير في {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            current_row
        )
    
    def build_complete_report(self, workbook: Workbook) -> Workbook:
        """
        بناء التقرير الكامل
        
        Args:
            workbook: دفتر العمل
            
        Returns:
            دفتر العمل المكتمل
        """
        self.create_dashboard_sheet(workbook)
        self.create_details_sheet(workbook)
        
        return workbook
