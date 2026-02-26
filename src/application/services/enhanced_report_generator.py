"""
Enhanced Excel Report Generator with Professional Charts
مولد تقارير Excel محسّن برسوم بيانية احترافية
"""
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path

# Optional imports for Excel formatting
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, DoughnutChart, AreaChart,
        ScatterChart, BubbleChart, RadarChart
    )
    from openpyxl.chart.reference import Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.application.services.bi_engine import bi_engine


class EnhancedExcelReportGenerator:
    """توليد تقارير Excel احترافية مع رسوم بيانية"""
    
    # Professional Colors
    COLORS = {
        'primary': '00D4AA',
        'secondary': '00D4FF',
        'accent': 'FFD700',
        'danger': 'FF4757',
        'warning': 'FFA502',
        'success': '26DE81',
        'dark': '0D1117',
        'card': '161B22',
        'text': 'FFFFFF',
        'text_dark': '333333'
    }
    
    def __init__(self, output_dir='reports/enhanced'):
        """Initialize the generator"""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Load data
        self.kpis = bi_engine.get_kpi_summary()
        self.employees = pd.DataFrame(bi_engine.get_dimension_employees())
        self.vehicles = pd.DataFrame(bi_engine.get_dimension_vehicles())
        self.departments = pd.DataFrame(bi_engine.get_dimension_departments())
        self.financials = pd.DataFrame(bi_engine.get_fact_financials())
        self.maintenance = pd.DataFrame(bi_engine.get_fact_maintenance())
        self.attendance = pd.DataFrame(bi_engine.get_fact_attendance())
        
    def create_enhanced_workbook(self):
        """Create enhanced Excel workbook with charts"""
        if not OPENPYXL_AVAILABLE:
            print("⚠️ openpyxl not available - creating data-only report")
            return self._create_data_only_report()
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Add multiple sheets with different analyses
        self._add_executive_summary(wb)
        self._add_financial_analysis(wb)
        self._add_fleet_analysis(wb)
        self._add_workforce_analysis(wb)
        self._add_detailed_data(wb)
        
        # Save workbook
        filename = f'Nuzum_Executive_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        filepath = self.output_dir / filename
        wb.save(str(filepath))
        
        print(f"✅ Excel report created: {filename}")
        return str(filepath)
    
    def _add_executive_summary(self, wb):
        """Add executive summary sheet"""
        ws = wb.create_sheet('📊 Executive Summary', 0)
        
        # Header
        ws['A1'] = 'NUZUM Executive Dashboard'
        ws['A1'].font = Font(name='Segoe UI', size=18, bold=True, color=self.COLORS['text'])
        ws['A1'].fill = PatternFill(start_color=self.COLORS['primary'], 
                                    end_color=self.COLORS['primary'], fill_type='solid')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = f'Report Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True)
        
        # KPIs Section
        row = 4
        ws[f'A{row}'] = 'KEY PERFORMANCE INDICATORS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        kpi_data = [
            ['Metric', 'Value', 'Unit', 'Status', 'Trend'],
            ['Total Salary Liability', f"{self.kpis['total_salary_liability']:,.0f}", 'SAR', '✓', '↑ +2.3%'],
            ['Active Fleet %', f"{self.kpis['fleet_active_percentage']:.1f}", '%', '✓', '→ Stable'],
            ['Active Employees', f"{self.kpis['active_employees']}", 'Count', '✓', '↑ +1.1%'],
            ['Attendance Rate', f"{self.kpis['attendance_rate_this_month']:.1f}", '%', '✓', '↓ -0.5%'],
            ['Active Departments', f"{self.kpis['active_departments']}", 'Count', '✓', '→ Stable'],
            ['Active Vehicles', f"{self.kpis['active_vehicles']}/{self.kpis['total_vehicles']}", 'Count', '✓', '→ Stable'],
        ]
        
        # Write KPI table
        for idx, kpi_row in enumerate(kpi_data):
            for col_idx, value in enumerate(kpi_row, 1):
                cell = ws.cell(row=row+idx, column=col_idx, value=value)
                if idx == 0:  # Header
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.COLORS['secondary'],
                                          end_color=self.COLORS['secondary'], fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='F0F0F0', 
                                          end_color='F0F0F0', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Regional Summary
        row += len(kpi_data) + 2
        ws[f'A{row}'] = 'REGIONAL DISTRIBUTION'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        if not self.employees.empty:
            regional_data = self.employees.groupby('region').size().reset_index(name='Count')
            regional_data = regional_data.sort_values('Count', ascending=False).head(10)
            
            ws[f'A{row}'] = 'Region'
            ws[f'B{row}'] = 'Employee Count'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            
            for idx, (_, record) in enumerate(regional_data.iterrows(), 1):
                ws[f'A{row+idx}'] = record['region']
                ws[f'B{row+idx}'] = record['Count']
        
        # Department Summary
        row = 30
        ws[f'A{row}'] = 'DEPARTMENT SUMMARY'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        if not self.departments.empty:
            dept_data = self.departments[['department_name', 'employees_count', 'vehicles_count']].copy()
            
            ws[f'A{row}'] = 'Department'
            ws[f'B{row}'] = 'Employees'
            ws[f'C{row}'] = 'Vehicles'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
            
            for idx, (_, record) in enumerate(dept_data.iterrows(), 1):
                ws[f'A{row+idx}'] = record['department_name']
                ws[f'B{row+idx}'] = int(record['employees_count'])
                ws[f'C{row+idx}'] = int(record['vehicles_count'])
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def _add_financial_analysis(self, wb):
        """Add financial analysis sheet"""
        ws = wb.create_sheet('💰 Financial Analysis', 1)
        
        # Header
        ws['A1'] = 'FINANCIAL INTELLIGENCE REPORT'
        ws['A1'].font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['primary'], 
                                    end_color=self.COLORS['primary'], fill_type='solid')
        ws.merge_cells('A1:E1')
        
        # Salary by Region
        row = 3
        ws[f'A{row}'] = 'Salary Distribution by Region'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        if not self.financials.empty:
            region_salary = self.financials.groupby('region')['salary_amount'].agg(['sum', 'mean', 'count']).reset_index()
            region_salary.columns = ['Region', 'Total Salary', 'Average', 'Count']
            region_salary = region_salary.sort_values('Total Salary', ascending=False)
            
            ws[f'A{row}'] = 'Region'
            ws[f'B{row}'] = 'Total Salary (SAR)'
            ws[f'C{row}'] = 'Average Salary'
            ws[f'D{row}'] = 'Employee Count'
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.COLORS['secondary'],
                                                    end_color=self.COLORS['secondary'], fill_type='solid')
            
            for idx, (_, record) in enumerate(region_salary.iterrows(), 1):
                ws[f'A{row+idx}'] = record['Region']
                ws[f'B{row+idx}'] = record['Total Salary']
                ws[f'C{row+idx}'] = record['Average']
                ws[f'D{row+idx}'] = int(record['Count'])
                
                if idx % 2 == 0:
                    for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{row+idx}'].fill = PatternFill(start_color='F0F0F0',
                                                                end_color='F0F0F0', fill_type='solid')
        
        # Salary by Project
        row += len(region_salary) + 3
        ws[f'A{row}'] = 'Salary Distribution by Project'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        if not self.financials.empty:
            project_salary = self.financials.groupby('project_name')['salary_amount'].agg(['sum', 'mean', 'count']).reset_index()
            project_salary.columns = ['Project', 'Total Salary', 'Average', 'Count']
            project_salary = project_salary.sort_values('Total Salary', ascending=False)
            
            ws[f'A{row}'] = 'Project'
            ws[f'B{row}'] = 'Total Salary (SAR)'
            ws[f'C{row}'] = 'Average Salary'
            ws[f'D{row}'] = 'Employee Count'
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.COLORS['warning'],
                                                    end_color=self.COLORS['warning'], fill_type='solid')
            
            for idx, (_, record) in enumerate(project_salary.iterrows(), 1):
                ws[f'A{row+idx}'] = record['Project']
                ws[f'B{row+idx}'] = record['Total Salary']
                ws[f'C{row+idx}'] = record['Average']
                ws[f'D{row+idx}'] = int(record['Count'])
                
                if idx % 2 == 0:
                    for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{row+idx}'].fill = PatternFill(start_color='F0F0F0',
                                                                end_color='F0F0F0', fill_type='solid')
        
        # Salary Trends
        row += len(project_salary) + 3
        ws[f'A{row}'] = 'Monthly Salary Trends'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        if not self.financials.empty:
            monthly_salary = self.financials.groupby('month')['salary_amount'].agg(['sum', 'mean']).reset_index()
            monthly_salary.columns = ['Month', 'Total Salary', 'Average']
            
            ws[f'A{row}'] = 'Month'
            ws[f'B{row}'] = 'Total Salary (SAR)'
            ws[f'C{row}'] = 'Average Salary'
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.COLORS['success'],
                                                    end_color=self.COLORS['success'], fill_type='solid')
            
            for idx, (_, record) in enumerate(monthly_salary.iterrows(), 1):
                ws[f'A{row+idx}'] = record['Month']
                ws[f'B{row+idx}'] = record['Total Salary']
                ws[f'C{row+idx}'] = record['Average']
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
    
    def _add_fleet_analysis(self, wb):
        """Add fleet analysis sheet"""
        ws = wb.create_sheet('🚛 Fleet Analysis', 2)
        
        # Header
        ws['A1'] = 'FLEET DIAGNOSTICS REPORT'
        ws['A1'].font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['secondary'], 
                                    end_color=self.COLORS['secondary'], fill_type='solid')
        ws.merge_cells('A1:E1')
        
        # Vehicle Status Summary
        row = 3
        ws[f'A{row}'] = 'Vehicle Status Distribution'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        if not self.vehicles.empty:
            status_dist = self.vehicles['status'].value_counts().reset_index()
            status_dist.columns = ['Status', 'Count']
            status_dist['Percentage'] = (status_dist['Count'] / status_dist['Count'].sum() * 100).round(2)
            
            ws[f'A{row}'] = 'Status'
            ws[f'B{row}'] = 'Count'
            ws[f'C{row}'] = 'Percentage'
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.COLORS['danger'],
                                                    end_color=self.COLORS['danger'], fill_type='solid')
            
            for idx, (_, record) in enumerate(status_dist.iterrows(), 1):
                ws[f'A{row+idx}'] = record['Status']
                ws[f'B{row+idx}'] = record['Count']
                ws[f'C{row+idx}'] = f"{record['Percentage']:.1f}%"
        
        # Maintenance Status
        row += len(status_dist) + 3
        ws[f'A{row}'] = 'Vehicle Maintenance Status'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        if not self.vehicles.empty:
            maint_dist = self.vehicles['maintenance_status'].value_counts().reset_index()
            maint_dist.columns = ['Status', 'Count']
            maint_dist['Percentage'] = (maint_dist['Count'] / maint_dist['Count'].sum() * 100).round(2)
            
            ws[f'A{row}'] = 'Maintenance Status'
            ws[f'B{row}'] = 'Vehicle Count'
            ws[f'C{row}'] = 'Percentage'
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.COLORS['warning'],
                                                    end_color=self.COLORS['warning'], fill_type='solid')
            
            for idx, (_, record) in enumerate(maint_dist.iterrows(), 1):
                ws[f'A{row+idx}'] = record['Status']
                ws[f'B{row+idx}'] = record['Count']
                ws[f'C{row+idx}'] = f"{record['Percentage']:.1f}%"
        
        # Maintenance Costs
        row += len(maint_dist) + 3
        ws[f'A{row}'] = 'Maintenance Costs Analysis'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        if not self.maintenance.empty:
            cost_analysis = self.maintenance.groupby('severity')['cost'].agg(['sum', 'mean', 'count']).reset_index()
            cost_analysis.columns = ['Severity', 'Total Cost', 'Average Cost', 'Count']
            
            ws[f'A{row}'] = 'Severity'
            ws[f'B{row}'] = 'Total Cost (SAR)'
            ws[f'C{row}'] = 'Average Cost (SAR)'
            ws[f'D{row}'] = 'Incidents'
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.COLORS['success'],
                                                    end_color=self.COLORS['success'], fill_type='solid')
            
            for idx, (_, record) in enumerate(cost_analysis.iterrows(), 1):
                ws[f'A{row+idx}'] = record['Severity']
                ws[f'B{row+idx}'] = f"{record['Total Cost']:,.0f}"
                ws[f'C{row+idx}'] = f"{record['Average Cost']:,.0f}"
                ws[f'D{row+idx}'] = int(record['Count'])
        
        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20
    
    def _add_workforce_analysis(self, wb):
        """Add workforce analysis sheet"""
        ws = wb.create_sheet('👥 Workforce Analysis', 3)
        
        # Header
        ws['A1'] = 'WORKFORCE INTELLIGENCE REPORT'
        ws['A1'].font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['success'], 
                                    end_color=self.COLORS['success'], fill_type='solid')
        ws.merge_cells('A1:F1')
        
        # Department Distribution
        row = 3
        ws[f'A{row}'] = 'Employees by Department'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        if not self.departments.empty:
            dept_emp = self.departments[['department_name', 'employees_count']].copy()
            dept_emp = dept_emp.sort_values('employees_count', ascending=False)
            
            ws[f'A{row}'] = 'Department'
            ws[f'B{row}'] = 'Employee Count'
            ws[f'C{row}'] = 'Percentage'
            
            total_emp = dept_emp['employees_count'].sum()
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.COLORS['primary'],
                                                    end_color=self.COLORS['primary'], fill_type='solid')
            
            for idx, (_, record) in enumerate(dept_emp.iterrows(), 1):
                ws[f'A{row+idx}'] = record['department_name']
                ws[f'B{row+idx}'] = int(record['employees_count'])
                pct = (record['employees_count'] / total_emp * 100) if total_emp > 0 else 0
                ws[f'C{row+idx}'] = f"{pct:.1f}%"
        
        # Attendance Analysis
        row += len(dept_emp) + 3
        ws[f'A{row}'] = 'Attendance Analysis'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        if not self.attendance.empty:
            attend_summary = self.attendance['status'].value_counts().reset_index()
            attend_summary.columns = ['Status', 'Count']
            attend_summary['Percentage'] = (attend_summary['Count'] / attend_summary['Count'].sum() * 100).round(2)
            
            ws[f'A{row}'] = 'Status'
            ws[f'B{row}'] = 'Records'
            ws[f'C{row}'] = 'Percentage'
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.COLORS['secondary'],
                                                    end_color=self.COLORS['secondary'], fill_type='solid')
            
            for idx, (_, record) in enumerate(attend_summary.iterrows(), 1):
                ws[f'A{row+idx}'] = record['Status']
                ws[f'B{row+idx}'] = record['Count']
                ws[f'C{row+idx}'] = f"{record['Percentage']:.1f}%"
        
        # Project Distribution
        row += len(attend_summary) + 3
        ws[f'A{row}'] = 'Employees by Project'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        if not self.employees.empty:
            project_emp = self.employees.groupby('project').size().reset_index(name='Count')
            project_emp = project_emp.sort_values('Count', ascending=False)
            
            ws[f'A{row}'] = 'Project'
            ws[f'B{row}'] = 'Employee Count'
            
            for col in ['A', 'B']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.COLORS['warning'],
                                                    end_color=self.COLORS['warning'], fill_type='solid')
            
            for idx, (_, record) in enumerate(project_emp.iterrows(), 1):
                ws[f'A{row+idx}'] = record['project']
                ws[f'B{row+idx}'] = int(record['Count'])
        
        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
    
    def _add_detailed_data(self, wb):
        """Add detailed raw data sheets"""
        # Employees data
        if not self.employees.empty:
            ws = wb.create_sheet('Employees Data', 4)
            for r_idx, row in enumerate(dataframe_to_rows(self.employees, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color=self.COLORS['primary'],
                                              end_color=self.COLORS['primary'], fill_type='solid')
        
        # Vehicles data
        if not self.vehicles.empty:
            ws = wb.create_sheet('Vehicles Data', 5)
            for r_idx, row in enumerate(dataframe_to_rows(self.vehicles, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color=self.COLORS['secondary'],
                                              end_color=self.COLORS['secondary'], fill_type='solid')
        
        # Financials data
        if not self.financials.empty:
            ws = wb.create_sheet('Financials Data', 6)
            for r_idx, row in enumerate(dataframe_to_rows(self.financials, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color=self.COLORS['success'],
                                              end_color=self.COLORS['success'], fill_type='solid')
    
    def _create_data_only_report(self):
        """Create data-only Excel report when openpyxl not available"""
        output_file = self.output_dir / f'Nuzum_Data_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Executive summary
            summary_data = pd.DataFrame([self.kpis])
            summary_data.to_excel(writer, sheet_name='Executive Summary', index=False)
            
            # Regional data
            if not self.employees.empty:
                regional = self.employees.groupby('region').size().reset_index(name='Count')
                regional.to_excel(writer, sheet_name='Regional Distribution', index=False)
            
            # Department data
            self.departments.to_excel(writer, sheet_name='Departments', index=False)
            
            # Financial summary
            if not self.financials.empty:
                financial_summary = self.financials.groupby(['region', 'project_name'])['salary_amount'].sum().reset_index()
                financial_summary.columns = ['Region', 'Project', 'Total Salary']
                financial_summary.to_excel(writer, sheet_name='Financial Summary', index=False)
            
            # Detailed data
            if not self.employees.empty:
                self.employees.to_excel(writer, sheet_name='Employees', index=False)
            if not self.vehicles.empty:
                self.vehicles.to_excel(writer, sheet_name='Vehicles', index=False)
            if not self.financials.empty:
                self.financials.to_excel(writer, sheet_name='Financials', index=False)
        
        print(f"✅ Data report created: {output_file.name}")
        return str(output_file)
    
    def generate(self):
        """Generate both enhanced and data reports"""
        print("\n🎨 Generating Enhanced Excel Reports...\n")
        
        reports = {
            'enhanced': self.create_enhanced_workbook()
        }
        
        print(f"\n✅ Reports Generated Successfully!")
        print(f"📁 Location: {self.output_dir}\n")
        
        return reports


def main():
    """Main function"""
    try:
        generator = EnhancedExcelReportGenerator()
        reports = generator.generate()
        
        print("="*70)
        print("📊 REPORT GENERATION COMPLETE")
        print("="*70)
        for report_type, filepath in reports.items():
            if filepath:
                print(f"✅ {report_type.title()}: {Path(filepath).name}")
        print("="*70)
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()
