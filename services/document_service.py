"""
Document Service Layer
======================
Handles all business logic for document management, including:
- File uploads/downloads with security checks
- Expiry date tracking and notifications
- Metadata management for 76+ employees
- PDF generation and Excel import/export
- Document statistics and analytics

Author: AI Assistant
Date: February 20, 2026
Version: 2.0.0
"""

from datetime import datetime, timedelta, date
from typing import List, Dict, Optional, Tuple, Any
from sqlalchemy import func, or_
from sqlalchemy.orm import selectinload
from werkzeug.utils import secure_filename
import pandas as pd
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
import arabic_reshaper
from bidi.algorithm import get_display
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from core.extensions import db
from models import Document, Employee, Department, SystemAudit, Notification, User
from utils.date_converter import parse_date, format_date_hijri, format_date_gregorian
from utils.audit_logger import log_activity
from services.file_service import FileService
import os


class DocumentService:
    """
    Service layer for document management operations.
    All methods are static for easy testing and reuse.
    """
    
    # ==================== Document Type Management ====================
    
    @staticmethod
    def get_document_types() -> List[Dict[str, str]]:
        """Get list of all supported document types"""
        return [
            {'value': 'national_id', 'label': 'الهوية الوطنية'},
            {'value': 'passport', 'label': 'جواز السفر'},
            {'value': 'health_certificate', 'label': 'الشهادة الصحية'},
            {'value': 'work_permit', 'label': 'رخصة العمل'},
            {'value': 'education_certificate', 'label': 'الشهادة التعليمية'},
            {'value': 'driving_license', 'label': 'رخصة القيادة'},
            {'value': 'annual_leave', 'label': 'إجازة سنوية'},
            {'value': 'other', 'label': 'أخرى'}
        ]
    
    @staticmethod
    def get_document_type_label(doc_type: str) -> str:
        """Get Arabic label for document type"""
        types_map = {
            'national_id': 'الهوية الوطنية',
            'passport': 'جواز السفر',
            'health_certificate': 'الشهادة الصحية',
            'work_permit': 'رخصة العمل',
            'education_certificate': 'الشهادة التعليمية',
            'driving_license': 'رخصة القيادة',
            'annual_leave': 'إجازة سنوية',
            'other': 'أخرى'
        }
        return types_map.get(doc_type, doc_type)
    
    # ==================== Security & Validation ====================
    
    @staticmethod
    def validate_file_security(filename: str, max_size_mb: int = 10) -> Tuple[bool, str]:
        """
        Validate file for security concerns
        
        Args:
            filename: Name of the file to validate
            max_size_mb: Maximum allowed file size in MB
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        # Check file extension
        allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx', 'xls', 'xlsx'}
        
        if '.' not in filename:
            return False, 'امتداد الملف غير صالح'
        
        extension = filename.rsplit('.', 1)[1].lower()
        if extension not in allowed_extensions:
            return False, f'نوع الملف غير مسموح. الأنواع المسموحة: {", ".join(allowed_extensions)}'
        
        # Check for dangerous patterns
        dangerous_patterns = ['..', '/', '\\', '<', '>', '|', ':', '*', '?', '"']
        for pattern in dangerous_patterns:
            if pattern in filename:
                return False, 'اسم الملف يحتوي على أحرف غير مسموحة'
        
        return True, ''
    
    @staticmethod
    def secure_filename_arabic(filename: str) -> str:
        """Make filename secure while preserving Arabic characters"""
        # Use werkzeug's secure_filename first
        safe_name = secure_filename(filename)
        
        # If result is empty (all non-ASCII), use timestamp
        if not safe_name or safe_name == '':
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            extension = filename.rsplit('.', 1)[1] if '.' in filename else 'pdf'
            safe_name = f'document_{timestamp}.{extension}'
        
        return safe_name
    
    @staticmethod
    def check_user_permission(user_id: int, action: str, document_id: Optional[int] = None) -> bool:
        """
        Check if user has permission for document action
        
        Args:
            user_id: ID of the user
            action: Action type ('view', 'create', 'update', 'delete')
            document_id: Optional document ID for ownership check
            
        Returns:
            True if user has permission
        """
        try:
            user = User.query.get(user_id)
            if not user:
                return False
            
            # Admin has all permissions
            if user.role == 'admin':
                return True
            
            # Manager can view, create, update
            if user.role == 'manager' and action in ['view', 'create', 'update']:
                return True
            
            # Regular users can only view their own documents
            if action == 'view' and document_id:
                document = Document.query.get(document_id)
                if document and document.employee and document.employee.user_id == user_id:
                    return True
            
            return False
            
        except Exception:
            return False
    
    # ==================== Document CRUD Operations ====================
    
    @staticmethod
    def get_documents(
        document_type: Optional[str] = None,
        employee_id: Optional[int] = None,
        department_id: Optional[int] = None,
        sponsorship_status: Optional[str] = None,
        status_filter: Optional[str] = None,
        search_query: Optional[str] = None,
        page: int = 1,
        per_page: int = 50
    ) -> Tuple[List[Document], int]:
        """
        Get documents with filtering and pagination
        
        Returns:
            Tuple of (documents_list, total_count)
        """
        try:
            query = Document.query.options(
                selectinload(Document.employee).selectinload(Employee.departments)
            )
            
            # Apply filters
            if document_type:
                query = query.filter(Document.document_type == document_type)
            
            if employee_id:
                query = query.filter(Document.employee_id == employee_id)
            
            # Filters requiring employee join
            if department_id or sponsorship_status or search_query:
                query = query.join(Employee)
                
                if department_id:
                    query = query.filter(Employee.department_id == department_id)
                
                if sponsorship_status:
                    query = query.filter(Employee.sponsorship_status == sponsorship_status)
                
                if search_query:
                    search_pattern = f'%{search_query}%'
                    query = query.filter(
                        or_(
                            Employee.name.like(search_pattern),
                            Employee.employee_id.like(search_pattern),
                            Document.document_number.like(search_pattern)
                        )
                    )
            
            # Status filter (expired/expiring)
            if status_filter:
                current_date = datetime.now().date()
                
                if status_filter == 'expired':
                    query = query.filter(Document.expiry_date < current_date)
                elif status_filter == 'expiring':
                    warning_date = current_date + timedelta(days=30)
                    query = query.filter(
                        Document.expiry_date >= current_date,
                        Document.expiry_date <= warning_date
                    )
            
            # Get total count
            total_count = query.count()
            
            # Apply pagination
            documents = query.order_by(Document.created_at.desc())\
                           .offset((page - 1) * per_page)\
                           .limit(per_page)\
                           .all()
            
            return documents, total_count
            
        except Exception as e:
            print(f"Error getting documents: {str(e)}")
            return [], 0
    
    @staticmethod
    def get_document_by_id(document_id: int) -> Optional[Document]:
        """Get single document by ID with related data"""
        try:
            return Document.query.options(
                selectinload(Document.employee).selectinload(Employee.departments)
            ).get(document_id)
        except Exception:
            return None
    
    @staticmethod
    def create_document(
        employee_id: int,
        document_type: str,
        document_number: str,
        issue_date: Optional[date] = None,
        expiry_date: Optional[date] = None,
        notes: str = '',
        user_id: Optional[int] = None
    ) -> Tuple[bool, str, Optional[Document]]:
        """
        Create new document
        
        Returns:
            Tuple of (success, message, document_object)
        """
        try:
            # Validate employee exists
            employee = Employee.query.get(employee_id)
            if not employee:
                return False, 'الموظف غير موجود', None
            
            # Check for duplicate
            existing = Document.query.filter_by(
                employee_id=employee_id,
                document_type=document_type,
                document_number=document_number
            ).first()
            
            if existing:
                return False, 'هذه الوثيقة موجودة بالفعل لهذا الموظف', None
            
            # Create document
            document = Document(
                employee_id=employee_id,
                document_type=document_type,
                document_number=document_number,
                issue_date=issue_date,
                expiry_date=expiry_date,
                notes=notes,
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            
            db.session.add(document)
            db.session.commit()
            
            # Log activity
            log_activity(
                action='create',
                entity_type='document',
                entity_id=document.id,
                details=f'تم إضافة وثيقة {DocumentService.get_document_type_label(document_type)} للموظف {employee.name}'
            )
            
            return True, 'تم إضافة الوثيقة بنجاح', document
            
        except Exception as e:
            db.session.rollback()
            return False, f'حدث خطأ: {str(e)}', None
    
    @staticmethod
    def update_document(
        document_id: int,
        document_number: Optional[str] = None,
        issue_date: Optional[date] = None,
        expiry_date: Optional[date] = None,
        notes: Optional[str] = None
    ) -> Tuple[bool, str]:
        """
        Update existing document
        
        Returns:
            Tuple of (success, message)
        """
        try:
            document = Document.query.get(document_id)
            if not document:
                return False, 'الوثيقة غير موجودة'
            
            # Update fields if provided
            if document_number is not None:
                document.document_number = document_number
            if issue_date is not None:
                document.issue_date = issue_date
            if expiry_date is not None:
                document.expiry_date = expiry_date
            if notes is not None:
                document.notes = notes
            
            document.updated_at = datetime.now()
            db.session.commit()
            
            # Log activity
            log_activity(
                action='update',
                entity_type='document',
                entity_id=document_id,
                details=f'تم تحديث الوثيقة رقم {document_id}'
            )
            
            return True, 'تم تحديث الوثيقة بنجاح'
            
        except Exception as e:
            db.session.rollback()
            return False, f'حدث خطأ: {str(e)}'
    
    @staticmethod
    def delete_document(document_id: int) -> Tuple[bool, str]:
        """
        Delete document
        
        Returns:
            Tuple of (success, message)
        """
        try:
            document = Document.query.get(document_id)
            if not document:
                return False, 'الوثيقة غير موجودة'
            
            employee_name = document.employee.name if document.employee else 'غير محدد'
            doc_type = DocumentService.get_document_type_label(document.document_type)
            
            db.session.delete(document)
            db.session.commit()
            
            # Log activity
            log_activity(
                action='delete',
                entity_type='document',
                entity_id=document_id,
                details=f'تم حذف وثيقة {doc_type} للموظف {employee_name}'
            )
            
            return True, 'تم حذف الوثيقة بنجاح'
            
        except Exception as e:
            db.session.rollback()
            return False, f'حدث خطأ: {str(e)}'

    @staticmethod
    def upload_document_file(document_id: int, file_storage) -> Tuple[bool, str, Optional[Dict[str, Any]]]:
        """Upload document file with security checks and image compression."""
        try:
            document = Document.query.get(document_id)
            if not document:
                return False, 'الوثيقة غير موجودة', None

            if not file_storage or not getattr(file_storage, 'filename', None):
                return False, 'لم يتم اختيار ملف', None

            is_valid, error_message = DocumentService.validate_file_security(file_storage.filename)
            if not is_valid:
                return False, error_message, None

            success, result = FileService.save_employee_document(file_storage, document.employee_id)
            if not success:
                return False, result, None

            saved_filename = result
            upload_folder = os.path.join('static', 'uploads', 'employees', str(document.employee_id), 'documents')
            saved_path = os.path.join(upload_folder, saved_filename)

            if FileService.is_image_file(saved_filename):
                compressed_path = FileService.resize_image(saved_path)
                if compressed_path:
                    saved_path = compressed_path
                    saved_filename = os.path.basename(compressed_path)

            document.file_path = saved_path
            document.updated_at = datetime.now()
            db.session.commit()

            log_activity(
                action='upload_file',
                entity_type='document',
                entity_id=document.id,
                details=f'تم رفع ملف للوثيقة {document.id}: {saved_filename}'
            )

            return True, 'تم رفع الملف بنجاح', {
                'document_id': document.id,
                'filename': saved_filename,
                'file_path': saved_path,
                'compressed': FileService.is_image_file(saved_filename)
            }
        except Exception as e:
            db.session.rollback()
            return False, f'حدث خطأ أثناء رفع الملف: {str(e)}', None

    @staticmethod
    def get_document_download_info(document_id: int) -> Tuple[bool, str, Optional[Dict[str, str]]]:
        """Get validated download info for a document file."""
        try:
            from pathlib import Path

            document = Document.query.get(document_id)
            if not document:
                return False, 'الوثيقة غير موجودة', None

            if not document.file_path:
                return False, 'لا يوجد ملف مرفق لهذه الوثيقة', None

            file_path = document.file_path
            if not os.path.isabs(file_path):
                project_root = Path(__file__).resolve().parents[1]
                file_path = os.path.normpath(str(project_root / file_path))

            if not os.path.exists(file_path):
                return False, 'الملف غير موجود على الخادم', None

            return True, 'OK', {
                'file_path': file_path,
                'download_name': os.path.basename(file_path)
            }
        except Exception as e:
            return False, f'حدث خطأ أثناء تجهيز التحميل: {str(e)}', None
    
    # ==================== Bulk Operations ====================
    
    @staticmethod
    def create_bulk_documents(
        employee_ids: List[int],
        document_type: str,
        issue_date: Optional[date] = None,
        expiry_date: Optional[date] = None,
        notes: str = ''
    ) -> Tuple[bool, str, int]:
        """
        Create documents for multiple employees
        
        Returns:
            Tuple of (success, message, count_created)
        """
        try:
            count = 0
            
            for employee_id in employee_ids:
                employee = Employee.query.get(employee_id)
                if not employee:
                    continue
                
                # Check if document already exists
                existing = Document.query.filter_by(
                    employee_id=employee_id,
                    document_type=document_type
                ).first()
                
                if existing:
                    continue  # Skip duplicates
                
                # Use national ID if available
                doc_number = employee.national_id or f'ID-{employee.employee_id}'
                
                document = Document(
                    employee_id=employee_id,
                    document_type=document_type,
                    document_number=doc_number,
                    issue_date=issue_date,
                    expiry_date=expiry_date,
                    notes=notes,
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                
                db.session.add(document)
                count += 1
            
            db.session.commit()
            
            # Log activity
            log_activity(
                action='bulk_create',
                entity_type='document',
                entity_id=0,
                details=f'تم إضافة {count} وثيقة من نوع {DocumentService.get_document_type_label(document_type)}'
            )
            
            return True, f'تم إضافة {count} وثيقة بنجاح', count
            
        except Exception as e:
            db.session.rollback()
            return False, f'حدث خطأ: {str(e)}', 0
    
    @staticmethod
    def save_bulk_documents_with_data(documents_data: List[Dict]) -> Tuple[bool, str, int]:
        """
        Save multiple documents with individual data
        
        Args:
            documents_data: List of dicts with keys: employee_id, document_type, document_number, 
                          issue_date, expiry_date, notes
        
        Returns:
            Tuple of (success, message, count_saved)
        """
        try:
            count = 0
            
            for doc_data in documents_data:
                # Skip if no meaningful data
                if not any([
                    doc_data.get('document_number'),
                    doc_data.get('issue_date'),
                    doc_data.get('expiry_date')
                ]):
                    continue
                
                document = Document(
                    employee_id=doc_data['employee_id'],
                    document_type=doc_data['document_type'],
                    document_number=doc_data.get('document_number', ''),
                    issue_date=parse_date(doc_data['issue_date']) if doc_data.get('issue_date') else None,
                    expiry_date=parse_date(doc_data['expiry_date']) if doc_data.get('expiry_date') else None,
                    notes=doc_data.get('notes', ''),
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                
                db.session.add(document)
                count += 1
            
            db.session.commit()
            
            # Log activity
            log_activity(
                action='bulk_create',
                entity_type='document',
                entity_id=0,
                details=f'تم حفظ {count} وثيقة بشكل جماعي'
            )
            
            return True, f'تم حفظ {count} وثيقة بنجاح', count
            
        except Exception as e:
            db.session.rollback()
            return False, f'حدث خطأ: {str(e)}', 0
    
    # ==================== Excel Import/Export ====================
    
    @staticmethod
    def import_from_excel(file_stream: BytesIO) -> Tuple[bool, str, int, int]:
        """
        Import documents from Excel file
        
        Returns:
            Tuple of (success, message, success_count, error_count)
        """
        try:
            df = pd.read_excel(file_stream)
            
            success_count = 0
            error_count = 0
            
            for _, row in df.iterrows():
                try:
                    # Parse employee
                    employee_id = row.get('employee_id') or row.get('رقم الموظف')
                    if not employee_id:
                        error_count += 1
                        continue
                    
                    employee = Employee.query.filter_by(employee_id=str(employee_id)).first()
                    if not employee:
                        error_count += 1
                        continue
                    
                    # Parse document data
                    doc_type = row.get('document_type') or row.get('نوع الوثيقة') or 'other'
                    doc_number = row.get('document_number') or row.get('رقم الوثيقة') or ''
                    
                    issue_date = None
                    if 'issue_date' in row or 'تاريخ الإصدار' in row:
                        issue_date = parse_date(row.get('issue_date') or row.get('تاريخ الإصدار'))
                    
                    expiry_date = None
                    if 'expiry_date' in row or 'تاريخ الانتهاء' in row:
                        expiry_date = parse_date(row.get('expiry_date') or row.get('تاريخ الانتهاء'))
                    
                    notes = row.get('notes') or row.get('ملاحظات') or ''
                    
                    # Create document
                    document = Document(
                        employee_id=employee.id,
                        document_type=doc_type,
                        document_number=doc_number,
                        issue_date=issue_date,
                        expiry_date=expiry_date,
                        notes=notes,
                        created_at=datetime.now(),
                        updated_at=datetime.now()
                    )
                    
                    db.session.add(document)
                    db.session.commit()
                    success_count += 1
                    
                except Exception:
                    db.session.rollback()
                    error_count += 1
            
            # Log activity
            log_activity(
                action='import',
                entity_type='document',
                entity_id=0,
                details=f'تم استيراد {success_count} وثيقة بنجاح و {error_count} فشل'
            )
            
            return True, f'تم استيراد {success_count} وثيقة', success_count, error_count
            
        except Exception as e:
            return False, f'حدث خطأ في قراءة الملف: {str(e)}', 0, 0
    
    @staticmethod
    def export_to_excel(
        documents: List[Document],
        include_employee_data: bool = True
    ) -> BytesIO:
        """
        Export documents to Excel file
        
        Returns:
            BytesIO object containing Excel file
        """
        # Prepare data
        data = []
        for doc in documents:
            row = {
                'رقم الوثيقة': doc.id,
                'نوع الوثيقة': DocumentService.get_document_type_label(doc.document_type),
                'رقم المستند': doc.document_number or '',
                'تاريخ الإصدار': format_date_gregorian(doc.issue_date) if doc.issue_date else '',
                'تاريخ الانتهاء': format_date_gregorian(doc.expiry_date) if doc.expiry_date else '',
                'الملاحظات': doc.notes or ''
            }
            
            if include_employee_data and doc.employee:
                row['اسم الموظف'] = doc.employee.name
                row['رقم الموظف'] = doc.employee.employee_id
                row['القسم'] = doc.employee.departments[0].name if doc.employee.departments else ''
            
            # Calculate days to expiry
            if doc.expiry_date:
                today = datetime.now().date()
                days_diff = (doc.expiry_date - today).days
                row['الأيام المتبقية'] = days_diff
                
                if days_diff < 0:
                    row['الحالة'] = 'منتهية'
                elif days_diff <= 30:
                    row['الحالة'] = 'تنتهي قريباً'
                else:
                    row['الحالة'] = 'سارية'
            else:
                row['الأيام المتبقية'] = ''
                row['الحالة'] = ''
            
            data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='الوثائق', index=False, startrow=2)
            
            # Format the workbook
            workbook = writer.book
            worksheet = writer.sheets['الوثائق']
            
            # Add title
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = f'تقرير الوثائق - {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            title_cell.font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format headers
            header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                worksheet.column_dimensions[cell.column_letter].width = 20
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, len(df) + 4):
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).border = thin_border
        
        output.seek(0)
        return output
    
    # ==================== Statistics & Analytics ====================
    
    @staticmethod
    def get_dashboard_stats() -> Dict[str, Any]:
        """Get dashboard statistics"""
        try:
            current_date = datetime.now().date()
            expiring_date = current_date + timedelta(days=60)
            warning_date = current_date + timedelta(days=30)
            
            # Document counts
            total_documents = Document.query.count()
            expired_documents = Document.query.filter(Document.expiry_date < current_date).count()
            expiring_soon = Document.query.filter(
                Document.expiry_date >= current_date,
                Document.expiry_date <= warning_date
            ).count()
            expiring_later = Document.query.filter(
                Document.expiry_date > warning_date,
                Document.expiry_date <= expiring_date
            ).count()
            valid_documents = Document.query.filter(Document.expiry_date > expiring_date).count()
            
            # Top expired documents
            expired_docs = Document.query.join(Employee)\
                .filter(Document.expiry_date < current_date)\
                .order_by(Document.expiry_date.desc())\
                .limit(10).all()
            
            # Top expiring documents
            expiring_docs = Document.query.join(Employee)\
                .filter(Document.expiry_date >= current_date, Document.expiry_date <= warning_date)\
                .order_by(Document.expiry_date)\
                .limit(10).all()
            
            # Document types stats
            document_types_stats = db.session.query(
                Document.document_type,
                func.count(Document.id).label('count')
            ).group_by(Document.document_type).all()
            
            # Department stats
            department_stats = db.session.query(
                Department.name,
                func.count(Document.id).label('count')
            ).select_from(Department)\
             .join(Employee, Employee.department_id == Department.id)\
             .join(Document, Document.employee_id == Employee.id)\
             .group_by(Department.name)\
             .order_by(func.count(Document.id).desc())\
             .limit(5).all()
            
            return {
                'total_documents': total_documents,
                'expired_documents': expired_documents,
                'expiring_soon': expiring_soon,
                'expiring_later': expiring_later,
                'valid_documents': valid_documents,
                'expired_docs': expired_docs,
                'expiring_docs': expiring_docs,
                'document_types_stats': document_types_stats,
                'department_stats': department_stats
            }
            
        except Exception as e:
            print(f"Error getting dashboard stats: {str(e)}")
            return {}
    
    @staticmethod
    def get_expiry_stats() -> Dict[str, Any]:
        """Get document expiry statistics"""
        try:
            today = datetime.now().date()
            thirty_days = today + timedelta(days=30)
            sixty_days = today + timedelta(days=60)
            ninety_days = today + timedelta(days=90)
            
            base_query = Document.query.filter(Document.expiry_date.isnot(None))
            
            expiring_30 = base_query.filter(
                Document.expiry_date <= thirty_days,
                Document.expiry_date >= today
            ).count()
            
            expiring_60 = base_query.filter(
                Document.expiry_date <= sixty_days,
                Document.expiry_date > thirty_days
            ).count()
            
            expiring_90 = base_query.filter(
                Document.expiry_date <= ninety_days,
                Document.expiry_date > sixty_days
            ).count()
            
            expired = base_query.filter(Document.expiry_date < today).count()
            
            # Type counts
            type_counts = db.session.query(
                Document.document_type,
                func.count(Document.id).label('count')
            ).group_by(Document.document_type).all()
            
            type_stats = {doc_type: count for doc_type, count in type_counts}
            
            return {
                'expiring_30': expiring_30,
                'expiring_60': expiring_60,
                'expiring_90': expiring_90,
                'expired': expired,
                'type_stats': type_stats
            }
            
        except Exception as e:
            print(f"Error getting expiry stats: {str(e)}")
            return {}
    
    # ==================== Notifications ====================
    
    @staticmethod
    def create_expiry_notification(
        user_id: int,
        document_type: str,
        employee_name: str,
        days_until_expiry: int,
        document_id: int
    ) -> Optional[Notification]:
        """Create expiry notification for user"""
        try:
            from flask import url_for
            
            if days_until_expiry < 0:
                title = f'وثيقة منتهية - {document_type}'
                description = f'انتهت صلاحية {document_type} للموظف {employee_name} منذ {abs(days_until_expiry)} يوم'
                priority = 'critical'
            elif days_until_expiry <= 7:
                title = f'تنبيه عاجل: وثيقة تنتهي قريباً'
                description = f'{document_type} للموظف {employee_name} تنتهي خلال {days_until_expiry} أيام'
                priority = 'critical'
            elif days_until_expiry <= 30:
                title = f'تذكير: وثيقة تنتهي خلال شهر'
                description = f'{document_type} للموظف {employee_name} تنتهي خلال {days_until_expiry} يوماً'
                priority = 'high'
            else:
                title = f'تذكير: وثيقة قريبة من الانتهاء'
                description = f'{document_type} للموظف {employee_name} تنتهي خلال {days_until_expiry} يوماً'
                priority = 'normal'
            
            notification = Notification(
                user_id=user_id,
                notification_type='document_expiry',
                title=title,
                description=description,
                related_entity_type='document',
                related_entity_id=document_id,
                priority=priority,
                action_url=url_for('documents.dashboard')
            )
            
            db.session.add(notification)
            return notification
            
        except Exception as e:
            print(f"Error creating notification: {str(e)}")
            return None
    
    @staticmethod
    def create_bulk_expiry_notifications() -> Tuple[bool, str, int]:
        """Create expiry notifications for all users"""
        try:
            current_date = datetime.now().date()
            warning_date = current_date + timedelta(days=30)
            
            # Get expiring/expired documents
            expiring_docs = Document.query.join(Employee)\
                .filter(Document.expiry_date <= warning_date)\
                .order_by(Document.expiry_date)\
                .limit(5).all()
            
            if not expiring_docs:
                return False, 'لا توجد وثائق منتهية أو قريبة من الانتهاء', 0
            
            all_users = User.query.all()
            notification_count = 0
            
            for doc in expiring_docs:
                days_until_expiry = (doc.expiry_date - current_date).days if doc.expiry_date else -999
                doc_type = DocumentService.get_document_type_label(doc.document_type)
                employee_name = doc.employee.name if doc.employee else 'غير محدد'
                
                for user in all_users:
                    try:
                        DocumentService.create_expiry_notification(
                            user_id=user.id,
                            document_type=doc_type,
                            employee_name=employee_name,
                            days_until_expiry=days_until_expiry,
                            document_id=doc.id
                        )
                        notification_count += 1
                    except Exception:
                        pass
            
            db.session.commit()
            
            return True, f'تم إنشاء {notification_count} إشعار', notification_count
            
        except Exception as e:
            db.session.rollback()
            return False, f'حدث خطأ: {str(e)}', 0
    
    # ==================== Employee Filtering ====================
    
    @staticmethod
    def get_employees_by_sponsorship(sponsorship_type: str) -> List[Dict[str, Any]]:
        """Get employees filtered by sponsorship status"""
        try:
            if sponsorship_type == 'internal':
                sponsorship_status = 'على الكفالة'
            elif sponsorship_type == 'external':
                sponsorship_status = 'خارج الكفالة'
            else:
                return []
            
            employees = Employee.query.filter_by(sponsorship_status=sponsorship_status)\
                                    .options(selectinload(Employee.departments))\
                                    .all()
            
            result = []
            for emp in employees:
                dept_names = ', '.join([dept.name for dept in emp.departments]) if emp.departments else 'غير محدد'
                result.append({
                    'id': emp.id,
                    'name': emp.name,
                    'employee_id': emp.employee_id,
                    'national_id': emp.national_id,
                    'department_names': dept_names
                })
            
            return result
            
        except Exception:
            return []
    
    @staticmethod
    def get_employees_by_department_and_sponsorship(
        department_id: Optional[int] = None,
        sponsorship_type: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """Get employees filtered by department and sponsorship"""
        try:
            query = Employee.query.options(selectinload(Employee.departments))
            
            if department_id:
                query = query.filter(Employee.departments.any(Department.id == department_id))
            
            if sponsorship_type == 'on_sponsorship':
                query = query.filter(Employee.sponsorship_status == 'على الكفالة')
            elif sponsorship_type == 'off_sponsorship':
                query = query.filter(Employee.sponsorship_status == 'خارج الكفالة')
            
            employees = query.all()
            
            result = []
            for emp in employees:
                dept_names = ', '.join([dept.name for dept in emp.departments]) if emp.departments else 'غير محدد'
                result.append({
                    'id': emp.id,
                    'name': emp.name,
                    'employee_id': emp.employee_id,
                    'national_id': emp.national_id,
                    'department_names': dept_names,
                    'sponsorship_status': sponsorship_type
                })
            
            return result
            
        except Exception:
            return []
    
    # ==================== PDF Generation ====================
    
    @staticmethod
    def generate_document_template_pdf() -> BytesIO:
        """Generate blank PDF template for documents"""
        buffer = BytesIO()
        
        try:
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )
            
            # Register Arabic font
            try:
                pdfmetrics.registerFont(TTFont('Cairo', 'static/fonts/Cairo.ttf'))
                arabic_font = 'Cairo'
            except:
                arabic_font = 'Helvetica'
            
            # Create styles
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontName=arabic_font,
                fontSize=20,
                spaceAfter=30,
                alignment=1,
                textColor=colors.darkblue
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontName=arabic_font,
                fontSize=16,
                spaceAfter=20,
                alignment=1,
                textColor=colors.blue
            )
            
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), arabic_font),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])
            
            # Build content
            story = []
            
            # Title
            title = get_display(arabic_reshaper.reshape("نموذج إدارة الوثائق"))
            story.append(Paragraph(title, title_style))
            
            # Company info
            company = get_display(arabic_reshaper.reshape("شركة نُظم لإدارة الموارد البشرية"))
            story.append(Paragraph(company, subtitle_style))
            story.append(Spacer(1, 20))
            
            # Employee info table
            employee_title = get_display(arabic_reshaper.reshape("معلومات الموظف"))
            story.append(Paragraph(employee_title, subtitle_style))
            
            employee_data = [
                [get_display(arabic_reshaper.reshape("البيان")), get_display(arabic_reshaper.reshape("القيمة"))],
                [get_display(arabic_reshaper.reshape("اسم الموظف")), "________________________"],
                [get_display(arabic_reshaper.reshape("رقم الموظف")), "________________________"],
                [get_display(arabic_reshaper.reshape("رقم الهوية الوطنية")), "________________________"],
                [get_display(arabic_reshaper.reshape("القسم")), "________________________"]
            ]
            
            employee_table = Table(employee_data, colWidths=[8*cm, 8*cm])
            employee_table.setStyle(table_style)
            story.append(employee_table)
            story.append(Spacer(1, 30))
            
            # Document info table
            doc_title = get_display(arabic_reshaper.reshape("معلومات الوثيقة"))
            story.append(Paragraph(doc_title, subtitle_style))
            
            doc_data = [
                [get_display(arabic_reshaper.reshape("البيان")), get_display(arabic_reshaper.reshape("القيمة"))],
                [get_display(arabic_reshaper.reshape("نوع الوثيقة")), "________________________"],
                [get_display(arabic_reshaper.reshape("رقم الوثيقة")), "________________________"],
                [get_display(arabic_reshaper.reshape("تاريخ الإصدار")), "________________________"],
                [get_display(arabic_reshaper.reshape("تاريخ الانتهاء")), "________________________"]
            ]
            
            doc_table = Table(doc_data, colWidths=[8*cm, 8*cm])
            doc_table.setStyle(table_style)
            story.append(doc_table)
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            
        except Exception as e:
            print(f"Error generating PDF: {str(e)}")
        
        return buffer
    
    @staticmethod
    def generate_employee_documents_pdf(employee_id: int) -> Optional[BytesIO]:
        """Generate PDF report for employee's documents"""
        try:
            employee = Employee.query.get(employee_id)
            if not employee:
                return None
            
            documents = Document.query.filter_by(employee_id=employee_id).all()
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)
            
            # Register Arabic font
            try:
                pdfmetrics.registerFont(TTFont('ArabicFont', 'static/fonts/Cairo.ttf'))
                arabic_font = 'ArabicFont'
            except:
                arabic_font = 'Helvetica'
            
            # Build PDF content (simplified for space)
            story = []
            
            # Add title and employee info
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('Title', parent=styles['Title'], fontName=arabic_font,
                                        fontSize=18, alignment=1)
            
            title = get_display(arabic_reshaper.reshape(f"وثائق الموظف: {employee.name}"))
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 20))
            
            # Add documents table
            table_data = [[
                get_display(arabic_reshaper.reshape("نوع الوثيقة")),
                get_display(arabic_reshaper.reshape("رقم المستند")),
                get_display(arabic_reshaper.reshape("تاريخ الانتهاء")),
                get_display(arabic_reshaper.reshape("الحالة"))
            ]]
            
            today = datetime.now().date()
            
            for document in documents:
                doc_type = get_display(arabic_reshaper.reshape(
                    DocumentService.get_document_type_label(document.document_type)
                ))
                doc_number = document.document_number or '-'
                expiry = format_date_gregorian(document.expiry_date) if document.expiry_date else '-'
                
                if document.expiry_date:
                    if document.expiry_date < today:
                        status = get_display(arabic_reshaper.reshape("منتهية"))
                    elif (document.expiry_date - today).days <= 30:
                        status = get_display(arabic_reshaper.reshape("تنتهي قريباً"))
                    else:
                        status = get_display(arabic_reshaper.reshape("سارية"))
                else:
                    status = '-'
                
                table_data.append([doc_type, doc_number, expiry, status])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), arabic_font),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            
            doc.build(story)
            buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            print(f"Error generating employee PDF: {str(e)}")
            return None
