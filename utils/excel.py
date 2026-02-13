import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, timedelta
from utils.date_converter import parse_date, format_date_gregorian, format_date_hijri
from calendar import monthrange
import xlsxwriter

def parse_employee_excel(file):
    """
    Parse Excel file containing employee data
    
    Args:
        file: The uploaded Excel file
        
    Returns:
        List of dictionaries containing employee data
    """
    try:
        # Reset file pointer to beginning
        file.seek(0)
        
        # Read the Excel file explicitly using openpyxl engine
        df = pd.read_excel(file, engine='openpyxl')
        
        # Debug: Print column names
        print(f"Excel columns: {df.columns.tolist()}")
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        # Check if DataFrame is empty
        if df.empty:
            raise ValueError("Excel file is empty or has no data")
        
        # Create a more flexible column mapping
        column_mappings = {
            'name': ['name', 'الاسم الكامل', 'اسم', 'الاسم', 'full name', 'employee name', 'Name'],
            'employee_id': ['رقم الموظف', 'employee_id', 'emp_id', 'emp id', 'Emp .N', 'Emp.N', 'EmpN'],
            'national_id': ['رقم الهوية الوطنية', 'national_id', 'id', 'ID .N', 'ID Number', 'هوية'],
            'mobile': ['رقم الجوال', 'mobile', 'phone', 'هاتف', 'جوال', 'No.Mobile', 'Mobil'],
            'job_title': ['المسمى الوظيفي', 'job_title', 'position', 'title', 'Job Title', 'وظيفة'],
            'status': ['الحالة الوظيفية', 'status', 'حالة', 'Status'],
            'location': ['الموقع', 'location', 'موقع', 'Location'],
            'project': ['المشروع', 'project', 'مشروع', 'Project'],
            'email': ['البريد الإلكتروني', 'email', 'بريد', 'Email'],
            'department': ['الأقسام', 'department', 'قسم', 'Department'],
            'join_date': ['تاريخ الانضمام', 'join_date', 'hire_date', 'انضمام'],
            'license_end_date': ['تاريخ انتهاء الإقامة', 'license_end_date', 'انتهاء الإقامة'],
            'contract_status': ['حالة العقد', 'contract_status', 'عقد'],
            'license_status': ['حالة الرخصة', 'license_status', 'رخصة'],
            'nationality': ['الجنسية', 'nationality', 'جنسية'],
            'notes': ['ملاحظات', 'notes', 'remarks', 'comments'],
            'mobilePersonal': ['الجوال الشخصي', 'mobile_personal', 'جوال شخصي']
        }
        
        # Map columns to their field names
        detected_columns = {}
        for col in df.columns:
            if isinstance(col, datetime):
                continue
                
            col_str = str(col).strip()
            
            # Check for matches in column mappings
            for field, variations in column_mappings.items():
                if col_str in variations:
                    detected_columns[field] = col
                    print(f"Detected '{field}' column: {col}")
                    break
        
        # If no columns detected, try to guess from position and content
        if not detected_columns:
            print("No columns detected by name, trying to guess from position...")
            columns_list = [col for col in df.columns if not isinstance(col, datetime)]
            
            # If we have enough columns, try to guess based on position
            if len(columns_list) >= 3:
                # Basic required fields
                detected_columns['name'] = columns_list[0]
                detected_columns['employee_id'] = columns_list[1] if len(columns_list) > 1 else None
                detected_columns['national_id'] = columns_list[2] if len(columns_list) > 2 else None
                
                # Optional fields
                if len(columns_list) > 3:
                    detected_columns['mobile'] = columns_list[3]
                if len(columns_list) > 4:
                    detected_columns['job_title'] = columns_list[4]
                
                print(f"Guessed columns: {detected_columns}")
        
        # Check for minimum required columns
        required_fields = ['name']
        missing_required = [field for field in required_fields if field not in detected_columns]
        
        if missing_required:
            raise ValueError(f"Required columns missing: {', '.join(missing_required)}. Available columns: {[c for c in df.columns if not isinstance(c, datetime)]}")
        
        # Process each row
        employees = []
        for idx, row in df.iterrows():
            try:
                # Skip completely empty rows
                if row.isnull().all():
                    continue
                
                # Check if name is present
                name_col = detected_columns.get('name')
                if name_col and pd.isna(row[name_col]):
                    continue
                
                # Create employee dictionary
                employee = {}
                
                # Add name (required)
                if name_col:
                    employee['name'] = str(row[name_col]).strip()
                
                # Add employee_id (auto-generate if missing)
                emp_id_col = detected_columns.get('employee_id')
                if emp_id_col and not pd.isna(row[emp_id_col]):
                    employee['employee_id'] = str(row[emp_id_col]).strip()
                else:
                    employee['employee_id'] = f"EMP{idx+1000}"
                
                # Add national_id (auto-generate if missing)
                national_id_col = detected_columns.get('national_id')
                if national_id_col and not pd.isna(row[national_id_col]):
                    employee['national_id'] = str(row[national_id_col]).strip()
                else:
                    employee['national_id'] = f"N{idx+5000:07d}"
                
                # Add mobile (auto-generate if missing)
                mobile_col = detected_columns.get('mobile')
                if mobile_col and not pd.isna(row[mobile_col]):
                    employee['mobile'] = str(row[mobile_col]).strip()
                else:
                    employee['mobile'] = f"05xxxxxxxx"
                
                # Add job_title (default if missing)
                job_title_col = detected_columns.get('job_title')
                if job_title_col and not pd.isna(row[job_title_col]):
                    employee['job_title'] = str(row[job_title_col]).strip()
                else:
                    employee['job_title'] = "موظف"
                
                # Add status (default to active)
                status_col = detected_columns.get('status')
                if status_col and not pd.isna(row[status_col]):
                    status_value = str(row[status_col]).lower().strip()
                    if status_value in ['active', 'نشط', 'فعال']:
                        employee['status'] = 'active'
                    elif status_value in ['inactive', 'غير نشط', 'غير فعال']:
                        employee['status'] = 'inactive'
                    elif status_value in ['on_leave', 'on leave', 'leave', 'إجازة', 'في إجازة']:
                        employee['status'] = 'on_leave'
                    else:
                        employee['status'] = 'active'
                else:
                    employee['status'] = 'active'
                
                # Add optional fields (excluding department which is handled separately)
                optional_fields = ['location', 'project', 'email', 'join_date', 
                                 'license_end_date', 'contract_status', 'license_status', 
                                 'nationality', 'notes', 'mobilePersonal']
                
                for field in optional_fields:
                    col = detected_columns.get(field)
                    if col and not pd.isna(row[col]):
                        employee[field] = str(row[col]).strip()
                
                # Handle department separately
                dept_col = detected_columns.get('department')
                if dept_col and not pd.isna(row[dept_col]):
                    employee['department'] = str(row[dept_col]).strip()
                
                # Debug: Print processed employee
                print(f"Processed employee {idx+1}: {employee.get('name', 'Unknown')}")
                
                employees.append(employee)
                
            except Exception as e:
                print(f"Error processing row {idx+1}: {str(e)}")
                continue
        
        if not employees:
            raise ValueError("No valid employee records found in the Excel file")
            
        print(f"Successfully parsed {len(employees)} employee records")
        return employees
    
    except Exception as e:
        import traceback
        print(f"Error parsing Excel: {str(e)}")
        print(traceback.format_exc())
        raise Exception(f"Error parsing Excel file: {str(e)}")

def export_employees_to_excel(employees, output=None):
    """
    Export employees to Excel file
    
    Args:
        employees: List of Employee objects
        output: BytesIO object to write to (optional)
        
    Returns:
        BytesIO object containing the Excel file
    """
    return generate_employee_excel(employees, output)
    
def generate_employee_excel(employees, output=None):
    """
    Generate Professional Excel file from employee data with Dashboard
    
    Args:
        employees: List of Employee objects
        output: BytesIO object to write to (optional)
        
    Returns:
        BytesIO object containing the Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter
        from collections import Counter
        
        if output is None:
            output = BytesIO()
        
        workbook = Workbook()
        
        # ===== ورقة Dashboard الاحترافية =====
        dashboard = workbook.active
        dashboard.title = "Dashboard"
        
        # حساب الإحصائيات
        total_employees = len(employees)
        active_employees = sum(1 for e in employees if e.status == 'active')
        inactive_employees = sum(1 for e in employees if e.status == 'inactive')
        on_leave_employees = sum(1 for e in employees if e.status == 'on_leave')
        
        # إحصائيات الأقسام
        dept_counter = Counter()
        for emp in employees:
            if emp.departments:
                for dept in emp.departments:
                    dept_counter[dept.name] += 1
            else:
                dept_counter['بدون قسم'] += 1
        
        # إحصائيات الوظائف
        job_counter = Counter(e.job_title for e in employees if e.job_title)
        
        # تنسيقات Dashboard
        title_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=16, name='Calibri')
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
        stat_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # عنوان Dashboard
        dashboard.merge_cells('A1:F1')
        title_cell = dashboard['A1']
        title_cell.value = f"تقرير الموظفين - لوحة التحكم | {datetime.now().strftime('%Y-%m-%d')}"
        title_cell.font = title_fill = Font(bold=True, color="FFFFFF", size=18, name='Calibri')
        title_cell.alignment = center_align
        title_cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        
        # كارت الإحصائيات الرئيسية
        dashboard.merge_cells('A3:B3')
        dashboard['A3'].value = "📊 إحصائيات عامة"
        dashboard['A3'].font = header_font
        dashboard['A3'].fill = header_fill
        dashboard['A3'].alignment = center_align
        
        stats_data = [
            ("إجمالي الموظفين", total_employees, "4472C4"),
            ("الموظفون النشطون", active_employees, "70AD47"),
            ("الموظفون غير النشطين", inactive_employees, "FFC000"),
            ("في إجازة", on_leave_employees, "ED7D31")
        ]
        
        row = 4
        for label, value, color in stats_data:
            dashboard.cell(row=row, column=1).value = label
            dashboard.cell(row=row, column=1).font = Font(bold=True, size=11)
            dashboard.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            
            dashboard.cell(row=row, column=2).value = value
            dashboard.cell(row=row, column=2).font = Font(bold=True, size=14, color=color)
            dashboard.cell(row=row, column=2).alignment = center_align
            dashboard.cell(row=row, column=2).fill = stat_fill
            row += 1
        
        # إحصائيات الأقسام
        dashboard.merge_cells('D3:F3')
        dashboard['D3'].value = "🏢 توزيع الموظفين حسب الأقسام"
        dashboard['D3'].font = header_font
        dashboard['D3'].fill = header_fill
        dashboard['D3'].alignment = center_align
        
        row = 4
        for dept_name, count in dept_counter.most_common(10):
            dashboard.cell(row=row, column=4).value = dept_name
            dashboard.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            dashboard.cell(row=row, column=5).value = count
            dashboard.cell(row=row, column=5).alignment = center_align
            dashboard.cell(row=row, column=5).fill = stat_fill
            
            # نسبة مئوية
            percentage = f"{(count/total_employees*100):.1f}%"
            dashboard.cell(row=row, column=6).value = percentage
            dashboard.cell(row=row, column=6).alignment = center_align
            dashboard.cell(row=row, column=6).font = Font(bold=True, color="4472C4")
            row += 1
        
        # أكثر الوظائف
        dashboard.merge_cells('A10:B10')
        dashboard['A10'].value = "💼 أكثر الوظائف شيوعاً"
        dashboard['A10'].font = header_font
        dashboard['A10'].fill = header_fill
        dashboard['A10'].alignment = center_align
        
        row = 11
        for job_title, count in job_counter.most_common(8):
            dashboard.cell(row=row, column=1).value = job_title
            dashboard.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            dashboard.cell(row=row, column=2).value = count
            dashboard.cell(row=row, column=2).alignment = center_align
            dashboard.cell(row=row, column=2).fill = stat_fill
            row += 1
        
        # ضبط عرض الأعمدة في Dashboard
        dashboard.column_dimensions['A'].width = 25
        dashboard.column_dimensions['B'].width = 15
        dashboard.column_dimensions['C'].width = 3
        dashboard.column_dimensions['D'].width = 25
        dashboard.column_dimensions['E'].width = 12
        dashboard.column_dimensions['F'].width = 12
        
        # ===== ورقة بيانات الموظفين الاحترافية =====
        employees_sheet = workbook.create_sheet(title="Employee Data")
        
        # العناوين الأساسية
        headers = [
            "الاسم الكامل", "رقم الموظف", "رقم الهوية", "رقم الجوال", 
            "المسمى الوظيفي", "الحالة", "الموقع", "المشروع", 
            "الأقسام", "البريد الإلكتروني", "تاريخ الانضمام"
        ]
        
        # كتابة العناوين
        for col_idx, header in enumerate(headers, start=1):
            cell = employees_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = center_align
            cell.border = border
        
        # كتابة البيانات
        for row_idx, employee in enumerate(employees, start=2):
            # البيانات الأساسية
            data_row = [
                employee.name,
                employee.employee_id,
                employee.national_id or "",
                employee.mobile or "",
                employee.job_title or "",
                employee.status or "",
                employee.location or "",
                employee.project or "",
                ', '.join([dept.name for dept in employee.departments]) if employee.departments else "",
                employee.email or "",
                employee.join_date.strftime('%Y-%m-%d') if employee.join_date else ""
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = employees_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = center_align
                cell.border = border
                
                # تلوين الصفوف بالتناوب
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # تلوين الحالة
                if col_idx == 6:  # عمود الحالة
                    if value == 'active':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif value == 'inactive':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")
                    elif value == 'on_leave':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True, color="9C6500")
        
        # ضبط عرض الأعمدة
        column_widths = [30, 15, 18, 16, 20, 12, 18, 18, 25, 25, 18]
        for idx, width in enumerate(column_widths, start=1):
            employees_sheet.column_dimensions[get_column_letter(idx)].width = width
        
        # ===== ورقة البيانات الكاملة =====
        full_data_sheet = workbook.create_sheet(title="Complete Data")
        
        # جميع البيانات بالترتيب المطلوب
        all_headers = [
            'الاسم الكامل', 'رقم الهوية الوطنية', 'رقم الموظف', 'الجنسية',
            'الجوال الشخصي', 'مقاس البنطلون', 'مقاس التيشرت', 'المسمى الوظيفي',
            'الحالة الوظيفية', 'نوع الجوال', 'رقم IMEI', 'رقم الجوال',
            'السيارة الحاليه', 'الموقع', 'المشروع', 'البريد الإلكتروني',
            'الأقسام', 'تاريخ الانضمام', 'تاريخ الميلاد', 'نوع الموظف',
            'نوع العقد', 'الراتب الأساسي', 'حالة العقد', 'حالة الرخصة',
            'حالة الكفالة', 'اسم الكفيل', 'رقم الإيبان', 'تفاصيل السكن', 'رابط موقع السكن',
            'ملف العرض الوظيفي', 'صورة الجواز', 'شهادة العنوان الوطني',
            'رابط العرض الوظيفي الخارجي', 'رابط صورة الجواز الخارجي', 'رابط شهادة العنوان الخارجي'
        ]
        
        for col_idx, header in enumerate(all_headers, start=1):
            cell = full_data_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = center_align
            cell.border = border
        
        for row_idx, employee in enumerate(employees, start=2):
            # جلب السيارة الحالية (المسلمة للموظف) من نظام إدارة السيارات
            current_vehicle = ""
            try:
                from models import VehicleHandover
                # البحث عن آخر سجل تسليم (delivery) للموظف
                latest_delivery = VehicleHandover.query.filter_by(
                    employee_id=employee.id
                ).filter(
                    VehicleHandover.handover_type.in_(['delivery', 'تسليم', 'handover'])
                ).order_by(VehicleHandover.handover_date.desc()).first()
                
                if latest_delivery and latest_delivery.vehicle:
                    current_vehicle = f"{latest_delivery.vehicle.plate_number}"
            except:
                pass
            
            # جلب بيانات الجهاز المحمول من نظام إدارة الأجهزة المحمولة
            mobile_type = ""
            mobile_imei = ""
            mobile_number = ""
            try:
                from models import DeviceAssignment, MobileDevice, SimCard
                
                # البحث عن التعيين النشط للموظف
                active_assignment = DeviceAssignment.query.filter_by(
                    employee_id=employee.id,
                    is_active=True
                ).first()
                
                print(f"🔍 الموظف {employee.name} (ID: {employee.id}): التعيين النشط = {active_assignment}")
                
                if active_assignment:
                    # جلب معلومات الجهاز
                    if active_assignment.device_id:
                        device = MobileDevice.query.get(active_assignment.device_id)
                        print(f"   📱 الجهاز: {device}")
                        if device:
                            mobile_type = f"{device.device_brand or ''} {device.device_model or ''}".strip()
                            mobile_imei = device.imei or ""
                            print(f"   ✅ نوع الجوال: {mobile_type}, IMEI: {mobile_imei}")
                    
                    # جلب رقم الهاتف من SIM Card
                    if active_assignment.sim_card_id:
                        sim = SimCard.query.get(active_assignment.sim_card_id)
                        print(f"   📞 SIM Card: {sim}")
                        if sim:
                            mobile_number = sim.phone_number or ""
                            print(f"   ✅ رقم الجوال: {mobile_number}")
                else:
                    print(f"   ❌ لا يوجد تعيين نشط للموظف")
                    
            except Exception as e:
                import traceback
                print(f"❌ خطأ في جلب بيانات الجهاز للموظف {employee.name} ({employee.id}): {str(e)}")
                print(traceback.format_exc())
            
            # إضافة روابط الملفات الجديدة
            job_offer_link = f"https://nuzum.site/static/{getattr(employee, 'job_offer_file', '')}" if getattr(employee, 'job_offer_file', '') else '-'
            passport_link = f"https://nuzum.site/static/{getattr(employee, 'passport_image_file', '')}" if getattr(employee, 'passport_image_file', '') else '-'
            national_address_link = f"https://nuzum.site/static/{getattr(employee, 'national_address_file', '')}" if getattr(employee, 'national_address_file', '') else '-'
            
            # إضافة الروابط الخارجية
            job_offer_external = getattr(employee, 'job_offer_link', '') or '-'
            passport_external = getattr(employee, 'passport_image_link', '') or '-'
            national_address_external = getattr(employee, 'national_address_link', '') or '-'
            
            all_data = [
                employee.name,  # 1. الاسم الكامل
                employee.national_id or "",  # 2. رقم الهوية الوطنية
                employee.employee_id,  # 3. رقم الموظف
                employee.nationality_rel.name_ar if hasattr(employee, 'nationality_rel') and employee.nationality_rel else (employee.nationality if hasattr(employee, 'nationality') else ""),  # 4. الجنسية
                getattr(employee, 'mobilePersonal', '') or '',  # 5. الجوال الشخصي
                getattr(employee, 'pants_size', '') or '',  # 6. مقاس البنطلون
                getattr(employee, 'shirt_size', '') or '',  # 7. مقاس التيشرت
                employee.job_title or "",  # 8. المسمى الوظيفي
                employee.status or "",  # 9. الحالة الوظيفية
                mobile_type,  # 10. نوع الجوال (من نظام إدارة الأجهزة)
                mobile_imei,  # 11. رقم IMEI (من نظام إدارة الأجهزة)
                mobile_number,  # 12. رقم الجوال (من نظام إدارة الأجهزة)
                current_vehicle,  # 13. السيارة الحالية (من نظام إدارة السيارات)
                employee.location or "",  # 14. الموقع
                employee.project or "",  # 15. المشروع
                employee.email or "",  # 16. البريد الإلكتروني
                ', '.join([dept.name for dept in employee.departments]) if employee.departments else "",  # 17. الأقسام
                employee.join_date.strftime('%Y-%m-%d') if employee.join_date else "",  # 18. تاريخ الانضمام
                employee.birth_date.strftime('%Y-%m-%d') if employee.birth_date else "",  # 19. تاريخ الميلاد
                getattr(employee, 'employee_type', '') or '',  # 20. نوع الموظف
                getattr(employee, 'contract_type', '') or '',  # 21. نوع العقد
                str(getattr(employee, 'basic_salary', '') or ''),  # 22. الراتب الأساسي
                getattr(employee, 'contract_status', '') or '',  # 23. حالة العقد
                getattr(employee, 'license_status', '') or '',  # 24. حالة الرخصة
                getattr(employee, 'sponsorship_status', '') or '',  # 25. حالة الكفالة
                getattr(employee, 'current_sponsor_name', '') or '',  # 26. اسم الكفيل
                getattr(employee, 'bank_iban', '') or '',  # 27. رقم الإيبان
                getattr(employee, 'residence_details', '') or '',  # 28. تفاصيل السكن
                getattr(employee, 'residence_location_url', '') or '',  # 29. رابط موقع السكن
                job_offer_link,  # 30. ملف العرض الوظيفي
                passport_link,  # 31. صورة الجواز
                national_address_link,  # 32. شهادة العنوان الوطني
                job_offer_external,  # 33. رابط العرض الوظيفي الخارجي
                passport_external,  # 34. رابط صورة الجواز الخارجي
                national_address_external  # 35. رابط شهادة العنوان الخارجي
            ]
            
            for col_idx, value in enumerate(all_data, start=1):
                cell = full_data_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = center_align
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # ضبط عرض الأعمدة
        for idx in range(1, len(all_headers) + 1):
            full_data_sheet.column_dimensions[get_column_letter(idx)].width = 18
        
        # حفظ الملف
        workbook.save(output)
        output.seek(0)
        return output
    
    except Exception as e:
        print(f"خطأ في إنشاء ملف Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise Exception(f"Error generating Excel file: {str(e)}")

def generate_vehicles_excel(vehicles, output=None):
    """
    Generate Professional Excel file from vehicles data with Dashboard
    
    Args:
        vehicles: List of Vehicle objects
        output: BytesIO object to write to (optional)
        
    Returns:
        BytesIO object containing the Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from collections import Counter
        
        if output is None:
            output = BytesIO()
        
        workbook = Workbook()
        
        # ===== ورقة Dashboard الاحترافية =====
        dashboard = workbook.active
        dashboard.title = "Dashboard"
        
        # حساب الإحصائيات
        total_vehicles = len(vehicles)
        
        # إحصائيات الحالة
        status_map = {
            'available': 'متاحة',
            'rented': 'مؤجرة',
            'in_workshop': 'في الورشة',
            'in_project': 'في المشروع',
            'accident': 'حادث',
            'sold': 'مباعة'
        }
        status_counter = Counter()
        for v in vehicles:
            status_ar = status_map.get(v.status, v.status)
            status_counter[status_ar] += 1
        
        # إحصائيات الشركة المصنعة
        make_counter = Counter(v.make for v in vehicles if v.make)
        
        # إحصائيات الموديل
        model_counter = Counter(v.model for v in vehicles if v.model)
        
        # إحصائيات السنوات
        year_counter = Counter(str(v.year) for v in vehicles if v.year)
        
        # تنسيقات Dashboard
        title_font = Font(bold=True, color="FFFFFF", size=18, name='Calibri')
        title_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
        stat_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # عنوان Dashboard
        dashboard.merge_cells('A1:F1')
        title_cell = dashboard['A1']
        title_cell.value = f"تقرير المركبات - لوحة التحكم | {datetime.now().strftime('%Y-%m-%d')}"
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = title_fill
        
        # كارت الإحصائيات الرئيسية
        dashboard.merge_cells('A3:B3')
        dashboard['A3'].value = "🚗 إحصائيات عامة"
        dashboard['A3'].font = header_font
        dashboard['A3'].fill = header_fill
        dashboard['A3'].alignment = center_align
        
        # الإحصائيات حسب الحالة
        row = 4
        dashboard.cell(row=row, column=1).value = "إجمالي المركبات"
        dashboard.cell(row=row, column=1).font = Font(bold=True, size=11)
        dashboard.cell(row=row, column=1).alignment = Alignment(horizontal='right')
        dashboard.cell(row=row, column=2).value = total_vehicles
        dashboard.cell(row=row, column=2).font = Font(bold=True, size=14, color="4472C4")
        dashboard.cell(row=row, column=2).alignment = center_align
        dashboard.cell(row=row, column=2).fill = stat_fill
        row += 1
        
        for status_ar, count in status_counter.most_common():
            dashboard.cell(row=row, column=1).value = status_ar
            dashboard.cell(row=row, column=1).font = Font(bold=True, size=11)
            dashboard.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            dashboard.cell(row=row, column=2).value = count
            dashboard.cell(row=row, column=2).font = Font(bold=True, size=12)
            dashboard.cell(row=row, column=2).alignment = center_align
            dashboard.cell(row=row, column=2).fill = stat_fill
            row += 1
        
        # إحصائيات الشركات المصنعة
        dashboard.merge_cells('D3:F3')
        dashboard['D3'].value = "🏭 توزيع المركبات حسب الشركة المصنعة"
        dashboard['D3'].font = header_font
        dashboard['D3'].fill = header_fill
        dashboard['D3'].alignment = center_align
        
        row = 4
        for make, count in make_counter.most_common(10):
            dashboard.cell(row=row, column=4).value = make
            dashboard.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            dashboard.cell(row=row, column=5).value = count
            dashboard.cell(row=row, column=5).alignment = center_align
            dashboard.cell(row=row, column=5).fill = stat_fill
            
            # نسبة مئوية
            percentage = f"{(count/total_vehicles*100):.1f}%"
            dashboard.cell(row=row, column=6).value = percentage
            dashboard.cell(row=row, column=6).alignment = center_align
            dashboard.cell(row=row, column=6).font = Font(bold=True, color="4472C4")
            row += 1
        
        # أكثر الموديلات شيوعاً
        dashboard.merge_cells('A12:B12')
        dashboard['A12'].value = "🚙 أكثر الموديلات شيوعاً"
        dashboard['A12'].font = header_font
        dashboard['A12'].fill = header_fill
        dashboard['A12'].alignment = center_align
        
        row = 13
        for model, count in model_counter.most_common(8):
            dashboard.cell(row=row, column=1).value = model
            dashboard.cell(row=row, column=1).alignment = Alignment(horizontal='right')
            dashboard.cell(row=row, column=2).value = count
            dashboard.cell(row=row, column=2).alignment = center_align
            dashboard.cell(row=row, column=2).fill = stat_fill
            row += 1
        
        # توزيع حسب سنة الصنع
        dashboard.merge_cells('D12:F12')
        dashboard['D12'].value = "📅 توزيع حسب سنة الصنع"
        dashboard['D12'].font = header_font
        dashboard['D12'].fill = header_fill
        dashboard['D12'].alignment = center_align
        
        row = 13
        for year, count in year_counter.most_common(8):
            dashboard.cell(row=row, column=4).value = year
            dashboard.cell(row=row, column=4).alignment = center_align
            dashboard.cell(row=row, column=5).value = count
            dashboard.cell(row=row, column=5).alignment = center_align
            dashboard.cell(row=row, column=5).fill = stat_fill
            
            percentage = f"{(count/total_vehicles*100):.1f}%"
            dashboard.cell(row=row, column=6).value = percentage
            dashboard.cell(row=row, column=6).alignment = center_align
            dashboard.cell(row=row, column=6).font = Font(bold=True, color="70AD47")
            row += 1
        
        # ضبط عرض الأعمدة في Dashboard
        dashboard.column_dimensions['A'].width = 25
        dashboard.column_dimensions['B'].width = 15
        dashboard.column_dimensions['C'].width = 3
        dashboard.column_dimensions['D'].width = 30
        dashboard.column_dimensions['E'].width = 12
        dashboard.column_dimensions['F'].width = 12
        
        # ===== ورقة بيانات المركبات الاحترافية =====
        vehicles_sheet = workbook.create_sheet(title="Vehicle Data")
        
        # العناوين الأساسية
        headers = [
            "رقم اللوحة", "الشركة المصنعة", "الموديل", "اللون",
            "سنة الصنع", "الحالة", "اسم السائق", "نوع السيارة",
            "تاريخ انتهاء الفحص", "تاريخ انتهاء الاستمارة"
        ]
        
        # كتابة العناوين
        for col_idx, header in enumerate(headers, start=1):
            cell = vehicles_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = center_align
            cell.border = border
        
        # كتابة البيانات
        for row_idx, vehicle in enumerate(vehicles, start=2):
            # البيانات الأساسية
            status_ar = status_map.get(vehicle.status, vehicle.status)
            data_row = [
                vehicle.plate_number or "",
                vehicle.make or "",
                vehicle.model or "",
                vehicle.color or "",
                str(vehicle.year) if vehicle.year else "",
                status_ar,
                vehicle.driver_name or "",
                vehicle.type_of_car or "",
                vehicle.inspection_expiry_date.strftime('%Y-%m-%d') if vehicle.inspection_expiry_date else "",
                vehicle.registration_expiry_date.strftime('%Y-%m-%d') if vehicle.registration_expiry_date else ""
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = vehicles_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = center_align
                cell.border = border
                
                # تلوين الصفوف بالتناوب
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # تلوين الحالة
                if col_idx == 6:  # عمود الحالة
                    if value == 'متاحة':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif value == 'في الورشة':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")
                    elif value in ['مؤجرة', 'في المشروع']:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True, color="9C6500")
        
        # ضبط عرض الأعمدة
        column_widths = [18, 20, 18, 15, 12, 15, 20, 22, 20, 20]
        for idx, width in enumerate(column_widths, start=1):
            vehicles_sheet.column_dimensions[get_column_letter(idx)].width = width
        
        # ===== ورقة البيانات الكاملة =====
        full_data_sheet = workbook.create_sheet(title="Complete Data")
        
        # جميع البيانات
        all_headers = [
            'رقم اللوحة', 'الشركة المصنعة', 'الموديل', 'اللون',
            'سنة الصنع', 'الحالة', 'اسم السائق', 'نوع السيارة',
            'المشروع', 'تاريخ انتهاء الفحص', 'تاريخ انتهاء الاستمارة',
            'تاريخ انتهاء التفويض', 'رابط Google Drive', 'الملاحظات', 'تاريخ الإضافة'
        ]
        
        for col_idx, header in enumerate(all_headers, start=1):
            cell = full_data_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = center_align
            cell.border = border
        
        for row_idx, vehicle in enumerate(vehicles, start=2):
            status_ar = status_map.get(vehicle.status, vehicle.status)
            all_data = [
                vehicle.plate_number or "",
                vehicle.make or "",
                vehicle.model or "",
                vehicle.color or "",
                str(vehicle.year) if vehicle.year else "",
                status_ar,
                vehicle.driver_name or "",
                vehicle.type_of_car or "",
                vehicle.project or "",
                vehicle.inspection_expiry_date.strftime('%Y-%m-%d') if vehicle.inspection_expiry_date else "",
                vehicle.registration_expiry_date.strftime('%Y-%m-%d') if vehicle.registration_expiry_date else "",
                vehicle.authorization_expiry_date.strftime('%Y-%m-%d') if vehicle.authorization_expiry_date else "",
                vehicle.drive_folder_link or "",
                vehicle.notes or "",
                vehicle.created_at.strftime('%Y-%m-%d') if vehicle.created_at else ""
            ]
            
            for col_idx, value in enumerate(all_data, start=1):
                cell = full_data_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = center_align
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # ضبط عرض الأعمدة
        for idx in range(1, len(all_headers) + 1):
            full_data_sheet.column_dimensions[get_column_letter(idx)].width = 20
        
        # ===== ورقة English Format =====
        english_sheet = workbook.create_sheet(title="English Format")
        
        # العناوين بالإنجليزية
        english_headers = [
            '#', 'Name', 'ID Num', 'EMP', 'Private Num', 'Work Num', 
            'Plate Num.', 'Owned By', 'TYPE OF VEHICLE', 'Project', 
            'Location', 'Start Date'
        ]
        
        # كتابة العناوين
        for col_idx, header in enumerate(english_headers, start=1):
            cell = english_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
            cell.fill = PatternFill(start_color="18B2B0", end_color="18B2B0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # كتابة البيانات
        for row_idx, vehicle in enumerate(vehicles, start=2):
            # الحصول على بيانات السائق والموظف المرتبط
            driver_name = vehicle.driver_name or ""
            employee_id_num = ""
            employee_num = ""
            private_num = ""
            work_num = ""
            project = vehicle.project or ""
            location = ""
            start_date = ""
            owner = ""
            
            # محاولة الحصول على بيانات السائق من نموذج Employee بناءً على الاسم
            if driver_name:
                from models import Employee
                driver = Employee.query.filter_by(name=driver_name).first()
                if driver:
                    employee_id_num = driver.national_id or ""
                    employee_num = driver.employee_id or ""
                    private_num = driver.mobilePersonal or ""
                    work_num = driver.mobile or ""
            
            # الحصول على المنطقة من حقل region في المركبة
            location = vehicle.region or ""
            
            # محاولة الحصول على تاريخ البدء من المشروع
            if vehicle.project:
                from models import VehicleProject
                project_obj = VehicleProject.query.filter_by(
                    project_name=vehicle.project
                ).first()
                if project_obj:
                    if project_obj.start_date:
                        start_date = project_obj.start_date.strftime('%Y-%m-%d')
            
            # الحصول على الشركة المالكة من حقل owned_by أو من سجلات الإيجار
            owner = vehicle.owned_by or ""
            if not owner:
                # إذا لم يتم تحديد الشركة المالكة، جرب من سجلات الإيجار
                from models import VehicleRental
                rental = VehicleRental.query.filter_by(
                    vehicle_id=vehicle.id, 
                    is_active=True
                ).first()
                if rental:
                    owner = rental.lessor_name or ""
            
            # البيانات
            data_row = [
                row_idx - 1,  # الرقم التسلسلي
                driver_name,
                employee_id_num,
                employee_num,
                private_num,
                work_num,
                vehicle.plate_number or "",
                owner,  # شركة الإيجار أو المالك
                f"{vehicle.make or ''} - {vehicle.model or ''}".strip(' -'),
                project,
                location,
                start_date
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = english_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # تلوين الصفوف بالتناوب
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # ضبط عرض الأعمدة
        column_widths = [8, 30, 15, 12, 15, 15, 15, 20, 25, 18, 15, 15]
        for idx, width in enumerate(column_widths, start=1):
            english_sheet.column_dimensions[get_column_letter(idx)].width = width
        
        # حفظ الملف
        workbook.save(output)
        output.seek(0)
        return output
    
    except Exception as e:
        print(f"خطأ في إنشاء ملف Excel للمركبات: {str(e)}")
        import traceback
        traceback.print_exc()
        raise Exception(f"Error generating vehicles Excel file: {str(e)}")

def parse_salary_excel(file, month, year):
    """
    Parse Excel file containing salary data
    
    Args:
        file: The uploaded Excel file
        month: The month for these salaries
        year: The year for these salaries
        
    Returns:
        List of dictionaries containing salary data
    """
    try:
        # Reset file pointer to beginning
        file.seek(0)
        
        # Read the Excel file explicitly using openpyxl engine
        df = pd.read_excel(file, engine='openpyxl')
        
        # Print column names for debugging
        print(f"Salary Excel columns: {df.columns.tolist()}")
        
        # Create a mapping for column detection
        column_mappings = {
            'employee_id': ['employee_id', 'employee id', 'emp id', 'employee number', 'emp no', 'emp.id', 'emp.no', 'emp .n', 'رقم الموظف', 'معرف الموظف', 'الرقم الوظيفي'],
            'basic_salary': ['basic_salary', 'basic salary', 'salary', 'راتب', 'الراتب', 'الراتب الأساسي'],
            'allowances': ['allowances', 'بدل', 'بدلات', 'البدلات'],
            'deductions': ['deductions', 'خصم', 'خصومات', 'الخصومات'],
            'bonus': ['bonus', 'مكافأة', 'علاوة', 'مكافآت'],
            'notes': ['notes', 'ملاحظات']
        }
        
        # Map columns to their normalized names
        detected_columns = {}
        for col in df.columns:
            if isinstance(col, datetime):
                continue
                
            col_str = str(col).lower().strip()
            
            # Check for exact column name or common variations
            for field, variations in column_mappings.items():
                if col_str in variations or any(var in col_str for var in variations):
                    detected_columns[field] = col
                    print(f"Detected '{field}' column: {col}")
                    break
        
        # Handle special case for Excel files with specific column names
        explicit_mappings = {
            'Employee ID': 'employee_id',
            'Emp .N': 'employee_id',  # شكل آخر لرقم الموظف
            'Basic Salary': 'basic_salary',
            'Allowances': 'allowances',
            'Deductions': 'deductions',
            'Bonus': 'bonus',
            'Notes': 'notes'
        }
        
        for excel_col, field in explicit_mappings.items():
            if excel_col in df.columns:
                detected_columns[field] = excel_col
                print(f"Explicitly mapped '{excel_col}' to '{field}'")
        
        # Print final column mapping
        print(f"Final salary column mapping: {detected_columns}")
        
        # تقسيم الحقول المطلوبة إلى أساسية وغير أساسية
        essential_fields = ['employee_id']  # رقم الموظف هو الأساسي الوحيد الضروري دائماً
        other_fields = ['basic_salary', 'allowances', 'deductions', 'bonus']  # يمكن وضع قيم افتراضية لها
        
        # التحقق من الحقول الأساسية
        missing_essential = [field for field in essential_fields if field not in detected_columns]
        if missing_essential:
            missing_str = ", ".join(missing_essential)
            raise ValueError(f"Required columns missing: {missing_str}. Available columns: {[c for c in df.columns if not isinstance(c, datetime)]}")
        
        # بالنسبة للحقول غير الأساسية المفقودة، سننشئ أعمدة وهمية تحتوي على قيم افتراضية
        for field in other_fields:
            if field not in detected_columns:
                print(f"Warning: Creating default column for: {field}")
                dummy_column_name = f"__{field}__default"
                df[dummy_column_name] = 0  # إنشاء عمود فارغ (0 للقيم المالية)
                detected_columns[field] = dummy_column_name  # تعيين العمود الوهمي للحقل
        
        # Process each row
        salaries = []
        for idx, row in df.iterrows():
            try:
                # Skip completely empty rows
                if row.isnull().all():
                    continue
                
                # Get employee_id field
                emp_id_col = detected_columns['employee_id']
                emp_id = row[emp_id_col]
                
                # Skip rows with missing employee_id
                if pd.isna(emp_id):
                    print(f"Skipping row {idx+1} due to missing employee ID")
                    continue
                
                # Try to convert employee_id to integer
                try:
                    employee_id = int(emp_id)
                except (ValueError, TypeError):
                    # If not convertible to int, use as string (could be employee code)
                    employee_id = str(emp_id).strip()
                
                # Get basic_salary field
                basic_salary_col = detected_columns['basic_salary']
                basic_salary_val = row[basic_salary_col]
                
                # تعامل مع القيم المفقودة أو غير الرقمية للراتب الأساسي
                if pd.isna(basic_salary_val) or not isinstance(basic_salary_val, (int, float)):
                    print(f"Row {idx+1}: Using default value of 0 for basic salary")
                    basic_salary_val = 0
                
                basic_salary = float(basic_salary_val)
                
                # Get optional fields with default values
                allowances = 0.0
                deductions = 0.0
                bonus = 0.0
                notes = ''
                
                # Extract allowances if column exists
                if 'allowances' in detected_columns and not pd.isna(row[detected_columns['allowances']]):
                    try:
                        allowances = float(row[detected_columns['allowances']])
                    except (ValueError, TypeError):
                        allowances = 0.0
                
                # Extract deductions if column exists
                if 'deductions' in detected_columns and not pd.isna(row[detected_columns['deductions']]):
                    try:
                        deductions = float(row[detected_columns['deductions']])
                    except (ValueError, TypeError):
                        deductions = 0.0
                
                # Extract bonus if column exists
                if 'bonus' in detected_columns and not pd.isna(row[detected_columns['bonus']]):
                    try:
                        bonus = float(row[detected_columns['bonus']])
                    except (ValueError, TypeError):
                        bonus = 0.0
                
                # Extract notes if column exists
                if 'notes' in detected_columns and not pd.isna(row[detected_columns['notes']]):
                    notes = str(row[detected_columns['notes']])
                
                # Calculate net salary
                net_salary = basic_salary + allowances + bonus - deductions
                
                # Create salary dictionary
                salary = {
                    'employee_id': employee_id,
                    'month': month,
                    'year': year,
                    'basic_salary': basic_salary,
                    'allowances': allowances,
                    'deductions': deductions,
                    'bonus': bonus,
                    'net_salary': net_salary
                }
                
                if notes:
                    salary['notes'] = notes
                
                print(f"Processed salary for employee ID: {employee_id}")
                salaries.append(salary)
                
            except Exception as e:
                print(f"Error processing salary row {idx+1}: {str(e)}")
                # Continue to next row instead of failing the entire import
                continue
        
        if not salaries:
            raise ValueError("No valid salary records found in the Excel file")
            
        return salaries
    
    except Exception as e:
        import traceback
        print(f"Error parsing salary Excel: {str(e)}")
        print(traceback.format_exc())
        raise Exception(f"Error parsing salary Excel file: {str(e)}")

def generate_comprehensive_employee_report(db_session, department_id=None, employee_id=None, month=None, year=None):
    """
    إنشاء تقرير شامل للموظفين مع كامل تفاصيل الرواتب والبيانات
    
    Args:
        db_session: جلسة قاعدة البيانات
        department_id: معرف القسم (اختياري للتصفية)
        employee_id: معرف الموظف (اختياري للتصفية)
        month: الشهر (اختياري للتصفية)
        year: السنة (اختياري للتصفية)
        
    Returns:
        كائن BytesIO يحتوي على ملف Excel
    """
    try:
        from models import Employee, Department, Salary, Attendance, Document
        from sqlalchemy import func
        from datetime import datetime, timedelta
        from dateutil.relativedelta import relativedelta
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Color
        from openpyxl.utils import get_column_letter
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting.rule import Rule
        from openpyxl.chart import BarChart, Reference, Series
        from openpyxl.chart.marker import DataPoint
        from openpyxl.drawing.image import Image
        
        # استعلام الموظفين مع التصفية المطلوبة
        query = db_session.query(Employee).join(Department)
        
        if department_id:
            query = query.filter(Employee.department_id == department_id)
        if employee_id:
            query = query.filter(Employee.id == employee_id)
            
        # الحصول على كل الموظفين المطلوبين
        employees = query.all()
        
        # البحث عن الرواتب المرتبطة بهذه الفترة
        salary_query = db_session.query(Salary).filter(Salary.employee_id.in_([e.id for e in employees]))
        if month:
            salary_query = salary_query.filter(Salary.month == month)
        if year:
            salary_query = salary_query.filter(Salary.year == year)
            
        salaries = salary_query.all()
        
        # تحديد الألوان والتنسيقات
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        total_row_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        total_row_font = Font(name="Arial", size=12, bold=True)
        
        normal_font = Font(name="Arial", size=11)
        highlight_font = Font(name="Arial", size=11, bold=True)
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        title_alignment = Alignment(horizontal='center', vertical='center')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='center', vertical='center')
        text_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # تنسيقات للخلايا المالية
        money_format = '#,##0.00 "ر.س"'
        percentage_format = '0.00%'
        date_format = 'yyyy-mm-dd'
        
        # إعداد البيانات المجمعة للموظفين
        employees_data = []
        salaries_by_employee = {}
        
        # تجميع الرواتب حسب الموظف
        for salary in salaries:
            if salary.employee_id not in salaries_by_employee:
                salaries_by_employee[salary.employee_id] = []
            salaries_by_employee[salary.employee_id].append(salary)
        
        # تجميع بيانات الموظفين مع الرواتب
        for employee in employees:
            emp_salaries = salaries_by_employee.get(employee.id, [])
            
            # حساب متوسط الراتب وأعلى وأدنى راتب
            basic_salaries = [s.basic_salary for s in emp_salaries] if emp_salaries else [0]
            net_salaries = [s.net_salary for s in emp_salaries] if emp_salaries else [0]
            
            avg_basic = sum(basic_salaries) / len(basic_salaries) if basic_salaries else 0
            avg_net = sum(net_salaries) / len(net_salaries) if net_salaries else 0
            max_net = max(net_salaries) if net_salaries else 0
            min_net = min(net_salaries) if net_salaries else 0
            
            # تجميع معلومات الموظف
            emp_data = {
                'معرف': employee.id,
                'رقم الموظف': employee.employee_id,
                'الاسم': employee.name,
                'القسم': ', '.join([dept.name for dept in employee.departments]) if employee.departments else 'بدون قسم',
                'الوظيفة': employee.job_title or '',
                'تاريخ التعيين': employee.hire_date,
                'الجنسية': employee.nationality or '',
                'الهاتف': employee.phone or '',
                'البريد الإلكتروني': employee.email or '',
                'الرقم الوطني/الإقامة': employee.national_id or '',
                'الحالة': employee.status or '',
                'متوسط الراتب الأساسي': avg_basic,
                'متوسط صافي الراتب': avg_net,
                'أعلى راتب': max_net,
                'أدنى راتب': min_net,
                'عدد الرواتب المسجلة': len(emp_salaries),
                'الملاحظات': employee.notes or ''
            }
            
            # إضافة تفاصيل آخر راتب
            if emp_salaries:
                # ترتيب الرواتب حسب السنة والشهر (تنازلياً)
                sorted_salaries = sorted(emp_salaries, key=lambda s: (s.year, s.month), reverse=True)
                latest_salary = sorted_salaries[0]
                
                emp_data.update({
                    'آخر راتب - الشهر': latest_salary.month,
                    'آخر راتب - السنة': latest_salary.year,
                    'آخر راتب - الأساسي': latest_salary.basic_salary,
                    'آخر راتب - البدلات': latest_salary.allowances,
                    'آخر راتب - الخصومات': latest_salary.deductions,
                    'آخر راتب - المكافآت': latest_salary.bonus,
                    'آخر راتب - الصافي': latest_salary.net_salary
                })
            
            employees_data.append(emp_data)
        
        # إنشاء ملف Excel باستخدام openpyxl
        output = BytesIO()
        with pd.ExcelWriter(path=output, engine='openpyxl') as writer:
            # ======= ورقة ملخص الموظفين =======
            emp_df = pd.DataFrame(employees_data)
            
            # ترتيب الأعمدة للتقرير الشامل
            columns_order = [
                'معرف', 'رقم الموظف', 'الاسم', 'القسم', 'الوظيفة', 'تاريخ التعيين', 
                'الجنسية', 'الرقم الوطني/الإقامة', 'الهاتف', 'البريد الإلكتروني', 'الحالة',
                'متوسط الراتب الأساسي', 'متوسط صافي الراتب', 'أعلى راتب', 'أدنى راتب', 
                'عدد الرواتب المسجلة',
                'آخر راتب - الشهر', 'آخر راتب - السنة', 'آخر راتب - الأساسي', 
                'آخر راتب - البدلات', 'آخر راتب - الخصومات', 'آخر راتب - المكافآت', 
                'آخر راتب - الصافي', 'الملاحظات'
            ]
            
            # استبعاد الأعمدة غير الموجودة
            actual_columns = [col for col in columns_order if col in emp_df.columns]
            emp_df = emp_df[actual_columns]
            
            # كتابة البيانات إلى الملف
            emp_df.to_excel(writer, sheet_name='ملخص الموظفين', index=False, startrow=2)
            
            # الحصول على ورقة العمل وتنسيقها
            summary_sheet = writer.sheets['ملخص الموظفين']
            
            # إضافة عنوان للتقرير
            summary_sheet.merge_cells(f'A1:{get_column_letter(len(actual_columns))}1')
            title_cell = summary_sheet.cell(1, 1)
            title_cell.value = "التقرير الشامل للموظفين مع تفاصيل الرواتب"
            title_cell.font = Font(name="Arial", size=16, bold=True, color="1F4E78")
            title_cell.alignment = title_alignment
            
            # تنسيق العناوين
            for col_idx, column_name in enumerate(actual_columns, 1):
                cell = summary_sheet.cell(3, col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
                # ضبط عرض العمود
                column_width = max((
                    emp_df[column_name].astype(str).map(len).max() if len(emp_df) > 0 else 0, 
                    len(column_name)
                )) + 4
                column_letter = get_column_letter(col_idx)
                summary_sheet.column_dimensions[column_letter].width = column_width
            
            # تنسيق البيانات
            for row_idx, (_, row) in enumerate(emp_df.iterrows(), 1):
                for col_idx, column_name in enumerate(actual_columns, 1):
                    cell = summary_sheet.cell(row_idx + 3, col_idx)
                    value = row[column_name]
                    cell.value = value
                    
                    # تنسيق خاص لأنواع البيانات المختلفة
                    if 'راتب' in column_name:
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
                    elif 'تاريخ' in column_name and value:
                        cell.number_format = date_format
                        cell.alignment = cell_alignment
                    else:
                        cell.alignment = text_alignment
                    
                    # تنسيق صفوف بديلة
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    
                    cell.font = normal_font
                    cell.border = thin_border
            
            # ======= ورقة تفاصيل الرواتب لكل موظف =======
            # تحضير بيانات الرواتب مفصلة حسب الموظف
            all_salary_data = []
            
            for employee in employees:
                emp_salaries = salaries_by_employee.get(employee.id, [])
                
                for salary in emp_salaries:
                    all_salary_data.append({
                        'معرف الموظف': employee.id,
                        'رقم الموظف': employee.employee_id,
                        'اسم الموظف': employee.name,
                        'القسم': ', '.join([dept.name for dept in employee.departments]) if employee.departments else 'بدون قسم',
                        'الشهر': salary.month,
                        'السنة': salary.year,
                        'الراتب الأساسي': salary.basic_salary,
                        'البدلات': salary.allowances,
                        'الخصومات': salary.deductions,
                        'المكافآت': salary.bonus,
                        'صافي الراتب': salary.net_salary,
                        'ملاحظات': salary.notes or ''
                    })
            
            if all_salary_data:
                # إنشاء DataFrame للرواتب
                salary_df = pd.DataFrame(all_salary_data)
                
                # ترتيب البيانات حسب القسم، الموظف، السنة، الشهر
                salary_df = salary_df.sort_values(by=['القسم', 'اسم الموظف', 'السنة', 'الشهر'], ascending=[True, True, False, False])
                
                # كتابة البيانات إلى ورقة عمل جديدة
                salary_df.to_excel(writer, sheet_name='تفاصيل الرواتب', index=False, startrow=2)
                
                # تنسيق ورقة تفاصيل الرواتب
                salary_sheet = writer.sheets['تفاصيل الرواتب']
                
                # إضافة عنوان
                salary_sheet.merge_cells(f'A1:{get_column_letter(len(salary_df.columns))}1')
                title_cell = salary_sheet.cell(1, 1)
                title_cell.value = "تفاصيل رواتب الموظفين"
                title_cell.font = Font(name="Arial", size=16, bold=True, color="1F4E78")
                title_cell.alignment = title_alignment
                
                # تنسيق العناوين
                for col_idx, column_name in enumerate(salary_df.columns, 1):
                    cell = salary_sheet.cell(3, col_idx)
                    cell.value = column_name
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    
                    # ضبط عرض العمود
                    column_width = max(salary_df[column_name].astype(str).map(len).max(), len(column_name)) + 4
                    column_letter = get_column_letter(col_idx)
                    salary_sheet.column_dimensions[column_letter].width = column_width
                
                # تجميع الصفوف حسب الموظف بألوان مختلفة
                current_employee = None
                color_index = 0
                colors = ["E6F2FF", "F2F2F2"]  # ألوان التناوب
                
                # تنسيق البيانات
                for row_idx, (_, row) in enumerate(salary_df.iterrows(), 1):
                    # تغيير اللون عند تغيير الموظف
                    if current_employee != row['اسم الموظف']:
                        current_employee = row['اسم الموظف']
                        color_index = (color_index + 1) % 2
                    
                    row_fill = PatternFill(start_color=colors[color_index], end_color=colors[color_index], fill_type="solid")
                    
                    for col_idx, column_name in enumerate(salary_df.columns, 1):
                        cell = salary_sheet.cell(row_idx + 3, col_idx)
                        cell.value = row[column_name]
                        
                        # تنسيق خاص لأنواع البيانات المختلفة
                        if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الراتب']:
                            cell.number_format = money_format
                            cell.alignment = cell_alignment
                        else:
                            cell.alignment = text_alignment
                        
                        cell.fill = row_fill
                        cell.font = normal_font
                        cell.border = thin_border
                
                # إضافة صف للمجاميع في نهاية الجدول
                total_row_idx = len(salary_df) + 4
                salary_sheet.cell(total_row_idx, 1).value = "المجموع الكلي"
                salary_sheet.cell(total_row_idx, 1).font = total_row_font
                salary_sheet.cell(total_row_idx, 1).alignment = text_alignment
                salary_sheet.cell(total_row_idx, 1).fill = total_row_fill
                salary_sheet.cell(total_row_idx, 1).border = thick_border
                
                # دمج خلايا المجموع
                merge_cols = 6  # دمج الخلايا الأولى للمجموع
                salary_sheet.merge_cells(f'A{total_row_idx}:{get_column_letter(merge_cols)}{total_row_idx}')
                
                # تنسيق وحساب المجاميع
                for col_idx, column_name in enumerate(salary_df.columns, 1):
                    cell = salary_sheet.cell(total_row_idx, col_idx)
                    cell.font = total_row_font
                    cell.fill = total_row_fill
                    cell.border = thick_border
                    
                    if col_idx <= merge_cols:
                        continue  # تخطي الخلايا المدمجة
                    
                    # حساب المجاميع للأعمدة المالية
                    if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الراتب']:
                        col_letter = get_column_letter(col_idx)
                        cell.value = f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})"
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
                
                # إضافة رسم بياني للرواتب حسب القسم
                try:
                    chart_sheet = writer.book.create_sheet(title="الرسوم البيانية")
                    
                    # إعداد بيانات الرسم البياني - متوسط الراتب حسب القسم
                    dept_avg_salary = salary_df.groupby('القسم')['صافي الراتب'].mean().reset_index()
                    dept_avg_salary.to_excel(writer, sheet_name="الرسوم البيانية", startrow=1, startcol=1, index=False)
                    
                    chart_sheet.cell(1, 1).value = "متوسط الرواتب حسب القسم"
                    chart_sheet.cell(1, 1).font = Font(name="Arial", size=14, bold=True)
                    
                    # إنشاء الرسم البياني
                    chart = BarChart()
                    chart.title = "متوسط الرواتب حسب القسم"
                    chart.y_axis.title = "متوسط الراتب (ر.س)"
                    chart.x_axis.title = "القسم"
                    
                    # تحديد نطاق البيانات
                    data = Reference(chart_sheet, min_col=3, min_row=2, max_row=2+len(dept_avg_salary))
                    cats = Reference(chart_sheet, min_col=2, min_row=3, max_row=2+len(dept_avg_salary))
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    
                    # إضافة الرسم البياني إلى الورقة
                    chart_sheet.add_chart(chart, "E5")
                    
                except Exception as chart_error:
                    print(f"حدث خطأ أثناء إنشاء الرسم البياني: {chart_error}")
            
            # ======= ورقة معلومات التقرير =======
            # إنشاء ورقة معلومات التقرير
            info_data = []
            
            # إضافة معلومات عامة
            info_data.append({
                'المعلومة': 'تاريخ التصدير',
                'القيمة': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # إضافة معلومات حول التصفية
            if department_id:
                dept = db_session.query(Department).get(department_id)
                info_data.append({
                    'المعلومة': 'تصفية حسب القسم',
                    'القيمة': dept.name if dept else department_id
                })
            
            if employee_id:
                emp = db_session.query(Employee).get(employee_id)
                info_data.append({
                    'المعلومة': 'تصفية حسب الموظف',
                    'القيمة': emp.name if emp else employee_id
                })
            
            if month:
                info_data.append({
                    'المعلومة': 'تصفية حسب الشهر',
                    'القيمة': month
                })
            
            if year:
                info_data.append({
                    'المعلومة': 'تصفية حسب السنة',
                    'القيمة': year
                })
            
            # إضافة إحصائيات عامة
            info_data.append({
                'المعلومة': 'إجمالي عدد الموظفين',
                'القيمة': len(employees)
            })
            
            info_data.append({
                'المعلومة': 'إجمالي عدد الرواتب المسجلة',
                'القيمة': len(salaries)
            })
            
            # حساب متوسطات الرواتب
            if salaries:
                avg_basic = sum(s.basic_salary for s in salaries) / len(salaries)
                avg_net = sum(s.net_salary for s in salaries) / len(salaries)
                
                info_data.append({
                    'المعلومة': 'متوسط الراتب الأساسي',
                    'القيمة': avg_basic
                })
                
                info_data.append({
                    'المعلومة': 'متوسط صافي الراتب',
                    'القيمة': avg_net
                })
                
                info_data.append({
                    'المعلومة': 'إجمالي مصاريف الرواتب',
                    'القيمة': sum(s.net_salary for s in salaries)
                })
            
            # إنشاء DataFrame للمعلومات
            info_df = pd.DataFrame(info_data)
            info_df.to_excel(writer, sheet_name='معلومات التقرير', index=False, startrow=2)
            
            # تنسيق ورقة المعلومات
            info_sheet = writer.sheets['معلومات التقرير']
            
            # إضافة عنوان للورقة
            info_sheet.merge_cells('A1:B1')
            info_sheet.cell(1, 1).value = "معلومات التقرير الشامل"
            info_sheet.cell(1, 1).font = Font(name="Arial", size=16, bold=True, color="1F4E78")
            info_sheet.cell(1, 1).alignment = title_alignment
            
            # تنسيق العناوين
            for col_idx, col_name in enumerate(info_df.columns, 1):
                info_sheet.cell(3, col_idx).value = col_name
                info_sheet.cell(3, col_idx).font = header_font
                info_sheet.cell(3, col_idx).fill = header_fill
                info_sheet.cell(3, col_idx).alignment = header_alignment
                info_sheet.cell(3, col_idx).border = thin_border
                
                # ضبط عرض العمود
                column_width = max(info_df[col_name].astype(str).map(len).max(), len(col_name)) + 4
                column_letter = get_column_letter(col_idx)
                info_sheet.column_dimensions[column_letter].width = column_width
            
            # تنسيق البيانات
            for row_idx, (_, row) in enumerate(info_df.iterrows(), 1):
                for col_idx, col_name in enumerate(info_df.columns, 1):
                    cell = info_sheet.cell(row_idx + 3, col_idx)
                    cell.value = row[col_name]
                    
                    # تنسيق خاص للقيم المالية
                    if 'متوسط' in row['المعلومة'] or 'إجمالي مصاريف' in row['المعلومة']:
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
                    else:
                        cell.alignment = text_alignment
                    
                    # تنسيق صفوف بديلة
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    
                    cell.font = normal_font
                    cell.border = thin_border
            
            # تعيين الصفحة الأولى كصفحة نشطة
            writer.book.active = writer.book.worksheets[0]
        
        output.seek(0)
        return output
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise Exception(f"خطأ في إنشاء التقرير الشامل: {str(e)}")

def generate_employee_salary_simple_excel(db_session, month=None, year=None, department_id=None):
    """
    إنشاء ملف Excel بسيط وواضح لبيانات الموظفين مع تفاصيل الرواتب
    مع تنسيق احترافي للجداول وألوان متناوبة للصفوف وتنسيق مالي للأرقام
    
    Args:
        db_session: جلسة قاعدة البيانات
        month: الشهر المطلوب (اختياري)
        year: السنة المطلوبة (اختياري)
        department_id: معرّف القسم (اختياري)
        
    Returns:
        كائن BytesIO يحتوي على ملف Excel
    """
    try:
        from models import Employee, Department, Salary
        from datetime import datetime
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # تحديد الألوان والتنسيقات
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        normal_font = Font(name="Arial", size=11)
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # تنسيقات للخلايا المالية
        money_format = '#,##0.00 "ر.س"'
        
        # استعلام الرواتب مع بيانات الموظفين بتحديد صريح للعلاقات
        query = db_session.query(Salary).\
            join(Employee, Salary.employee_id == Employee.id).\
            join(Department, Employee.department_id == Department.id)
        
        # تطبيق الفلاتر
        if department_id:
            query = query.filter(Department.id == department_id)
        if month:
            query = query.filter(Salary.month == month)
        if year:
            query = query.filter(Salary.year == year)
            
        # ترتيب البيانات حسب القسم، اسم الموظف، السنة والشهر
        query = query.order_by(Department.name, Employee.name, Salary.year.desc(), Salary.month.desc())
        
        # الحصول على النتائج
        results = query.all()
        
        # جمع البيانات في قائمة
        employee_data = []
        
        for salary in results:
            employee = salary.employee
            departments_list = employee.departments if employee.departments else []
            department_name = ', '.join([dept.name for dept in departments_list]) if departments_list else 'بدون قسم'
            
            data = {
                'اسم الموظف': employee.name,
                'رقم الموظف': employee.employee_id,
                'رقم الهوية': employee.national_id or '',
                'القسم': department_name,
                'الشهر': salary.month,
                'السنة': salary.year,
                'الراتب الأساسي': salary.basic_salary,
                'البدلات': salary.allowances,
                'الخصومات': salary.deductions,
                'المكافآت': salary.bonus,
                'صافي الراتب': salary.net_salary,
                'ملاحظات': salary.notes or ''
            }
            
            employee_data.append(data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        
        with pd.ExcelWriter(path=output, engine='openpyxl') as writer:
            # إنشاء DataFrame
            df = pd.DataFrame(employee_data)
            
            # ترتيب الأعمدة بالشكل المطلوب
            columns_order = [
                'اسم الموظف', 'رقم الموظف', 'رقم الهوية', 'القسم',
                'الشهر', 'السنة', 'الراتب الأساسي', 'البدلات',
                'الخصومات', 'المكافآت', 'صافي الراتب', 'ملاحظات'
            ]
            
            # ترتيب الأعمدة حسب الترتيب المحدد
            df = df[columns_order]
            
            # كتابة البيانات إلى الملف
            df.to_excel(writer, sheet_name='بيانات الموظفين والرواتب', index=False)
            
            # الحصول على ورقة العمل وتنسيقها
            sheet = writer.sheets['بيانات الموظفين والرواتب']
            
            # تنسيق العناوين
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = sheet.cell(1, col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # ضبط عرض العمود
                column_letter = get_column_letter(col_idx)
                max_length = max(df[col_name].astype(str).map(len).max(), len(col_name)) + 2
                sheet.column_dimensions[column_letter].width = max_length
            
            # تنسيق البيانات
            for row_idx, _ in enumerate(df.iterrows(), 2):  # بدء من الصف 2 (بعد العناوين)
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = sheet.cell(row_idx, col_idx)
                    
                    # تنسيق الخلايا المالية
                    if col_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الراتب']:
                        cell.number_format = money_format
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = right_alignment
                    
                    # تنسيق عام
                    cell.font = normal_font
                    cell.border = thin_border
                    
                    # تلوين الصفوف بالتناوب
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # إضافة صف المجموع في النهاية
            total_row = len(df) + 2
            
            # إضافة نص "المجموع" في أول خلية
            sheet.cell(total_row, 1).value = "المجموع"
            sheet.cell(total_row, 1).font = Font(name="Arial", size=12, bold=True)
            sheet.cell(total_row, 1).alignment = right_alignment
            sheet.cell(total_row, 1).border = thin_border
            sheet.cell(total_row, 1).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # دمج الخلايا من العمود 1 إلى العمود 6
            for col_idx in range(2, 7):
                cell = sheet.cell(total_row, col_idx)
                cell.border = thin_border
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
            
            # حساب المجاميع للأعمدة المالية
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الراتب']:
                    col_letter = get_column_letter(col_idx)
                    cell = sheet.cell(total_row, col_idx)
                    cell.value = f"=SUM({col_letter}2:{col_letter}{total_row-1})"
                    cell.font = Font(name="Arial", size=12, bold=True)
                    cell.number_format = money_format
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                elif col_idx > 6:
                    # تنسيق باقي الخلايا في صف المجموع
                    cell = sheet.cell(total_row, col_idx)
                    cell.border = thin_border
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # إضافة معلومات الفلترة في أعلى الورقة
            info_row = sheet.max_row + 2
            
            # إضافة عنوان المعلومات
            sheet.cell(info_row, 1).value = "معلومات التقرير:"
            sheet.cell(info_row, 1).font = Font(name="Arial", size=12, bold=True)
            
            # إضافة تفاصيل الفلترة
            info_row += 1
            sheet.cell(info_row, 1).value = "تاريخ التصدير:"
            sheet.cell(info_row, 2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            info_row += 1
            filter_text = []
            if month:
                filter_text.append(f"الشهر: {month}")
            if year:
                filter_text.append(f"السنة: {year}")
            if department_id:
                dept = db_session.query(Department).get(department_id)
                if dept:
                    filter_text.append(f"القسم: {dept.name}")
            
            sheet.cell(info_row, 1).value = "الفلاتر المطبقة:"
            sheet.cell(info_row, 2).value = " | ".join(filter_text) if filter_text else "كافة البيانات"
            
            info_row += 1
            sheet.cell(info_row, 1).value = "عدد السجلات:"
            sheet.cell(info_row, 2).value = len(df)
        
        output.seek(0)
        return output
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise Exception(f"خطأ في إنشاء ملف Excel: {str(e)}")

def generate_salary_excel(salaries, filter_description=None):
    """
    إنشاء ملف Excel من بيانات الرواتب مع تنظيم وتجميع حسب القسم وتنسيق ممتاز
    
    Args:
        salaries: قائمة كائنات Salary 
        filter_description: وصف مرشحات البحث المستخدمة (اختياري)
        
    Returns:
        كائن BytesIO يحتوي على ملف Excel
    """
    try:
        from datetime import datetime
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Color
        from openpyxl.utils import get_column_letter
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting.rule import Rule
        from openpyxl.drawing.image import Image
        
        # تحديد الألوان والتنسيقات
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        total_row_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        total_row_font = Font(name="Arial", size=12, bold=True)
        
        normal_font = Font(name="Arial", size=11)
        highlight_font = Font(name="Arial", size=11, bold=True)
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        title_alignment = Alignment(horizontal='center', vertical='center')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='center', vertical='center')
        text_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # تنسيقات للخلايا المالية
        money_format = '#,##0.00 "ر.س"'
        percentage_format = '0.00%'
        
        # تجميع البيانات حسب القسم
        departments_data = {}
        for salary in salaries:
            dept_name = salary.employee.department.name if salary.employee.department else 'بدون قسم'
            if dept_name not in departments_data:
                departments_data[dept_name] = []
            
            # إضافة بيانات الراتب إلى القسم المناسب
            departments_data[dept_name].append({
                'معرف': salary.id,
                'اسم الموظف': salary.employee.name,
                'رقم الموظف': salary.employee.employee_id,
                'رقم الهوية': salary.employee.national_id or '',
                'الوظيفة': salary.employee.job_title or '',
                'القسم': dept_name,
                'الشهر': salary.month,
                'السنة': salary.year,
                'الراتب الأساسي': salary.basic_salary,
                'البدلات': salary.allowances,
                'الخصومات': salary.deductions,
                'المكافآت': salary.bonus,
                'أيام الحضور': salary.present_days if salary.attendance_calculated else '-',
                'أيام الغياب': salary.absent_days if salary.attendance_calculated else '-',
                'خصم الغياب': salary.attendance_deduction if salary.attendance_calculated else 0,
                'صافي الراتب': salary.net_salary,
                'ملاحظات': salary.notes or ''
            })
        
        # إنشاء ملف Excel باستخدام openpyxl
        output = BytesIO()
        with pd.ExcelWriter(path=output, engine='openpyxl') as writer:
            # بيانات الملخص
            summary_data = []
            total_salaries = 0
            total_basic = 0
            total_allowances = 0
            total_deductions = 0
            total_bonus = 0
            total_net = 0
            
            # حساب المجاميع لكل قسم
            for dept_name, dept_salaries in departments_data.items():
                dept_count = len(dept_salaries)
                dept_basic_sum = sum(s['الراتب الأساسي'] for s in dept_salaries)
                dept_allowances_sum = sum(s['البدلات'] for s in dept_salaries)
                dept_deductions_sum = sum(s['الخصومات'] for s in dept_salaries)
                dept_bonus_sum = sum(s['المكافآت'] for s in dept_salaries)
                dept_net_sum = sum(s['صافي الراتب'] for s in dept_salaries)
                
                summary_data.append({
                    'القسم': dept_name,
                    'عدد الموظفين': dept_count,
                    'إجمالي الرواتب الأساسية': dept_basic_sum,
                    'إجمالي البدلات': dept_allowances_sum,
                    'إجمالي الخصومات': dept_deductions_sum,
                    'إجمالي المكافآت': dept_bonus_sum,
                    'إجمالي صافي الرواتب': dept_net_sum
                })
                
                # تحديث المجاميع الكلية
                total_salaries += dept_count
                total_basic += dept_basic_sum
                total_allowances += dept_allowances_sum
                total_deductions += dept_deductions_sum
                total_bonus += dept_bonus_sum
                total_net += dept_net_sum
            
            # إضافة صف المجموع الكلي
            summary_data.append({
                'القسم': 'الإجمالي',
                'عدد الموظفين': total_salaries,
                'إجمالي الرواتب الأساسية': total_basic,
                'إجمالي البدلات': total_allowances,
                'إجمالي الخصومات': total_deductions,
                'إجمالي المكافآت': total_bonus,
                'إجمالي صافي الرواتب': total_net
            })
            
            # إنشاء DataFrame للملخص
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='ملخص الرواتب', index=False)
            
            # تنسيق ورقة الملخص
            summary_sheet = writer.sheets['ملخص الرواتب']
            
            # إضافة عنوان للتقرير
            summary_sheet.merge_cells('A1:G1')
            summary_sheet.cell(1, 1).value = "تقرير ملخص الرواتب"
            summary_sheet.cell(1, 1).font = Font(name="Arial", size=16, bold=True, color="1F4E78")
            summary_sheet.cell(1, 1).alignment = title_alignment
            
            # إضافة معلومات الفلترة تحت العنوان
            if filter_description:
                summary_sheet.merge_cells('A2:G2')
                summary_sheet.cell(2, 1).value = "مرشحات البحث: " + " - ".join(filter_description)
                summary_sheet.cell(2, 1).font = Font(name="Arial", size=12, italic=True)
                summary_sheet.cell(2, 1).alignment = title_alignment
                
                # ضبط العنوان ليبدأ من الصف الثالث
                title_row_offset = 3
            else:
                # ضبط العنوان ليبدأ من الصف الثاني
                title_row_offset = 2
            
            # الحصول على عدد الصفوف والأعمدة في البيانات
            num_rows = len(summary_data) + 1  # +1 للعنوان
            num_cols = len(summary_df.columns)
            
            # تنسيق الترويسات
            for col_idx, column_name in enumerate(summary_df.columns, 1):
                cell = summary_sheet.cell(title_row_offset, col_idx)
                cell.value = column_name
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
                # ضبط عرض العمود
                column_width = max(summary_df[column_name].astype(str).map(len).max(), len(column_name)) + 4
                column_letter = get_column_letter(col_idx)
                summary_sheet.column_dimensions[column_letter].width = column_width
            
            # تنسيق البيانات
            for row_idx, row in enumerate(summary_data, 1):
                is_total_row = row_idx == len(summary_data)
                
                for col_idx, (column_name, value) in enumerate(zip(summary_df.columns, row.values()), 1):
                    cell = summary_sheet.cell(title_row_offset + row_idx, col_idx)
                    cell.value = value
                    
                    # تنسيق خاص للصف الأخير (الإجمالي)
                    if is_total_row:
                        cell.font = total_row_font
                        cell.fill = total_row_fill
                        cell.border = thick_border
                    else:
                        cell.font = normal_font
                        cell.border = thin_border
                    
                    # تنسيق خاص للأعمدة المالية
                    if 'إجمالي' in column_name or column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الراتب']:
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
                    else:
                        cell.alignment = text_alignment
            
            # إنشاء ورقة عمل مفصلة لكل قسم
            for dept_name, dept_salaries in departments_data.items():
                dept_df = pd.DataFrame(dept_salaries)
                
                # ترتيب الأعمدة بشكل منطقي
                ordered_columns = [
                    'معرف', 'اسم الموظف', 'رقم الموظف', 'رقم الهوية', 'الوظيفة', 'القسم',
                    'الشهر', 'السنة', 'الراتب الأساسي', 'البدلات', 'الخصومات',
                    'المكافآت', 'أيام الحضور', 'أيام الغياب', 'خصم الغياب', 'صافي الراتب', 'ملاحظات'
                ]
                
                # إعادة ترتيب الأعمدة واستبعاد الأعمدة غير الموجودة
                actual_columns = [col for col in ordered_columns if col in dept_df.columns]
                dept_df = dept_df[actual_columns]
                
                # إضافة ورقة العمل للقسم
                sheet_name = dept_name[:31]  # تقليص اسم الورقة إلى 31 حرف (أقصى طول مسموح في Excel)
                dept_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)  # بدء من الصف الثالث لإتاحة مساحة للعنوان
                
                # الحصول على sheet وتنسيق عنوان القسم
                dept_sheet = writer.sheets[sheet_name]
                
                # دمج الخلايا للعنوان
                num_dept_cols = len(actual_columns)
                dept_sheet.merge_cells(f'A1:{get_column_letter(num_dept_cols)}1')
                title_cell = dept_sheet.cell(1, 1)
                title_cell.value = f"تفاصيل رواتب قسم {dept_name}"
                title_cell.font = Font(name="Arial", size=16, bold=True, color="1F4E78")
                title_cell.alignment = title_alignment
                
                # تنسيق الترويسات
                for col_idx, column_name in enumerate(actual_columns, 1):
                    cell = dept_sheet.cell(3, col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    
                    # ضبط عرض العمود
                    column_width = max(dept_df[column_name].astype(str).map(len).max(), len(column_name)) + 4
                    column_letter = get_column_letter(col_idx)
                    dept_sheet.column_dimensions[column_letter].width = column_width
                
                # تنسيق بيانات الموظفين
                for row_idx in range(len(dept_df)):
                    for col_idx, column_name in enumerate(actual_columns, 1):
                        cell = dept_sheet.cell(row_idx + 4, col_idx)  # +4 للترويسة والعنوان
                        
                        # تنسيق خاص للخلايا المالية
                        if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'خصم الغياب', 'صافي الراتب']:
                            cell.number_format = money_format
                            cell.alignment = cell_alignment
                        elif column_name in ['أيام الحضور', 'أيام الغياب']:
                            cell.alignment = cell_alignment
                        else:
                            cell.alignment = text_alignment
                        
                        cell.font = normal_font
                        cell.border = thin_border
                
                # إضافة صف للمجاميع في نهاية جدول القسم
                total_row_idx = len(dept_df) + 4
                dept_sheet.cell(total_row_idx, 1).value = "المجموع"
                dept_sheet.cell(total_row_idx, 1).font = total_row_font
                dept_sheet.cell(total_row_idx, 1).alignment = text_alignment
                
                # تنسيق صف المجموع وحساب المجاميع للأعمدة المالية
                for col_idx, column_name in enumerate(actual_columns, 1):
                    cell = dept_sheet.cell(total_row_idx, col_idx)
                    cell.font = total_row_font
                    cell.fill = total_row_fill
                    cell.border = thick_border
                    
                    # حساب المجاميع للأعمدة المالية
                    if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الراتب']:
                        col_letter = get_column_letter(col_idx)
                        cell.value = f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})"
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
                
                # إضافة قواعد تنسيق شرطية للخلايا
                # تلوين الخلايا ذات القيم السالبة باللون الأحمر
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                red_font = Font(color="9C0006")
                
                # تلوين الخلايا ذات القيم الموجبة باللون الأخضر
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                green_font = Font(color="006100")
                
                # تطبيق التنسيق الشرطي على أعمدة المال
                for col_idx, column_name in enumerate(actual_columns, 1):
                    if column_name in ['الراتب الأساسي', 'البدلات', 'المكافآت', 'صافي الراتب']:
                        col_letter = get_column_letter(col_idx)
                        
                        # قاعدة للقيم العالية (أعلى من المتوسط)
                        high_rule = Rule(
                            type="cellIs",
                            operator="greaterThan",
                            formula=[f"AVERAGE({col_letter}4:{col_letter}{total_row_idx-1})"],
                            dxf=DifferentialStyle(fill=green_fill, font=green_font)
                        )
                        dept_sheet.conditional_formatting.add(f"{col_letter}4:{col_letter}{total_row_idx-1}", high_rule)
                    
                    elif column_name == 'الخصومات':
                        col_letter = get_column_letter(col_idx)
                        
                        # قاعدة للخصومات العالية
                        high_deduction_rule = Rule(
                            type="cellIs",
                            operator="greaterThan",
                            formula=[f"AVERAGE({col_letter}4:{col_letter}{total_row_idx-1})"],
                            dxf=DifferentialStyle(fill=red_fill, font=red_font)
                        )
                        dept_sheet.conditional_formatting.add(f"{col_letter}4:{col_letter}{total_row_idx-1}", high_deduction_rule)
            
            # إنشاء ورقة لجميع الرواتب
            all_data = []
            for dept_salaries in departments_data.values():
                all_data.extend(dept_salaries)
            
            if all_data:
                all_df = pd.DataFrame(all_data)
                
                # ترتيب الأعمدة بشكل منطقي
                all_columns = [col for col in ordered_columns if col in all_df.columns]
                all_df = all_df[all_columns]
                
                # إضافة ورقة كل الرواتب
                all_df.to_excel(writer, sheet_name='جميع الرواتب', index=False, startrow=2)
                
                # تنسيق ورقة كل الرواتب
                all_sheet = writer.sheets['جميع الرواتب']
                
                # دمج الخلايا للعنوان
                num_all_cols = len(all_columns)
                all_sheet.merge_cells(f'A1:{get_column_letter(num_all_cols)}1')
                title_cell = all_sheet.cell(1, 1)
                title_cell.value = "قائمة كاملة بالرواتب"
                title_cell.font = Font(name="Arial", size=16, bold=True, color="1F4E78")
                title_cell.alignment = title_alignment
                
                # تنسيق الترويسات
                for col_idx, column_name in enumerate(all_columns, 1):
                    cell = all_sheet.cell(3, col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    
                    # ضبط عرض العمود
                    column_width = max(all_df[column_name].astype(str).map(len).max(), len(column_name)) + 4
                    column_letter = get_column_letter(col_idx)
                    all_sheet.column_dimensions[column_letter].width = column_width
                
                # تنسيق كافة البيانات
                for row_idx in range(len(all_df)):
                    for col_idx, column_name in enumerate(all_columns, 1):
                        cell = all_sheet.cell(row_idx + 4, col_idx)  # +4 للترويسة والعنوان
                        
                        # تنسيق خاص للخلايا المالية
                        if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'خصم الغياب', 'صافي الراتب']:
                            cell.number_format = money_format
                            cell.alignment = cell_alignment
                        elif column_name in ['أيام الحضور', 'أيام الغياب']:
                            cell.alignment = cell_alignment
                        else:
                            cell.alignment = text_alignment
                        
                        # تمييز الصفوف بألوان متناوبة
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        
                        cell.font = normal_font
                        cell.border = thin_border
                
                # إضافة صف للمجاميع في نهاية الجدول
                total_row_idx = len(all_df) + 4
                all_sheet.cell(total_row_idx, 1).value = "المجموع الكلي"
                all_sheet.cell(total_row_idx, 1).font = total_row_font
                all_sheet.cell(total_row_idx, 1).alignment = text_alignment
                
                # تنسيق صف المجموع وحساب المجاميع للأعمدة المالية
                for col_idx, column_name in enumerate(all_columns, 1):
                    cell = all_sheet.cell(total_row_idx, col_idx)
                    cell.font = total_row_font
                    cell.fill = total_row_fill
                    cell.border = thick_border
                    
                    # حساب المجاميع للأعمدة المالية
                    if column_name in ['الراتب الأساسي', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الراتب']:
                        col_letter = get_column_letter(col_idx)
                        cell.value = f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})"
                        cell.number_format = money_format
                        cell.alignment = cell_alignment
            
            # إنشاء داش بورد احترافي لمعلومات التقرير
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.chart.label import DataLabelList
            
            info_sheet = writer.book.create_sheet('معلومات التقرير')
            
            # ==== العنوان الرئيسي ====
            info_sheet.merge_cells('A1:J1')
            title_cell = info_sheet.cell(1, 1)
            title_cell.value = "📊 داش بورد تحليل الرواتب"
            title_cell.font = Font(name="Arial", size=22, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            info_sheet.row_dimensions[1].height = 35
            
            # ==== معلومات التصفية ====
            info_sheet.merge_cells('A2:J2')
            filter_cell = info_sheet.cell(2, 1)
            filter_text = ' | '.join(filter_description) if filter_description else 'جميع البيانات'
            filter_cell.value = f"المرشحات: {filter_text}"
            filter_cell.font = Font(name="Arial", size=11, italic=True, color="666666")
            filter_cell.alignment = Alignment(horizontal='center', vertical='center')
            info_sheet.row_dimensions[2].height = 20
            
            # ==== بطاقات المؤشرات الرئيسية (KPI Cards) ====
            kpi_row = 4
            kpis = [
                {'title': 'إجمالي الموظفين', 'value': total_salaries, 'color': '4472C4', 'icon': '👥'},
                {'title': 'عدد الأقسام', 'value': len(departments_data), 'color': '70AD47', 'icon': '🏢'},
                {'title': 'متوسط الراتب', 'value': f'{(total_net / total_salaries if total_salaries > 0 else 0):,.0f} ر.س', 'color': 'FFC000', 'icon': '💰'},
                {'title': 'إجمالي المصاريف', 'value': f'{total_net:,.0f} ر.س', 'color': 'E74C3C', 'icon': '💵'}
            ]
            
            col_start = 1
            for idx, kpi in enumerate(kpis):
                col = col_start + (idx * 3)
                
                # دمج خلايا البطاقة
                info_sheet.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row+2, end_column=col+1)
                
                # الأيقونة والعنوان
                header_cell = info_sheet.cell(kpi_row, col)
                header_cell.value = f"{kpi['icon']} {kpi['title']}"
                header_cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
                header_cell.fill = PatternFill(start_color=kpi['color'], end_color=kpi['color'], fill_type="solid")
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.border = Border(
                    left=Side(style='medium', color='000000'),
                    right=Side(style='medium', color='000000'),
                    top=Side(style='medium', color='000000'),
                    bottom=Side(style='medium', color='000000')
                )
                
                # القيمة
                value_cell = info_sheet.cell(kpi_row+3, col)
                info_sheet.merge_cells(start_row=kpi_row+3, start_column=col, end_row=kpi_row+3, end_column=col+1)
                value_cell.value = kpi['value']
                value_cell.font = Font(name="Arial", size=18, bold=True, color=kpi['color'])
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                value_cell.border = Border(
                    left=Side(style='medium', color='000000'),
                    right=Side(style='medium', color='000000'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='medium', color='000000')
                )
                
                info_sheet.row_dimensions[kpi_row+3].height = 40
            
            # ==== جدول تفصيل الأقسام ====
            table_row = kpi_row + 6
            info_sheet.cell(table_row, 1).value = "📋 تفصيل الأقسام"
            info_sheet.cell(table_row, 1).font = Font(name="Arial", size=14, bold=True, color="1F4E78")
            
            # رؤوس الجدول
            headers = ['القسم', 'عدد الموظفين', 'الرواتب الأساسية', 'البدلات', 'الخصومات', 'المكافآت', 'صافي الرواتب']
            for col_idx, header in enumerate(headers, 1):
                cell = info_sheet.cell(table_row + 1, col_idx)
                cell.value = header
                cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # بيانات الأقسام
            data_row = table_row + 2
            for dept_name, dept_salaries in departments_data.items():
                dept_count = len(dept_salaries)
                dept_basic = sum(s['الراتب الأساسي'] for s in dept_salaries)
                dept_allowances = sum(s['البدلات'] for s in dept_salaries)
                dept_deductions = sum(s['الخصومات'] for s in dept_salaries)
                dept_bonus = sum(s['المكافآت'] for s in dept_salaries)
                dept_net = sum(s['صافي الراتب'] for s in dept_salaries)
                
                row_data = [dept_name, dept_count, dept_basic, dept_allowances, dept_deductions, dept_bonus, dept_net]
                for col_idx, value in enumerate(row_data, 1):
                    cell = info_sheet.cell(data_row, col_idx)
                    cell.value = value
                    if col_idx > 2:
                        cell.number_format = '#,##0 "ر.س"'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.fill = PatternFill(start_color="E7E6E6" if data_row % 2 == 0 else "FFFFFF", end_color="E7E6E6" if data_row % 2 == 0 else "FFFFFF", fill_type="solid")
                data_row += 1
            
            # ==== مخطط دائري لتوزيع الموظفين ====
            pie_chart = PieChart()
            pie_chart.title = "توزيع الموظفين حسب القسم"
            pie_chart.style = 10
            
            data = Reference(info_sheet, min_col=2, min_row=table_row+1, max_row=data_row-1)
            cats = Reference(info_sheet, min_col=1, min_row=table_row+2, max_row=data_row-1)
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(cats)
            pie_chart.height = 10
            pie_chart.width = 15
            
            info_sheet.add_chart(pie_chart, f"A{data_row + 2}")
            
            # ==== مخطط عمودي للرواتب ====
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.title = "مقارنة الرواتب حسب القسم"
            bar_chart.y_axis.title = 'المبلغ (ر.س)'
            bar_chart.x_axis.title = 'القسم'
            bar_chart.style = 11
            
            data = Reference(info_sheet, min_col=7, min_row=table_row+1, max_row=data_row-1)
            cats = Reference(info_sheet, min_col=1, min_row=table_row+2, max_row=data_row-1)
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)
            bar_chart.height = 10
            bar_chart.width = 15
            
            info_sheet.add_chart(bar_chart, f"H{data_row + 2}")
            
            # ==== معلومات إضافية ====
            info_row = data_row + 20
            info_sheet.cell(info_row, 1).value = "📅 تاريخ التصدير:"
            info_sheet.cell(info_row, 1).font = Font(name="Arial", size=10, bold=True)
            info_sheet.cell(info_row, 2).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            info_sheet.cell(info_row, 2).font = Font(name="Arial", size=10)
            
            # ضبط عرض الأعمدة
            for col in range(1, 11):
                info_sheet.column_dimensions[get_column_letter(col)].width = 15
            
            # تعيين الصفحة الأولى كصفحة نشطة
            writer.book.active = writer.book.worksheets[0]
        
        output.seek(0)
        return output
    
    except Exception as e:
        raise Exception(f"خطأ في إنشاء ملف Excel: {str(e)}")

def parse_document_excel(file):
    """
    Parse Excel file containing document data
    
    Args:
        file: The uploaded Excel file
        
    Returns:
        List of dictionaries containing document data
    """
    try:
        # Reset file pointer to beginning
        file.seek(0)
        
        # Read the Excel file explicitly using openpyxl engine
        # Try to detect if this is a report-style export with header rows
        try:
            df = pd.read_excel(file, engine='openpyxl')
            
            # Check if first row looks like a report title
            if len(df.columns) > 0 and str(df.columns[0]).startswith('تقرير'):
                # This is a report-style export, try reading from row 2 or find actual header
                df_test = pd.read_excel(file, engine='openpyxl', header=None)
                
                # Look for a row that contains document field names
                header_row = None
                for i in range(min(10, len(df_test))):  # Check first 10 rows
                    row_values = [str(val).lower() for val in df_test.iloc[i].values if pd.notna(val)]
                    if any('موظف' in val or 'employee' in val for val in row_values):
                        header_row = i
                        break
                
                if header_row is not None:
                    df = pd.read_excel(file, engine='openpyxl', header=header_row)
                else:
                    # Try standard document import format
                    df = pd.read_excel(file, engine='openpyxl', header=0)
            
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            # Fallback to standard reading
            df = pd.read_excel(file, engine='openpyxl')
        
        # Print column names for debugging
        print(f"Document Excel columns: {df.columns.tolist()}")
        
        # Create a mapping for column detection - include export format columns
        column_mappings = {
            'employee_id': ['employee_id', 'employee id', 'emp id', 'employee number', 'emp no', 'emp.id', 'emp.no', 'رقم الموظف', 'معرف الموظف', 'الرقم الوظيفي', 'موظف', 'اسم الموظف'],
            'document_type': ['document_type', 'document type', 'type', 'doc type', 'نوع الوثيقة', 'نوع المستند', 'النوع', 'نوع الوثيقة'],
            'document_number': ['document_number', 'document no', 'doc number', 'doc no', 'رقم الوثيقة', 'رقم المستند'],
            'issue_date': ['issue_date', 'issue date', 'start date', 'تاريخ الإصدار', 'تاريخ البدء'],
            'expiry_date': ['expiry_date', 'expiry date', 'end date', 'valid until', 'تاريخ الانتهاء', 'صالح حتى'],
            'notes': ['notes', 'comments', 'remarks', 'ملاحظات', 'تعليقات']
        }
        
        # Map columns to their normalized names
        detected_columns = {}
        for col in df.columns:
            if isinstance(col, datetime):
                continue
                
            col_str = str(col).lower().strip()
            
            # Check for exact column name or common variations
            for field, variations in column_mappings.items():
                if col_str in variations or any(var in col_str for var in variations):
                    detected_columns[field] = col
                    print(f"Detected '{field}' column: {col}")
                    break
        
        # Handle special case for Excel files with specific column names
        explicit_mappings = {
            'Employee ID': 'employee_id',
            'Document Type': 'document_type',
            'Document Number': 'document_number',
            'Issue Date': 'issue_date',
            'Expiry Date': 'expiry_date',
            'Notes': 'notes'
        }
        
        for excel_col, field in explicit_mappings.items():
            if excel_col in df.columns:
                detected_columns[field] = excel_col
                print(f"Explicitly mapped '{excel_col}' to '{field}'")
        
        # Print final column mapping
        print(f"Final document column mapping: {detected_columns}")
        
        # تقسيم الحقول المطلوبة إلى أساسية وغير أساسية
        essential_fields = ['employee_id', 'document_type', 'document_number']  # الحقول الأساسية التي يجب توفرها
        other_fields = ['issue_date', 'expiry_date', 'notes']  # الحقول التي يمكن إضافة قيم افتراضية لها
        
        # التحقق من الحقول الأساسية
        missing_essential = [field for field in essential_fields if field not in detected_columns]
        if missing_essential:
            missing_str = ", ".join(missing_essential)
            raise ValueError(f"Required columns missing: {missing_str}. Available columns: {[c for c in df.columns if not isinstance(c, datetime)]}")
        
        # بالنسبة للحقول غير الأساسية المفقودة، سننشئ أعمدة وهمية تحتوي على قيم افتراضية
        for field in other_fields:
            if field not in detected_columns:
                print(f"Warning: Creating default column for: {field}")
                dummy_column_name = f"__{field}__default"
                # إذا كان الحقل هو تاريخ، نضيف تاريخ افتراضي (اليوم للإصدار، وبعد سنة للانتهاء)
                if field == 'issue_date':
                    default_value = datetime.now()
                elif field == 'expiry_date':
                    default_value = datetime.now() + timedelta(days=365)
                else:
                    default_value = ''  # للملاحظات
                
                df[dummy_column_name] = default_value
                detected_columns[field] = dummy_column_name  # تعيين العمود الوهمي للحقل
        
        # Process each row
        documents = []
        for idx, row in df.iterrows():
            try:
                # Skip completely empty rows
                if row.isnull().all():
                    continue
                
                # Get employee_id field
                emp_id_col = detected_columns['employee_id']
                emp_id = row[emp_id_col]
                
                # Skip rows with missing employee_id
                if pd.isna(emp_id):
                    print(f"Skipping row {idx+1} due to missing employee ID")
                    continue
                
                # Try to convert employee_id to integer
                try:
                    employee_id = int(emp_id)
                except (ValueError, TypeError):
                    # If not convertible to int, use as string (could be employee code)
                    employee_id = str(emp_id).strip()
                
                # Get document type and number
                doc_type_col = detected_columns['document_type']
                doc_type = row[doc_type_col]
                
                doc_number_col = detected_columns['document_number']
                doc_number = row[doc_number_col]
                
                # Skip rows with missing document type or number
                if pd.isna(doc_type) or pd.isna(doc_number):
                    print(f"Skipping row {idx+1} due to missing document type or number")
                    continue
                
                # Get dates and parse them
                issue_date_col = detected_columns['issue_date']
                expiry_date_col = detected_columns['expiry_date']
                
                # تعامل مع التواريخ المفقودة - استخدام تاريخ اليوم للإصدار وبعد سنة للانتهاء
                if pd.isna(row[issue_date_col]):
                    print(f"Row {idx+1}: Using today's date for missing issue date")
                    issue_date_val = datetime.now()
                else:
                    issue_date_val = row[issue_date_col]
                    
                if pd.isna(row[expiry_date_col]):
                    print(f"Row {idx+1}: Using date one year from today for missing expiry date")
                    expiry_date_val = datetime.now() + timedelta(days=365)
                else:
                    expiry_date_val = row[expiry_date_col]
                
                try:
                    # Handle different date formats and convert to datetime
                    if isinstance(issue_date_val, datetime):
                        issue_date = issue_date_val
                    else:
                        issue_date = parse_date(str(issue_date_val))
                        
                    if isinstance(expiry_date_val, datetime):
                        expiry_date = expiry_date_val
                    else:
                        expiry_date = parse_date(str(expiry_date_val))
                    
                    # استخدام تواريخ افتراضية في حالة فشل تحليل التواريخ
                    if not issue_date:
                        print(f"Row {idx+1}: Using today's date due to invalid issue date format")
                        issue_date = datetime.now()
                        
                    if not expiry_date:
                        print(f"Row {idx+1}: Using date one year from today due to invalid expiry date format")
                        expiry_date = datetime.now() + timedelta(days=365)
                        
                except (ValueError, TypeError) as e:
                    print(f"Row {idx+1}: Date parsing error: {str(e)}, using default dates")
                    issue_date = datetime.now()
                    expiry_date = datetime.now() + timedelta(days=365)
                
                # Create document dictionary
                document = {
                    'employee_id': employee_id,
                    'document_type': str(doc_type).strip(),
                    'document_number': str(doc_number).strip(),
                    'issue_date': issue_date,
                    'expiry_date': expiry_date
                }
                
                # Add notes if present
                if 'notes' in detected_columns and not pd.isna(row[detected_columns['notes']]):
                    document['notes'] = str(row[detected_columns['notes']])
                
                print(f"Processed document for employee ID: {employee_id}, type: {document['document_type']}")
                documents.append(document)
                
            except Exception as e:
                print(f"Error processing document row {idx+1}: {str(e)}")
                # Continue to next row instead of failing the entire import
                continue
        
        if not documents:
            raise ValueError("No valid document records found in the Excel file")
            
        return documents
    
    except Exception as e:
        import traceback
        print(f"Error parsing document Excel: {str(e)}")
        print(traceback.format_exc())
        raise Exception(f"Error parsing document Excel file: {str(e)}")

def export_employee_attendance_to_excel(employee, month=None, year=None):
    """
    تصدير بيانات الحضور لموظف معين إلى ملف إكسل
    
    Args:
        employee: كائن الموظف
        month: الشهر (اختياري)
        year: السنة (اختياري)
        
    Returns:
        BytesIO object containing the Excel file
    """
    try:
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        
        # اختصار اسم ورقة Excel لتكون أقل من 31 حرفاً (الحد الأقصى المسموح به)
        sheet_name = f"Attendance-{employee.employee_id}"
        worksheet = workbook.add_worksheet(sheet_name)
        
        # تنسيقات الخلايا
        header_format = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#4F81BD',
            'font_color': 'white',
            'border': 1
        })
        
        date_header_format = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#4F81BD',
            'font_color': 'white',
            'border': 1,
            'text_wrap': True  # لسماح النص بالانتقال للسطر التالي
        })
        
        normal_format = workbook.add_format({
            'align': 'center',
            'border': 1
        })
        
        present_format = workbook.add_format({
            'align': 'center',
            'bg_color': '#C6EFCE',
            'font_color': '#006100',
            'border': 1
        })
        
        absent_format = workbook.add_format({
            'align': 'center',
            'bg_color': '#FFC7CE',
            'font_color': '#9C0006',
            'border': 1
        })
        
        leave_format = workbook.add_format({
            'align': 'center',
            'bg_color': '#FFEB9C',
            'font_color': '#9C5700',
            'border': 1
        })
        
        sick_format = workbook.add_format({
            'align': 'center',
            'bg_color': '#FFCC99',
            'font_color': '#974706',
            'border': 1
        })
        
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # تحديد الشهر والسنة إذا لم يتم توفيرهما
        current_date = datetime.now()
        if not year:
            year = current_date.year
        if not month:
            month = current_date.month
            
        # الحصول على عدد أيام الشهر
        _, days_in_month = monthrange(year, month)
        
        # إنشاء عنوان الملف
        title = f"سجل حضور {employee.name}"
        
        # إضافة عنوان
        worksheet.merge_range('A1:H1', title, title_format)
        
        # تحديد أسماء الأعمدة الثابتة وترتيبها كما في الصورة
        col_headers = ["Name", "ID Number", "Emp. No.", "Job Title", "Location", "Project", "Total"]
        
        # كتابة العناوين الرئيسية
        for col_idx, header in enumerate(col_headers):
            worksheet.write(2, col_idx, header, header_format)
        
        # تحديد نطاق التواريخ للشهر المحدد
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        # استخدام استعلام أكثر فعالية للحصول على سجلات الحضور
        from app import db
        from models import Attendance
        
        # استعلام من قاعدة البيانات مباشرة
        attendances = db.session.query(Attendance).filter(
            Attendance.employee_id == employee.id,
            Attendance.date >= start_date,
            Attendance.date <= end_date
        ).order_by(Attendance.date).all()
        
        # تخزين بيانات الحضور في قاموس للوصول السريع وإنشاء قائمة بالتواريخ الفعلية للحضور
        attendance_data = {}
        date_list = []  # فقط التواريخ التي يوجد بها سجلات حضور
        
        for attendance in attendances:
            # تخزين معلومات الحضور
            attendance_data[attendance.date] = {
                'status': attendance.status,
                'notes': attendance.notes if hasattr(attendance, 'notes') else None
            }
            # إضافة التاريخ إلى قائمة التواريخ
            date_list.append(attendance.date)
        
        # إعداد أيام الأسبوع
        weekdays = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        
        # كتابة عناوين أيام الأسبوع
        first_date_col = len(col_headers)
        
        # إنشاء الصف الأول بعناوين أيام الأسبوع
        for col_idx, date in enumerate(date_list):
            # تنسيق عنوان اليوم ليظهر كما في الصورة المرفقة: يوم الأسبوع مع التاريخ (مثال: Mon 01/04/2025)
            day_of_week = weekdays[date.weekday()]
            date_str = date.strftime("%d/%m/%Y")
            day_header = f"{day_of_week}\n{date_str}"
            
            col = first_date_col + col_idx
            worksheet.write(2, col, day_header, date_header_format)
        
        # كتابة معلومات الموظف والحضور
        row = 3  # البدء من الصف الرابع بعد عنوان الجدول
        
        # استخراج معلومات لموقع العمل والمشروع
        location = "AL QASSIM"  # قيمة افتراضية أو استخراجها من الموظف
        if hasattr(employee, 'departments') and employee.departments:
            location = employee.departments[0].name[:20]  # استخدام اسم أول قسم كموقع
            
        project = "ARAMEX"  # قيمة افتراضية، يمكن استخراجها من بيانات الموظف
        
        # عدد الحضور
        present_days = 0
        
        # كتابة معلومات الموظف
        worksheet.write(row, 0, employee.name, normal_format)  # Name
        worksheet.write(row, 1, employee.national_id or "", normal_format)  # ID Number
        worksheet.write(row, 2, employee.employee_id or "", normal_format)  # Emp. No.
        worksheet.write(row, 3, employee.job_title or "courier", normal_format)  # Job Title
        worksheet.write(row, 4, location, normal_format)  # Location
        worksheet.write(row, 5, project, normal_format)  # Project
        
        # كتابة سجلات الحضور لكل يوم
        for col_idx, date in enumerate(date_list):
            col = first_date_col + col_idx  # بداية أعمدة التواريخ
            cell_value = ""  # القيمة الافتراضية فارغة
            cell_format = normal_format
            
            if date in attendance_data:
                att_data = attendance_data[date]
                
                if att_data['status'] == 'present':
                    cell_value = "P"  # استخدام حرف P للحضور
                    cell_format = present_format
                    present_days += 1
                elif att_data['status'] == 'absent':
                    cell_value = "A"  # استخدام حرف A للغياب
                    cell_format = absent_format
                elif att_data['status'] == 'leave':
                    cell_value = "L"  # استخدام حرف L للإجازة
                    cell_format = leave_format
                elif att_data['status'] == 'sick':
                    cell_value = "S"  # استخدام حرف S للمرض
                    cell_format = sick_format
            # لا نحتاج للحالة else هنا لأن date_list تحتوي فقط على التواريخ التي لها سجلات حضور فعلية
            
            worksheet.write(row, col, cell_value, cell_format)
        
        # كتابة إجمالي أيام الحضور
        worksheet.write(row, 6, present_days, normal_format)  # Total
        
        # إضافة تفسير للرموز المستخدمة في صفحة منفصلة
        legend_sheet = workbook.add_worksheet('دليل الرموز')
        
        # تنسيق العناوين
        legend_title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # تنسيق الشرح
        description_format = workbook.add_format({
            'align': 'right',
            'valign': 'vcenter',
            'text_wrap': True
        })
        
        # كتابة العنوان
        legend_sheet.merge_range('A1:B1', 'دليل رموز الحضور والغياب', legend_title_format)
        
        # ضبط عرض الأعمدة
        legend_sheet.set_column(0, 0, 10)
        legend_sheet.set_column(1, 1, 40)
        
        # إضافة تفسير الرموز
        legend_sheet.write(2, 0, 'P', present_format)
        legend_sheet.write(2, 1, 'حاضر (Present)', description_format)
        
        legend_sheet.write(3, 0, 'A', absent_format)
        legend_sheet.write(3, 1, 'غائب (Absent)', description_format)
        
        legend_sheet.write(4, 0, 'L', leave_format)
        legend_sheet.write(4, 1, 'إجازة (Leave)', description_format)
        
        legend_sheet.write(5, 0, 'S', sick_format)
        legend_sheet.write(5, 1, 'مرضي (Sick Leave)', description_format)
        
        # ضبط عرض الأعمدة في الصفحة الرئيسية
        worksheet.set_column(0, 0, 30)  # عمود الاسم
        worksheet.set_column(1, 1, 15)  # عمود ID Number
        worksheet.set_column(2, 2, 10)  # عمود Emp. No.
        worksheet.set_column(3, 3, 15)  # عمود Job Title
        worksheet.set_column(4, 4, 13)  # عمود Location
        worksheet.set_column(5, 5, 13)  # عمود Project
        worksheet.set_column(6, 6, 8)   # عمود Total
        worksheet.set_column(first_date_col, first_date_col + len(date_list) - 1, 5)  # أعمدة التواريخ
        
        workbook.close()
        output.seek(0)
        return output
        
    except Exception as e:
        import traceback
        print(f"Error generating employee attendance Excel file: {str(e)}")
        print(traceback.format_exc())
        raise

def export_attendance_by_department(employees, attendances, start_date, end_date=None):
    """
    تصدير بيانات الحضور إلى ملف إكسل في صيغة جدول
    حيث تكون معلومات الموظفين في الأعمدة الأولى
    وتواريخ الحضور في الأعمدة الباقية مع استخدام P للحضور

    Args:
        employees: قائمة بجميع الموظفين
        attendances: قائمة بسجلات الحضور
        start_date: تاريخ البداية
        end_date: تاريخ النهاية (اختياري، إذا لم يتم تحديده سيتم استخدام تاريخ البداية فقط)

    Returns:
        BytesIO: كائن يحتوي على ملف اكسل
    """
    try:
        output = BytesIO()
        
        # إنشاء ملف إكسل جديد باستخدام xlsxwriter
        workbook = xlsxwriter.Workbook(output)
        
        # تعريف التنسيقات
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#00B0B0',  # لون أخضر فاتح مائل للأزرق
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })
        
        date_header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#00B0B0',  # لون أخضر فاتح مائل للأزرق
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })
        
        normal_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        right_aligned_format = workbook.add_format({
            'border': 1,
            'align': 'right',
            'valign': 'vcenter'
        })
        
        present_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'bold': True,
            'font_color': '#006100'  # اللون الأخضر لحرف P
        })
        
        absent_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'bold': True,
            'font_color': '#FF0000'  # اللون الأحمر لحرف A
        })
        
        leave_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_color': '#FF9900'  # اللون البرتقالي لحرف L
        })
        
        sick_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_color': '#0070C0'  # اللون الأزرق لحرف S
        })
        
        # تحديد الفترة الزمنية
        if end_date is None:
            end_date = start_date
        
        # تحديد قائمة التواريخ
        date_list = []
        current_date = start_date
        while current_date <= end_date:
            date_list.append(current_date)
            current_date += timedelta(days=1)
        
        # تنظيم الموظفين حسب الأقسام
        departments = {}
        for employee in employees:
            dept_name = ', '.join([dept.name for dept in employee.departments]) if employee.departments else 'بدون قسم'
            if dept_name not in departments:
                departments[dept_name] = []
            departments[dept_name].append(employee)
        
        # تنظيم بيانات الحضور حسب الموظفين
        attendance_data = {}
        for attendance in attendances:
            emp_id = attendance.employee_id
            if emp_id not in attendance_data:
                attendance_data[emp_id] = {}
            
            # تخزين حالة الحضور لهذا اليوم
            attendance_data[emp_id][attendance.date] = {
                'status': attendance.status
            }
        
        # عمل قائمة بأيام الأسبوع للعناوين
        weekdays = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        
        # إنشاء ورقة عمل لكل قسم
        for dept_name, dept_employees in departments.items():
            # تحويل الاسم ليكون صالحًا كاسم ورقة Excel
            sheet_name = dept_name[:31]  # Excel يدعم حد أقصى 31 حرف لأسماء الأوراق
            worksheet = workbook.add_worksheet(sheet_name)
            
            # تحديد أسماء الأعمدة الثابتة وترتيبها كما في الصورة
            col_headers = ["Name", "ID Number", "Emp. No.", "Job Title", "No. Mobile", "Car", "Location", "Project", "Total"]
            
            # كتابة العناوين الرئيسية
            for col_idx, header in enumerate(col_headers):
                worksheet.write(1, col_idx, header, header_format)
            
            # كتابة عناوين أيام الأسبوع
            first_date_col = len(col_headers)
            
            # إنشاء الصف الأول بعناوين أيام الأسبوع
            for col_idx, date in enumerate(date_list):
                # تنسيق عنوان اليوم ليظهر كما في الصورة المرفقة: يوم الأسبوع مع التاريخ (مثال: Mon 01/04/2025)
                day_of_week = weekdays[date.weekday()]
                date_str = date.strftime("%d/%m/%Y")
                day_header = f"{day_of_week}\n{date_str}"
                
                col = first_date_col + col_idx
                worksheet.write(0, col, day_header, date_header_format)
                
                # لم نعد بحاجة لكتابة التاريخ في صف منفصل لأننا دمجناه مع اسم اليوم
            
            # ضبط عرض الأعمدة
            worksheet.set_column(0, 0, 30)  # عمود الاسم
            worksheet.set_column(1, 1, 15)  # عمود ID Number
            worksheet.set_column(2, 2, 10)  # عمود Emp. No.
            worksheet.set_column(3, 3, 15)  # عمود Job Title
            worksheet.set_column(4, 4, 15)  # عمود No. Mobile
            worksheet.set_column(5, 5, 13)  # عمود Car
            worksheet.set_column(6, 6, 13)  # عمود Location
            worksheet.set_column(7, 7, 13)  # عمود Project
            worksheet.set_column(8, 8, 8)   # عمود Total
            worksheet.set_column(first_date_col, first_date_col + len(date_list) - 1, 5)  # أعمدة التواريخ
            
            # كتابة بيانات الموظفين وسجلات الحضور
            for row_idx, employee in enumerate(sorted(dept_employees, key=lambda e: e.name)):
                row = row_idx + 2  # صف البيانات (بعد صفي العناوين)
                
                # استخراج رقم الهاتف من المعلومات الإضافية إن وجد
                phone_number = ""
                if hasattr(employee, 'phone'):
                    phone_number = employee.phone
                
                # كتابة معلومات الموظف
                worksheet.write(row, 0, employee.name, normal_format)  # Name
                worksheet.write(row, 1, employee.national_id or "", normal_format)  # ID Number
                worksheet.write(row, 2, employee.employee_id or "", normal_format)  # Emp. No.
                worksheet.write(row, 3, employee.job_title or "courier", normal_format)  # Job Title
                worksheet.write(row, 4, phone_number, normal_format)  # No. Mobile
                
                # معلومات إضافية (قد تحتاج لتكييفها حسب هيكل البيانات الفعلي)
                car = ""  # يمكن إضافة منطق لاستخراج رقم السيارة إن وجد
                worksheet.write(row, 5, car, normal_format)  # Car
                
                # أحضر اسم الموقع من القسم
                location = "AL QASSIM"  # قيمة افتراضية أو استخراجها من الموظف
                if employee.departments:
                    location = employee.departments[0].name[:20]  # استخدام اسم أول قسم كموقع
                worksheet.write(row, 6, location, normal_format)  # Location
                
                # اسم المشروع
                project = "ARAMEX"  # قيمة افتراضية، يمكن استخراجها من بيانات الموظف
                worksheet.write(row, 7, project, normal_format)  # Project
                
                # عداد للحضور
                present_days = 0
                
                # كتابة سجلات الحضور لكل يوم
                for col_idx, date in enumerate(date_list):
                    col = first_date_col + col_idx  # بداية أعمدة التواريخ
                    cell_value = ""  # القيمة الافتراضية فارغة
                    cell_format = normal_format
                    
                    if employee.id in attendance_data and date in attendance_data[employee.id]:
                        att_data = attendance_data[employee.id][date]
                        
                        if att_data['status'] == 'present':
                            cell_value = "P"  # استخدام حرف P للحضور
                            cell_format = present_format
                            present_days += 1
                        elif att_data['status'] == 'absent':
                            cell_value = "A"  # استخدام حرف A للغياب
                            cell_format = absent_format
                        elif att_data['status'] == 'leave':
                            cell_value = "L"  # استخدام حرف L للإجازة
                            cell_format = leave_format
                        elif att_data['status'] == 'sick':
                            cell_value = "S"  # استخدام حرف S للمرض
                            cell_format = sick_format
                    else:
                        # إذا لم يوجد سجل لهذا اليوم، نفترض أنه حاضر (كما في الصورة المرفقة)
                        cell_value = "P"
                        cell_format = present_format
                        present_days += 1
                    
                    worksheet.write(row, col, cell_value, cell_format)
                
                # كتابة إجمالي أيام الحضور
                worksheet.write(row, 8, present_days, normal_format)  # Total
        
        # إضافة تفسير للرموز المستخدمة في صفحة منفصلة
        legend_sheet = workbook.add_worksheet('دليل الرموز')
        
        # تنسيق العناوين
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # تنسيق الشرح
        description_format = workbook.add_format({
            'align': 'right',
            'valign': 'vcenter',
            'text_wrap': True
        })
        
        # كتابة العنوان
        legend_sheet.merge_range('A1:B1', 'دليل رموز الحضور والغياب', title_format)
        
        # ضبط عرض الأعمدة
        legend_sheet.set_column(0, 0, 10)
        legend_sheet.set_column(1, 1, 40)
        
        # إضافة تفسير الرموز
        legend_sheet.write(2, 0, 'P', present_format)
        legend_sheet.write(2, 1, 'حاضر (Present)', description_format)
        
        legend_sheet.write(3, 0, 'A', absent_format)
        legend_sheet.write(3, 1, 'غائب (Absent)', description_format)
        
        legend_sheet.write(4, 0, 'L', leave_format)
        legend_sheet.write(4, 1, 'إجازة (Leave)', description_format)
        
        legend_sheet.write(5, 0, 'S', sick_format)
        legend_sheet.write(5, 1, 'مرضي (Sick Leave)', description_format)
        
        # إغلاق الملف وإعادة المخرجات
        workbook.close()
        output.seek(0)
        return output
    
    except Exception as e:
        import traceback
        print(f"Error generating attendance Excel file: {str(e)}")
        print(traceback.format_exc())
        raise
