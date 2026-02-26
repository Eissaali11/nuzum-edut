from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ExcelStyles:
    """Centralized styling for Excel exports to ensure Zero-Regression"""
    
    @staticmethod
    def get_title_font():
        return Font(bold=True, color="FFFFFF", size=18, name="Calibri")
        
    @staticmethod
    def get_title_fill():
        return PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        
    @staticmethod
    def get_header_font():
        return Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        
    @staticmethod
    def get_header_fill():
        return PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
    @staticmethod
    def get_stat_fill():
        return PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
    @staticmethod
    def get_center_align():
        return Alignment(horizontal="center", vertical="center")
        
    @staticmethod
    def get_right_align():
        return Alignment(horizontal="right", vertical="center")
        
    @staticmethod
    def get_border():
        return Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    @staticmethod
    def apply_rtl(worksheet, engine="openpyxl"):
        """Apply Right-to-Left orientation for Arabic support"""
        if engine == "openpyxl":
            worksheet.sheet_view.rightToLeft = True
        elif engine == "xlsxwriter":
            worksheet.right_to_left()

    @staticmethod
    def apply_header_style(worksheet, headers, row=1):
        """Apply standard header styling to a row"""
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col_num)
            cell.value = header
            cell.font = ExcelStyles.get_header_font()
            cell.fill = ExcelStyles.get_header_fill()
            cell.alignment = ExcelStyles.get_center_align()
            cell.border = ExcelStyles.get_border()

