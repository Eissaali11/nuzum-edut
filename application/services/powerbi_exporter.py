"""
Power BI Excel Exporter
=======================
تصدير البيانات إلى ملف Excel متعدد الأوراق لاستخدامه في Power BI
"""
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from application.services.bi_engine import bi_engine


class PowerBIExporter:
    """مُصدّر البيانات لـ Power BI"""
    
    def __init__(self):
        """تهيئة المُصدّر"""
        self.workbook = Workbook()
        # حذف الورقة الافتراضية
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # تنسيقات
        self.header_fill = PatternFill(start_color='00D4AA', end_color='00D4AA', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _style_header_row(self, sheet, row_num: int = 1):
        """تنسيق صف العناوين"""
        for cell in sheet[row_num]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.border
    
    def _auto_size_columns(self, sheet):
        """ضبط عرض الأعمدة تلقائياً"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def add_dimension_employees_sheet(self):
        """إضافة ورقة DIM_Employees"""
        data = bi_engine.get_dimension_employees()
        
        if not data:
            return
        
        # إنشاء DataFrame
        df = pd.DataFrame(data)
        
        # إنشاء ورقة العمل
        sheet = self.workbook.create_sheet('DIM_Employees')
        
        # كتابة العناوين
        for col_num, column_title in enumerate(df.columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
        
        # كتابة البيانات
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
        
        # تنسيق
        self._style_header_row(sheet)
        self._auto_size_columns(sheet)
        
        # تجميد الصف الأول
        sheet.freeze_panes = 'A2'
    
    def add_dimension_vehicles_sheet(self):
        """إضافة ورقة DIM_Vehicles"""
        data = bi_engine.get_dimension_vehicles()
        
        if not data:
            return
        
        df = pd.DataFrame(data)
        sheet = self.workbook.create_sheet('DIM_Vehicles')
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
        
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
        
        self._style_header_row(sheet)
        self._auto_size_columns(sheet)
        sheet.freeze_panes = 'A2'
    
    def add_dimension_departments_sheet(self):
        """إضافة ورقة DIM_Departments"""
        data = bi_engine.get_dimension_departments()
        
        if not data:
            return
        
        df = pd.DataFrame(data)
        sheet = self.workbook.create_sheet('DIM_Departments')
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
        
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
        
        self._style_header_row(sheet)
        self._auto_size_columns(sheet)
        sheet.freeze_panes = 'A2'
    
    def add_fact_financials_sheet(self):
        """إضافة ورقة FACT_Financials"""
        data = bi_engine.get_fact_financials()
        
        if not data:
            return
        
        df = pd.DataFrame(data)
        sheet = self.workbook.create_sheet('FACT_Financials')
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
        
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
        
        self._style_header_row(sheet)
        self._auto_size_columns(sheet)
        sheet.freeze_panes = 'A2'
    
    def add_fact_maintenance_sheet(self):
        """إضافة ورقة FACT_Maintenance"""
        data = bi_engine.get_fact_maintenance()
        
        if not data:
            return
        
        df = pd.DataFrame(data)
        sheet = self.workbook.create_sheet('FACT_Maintenance')
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
        
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
        
        self._style_header_row(sheet)
        self._auto_size_columns(sheet)
        sheet.freeze_panes = 'A2'
    
    def add_fact_attendance_sheet(self):
        """إضافة ورقة FACT_Attendance"""
        data = bi_engine.get_fact_attendance()
        
        if not data:
            return
        
        df = pd.DataFrame(data)
        sheet = self.workbook.create_sheet('FACT_Attendance')
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
        
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
        
        self._style_header_row(sheet)
        self._auto_size_columns(sheet)
        sheet.freeze_panes = 'A2'
    
    def add_kpi_summary_sheet(self):
        """إضافة ورقة KPI Summary"""
        kpis = bi_engine.get_kpi_summary()
        
        sheet = self.workbook.create_sheet('KPI_Summary')
        
        # العنوان
        sheet['A1'] = 'KPI Metric'
        sheet['B1'] = 'Value'
        
        # البيانات
        row = 2
        for key, value in kpis.items():
            sheet[f'A{row}'] = key.replace('_', ' ').title()
            sheet[f'B{row}'] = value
            row += 1
        
        self._style_header_row(sheet)
        self._auto_size_columns(sheet)
        sheet.freeze_panes = 'A2'
    
    def add_metadata_sheet(self):
        """إضافة ورقة Metadata"""
        sheet = self.workbook.create_sheet('Metadata', 0)  # أول ورقة
        
        sheet['A1'] = 'Nuzum Business Intelligence Export'
        sheet['A1'].font = Font(bold=True, size=14)
        
        sheet['A3'] = 'Export Information'
        sheet['A3'].font = Font(bold=True)
        
        sheet['A4'] = 'Export Date:'
        sheet['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        sheet['A5'] = 'Data As Of:'
        sheet['B5'] = bi_engine.today.isoformat()
        
        sheet['A6'] = 'Export Type:'
        sheet['B6'] = 'Power BI Star Schema'
        
        sheet['A8'] = 'Sheets Included:'
        sheet['A8'].font = Font(bold=True)
        
        sheets_info = [
            'DIM_Employees: Employee dimension (full details, projects)',
            'DIM_Vehicles: Vehicle dimension (maintenance status, regions)',
            'DIM_Departments: Department dimension',
            'FACT_Financials: Financial facts (salaries, bonuses, costs)',
            'FACT_Maintenance: Maintenance facts',
            'FACT_Attendance: Attendance facts',
            'KPI_Summary: Key Performance Indicators'
        ]
        
        for idx, info in enumerate(sheets_info, 9):
            sheet[f'A{idx}'] = info
        
        sheet['A17'] = 'Region Standardization:'
        sheet['A17'].font = Font(bold=True)
        sheet['A18'] = 'All locations have been standardized to English region names'
        sheet['A19'] = 'Compatible with Power BI Map Visuals'
        
        self._auto_size_columns(sheet)
    
    def generate(self) -> BytesIO:
        """
        توليد ملف Excel كامل
        Returns: BytesIO object يحتوي على الملف
        """
        # إضافة جميع الأوراق
        self.add_metadata_sheet()
        self.add_dimension_employees_sheet()
        self.add_dimension_vehicles_sheet()
        self.add_dimension_departments_sheet()
        self.add_fact_financials_sheet()
        self.add_fact_maintenance_sheet()
        self.add_fact_attendance_sheet()
        self.add_kpi_summary_sheet()
        
        # حفظ في الذاكرة
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def get_filename(self) -> str:
        """الحصول على اسم الملف"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f'nuzum_powerbi_export_{timestamp}.xlsx'


def export_to_powerbi() -> tuple:
    """
    تصدير البيانات لـ Power BI
    يُرجع أحدث ملف Executive Report من instance/reports/
    Returns: (buffer, filename, mimetype)
    """
    from pathlib import Path
    import os
    
    # المجلد الذي يحتوي على الملفات المُولدة
    reports_dir = Path('instance/reports')
    
    # البحث عن أحدث ملف Executive_Report_*.xlsx
    if reports_dir.exists():
        report_files = sorted(
            reports_dir.glob('Executive_Report_*.xlsx'),
            key=os.path.getctime,
            reverse=True
        )
        
        if report_files:
            latest_file = report_files[0]
            
            # قراءة الملف إلى BytesIO buffer
            with open(latest_file, 'rb') as f:
                buffer = BytesIO(f.read())
            
            filename = latest_file.name
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            return buffer, filename, mimetype
    
    # fallback: إنشاء ملف جديد إذا لم يوجد ملف حالي
    exporter = PowerBIExporter()
    buffer = exporter.generate()
    filename = exporter.get_filename()
    mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return buffer, filename, mimetype
