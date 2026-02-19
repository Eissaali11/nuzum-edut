"""
نظام الألوان والتنسيقات الاحترافية
===================================
تعريف جميع أنماط Excel المستخدمة في التقارير
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dataclasses import dataclass


@dataclass
class ColorPalette:
    """لوحة الألوان الاحترافية"""
    # الألوان الأساسية
    NAVY_BLUE = '0D1117'      # أزرق بحري
    EMERALD = '00D4AA'         # أخضر زمردي
    CYAN = '00D4FF'             # سماوي
    GOLD = 'FFD700'             # ذهب
    LIGHT_GRAY = 'ECEFF1'       # رمادي فاتح
    WHITE = 'FFFFFF'            # أبيض
    
    # ألوان محايدة
    DARK_GRAY = '37474F'        # رمادي غامق
    MEDIUM_GRAY = '90A4AE'      # رمادي متوسط
    
    # ألوان الإشارات
    SUCCESS = '4CAF50'          # أخضر النجاح
    WARNING = 'FFA726'          # برتقالي التحذير
    DANGER = 'EF5350'           # أحمر الخطر
    INFO = '29B6F6'             # أزرق المعلومات


class ExcelStyles:
    """فئة توفر جميع أنماط Excel"""
    
    # ==================== الخطوط ====================
    @staticmethod
    def header_font() -> Font:
        """خط رؤوس العناوين الرئيسية"""
        return Font(
            name='Calibri',
            size=14,
            bold=True,
            color=ColorPalette.WHITE
        )
    
    @staticmethod
    def subheader_font() -> Font:
        """خط العناوين الفرعية"""
        return Font(
            name='Calibri',
            size=12,
            bold=True,
            color=ColorPalette.DARK_GRAY
        )
    
    @staticmethod
    def title_font() -> Font:
        """خط العنوان الرئيسي للتقرير"""
        return Font(
            name='Calibri',
            size=18,
            bold=True,
            color=ColorPalette.NAVY_BLUE
        )
    
    @staticmethod
    def normal_font() -> Font:
        """خط النصوص العادية"""
        return Font(
            name='Calibri',
            size=11,
            color=ColorPalette.DARK_GRAY
        )
    
    @staticmethod
    def metric_font() -> Font:
        """خط الأرقام والمقاييس"""
        return Font(
            name='Calibri',
            size=12,
            bold=True,
            color=ColorPalette.NAVY_BLUE
        )
    
    # ==================== الملء والحشو ====================
    @staticmethod
    def header_fill() -> PatternFill:
        """حشو رؤوس الجداول"""
        return PatternFill(
            start_color=ColorPalette.NAVY_BLUE,
            end_color=ColorPalette.NAVY_BLUE,
            fill_type='solid'
        )
    
    @staticmethod
    def subheader_fill() -> PatternFill:
        """حشو العناوين الفرعية"""
        return PatternFill(
            start_color=ColorPalette.EMERALD,
            end_color=ColorPalette.EMERALD,
            fill_type='solid'
        )
    
    @staticmethod
    def accent_fill() -> PatternFill:
        """حشو الخلايا المهمة"""
        return PatternFill(
            start_color=ColorPalette.CYAN,
            end_color=ColorPalette.CYAN,
            fill_type='solid'
        )
    
    @staticmethod
    def metric_fill() -> PatternFill:
        """حشو خلايا المقاييس"""
        return PatternFill(
            start_color=ColorPalette.LIGHT_GRAY,
            end_color=ColorPalette.LIGHT_GRAY,
            fill_type='solid'
        )
    
    @staticmethod
    def success_fill() -> PatternFill:
        """حشو الحالة الناجحة"""
        return PatternFill(
            start_color=ColorPalette.SUCCESS,
            end_color=ColorPalette.SUCCESS,
            fill_type='solid'
        )
    
    @staticmethod
    def warning_fill() -> PatternFill:
        """حشو التحذير"""
        return PatternFill(
            start_color=ColorPalette.WARNING,
            end_color=ColorPalette.WARNING,
            fill_type='solid'
        )
    
    # ==================== المحاذاة ====================
    @staticmethod
    def center_alignment() -> Alignment:
        """محاذاة وسط"""
        return Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
    
    @staticmethod
    def right_alignment() -> Alignment:
        """محاذاة يمين (للنصوص العربية)"""
        return Alignment(
            horizontal='right',
            vertical='center',
            wrap_text=True
        )
    
    @staticmethod
    def left_alignment() -> Alignment:
        """محاذاة يسار"""
        return Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True
        )
    
    # ==================== الحدود ====================
    @staticmethod
    def thin_border() -> Border:
        """حد رفيع"""
        side = Side(style='thin', color=ColorPalette.MEDIUM_GRAY)
        return Border(
            left=side,
            right=side,
            top=side,
            bottom=side
        )
    
    @staticmethod
    def thick_border() -> Border:
        """حد سميك"""
        side = Side(style='medium', color=ColorPalette.NAVY_BLUE)
        return Border(
            left=side,
            right=side,
            top=side,
            bottom=side
        )
    
    # ==================== مجموعات الأنماط ====================
    @staticmethod
    def header_style() -> dict:
        """نمط العنوان الكامل"""
        return {
            'font': ExcelStyles.header_font(),
            'fill': ExcelStyles.header_fill(),
            'alignment': ExcelStyles.center_alignment(),
            'border': ExcelStyles.thin_border()
        }
    
    @staticmethod
    def subheader_style() -> dict:
        """نمط العنوان الفرعي الكامل"""
        return {
            'font': ExcelStyles.subheader_font(),
            'fill': ExcelStyles.subheader_fill(),
            'alignment': ExcelStyles.right_alignment(),
            'border': ExcelStyles.thin_border()
        }
    
    @staticmethod
    def metric_card_style() -> dict:
        """نمط بطاقة المقياس"""
        return {
            'font': ExcelStyles.metric_font(),
            'fill': ExcelStyles.metric_fill(),
            'alignment': ExcelStyles.center_alignment(),
            'border': ExcelStyles.thick_border()
        }
    
    @staticmethod
    def data_cell_style() -> dict:
        """نمط خلية البيانات العادية"""
        return {
            'font': ExcelStyles.normal_font(),
            'alignment': ExcelStyles.right_alignment(),
            'border': ExcelStyles.thin_border()
        }
    
    @staticmethod
    def number_cell_style() -> dict:
        """نمط خلية الأرقام"""
        style = ExcelStyles.data_cell_style()
        style['alignment'] = ExcelStyles.center_alignment()
        return style


class TextFormatters:
    """فئة معالجة صيغ الأرقام والنصوص"""
    
    # صيغ الأرقام
    CURRENCY_SAR = '[DBNum1][$￼-AR]#,##0.00;[DBNum1][$￼-AR]-#,##0.00'
    PERCENTAGE = '0.00%'
    NUMBER_WITH_COMMA = '#,##0'
    DECIMAL_TWO = '0.00'
    
    @staticmethod
    def format_currency(value: float) -> str:
        """تنسيق العملة SAR"""
        return f'﷼ {value:,.2f}'
    
    @staticmethod
    def format_percentage(value: float) -> str:
        """تنسيق النسبة المئوية"""
        return f'{value:.2f}%'
    
    @staticmethod
    def format_number(value: float) -> str:
        """تنسيق الأرقام برموز الفاصل"""
        return f'{value:,.0f}'
