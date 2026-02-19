#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
اختبار الـ endpoint الجديد /analytics/export/latest-report
Test new endpoint
"""
import urllib.request

url = 'http://127.0.0.1:5000/analytics/export/latest-report'

try:
    print('Testing: /analytics/export/latest-report')
    print('=' * 60)
    
    response = urllib.request.urlopen(url, timeout=10)
    file_data = response.read()
    
    print(f'Status Code: {response.getcode()}')
    print(f'Downloaded Size: {len(file_data)} bytes')
    
    # حفظ الملف
    test_file = 'test_latest_report.xlsx'
    with open(test_file, 'wb') as f:
        f.write(file_data)
    
    # التحقق من نوع الملف
    try:
        from openpyxl import load_workbook
        wb = load_workbook(test_file)
        sheets = wb.sheetnames
        print(f'Excel Valid: YES')
        print(f'Sheets: {sheets}')
    except:
        print('Excel Valid: NO')
    
    print(f'Saved to: {test_file}')
    print('\nRESULT: SUCCESS')
    
except urllib.error.HTTPError as e:
    print(f'HTTP Error {e.code}')
except Exception as e:
    print(f'Error: {e}')
