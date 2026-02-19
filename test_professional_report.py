"""
اختبار نظام تصدير Excel الاحترافي
==================================
اختبار شامل لجميع مكونات النظام
"""

import sys
from pathlib import Path

# إضافة المسار
sys.path.insert(0, str(Path(__file__).parent))

from application.services.excel.exporter import ExcelExporter
from application.services.excel.styles import ExcelStyles, ColorPalette
from application.services.excel.data_extractor import DataExtractor
from openpyxl import load_workbook
from io import BytesIO


class TestReportSystem:
    """فئة اختبار نظام التقارير"""
    
    def __init__(self):
        """تهيئة الاختبار"""
        self.exporter = ExcelExporter()
        self.styles = ExcelStyles()
        self.data_extractor = DataExtractor()
        self.passed = 0
        self.failed = 0
    
    def test_styles_module(self) -> bool:
        """اختبار وحدة الأنماط"""
        try:
            print("\n" + "="*60)
            print("اختبار وحدة الأنماط")
            print("="*60)
            
            # اختبار الألوان
            assert ColorPalette.NAVY_BLUE == '0D1117'
            print("✅ لوحة الألوان تعمل بشكل صحيح")
            
            # اختبار الخطوط
            header_font = self.styles.header_font()
            assert header_font.size == 14
            assert header_font.bold == True
            print("✅ الخطوط تم تعريفها بشكل صحيح")
            
            # اختبار الملء
            header_fill = self.styles.header_fill()
            assert header_fill.start_color is not None
            print("✅ الملء يعمل بشكل صحيح")
            
            # اختبار المحاذاة
            alignment = self.styles.center_alignment()
            assert alignment.horizontal == 'center'
            print("✅ المحاذاة بشكل صحيح")
            
            # اختبار الحدود
            border = self.styles.thin_border()
            assert border.left is not None
            print("✅ الحدود معرفة بشكل صحيح")
            
            self.passed += 1
            return True
        except Exception as e:
            print(f"❌ فشل الاختبار: {e}")
            self.failed += 1
            return False
    
    def test_data_extraction(self) -> bool:
        """اختبار استخراج البيانات"""
        try:
            print("\n" + "="*60)
            print("اختبار استخراج البيانات")
            print("="*60)
            
            # اختبار مقاييس لوحة القيادة
            metrics = self.data_extractor.get_dashboard_metrics()
            assert 'total_records' in metrics
            assert 'today_records' in metrics
            assert 'growth_rate' in metrics
            print("✅ مقاييس لوحة القيادة تم استخراجها")
            
            # اختبار توزيع الفئات
            categories = self.data_extractor.get_category_distribution()
            assert not categories.empty
            assert 'الفئة' in categories.columns
            print("✅ توزيع الفئات تم استخراجه")
            
            # اختبار الاتجاهات الشهرية
            trends = self.data_extractor.get_monthly_trends()
            assert not trends.empty
            assert 'الشهر' in trends.columns
            print("✅ الاتجاهات الشهرية تم استخراجها")
            
            # اختبار بيانات الأداء
            performance = self.data_extractor.get_performance_data()
            assert not performance.empty
            print("✅ بيانات الأداء تم استخراجها")
            
            # اختبار الأفضل أداءً
            top = self.data_extractor.get_top_performers(5)
            assert len(top) <= 5
            print("✅ الأفضل أداءً تم استخراجهم")
            
            self.passed += 1
            return True
        except Exception as e:
            print(f"❌ فشل الاختبار: {e}")
            self.failed += 1
            return False
    
    def test_report_generation(self) -> bool:
        """اختبار إنشاء التقرير"""
        try:
            print("\n" + "="*60)
            print("اختبار إنشاء التقرير")
            print("="*60)
            
            # إنشاء التقرير
            buffer, filename = self.exporter.generate_report()
            assert buffer is not None
            assert filename.startswith('Report_')
            assert filename.endswith('.xlsx')
            print(f"✅ تم إنشاء التقرير: {filename}")
            
            # التحقق من حجم الملف
            buffer_size = len(buffer.getvalue())
            assert buffer_size > 0
            print(f"✅ حجم الملف: {buffer_size} بايت")
            
            # التحقق من محتوى ملف Excel
            buffer.seek(0)
            workbook = load_workbook(buffer)
            sheets = workbook.sheetnames
            assert 'لوحة العمل' in sheets
            assert 'التفاصيل' in sheets
            print(f"✅ الأوراق الموجودة: {sheets}")
            
            # التحقق من محتوى الورقة الأولى
            dashboard = workbook['لوحة العمل']
            assert dashboard.max_row > 0
            print(f"✅ لوحة العمل تحتوي على {dashboard.max_row} صف")
            
            self.passed += 1
            return True
        except Exception as e:
            print(f"❌ فشل الاختبار: {e}")
            import traceback
            traceback.print_exc()
            self.failed += 1
            return False
    
    def test_export_buffer(self) -> bool:
        """اختبار تصدير Buffer"""
        try:
            print("\n" + "="*60)
            print("اختبار تصدير Buffer")
            print("="*60)
            
            # اختبار التصدير
            buffer, filename, mimetype = self.exporter.export_to_buffer()
            
            assert buffer is not None
            assert len(buffer.getvalue()) > 0
            print(f"✅ تم التصدير بنجاح")
            
            assert filename.endswith('.xlsx')
            print(f"✅ اسم الملف: {filename}")
            
            assert 'spreadsheet' in mimetype
            print(f"✅ نوع ملف MIME: {mimetype}")
            
            # التحقق من صحة Excel
            buffer.seek(0)
            wb = load_workbook(buffer)
            assert len(wb.sheetnames) >= 2
            print(f"✅ ملف Excel صحيح مع {len(wb.sheetnames)} أوراق")
            
            self.passed += 1
            return True
        except Exception as e:
            print(f"❌ فشل الاختبار: {e}")
            import traceback
            traceback.print_exc()
            self.failed += 1
            return False
    
    def test_latest_report_retrieval(self) -> bool:
        """اختبار استرجاع أحدث تقرير"""
        try:
            print("\n" + "="*60)
            print("اختبار استرجاع أحدث تقرير")
            print("="*60)
            
            # إنشاء تقرير أولاً
            self.exporter.export_to_buffer()
            
            # محاولة استرجاع أحدث تقرير
            result = self.exporter.get_latest_report()
            
            if result:
                buffer, filename, mimetype = result
                assert buffer is not None
                print(f"✅ تم استرجاع أحدث تقرير: {filename}")
                
                # التحقق من صحته
                buffer.seek(0)
                wb = load_workbook(buffer)
                assert len(wb.sheetnames) > 0
                print(f"✅ التقرير يحتوي على {len(wb.sheetnames)} أوراق")
            else:
                print("⚠️ لا توجد تقارير سابقة")
            
            self.passed += 1
            return True
        except Exception as e:
            print(f"❌ فشل الاختبار: {e}")
            import traceback
            traceback.print_exc()
            self.failed += 1
            return False
    
    def test_cleanup_old_reports(self) -> bool:
        """اختبار تنظيف التقارير القديمة"""
        try:
            print("\n" + "="*60)
            print("اختبار تنظيف التقارير القديمة")
            print("="*60)
            
            # إنشاء عدة تقارير
            for i in range(3):
                self.exporter.export_to_buffer()
            
            # عد الملفات
            report_files = list(self.exporter.reports_dir.glob('Report_*.xlsx'))
            print(f"  التقارير قبل التنظيف: {len(report_files)}")
            
            # تنظيف مع الحفاظ على 2 فقط
            self.exporter.cleanup_old_reports(keep_count=2)
            
            # التحقق
            remaining = list(self.exporter.reports_dir.glob('Report_*.xlsx'))
            print(f"  التقارير بعد التنظيف: {len(remaining)}")
            assert len(remaining) <= 2
            print("✅ تم تنظيف التقارير القديمة بنجاح")
            
            self.passed += 1
            return True
        except Exception as e:
            print(f"❌ فشل الاختبار: {e}")
            import traceback
            traceback.print_exc()
            self.failed += 1
            return False
    
    def run_all_tests(self) -> None:
        """تشغيل جميع الاختبارات"""
        print("\n" + "🚀 "*30)
        print("  اختبار نظام تصدير Excel الاحترافي")
        print("🚀 "*30)
        
        self.test_styles_module()
        self.test_data_extraction()
        self.test_report_generation()
        self.test_export_buffer()
        self.test_latest_report_retrieval()
        self.test_cleanup_old_reports()
        
        # ملخص النتائج
        self.print_summary()
    
    def print_summary(self) -> None:
        """طباعة ملخص الاختبارات"""
        total = self.passed + self.failed
        percentage = (self.passed / total * 100) if total > 0 else 0
        
        print("\n" + "="*60)
        print("ملخص الاختبارات")
        print("="*60)
        print(f"✅ نجح:     {self.passed}")
        print(f"❌ فشل:     {self.failed}")
        print(f"📊 الكلي:   {total}")
        print(f"📈 النسبة:  {percentage:.1f}%")
        print("="*60)
        
        if self.failed == 0:
            print("\n🎉 جميع الاختبارات نجحت! التقرير جاهز للاستخدام.")
        else:
            print(f"\n⚠️  هناك {self.failed} اختبار فشل. تحقق من الأخطاء أعلاه.")


def main():
    """الدالة الرئيسية"""
    tester = TestReportSystem()
    tester.run_all_tests()


if __name__ == '__main__':
    main()
