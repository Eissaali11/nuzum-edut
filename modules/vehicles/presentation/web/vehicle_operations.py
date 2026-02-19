from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, make_response
from modules.vehicles.domain.models import Vehicle, VehicleHandover, VehicleWorkshop, VehicleExternalSafetyCheck
from modules.vehicles.domain import VehicleMaintenance
from domain.employees.models import Employee, Department
from core.domain.models import User, UserRole
from core.extensions import db
from flask_login import current_user, login_required
from sqlalchemy import or_, and_, func
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl.utils
import io

from modules.vehicles.application.vehicle_management_service import (
    build_vehicle_operations_context,
    build_vehicle_operations_export_operations,
)

vehicle_operations_bp = Blueprint('vehicle_operations', __name__)

@vehicle_operations_bp.route('/')
@login_required
def vehicle_operations_list():
    """صفحة عرض جميع عمليات السيارات مع فلاتر شاملة"""

    # جلب الفلاتر من الطلب
    vehicle_filter = request.args.get('vehicle_filter', '').strip()
    operation_type = request.args.get('operation_type', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    employee_filter = request.args.get('employee_filter', '').strip()
    department_filter = request.args.get('department_filter', '').strip()

    try:
        context = build_vehicle_operations_context(
            current_user=current_user,
            vehicle_filter=vehicle_filter,
            operation_type=operation_type,
            date_from=date_from,
            date_to=date_to,
            employee_filter=employee_filter,
            department_filter=department_filter,
        )
        return render_template('vehicle_operations_list.html', **context)

    except Exception as e:
        flash(f'حدث خطأ في جلب البيانات: {str(e)}', 'error')
        return render_template('vehicle_operations_list.html',
                             operations=[],
                             vehicles=[],
                             departments=[],
                             vehicle_filter='',
                             operation_type='',
                             date_from='',
                             date_to='',
                             employee_filter='',
                             department_filter='',
                             total_operations=0,
                             handover_count=0,
                             workshop_count=0,
                             safety_count=0,
                             maintenance_count=0)

@vehicle_operations_bp.route('/test')
def test_page():
    """صفحة اختبار للتحقق من الوصول"""
    from flask import jsonify
    
    # اختبار جلب البيانات من قاعدة البيانات
    vehicle_count = Vehicle.query.count()
    handover_count = VehicleHandover.query.count()
    workshop_count = VehicleWorkshop.query.count()
    safety_count = VehicleExternalSafetyCheck.query.count()
    
    return jsonify({
        'message': 'صفحة العمليات تعمل!',
        'user_authenticated': current_user.is_authenticated if current_user else False,
        'database_stats': {
            'vehicles': vehicle_count,
            'handovers': handover_count,
            'workshops': workshop_count,
            'safety_checks': safety_count
        }
    })

@vehicle_operations_bp.route('/api/vehicle-operations/export')
@login_required
def export_vehicle_operations():
    """تصدير عمليات السيارة إلى Excel"""
    try:
        # جلب الفلاتر من الطلب
        vehicle_filter = request.args.get('vehicle_filter', '').strip()
        operation_type = request.args.get('operation_type', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        employee_filter = request.args.get('employee_filter', '').strip()
        department_filter = request.args.get('department_filter', '').strip()

        operations = build_vehicle_operations_export_operations(
            vehicle_filter=vehicle_filter,
            operation_type=operation_type,
            date_from=date_from,
            date_to=date_to,
            employee_filter=employee_filter,
            department_filter=department_filter,
        )

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "عمليات السيارات"

        # تعيين الاتجاه من اليمين إلى اليسار
        ws.sheet_view.rightToLeft = True

        # إعداد الألوان والخطوط
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A5C', end_color='1E3A5C', fill_type='solid')
        data_font = Font(name='Arial', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # إضافة العناوين
        headers = [
            'رقم السيارة',
            'نوع العملية', 
            'تاريخ العملية',
            'اسم الشخص/الفني',
            'التفاصيل',
            'الحالة',
            'القسم'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # إضافة البيانات
        for row_idx, operation in enumerate(operations, 2):
            data = [
                operation['vehicle_plate'],
                operation['type_ar'],
                operation['operation_date'].strftime('%Y-%m-%d') if operation['operation_date'] else 'غير محدد',
                operation['person_name'],
                operation['details'],
                operation['status'],
                operation['department']
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        # تعديل عرض الأعمدة
        column_widths = [15, 15, 15, 20, 40, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # حفظ الملف في الذاكرة
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # إنشاء الاستجابة
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=vehicle_operations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response

    except Exception as e:
        print(f"خطأ في تصدير عمليات السيارة: {e}")
        return jsonify({'message': f'حدث خطأ أثناء التصدير: {str(e)}'}), 500

@vehicle_operations_bp.route('/export')
def export_simple():
    """تصدير مبسط بدون مصادقة للاختبار"""
    try:
        # جلب الفلاتر من الطلب
        vehicle_filter = request.args.get('vehicle_filter', '').strip()
        operation_type = request.args.get('operation_type', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()

        operations = build_vehicle_operations_export_operations(
            vehicle_filter=vehicle_filter,
            operation_type=operation_type,
            date_from=date_from,
            date_to=date_to,
        )

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "عمليات السيارات"

        # تعيين الاتجاه من اليمين إلى اليسار
        ws.sheet_view.rightToLeft = True

        # إعداد الألوان والخطوط
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A5C', end_color='1E3A5C', fill_type='solid')
        data_font = Font(name='Arial', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # إضافة العناوين
        headers = [
            'رقم السيارة',
            'نوع العملية', 
            'تاريخ العملية',
            'اسم الشخص/الفني',
            'التفاصيل',
            'الحالة',
            'القسم'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # إضافة البيانات
        for row_idx, operation in enumerate(operations, 2):
            data = [
                operation['vehicle_plate'],
                operation['type_ar'],
                operation['operation_date'].strftime('%Y-%m-%d') if operation['operation_date'] else 'غير محدد',
                operation['person_name'],
                operation['details'],
                operation['status'],
                operation['department']
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        # تعديل عرض الأعمدة
        column_widths = [15, 15, 15, 20, 40, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # حفظ الملف في الذاكرة
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # إنشاء الاستجابة
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=vehicle_operations_filtered_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response

    except Exception as e:
        print(f"خطأ في تصدير عمليات السيارة المبسط: {e}")
        return f"خطأ في التصدير: {str(e)}", 500

@vehicle_operations_bp.route('/test-operations')
def test_operations():
    """اختبار عرض العمليات بدون مصادقة"""
    vehicle_filter = request.args.get('vehicle_filter', '').strip()
    
    # جلب عمليات التسليم والاستلام للاختبار
    handover_query = VehicleHandover.query.join(Vehicle, VehicleHandover.vehicle_id == Vehicle.id)
    
    if vehicle_filter:
        handover_query = handover_query.filter(Vehicle.plate_number.ilike(f'%{vehicle_filter}%'))
    
    handovers = handover_query.all()
    
    operations = []
    for handover in handovers:
        operations.append({
            'id': handover.id,
            'type': 'handover',
            'vehicle_plate': handover.vehicle.plate_number if handover.vehicle else 'غير محدد',
            'operation_date': str(handover.handover_date),
            'person_name': handover.person_name,
            'details': f"{handover.handover_type or 'تسليم/استلام'} - الوقود: {handover.fuel_level or 'غير محدد'}",
        })
    
    return jsonify({
        'message': f'تم العثور على {len(operations)} عملية',
        'filter': vehicle_filter,
        'operations': operations
    })

