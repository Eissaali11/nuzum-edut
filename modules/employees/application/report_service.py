"""Employee reporting application services."""
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from io import BytesIO
from math import radians, sin, cos, sqrt, atan2
from typing import Any, Dict, Optional

from core.extensions import db
from models import Employee, EmployeeLocation, SystemAudit
from modules.employees.application.file_service import UPLOAD_FOLDER
from modules.employees.application.tracking_service import format_time_12hr_arabic
from utils.employee_basic_report import generate_employee_basic_pdf
from utils.employee_comprehensive_report_updated import (
    generate_employee_comprehensive_pdf,
    generate_employee_comprehensive_excel,
)
from utils.excel import export_employee_attendance_to_excel


@dataclass
class ReportResult:
    success: bool
    message: str
    category: str
    output: Any = None
    filename: str = ""
    mimetype: str = ""
    data: Dict[str, Any] = field(default_factory=dict)


def export_basic_report_pdf(employee_id: int) -> ReportResult:
    try:
        employee = Employee.query.get(employee_id)
        if not employee:
            return ReportResult(False, "الموظف غير موجود", "danger")

        pdf_buffer = generate_employee_basic_pdf(employee_id)
        if not pdf_buffer:
            return ReportResult(False, "لم يتم العثور على بيانات كافية لإنشاء التقرير", "warning")

        current_date = datetime.now().strftime("%Y%m%d")
        filename = f"تقرير_أساسي_{employee.name}_{current_date}.pdf"

        audit = SystemAudit(
            action="export",
            entity_type="employee_basic_report",
            entity_id=employee.id,
            details=f"تم تصدير التقرير الأساسي للموظف: {employee.name}",
        )
        db.session.add(audit)
        db.session.commit()

        return ReportResult(
            True,
            "تم تصدير التقرير الأساسي بنجاح",
            "success",
            output=pdf_buffer,
            filename=filename,
            mimetype="application/pdf",
        )
    except Exception as exc:
        db.session.rollback()
        return ReportResult(False, f"خطأ في تصدير PDF: {exc}", "danger")


def export_comprehensive_report_pdf(employee_id: int) -> ReportResult:
    try:
        employee = Employee.query.get(employee_id)
        if not employee:
            return ReportResult(False, "الموظف غير موجود", "danger")

        output = generate_employee_comprehensive_pdf(employee_id)
        if not output:
            return ReportResult(False, "لم يتم العثور على بيانات كافية لإنشاء التقرير", "warning")

        filename = f"تقرير_شامل_{employee.name}_{datetime.now().strftime('%Y%m%d')}.pdf"

        audit = SystemAudit(
            action="export",
            entity_type="employee_report",
            entity_id=employee.id,
            details=f"تم إنشاء تقرير شامل للموظف: {employee.name}",
        )
        db.session.add(audit)
        db.session.commit()

        return ReportResult(
            True,
            "تم إنشاء التقرير الشامل بنجاح",
            "success",
            output=output,
            filename=filename,
            mimetype="application/pdf",
        )
    except Exception as exc:
        db.session.rollback()
        return ReportResult(
            False, f"حدث خطأ أثناء إنشاء التقرير الشامل: {exc}", "danger"
        )


def export_comprehensive_report_excel(employee_id: int) -> ReportResult:
    try:
        employee = Employee.query.get(employee_id)
        if not employee:
            return ReportResult(False, "الموظف غير موجود", "danger")

        output = generate_employee_comprehensive_excel(employee_id)
        if not output:
            return ReportResult(False, "لم يتم العثور على بيانات كافية لإنشاء التقرير", "warning")

        filename = f"تقرير_شامل_{employee.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        audit = SystemAudit(
            action="export",
            entity_type="employee_report_excel",
            entity_id=employee.id,
            details=f"تم تصدير تقرير شامل (إكسل) للموظف: {employee.name}",
        )
        db.session.add(audit)
        db.session.commit()

        return ReportResult(
            True,
            "تم إنشاء التقرير الشامل (إكسل) بنجاح",
            "success",
            output=output,
            filename=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        db.session.rollback()
        return ReportResult(
            False, f"حدث خطأ أثناء إنشاء التقرير الشامل (إكسل): {exc}", "danger"
        )


def export_attendance_excel(
    employee_id: int, month: Optional[str], year: Optional[str]
) -> ReportResult:
    try:
        employee = Employee.query.get(employee_id)
        if not employee:
            return ReportResult(False, "الموظف غير موجود", "danger")

        warnings = []
        month_int = None
        year_int = None

        if month:
            try:
                month_int = int(month)
            except (ValueError, TypeError):
                warnings.append("قيمة الشهر غير صالحة، تم استخدام الشهر الحالي")
        if year:
            try:
                year_int = int(year)
            except (ValueError, TypeError):
                warnings.append("قيمة السنة غير صالحة، تم استخدام السنة الحالية")

        output = export_employee_attendance_to_excel(employee, month_int, year_int)

        current_date = datetime.now().strftime("%Y%m%d")
        if month_int and year_int:
            filename = (
                f"attendance_{employee.name}_{year_int}_{month_int}_{current_date}.xlsx"
            )
        else:
            current_month = datetime.now().month
            current_year = datetime.now().year
            filename = (
                f"attendance_{employee.name}_{current_year}_{current_month}_{current_date}.xlsx"
            )

        audit = SystemAudit(
            action="export",
            entity_type="attendance",
            entity_id=employee.id,
            details=f"تم تصدير سجل الحضور للموظف: {employee.name}",
        )
        db.session.add(audit)
        db.session.commit()

        return ReportResult(
            True,
            "تم تصدير ملف الحضور بنجاح",
            "success",
            output=output,
            filename=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            data={"warnings": warnings},
        )
    except Exception as exc:
        db.session.rollback()
        return ReportResult(
            False, f"حدث خطأ أثناء تصدير ملف الحضور: {exc}", "danger"
        )


def export_track_history_pdf(employee_id: int) -> ReportResult:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display

    def prepare_arabic(text):
        if not text:
            return ""
        return get_display(reshape(str(text)))

    employee = Employee.query.get(employee_id)
    if not employee:
        return ReportResult(False, "الموظف غير موجود", "danger")

    cutoff_time = datetime.utcnow() - timedelta(hours=24)
    locations = (
        EmployeeLocation.query.filter(
            EmployeeLocation.employee_id == employee_id,
            EmployeeLocation.recorded_at >= cutoff_time,
        )
        .order_by(EmployeeLocation.recorded_at.asc())
        .all()
    )

    pdfmetrics.registerFont(TTFont("Amiri", "static/fonts/Amiri-Regular.ttf"))
    pdfmetrics.registerFont(TTFont("AmiriBold", "static/fonts/Amiri-Bold.ttf"))
    pdfmetrics.registerFontFamily("Amiri", normal="Amiri", bold="AmiriBold")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    story = []

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=getSampleStyleSheet()["Heading1"],
        fontName="AmiriBold",
        fontSize=24,
        textColor=colors.HexColor("#1e1b4b"),
        alignment=TA_CENTER,
        spaceAfter=15,
        spaceBefore=10,
    )

    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=getSampleStyleSheet()["Heading2"],
        fontName="AmiriBold",
        fontSize=16,
        textColor=colors.HexColor("#4f46e5"),
        alignment=TA_RIGHT,
        spaceAfter=12,
        spaceBefore=8,
    )

    normal_style = ParagraphStyle(
        "CustomNormal",
        parent=getSampleStyleSheet()["Normal"],
        fontName="Amiri",
        fontSize=12,
        alignment=TA_RIGHT,
        rightIndent=0,
        leftIndent=0,
        textColor=colors.HexColor("#374151"),
        leading=18,
    )

    title_text = prepare_arabic(f"سجل تحركات الموظف - {employee.name}")
    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 0.5 * cm))

    info_data = [
        [prepare_arabic("رقم الموظف:"), prepare_arabic(str(employee.employee_id))],
        [prepare_arabic("الاسم:"), prepare_arabic(employee.name)],
        [prepare_arabic("عدد النقاط:"), str(len(locations))],
        [prepare_arabic("التاريخ:"), datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]

    if employee.departments:
        info_data.insert(
            2, [prepare_arabic("القسم:"), prepare_arabic(employee.departments[0].name)]
        )

    info_table = Table(info_data, colWidths=[4.5 * cm, 12 * cm])
    info_table.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), "Amiri", 12),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#4f46e5")),
                ("BACKGROUND", (1, 0), (1, -1), colors.white),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
                ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor("#1e1b4b")),
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 1.5, colors.HexColor("#c7d2fe")),
                ("LEFTPADDING", (0, 0), (-1, -1), 15),
                ("RIGHTPADDING", (0, 0), (-1, -1), 15),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ("ROUNDEDCORNERS", [5, 5, 5, 5]),
            ]
        )
    )

    story.append(info_table)
    story.append(Spacer(1, 1 * cm))

    if locations:
        max_speed = max([float(loc.speed_kmh) if loc.speed_kmh else 0 for loc in locations])
        total_distance = 0
        vehicle_count = sum([1 for loc in locations if loc.vehicle_id])

        for i in range(1, len(locations)):
            prev = locations[i - 1]
            curr = locations[i]
            lat1, lon1 = float(prev.latitude), float(prev.longitude)
            lat2, lon2 = float(curr.latitude), float(curr.longitude)

            R = 6371
            dlat = radians(lat2 - lat1)
            dlon = radians(lon2 - lon1)
            a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(
                dlon / 2
            ) ** 2
            c = 2 * atan2(sqrt(a), sqrt(1 - a))
            total_distance += R * c

        subtitle = prepare_arabic("إحصائيات التحركات")
        story.append(Paragraph(subtitle, subtitle_style))
        story.append(Spacer(1, 0.3 * cm))

        stats_data = [
            [prepare_arabic("إجمالي المسافة:"), f"{total_distance:.2f} " + prepare_arabic("كم")],
            [prepare_arabic("أقصى سرعة:"), f"{max_speed:.1f} " + prepare_arabic("كم/س")],
            [prepare_arabic("عدد النقاط على سيارة:"), str(vehicle_count)],
        ]

        stats_table = Table(stats_data, colWidths=[4.5 * cm, 12 * cm])
        stats_table.setStyle(
            TableStyle(
                [
                    ("FONT", (0, 0), (-1, -1), "Amiri", 12),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#10b981")),
                    ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#d1fae5")),
                    ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
                    ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor("#065f46")),
                    ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 1.5, colors.HexColor("#6ee7b7")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 15),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 15),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )

        story.append(stats_table)
        story.append(Spacer(1, 1 * cm))

        subtitle2 = prepare_arabic("سجل التحركات التفصيلي")
        story.append(Paragraph(subtitle2, subtitle_style))
        story.append(Spacer(1, 0.3 * cm))

        data = [
            [
                prepare_arabic("#"),
                prepare_arabic("الوقت"),
                prepare_arabic("الإحداثيات"),
                prepare_arabic("السرعة"),
                prepare_arabic("السيارة"),
            ]
        ]

        for idx, loc in enumerate(locations, 1):
            coords = f"{float(loc.latitude):.6f}, {float(loc.longitude):.6f}"
            coords_link = (
                f"<link href=\"https://www.google.com/maps?q={float(loc.latitude)},{float(loc.longitude)}\" "
                f"color=\"#2563eb\"><u>{coords}</u></link>"
            )

            speed_val = (
                f"{float(loc.speed_kmh):.1f} " + prepare_arabic("كم/س")
                if loc.speed_kmh and float(loc.speed_kmh) > 0
                else "-"
            )

            vehicle_info = "-"
            if loc.vehicle_id and loc.vehicle:
                vehicle_info = prepare_arabic(
                    f"{loc.vehicle.plate_number} - {loc.vehicle.make}"
                )

            time_str = format_time_12hr_arabic(loc.recorded_at)

            data.append(
                [
                    str(idx),
                    time_str,
                    Paragraph(coords_link, normal_style),
                    speed_val,
                    vehicle_info,
                ]
            )

        table = Table(data, colWidths=[1.2 * cm, 3.5 * cm, 5 * cm, 3 * cm, 4.5 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("FONT", (0, 0), (-1, 0), "AmiriBold", 13),
                    ("FONT", (0, 1), (-1, -1), "Amiri", 11),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#1e1b4b")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("ALIGN", (2, 1), (2, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#c7d2fe")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#eef2ff")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )

        story.append(table)
    else:
        no_data_text = prepare_arabic("لا توجد بيانات تتبع خلال آخر 24 ساعة")
        story.append(Paragraph(no_data_text, normal_style))

    doc.build(story)
    buffer.seek(0)

    filename = f"track_history_{employee.employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return ReportResult(
        True,
        "تم تصدير سجل التحركات بنجاح",
        "success",
        output=buffer,
        filename=filename,
        mimetype="application/pdf",
    )


def export_track_history_excel(employee_id: int) -> ReportResult:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.drawing.image import Image as XLImage
    import os
    import requests

    employee = Employee.query.get(employee_id)
    if not employee:
        return ReportResult(False, "الموظف غير موجود", "danger")

    cutoff_time = datetime.utcnow() - timedelta(hours=24)
    locations = (
        EmployeeLocation.query.filter(
            EmployeeLocation.employee_id == employee_id,
            EmployeeLocation.recorded_at >= cutoff_time,
        )
        .order_by(EmployeeLocation.recorded_at.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "سجل التحركات"
    ws.right_to_left = True

    ws.merge_cells("A1:J1")
    ws["A1"] = f"📍 سجل تحركات الموظف - {employee.name}"
    ws["A1"].font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = GradientFill(stop=("4F46E5", "7C3AED"))
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:J2")
    ws["A2"] = f"التقرير الشامل - {datetime.now().strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(name="Arial", size=12, italic=True, color="6366F1")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    current_row = 4

    ws.merge_cells(f"A{current_row}:D{current_row}")
    ws[f"A{current_row}"] = "📋 معلومات الموظف"
    ws[f"A{current_row}"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws[f"A{current_row}"].fill = PatternFill(
        start_color="6366F1", end_color="6366F1", fill_type="solid"
    )
    ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 28

    current_row += 1
    info_data = [
        ["رقم الموظف:", employee.employee_id, "الاسم:", employee.name],
        [
            "القسم:",
            employee.departments[0].name if employee.departments else "-",
            "تاريخ التقرير:",
            format_time_12hr_arabic(datetime.now()),
        ],
    ]

    for row_data in info_data:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color="C7D2FE"),
                right=Side(style="thin", color="C7D2FE"),
                top=Side(style="thin", color="C7D2FE"),
                bottom=Side(style="thin", color="C7D2FE"),
            )
            if col_idx % 2 == 1:
                cell.font = Font(name="Arial", size=11, bold=True, color="312E81")
                cell.fill = PatternFill(
                    start_color="E0E7FF", end_color="E0E7FF", fill_type="solid"
                )
            else:
                cell.font = Font(name="Arial", size=11, color="1E1B4B")
                cell.fill = PatternFill(
                    start_color="F5F7FF", end_color="F5F7FF", fill_type="solid"
                )
        current_row += 1

    current_row += 1

    if locations:
        max_speed = max([float(loc.speed_kmh) if loc.speed_kmh else 0 for loc in locations])
        total_distance = 0
        vehicle_count = sum([1 for loc in locations if loc.vehicle_id])

        for i in range(1, len(locations)):
            prev = locations[i - 1]
            curr = locations[i]
            lat1, lon1 = float(prev.latitude), float(prev.longitude)
            lat2, lon2 = float(curr.latitude), float(curr.longitude)

            R = 6371
            dlat = radians(lat2 - lat1)
            dlon = radians(lon2 - lon1)
            a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(
                dlon / 2
            ) ** 2
            c = 2 * atan2(sqrt(a), sqrt(1 - a))
            total_distance += R * c

        ws.merge_cells(f"A{current_row}:D{current_row}")
        ws[f"A{current_row}"] = "📊 إحصائيات التحركات"
        ws[f"A{current_row}"].font = Font(
            name="Arial", size=14, bold=True, color="FFFFFF"
        )
        ws[f"A{current_row}"].fill = PatternFill(
            start_color="10B981", end_color="10B981", fill_type="solid"
        )
        ws[f"A{current_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[current_row].height = 28

        current_row += 1
        stats_data = [
            ["عدد نقاط التتبع:", len(locations), "إجمالي المسافة:", f"{total_distance:.2f} كم"],
            ["أقصى سرعة:", f"{max_speed:.1f} كم/س", "نقاط على سيارة:", vehicle_count],
        ]

        for row_data in stats_data:
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="medium", color="10B981"),
                    right=Side(style="medium", color="10B981"),
                    top=Side(style="thin", color="D1FAE5"),
                    bottom=Side(style="thin", color="D1FAE5"),
                )
                if col_idx % 2 == 1:
                    cell.font = Font(name="Arial", size=12, bold=True, color="065F46")
                    cell.fill = PatternFill(
                        start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
                    )
                else:
                    cell.font = Font(name="Arial", size=13, bold=True, color="059669")
                    cell.fill = PatternFill(
                        start_color="A7F3D0", end_color="A7F3D0", fill_type="solid"
                    )
            current_row += 1

        current_row += 1

        map_row = current_row
        ws.merge_cells(f"F{map_row}:J{map_row}")
        ws[f"F{map_row}"] = "🗺️ خريطة المسار"
        ws[f"F{map_row}"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        ws[f"F{map_row}"].fill = PatternFill(
            start_color="F59E0B", end_color="F59E0B", fill_type="solid"
        )
        ws[f"F{map_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[map_row].height = 28

        lats = [float(loc.latitude) for loc in locations]
        lons = [float(loc.longitude) for loc in locations]
        center_lat = sum(lats) / len(lats)
        center_lon = sum(lons) / len(lons)

        min_lat, max_lat = min(lats), max(lats)
        min_lon, max_lon = min(lons), max(lons)

        try:
            zoom_level = 12
            lat_diff = max_lat - min_lat
            lon_diff = max_lon - min_lon

            if lat_diff > 0.5 or lon_diff > 0.5:
                zoom_level = 10
            elif lat_diff < 0.05 and lon_diff < 0.05:
                zoom_level = 14

            markers = ""
            for i, (lat, lon) in enumerate(zip(lats, lons)):
                if i == 0:
                    markers += f"{center_lat},{center_lon},lightblue1"
                elif i == len(lats) - 1:
                    markers += f"|{lat},{lon},lightblue2"
                elif i % 5 == 0:
                    markers += f"|{lat},{lon},lightblue3"

            map_url = (
                "https://staticmap.openstreetmap.de/staticmap.php"
                f"?center={center_lat},{center_lon}&zoom={zoom_level}&size=800x500"
                f"&maptype=mapnik&markers={markers}"
            )

            response = requests.get(map_url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
            if response.status_code == 200 and len(response.content) > 1000:
                map_upload_folder = os.path.join(UPLOAD_FOLDER, "maps")
                os.makedirs(map_upload_folder, exist_ok=True)

                map_filename = (
                    f"track_map_{employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                )
                map_filepath = os.path.join(map_upload_folder, map_filename)

                with open(map_filepath, "wb") as file_handle:
                    file_handle.write(response.content)

                if os.path.exists(map_filepath) and os.path.getsize(map_filepath) > 0:
                    img = XLImage(map_filepath)
                    img.width = 450
                    img.height = 300

                    ws.add_image(img, f"F{map_row + 1}")

                    for i in range(map_row + 1, map_row + 16):
                        ws.row_dimensions[i].height = 20
                else:
                    raise Exception("فشل في حفظ صورة الخريطة")
            else:
                raise Exception("فشل تحميل الخريطة")
        except Exception:
            ws.merge_cells(f"F{map_row + 1}:J{map_row + 5}")
            ws[
                f"F{map_row + 1}"
            ] = "🗺️ عرض الخريطة\n\nانقر هنا لفتح الخريطة في Google Maps"
            ws[f"F{map_row + 1}"].font = Font(
                name="Arial", size=12, bold=True, color="2563EB", underline="single"
            )
            ws[f"F{map_row + 1}"].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            google_maps_url = (
                f"https://www.google.com/maps/dir/{lats[0]},{lons[0]}/{lats[-1]},{lons[-1]}"
            )
            ws[f"F{map_row + 1}"].hyperlink = google_maps_url
            ws[f"F{map_row + 1}"].fill = PatternFill(
                start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"
            )

            for i in range(map_row + 1, map_row + 6):
                ws.row_dimensions[i].height = 25

        table_start_row = max(current_row + 1, map_row + 17)

        ws.merge_cells(f"A{table_start_row}:J{table_start_row}")
        ws[f"A{table_start_row}"] = "📝 سجل التحركات التفصيلي"
        ws[f"A{table_start_row}"].font = Font(
            name="Arial", size=14, bold=True, color="FFFFFF"
        )
        ws[f"A{table_start_row}"].fill = PatternFill(
            start_color="EC4899", end_color="EC4899", fill_type="solid"
        )
        ws[f"A{table_start_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[table_start_row].height = 28

        table_start_row += 1

        headers = [
            "#",
            "الوقت",
            "خط العرض",
            "خط الطول",
            "السرعة (كم/س)",
            "الحالة",
            "السيارة",
            "الدقة (م)",
            "رابط الموقع",
            "ملاحظات",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start_row, column=col_idx)
            cell.value = header
            cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F46E5", end_color="4F46E5", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="medium", color="312E81"),
                right=Side(style="medium", color="312E81"),
                top=Side(style="medium", color="312E81"),
                bottom=Side(style="medium", color="312E81"),
            )
        ws.row_dimensions[table_start_row].height = 30

        for idx, loc in enumerate(locations, 1):
            row = table_start_row + idx

            speed_val = float(loc.speed_kmh) if loc.speed_kmh else 0

            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = format_time_12hr_arabic(loc.recorded_at)
            ws.cell(row=row, column=3).value = float(loc.latitude)
            ws.cell(row=row, column=4).value = float(loc.longitude)
            ws.cell(row=row, column=5).value = f"{speed_val:.1f}" if speed_val > 0 else "-"

            if speed_val > 100:
                ws.cell(row=row, column=6).value = "⚠️ سرعة عالية"
                status_color = "FEE2E2"
            elif speed_val > 60:
                ws.cell(row=row, column=6).value = "⚡ متوسطة"
                status_color = "FEF3C7"
            elif speed_val > 0:
                ws.cell(row=row, column=6).value = "✅ عادية"
                status_color = "D1FAE5"
            else:
                ws.cell(row=row, column=6).value = "⏸️ متوقف"
                status_color = "E0E7FF"

            if loc.vehicle_id and loc.vehicle:
                ws.cell(row=row, column=7).value = (
                    f"🚗 {loc.vehicle.plate_number} - {loc.vehicle.make}"
                )
            else:
                ws.cell(row=row, column=7).value = "-"

            ws.cell(row=row, column=8).value = (
                f"{float(loc.accuracy_m):.1f}" if loc.accuracy_m else "-"
            )

            maps_link = f"https://www.google.com/maps?q={float(loc.latitude)},{float(loc.longitude)}"
            ws.cell(row=row, column=9).value = "📍 عرض الموقع"
            ws.cell(row=row, column=9).hyperlink = maps_link
            ws.cell(row=row, column=9).font = Font(
                name="Arial", size=10, color="2563EB", underline="single", bold=True
            )

            if speed_val > 120:
                ws.cell(row=row, column=10).value = "⚠️ تجاوز السرعة القصوى"
            elif loc.accuracy_m and float(loc.accuracy_m) > 50:
                ws.cell(row=row, column=10).value = "⚠️ دقة منخفضة"
            else:
                ws.cell(row=row, column=10).value = "-"

            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(
                    left=Side(style="thin", color="C7D2FE"),
                    right=Side(style="thin", color="C7D2FE"),
                    top=Side(style="thin", color="C7D2FE"),
                    bottom=Side(style="thin", color="C7D2FE"),
                )

                if col == 6:
                    cell.fill = PatternFill(
                        start_color=status_color, end_color=status_color, fill_type="solid"
                    )
                elif idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="F5F7FF", end_color="F5F7FF", fill_type="solid"
                    )
                else:
                    cell.fill = PatternFill(
                        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                    )

                if col == 5 and speed_val > 100:
                    cell.font = Font(name="Arial", size=11, bold=True, color="DC2626")

            ws.row_dimensions[row].height = 22
    else:
        ws.merge_cells(f"A{current_row}:J{current_row}")
        ws[f"A{current_row}"] = "⚠️ لا توجد بيانات تتبع خلال آخر 24 ساعة"
        ws[f"A{current_row}"].font = Font(name="Arial", size=14, bold=True, color="DC2626")
        ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 40

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"track_history_{employee.employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return ReportResult(
        True,
        "تم تصدير سجل التحركات (إكسل) بنجاح",
        "success",
        output=output,
        filename=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
